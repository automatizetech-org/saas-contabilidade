"""
Interface para gerenciar empresas (CPF/CNPJ + senha) e rodar o bot de
download de XML de NFS-e Emitidas/Recebidas com Playwright.

Estilo e layout espelhados do interface exemplo.py: lista de empresas
com busca, botoes de gerenciar (adicionar/editar/excluir), log em caixa
com borda azul clara e botoes de controle na base.
"""

import base64
import html
import hashlib
import json
import math
import os
import re
import shutil
import subprocess
import sys
import atexit
import signal
import tempfile
import threading
import time
import xml.etree.ElementTree as ET
from collections import deque
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple
from urllib.parse import urljoin
from zipfile import ZipFile, BadZipFile

import requests
from PIL import Image, ImageChops, ImageDraw, ImageFilter, ImageQt
from PySide6.QtCore import (
    Qt,
    QEasingCurve,
    QDate,
    QEvent,
    QObject,
    QPropertyAnimation,
    QDateTime,
    QThread,
    QTimer,
    QUrl,
    Signal,
    QSize,
)
from PySide6.QtGui import QDesktopServices, QIcon, QPainter, QPixmap, QColor, QAction
from PySide6.QtWidgets import (
    QApplication,
    QCheckBox,
    QDialog,
    QDialogButtonBox,
    QGraphicsDropShadowEffect,
    QFrame,
    QHBoxLayout,
    QInputDialog,
    QLabel,
    QLineEdit,
    QDateEdit,
    QDateTimeEdit,
    QMainWindow,
    QMessageBox,
    QSpacerItem,
    QSizePolicy,
    QTextEdit,
    QFileDialog,
    QPushButton,
    QComboBox,
    QRadioButton,
    QListWidget,
    QListWidgetItem,
    QScrollArea,
    QStackedLayout,
    QStackedWidget,
    QSystemTrayIcon,
    QVBoxLayout,
    QWidget,
)
from supabase import create_client

# --------------------------------------------------------------------
# Constantes e caminhos
# --------------------------------------------------------------------

RUNTIME_FOLDER_NAME = "nfs_padrao"


def _resolve_runtime_base_dir() -> Path:
    explicit = os.environ.get("ROBOT_SCRIPT_DIR", "").strip().rstrip(os.sep)
    if explicit:
        return Path(explicit).resolve()

    current_dir = Path(__file__).resolve().parent
    candidates: List[Path] = []
    robots_root = (os.getenv("ROBOTS_ROOT_PATH") or "").strip()

    if robots_root:
        candidates.append(Path(robots_root) / RUNTIME_FOLDER_NAME)
    candidates.append(Path.home() / "Documents" / "ROBOS" / RUNTIME_FOLDER_NAME)
    candidates.append(current_dir)

    seen: set[str] = set()
    for candidate in candidates:
        try:
            resolved = candidate.resolve()
        except Exception:
            resolved = candidate
        key = str(resolved).lower()
        if key in seen:
            continue
        seen.add(key)
        if (resolved / "data" / "json").exists():
            return resolved

    return current_dir


def resolve_base_dir() -> Path:
    """
    Retorna a pasta base de assets tanto no .py quanto no .exe.
    Prioriza `_internal` ao rodar instalado, senao cai para _MEIPASS ou pasta do executavel.
    """
    if getattr(sys, "frozen", False):
        exe_dir = Path(sys.executable).resolve().parent
        internal_dir = exe_dir / "_internal"
        if internal_dir.is_dir():
            return internal_dir

        meipass_dir = getattr(sys, "_MEIPASS", None)
        if meipass_dir:
            return Path(meipass_dir)

        return exe_dir

    return _resolve_runtime_base_dir()


RUNTIME_LOGS_DIR = _resolve_runtime_base_dir() / "data" / "logs"
RUNTIME_LOG_PATH = RUNTIME_LOGS_DIR / "runtime.log"


def append_runtime_log(message: str) -> None:
    try:
        RUNTIME_LOGS_DIR.mkdir(parents=True, exist_ok=True)
        with RUNTIME_LOG_PATH.open("a", encoding="utf-8") as handle:
            handle.write(str(message).rstrip() + "\n")
    except Exception:
        pass


def emit_terminal_log(message: str) -> str:
    text = str(message or "").rstrip()
    timestamp = datetime.now().strftime("%H:%M:%S")
    line = f"[{timestamp}] [NFS] {text}" if text else f"[{timestamp}] [NFS]"
    try:
        print(line, flush=True)
    except Exception:
        pass
    return line


def _json_runtime_utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _json_runtime_ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _json_runtime_atomic_write_json(path: Path, payload: dict[str, Any]) -> None:
    _json_runtime_ensure_parent(path)
    temp_path = path.with_suffix(path.suffix + ".tmp")
    temp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    temp_path.replace(path)


class JsonRobotRuntime:
    def __init__(self, technical_id: str, display_name: str, base_dir: Path) -> None:
        self.technical_id = technical_id
        self.display_name = display_name
        self.base_dir = base_dir
        self.last_status = "inactive"
        self.json_dir = self.base_dir / "data" / "json"
        self.job_path = self.json_dir / "job.json"
        self.result_path = self.json_dir / "result.json"
        self.heartbeat_path = self.json_dir / "heartbeat.json"

    def register_robot(self) -> str:
        self.write_heartbeat(status="active")
        return self.technical_id

    def load_job(self) -> Optional[dict[str, Any]]:
        if not self.job_path.exists():
            return None
        try:
            payload = json.loads(self.job_path.read_text(encoding="utf-8"))
        except Exception:
            return None
        if not isinstance(payload, dict):
            return None
        if not payload.get("id"):
            payload["id"] = payload.get("execution_request_id") or payload.get("job_id")
        if not payload.get("job_id"):
            payload["job_id"] = payload.get("id")
        if not payload.get("execution_request_id"):
            payload["execution_request_id"] = payload.get("id")
        job_execution_id = str(
            payload.get("execution_request_id")
            or payload.get("job_id")
            or payload.get("id")
            or ""
        ).strip()
        if not job_execution_id:
            return None
        if self.result_path.exists():
            try:
                result_payload = json.loads(self.result_path.read_text(encoding="utf-8"))
            except Exception:
                result_payload = None
            if isinstance(result_payload, dict):
                result_execution_id = str(
                    result_payload.get("execution_request_id")
                    or result_payload.get("job_id")
                    or result_payload.get("event_id")
                    or ""
                ).strip()
                if result_execution_id and result_execution_id == job_execution_id:
                    return None
        return payload

    def load_job_companies(
        self,
        job: Optional[dict[str, Any]],
        company_ids: Optional[List[str]] = None,
    ) -> List[dict[str, Any]]:
        if not isinstance(job, dict):
            return []
        companies = job.get("companies")
        if not isinstance(companies, list):
            return []
        wanted = {str(company_id) for company_id in (company_ids or []) if str(company_id).strip()}
        rows: List[dict[str, Any]] = []
        for row in companies:
            if not isinstance(row, dict):
                continue
            company_id = str(row.get("company_id") or row.get("id") or "").strip()
            if wanted and company_id not in wanted:
                continue
            rows.append(row)
        return rows

    def write_heartbeat(
        self,
        *,
        status: str,
        current_job_id: Optional[str] = None,
        current_execution_request_id: Optional[str] = None,
        message: Optional[str] = None,
        progress: Optional[dict[str, Any]] = None,
        extra: Optional[dict[str, Any]] = None,
    ) -> None:
        self.last_status = status
        payload: dict[str, Any] = {
            "robot_technical_id": self.technical_id,
            "display_name": self.display_name,
            "status": status,
            "updated_at": _json_runtime_utc_now_iso(),
            "current_job_id": current_job_id,
            "current_execution_request_id": current_execution_request_id,
            "message": message,
            "progress": progress or {},
        }
        if extra:
            payload.update(extra)
        _json_runtime_atomic_write_json(self.heartbeat_path, payload)

    def write_result(
        self,
        *,
        job: Optional[dict[str, Any]],
        success: bool,
        error_message: Optional[str] = None,
        summary: Optional[dict[str, Any]] = None,
        payload: Optional[dict[str, Any]] = None,
        company_results: Optional[List[dict[str, Any]]] = None,
    ) -> None:
        execution_request_id = None
        job_id = None
        if isinstance(job, dict):
            execution_request_id = str(job.get("execution_request_id") or job.get("id") or "").strip() or None
            job_id = str(job.get("job_id") or job.get("id") or "").strip() or execution_request_id

        event_id = execution_request_id or str(uuid.uuid4())
        result_payload: dict[str, Any] = {
            "event_id": event_id,
            "job_id": job_id or event_id,
            "execution_request_id": execution_request_id,
            "robot_technical_id": self.technical_id,
            "status": "completed" if success else "failed",
            "started_at": _json_runtime_utc_now_iso(),
            "finished_at": _json_runtime_utc_now_iso(),
            "error_message": error_message,
            "summary": summary or {},
            "company_results": company_results or [],
            "payload": payload or {},
        }
        _json_runtime_atomic_write_json(self.result_path, result_payload)
        self.write_heartbeat(
            status="active",
            current_job_id=None,
            current_execution_request_id=None,
            message="result_ready",
        )


LOGIN_URL = "https://www.nfse.gov.br/EmissorNacional/Login?ReturnUrl=%2fEmissorNacional"
LOGIN_ORIGIN = "https://www.nfse.gov.br"
EMITIDAS_URL = "https://www.nfse.gov.br/EmissorNacional/Notas/Emitidas"
RECEBIDAS_URL = "https://www.nfse.gov.br/EmissorNacional/Notas/Recebidas"

BASE_DIR = resolve_base_dir()

# Pasta base dos robôs na VM: .env compartilhado (SERVER_API_URL, SUPABASE_*, etc.)
def resolve_robots_base_env_dir() -> Path:
    candidates: List[Path] = []
    env_root = (os.environ.get("ROBOTS_ROOT_PATH") or "").strip()
    robot_script_dir = (os.environ.get("ROBOT_SCRIPT_DIR") or "").strip()

    if env_root:
        candidates.append(Path(env_root))
    if robot_script_dir:
        candidates.append(Path(robot_script_dir).resolve().parent)
    candidates.append(BASE_DIR.parent)
    if getattr(sys, "frozen", False):
        exe_dir = Path(sys.executable).resolve().parent
        candidates.append(exe_dir.parent)
        candidates.append(exe_dir)

    seen: set[str] = set()
    for candidate in candidates:
        try:
            resolved = candidate.resolve()
        except Exception:
            resolved = candidate
        key = str(resolved).lower()
        if key in seen:
            continue
        seen.add(key)
        if (resolved / ".env").exists() or (resolved / ".env.example").exists():
            return resolved

    return Path(env_root).resolve() if env_root else BASE_DIR.parent.resolve()


ROBOTS_BASE_ENV_DIR = resolve_robots_base_env_dir()

# Carrega .env: primeiro da base dos robôs (VM), depois da pasta do script/.exe (override local)
try:
    from dotenv import load_dotenv
    env_robos = ROBOTS_BASE_ENV_DIR / ".env"
    if env_robos.exists():
        load_dotenv(env_robos)
    elif (ROBOTS_BASE_ENV_DIR / ".env.example").exists():
        load_dotenv(ROBOTS_BASE_ENV_DIR / ".env.example")
    env_path = BASE_DIR / ".env"
    if not env_path.exists():
        env_path = BASE_DIR / ".env.example"
    if env_path.exists():
        load_dotenv(env_path)
    if getattr(sys, "frozen", False):
        exe_dir = Path(sys.executable).resolve().parent
        if exe_dir != ROBOTS_BASE_ENV_DIR:
            load_dotenv(exe_dir / ".env")
            if not (exe_dir / ".env").exists():
                load_dotenv(exe_dir / ".env.example")
except ImportError:
    pass

CONNECTOR_SECRET = (os.environ.get("CONNECTOR_SECRET") or "").strip()


def build_server_api_headers(url_base: str) -> Dict[str, str]:
    headers: Dict[str, str] = {}
    if "ngrok" in url_base.lower():
        headers["ngrok-skip-browser-warning"] = "true"
    if CONNECTOR_SECRET:
        headers["Authorization"] = f"Bearer {hashlib.sha256(CONNECTOR_SECRET.encode('utf-8')).hexdigest()}"
    return headers

DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)

JSON_DIR = DATA_DIR / "json"
JSON_DIR.mkdir(exist_ok=True)
COMPANIES_JSON = JSON_DIR / "companies.json"
OUTPUT_PATH_JSON = JSON_DIR / "path.json"
PLAYWRIGHT_DIR = DATA_DIR / "ms-playwright"
os.environ.setdefault("PLAYWRIGHT_BROWSERS_PATH", str(PLAYWRIGHT_DIR))
PNG_DIR = DATA_DIR / "image"
ICO_DIR = DATA_DIR / "ico"
EYE_OPEN_PNG = PNG_DIR / "olho aberto.png"
EYE_CLOSED_PNG = PNG_DIR / "olho fechado.png"
LOGO_ICON = ICO_DIR / "app.ico"
LOGO_PNG = PNG_DIR / "logo.png"

DEFAULT_HEADLESS = False

AUTH_PASSWORD = "password"
AUTH_CERTIFICATE = "certificate"

DATE_MODE_PREVIOUS_MONTH = "previous_month"
DATE_MODE_PREVIOUS_DAY = "previous_day"


def _current_json_job() -> Optional[Dict[str, Any]]:
    return ACTIVE_JSON_JOB if isinstance(ACTIVE_JSON_JOB, dict) else None


def _set_current_json_job(job: Optional[Dict[str, Any]]) -> None:
    global ACTIVE_JSON_JOB
    ACTIVE_JSON_JOB = job if isinstance(job, dict) else None


def _set_last_json_result_summary(summary: Optional[Dict[str, Any]]) -> None:
    global LAST_JSON_RESULT_SUMMARY
    LAST_JSON_RESULT_SUMMARY = summary if isinstance(summary, dict) else None


def _consume_last_json_result_summary() -> Optional[Dict[str, Any]]:
    global LAST_JSON_RESULT_SUMMARY
    summary = LAST_JSON_RESULT_SUMMARY if isinstance(LAST_JSON_RESULT_SUMMARY, dict) else None
    LAST_JSON_RESULT_SUMMARY = None
    return summary


def _clear_pending_result_operations() -> None:
    PENDING_RESULT_OPERATIONS.clear()


def _append_result_operation(operation: Dict[str, Any]) -> None:
    if isinstance(operation, dict):
        PENDING_RESULT_OPERATIONS.append(operation)


def _consume_result_operations() -> List[Dict[str, Any]]:
    operations = list(PENDING_RESULT_OPERATIONS)
    PENDING_RESULT_OPERATIONS.clear()
    return operations


def ensure_openssl_legacy_provider() -> None:
    """
    Garante que o Playwright (driver Node.js) carregue o provider legacy do OpenSSL.
    Necessario para certificados antigos (ex.: algoritmos depreciados no OpenSSL 3).
    """
    node_opts = os.environ.get("NODE_OPTIONS", "")
    flag = "--openssl-legacy-provider"
    if flag in node_opts:
        return
    node_opts = f"{node_opts} {flag}".strip() if node_opts else flag
    os.environ["NODE_OPTIONS"] = node_opts


def find_logo_image_path() -> Optional[Path]:
    """
    Resolve o caminho do logo (PNG) considerando execuçao em modo portátil/_internal,
    mantendo compatibilidade com a busca usada no sefaz.py.
    """
    bases: List[Path] = []
    if hasattr(sys, "_MEIPASS"):
        bases.append(Path(getattr(sys, "_MEIPASS")))
    if getattr(sys, "frozen", False):
        exe_dir = Path(sys.executable).resolve().parent
        bases.append(exe_dir)
        bases.append(exe_dir / "_internal")
    bases.append(BASE_DIR)

    seen = set()
    unique_bases = []
    for b in bases:
        if b in seen:
            continue
        seen.add(b)
        unique_bases.append(b)

    candidates = [
        "data/IMAGE/logo 2.png",
        "data/Image/logo 2.png",
        "data/image/logo 2.png",
        "IMAGE/logo 2.png",
        "Image/logo 2.png",
        "image/logo 2.png",
        "data/IMAGE/logo.png",
        "data/Image/logo.png",
        "data/image/logo.png",
        "IMAGE/logo.png",
        "Image/logo.png",
        "image/logo.png",
        "logo.png",
    ]

    for base in unique_bases:
        for rel in candidates:
            candidate = (base / rel).resolve()
            if candidate.exists():
                return candidate

    if LOGO_PNG.exists():
        return LOGO_PNG
    return None


def find_money_image_path() -> Optional[Path]:
    """
    Busca um ícone de dinheiro/moeda para usar em destaques de valor.
    """
    bases: List[Path] = []
    if hasattr(sys, "_MEIPASS"):
        bases.append(Path(getattr(sys, "_MEIPASS")))
    if getattr(sys, "frozen", False):
        exe_dir = Path(sys.executable).resolve().parent
        bases.append(exe_dir)
        bases.append(exe_dir / "_internal")
    bases.append(BASE_DIR)

    seen = set()
    unique_bases = []
    for b in bases:
        if b in seen:
            continue
        seen.add(b)
        unique_bases.append(b)

    candidates = [
        "data/image/dinheiro.png",
        "data/Image/dinheiro.png",
        "data/IMAGE/dinheiro.png",
        "image/dinheiro.png",
        "Image/dinheiro.png",
        "IMAGE/dinheiro.png",
        "data/image/money.png",
        "image/money.png",
        "data/image/moeda.png",
        "image/moeda.png",
    ]
    for base in unique_bases:
        for rel in candidates:
            p = (base / rel).resolve()
            if p.exists():
                return p
    return None


def load_png_icon(*names: str) -> Optional[QIcon]:
    """
    Carrega um ícone da pasta de imagens, se existir (aceita variações simples de nome).
    """
    bases: List[Path] = []
    if hasattr(sys, "_MEIPASS"):
        bases.append(Path(getattr(sys, "_MEIPASS")))
    if getattr(sys, "frozen", False):
        exe_dir = Path(sys.executable).resolve().parent
        bases.append(exe_dir)
        bases.append(exe_dir / "_internal")
    bases.append(BASE_DIR)

    candidates = []
    for name in names or ("",):
        clean = name.strip().lower().replace(" ", "_")
        for ext in (".png", ".ico"):
            if clean:
                candidates.append(clean + ext)
            if name.strip():
                candidates.append(name.strip() + ext)
    candidates = list(dict.fromkeys(candidates))  # unique order

    for base in bases:
        for rel in (
            "data/image",
            "data/Image",
            "data/IMAGE",
            "image",
            "Image",
            "IMAGE",
        ):
            folder = (base / rel).resolve()
            for cand in candidates:
                p = folder / cand
                if p.exists():
                    try:
                        return QIcon(str(p))
                    except Exception:
                        continue
    return None


class GlowHoverFilter(QObject):
    """
    Adiciona um efeito de brilho suave ao passar o mouse sobre botoes.
    """

    def __init__(self, color: str = "#4da3ff", strength: int = 26, parent=None):
        super().__init__(parent)
        self.color = QColor(color)
        self.strength = strength

    def eventFilter(self, obj, event):  # noqa: N802
        if event.type() == QEvent.Enter:
            eff = getattr(obj, "_glow_effect", None)
            if eff is None:
                eff = QGraphicsDropShadowEffect(obj)
                eff.setBlurRadius(self.strength)
                eff.setOffset(0, 0)
                obj._glow_effect = eff
                obj.setGraphicsEffect(eff)
            eff.setColor(self.color)
            eff.setEnabled(True)
        elif event.type() == QEvent.Leave:
            eff = getattr(obj, "_glow_effect", None)
            if eff:
                eff.setEnabled(False)
        return False


def attach_glow(widget, color: str = "#4da3ff", strength: int = 26) -> None:
    filt = GlowHoverFilter(color, strength, widget)
    widget._glow_filter = filt  # manter referencia
    widget.installEventFilter(filt)

# --------------------------------------------------------------------
# Supabase (service role para o robô)
# --------------------------------------------------------------------

SUPABASE_URL = os.environ.get("SUPABASE_URL", "").strip()
def _get_supabase_service_role_key() -> str:
    for env_name in (
        "SUPABASE_SERVICE_ROLE_KEY",
        "SERVICE_ROLE_KEY",
        "SUPABASE_KEY",
        "SUPABASE_SECRET_KEY",
        "NEXT_PUBLIC_SUPABASE_SERVICE_ROLE_KEY",
    ):
        value = (os.getenv(env_name) or "").strip()
        if value:
            return value
    return ""

SUPABASE_SERVICE_ROLE_KEY = _get_supabase_service_role_key()
LICENSE_PATH = JSON_DIR / "license.json"

AUTO_UPDATE_APP_VERSION = "1.0.0"
UPDATER_SETTINGS_PATH = JSON_DIR / "updater.json"
UPDATER_LOG_PATH = JSON_DIR / "updater.log"
DEFAULT_UPDATER_SETTINGS: Dict[str, Any] = {
    "enabled": True,
    "check_on_startup": True,
    "github_repo": "VectorG0ld/CODIGOS",
    "asset_name": "nfs_padrao_update.zip",
    "allow_prerelease": False,
    "current_version": AUTO_UPDATE_APP_VERSION,
}
UPDATER_JSON_FILES_TO_PRESERVE = [
    "license.json",
    "config.json",
]

# Integração com painel admin (agendador + status)
ROBOT_TECHNICAL_ID = "nfs_padrao"
ROBOT_DISPLAY_NAME_DEFAULT = "NFS Padrão"
JSON_RUNTIME = JsonRobotRuntime(
    ROBOT_TECHNICAL_ID,
    ROBOT_DISPLAY_NAME_DEFAULT,
    Path(__file__).resolve().parent,
)
ACTIVE_JSON_JOB: Optional[Dict[str, Any]] = None
LAST_JSON_RESULT_SUMMARY: Optional[Dict[str, Any]] = None
PENDING_RESULT_OPERATIONS: List[Dict[str, Any]] = []
_PROCESS_SHUTDOWN_MARKED_INACTIVE = False
_WINDOWS_CONSOLE_HANDLER = None


def _mark_process_inactive(window: Any | None = None, reason: str = "process_exit") -> None:
    global _PROCESS_SHUTDOWN_MARKED_INACTIVE
    if _PROCESS_SHUTDOWN_MARKED_INACTIVE:
        return
    _PROCESS_SHUTDOWN_MARKED_INACTIVE = True
    try:
        JSON_RUNTIME.write_heartbeat(
            status="inactive",
            current_job_id=None,
            current_execution_request_id=None,
            message=reason,
        )
    except Exception:
        pass
    try:
        if window is not None and getattr(window, "_robot_id", None):
            update_robot_status(
                getattr(window, "_robot_supabase_url", "") or "",
                getattr(window, "_robot_supabase_key", "") or "",
                getattr(window, "_robot_id", "") or "",
                "inactive",
            )
    except Exception:
        pass


def _install_process_shutdown_handlers(window: Any | None = None) -> None:
    global _WINDOWS_CONSOLE_HANDLER

    def _handle_signal(signum, _frame) -> None:
        _mark_process_inactive(window, f"signal_{signum}")
        raise SystemExit(0)

    atexit.register(lambda: _mark_process_inactive(window, "atexit"))
    for signal_name in ("SIGINT", "SIGTERM", "SIGBREAK"):
        sig = getattr(signal, signal_name, None)
        if sig is None:
            continue
        try:
            signal.signal(sig, _handle_signal)
        except Exception:
            pass

    if os.name == "nt":
        try:
            import ctypes

            handler_type = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_uint)

            @handler_type
            def _console_handler(ctrl_type: int) -> bool:
                _mark_process_inactive(window, f"console_ctrl_{ctrl_type}")
                return False

            ctypes.windll.kernel32.SetConsoleCtrlHandler(_console_handler, True)
            _WINDOWS_CONSOLE_HANDLER = _console_handler
        except Exception:
            pass

# --------------------------------------------------------------------
# Helpers de documento (CPF/CNPJ) e persistencia
# --------------------------------------------------------------------
from playwright.sync_api import (
    Locator,
    sync_playwright,
    TimeoutError as PlaywrightTimeoutError,
)

# --------------------------------------------------------------------
# Estilos de botoes (mesmo padrao do sefaz.py)
# --------------------------------------------------------------------


def button_style(base: str, hover: str, pressed: str) -> str:
    """
    Retorna stylesheet para um botao no mesmo estilo do sefaz.py (gradiente, borda suave).
    """
    return f"""
    QPushButton {{
        font: 9pt 'Verdana';
        font-weight: bold;
        color: #E8F4FF;
        padding: 9px 14px;
        border-radius: 10px;
        border: 1px solid rgba(255,255,255,0.12);
        background: qlineargradient(x1:0,y1:0, x2:1,y2:1,
            stop:0 {base}, stop:1 #0F1E35);
    }}
    QPushButton:hover {{
        background: qlineargradient(x1:0,y1:0, x2:1,y2:1,
            stop:0 {hover}, stop:1 #12304E);
    }}
    QPushButton:pressed {{
        background: {pressed};
        padding-top: 11px;
        padding-bottom: 7px;
    }}
    """

# --------------------------------------------------------------------
# Helpers de documento (CPF/CNPJ) e persistencia
# --------------------------------------------------------------------


def only_digits(text: str) -> str:
    return "".join(ch for ch in text if ch.isdigit())


def format_cpf(cpf: str) -> str:
    digits = only_digits(cpf)
    if len(digits) != 11:
        return digits
    return f"{digits[0:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:11]}"


def format_cnpj(cnpj: str) -> str:
    digits = only_digits(cnpj)
    if len(digits) != 14:
        return digits
    return f"{digits[0:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:14]}"


def format_document(text: str) -> str:
    digits = only_digits(text)
    if len(digits) <= 11:
        return format_cpf(digits)
    return format_cnpj(digits)


def normalize_company_name(name: str) -> str:
    return (name or "").strip().upper()


# Tags mais comuns em NFS-e para valor de serviços (emitidas/recebidas).
NFSE_VALUE_TAGS = [
    "valorservico",
    "valorservicos",
    "valorserviço",
    "valorserviços",
    "vlrservico",
    "vlrservicos",
    "vliq",
    "vliquidonfse",
    "vliqservico",
    "vliqservicos",
    "vliqserv",
    "valorliqservico",
    "valorliqservicos",
    "valorliquidonfse",
    "vlrliqnfse",
    "valor_nfse",
    "valornfse",
    "valortotal",
    "valortotalservicos",
    "vlrtotal",
    "valorliquido",
    "valorliquidonota",
    "valorliquidoprestacao",
    "vservicos",
    "vservico",
    "vserv",
    "vnf",
]


def _parse_brl_decimal(text: str) -> Optional[Decimal]:
    clean = re.sub(r"[^\d,.\-]", "", text or "")
    if not clean:
        return None
    # Se tiver virgula, assume como decimal; remove separador de milhar.
    if "," in clean and "." in clean:
        if clean.rfind(",") > clean.rfind("."):
            clean = clean.replace(".", "").replace(",", ".")
        else:
            clean = clean.replace(",", "")
    elif "," in clean:
        clean = clean.replace(".", "").replace(",", ".")
    try:
        return Decimal(clean)
    except InvalidOperation:
        return None


def extract_nfse_value(xml_path: Path) -> Optional[float]:
    """
    Extrai o valor de servicos de uma NFS-e, priorizando tags conhecidas.
    Considera texto do elemento e atributos (ex.: ValorTotal="123,45").
    Fallback: coleta qualquer tag com 'valor'/'total'/'liq' e retorna o maior valor (evita impostos).
    Retorna float ou None se nao conseguir ler.
    """
    try:
        tree = ET.parse(xml_path)
    except Exception:
        return None
    root = tree.getroot()

    def tag_clean(el_tag: str) -> str:
        tag = el_tag.lower()
        if "}" in tag:
            tag = tag.split("}", 1)[1]
        return tag

    def collect_numeric_candidates(el: ET.Element, tag: str) -> List[Optional[Decimal]]:
        out: List[Optional[Decimal]] = []
        txt = (el.text or "").strip()
        if txt:
            out.append(_parse_brl_decimal(txt))
        for _attr, attr_val in el.attrib.items():
            if attr_val and isinstance(attr_val, str):
                out.append(_parse_brl_decimal(attr_val.strip()))
        return out

    priority_order = [
        "valorservicos",
        "valorservico",
        "valor_servicos",
        "valor_servico",
        "vliqservicos",
        "vliqservico",
        "vliqserv",
        "valorliqservico",
        "valorliqservicos",
        "valorliquidonfse",
        "vliquidonfse",
        "vliq",
        "valor_nfse",
        "valornfse",
        "valortotalservicos",
        "valortotal",
        "vservicos",
        "vservico",
        "vserv",
        "vnf",
        "valorliquido",
        "valorliquidonota",
        "vlrtotal",
        "vlrservico",
        "vlrservicos",
    ]
    priority_index = {name: idx for idx, name in enumerate(priority_order)}

    best_val: Optional[Decimal] = None
    best_score = -1
    fallback_vals: List[Decimal] = []

    for el in root.iter():
        tag = tag_clean(el.tag)
        for val in collect_numeric_candidates(el, tag):
            if val is None:
                continue
            if val <= 0:
                continue
            score = -1
            if tag in priority_index:
                score = 100 - priority_index[tag]
            elif any(kw in tag for kw in NFSE_VALUE_TAGS):
                score = 80
            else:
                if "valor" in tag and ("serv" in tag or "liq" in tag or "total" in tag or "nfse" in tag):
                    score = 60
                elif "valor" in tag or "total" in tag:
                    score = 50
                elif "serv" in tag or "vliq" in tag or "vserv" in tag:
                    score = 40
            if score >= 0:
                if score > best_score or (score == best_score and val > (best_val or Decimal("0"))):
                    best_score = score
                    best_val = val
            if score >= 0 or "valor" in tag or "total" in tag or "liq" in tag:
                fallback_vals.append(val)

    if best_val is not None:
        return float(best_val)
    if fallback_vals:
        return float(max(fallback_vals))
    return None


def _find_local_text(root: ET.Element, names: Tuple[str, ...]) -> Optional[str]:
    """
    Busca o primeiro elemento cujo local-name casado (ignorando namespace) esteja em names.
    """
    targets = {name.lower() for name in names}
    for element in root.iter():
        tag = element.tag
        if isinstance(tag, str):
            local = tag.split("}", 1)[-1] if "}" in tag else tag
        else:
            continue
        if local.lower() in targets:
            text = (element.text or "").strip()
            if text:
                return text
    return None


def _inspect_nfse_status(xml_path: Path) -> Dict[str, Optional[str]]:
    """
    Extrai metadados de situacao da NFS-e com base no XML baixado.
    Retorna informacoes de codigo de situacao, numero, chave de acesso e classificacao tratada.
    """
    info = {
        "status": "active",
        "status_code": None,
        "status_text": None,
        "note_number": None,
        "access_key": None,
    }
    try:
        tree = ET.parse(xml_path)
    except Exception:
        return info

    root = tree.getroot()
    status_code = _find_local_text(root, ("tsSituacaoNFSe", "tsSituacaoNfse", "situacaoNFSe", "situacao"))
    status_text = _find_local_text(root, ("situacao", "situacaoNFSe", "situacaoNfse", "tsSituacaoNFSe"))
    cancel_flag = _find_local_text(root, ("cancelada",))
    note_number = _find_local_text(root, ("Numero", "numero", "nNFSe", "nNfse", "nNF", "nNFSe", "numeroNFSe"))
    access_key = _find_local_text(
        root,
        (
            "ChaveAcesso",
            "chaveacesso",
            "Chave",
            "ChaveNFSe",
            "ChaveNfse",
            "chavenfse",
            "ChaveNfse",
            "chaveNfse",
            "refNF",
            "refNf",
            "CodigoVerificacao",
            "codigoVerificacao",
        ),
    )

    info["status_code"] = status_code
    info["status_text"] = status_text
    info["note_number"] = note_number
    info["access_key"] = access_key

    if status_code:
        code = status_code.strip()
        if code == "3":
            info["status"] = "cancelled"
        elif code == "4":
            info["status"] = "substituted"
        elif code == "2":
            info["status"] = "active"
    elif cancel_flag:
        if cancel_flag.strip().lower() in {"s", "sim", "true", "1"}:
            info["status"] = "cancelled"
    elif status_text:
        normalized = status_text.lower()
        if "cancel" in normalized or "cancelada" in normalized:
            info["status"] = "cancelled"
        elif "substituída" in normalized or "substituida" in normalized:
            info["status"] = "substituted"

    return info


def extract_nfse_service_codes(xml_path: Path) -> List[Tuple[str, str]]:
    """
    Extrai pares (cTribNac, xTribNac) do XML da NFS-e.
    """
    try:
        tree = ET.parse(xml_path)
    except Exception:
        return []

    root = tree.getroot()

    def _local_name(element: ET.Element) -> str:
        tag = element.tag
        if not isinstance(tag, str):
            return ""
        return tag.split("}", 1)[-1].lower()

    extracted: List[Tuple[str, str]] = []
    seen: set = set()
    global_desc = _find_local_text(root, ("xTribNac",)) or ""

    for parent in root.iter():
        children = list(parent)
        if not children:
            continue
        code_values = [
            (child.text or "").strip()
            for child in children
            if _local_name(child) == "ctribnac" and (child.text or "").strip()
        ]
        if not code_values:
            continue
        desc = ""
        for child in children:
            if _local_name(child) == "xtribnac":
                desc = (child.text or "").strip()
                if desc:
                    break
        for code in code_values:
            key = (code, desc)
            if key in seen:
                continue
            seen.add(key)
            extracted.append(key)

    if extracted:
        if global_desc:
            normalized: List[Tuple[str, str]] = []
            normalized_seen: set = set()
            for code, desc in extracted:
                desc_final = desc or global_desc
                key = (code, desc_final)
                if key in normalized_seen:
                    continue
                normalized_seen.add(key)
                normalized.append(key)
            return normalized
        return extracted

    fallback_code = _find_local_text(root, ("cTribNac",))
    if not fallback_code:
        return []
    fallback_desc = global_desc
    return [(fallback_code, fallback_desc)]


def format_ctribnac_display(code: str) -> str:
    """
    Formata cTribNac para exibicao no padrao XX.XX.XX.
    """
    raw = (code or "").strip()
    digits = only_digits(raw)
    if len(digits) == 5:
        digits = f"0{digits}"
    if len(digits) == 6:
        return f"{digits[:2]}.{digits[2:4]}.{digits[4:6]}"
    if len(digits) == 4:
        return f"{digits[:2]}.{digits[2:]}"
    return raw or "-"


EMISSION_DATE_FORMATS = (
    "%d/%m/%Y",
    "%d/%m/%Y %H:%M:%S",
    "%d/%m/%Y %H:%M:%S.%f",
    "%Y-%m-%d",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M:%S.%f",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%dT%H:%M:%S.%f",
    "%Y-%m-%dT%H:%M:%S%z",
    "%Y-%m-%dT%H:%M:%S.%f%z",
    "%Y%m%d",
    "%Y%m%d%H%M%S",
    "%Y%m%d%H%M",
)
PRIORITY_EMISSION_TAGS = (
    "dataemissao",
    "dhproc",
    "datahoradeemissao",
    "dataemissaonfse",
    "dataemissaoprestacao",
    "dataemissaonotafiscal",
    "demissao",
    "dtemissao",
    "dhemissao",
    "dhemi",
    "dhemit",
)


def _normalize_for_strptime(value: str) -> str:
    cleaned = value.strip()
    if cleaned.endswith("Z"):
        cleaned = f"{cleaned[:-1]}+0000"
    if len(cleaned) >= 6 and cleaned[-3] == ":" and cleaned[-6] in "+-":
        cleaned = f"{cleaned[:-6]}{cleaned[-6:-3]}{cleaned[-2:]}"
    return cleaned


def _try_parse_emission_date(text: str) -> Optional[date]:
    value = text.strip()
    if not value:
        return None
    iso_candidate = value
    if iso_candidate.endswith("Z"):
        iso_candidate = f"{iso_candidate[:-1]}+00:00"
    try:
        parsed = datetime.fromisoformat(iso_candidate)
        return parsed.date()
    except Exception:
        pass
    for fmt in EMISSION_DATE_FORMATS:
        candidate = value
        if fmt.endswith("%z"):
            candidate = _normalize_for_strptime(candidate)
        else:
            candidate = candidate.replace("Z", "")
        try:
            parsed = datetime.strptime(candidate, fmt)
            return parsed.date()
        except Exception:
            continue
    digits_only = re.sub(r"[^\d]", "", value)
    if len(digits_only) == 8:
        try:
            parsed = datetime.strptime(digits_only, "%Y%m%d")
            return parsed.date()
        except Exception:
            pass
    return None


def _extract_xml_emission_date(xml_path: Path) -> Optional[date]:
    try:
        tree = ET.parse(xml_path)
    except Exception:
        return None

    root = tree.getroot()
    best_date: Optional[date] = None
    best_score = (1, 0, len(PRIORITY_EMISSION_TAGS))

    def _local_name(element: ET.Element) -> str:
        tag = element.tag
        if not isinstance(tag, str):
            return ""
        return tag.split("}", 1)[-1].lower()

    def _stack_contains(stack: List[str], keyword: str) -> bool:
        lowercase = keyword.lower()
        return any(lowercase in local for local in stack)

    def _visit(element: ET.Element, ancestors: List[str]) -> None:
        nonlocal best_date, best_score
        local = _local_name(element)
        text = (element.text or "").strip()
        if text:
            rank = PRIORITY_EMISSION_TAGS.index(local) if local in PRIORITY_EMISSION_TAGS else None
            if rank is None:
                if "emiss" not in local or "emissor" in local:
                    pass
                else:
                    rank = len(PRIORITY_EMISSION_TAGS)
            if rank is not None:
                parsed = _try_parse_emission_date(text)
                if parsed is not None:
                    stack_names = ancestors + [local]
                    has_nfse = _stack_contains(stack_names, "nfse")
                    has_dps = _stack_contains(stack_names, "dps")
                    nfse_flag = 0 if has_nfse else 1
                    dps_penalty = 1 if has_dps else 0
                    score = (nfse_flag, dps_penalty, rank)
                    if score < best_score:
                        best_score = score
                        best_date = parsed
                        if nfse_flag == 0 and rank == 0:
                            return
        new_ancestors = ancestors + [local]
        for child in element:
            _visit(child, new_ancestors)

    _visit(root, [])
    return best_date


def is_valid_cpf(cpf: str) -> bool:
    digits = only_digits(cpf)
    if len(digits) != 11 or digits == digits[0] * 11:
        return False

    def calc_digit(slice_len: int, factor: int) -> int:
        total = sum(int(digits[i]) * (factor - i) for i in range(slice_len))
        rest = (total * 10) % 11
        return 0 if rest == 10 else rest

    d1 = calc_digit(9, 10)
    d2 = calc_digit(10, 11)
    return d1 == int(digits[9]) and d2 == int(digits[10])


def is_valid_cnpj(cnpj: str) -> bool:
    digits = only_digits(cnpj)
    if len(digits) != 14 or digits == digits[0] * 14:
        return False

    def calc_digit(length: int) -> int:
        weights = [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
        total = sum(int(digits[i]) * weights[i + 13 - length] for i in range(length))
        rest = total % 11
        return 0 if rest < 2 else 11 - rest

    d1 = calc_digit(12)
    d2 = calc_digit(13)
    return d1 == int(digits[12]) and d2 == int(digits[13])


def is_valid_document(text: str) -> bool:
    digits = only_digits(text)
    if len(digits) == 11:
        return is_valid_cpf(digits)
    if len(digits) == 14:
        return is_valid_cnpj(digits)
    return False


def safe_folder_name(name: str) -> str:
    """
    Sanitiza nome de pasta removendo caracteres proibidos e espacos extras.
    """
    invalid = '<>:"/\\|?*'
    cleaned = "".join(c for c in (name or "").strip() if c not in invalid)
    cleaned = cleaned.strip().rstrip(". ")
    return cleaned or "sem_nome"


DEFAULT_FOLDER_STRUCTURE: Dict[str, Any] = {
    "year": False,
    "month": False,
    "day": False,
    "usar_pasta_cliente": False,
    "pastas_cliente": [],
    "nome_pasta_cliente": "",
}


def normalizar_estrutura_pastas(raw: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Normaliza a configuracao de estrutura de pastas, garantindo chaves esperadas.
    """
    cfg = DEFAULT_FOLDER_STRUCTURE.copy()

    def _as_bool(val: Any, default: bool) -> bool:
        if isinstance(val, str):
            v = val.strip().lower()
            if v in {"1", "true", "sim", "yes", "y"}:
                return True
            if v in {"0", "false", "nao", "não", "no", "n"}:
                return False
        if val is None:
            return default
        return bool(val)

    if isinstance(raw, dict):
        cfg["year"] = _as_bool(raw.get("year"), cfg["year"])
        cfg["month"] = _as_bool(raw.get("month"), cfg["month"])
        cfg["day"] = _as_bool(raw.get("day"), cfg["day"])
        cfg["usar_pasta_cliente"] = _as_bool(raw.get("usar_pasta_cliente"), cfg["usar_pasta_cliente"])

        pastas: List[str] = []
        raw_pastas = raw.get("pastas_cliente", None)
        if isinstance(raw_pastas, list):
            for it in raw_pastas:
                s = str(it or "").strip()
                if s:
                    pastas.append(s)
        elif isinstance(raw_pastas, str):
            for p in re.split(r"[;|\n]+", raw_pastas):
                s = (p or "").strip()
                if s:
                    pastas.append(s)

        legacy = str(raw.get("nome_pasta_cliente", "") or "").strip()
        if legacy:
            parts = re.split(r"[\\/]+", legacy) if ("\\" in legacy or "/" in legacy) else [legacy]
            parts = [x.strip() for x in parts if x and x.strip()]
            for p in reversed(parts):
                if not any((x or "").strip().lower() == p.lower() for x in pastas):
                    pastas.insert(0, p)

        cfg["pastas_cliente"] = pastas
        cfg["nome_pasta_cliente"] = pastas[0] if pastas else ""

    return cfg


def load_companies() -> List[Dict[str, str]]:
    if COMPANIES_JSON.exists():
        try:
            data = json.load(open(COMPANIES_JSON, "r", encoding="utf-8"))
            if isinstance(data, list):
                norm = []
                for entry in data:
                    if isinstance(entry, dict):
                        auth_mode = entry.get("auth_mode") or (
                            AUTH_CERTIFICATE if entry.get("cert_blob_b64") else AUTH_PASSWORD
                        )
                        if auth_mode not in (AUTH_PASSWORD, AUTH_CERTIFICATE):
                            auth_mode = AUTH_PASSWORD
                        norm.append(
                            {
                                "name": entry.get("name", "").strip(),
                                "doc": only_digits(entry.get("doc", "")),
                                "password": entry.get("password", ""),
                                "auth_mode": auth_mode,
                                "cert_path": entry.get("cert_path", "") or "",
                                "cert_password": entry.get("cert_password", "") or "",
                                "cert_blob_b64": entry.get("cert_blob_b64", "") or "",
                            }
                        )
                return norm
        except Exception:
            pass
    return []


def load_companies_from_supabase(supabase_url: str, supabase_anon_key: str) -> List[Dict[str, str]]:
    """
    Carrega empresas para o Robô NFS: primeiro as marcadas em company_robot_config (enabled);
    se nenhuma, usa fallback com empresas ativas do dashboard (auth/cert da empresa).
    """
    if not (supabase_url and supabase_anon_key):
        return []
    try:
        client = create_client(supabase_url.strip(), supabase_anon_key.strip())
        r = client.table("company_robot_config").select(
            "auth_mode,nfs_password,companies(id,name,document,auth_mode,cert_blob_b64,cert_password)"
        ).eq("robot_technical_id", ROBOT_TECHNICAL_ID).eq("enabled", True).execute()
        if getattr(r, "data", None):
            norm: List[Dict[str, str]] = []
            for row in r.data:
                company = row.get("companies")
                if not company:
                    continue
                name = (company.get("name") or "").strip()
                doc = only_digits(company.get("document") or "")
                cfg_auth = (row.get("auth_mode") or "password").strip().lower()
                if cfg_auth not in ("password", "certificate"):
                    cfg_auth = "password"
                if cfg_auth == "certificate":
                    auth_mode = AUTH_CERTIFICATE
                    password = ""
                    cert_password = (company.get("cert_password") or "")
                    cert_blob_b64 = (company.get("cert_blob_b64") or "")
                else:
                    auth_mode = AUTH_PASSWORD
                    password = (row.get("nfs_password") or "").strip()
                    cert_password = ""
                    cert_blob_b64 = ""
                norm.append({
                    "id": company.get("id"),
                    "name": name,
                    "doc": doc,
                    "password": password,
                    "auth_mode": auth_mode,
                    "cert_path": "",
                    "cert_password": cert_password,
                    "cert_blob_b64": cert_blob_b64,
                })
            norm.sort(key=lambda x: (x.get("name") or "").lower())
            return norm
        # Fallback: nenhuma empresa em company_robot_config — carrega ativas do dashboard (auth da empresa)
        comps = client.table("companies").select(
            "id,name,document,auth_mode,cert_blob_b64,cert_password"
        ).eq("active", True).order("name").execute()
        if not getattr(comps, "data", None):
            return []
        norm = []
        for row in comps.data:
            cid = row.get("id")
            name = (row.get("name") or "").strip()
            doc = only_digits(row.get("document") or "")
            company_auth = (row.get("auth_mode") or "password").strip().lower()
            if company_auth not in ("password", "certificate"):
                company_auth = "password"
            if company_auth == "certificate" and row.get("cert_blob_b64"):
                auth_mode = AUTH_CERTIFICATE
                password = ""
                cert_password = (row.get("cert_password") or "")
                cert_blob_b64 = (row.get("cert_blob_b64") or "")
            else:
                auth_mode = AUTH_PASSWORD
                password = ""
                cert_password = ""
                cert_blob_b64 = ""
            norm.append({
                "id": cid,
                "name": name,
                "doc": doc,
                "password": password,
                "auth_mode": auth_mode,
                "cert_path": "",
                "cert_password": cert_password,
                "cert_blob_b64": cert_blob_b64,
            })
        norm.sort(key=lambda x: (x.get("name") or "").lower())
        return norm
    except Exception:
        return []


def load_companies_from_supabase_by_ids(
    supabase_url: str, supabase_anon_key: str, company_ids: List[str]
) -> List[Dict[str, str]]:
    """
    Carrega empresas do Supabase pelos IDs (para execução via agendador).
    Usa company_robot_config quando existir; senão usa auth_mode/cert da própria empresa (companies).
    """
    if company_ids:
        snapshot_rows = JSON_RUNTIME.load_job_companies(_current_json_job(), company_ids)
        if snapshot_rows:
            normalized_rows: List[Dict[str, str]] = []
            for row in snapshot_rows:
                auth_mode = str(row.get("auth_mode") or AUTH_PASSWORD).strip().lower()
                if auth_mode not in (AUTH_PASSWORD, AUTH_CERTIFICATE):
                    auth_mode = AUTH_PASSWORD
                normalized_rows.append(
                    {
                        "id": str(row.get("company_id") or row.get("id") or ""),
                        "name": str(row.get("name") or "").strip(),
                        "doc": only_digits(row.get("document") or row.get("doc") or ""),
                        "password": str(row.get("password") or row.get("nfs_password") or "").strip(),
                        "auth_mode": auth_mode,
                        "cert_path": "",
                        "cert_password": str(row.get("cert_password") or "").strip(),
                        "cert_blob_b64": str(row.get("cert_blob_b64") or ""),
                    }
                )
            return normalized_rows
    if not (supabase_url and supabase_anon_key and company_ids):
        return []
    try:
        client = create_client(supabase_url.strip(), supabase_anon_key.strip())
        configs = client.table("company_robot_config").select(
            "company_id,auth_mode,nfs_password"
        ).eq("robot_technical_id", ROBOT_TECHNICAL_ID).in_("company_id", company_ids).execute()
        config_by_company: Dict[str, Dict[str, Any]] = {}
        if getattr(configs, "data", None):
            for c in configs.data:
                config_by_company[c["company_id"]] = c
        r = client.table("companies").select(
            "id,name,document,auth_mode,cert_blob_b64,cert_password"
        ).in_("id", company_ids).eq("active", True).execute()
        if not getattr(r, "data", None):
            return []
        norm = []
        for row in r.data:
            cid = row.get("id")
            cfg = config_by_company.get(cid) if cid else None
            name = (row.get("name") or "").strip()
            doc = only_digits(row.get("document") or "")
            if cfg:
                cfg_auth = (cfg.get("auth_mode") or "password").strip().lower()
                if cfg_auth not in ("password", "certificate"):
                    cfg_auth = "password"
                if cfg_auth == "certificate":
                    auth_mode = AUTH_CERTIFICATE
                    password = ""
                    cert_password = (row.get("cert_password") or "")
                    cert_blob_b64 = (row.get("cert_blob_b64") or "")
                else:
                    auth_mode = AUTH_PASSWORD
                    password = (cfg.get("nfs_password") or "").strip()
                    cert_password = ""
                    cert_blob_b64 = ""
            else:
                # Fallback: sem company_robot_config, usa auth da empresa (companies)
                company_auth = (row.get("auth_mode") or "password").strip().lower()
                if company_auth not in ("password", "certificate"):
                    company_auth = "password"
                if company_auth == "certificate" and row.get("cert_blob_b64"):
                    auth_mode = AUTH_CERTIFICATE
                    password = ""
                    cert_password = (row.get("cert_password") or "")
                    cert_blob_b64 = (row.get("cert_blob_b64") or "")
                else:
                    auth_mode = AUTH_PASSWORD
                    password = ""
                    cert_password = ""
                    cert_blob_b64 = ""
            norm.append({
                "id": cid,
                "name": name,
                "doc": doc,
                "password": password,
                "auth_mode": auth_mode,
                "cert_path": "",
                "cert_password": cert_password,
                "cert_blob_b64": cert_blob_b64,
            })
        return norm
    except Exception:
        return []


def save_companies(companies: List[Dict[str, str]]) -> None:
    COMPANIES_JSON.parent.mkdir(exist_ok=True)
    clean: List[Dict[str, str]] = []
    for entry in companies:
        clean.append(
            {
                "name": entry.get("name", ""),
                "doc": only_digits(entry.get("doc", "")),
                "password": entry.get("password", ""),
                "auth_mode": entry.get("auth_mode", AUTH_PASSWORD),
                "cert_password": entry.get("cert_password", ""),
                "cert_blob_b64": entry.get("cert_blob_b64", ""),
            }
        )
    with open(COMPANIES_JSON, "w", encoding="utf-8") as f:
        json.dump(clean, f, indent=2, ensure_ascii=False)


def _read_output_config() -> Dict[str, Any]:
    if OUTPUT_PATH_JSON.exists():
        try:
            with open(OUTPUT_PATH_JSON, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                return data
        except Exception:
            pass
    return {}


def _write_output_config(payload: Dict[str, Any]) -> None:
    OUTPUT_PATH_JSON.parent.mkdir(exist_ok=True)
    with open(OUTPUT_PATH_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)


def load_paths() -> Tuple[Optional[Path], Optional[Path]]:
    data = _read_output_config()
    base = Path(str(data["output_base"])) if data.get("output_base") else None
    report = Path(str(data["report_path"])) if data.get("report_path") else None
    return base, report or base


def load_path_preferences() -> Dict[str, Any]:
    data = _read_output_config()
    prefs = data.get("preferences")
    if isinstance(prefs, dict):
        return dict(prefs)
    return {}


def save_paths(output_base: Path, report_path: Optional[Path]) -> None:
    payload = _read_output_config()
    payload["output_base"] = str(output_base)
    if report_path:
        payload["report_path"] = str(report_path)
    else:
        payload.pop("report_path", None)
    _write_output_config(payload)


def save_path_preferences(preferences: Dict[str, Any]) -> None:
    payload = _read_output_config()
    payload["preferences"] = dict(preferences)
    _write_output_config(payload)


def load_output_base_path() -> Optional[Path]:
    """
    Compat: retorna apenas a pasta de XML.
    """
    base, _ = load_paths()
    return base


def save_output_base_path(path: Path) -> None:
    """
    Compat: salva a pasta de XML preservando a de relatorio (se houver).
    """
    _, report = load_paths()
    save_paths(path, report)


def fetch_robot_config_from_api() -> Optional[Dict[str, Any]]:
    """
    Obtém base_path, segment_path, date_rule e notes_mode do server-api (config global no dashboard).
    GET {SERVER_API_URL}/api/robot-config?technical_id=xxx
    Retorna dict com base_path, segment_path, date_rule, notes_mode, folder_structure; ou None se falhar.
    """
    url_base = (os.environ.get("FOLDER_STRUCTURE_API_URL") or os.environ.get("SERVER_API_URL") or "").strip().rstrip("/")
    if not url_base:
        return None
    try:
        url = f"{url_base}/api/robot-config"
        params = {"technical_id": ROBOT_TECHNICAL_ID}
        headers = build_server_api_headers(url_base)
        r = requests.get(url, params=params, headers=headers, timeout=15)
        r.raise_for_status()
        return r.json()
    except Exception:
        return None


# Cache da config da API (base_path, segment_path, date_rule) para não depender do .env
_robot_api_config: Optional[Dict[str, Any]] = None


def get_robot_api_config(force_refresh: bool = False) -> Optional[Dict[str, Any]]:
    """Retorna a config do robô vinda da API (cache). Permite refresh explícito."""
    global _robot_api_config
    if force_refresh or _robot_api_config is None:
        _robot_api_config = fetch_robot_config_from_api()
    return _robot_api_config


def get_resolved_output_base() -> Optional[Path]:
    """
    Pasta base para gravação: primeiro tenta base_path da API (dashboard); senão BASE_PATH do .env.
    """
    cfg = get_robot_api_config()
    if (not cfg or not (cfg.get("base_path") or "").strip()):
        cfg = get_robot_api_config(force_refresh=True)
    if cfg and (cfg.get("base_path") or "").strip():
        return Path((cfg["base_path"] or "").strip())
    base_env = os.environ.get("BASE_PATH", "").strip()
    if base_env:
        return Path(base_env)
    return None


def fetch_central_folder_structure(path_logical: Optional[str] = None) -> Tuple[Optional[List[str]], Optional[str]]:
    """
    Obtém da API a estrutura de pastas e retorna (segmentos para este robô, date_rule).
    path_logical: caminho do departamento (ex. FISCAL/NFS); se None, usa config da API ou ROBOT_SEGMENT_PATH do .env.
    """
    cfg = get_robot_api_config()
    if (not cfg or not (cfg.get("segment_path") or "").strip()):
        cfg = get_robot_api_config(force_refresh=True)
    path_logical = path_logical or (cfg or {}).get("segment_path") or os.environ.get("ROBOT_SEGMENT_PATH", "FISCAL/NFS").strip()
    if not path_logical:
        path_logical = "FISCAL/NFS"
    url_base = (os.environ.get("FOLDER_STRUCTURE_API_URL") or os.environ.get("SERVER_API_URL") or "").strip().rstrip("/")
    if not url_base:
        return (None, None)
    # Se já temos config da API com date_rule e folder_structure, podemos montar os segmentos a partir do path_logical
    cfg = cfg or get_robot_api_config()
    if cfg and isinstance(cfg.get("folder_structure"), list) and cfg.get("date_rule") is not None:
        nodes = cfg["folder_structure"]
        parts = [p.strip() for p in path_logical.split("/") if p.strip()]
        if not parts:
            return (None, None)
        segment_path: List[str] = []
        date_rule: Optional[str] = cfg.get("date_rule")
        parent_id: Optional[str] = None
        for part in parts:
            found = None
            for n in nodes:
                n_pid = n.get("parent_id")
                name_match = (n.get("slug") or n.get("name") or "").strip().upper() == part.upper()
                parent_ok = (n_pid or None) == (parent_id or None)
                if parent_ok and name_match:
                    found = n
                    break
            if not found:
                break
            seg = (found.get("slug") or found.get("name") or "").strip()
            if seg:
                segment_path.append(seg)
            parent_id = found["id"]
        if segment_path:
            return (segment_path, date_rule)
    # Fallback: chamar /api/folder-structure como antes
    try:
        r = requests.get(
            f"{url_base}/api/folder-structure",
            headers=build_server_api_headers(url_base),
            timeout=10,
        )
        r.raise_for_status()
        data = r.json()
        nodes = data.get("nodes") or []
    except Exception:
        return (None, None)
    if not nodes:
        return (None, None)
    parts = [p.strip() for p in path_logical.split("/") if p.strip()]
    if not parts:
        return (None, None)
    segment_path = []
    date_rule = None
    parent_id = None
    for part in parts:
        found = None
        for n in nodes:
            n_pid = n.get("parent_id")
            name_match = (n.get("slug") or n.get("name") or "").strip().upper() == part.upper()
            parent_ok = (n_pid or None) == (parent_id or None)
            if parent_ok and name_match:
                found = n
                break
        if not found:
            return (None, None)
        seg = (found.get("slug") or found.get("name") or "").strip()
        if seg:
            segment_path.append(seg)
        date_rule = found.get("date_rule")
        parent_id = found["id"]
    return (segment_path if segment_path else None, date_rule)


def get_robot_supabase(preferences: Optional[Dict[str, Any]] = None) -> Tuple[Optional[str], Optional[str]]:
    """Retorna (url, anon_key) do Supabase para o robô: apenas .env (config só no dashboard)."""
    url = os.environ.get("SUPABASE_URL", "").strip()
    key = _get_supabase_service_role_key()
    if url and key:
        return (url, key)
    return (None, None)


def fetch_robot_config(supabase_url: str, supabase_anon_key: str) -> Optional[Dict[str, Any]]:
    """Retorna segment_path e notes_mode do robô a partir da tabela robots no Supabase."""
    try:
        client = create_client(supabase_url.strip(), supabase_anon_key.strip())
        r = client.table("robots").select("segment_path, notes_mode").eq("technical_id", ROBOT_TECHNICAL_ID).limit(1).execute()
        rows = getattr(r, "data", None) or []
        if not rows:
            return None
        row = rows[0]
        seg = (row.get("segment_path") or "").strip() or None
        mode = (row.get("notes_mode") or "").strip() or None
        return {"segment_path": seg, "notes_mode": mode}
    except Exception:
        return None


def register_robot(supabase_url: str, supabase_anon_key: str) -> Optional[str]:
    try:
        return JSON_RUNTIME.register_robot()
    except Exception as e:
        print(f"[Robô] Falha ao iniciar runtime JSON: {e}", file=sys.stderr)
        return None


def update_robot_heartbeat(supabase_url: str, supabase_anon_key: str, robot_id: str) -> None:
    try:
        job = _current_json_job()
        JSON_RUNTIME.write_heartbeat(
            status="processing" if job else "active",
            current_job_id=(job or {}).get("job_id"),
            current_execution_request_id=(job or {}).get("execution_request_id"),
        )
    except Exception:
        pass


def update_robot_status(
    supabase_url: str, supabase_anon_key: str, robot_id: str, status: str
) -> None:
    try:
        job = _current_json_job()
        JSON_RUNTIME.write_heartbeat(
            status=status,
            current_job_id=(job or {}).get("job_id"),
            current_execution_request_id=(job or {}).get("execution_request_id"),
        )
    except Exception:
        pass


def update_robot_last_period_end(
    supabase_url: str, supabase_anon_key: str, robot_id: str, period_end: Optional[str]
) -> None:
    return


def _get_active_schedule_rule_ids(
    client: Any,
) -> Optional[list]:
    """Retorna lista de ids de schedule_rules ativas (status=active, run_daily=true). None em erro."""
    try:
        r = (
            client.table("schedule_rules")
            .select("id")
            .eq("status", "active")
            .eq("run_daily", True)
            .execute()
        )
        rows = getattr(r, "data", None) or []
        return [row["id"] for row in rows if row.get("id")]
    except Exception:
        return None


def claim_execution_request(
    supabase_url: str,
    supabase_anon_key: str,
    robot_id: str,
    log_callback: Optional[Callable[[str], None]] = None,
) -> Optional[Dict[str, Any]]:
    try:
        job = JSON_RUNTIME.load_job()
        if not job:
            return None
        _set_current_json_job(job)
        _clear_pending_result_operations()
        JSON_RUNTIME.write_heartbeat(
            status="processing",
            current_job_id=job.get("job_id"),
            current_execution_request_id=job.get("execution_request_id"),
            message="job_loaded",
        )
        return job
    except Exception as e:
        msg = f"[Robô] Erro ao buscar job da fila (agendador): {e}"
        if log_callback:
            log_callback(msg)
        print(msg, file=sys.stderr)
        return None


def complete_execution_request(
    supabase_url: str, supabase_anon_key: str, request_id: str, success: bool, error_message: Optional[str] = None
) -> None:
    try:
        job = _current_json_job()
        summary = _consume_last_json_result_summary()
        operations = _consume_result_operations()
        JSON_RUNTIME.write_result(
            job=job if isinstance(job, dict) and str(job.get("execution_request_id") or job.get("id") or "") == str(request_id) else {"execution_request_id": request_id, "job_id": request_id},
            success=success,
            error_message=error_message,
            summary=summary,
            payload={"operations": operations},
        )
        _set_current_json_job(None)
    except Exception:
        pass


def _service_codes_ranking(section: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Retorna lista de códigos de serviço da seção (emitidas ou recebidas) ordenada por total_value desc."""
    items = (section or {}).get("service_codes") or []
    result: List[Dict[str, Any]] = []
    for item in items:
        code = str(item.get("code", "")).strip()
        desc = str(item.get("description", "")).strip()
        try:
            val = float(item.get("total_value") or 0)
        except (TypeError, ValueError):
            val = 0.0
        result.append({"code": code, "description": desc, "total_value": round(val, 2)})
    result.sort(key=lambda it: (-(it["total_value"] or 0), str(it.get("code", "")), str(it.get("description", ""))))
    return result


def _merge_service_codes_for_stats(emitidas: Dict[str, Any], recebidas: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Junta service_codes de emitidas e recebidas, somando total_value por (code, description)."""
    index: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for section in (emitidas or {}, recebidas or {}):
        for item in section.get("service_codes") or []:
            code = str(item.get("code", "")).strip()
            desc = str(item.get("description", "")).strip()
            key = (code, desc)
            if key not in index:
                index[key] = {"code": code, "description": desc, "total_value": 0.0}
            try:
                index[key]["total_value"] += float(item.get("total_value") or 0)
            except (TypeError, ValueError):
                pass
    result = list(index.values())
    result.sort(key=lambda it: (only_digits(str(it.get("code", "")).zfill(6)), str(it.get("description", ""))))
    return result


def upsert_nfs_stats(
    supabase_url: str,
    supabase_anon_key: str,
    job: Dict[str, Any],
    summary: Dict[str, Any],
) -> None:
    """Envia totais e ranking de códigos de serviço NFS para o Supabase (nfs_stats)."""
    companies_summary = summary.get("companies") or []
    period_start_raw = job.get("period_start")
    if period_start_raw is None:
        return
    if hasattr(period_start_raw, "isoformat"):
        period_start = period_start_raw.isoformat()[:10]
    else:
        period_start = (str(period_start_raw).strip().split("T")[0].split(" ")[0]) if period_start_raw else ""
    if not period_start or len(period_start) < 7:
        return
    period = period_start[:7]
    if not re.match(r"^\d{4}-\d{2}$", period):
        return
    rows: List[Dict[str, Any]] = []
    for comp in companies_summary:
        company_id = str(comp.get("company_id") or "").strip()
        if not company_id:
            continue
        emitidas = comp.get("emitidas") or {}
        recebidas = comp.get("recebidas") or {}
        qty_emitidas = int(emitidas.get("downloaded") or 0)
        qty_recebidas = int(recebidas.get("downloaded") or 0)
        try:
            valor_emitidas = float(emitidas.get("total_value") or 0)
        except (TypeError, ValueError):
            valor_emitidas = 0.0
        try:
            valor_recebidas = float(recebidas.get("total_value") or 0)
        except (TypeError, ValueError):
            valor_recebidas = 0.0
        service_codes = _merge_service_codes_for_stats(emitidas, recebidas)
        service_codes_emitidas = _service_codes_ranking(emitidas)
        service_codes_recebidas = _service_codes_ranking(recebidas)
        rows.append({
            "company_id": str(company_id),
            "period": period,
            "qty_emitidas": qty_emitidas,
            "qty_recebidas": qty_recebidas,
            "valor_emitidas": round(valor_emitidas, 2),
            "valor_recebidas": round(valor_recebidas, 2),
            "service_codes": service_codes,
            "service_codes_emitidas": service_codes_emitidas,
            "service_codes_recebidas": service_codes_recebidas,
        })
    if not rows:
        return
    if _current_json_job():
        _append_result_operation(
            {
                "kind": "upsert_rows",
                "table": "nfs_stats",
                "on_conflict": "office_id,company_id,period",
                "rows": rows,
            }
        )
        return
    try:
        client = create_client(supabase_url.strip(), supabase_anon_key.strip())
    except Exception:
        raise
    client.rpc("nfs_stats_upsert_batch", {"rows": rows}).execute()


def fetch_robot_display_config(supabase_url: str, supabase_anon_key: str) -> Optional[Dict[str, Any]]:
    """Retorna a config de exibição do robô (empresas, período, modo) para sincronizar com o dashboard."""
    try:
        client = create_client(supabase_url.strip(), supabase_anon_key.strip())
        office_id = ""
        current_job = _current_json_job()
        if isinstance(current_job, dict):
            office_id = str(current_job.get("office_id") or "").strip()
        if not office_id:
            try:
                office_id = str((_bridge_get_office_context(client) or {}).get("office_id") or "").strip()
            except Exception:
                office_id = ""
        query = client.table("robot_display_config").select("*").eq("robot_technical_id", ROBOT_TECHNICAL_ID)
        if office_id:
            query = query.eq("office_id", office_id)
        r = query.limit(1).execute()
        rows = getattr(r, "data", None) or []
        if not rows:
            return None
        return rows[0]
    except Exception:
        return None


def friendly_path_display(path: Path, keep_parts: int = 2) -> str:
    """
    Retorna uma versao encurtada do caminho, exibindo apenas os ultimos segmentos.
    """
    try:
        p = Path(path)
    except TypeError:
        return str(path)

    parts = p.parts
    if not parts:
        return str(p)

    drive = ""
    tail_parts = list(parts)
    if os.name == "nt" and tail_parts and tail_parts[0].endswith(":\\"):
        drive = tail_parts.pop(0)

    if len(tail_parts) <= keep_parts:
        tail = Path(*tail_parts)
        return f"{drive}{tail}" if drive else str(p)

    tail = Path(*tail_parts[-keep_parts:])
    prefix = f"{drive}..." if drive else "..."
    return f"{prefix}{os.sep}{tail}"


# --------------------------------------------------------------------
# Auto update (GitHub Releases)
# --------------------------------------------------------------------


def _updater_log(msg: str) -> None:
    try:
        UPDATER_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(UPDATER_LOG_PATH, "a", encoding="utf-8") as f:
            stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            f.write(f"[{stamp}] {msg}\n")
    except Exception:
        pass


def _read_json_dict(path: Path) -> Dict[str, Any]:
    if path.exists():
        try:
            data = json.load(open(path, "r", encoding="utf-8"))
            if isinstance(data, dict):
                return data
        except Exception:
            return {}
    return {}


def _save_json_dict(path: Path, payload: Dict[str, Any]) -> None:
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
    except Exception:
        pass


def _coerce_bool(value: Any, default: bool) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        v = value.strip().lower()
        if v in {"1", "true", "yes", "y", "sim"}:
            return True
        if v in {"0", "false", "no", "n", "nao", "não"}:
            return False
    return default


def load_updater_settings() -> Dict[str, Any]:
    cfg = dict(DEFAULT_UPDATER_SETTINGS)
    raw = _read_json_dict(UPDATER_SETTINGS_PATH)
    for key in cfg.keys():
        if key in raw:
            cfg[key] = raw[key]

    cfg["enabled"] = _coerce_bool(cfg.get("enabled"), True)
    cfg["check_on_startup"] = _coerce_bool(cfg.get("check_on_startup"), True)
    cfg["allow_prerelease"] = _coerce_bool(cfg.get("allow_prerelease"), False)
    cfg["github_repo"] = str(cfg.get("github_repo") or "").strip().strip("/")
    cfg["asset_name"] = str(cfg.get("asset_name") or "").strip()
    cfg["current_version"] = str(cfg.get("current_version") or AUTO_UPDATE_APP_VERSION).strip() or AUTO_UPDATE_APP_VERSION

    if not UPDATER_SETTINGS_PATH.exists():
        _save_json_dict(UPDATER_SETTINGS_PATH, cfg)

    return cfg


def _version_tuple(version_text: str) -> Tuple[int, ...]:
    nums = re.findall(r"\d+", str(version_text or ""))
    if not nums:
        return (0,)
    return tuple(int(n) for n in nums[:4])


def _is_remote_version_newer(remote_version: str, local_version: str) -> bool:
    return _version_tuple(remote_version) > _version_tuple(local_version)


def _fetch_github_release(repo: str, allow_prerelease: bool) -> Optional[Dict[str, Any]]:
    headers = {
        "Accept": "application/vnd.github+json",
        "User-Agent": f"nfs-updater/{AUTO_UPDATE_APP_VERSION}",
    }
    timeout = (4, 8)

    if allow_prerelease:
        url = f"https://api.github.com/repos/{repo}/releases?per_page=10"
        resp = requests.get(url, headers=headers, timeout=timeout)
        resp.raise_for_status()
        payload = resp.json()
        if not isinstance(payload, list):
            return None
        for rel in payload:
            if not isinstance(rel, dict):
                continue
            if rel.get("draft"):
                continue
            return rel
        return None

    url = f"https://api.github.com/repos/{repo}/releases/latest"
    resp = requests.get(url, headers=headers, timeout=timeout)
    resp.raise_for_status()
    payload = resp.json()
    return payload if isinstance(payload, dict) else None


def _select_release_asset(release_data: Dict[str, Any], preferred_name: str) -> Optional[Dict[str, Any]]:
    assets = release_data.get("assets")
    if not isinstance(assets, list):
        return None

    cleaned_pref = (preferred_name or "").strip().lower()
    if cleaned_pref:
        for asset in assets:
            if not isinstance(asset, dict):
                continue
            name = str(asset.get("name") or "").strip().lower()
            if name == cleaned_pref:
                return asset
        return None

    for asset in assets:
        if not isinstance(asset, dict):
            continue
        name = str(asset.get("name") or "").strip().lower()
        if name.endswith(".zip"):
            return asset

    for asset in assets:
        if isinstance(asset, dict):
            return asset
    return None


def _download_file(download_url: str, target_path: Path) -> str:
    sha256 = hashlib.sha256()
    with requests.get(download_url, stream=True, timeout=(8, 120)) as resp:
        resp.raise_for_status()
        with open(target_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=1024 * 512):
                if not chunk:
                    continue
                f.write(chunk)
                sha256.update(chunk)
    return sha256.hexdigest()


def _find_payload_root(extract_dir: Path, exe_name: str) -> Optional[Path]:
    root_has_internal = (extract_dir / "_internal").is_dir()
    root_has_exe = (extract_dir / exe_name).is_file() or bool(list(extract_dir.glob("*.exe")))
    if root_has_internal and root_has_exe:
        return extract_dir

    subdirs = [p for p in extract_dir.iterdir() if p.is_dir()]
    for sub in subdirs:
        has_internal = (sub / "_internal").is_dir()
        has_exe = (sub / exe_name).is_file() or bool(list(sub.glob("*.exe")))
        if has_internal and has_exe:
            return sub
    return None


def _ps_quote(text: str) -> str:
    return "'" + str(text).replace("'", "''") + "'"


def _build_update_script(
    app_dir: Path,
    exe_path: Path,
    source_dir: Path,
    temp_root: Path,
    pid_to_wait: int,
) -> str:
    files_arr = ", ".join(_ps_quote(x) for x in UPDATER_JSON_FILES_TO_PRESERVE)
    return f"""$ErrorActionPreference = 'SilentlyContinue'
$appDir = {_ps_quote(str(app_dir))}
$exePath = {_ps_quote(str(exe_path))}
$sourceDir = {_ps_quote(str(source_dir))}
$tempRoot = {_ps_quote(str(temp_root))}
$pidToWait = {pid_to_wait}
$jsonFiles = @({files_arr})
$jsonDir = Join-Path $appDir '_internal\\data\\json'
$backupDir = Join-Path $tempRoot 'json_backup'

try {{
    New-Item -Path $backupDir -ItemType Directory -Force | Out-Null
    foreach ($file in $jsonFiles) {{
        $src = Join-Path $jsonDir $file
        if (Test-Path -LiteralPath $src) {{
            Copy-Item -LiteralPath $src -Destination (Join-Path $backupDir $file) -Force
        }}
    }}

    for ($i = 0; $i -lt 240; $i++) {{
        if (-not (Get-Process -Id $pidToWait -ErrorAction SilentlyContinue)) {{
            break
        }}
        Start-Sleep -Milliseconds 500
    }}

    Get-ChildItem -LiteralPath $sourceDir | ForEach-Object {{
        $dest = Join-Path $appDir $_.Name
        Copy-Item -LiteralPath $_.FullName -Destination $dest -Recurse -Force
    }}

    New-Item -Path $jsonDir -ItemType Directory -Force | Out-Null
    foreach ($file in $jsonFiles) {{
        $bk = Join-Path $backupDir $file
        if (Test-Path -LiteralPath $bk) {{
            Copy-Item -LiteralPath $bk -Destination (Join-Path $jsonDir $file) -Force
        }}
    }}
}} catch {{
}}

Start-Process -FilePath $exePath -WorkingDirectory $appDir -ArgumentList '--skip-startup-update'
Start-Sleep -Milliseconds 900
Remove-Item -LiteralPath $tempRoot -Recurse -Force -ErrorAction SilentlyContinue
"""


def apply_startup_update_if_available() -> bool:
    if os.name != "nt":
        return False
    if not getattr(sys, "frozen", False):
        return False
    if "--skip-startup-update" in sys.argv:
        return False
    if os.environ.get("NFS_SKIP_AUTOUPDATE", "").strip() == "1":
        return False

    settings = load_updater_settings()
    if not settings.get("enabled", True):
        return False
    if not settings.get("check_on_startup", True):
        return False

    repo = str(settings.get("github_repo") or "").strip().strip("/")
    if not repo:
        _updater_log("Auto update ignorado: github_repo vazio.")
        return False
    if repo.upper().startswith("SEU_"):
        _updater_log("Auto update ignorado: github_repo com placeholder.")
        return False

    preferred_asset = str(settings.get("asset_name") or "").strip()
    allow_prerelease = bool(settings.get("allow_prerelease"))

    try:
        release_data = _fetch_github_release(repo, allow_prerelease=allow_prerelease)
    except Exception as exc:  # noqa: BLE001
        _updater_log(f"Falha ao consultar release: {exc}")
        return False
    if not release_data:
        return False

    tag = str(release_data.get("tag_name") or "").strip()
    if not tag:
        _updater_log("Release sem tag_name.")
        return False
    remote_version = re.sub(r"^[vV]", "", tag)
    local_version = str(settings.get("current_version") or AUTO_UPDATE_APP_VERSION).strip() or AUTO_UPDATE_APP_VERSION
    if not _is_remote_version_newer(remote_version, local_version):
        return False

    asset = _select_release_asset(release_data, preferred_asset)
    if not asset:
        if preferred_asset:
            _updater_log(f"Asset '{preferred_asset}' nao encontrado na release {tag}.")
        else:
            _updater_log(f"Nenhum asset elegivel encontrado na release {tag}.")
        return False

    asset_name = str(asset.get("name") or "").strip()
    download_url = str(asset.get("browser_download_url") or "").strip()
    if not download_url:
        _updater_log(f"Asset sem URL de download: {asset_name or '(sem nome)'}")
        return False

    app_dir = Path(sys.executable).resolve().parent
    exe_path = Path(sys.executable).resolve()
    temp_root = Path(tempfile.mkdtemp(prefix="nfs_update_"))
    zip_path = temp_root / "release.zip"
    extract_dir = temp_root / "extract"
    extract_dir.mkdir(parents=True, exist_ok=True)

    _updater_log(
        f"Atualizacao detectada: local={local_version} remoto={remote_version} "
        f"asset={asset_name or '(sem nome)'}"
    )

    try:
        checksum = _download_file(download_url, zip_path)
        _updater_log(f"Pacote baixado ({zip_path.name}) sha256={checksum}")

        with ZipFile(zip_path, "r") as zf:
            zf.extractall(extract_dir)
    except BadZipFile:
        _updater_log("Falha na atualizacao: zip invalido.")
        shutil.rmtree(temp_root, ignore_errors=True)
        return False
    except Exception as exc:  # noqa: BLE001
        _updater_log(f"Falha no download/extracao: {exc}")
        shutil.rmtree(temp_root, ignore_errors=True)
        return False

    payload_root = _find_payload_root(extract_dir, exe_path.name)
    if payload_root is None:
        _updater_log("Falha na atualizacao: pacote sem .exe + _internal.")
        shutil.rmtree(temp_root, ignore_errors=True)
        return False

    script_path = temp_root / "apply_update.ps1"
    script = _build_update_script(app_dir, exe_path, payload_root, temp_root, os.getpid())
    try:
        script_path.write_text(script, encoding="utf-8-sig")
        creationflags = (
            getattr(subprocess, "DETACHED_PROCESS", 0)
            | getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)
            | getattr(subprocess, "CREATE_NO_WINDOW", 0)
        )
        subprocess.Popen(
            [
                "powershell",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-WindowStyle",
                "Hidden",
                "-File",
                str(script_path),
            ],
            creationflags=creationflags,
            close_fds=True,
        )
        _updater_log(f"Updater iniciado para versao {remote_version}.")
        return True
    except Exception as exc:  # noqa: BLE001
        _updater_log(f"Nao foi possivel iniciar updater: {exc}")
        shutil.rmtree(temp_root, ignore_errors=True)
        return False


# --------------------------------------------------------------------
# CNPJ lookup (rate limited)
# --------------------------------------------------------------------

CNPJ_LOOKUP_WINDOW = 20  # segundos
CNPJ_LOOKUP_MAX = 3
_CNPJ_LOOKUP_HISTORY: deque[float] = deque()
_CNPJ_LOOKUP_LOCK = threading.Lock()


def _cnpj_prune_history(now: float) -> None:
    with _CNPJ_LOOKUP_LOCK:
        while _CNPJ_LOOKUP_HISTORY and now - _CNPJ_LOOKUP_HISTORY[0] > CNPJ_LOOKUP_WINDOW:
            _CNPJ_LOOKUP_HISTORY.popleft()


def cnpj_wait_time() -> float:
    """
    Retorna segundos a aguardar para respeitar o limite (3 consultas a cada 20s).
    """
    now = time.monotonic()
    _cnpj_prune_history(now)
    with _CNPJ_LOOKUP_LOCK:
        if len(_CNPJ_LOOKUP_HISTORY) >= CNPJ_LOOKUP_MAX:
            return max(CNPJ_LOOKUP_WINDOW - (now - _CNPJ_LOOKUP_HISTORY[0]), 0.0)
    return 0.0


def cnpj_register_call() -> None:
    now = time.monotonic()
    _cnpj_prune_history(now)
    with _CNPJ_LOOKUP_LOCK:
        _CNPJ_LOOKUP_HISTORY.append(now)


# --------------------------------------------------------------------
# Licenca: armazenamento local + validacao Supabase
# --------------------------------------------------------------------


def _get_supabase_client():
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        raise RuntimeError("Supabase não configurado. Defina SUPABASE_URL e SUPABASE_SERVICE_ROLE_KEY.")
    return create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)


def _get_license_supabase_client():
    """Cliente Supabase usado apenas para validar licença (projeto separado)."""
    url = (os.environ.get("LICENSE_SUPABASE_URL") or os.environ.get("SUPABASE_URL") or "").strip()
    anon = (os.environ.get("LICENSE_SUPABASE_ANON_KEY") or os.environ.get("SUPABASE_ANON_KEY") or "").strip()
    if not url or not anon:
        raise RuntimeError(
            "Supabase (licenca) nao configurado. Defina LICENSE_SUPABASE_URL e LICENSE_SUPABASE_ANON_KEY."
        )
    return create_client(url, anon)


def _load_license_local() -> Dict[str, Any]:
    if LICENSE_PATH.exists():
        try:
            return json.load(open(LICENSE_PATH, "r", encoding="utf-8")) or {}
        except Exception:
            return {}
    return {}


def _save_license_local(data: Dict[str, Any]) -> None:
    LICENSE_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(LICENSE_PATH, "w", encoding="utf-8") as f:
        json.dump(data or {}, f, ensure_ascii=False, indent=2)


def get_saved_key() -> str:
    return str(_load_license_local().get("license_key", "")).strip()


def persist_key(key: str, meta: Optional[Dict[str, Any]] = None) -> None:
    payload = _load_license_local()
    payload["license_key"] = (key or "").strip()
    if meta:
        payload.update(meta)
    _save_license_local(payload)


def check_license_with_supabase(key: str) -> Tuple[bool, str, Dict[str, Any]]:
    key = (key or "").strip()
    if not key:
        return False, "Informe a chave de licenca.", {}
    try:
        client = _get_license_supabase_client()
        res = client.rpc("verify_license", {"p_key": key}).execute()
        rows = res.data if hasattr(res, "data") else []
    except Exception as exc:  # noqa: BLE001
        return False, f"Falha ao contatar o servidor de licenca: {exc}", {}

    if not rows:
        return False, "Chave invalida, expirada ou inativa.", {}

    row = rows[0] or {}
    meta = {"expires_at": row.get("expires_at"), "licensee": row.get("client_name") or ""}
    return True, "Licenca valida.", meta


def _warn_if_expiring_soon(parent=None) -> None:
    meta = _load_license_local()
    exp = str(meta.get("expires_at") or "").strip()
    if not exp:
        return
    try:
        if "T" in exp:
            exp_dt = datetime.fromisoformat(exp.replace("Z", "+00:00"))
        else:
            exp_dt = datetime.strptime(exp, "%Y-%m-%d")
        days = (exp_dt.date() - datetime.now(exp_dt.tzinfo or None).date()).days
        if days <= 7:
            QMessageBox.warning(
                parent,
                "Licenca perto de vencer",
                f"Sua licenca expira em {exp_dt.date().strftime('%d/%m/%Y')} (em {max(days,0)} dia(s)).",
            )
    except Exception:
        pass


# --------------------------------------------------------------------
# UI widgets
# --------------------------------------------------------------------


class LicenseDialog(QDialog):
    ACCENT = "#7C3AED"
    BG1 = "#0B1220"
    BG2 = "#111827"
    CARD = "#0F172A"
    STROKE = "rgba(255,255,255,0.08)"
    TEXT = "rgba(255,255,255,0.92)"
    MUTED = "rgba(255,255,255,0.65)"

    def __init__(self, parent=None, preset_key: str = ""):
        super().__init__(parent)
        self.setObjectName("LicenseDialog")
        self.setModal(True)
        self.setMinimumWidth(520)
        self.setWindowTitle("Ativacao da Licenca")
        try:
            if LOGO_PNG.exists():
                self.setWindowIcon(QIcon(str(LOGO_PNG)))
        except Exception:
            pass

        self.setStyleSheet(
            f"""
        QDialog#LicenseDialog {{
            background: qlineargradient(x1:0,y1:0, x2:1,y2:1,
                        stop:0 {self.BG1}, stop:1 {self.BG2});
        }}
        QLabel, QLineEdit, QPushButton {{
            color: {self.TEXT};
            font-family: "Segoe UI", "Inter", "Roboto", "Ubuntu", "Arial";
            font-size: 14px;
        }}
        QFrame#Card {{
            background: {self.CARD};
            border: 1px solid {self.STROKE};
            border-radius: 18px;
        }}
        QLabel#Title {{
            font-size: 22px;
            font-weight: 700;
            color: {self.TEXT};
        }}
        QLabel#Subtitle {{
            font-size: 13px;
            color: {self.MUTED};
        }}
        QLineEdit#KeyEdit {{
            background: #0B1220;
            border: 1px solid #243041;
            border-radius: 12px;
            padding: 12px 14px;
            selection-background-color: {self.ACCENT};
            selection-color: white;
        }}
        QLineEdit#KeyEdit:focus {{
            border: 1px solid {self.ACCENT};
            box-shadow: 0 0 0 3px rgba(124,58,237,0.25);
        }}
        QPushButton#PrimaryButton {{
            background: {self.ACCENT};
            border: none;
            border-radius: 12px;
            padding: 10px 16px;
            font-weight: 600;
            color: white;
        }}
        QPushButton#PrimaryButton:hover {{
            background: #6D28D9;
        }}
        QPushButton#PrimaryButton:disabled {{
            background: #5B21B6;
            opacity: 0.6;
        }}
        QPushButton#GhostButton {{
            background: transparent;
            border: 1px solid {self.STROKE};
            border-radius: 12px;
            padding: 10px 16px;
            color: {self.MUTED};
        }}
        QPushButton#GhostButton:hover {{
            border: 1px solid #2A374B;
            color: {self.TEXT};
        }}
        QPushButton#LinkButton {{
            background: transparent;
            border: none;
            color: {self.ACCENT};
            font-weight: 600;
        }}
        QPushButton#LinkButton:hover {{
            text-decoration: underline;
        }}
        QLabel#Msg {{
            color: {self.MUTED};
        }}
        """
        )

        root = QVBoxLayout(self)
        root.setContentsMargins(20, 20, 20, 20)

        card = QFrame(self)
        card.setObjectName("Card")
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(28, 28, 28, 28)
        card_layout.setSpacing(16)

        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(32)
        shadow.setColor(Qt.black)
        shadow.setOffset(0, 12)
        card.setGraphicsEffect(shadow)

        header = QHBoxLayout()
        header.setSpacing(14)

        logo_lbl = QLabel(card)
        logo_pix = None
        try:
            if LOGO_PNG.exists():
                logo_pix = QPixmap(str(LOGO_PNG))
        except Exception:
            pass
        if not logo_pix or logo_pix.isNull():
            logo_pix = QPixmap(64, 64)
            logo_pix.fill(Qt.transparent)
        logo_lbl.setPixmap(logo_pix.scaled(48, 48, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        header.addWidget(logo_lbl, 0, Qt.AlignTop)

        title_box = QVBoxLayout()
        title = QLabel("Ative sua licenca", card)
        title.setObjectName("Title")
        self.subtitle = QLabel("Digite sua chave para liberar todas as funcionalidades.", card)
        self.subtitle.setObjectName("Subtitle")
        self.subtitle.setWordWrap(True)
        title_box.addWidget(title)
        title_box.addWidget(self.subtitle)
        header.addLayout(title_box)
        header.addStretch(1)
        card_layout.addLayout(header)

        key_row = QHBoxLayout()
        key_row.setSpacing(10)

        self.edt = QLineEdit(card)
        self.edt.setObjectName("KeyEdit")
        self.edt.setPlaceholderText("XXXXX-XXXXX-XXXXX-XXXXX")
        if preset_key:
            self.edt.setText(preset_key)
        self.edt.returnPressed.connect(self._on_activate)

        paste_btn = QPushButton("Colar", card)
        paste_btn.setObjectName("GhostButton")
        paste_btn.setCursor(Qt.PointingHandCursor)
        paste_btn.clicked.connect(self._paste_from_clipboard)

        key_row.addWidget(self.edt, 1)
        key_row.addWidget(paste_btn, 0)
        card_layout.addLayout(key_row)

        self.msg = QLabel("", card)
        self.msg.setObjectName("Msg")
        self.msg.setWordWrap(True)
        card_layout.addWidget(self.msg)

        card_layout.addItem(QSpacerItem(0, 8, QSizePolicy.Minimum, QSizePolicy.Expanding))

        btns_row = QHBoxLayout()
        btns_row.addStretch(1)

        self.btn_cancel = QPushButton("Sair", card)
        self.btn_cancel.setObjectName("GhostButton")
        self.btn_cancel.setCursor(Qt.PointingHandCursor)
        self.btn_cancel.clicked.connect(self.reject)

        self.btn_ok = QPushButton("Ativar", card)
        self.btn_ok.setObjectName("PrimaryButton")
        self.btn_ok.setCursor(Qt.PointingHandCursor)
        self.btn_ok.clicked.connect(self._on_activate)

        btns_row.addWidget(self.btn_cancel)
        btns_row.addWidget(self.btn_ok)
        card_layout.addLayout(btns_row)

        help_row = QHBoxLayout()
        help_row.addStretch(1)
        help_lbl = QLabel(
            '<span style="color:rgba(255,255,255,0.55)">Precisa de ajuda?</span> '
            '<a style="color:#7C3AED; text-decoration:none;" href="https://wa.me/5562993626968"><b>Fale com o suporte</b></a>',
            card,
        )
        help_lbl.setTextFormat(Qt.RichText)
        help_lbl.setOpenExternalLinks(True)
        help_row.addWidget(help_lbl, 0, Qt.AlignRight)
        card_layout.addLayout(help_row)

        root.addWidget(card)

        self.setWindowOpacity(0.0)
        QTimer.singleShot(0, self._animate_entrance)

    def set_revoked_mode(self, reason: Optional[str] = None) -> None:
        base = "Sua licenca foi revogada. Entre em contato com o suporte para reativar."
        if reason:
            base = f"{base}\n{reason}"
        try:
            self.subtitle.setText(base)
        except Exception:
            pass
        try:
            self.msg.clear()
        except Exception:
            pass

    def _paste_from_clipboard(self) -> None:
        cb = QApplication.clipboard()
        if cb:
            self.edt.setText(cb.text().strip())

    def _animate_entrance(self) -> None:
        anim = QPropertyAnimation(self, b"windowOpacity", self)
        anim.setDuration(220)
        anim.setStartValue(0.0)
        anim.setEndValue(1.0)
        anim.setEasingCurve(QEasingCurve.OutCubic)
        anim.start(QPropertyAnimation.DeleteWhenStopped)

    def _on_activate(self) -> None:
        key = self.edt.text().strip()
        self._set_busy(True)
        ok, msg, meta = check_license_with_supabase(key)
        self._set_busy(False)
        self.msg.setText(msg or "")
        if ok:
            persist_key(key, meta)
            os.environ["AUTOMATIZE_VERIFIED"] = "1"
            self.accept()
            return
        low = (msg or "").lower()
        if "desativad" in low or "inativa" in low or "revog" in low:
            self.set_revoked_mode(msg)

    def _set_busy(self, busy: bool) -> None:
        self.btn_ok.setEnabled(not busy)
        self.btn_cancel.setEnabled(not busy)
        self.edt.setEnabled(not busy)


class ConfirmDialog(QDialog):
    def __init__(
        self,
        title: str = "Confirmacao",
        message: str = "Deseja continuar?",
        primary_text: str = "OK",
        secondary_text: str = "Cancelar",
        primary_colors: Optional[Tuple[str, str, str]] = None,
        secondary_colors: Optional[Tuple[str, str, str]] = None,
        parent=None,
    ):
        super().__init__(parent)
        self.setObjectName("ConfirmDialog")
        self.setModal(True)
        self.setMinimumWidth(520)
        self.setWindowTitle(title)
        try:
            if LOGO_ICON.exists():
                self.setWindowIcon(QIcon(str(LOGO_ICON)))
        except Exception:
            pass
        self.choice: Optional[str] = None
        self._primary_colors = primary_colors or ("#7C3AED", "#6D28D9", "#5B21B6")
        self._secondary_colors = secondary_colors or ("#2980B9", "#3498DB", "#2471A3")

        self.setStyleSheet(
            f"""
        QDialog#ConfirmDialog {{
            background: #0f1722;
        }}
        QLabel {{ color: #ECF0F1; font:10pt 'Verdana'; }}
        QFrame#Card {{ background:#162231; border:1px solid #22344a; border-radius:14px; }}
        QLabel#Title {{
            font-size: 18px;
            font-weight: 700;
            color: #ECF0F1;
        }}
        QLabel#Msg {{
            font-size: 12px;
            color: #95a5c8;
        }}
        """
        )

        root = QVBoxLayout(self)
        root.setContentsMargins(20, 20, 20, 20)

        card = QFrame(self)
        card.setObjectName("Card")
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(28, 28, 28, 28)
        card_layout.setSpacing(16)

        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(32)
        shadow.setColor(Qt.black)
        shadow.setOffset(0, 12)
        card.setGraphicsEffect(shadow)

        lbl_title = QLabel(title, card)
        lbl_title.setObjectName("Title")
        card_layout.addWidget(lbl_title)

        lbl_msg = QLabel(message, card)
        lbl_msg.setObjectName("Msg")
        lbl_msg.setWordWrap(True)
        card_layout.addWidget(lbl_msg)

        card_layout.addItem(QSpacerItem(0, 8, QSizePolicy.Minimum, QSizePolicy.Expanding))

        btns = QHBoxLayout()
        btns.addStretch(1)

        self.btn_secondary = QPushButton(secondary_text, card)
        self.btn_secondary.setCursor(Qt.PointingHandCursor)
        self.btn_secondary.setStyleSheet(button_style(*self._secondary_colors))
        self.btn_secondary.clicked.connect(self._on_secondary)

        self.btn_primary = QPushButton(primary_text, card)
        self.btn_primary.setCursor(Qt.PointingHandCursor)
        self.btn_primary.setStyleSheet(button_style(*self._primary_colors))
        self.btn_primary.clicked.connect(self._on_primary)

        attach_glow(self.btn_secondary, self._secondary_colors[1] if len(self._secondary_colors) > 1 else "#4da3ff")
        attach_glow(self.btn_primary, self._primary_colors[1] if len(self._primary_colors) > 1 else "#7C3AED")

        btns.addWidget(self.btn_secondary)
        btns.addWidget(self.btn_primary)
        card_layout.addLayout(btns)

        root.addWidget(card)

    def _on_primary(self) -> None:
        self.choice = "primary"
        self.accept()

    def _on_secondary(self) -> None:
        self.choice = "secondary"
        self.accept()


class InfoDialog(QDialog):
    def __init__(
        self,
        title: str = "Informacao",
        message: str = "",
        button_text: str = "OK",
        button_colors: Optional[Tuple[str, str, str]] = None,
        parent=None,
    ):
        super().__init__(parent)
        self.setObjectName("InfoDialog")
        self.setModal(True)
        self.setMinimumWidth(520)
        self.setWindowTitle(title)
        try:
            if LOGO_ICON.exists():
                self.setWindowIcon(QIcon(str(LOGO_ICON)))
        except Exception:
            pass

        colors = button_colors or ("#27AE60", "#2ECC71", "#1E8449")
        self.setStyleSheet(
            """
            QDialog#InfoDialog { background:#0f1722; }
            QLabel { color:#ECF0F1; font:10pt 'Verdana'; }
            QFrame#Card { background:#162231; border:1px solid #22344a; border-radius:14px; }
            QLabel#Title { font-size: 18px; font-weight: 700; color:#ECF0F1; }
            QLabel#Msg { font-size: 12px; color:#95a5c8; }
            """
        )

        root = QVBoxLayout(self)
        root.setContentsMargins(20, 20, 20, 20)

        card = QFrame(self)
        card.setObjectName("Card")
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(28, 28, 28, 28)
        card_layout.setSpacing(16)

        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(32)
        shadow.setColor(Qt.black)
        shadow.setOffset(0, 12)
        card.setGraphicsEffect(shadow)

        lbl_title = QLabel(title, card)
        lbl_title.setObjectName("Title")
        card_layout.addWidget(lbl_title)

        lbl_msg = QLabel(message, card)
        lbl_msg.setObjectName("Msg")
        lbl_msg.setWordWrap(True)
        card_layout.addWidget(lbl_msg)

        btns = QHBoxLayout()
        btns.addStretch(1)
        btn_ok = QPushButton(button_text, card)
        btn_ok.setCursor(Qt.PointingHandCursor)
        btn_ok.setStyleSheet(button_style(*colors))
        attach_glow(btn_ok, colors[1] if len(colors) > 1 else "#2ECC71")
        btn_ok.clicked.connect(self.accept)
        btns.addWidget(btn_ok)
        card_layout.addLayout(btns)

        root.addWidget(card)


def ensure_license_valid(parent_app: QApplication) -> bool:
    last_fail_msg = None
    saved = get_saved_key()
    if saved:
        ok, msg, meta = check_license_with_supabase(saved)
        if ok:
            persist_key(saved, meta)
            os.environ["AUTOMATIZE_VERIFIED"] = "1"
            _warn_if_expiring_soon()
            return True
        last_fail_msg = msg

    while True:
        dlg = LicenseDialog(preset_key=get_saved_key() or "")
        low = (last_fail_msg or "").lower()
        if "desativad" in low or "inativa" in low or "revog" in low:
            dlg.set_revoked_mode(last_fail_msg)
        if dlg.exec() == QDialog.Accepted:
            _warn_if_expiring_soon()
            return True
        confirm = ConfirmDialog(
            title="E necessario ativar a licenca",
            message="Para continuar, voce precisa ativar a licenca.\nDeseja encerrar o aplicativo agora?",
            primary_text="Encerrar",
            secondary_text="Voltar e ativar",
        )
        confirm.exec()
        if confirm.choice == "primary":
            return False
# --------------------------------------------------------------------


def is_scheduler_mode_enabled() -> bool:
    return str(os.environ.get("AUTOMATIZE_SCHEDULER_MODE") or "").strip().lower() in {
        "1",
        "true",
        "yes",
        "sim",
        "on",
    }

class LogFrame(QFrame):
    """
    Caixa de log com marca d'agua do logo (mesma dinamica usada no sefaz.py).
    """

    def __init__(self, watermark_path: Optional[Path] = None, height: int = 240):
        super().__init__()
        self._pixmap = self._load_pixmap(watermark_path)
        self.setMinimumHeight(height)
        self.setStyleSheet(
            "QFrame { border:1px solid rgba(52,73,94,0.65); border-radius:12px; background:rgba(12,24,40,0.85); }"
        )
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        self.text = QTextEdit(self)
        self.text.setReadOnly(True)
        self.text.setAcceptRichText(True)
        self.text.setPlaceholderText("Logs da execucao...")
        self.text.setStyleSheet(
            "QTextEdit{background:transparent;color:#E8F4FF;font:10pt Consolas,'Courier New';padding:10px;}"
        )
        layout.addWidget(self.text)

    def _soften(self, img: Image.Image) -> Image.Image:
        if img.mode != "RGBA":
            img = img.convert("RGBA")
        w, h = img.size
        alpha = img.split()[-1].point(lambda p: int(p * 0.22))
        cx, cy = w / 2.0, h / 2.0
        half_x = max(1.0, cx)
        grad_x = Image.new("L", (w, 1), 0)
        gx = grad_x.load()
        for x in range(w):
            t = abs((x + 0.5 - cx) / half_x)
            t = min(1.0, t)
            fall = 1.0 - (t ** 1.8)
            gx[x, 0] = int(255 * fall)
        grad_x = grad_x.resize((w, h))
        alpha = ImageChops.multiply(alpha, grad_x)
        alpha = alpha.filter(ImageFilter.GaussianBlur(max(8, int(min(w, h) * 0.05))))
        img.putalpha(alpha)
        return img

    def _load_pixmap(self, watermark_path: Optional[Path]) -> QPixmap:
        try:
            if watermark_path and Path(watermark_path).exists():
                img = Image.open(watermark_path)
                img = self._soften(img)
                return QPixmap.fromImage(ImageQt.ImageQt(img))
        except Exception:
            pass
        return QPixmap()

    def append(self, msg: str) -> None:
        bar = self.text.verticalScrollBar()
        at_bottom = bar.value() >= (bar.maximum() - 2)
        old_value = bar.value()
        self.text.append(msg)
        if at_bottom:
            bar.setValue(bar.maximum())
        else:
            bar.setValue(old_value)

    def paintEvent(self, event):  # noqa: N802
        super().paintEvent(event)
        if getattr(self, "_pixmap", None) and not self._pixmap.isNull():
            painter = QPainter(self)
            sz = self.size()
            target_w = int(sz.width() * 1.35)
            target_h = int(sz.height() * 1.35)
            scaled = self._pixmap.scaled(target_w, target_h, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            x = (sz.width() - scaled.width()) // 2
            y = (sz.height() - scaled.height()) // 2
            painter.drawPixmap(x, y, scaled)


class CompanyItem(QWidget):
    """Item de empresa na lista do robô: somente leitura (espelho do dashboard)."""
    def __init__(self, name: str, doc: str, auth_mode: str):
        super().__init__()
        self.name = normalize_company_name(name)
        self.doc = only_digits(doc)
        self.auth_mode = auth_mode
        display_doc = format_document(self.doc) if self.doc else ""
        layout = QHBoxLayout(self)
        suffix = " [certificado]" if auth_mode == AUTH_CERTIFICATE else ""
        label_name = self.name
        label = f"{label_name} - {display_doc}{suffix}" if display_doc else f"{label_name}{suffix}"
        self.label = QLabel(label)
        self.label.setStyleSheet("QLabel { color:#ECF0F1; font:9pt Verdana; font-weight:bold; }")
        layout.addWidget(self.label)
        layout.addStretch()


class CnpjLookupWorker(QThread):
    status = Signal(str)
    result = Signal(str, str)  # status, payload

    def __init__(self, cnpj: str, parent=None):
        super().__init__(parent)
        self.cnpj = only_digits(cnpj)

    def run(self) -> None:
        # Respeita limite de 3 consultas a cada 20s, aguardando se necessario.
        while True:
            wait = cnpj_wait_time()
            if wait <= 0:
                break
            self.status.emit(f"Aguardando limite da API: {math.ceil(wait)}s")
            time.sleep(wait)
        cnpj_register_call()
        self.status.emit("Consultando API da Receita...")
        try:
            resp = requests.get(f"https://brasilapi.com.br/api/cnpj/v1/{self.cnpj}", timeout=15)
            if resp.status_code != 200:
                self.result.emit("error", f"API retornou {resp.status_code}. Tente novamente.")
                return
            data = resp.json()
            name = (data.get("razao_social") or data.get("nome_fantasia") or "").strip()
            if not name:
                self.result.emit("error", "Nome nao retornado pela Receita.")
                return
            self.result.emit("ok", name)
        except requests.Timeout:
            self.result.emit("error", "Timeout na consulta da Receita.")
        except Exception as exc:  # noqa: BLE001
            self.result.emit("error", f"Falha na consulta: {exc}")


class CompanyEditDialog(QDialog):
    def __init__(
        self,
        name: str = "",
        doc: str = "",
        password: str = "",
        auth_mode: str = AUTH_PASSWORD,
        cert_path: str = "",
        cert_password: str = "",
        cert_blob_b64: str = "",
        parent=None,
    ):
        super().__init__(parent)
        self.setWindowTitle("Empresa")
        self.setFixedWidth(420)
        self.setStyleSheet(
            "QDialog { background:#0f1722; } "
            "QLabel { color:#ECF0F1; font:9pt Verdana; font-weight:bold; }"
        )
        self.auth_mode = auth_mode if auth_mode in (AUTH_PASSWORD, AUTH_CERTIFICATE) else AUTH_PASSWORD
        self.name_edit = QLineEdit(name)
        self.doc_edit = QLineEdit(self._pretty_format_doc(only_digits(doc)))
        self.doc_edit.setPlaceholderText("00.000.000/0000-00")
        self.pass_edit = QLineEdit(password)
        self.pass_edit.setEchoMode(QLineEdit.Password)
        self.cert_blob_b64 = cert_blob_b64 or ""
        self.status = QLabel("")
        self.status.setStyleSheet("color:#e74c3c;")
        self.lookup_thread: Optional[CnpjLookupWorker] = None
        self._lookup_message: str = ""
        self._base_dialog_height: Optional[int] = None
        self._cert_dialog_height: Optional[int] = None
        self._pass_holder_height: Optional[int] = None
        self._cert_holder_height: Optional[int] = None

        for widget in (self.name_edit, self.doc_edit, self.pass_edit):
            widget.setStyleSheet("background:#34495E;color:#ECF0F1;border-radius:6px;padding:6px;font:9pt 'Verdana';")

        header = QLabel("Dados da empresa")
        header.setStyleSheet("color:#5dade2; font:12pt 'Verdana'; font-weight:bold; padding:4px 0;")

        form_frame = QFrame()
        form_frame.setStyleSheet(
            "QFrame { background:#162231; border:1px solid #22344a; border-radius:10px; padding:10px; }"
        )
        form = QVBoxLayout(form_frame)
        form.setSpacing(6)
        form.addWidget(QLabel("Nome"))
        form.addWidget(self.name_edit)
        form.addWidget(QLabel("CPF/CNPJ"))
        form.addWidget(self.doc_edit)
        form.addWidget(QLabel("Metodo de acesso"))

        auth_row = QHBoxLayout()
        auth_row.setSpacing(6)
        self.radio_password = QRadioButton("Usuario e senha")
        self.radio_certificate = QRadioButton("Certificado digital (A1)")
        for rb in (self.radio_password, self.radio_certificate):
            rb.setStyleSheet("QRadioButton { color:#ECF0F1; font:9pt Verdana; font-weight:bold; }")
        self.radio_password.setChecked(self.auth_mode == AUTH_PASSWORD)
        self.radio_certificate.setChecked(self.auth_mode == AUTH_CERTIFICATE)
        self.radio_password.toggled.connect(self._on_auth_mode_changed)
        self.radio_certificate.toggled.connect(self._on_auth_mode_changed)
        auth_row.addWidget(self.radio_password)
        auth_row.addWidget(self.radio_certificate)
        form.addLayout(auth_row)

        self.pass_container = QFrame()
        self.pass_container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        pass_layout = QVBoxLayout(self.pass_container)
        pass_layout.setContentsMargins(0, 0, 0, 0)
        pass_layout.setSpacing(3)
        pass_layout.addWidget(QLabel("Senha"))

        pwd_row = QHBoxLayout()
        pwd_row.setSpacing(4)
        pwd_row.addWidget(self.pass_edit, stretch=1)
        self.eye_btn = QPushButton()
        self.eye_btn.setCheckable(True)
        self.eye_btn.setFixedWidth(38)
        self.eye_btn.setStyleSheet(
            "QPushButton { background:#22344a; color:#ECF0F1; border:1px solid #2d4a68; border-radius:6px; padding:4px; }"
        )
        self.eye_btn.setIcon(QIcon(str(EYE_CLOSED_PNG)) if EYE_CLOSED_PNG.exists() else QIcon())
        self.eye_btn.setIconSize(QSize(18, 18))
        self.eye_btn.toggled.connect(self._toggle_password)
        pwd_row.addWidget(self.eye_btn)
        pass_layout.addLayout(pwd_row)

        self.cert_container = QFrame()
        self.cert_container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        cert_layout = QVBoxLayout(self.cert_container)
        cert_layout.setContentsMargins(0, 0, 0, 0)
        cert_layout.setSpacing(5)
        cert_layout.addWidget(QLabel("Certificado (.pfx)"))

        cert_path_row = QHBoxLayout()
        cert_path_row.setSpacing(4)
        self.cert_path_edit = QLineEdit(cert_path)
        self.cert_path_edit.setStyleSheet(
            "background:#34495E;color:#ECF0F1;border-radius:6px;padding:6px;font:9pt 'Verdana';"
        )
        cert_path_row.addWidget(self.cert_path_edit, stretch=1)
        browse_btn = QPushButton("...")
        browse_btn.setFixedWidth(64)
        browse_btn.setStyleSheet(button_style("#2980B9", "#3498DB", "#2471A3"))
        browse_btn.clicked.connect(self._browse_cert)
        cert_path_row.addWidget(browse_btn)
        cert_layout.addLayout(cert_path_row)

        cert_layout.addWidget(QLabel("Senha do certificado"))
        self.cert_pass_edit = QLineEdit(cert_password)
        self.cert_pass_edit.setEchoMode(QLineEdit.Password)
        self.cert_pass_edit.setStyleSheet(
            "background:#34495E;color:#ECF0F1;border-radius:6px;padding:6px;font:9pt 'Verdana';"
        )
        cert_layout.addWidget(self.cert_pass_edit)

        self.auth_stack = QStackedLayout()
        self.auth_stack.addWidget(self.pass_container)
        self.auth_stack.addWidget(self.cert_container)
        self.auth_holder = QWidget()
        holder_layout = QVBoxLayout(self.auth_holder)
        holder_layout.setContentsMargins(0, 0, 0, 0)
        holder_layout.setSpacing(0)
        holder_layout.addLayout(self.auth_stack)
        form.addWidget(self.auth_holder)

        form.addWidget(self.status)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        self.ok_button = buttons.button(QDialogButtonBox.Ok)

        for btn in buttons.buttons():
            role = buttons.buttonRole(btn)
            if role == QDialogButtonBox.AcceptRole:
                btn.setStyleSheet(button_style("#27AE60", "#2ECC71", "#1E8449"))
                btn.setText("OK")
                attach_glow(btn, "#2ECC71")
            else:
                btn.setStyleSheet(button_style("#5D6D7E", "#707B7C", "#4A545B"))
                btn.setText("Cancelar")
                attach_glow(btn, "#9CA3AF")
                attach_glow(btn, "#9CA3AF")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)
        layout.addWidget(header)
        layout.addWidget(form_frame)
        layout.addWidget(buttons, alignment=Qt.AlignRight)

        self.doc_edit.textChanged.connect(self._on_doc_changed)
        self.doc_edit.editingFinished.connect(self._on_doc_finished)
        self.name_edit.textChanged.connect(self._update_state)
        self.pass_edit.textChanged.connect(self._update_state)
        self.cert_pass_edit.textChanged.connect(self._update_state)
        self.cert_path_edit.textChanged.connect(self._update_state)

        self._on_auth_mode_changed()
        self._update_state()

    def _on_doc_changed(self, text: str) -> None:
        self._lookup_message = ""
        digits = only_digits(text)
        formatted = self._pretty_format_doc(digits)
        if text != formatted:
            self.doc_edit.blockSignals(True)
            self.doc_edit.setText(formatted)
            # Mantem o cursor no fim ao aplicar a mascara para nao pular para o meio.
            self.doc_edit.setCursorPosition(len(formatted))
            self.doc_edit.blockSignals(False)
        self._update_state()

    def _pretty_format_doc(self, digits: str) -> str:
        if not digits:
            return ""
        if len(digits) <= 3:
            return digits
        if len(digits) <= 6:
            return f"{digits[:3]}.{digits[3:]}"
        if len(digits) <= 9:
            return f"{digits[:3]}.{digits[3:6]}.{digits[6:]}"
        if len(digits) <= 11:
            return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"
        if len(digits) <= 12:
            return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:]}"
        return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:14]}"

    def _on_doc_finished(self) -> None:
        if self.name_edit.text().strip():
            return
        digits = only_digits(self.doc_edit.text())
        if len(digits) != 14 or not is_valid_cnpj(digits):
            return
        if self.lookup_thread and self.lookup_thread.isRunning():
            return
        self._lookup_message = "Consultando API da Receita..."
        self._update_state()
        worker = CnpjLookupWorker(digits, parent=self)
        self.lookup_thread = worker
        worker.status.connect(self._on_lookup_status)
        worker.result.connect(self._on_lookup_result)
        worker.finished.connect(self._on_lookup_finished)
        worker.start()

    def _on_lookup_status(self, msg: str) -> None:
        self._lookup_message = msg
        self._update_state()

    def _on_lookup_result(self, status: str, payload: str) -> None:
        if status == "ok":
            if not self.name_edit.text().strip():
                self.name_edit.setText(payload)
            self._lookup_message = ""
        else:
            self._lookup_message = payload or "Nao foi possivel consultar o CNPJ."
        self._update_state()

    def _on_lookup_finished(self) -> None:
        self.lookup_thread = None

    def closeEvent(self, event) -> None:
        if self.lookup_thread and self.lookup_thread.isRunning():
            self.lookup_thread.requestInterruption()
            self.lookup_thread.wait(3000)
        super().closeEvent(event)

    def _update_state(self) -> None:
        name_ok = bool(self.name_edit.text().strip())
        doc = self.doc_edit.text()
        doc_ok = is_valid_document(doc)
        using_cert = self.radio_certificate.isChecked()
        cert_path_txt = self.cert_path_edit.text().strip()
        cert_file_exists = Path(cert_path_txt).exists() if cert_path_txt else False
        cert_has_blob = bool(self.cert_blob_b64)
        cert_ok = cert_file_exists or cert_has_blob
        cert_pwd_ok = bool(self.cert_pass_edit.text())
        pwd_ok = bool(self.pass_edit.text()) if not using_cert else (cert_ok and cert_pwd_ok)

        if not doc:
            msg = "Informe CPF ou CNPJ."
        elif not doc_ok:
            msg = "Documento invalido."
        elif using_cert and not cert_ok:
            msg = "Selecione o certificado (.pfx)."
        elif using_cert and not cert_pwd_ok:
            msg = "Informe a senha do certificado."
        elif not using_cert and not self.pass_edit.text():
            msg = "Informe a senha."
        else:
            msg = ""
        if self._lookup_message:
            msg = self._lookup_message
        self.status.setText(msg)
        self.ok_button.setEnabled(name_ok and doc_ok and pwd_ok)

    def get_data(self) -> Dict[str, str]:
        using_cert = self.radio_certificate.isChecked()
        mode = AUTH_CERTIFICATE if using_cert else AUTH_PASSWORD
        cert_path_txt = self.cert_path_edit.text().strip()
        cert_blob_b64 = self.cert_blob_b64 if using_cert else ""
        if using_cert and cert_path_txt:
            try:
                cert_blob_b64 = base64.b64encode(Path(cert_path_txt).read_bytes()).decode("ascii")
            except Exception:
                cert_blob_b64 = self.cert_blob_b64
        return {
            "name": normalize_company_name(self.name_edit.text()),
            "doc": only_digits(self.doc_edit.text()),
            "password": "" if using_cert else self.pass_edit.text(),
            "auth_mode": mode,
            "cert_path": "",
            "cert_password": self.cert_pass_edit.text() if using_cert else "",
            "cert_blob_b64": cert_blob_b64 if using_cert else "",
        }

    def _toggle_password(self, checked: bool) -> None:
        self.pass_edit.setEchoMode(QLineEdit.Normal if checked else QLineEdit.Password)
        if EYE_OPEN_PNG.exists() and EYE_CLOSED_PNG.exists():
            self.eye_btn.setIcon(QIcon(str(EYE_OPEN_PNG if checked else EYE_CLOSED_PNG)))
        else:
            self.eye_btn.setText("Ver" if checked else "Ocultar")

    def _on_auth_mode_changed(self) -> None:
        using_cert = self.radio_certificate.isChecked()
        if hasattr(self, "auth_stack"):
            self.auth_stack.setCurrentWidget(self.cert_container if using_cert else self.pass_container)
        self._set_auth_holder_height(using_cert)
        self._update_state()

    def _browse_cert(self) -> None:
        fname, _ = QFileDialog.getOpenFileName(
            self,
            "Selecione o certificado (.pfx)",
            str(Path.home()),
            "Certificado (*.pfx);;Todos os arquivos (*)",
        )
        if fname:
            self.cert_path_edit.setText(fname)

    def _set_auth_holder_height(self, using_cert: Optional[bool] = None) -> None:
        if using_cert is None:
            using_cert = self.radio_certificate.isChecked()
        if self._base_dialog_height is None:
            self._pass_holder_height = self.pass_container.sizeHint().height()
            self._cert_holder_height = self.cert_container.sizeHint().height()
            self._base_dialog_height = self.sizeHint().height()
            extra = max(0, (self._cert_holder_height or 0) - (self._pass_holder_height or 0))
            self._cert_dialog_height = self._base_dialog_height + extra

        target_holder = self._cert_holder_height if using_cert else self._pass_holder_height
        if target_holder is not None:
            self.auth_holder.setMinimumHeight(target_holder)
            self.auth_holder.setMaximumHeight(target_holder)

        target_dialog_height = self._cert_dialog_height if using_cert else self._base_dialog_height
        if target_dialog_height is not None:
            self.setMinimumHeight(target_dialog_height)
            self.setMaximumHeight(target_dialog_height)
            self.resize(self.width(), target_dialog_height)


class CompanyPickerDialog(QDialog):
    def __init__(
        self,
        companies: List[Tuple[str, str, str]],
        title: str,
        action_text: str,
        accent: str = "#2980B9",
        accent_hover: Optional[str] = None,
        accent_pressed: Optional[str] = None,
        multi_select: bool = False,
        parent=None,
    ):
        super().__init__(parent)
        self.selected_index: Optional[int] = None
        self.selected_indexes: List[int] = []
        self.multi_select = bool(multi_select)
        self.setWindowTitle(title)
        self.setModal(True)
        self.setMinimumWidth(460)
        self.setStyleSheet(
            """
            QDialog { background:#0b1220; }
            QLabel { color:#E8F4FF; font:10pt Verdana; font-weight:bold; }
            QLineEdit {
                background:#132033;
                color:#E8F4FF;
                border-radius:8px;
                padding:8px;
                border:1px solid #1f3b5b;
            }
            QListWidget {
                background:#0c1624;
                color:#E8F4FF;
                border:1px solid #1f3b5b;
                border-radius:10px;
                padding:6px;
            }
            QListWidget::item { padding:10px 8px; margin:2px; }
            QListWidget::item:selected { background:#1f2f46; border-radius:6px; }
            QPushButton { font:10pt Verdana; color:#fff; padding:10px 14px; border-radius:10px; }
            """
        )

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)

        header = QLabel(title)
        header.setStyleSheet("color:#5dade2; font:12pt 'Verdana'; font-weight:bold; padding-bottom:4px;")
        layout.addWidget(header)

        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Pesquisar empresa...")
        layout.addWidget(self.search_edit)

        self.list = QListWidget()
        if self.multi_select:
            self.list.setSelectionMode(QListWidget.ExtendedSelection)
        else:
            self.list.setSelectionMode(QListWidget.SingleSelection)
        self.list.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        layout.addWidget(self.list)

        for idx, (name, doc, meta) in enumerate(companies):
            label = f"{name}\n{doc}"
            if meta:
                label = f"{name} • {meta}\n{doc}"
            item = QListWidgetItem(label)
            item.setData(Qt.UserRole, idx)
            item.setToolTip(name)
            self.list.addItem(item)
        if self.list.count():
            self.list.setCurrentRow(0)

        btns = QHBoxLayout()
        btns.addStretch(1)
        btn_cancel = QPushButton("Cancelar")
        btn_cancel.setStyleSheet(button_style("#5D6D7E", "#707B7C", "#4A545B"))
        btn_cancel.clicked.connect(self.reject)

        btn_ok = QPushButton(action_text)
        btn_ok.setStyleSheet(button_style(accent, accent_hover or accent, accent_pressed or accent))
        btn_ok.clicked.connect(self._accept_current)

        attach_glow(btn_cancel, "#9CA3AF")
        attach_glow(btn_ok, accent or "#3498DB")

        btns.addWidget(btn_cancel)
        btns.addWidget(btn_ok)
        layout.addLayout(btns)

        self.search_edit.textChanged.connect(self._filter)
        self.list.itemDoubleClicked.connect(lambda _: self._accept_current())

    def _filter(self, text: str) -> None:
        low = text.lower()
        for i in range(self.list.count()):
            item = self.list.item(i)
            visible = low in item.text().lower()
            item.setHidden(not visible)

    def _accept_current(self) -> None:
        if self.multi_select:
            selected = self.list.selectedItems()
            if not selected:
                current = self.list.currentItem()
                if current:
                    selected = [current]
            self.selected_indexes = [int(it.data(Qt.UserRole)) for it in selected if it.data(Qt.UserRole) is not None]
            if not self.selected_indexes:
                return
            self.selected_indexes = sorted(dict.fromkeys(self.selected_indexes))
            self.selected_index = self.selected_indexes[0]
        else:
            item = self.list.currentItem()
            if not item:
                return
            self.selected_index = item.data(Qt.UserRole)
        self.accept()

    def get_index(self) -> Optional[int]:
        return self.selected_index

    def get_indexes(self) -> List[int]:
        return list(self.selected_indexes)

    def keyPressEvent(self, event) -> None:  # noqa: N802
        if self.multi_select and (event.modifiers() & Qt.ControlModifier) and event.key() == Qt.Key_A:
            self.list.selectAll()
            event.accept()
            return
        super().keyPressEvent(event)


class ManagerDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.choice: Optional[str] = None
        self.setWindowTitle("Gerenciador de Empresas")
        self.setStyleSheet(
            """
            QDialog { background:#0f1722; }
            QLabel { color:#ECF0F1; font:10pt Verdana; font-weight:bold; }
            QPushButton { font:10pt Verdana; color:#fff; padding:10px 14px; border-radius:8px; }
            """
        )

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)

        header = QLabel("Ações disponíveis")
        header.setStyleSheet("color:#5dade2; font:12pt 'Verdana'; font-weight:bold; padding-bottom:6px;")
        layout.addWidget(header)

        box = QFrame()
        box.setStyleSheet(
            "QFrame { background:#162231; border:1px solid #22344a; border-radius:10px; padding:12px; }"
        )
        box_layout = QVBoxLayout(box)
        box_layout.setSpacing(8)
        box_layout.addWidget(QLabel("Escolha o que deseja fazer:"))

        btn_path = QPushButton("Estrutura de pastas")
        btn_path.setStyleSheet(button_style("#F1C40F", "#F4D03F", "#D4AC0D"))
        icon_path = load_png_icon("caminho")
        if icon_path:
            btn_path.setIcon(icon_path)
        attach_glow(btn_path, "#F4D03F")
        btn_path.clicked.connect(lambda: self._choose("path"))

        for b in (btn_path,):
            box_layout.addWidget(b)

        layout.addWidget(box)

        btns = QHBoxLayout()
        btns.setContentsMargins(0, 4, 0, 0)
        btns.setSpacing(10)
        btns.addStretch(1)

        btn_cancel = QPushButton("Cancelar")
        btn_cancel.setCursor(Qt.PointingHandCursor)
        btn_cancel.setStyleSheet(button_style("#5D6D7E", "#707B7C", "#4A545B"))
        attach_glow(btn_cancel, "#9CA3AF")
        btn_cancel.clicked.connect(self.reject)

        btns.addWidget(btn_cancel)
        layout.addLayout(btns)

    def _choose(self, action: str) -> None:
        self.choice = action
        self.accept()

class PathDialog(QDialog):
    def __init__(self, current: Path, current_report: Optional[Path], folder_options: Optional[Dict[str, Any]] = None, parent=None, preferences: Optional[Dict[str, Any]] = None, read_only: bool = False, segment_path: Optional[str] = None):
        super().__init__(parent)
        self.setWindowTitle("Estrutura de pastas" if read_only else "Pasta de XMLs")
        self.setMinimumWidth(520)
        self.setStyleSheet(
            """
            QDialog { background:#0f1722; }
            QLabel { color:#ECF0F1; font:9pt Verdana; font-weight:bold; }
            QPushButton { font:9pt Verdana; color:#fff; padding:8px 12px; border-radius:8px; }
            QLineEdit { background:#34495E; color:#ECF0F1; border-radius:6px; padding:6px; font-weight:bold; }
            """
        )
        self._resolved_xml = current
        self._resolved_report = current_report or current
        if read_only:
            self._build_read_only_layout(current, current_report, segment_path)
            return

        self.status = QLabel("")
        self.status.setStyleSheet("color:#e74c3c;")

        self.line = QLineEdit(str(current))
        browse = QPushButton("...")
        browse.setFixedWidth(52)
        browse.setStyleSheet(button_style("#2980B9", "#3498DB", "#2471A3"))
        browse.clicked.connect(lambda: self._browse(self.line))

        self.line_report = QLineEdit(str(current_report or current))
        browse_report = QPushButton("...")
        browse_report.setFixedWidth(52)
        browse_report.setStyleSheet(button_style("#9B59B6", "#AF7AC5", "#7D3C98"))
        browse_report.clicked.connect(lambda: self._browse(self.line_report))

        row = QHBoxLayout()
        row.addWidget(self.line, stretch=1)
        row.addWidget(browse)

        row_report = QHBoxLayout()
        row_report.addWidget(self.line_report, stretch=1)
        row_report.addWidget(browse_report)

        btns = QHBoxLayout()
        btns.setContentsMargins(0, 4, 0, 0)
        btns.setSpacing(10)
        btns.addStretch(1)

        self.btn_ok = QPushButton("OK")
        self.btn_ok.setCursor(Qt.PointingHandCursor)
        self.btn_ok.setDefault(True)
        self.btn_ok.setStyleSheet(button_style("#27AE60", "#2ECC71", "#1E8449"))
        attach_glow(self.btn_ok, "#2ECC71")
        self.btn_ok.clicked.connect(self._on_accept)

        btn_cancel = QPushButton("Cancelar")
        btn_cancel.setCursor(Qt.PointingHandCursor)
        btn_cancel.setStyleSheet(button_style("#5D6D7E", "#707B7C", "#4A545B"))
        attach_glow(btn_cancel, "#9CA3AF")
        btn_cancel.clicked.connect(self.reject)

        btns.addWidget(btn_cancel)
        btns.addWidget(self.btn_ok)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)
        layout.addWidget(QLabel("Informe o caminho para salvar os XMLs:"))
        layout.addLayout(row)
        layout.addWidget(QLabel("Informe o caminho para salvar os relatorios (PDF):"))
        layout.addLayout(row_report)
        layout.addWidget(self.status)

        # Preview: onde os arquivos serao salvos (estrutura centralizada ou local)
        self._central_segment_path, self._central_date_rule = fetch_central_folder_structure()
        preview_label = QLabel("Onde os arquivos serao salvos:")
        preview_label.setStyleSheet("color:#7DD3FC;font:9pt Verdana;font-weight:bold;margin-top:8px;")
        layout.addWidget(preview_label)
        self.preview_text = QTextEdit()
        self.preview_text.setReadOnly(True)
        self.preview_text.setMaximumHeight(140)
        self.preview_text.setStyleSheet(
            "QTextEdit { background:#0F172A; color:#94A3B8; border:1px solid #334155; border-radius:6px; padding:8px; font-family:Consolas,Courier; font-size:9pt; }"
        )
        layout.addWidget(self.preview_text)
        self._update_structure_preview()

        def _on_path_changed() -> None:
            self._update_structure_preview()
        self.line.textChanged.connect(_on_path_changed)

        # Rodar em servidor (VM):
        prefs = preferences or {}
        self.chk_rodar_servidor = QCheckBox("Rodar em servidor")
        self.chk_rodar_servidor.setChecked(bool(prefs.get("rodar_em_servidor")))
        self.chk_rodar_servidor.setStyleSheet("color:#ECF0F1;font:9pt Verdana;")
        self.chk_rodar_servidor.setToolTip("Quando ativo, ao abrir o robô inicia contagem regressiva de 10s e executa com todas as empresas (Emitidas+Recebidas, dia atual). Empresas sao carregadas do Supabase.")
        layout.addWidget(self.chk_rodar_servidor)

        self.le_supabase_url = QLineEdit(prefs.get("supabase_url", ""))
        self.le_supabase_url.setPlaceholderText("SUPABASE_URL (ex: https://xxx.supabase.co)")
        self.le_supabase_url.setStyleSheet("background:#34495E;color:#ECF0F1;border-radius:6px;padding:6px;font:9pt Verdana;")
        self.le_supabase_anon_key = QLineEdit(prefs.get("supabase_anon_key", ""))
        self.le_supabase_anon_key.setPlaceholderText("SUPABASE_SERVICE_ROLE_KEY")
        self.le_supabase_anon_key.setEchoMode(QLineEdit.Password)
        self.le_supabase_anon_key.setStyleSheet("background:#34495E;color:#ECF0F1;border-radius:6px;padding:6px;font:9pt Verdana;")

        server_layout = QVBoxLayout()
        server_layout.setSpacing(4)
        server_layout.addWidget(QLabel("SUPABASE_URL"))
        server_layout.addWidget(self.le_supabase_url)
        server_layout.addWidget(QLabel("SUPABASE_SERVICE_ROLE_KEY"))
        server_layout.addWidget(self.le_supabase_anon_key)
        self.server_container = QWidget()
        self.server_container.setLayout(server_layout)
        layout.addWidget(self.server_container)
        self.server_container.setVisible(self.chk_rodar_servidor.isChecked())
        self.chk_rodar_servidor.toggled.connect(lambda checked: self.server_container.setVisible(checked))

        layout.addWidget(QLabel("Estrutura de pastas para salvar XMLs:"))

        folder_base = normalizar_estrutura_pastas(folder_options)
        struct_layout = QVBoxLayout()
        struct_layout.setSpacing(6)

        self.chk_pasta_cliente = QCheckBox("Pasta de opção do cliente")
        self.chk_pasta_cliente.setChecked(bool(folder_base.get("usar_pasta_cliente")))
        self.chk_pasta_cliente.setStyleSheet("color:#ECF0F1;font:9pt Verdana;")
        struct_layout.addWidget(self.chk_pasta_cliente)

        self._pastas_cliente_state: List[str] = []
        pastas_cliente_initial = list(folder_base.get("pastas_cliente") or [])
        if not pastas_cliente_initial:
            legacy_one = (folder_base.get("nome_pasta_cliente") or "").strip()
            if legacy_one:
                pastas_cliente_initial = [legacy_one]

        row_custom = QHBoxLayout()
        row_custom.addSpacing(18)
        self.le_pasta_cliente = QLineEdit("")
        self.le_pasta_cliente.setPlaceholderText("Adicionar pasta (ex.: DEPARTAMENTO FISCAL)")
        self.le_pasta_cliente.setEnabled(self.chk_pasta_cliente.isChecked())

        self.btn_add_pasta = QPushButton("+")
        self.btn_add_pasta.setFixedSize(38, 30)
        self.btn_add_pasta.setCursor(Qt.PointingHandCursor)
        self.btn_add_pasta.setStyleSheet(button_style("#2980B9", "#3498DB", "#2471A3"))
        attach_glow(self.btn_add_pasta, "#3498DB")
        self.btn_add_pasta.setEnabled(self.chk_pasta_cliente.isChecked())
        row_custom.addWidget(self.le_pasta_cliente)
        row_custom.addWidget(self.btn_add_pasta)
        struct_layout.addLayout(row_custom)

        self._lista_pastas_widget = QWidget()
        lista_layout = QVBoxLayout(self._lista_pastas_widget)
        lista_layout.setContentsMargins(30, 0, 0, 0)
        lista_layout.setSpacing(4)

        def _pasta_existe(n: str) -> bool:
            return any((x or "").strip().lower() == (n or "").strip().lower() for x in self._pastas_cliente_state)

        def _add_pasta_item(nome: str) -> None:
            nome = (nome or "").strip()
            if not nome or _pasta_existe(nome):
                return
            self._pastas_cliente_state.append(nome)

            item = QWidget()
            item_row = QHBoxLayout(item)
            item_row.setContentsMargins(0, 0, 0, 0)
            item_row.setSpacing(6)

            lbl = QLabel(nome)
            lbl.setStyleSheet("QLabel { color: #BFE6FF; font-family: Verdana; }")

            btn_del = QPushButton("-")
            btn_del.setFixedSize(34, 26)
            btn_del.setCursor(Qt.PointingHandCursor)
            btn_del.setStyleSheet(button_style("#34495E", "#5D6D7E", "#2C3E50"))
            attach_glow(btn_del, "#9CA3AF")
            item_row.addWidget(lbl)
            item_row.addStretch()
            item_row.addWidget(btn_del)

            def _remove() -> None:
                try:
                    for i, v in enumerate(list(self._pastas_cliente_state)):
                        if (v or "").strip().lower() == nome.lower():
                            self._pastas_cliente_state.pop(i)
                            break
                except Exception:
                    pass
                item.setParent(None)
                item.deleteLater()
                self._update_structure_preview()

            btn_del.clicked.connect(_remove)
            lista_layout.addWidget(item)
            self._update_structure_preview()

        for p in pastas_cliente_initial:
            _add_pasta_item(p)

        struct_layout.addWidget(self._lista_pastas_widget)

        def _on_add_click() -> None:
            raw_txt = self.le_pasta_cliente.text().strip()
            if not raw_txt:
                return
            parts = re.split(r"[;|\n]+", raw_txt)
            for part in parts:
                part = (part or "").strip()
                if not part:
                    continue
                segs = re.split(r"[\\/]+", part) if ("\\" in part or "/" in part) else [part]
                for s in segs:
                    s = (s or "").strip()
                    if s:
                        _add_pasta_item(s)
            self.le_pasta_cliente.clear()

        self.btn_add_pasta.clicked.connect(_on_add_click)

        def _toggle_pasta_cliente(enabled: bool) -> None:
            self.le_pasta_cliente.setEnabled(enabled)
            self.btn_add_pasta.setEnabled(enabled)
            self._lista_pastas_widget.setEnabled(enabled)

        self.chk_pasta_cliente.toggled.connect(_toggle_pasta_cliente)
        _toggle_pasta_cliente(self.chk_pasta_cliente.isChecked())
        self.chk_pasta_cliente.toggled.connect(lambda: self._update_structure_preview())

        self.chk_year = QCheckBox("Separar por ano")
        self.chk_year.setChecked(bool(folder_base.get("year")))
        self.chk_month = QCheckBox("Separar por mês")
        self.chk_month.setChecked(bool(folder_base.get("month")))
        self.chk_day = QCheckBox("Separar por dia")
        self.chk_day.setChecked(bool(folder_base.get("day")))
        folder_layout = QHBoxLayout()
        folder_layout.setSpacing(16)
        for chk in (self.chk_year, self.chk_month, self.chk_day):
            chk.setStyleSheet("color:#ECF0F1;font:9pt Verdana;")
            folder_layout.addWidget(chk)
        struct_layout.addLayout(folder_layout)
        layout.addLayout(struct_layout)
        layout.addLayout(btns)

    def _browse(self, target_line: QLineEdit) -> None:
        dlg_path = QFileDialog.getExistingDirectory(self, "Escolha a pasta")
        if dlg_path:
            target_line.setText(dlg_path)

    def _update_structure_preview(self) -> None:
        base_resolved = get_resolved_output_base()
        base_from_api_or_env = str(base_resolved) if base_resolved else ""
        base_txt = self.line.text().strip()
        base = base_from_api_or_env or base_txt or os.environ.get("BASE_PATH", "").strip()
        if not base:
            self.preview_text.setPlainText("Defina a pasta base acima, ou no painel Admin (Pasta base na VM), ou BASE_PATH no .env")
            return
        lines = [base]
        if self._central_segment_path:
            lines.append("  └── [nome_empresa]   ← mesmo nome do dashboard")
            indent = "      "
            for i, seg in enumerate(self._central_segment_path):
                prefix = "└── " if i == len(self._central_segment_path) - 1 else "├── "
                lines.append(f"{indent}{prefix} {seg}")
                indent += "    "
            date_rule = self._central_date_rule or ""
            if date_rule == "year_month_day":
                date_ex = "2026\\03\\06"
            elif date_rule == "year_month":
                date_ex = "2026\\03"
            elif date_rule == "year":
                date_ex = "2026"
            else:
                date_ex = ""
            seg_indent = indent
            lines.append(f"{seg_indent}├── Emitidas" + (f"\\{date_ex}" if date_ex else ""))
            lines.append(f"{seg_indent}└── Recebidas" + (f"\\{date_ex}" if date_ex else ""))
            if base_from_api_or_env:
                lines.append("")
                lines.append("(Pasta base do painel/API ou .env)")
        else:
            lines.append("  └── [nome_empresa]")
            lines.append("      ├── Emitidas")
            lines.append("      └── Recebidas")
            chk = getattr(self, "chk_pasta_cliente", None)
            pastas = getattr(self, "_pastas_cliente_state", [])
            if chk and chk.isChecked() and pastas:
                extra = "  ".join(pastas[:3])
                if len(pastas) > 3:
                    extra += " ..."
                lines.append("")
                lines.append(f"Subpastas (cliente): {extra}")
        self.preview_text.setPlainText("\n".join(lines))

    def _on_accept(self) -> None:
        path_txt = self.line.text().strip()
        report_txt = self.line_report.text().strip()
        if not path_txt:
            self.status.setText("Informe um caminho.")
            return
        p = Path(path_txt)
        if not p.exists():
            self.status.setText("Caminho inexistente.")
            return
        p_report = Path(report_txt) if report_txt else p
        if not p_report.exists():
            self.status.setText("Caminho de relatorio inexistente.")
            return
        self.status.setText("")
        self._resolved_xml = p
        self._resolved_report = p_report
        self.accept()

    def get_path(self) -> Path:
        return Path(self.line.text().strip())

    def get_paths(self) -> Tuple[Path, Path]:
        return self._resolved_xml, self._resolved_report

    def get_folder_options(self) -> Dict[str, Any]:
        return {
            "year": self.chk_year.isChecked(),
            "month": self.chk_month.isChecked(),
            "day": self.chk_day.isChecked(),
            "usar_pasta_cliente": self.chk_pasta_cliente.isChecked(),
            "pastas_cliente": list(self._pastas_cliente_state),
            "nome_pasta_cliente": self._pastas_cliente_state[0] if self._pastas_cliente_state else "",
            "rodar_em_servidor": self.chk_rodar_servidor.isChecked(),
            "supabase_url": self.le_supabase_url.text().strip(),
            "supabase_anon_key": self.le_supabase_anon_key.text().strip(),
        }

    def _build_read_only_layout(self, current: Path, current_report: Optional[Path], segment_path: Optional[str]) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)
        layout.addWidget(QLabel("Pasta base (definida no .env):"))
        layout.addWidget(QLabel(str(current)))
        layout.addWidget(QLabel("Pasta de relatórios (departamento):"))
        layout.addWidget(QLabel(str(current_report or current)))
        central_segments, central_date_rule = fetch_central_folder_structure(segment_path)
        preview_label = QLabel("Onde os arquivos serão salvos:")
        preview_label.setStyleSheet("color:#7DD3FC;font:9pt Verdana;font-weight:bold;margin-top:8px;")
        layout.addWidget(preview_label)
        preview_text = QTextEdit()
        preview_text.setReadOnly(True)
        preview_text.setMaximumHeight(160)
        preview_text.setStyleSheet(
            "QTextEdit { background:#0F172A; color:#94A3B8; border:1px solid #334155; border-radius:6px; padding:8px; font-family:Consolas,Courier; font-size:9pt; }"
        )
        base = str(current)
        lines = [base] if base else ["(Pasta base do painel ou .env)"]
        if central_segments:
            lines.append("  └── [nome_empresa]   ← mesmo nome do dashboard")
            indent = "      "
            for i, seg in enumerate(central_segments):
                prefix = "└── " if i == len(central_segments) - 1 else "├── "
                lines.append(f"{indent}{prefix} {seg}")
                indent += "    "
            date_ex = "2026\\03\\06" if (central_date_rule == "year_month_day") else "2026\\03" if (central_date_rule == "year_month") else "2026" if (central_date_rule == "year") else ""
            lines.append(f"{indent}├── Emitidas" + (f"\\{date_ex}" if date_ex else ""))
            lines.append(f"{indent}└── Recebidas" + (f"\\{date_ex}" if date_ex else ""))
        else:
            lines.append("  └── [nome_empresa]")
            lines.append("      ├── Emitidas")
            lines.append("      └── Recebidas")
        preview_text.setPlainText("\n".join(lines))
        layout.addWidget(preview_text)
        btn_close = QPushButton("Fechar")
        btn_close.setCursor(Qt.PointingHandCursor)
        btn_close.setStyleSheet(button_style("#5D6D7E", "#707B7C", "#4A545B"))
        attach_glow(btn_close, "#9CA3AF")
        btn_close.clicked.connect(self.reject)
        layout.addWidget(btn_close)



# --------------------------------------------------------------------
# Worker (Playwright) rodando em thread separada
# --------------------------------------------------------------------


@dataclass
class Company:
    name: str
    doc: str
    company_id: str = ""
    password: str = ""
    auth_mode: str = AUTH_PASSWORD
    cert_path: Optional[Path] = None
    cert_password: str = ""
    cert_data: Optional[bytes] = None



class DownloadThread(QThread):
    log = Signal(str)
    finished = Signal()
    summary_ready = Signal(object)

    def __init__(
        self,
        companies: List[Company],
        headless: bool,
        output_base: Path,
        period_start: datetime,
        period_end: datetime,
        include_emitidas: bool = True,
        include_recebidas: bool = True,
        folder_structure: Optional[Dict[str, Any]] = None,
        central_segment_path: Optional[List[str]] = None,
        central_date_rule: Optional[str] = None,
        segment_path_from_dashboard: Optional[str] = None,
    ):
        super().__init__()
        self.companies = companies
        self.headless = headless
        self.output_base = output_base
        self.period_start = period_start
        self.period_end = period_end
        self.include_emitidas = include_emitidas
        self.include_recebidas = include_recebidas
        self.folder_structure = normalizar_estrutura_pastas(folder_structure)
        self._central_segment_path = central_segment_path
        self._central_date_rule = central_date_rule
        self._segment_path_from_dashboard = (segment_path_from_dashboard or "").strip() or None
        self._stop_requested = False
        self._first_company = True
        self.summary_data: Dict[str, Any] = {}

    def request_stop(self) -> None:
        self._stop_requested = True
        self.requestInterruption()

    def _resolve_chromium_executable(self, pw) -> Optional[Path]:
        """
        Garante que o Playwright use o Chromium existente mesmo que o nome da pasta mude
        (ex.: chromium-1187, chromium-1200). Tenta o caminho padrao e, se faltar,
        escolhe a revisao mais nova encontrada em PLAYWRIGHT_BROWSERS_PATH.
        """
        try:
            default_path = Path(pw.chromium.executable_path)
            if default_path.exists():
                return default_path
        except Exception:
            default_path = None

        root = Path(os.environ.get("PLAYWRIGHT_BROWSERS_PATH", str(PLAYWRIGHT_DIR)))
        candidates = sorted(root.glob("chromium-*"), reverse=True)
        for cand in candidates:
            if os.name == "nt":
                exe = cand / "chrome-win64" / "chrome.exe"
                if not exe.exists():
                    exe = cand / "chrome-win" / "chrome.exe"
            else:
                exe = cand / "chrome-linux" / "chrome"
            if exe.exists():
                exe_label = friendly_path_display(exe, keep_parts=3)
                self.log.emit(f"ℹ️ Chromium: {exe_label}")
                return exe

        if default_path:
            self.log.emit("⚠️ Chromium padrao nao encontrado; usando caminho esperado.")
            return default_path
        self.log.emit("⛔ Chromium nao encontrado no PLAYWRIGHT_BROWSERS_PATH.")
        return None

    def run(self) -> None:
        if self.include_emitidas and self.include_recebidas:
            mode_label = "ambas"
        elif self.include_emitidas:
            mode_label = "emitidas"
        elif self.include_recebidas:
            mode_label = "recebidas"
        else:
            mode_label = "nenhuma"
        self.summary_data = {
            "companies": [],
            "started_at": datetime.now().isoformat(),
            "output_base": str(self.output_base),
            "headless": self.headless,
            "period": {
                "start": self.period_start.isoformat(),
                "end": self.period_end.isoformat(),
            },
            "mode": mode_label,
            "include_emitidas": self.include_emitidas,
            "include_recebidas": self.include_recebidas,
        }
        ensure_openssl_legacy_provider()
        try:
            with sync_playwright() as p:
                for company in self.companies:
                    if self._stop_requested or self.isInterruptionRequested():
                        break
                    try:
                        self._run_company(p, company)
                    except Exception as exc:  # noqa: BLE001
                        folder_label = safe_folder_name(company.name or company.doc or "sem_nome")
                        summary = self._init_company_summary(company, folder_label)
                        summary["errors"].append(str(exc))
                        summary["status"] = "error"
                        summary["finished_at"] = datetime.now().isoformat()
                        for key in ("emitidas", "recebidas"):
                            if summary[key].get("status") == "pending":
                                summary[key]["status"] = "error"
                        self.summary_data["companies"].append(summary)
                        msg_lower = str(exc).lower()
                        if "certificate" in msg_lower or "mac verify" in msg_lower or "client certificate" in msg_lower:
                            self.log.emit(f"[ERRO] {company.name}: falha ao carregar certificado digital ({exc}). Verifique o arquivo/senha do certificado ou use login por senha.")
                        else:
                            self.log.emit(f"[ERRO] {company.name}: falha inesperada - {exc}")
        except Exception as exc:  # noqa: BLE001
            self.log.emit(f"[ERRO] Falha geral: {exc}")
        finally:
            self.summary_data["finished_at"] = datetime.now().isoformat()
            self.summary_data["totals"] = self._compute_totals(self.summary_data["companies"])
            self.summary_data["period_start"] = self.period_start.strftime("%Y-%m-%d")
            self.summary_data["period_end"] = self.period_end.strftime("%Y-%m-%d")
            self.summary_ready.emit(self.summary_data)
            self.finished.emit()

    def _init_section(self, label: str) -> Dict[str, Any]:
        return {
            "label": label,
            "status": "pending",
            "row_count": 0,
            "downloaded": 0,
            "pdf_downloaded": 0,
            "files": [],
            "pdf_files": [],
            "errors": [],
            "notes": [],
            "started_at": "",
            "finished_at": "",
            "total_value": 0.0,
            "excluded_value": 0.0,
            "invalid_notes": [],
            "service_codes": [],
        }

    def _init_company_summary(self, company: Company, folder_label: str) -> Dict[str, Any]:
        return {
            "company_id": company.company_id,
            "name": company.name,
            "doc": format_document(company.doc),
            "folder": folder_label,
            "status": "pending",
            "started_at": datetime.now().isoformat(),
            "finished_at": "",
            "errors": [],
            "emitidas": self._init_section("Emitidas"),
            "recebidas": self._init_section("Recebidas"),
        }

    def _company_status(self, summary: Dict[str, Any]) -> str:
        sections: List[Dict[str, Any]] = []
        if self.include_emitidas:
            sections.append(summary.get("emitidas", {}))
        if self.include_recebidas:
            sections.append(summary.get("recebidas", {}))
        if not sections:
            sections = [summary.get("emitidas", {}), summary.get("recebidas", {})]
        if summary.get("errors"):
            return "error"
        if any(sec.get("status") == "error" for sec in sections):
            return "error"
        if any(sec.get("status") == "partial" for sec in sections):
            return "partial"
        if all(sec.get("status") == "empty" for sec in sections):
            return "empty"
        if all(sec.get("downloaded", 0) == 0 for sec in sections):
            return "empty"
        return "success"

    def _compute_totals(self, companies: List[Dict[str, Any]]) -> Dict[str, Any]:
        totals = {
            "companies": len(companies),
            "by_status": {"success": 0, "partial": 0, "error": 0, "empty": 0},
            "downloads": {"emitidas": 0, "recebidas": 0},
            "amounts": {"emitidas": 0.0, "recebidas": 0.0},
        }
        for comp in companies:
            status = comp.get("status", "")
            if status in totals["by_status"]:
                totals["by_status"][status] += 1
            totals["downloads"]["emitidas"] += comp.get("emitidas", {}).get("downloaded", 0)
            totals["downloads"]["recebidas"] += comp.get("recebidas", {}).get("downloaded", 0)
            totals["amounts"]["emitidas"] += float(comp.get("emitidas", {}).get("total_value", 0) or 0)
            totals["amounts"]["recebidas"] += float(comp.get("recebidas", {}).get("total_value", 0) or 0)
        return totals

    def _merge_service_codes(self, section: Dict[str, Any], xml_path: Path, note_value: Optional[float]) -> None:
        entries = extract_nfse_service_codes(xml_path)
        if not entries:
            return
        bucket = section.setdefault("service_codes", [])
        index: Dict[Tuple[str, str], Dict[str, Any]] = {}
        for item in bucket:
            code = str(item.get("code", "")).strip()
            desc = str(item.get("description", "")).strip()
            index[(code, desc)] = item
        for code, desc in entries:
            code_clean = (code or "").strip()
            desc_clean = (desc or "").strip()
            key = (code_clean, desc_clean)
            item = index.get(key)
            if not item:
                item = {"code": code_clean, "description": desc_clean, "total_value": 0.0}
                index[key] = item
                bucket.append(item)
            current = Decimal(str(item.get("total_value", 0) or 0))
            add_value = Decimal(str(note_value or 0))
            item["total_value"] = float(current + add_value)
        bucket.sort(key=lambda it: (only_digits(str(it.get("code", "")).zfill(6)), str(it.get("description", ""))))

    def _run_company(self, pw, company: Company) -> None:
        doc_fmt = format_document(company.doc)
        folder_label = safe_folder_name(company.name or company.doc or "sem_nome")
        if not self._first_company:
            self.log.emit("")
        self._first_company = False
        auth_label = "certificado" if company.auth_mode == AUTH_CERTIFICATE else "senha"
        self.log.emit(f"🚀 {company.name} ({auth_label})")

        summary = self._init_company_summary(company, folder_label)
        chromium_kwargs: Dict[str, Any] = {"headless": self.headless}
        exe_path = self._resolve_chromium_executable(pw)
        if exe_path:
            chromium_kwargs["executable_path"] = str(exe_path)
        browser = pw.chromium.launch(**chromium_kwargs)
        context_kwargs = {}
        if company.auth_mode == AUTH_CERTIFICATE:
            cert_entry = {
                "origin": LOGIN_ORIGIN,
                "passphrase": company.cert_password or None,
            }
            if company.cert_data:
                cert_entry["pfx"] = company.cert_data
            elif company.cert_path:
                cert_entry["pfxPath"] = str(company.cert_path)
            context_kwargs["client_certificates"] = [cert_entry]
        context = browser.new_context(**context_kwargs)
        page = context.new_page()

        try:
            self._login(page, company)
            company_base = self._company_base_dir(folder_label)
            emitidas_dir = self._section_dir(company_base, "Emitidas")
            recebidas_dir = self._section_dir(company_base, "Recebidas")
            if self.include_emitidas:
                summary["emitidas"] = self._download_xmls(page, "Emitidas", EMITIDAS_URL, emitidas_dir, summary["emitidas"])
            else:
                now = datetime.now().isoformat()
                summary["emitidas"]["status"] = "empty"
                summary["emitidas"]["started_at"] = now
                summary["emitidas"]["finished_at"] = now
                summary["emitidas"]["notes"].append("Nao solicitado neste modo.")
            if self.include_recebidas and not self._stop_requested and not self.isInterruptionRequested():
                summary["recebidas"] = self._download_xmls(page, "Recebidas", RECEBIDAS_URL, recebidas_dir, summary["recebidas"])
            elif not self.include_recebidas:
                now = datetime.now().isoformat()
                summary["recebidas"]["status"] = "empty"
                summary["recebidas"]["started_at"] = now
                summary["recebidas"]["finished_at"] = now
                summary["recebidas"]["notes"].append("Nao solicitado neste modo.")
        except Exception as exc:  # noqa: BLE001
            self.log.emit(f"[ERRO] {company.name}: erro inesperado ({exc})")
            summary["errors"].append(str(exc))
            for key in ("emitidas", "recebidas"):
                if summary[key].get("status") == "pending":
                    summary[key]["status"] = "error"
        finally:
            if self._stop_requested or self.isInterruptionRequested():
                if summary["emitidas"]["status"] == "pending":
                    summary["emitidas"]["status"] = "partial"
                    summary["emitidas"]["notes"].append("Interrompido pelo usuario.")
                if summary["recebidas"]["status"] == "pending":
                    summary["recebidas"]["status"] = "partial"
                    summary["recebidas"]["notes"].append("Interrompido pelo usuario.")
            summary["finished_at"] = datetime.now().isoformat()
            summary["status"] = self._company_status(summary)
            self.summary_data["companies"].append(summary)
            context.close()
            browser.close()

    def _company_base_dir(self, folder_label: str) -> Path:
        # Estrutura vem do dashboard: API (folder_structure) ou Supabase (robots.segment_path). Nada fixo no robô.
        path = self.output_base / folder_label
        if self._central_segment_path:
            for seg in self._central_segment_path:
                path /= seg
        elif self._segment_path_from_dashboard:
            for part in (p.strip() for p in self._segment_path_from_dashboard.split("/") if p.strip()):
                path /= part
        return path

    def _section_dir(self, company_base: Path, section_label: str, target_date: Optional[date] = None) -> Path:
        path = company_base / section_label
        use_year = use_month = use_day = False
        if self._central_date_rule:
            if self._central_date_rule == "year":
                use_year = True
            elif self._central_date_rule == "year_month":
                use_year = use_month = True
            elif self._central_date_rule == "year_month_day":
                use_year = use_month = use_day = True
        else:
            fs = self.folder_structure or {}
            use_year = bool(fs.get("year"))
            use_month = bool(fs.get("month"))
            use_day = bool(fs.get("day"))
        if target_date is None or not (use_year or use_month or use_day):
            return path
        year = f"{target_date.year:04d}"
        month = f"{target_date.month:02d}"
        day = f"{target_date.day:02d}"
        if use_year:
            path /= year
        if use_month:
            path /= month
        if use_day:
            if use_year and not use_month:
                path /= f"{day}-{month}"
            elif use_year and use_month:
                path /= day
            elif use_month:
                path /= day
            else:
                path /= f"{day}-{month}-{year}"
        return path

    def _move_file_to_directory(self, src_path: Path, dest_dir: Path) -> Path:
        try:
            if src_path.parent == dest_dir:
                return src_path
            dest_dir.mkdir(parents=True, exist_ok=True)
            dest_path = dest_dir / src_path.name
            src_path.replace(dest_path)
            return dest_path
        except Exception as exc:  # noqa: BLE001
            self.log.emit(f"â›” Erro ao mover {src_path.name} para pasta de data ({exc})")
            return src_path

    def _login(self, page, company: Company) -> None:
        page.goto(LOGIN_URL, wait_until="networkidle")
        if company.auth_mode == AUTH_CERTIFICATE:
            self._login_with_certificate(page, company)
            return
        page.wait_for_selector("input#Inscricao", timeout=20000)
        page.fill("input#Inscricao", format_document(company.doc))
        page.fill("input#Senha", company.password)
        page.click("button[type='submit']")
        page.wait_for_load_state("networkidle")
        page.wait_for_timeout(2000)
        self.log.emit(f"✅ {company.name}: login concluido")

    def _login_with_certificate(self, page, company: Company) -> None:
        try:
            cert_link = page.wait_for_selector("a.img-certificado", timeout=20000)
        except PlaywrightTimeoutError:
            self.log.emit(f"⚠️ {company.name}: opcao de certificado nao encontrada.")
            raise
        try:
            with page.expect_navigation(wait_until="networkidle"):
                cert_link.click()
        except PlaywrightTimeoutError:
            self.log.emit(f"⚠️ {company.name}: clique em certificado nao navegou.")
            raise
        page.wait_for_load_state("networkidle")
        page.wait_for_timeout(2000)
        self.log.emit(f"✅ {company.name}: login via certificado concluido")

    def _apply_period_filter(self, page, page_name: str) -> None:
        start_str = self.period_start.strftime("%d/%m/%Y")
        end_str = self.period_end.strftime("%d/%m/%Y")
        date_selector = (
            "input[type='date'], input[placeholder*='dd'], input[placeholder*='DD'], "
            "input[aria-label*='Data'], input[name*='data'], input[name*='Data']"
        )
        try:
            page.wait_for_selector(date_selector, timeout=8000)
        except PlaywrightTimeoutError:
            self.log.emit(f"⚠️ {page_name}: filtros de data nao carregaram.")
            return

        date_inputs = page.locator(date_selector)
        count = date_inputs.count()
        if count >= 2:
            try:
                date_inputs.nth(0).fill(start_str)
                date_inputs.nth(1).fill(end_str)
            except Exception as exc:  # noqa: BLE001
                self.log.emit(f"⚠️ {page_name}: nao foi possivel preencher datas ({exc}).")
                return
        else:
            self.log.emit(f"⚠️ {page_name}: campos de data nao identificados para filtrar.")
            return

        filter_btn = page.locator(
            "button:has-text('Filtrar'), "
            "button:has-text('Pesquisar'), "
            "button:has-text('Consultar'), "
            "input[type='submit'][value*='Filtrar'], "
            "input[type='submit'][value*='Pesquisar'], "
            "a:has-text('Filtrar'), "
            "a:has-text('Pesquisar')"
        )
        try:
            if filter_btn.count() > 0:
                filter_btn.first.click()
                self.log.emit(f"ℹ️ {page_name}: filtro {start_str}–{end_str}")
            else:
                self.log.emit(f"ℹ️ {page_name}: datas preenchidas (filtro sem botao).")
        except Exception as exc:  # noqa: BLE001
            self.log.emit(f"⚠️ {page_name}: filtro nao aplicado ({exc}).")

        page.wait_for_load_state("networkidle")
        page.wait_for_timeout(1500)

    def _advance_to_next_page(self, page, page_name: str) -> bool:
        pagination = page.locator("ul.pagination")
        if pagination.count():
            active_li = pagination.locator("li.active")
            if active_li.count():
                next_li = active_li.first.locator("xpath=following-sibling::li[not(contains(@class,'disabled'))][1]")
                if next_li.count():
                    next_target = next_li.first.locator("a")
                    if next_target.count():
                        return self._click_next_link(next_target.first, page, page_name)

        selectors = [
            "a[rel='next']",
            "a:has-text('Próxima')",
            "a:has-text('Proxima')",
            "a:has-text('Next')",
            "a:has-text('>')",
            "a:has-text('»')",
            "button:has-text('Próxima')",
            "button:has-text('Next')",
            "li.next:not(.disabled) a",
        ]
        for sel in selectors:
            locator = page.locator(sel)
            if locator.count() == 0:
                continue
            target = locator.first
            try:
                if not target.is_visible() or not target.is_enabled():
                    continue
            except Exception:
                continue
            if self._is_disabled(target):
                continue
            if self._click_next_link(target, page, page_name):
                return True
        return False

    def _is_disabled(self, target) -> bool:
        for attr in ("disabled", "aria-disabled"):
            try:
                value = target.get_attribute(attr)
            except Exception:
                value = None
            if value and value.strip().lower() in {"true", "disabled"}:
                return True
        return False

    def _click_next_link(self, target, page, page_name: str) -> bool:
        try:
            target.click()
        except Exception as exc:  # noqa: BLE001
            self.log.emit(f"⚠️ {page_name}: falha ao avançar para a próxima página ({exc}).")
            return False
        try:
            page.wait_for_load_state("networkidle", timeout=8000)
        except PlaywrightTimeoutError:
            pass
        page.wait_for_timeout(1200)
        self.log.emit(f"➡️ {page_name}: próxima página")
        return True

    def _row_status_from_dom(self, row: Locator) -> Optional[str]:
        def match_text(value: Optional[str]) -> Optional[str]:
            if not value:
                return None
            text = value.strip().lower()
            if not text:
                return None
            if "cancel" in text:
                return "cancelled"
            if "substitu" in text:
                return "substituted"
            return None

        try:
            situacao_attr = row.get_attribute("data-situacao")
        except Exception:
            situacao_attr = None
        status = match_text(situacao_attr)
        if status:
            return status

        try:
            situ_cell = row.locator("td.td-situacao")
            if situ_cell.count():
                status = match_text(situ_cell.first.text_content())
                if status:
                    return status
                icon = situ_cell.first.locator("img")
                if icon.count():
                    for attr_name in ("data-original-title", "title", "alt"):
                        status = match_text(icon.first.get_attribute(attr_name))
                        if status:
                            return status
        except Exception:
            pass
        return None

    def _download_xmls(self, page, page_name: str, url: str, output_dir: Path, section: Dict[str, Any]) -> Dict[str, Any]:
        section["label"] = page_name
        section["started_at"] = datetime.now().isoformat()
        section.setdefault("pdf_downloaded", 0)
        section.setdefault("pdf_files", [])
        section.setdefault("total_value", 0.0)
        section.setdefault("invalid_notes", [])
        section.setdefault("excluded_value", 0.0)
        section.setdefault("service_codes", [])
        try:
            page.goto(url, wait_until="networkidle")
            page.wait_for_timeout(1500)
            self._apply_period_filter(page, page_name)
            try:
                page.wait_for_selector("table.table, span.sem-registros", timeout=15000)
            except PlaywrightTimeoutError:
                if page.locator("text='Nenhum registro encontrado'").count() > 0:
                    self.log.emit(f"ℹ️ {page_name}: nenhum registro.")
                    section["status"] = "empty"
                    section["notes"].append("Nenhum registro encontrado.")
                    return section
                self.log.emit(f"⚠️ {page_name}: nenhum painel encontrado na pagina.")
                section["status"] = "error"
                section["errors"].append("Painel nao encontrado na pagina.")
                return section

            if page.locator("span.sem-registros").count() > 0 or page.locator("text='Nenhum registro encontrado'").count() > 0:
                self.log.emit(f"ℹ️ {page_name}: nenhum registro.")
                section["status"] = "empty"
                section["notes"].append("Nenhum registro encontrado.")
                return section

            rows_selector = "table.table tbody tr"
            total_rows = 0
            page_index = 0
            processed_rows = 0
            output_dir.mkdir(parents=True, exist_ok=True)
            section_root = output_dir
            company_root = section_root.parent
            section_label = section_root.name
            while True:
                if self._stop_requested or self.isInterruptionRequested():
                    section["notes"].append("Interrompido pelo usuario.")
                    break
                rows = page.locator(rows_selector)
                row_count = rows.count()
                page_index += 1

                if row_count == 0:
                    if total_rows == 0:
                        self.log.emit(f"ℹ️ {page_name}: nenhum registro.")
                        section["status"] = "empty"
                        section["notes"].append("Nenhum registro listado.")
                        section["row_count"] = 0
                        return section
                    break

                total_rows += row_count
                self.log.emit(f"📥 {page_name} (p{page_index}): {row_count} notas → {output_dir.name}")

                for idx in range(row_count):
                    if self._stop_requested or self.isInterruptionRequested():
                        section["notes"].append("Interrompido pelo usuario.")
                        break
                    row_seq = processed_rows + 1
                    row_errors: List[str] = []
                    xml_path: Optional[Path] = None
                    pdf_path: Optional[Path] = None
                    xml_attempt: Optional[int] = None
                    pdf_attempt: Optional[int] = None
                    row_status_hint = self._row_status_from_dom(rows.nth(idx))
                    emission_date: Optional[date] = None
                    final_dir: Optional[Path] = None

                    def attempt_download(kind: str, target, fallback_name: str, row_errors: List[str]) -> Tuple[Optional[Path], Optional[int]]:
                        max_attempts = 3
                        last_error: Optional[str] = None
                        attempt_errors: List[str] = []
                        def compact_exc(exc: Exception) -> str:
                            text = str(exc)
                            if "Call log:" in text:
                                text = text.split("Call log:", 1)[0].strip()
                            if "\n" in text:
                                text = text.splitlines()[0].strip()
                            if text:
                                return f"{exc.__class__.__name__}: {text}"
                            return exc.__class__.__name__
                        for attempt in range(1, max_attempts + 1):
                            try:
                                if target.count() == 0:
                                    msg = f"{page_name} #{row_seq}: link {kind} nao encontrado no DOM."
                                    self.log.emit(f"[WARN] {msg}")
                                    attempt_errors.append(msg)
                                    break
                                href = target.first.get_attribute("href")
                                if not href:
                                    msg = f"{page_name} #{row_seq}: link {kind} sem href (tentativa {attempt})."
                                    self.log.emit(f"[WARN] {msg}")
                                    attempt_errors.append(msg)
                                    continue
                                url = urljoin(page.url, href)
                                if kind == "XML":
                                    if attempt == 1:
                                        timeout_ms = 1000
                                    elif attempt == 2:
                                        timeout_ms = 1500
                                    else:
                                        timeout_ms = 2500
                                else:
                                    if attempt == 1:
                                        timeout_ms = 2500
                                    elif attempt == 2:
                                        timeout_ms = 4000
                                    else:
                                        timeout_ms = 6000
                                try:
                                    resp = page.request.get(url, timeout=timeout_ms)
                                except Exception as exc:  # noqa: BLE001
                                    last_error = (
                                        f"{page_name} #{row_seq}: erro ao baixar {kind} via request "
                                        f"(tentativa {attempt}/{max_attempts}) ({compact_exc(exc)})"
                                    )
                                    self.log.emit(f"[ERRO] {last_error}")
                                    continue
                                if not resp.ok:
                                    msg = f"{page_name} #{row_seq}: resposta HTTP {resp.status} ao baixar {kind} (tentativa {attempt})."
                                    self.log.emit(f"[WARN] {msg}")
                                    attempt_errors.append(msg)
                                    continue
                                try:
                                    body = resp.body()
                                except Exception as exc:  # noqa: BLE001
                                    last_error = (
                                        f"{page_name} #{row_seq}: falha ao ler conteudo {kind} "
                                        f"(tentativa {attempt}/{max_attempts}) ({compact_exc(exc)})"
                                    )
                                    self.log.emit(f"[ERRO] {last_error}")
                                    continue
                                sample = body.lstrip()[:500].lower()
                                if kind == "PDF" and not sample.startswith(b"%pdf"):
                                    msg = f"{page_name} #{row_seq}: conteudo inesperado ao baixar PDF (tentativa {attempt})."
                                    self.log.emit(f"[WARN] {msg}")
                                    attempt_errors.append(msg)
                                    continue
                                if kind == "XML" and b"<nfse" not in sample and b"<?xml" not in sample:
                                    msg = f"{page_name} #{row_seq}: conteudo inesperado ao baixar XML (tentativa {attempt})."
                                    self.log.emit(f"[WARN] {msg}")
                                    attempt_errors.append(msg)
                                    continue
                                final_path = output_dir / fallback_name
                                final_path.write_bytes(body)
                                return final_path, attempt
                            except Exception as exc:  # noqa: BLE001
                                last_error = (
                                    f"{page_name} #{row_seq}: erro ao baixar {kind} "
                                    f"(tentativa {attempt}/{max_attempts}) ({compact_exc(exc)})"
                                )
                                self.log.emit(f"[ERRO] {last_error}")
                        menu_attempts = 2
                        for menu_attempt in range(1, menu_attempts + 1):
                            try:
                                target_row = rows.nth(idx)
                                menu_button = target_row.locator(".menu-suspenso-tabela .icone-trigger")
                                if menu_button.count() == 0:
                                    msg = f"{page_name} #{row_seq}: menu nao encontrado para {kind}."
                                    self.log.emit(f"[WARN] {msg}")
                                    attempt_errors.append(msg)
                                    break
                                menu_button.click()
                                try:
                                    target.first.wait_for(state="visible", timeout=3000)
                                except PlaywrightTimeoutError:
                                    msg = f"{page_name} #{row_seq}: menu nao abriu para {kind} (tentativa {menu_attempt})."
                                    self.log.emit(f"[WARN] {msg}")
                                    attempt_errors.append(msg)
                                    continue
                                with page.expect_download(timeout=45000) as dl_info:
                                    target.first.click()
                                    download = dl_info.value
                                    suggested = download.suggested_filename or fallback_name
                                    final_path = output_dir / suggested
                                    download.save_as(final_path)
                                    return final_path, menu_attempt
                            except Exception as exc:  # noqa: BLE001
                                last_error = (
                                    f"{page_name} #{row_seq}: erro ao baixar {kind} pelo menu "
                                    f"(tentativa {menu_attempt}/{menu_attempts}) ({compact_exc(exc)})"
                                )
                                self.log.emit(f"[ERRO] {last_error}")
                        if last_error:
                            row_errors.extend(attempt_errors or [last_error])
                        elif attempt_errors:
                            row_errors.extend(attempt_errors)
                        return None, None
                    stop_now = False
                    row = rows.nth(idx)
                    download_link = row.locator("a:has-text('Download XML')")
                    pdf_link = row.locator(
                        "a:has-text('Download PDF'), "
                        "a:has-text('Download Danfe'), "
                        "a:has-text('Baixar PDF'), "
                        "a:has-text('Imprimir'), "
                        "a:has-text('DANF'), "
                        "a[href*='pdf'], "
                        "a[href*='danf']"
                    )
                    access_key = None
                    try:
                        if download_link.count():
                            href = download_link.first.get_attribute("href")
                            if href:
                                match = re.search(r"/Download/(?:NFSe|DANFSe)/([^/?#]+)", href)
                                if match:
                                    access_key = match.group(1)
                        if not access_key and pdf_link.count():
                            href = pdf_link.first.get_attribute("href")
                            if href:
                                match = re.search(r"/Download/(?:NFSe|DANFSe)/([^/?#]+)", href)
                                if match:
                                    access_key = match.group(1)
                    except Exception:
                        access_key = None
                    if not access_key:
                        access_key = f"{page_name.lower()}_{row_seq}"
                    while not xml_path:
                        if self._stop_requested or self.isInterruptionRequested():
                            section["notes"].append("Interrompido pelo usuario.")
                            stop_now = True
                            break
                        xml_path, xml_attempt = attempt_download(
                            "XML",
                            download_link,
                            f"{access_key}.xml",
                            row_errors,
                        )
                        if not xml_path:
                            self.log.emit(f"[WARN] {page_name} #{row_seq}: XML nao baixou, tentando novamente.")
                    if stop_now:
                        break
                    section["downloaded"] += 1
                    emission_date = _extract_xml_emission_date(xml_path)
                    final_dir = self._section_dir(company_root, section_label, emission_date)
                    xml_path = self._move_file_to_directory(xml_path, final_dir)
                    section["files"].append(xml_path.name)
                    val = extract_nfse_value(xml_path)
                    status_info = _inspect_nfse_status(xml_path)
                    if row_status_hint:
                        status_info["status"] = row_status_hint
                    invalid_status = status_info.get("status") in {"cancelled", "substituted"}
                    note_number = status_info.get("note_number") or xml_path.stem
                    if invalid_status:
                        excluded_value = float(section.get("excluded_value", 0) or 0.0)
                        excluded_value += float(val or 0.0)
                        section["excluded_value"] = excluded_value
                        section["invalid_notes"].append(
                            {
                                "section": section.get("label") or page_name,
                                "status": status_info.get("status"),
                                "note_number": note_number,
                                "access_key": status_info.get("access_key"),
                                "value": float(val or 0.0),
                            }
                        )
                        status_label_pt = "cancelada" if status_info.get("status") == "cancelled" else "substituida"
                        access_key_display = status_info.get("access_key") or "-"
                        self.log.emit(
                            f"[WARN] {page_name} #{row_seq} {note_number}: nota {status_label_pt} (chave {access_key_display}) excluida do total."
                        )
                    elif val is not None:
                        try:
                            section["total_value"] = float(
                                Decimal(str(section.get("total_value", 0) or 0)) + Decimal(str(val))
                            )
                        except Exception:
                            pass
                    if not invalid_status:
                        self._merge_service_codes(section, xml_path, val)
                    while not pdf_path:
                        if self._stop_requested or self.isInterruptionRequested():
                            section["notes"].append("Interrompido pelo usuario.")
                            stop_now = True
                            break
                        pdf_path, pdf_attempt = attempt_download(
                            "PDF",
                            pdf_link,
                            f"{access_key}.pdf",
                            row_errors,
                        )
                        if not pdf_path:
                            self.log.emit(f"[WARN] {page_name} #{row_seq}: PDF nao baixou, tentando novamente.")
                    if stop_now:
                        break
                    pdf_target = final_dir or section_root
                    pdf_path = self._move_file_to_directory(pdf_path, pdf_target)
                    section["pdf_downloaded"] += 1
                    section["pdf_files"].append(pdf_path.name)
                    if xml_path and pdf_path:
                        note_label = (xml_path or pdf_path).stem  # type: ignore[union-attr]
                        parts: List[str] = []
                        if xml_path:
                            attempt_info = f" (tentativa {xml_attempt})" if xml_attempt and xml_attempt > 1 else ""
                            parts.append(f"XML{attempt_info}")
                        if pdf_path:
                            attempt_info = f" (tentativa {pdf_attempt})" if pdf_attempt and pdf_attempt > 1 else ""
                            parts.append(f"PDF{attempt_info}")
                        details = "/".join(parts) if parts else ""
                        attempt_tags: List[str] = []
                        if xml_path and xml_attempt and xml_attempt > 1:
                            attempt_tags.append(f"xml t{xml_attempt}")
                        if pdf_path and pdf_attempt and pdf_attempt > 1:
                            attempt_tags.append(f"pdf t{pdf_attempt}")
                        attempt_info = f" ({', '.join(attempt_tags)})" if attempt_tags else ""
                        self.log.emit(f"✅ {page_name} #{row_seq} {note_label} ({details}{attempt_info})")
                        processed_rows += 1

                    if row_errors:
                        section["errors"].extend(row_errors)

                if self._stop_requested or self.isInterruptionRequested():
                    break
                if not self._advance_to_next_page(page, page_name):
                    break
            section["row_count"] = total_rows

            if section["errors"] and section["downloaded"] > 0:
                section["status"] = "partial"
            elif section["errors"]:
                section["status"] = "error"
            elif section["downloaded"] == 0:
                if self._stop_requested or self.isInterruptionRequested():
                    section["status"] = "partial"
                    section["notes"].append("Interrompido pelo usuario.")
                else:
                    section["status"] = "empty"
                    section["notes"].append("Nenhum registro listado.")
            else:
                if self._stop_requested or self.isInterruptionRequested():
                    section["status"] = "partial"
                    section["notes"].append("Interrompido pelo usuario.")
                else:
                    section["status"] = "success"
            section["finished_at"] = datetime.now().isoformat()
            return section
        except Exception as exc:  # noqa: BLE001
            self.log.emit(f"⛔ {page_name}: erro inesperado ({exc})")
            section["errors"].append(f"Erro inesperado: {exc}")
            section["status"] = "error"
            section["finished_at"] = datetime.now().isoformat()
            return section

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("NFS-e - Download de XML/PDF")
        self.resize(900, 700)
        if LOGO_ICON.exists():
            self.setWindowIcon(QIcon(str(LOGO_ICON)))
        self.setStyleSheet(
            """
            QMainWindow { background:qlineargradient(x1:0,y1:0, x2:1,y2:1, stop:0 #0f1722, stop:1 #111827); }
            QWidget { color:#ECF0F1; }
            QScrollArea { border:none; }
            QLabel { font-weight:bold; font:9pt 'Verdana'; }
            """
        )

        self.companies: List[Dict[str, str]] = []
        self.items: List[CompanyItem] = []
        self.worker: Optional[DownloadThread] = None
        self.output_base: Optional[Path] = get_resolved_output_base()
        api_cfg = get_robot_api_config()
        self._segment_path = (api_cfg.get("segment_path") if api_cfg else None) or os.environ.get("ROBOT_SEGMENT_PATH", "FISCAL/NFS").strip() or "FISCAL/NFS"
        self.report_path: Optional[Path] = None
        self._default_notes_mode: Optional[str] = None
        url, key = get_robot_supabase()
        self._robot_id = register_robot(url or "", key or "")
        if url and key:
            self.companies = load_companies_from_supabase(url, key)
            if self._robot_id:
                # Preferir notes_mode da API (dashboard); senão buscar do Supabase
                if api_cfg and api_cfg.get("notes_mode") in ("recebidas", "emitidas", "both"):
                    self._default_notes_mode = api_cfg["notes_mode"]
                else:
                    cfg = fetch_robot_config(url, key)
                    if cfg:
                        if cfg.get("segment_path"):
                            self._segment_path = cfg["segment_path"]
                        if cfg.get("notes_mode") in ("recebidas", "emitidas", "both"):
                            self._default_notes_mode = cfg["notes_mode"]
            if self.output_base:
                seg_slug = (self._segment_path or "FISCAL/NFS").replace("/", os.sep)
                self.report_path = self.output_base / seg_slug
        else:
            self._robot_id = self._robot_id or None
        self.preferences: Dict[str, Any] = load_path_preferences()
        default_mode = self.preferences.get("default_date_mode", DATE_MODE_PREVIOUS_MONTH)
        if default_mode not in {DATE_MODE_PREVIOUS_MONTH, DATE_MODE_PREVIOUS_DAY}:
            default_mode = DATE_MODE_PREVIOUS_MONTH
        self.default_date_mode = default_mode
        folder_opts = self.preferences.get("folder_structure") or {}
        self.folder_structure = normalizar_estrutura_pastas(folder_opts)
        self.last_summary: Optional[Dict[str, Any]] = None
        self._watermark_path = find_logo_image_path()

        self._robot_supabase_url: Optional[str] = None
        self._robot_supabase_key: Optional[str] = None
        self._current_job_id: Optional[str] = None
        self._current_job: Optional[Dict[str, Any]] = None
        self._heartbeat_timer = QTimer(self)
        self._poll_timer = QTimer(self)
        self._display_config_timer = QTimer(self)
        self._last_display_config_updated_at: Optional[str] = None
        self._heartbeat_timer.timeout.connect(self._on_robot_heartbeat)
        self._poll_timer.timeout.connect(self._on_robot_poll_job)
        self._display_config_timer.timeout.connect(self._on_display_config_poll)
        if self._robot_id:
            self._robot_supabase_url = url or None
            self._robot_supabase_key = key or None
            self._heartbeat_timer.start(30000)
            self._poll_timer.start(5000)
            if url and key:
                self._display_config_timer.start(2000)
                QTimer.singleShot(500, self._on_display_config_poll)
            QTimer.singleShot(2000, self._on_robot_poll_job)
            print("[Robô] Conectado ao painel. Status: ativo.", file=sys.stderr)

        if not url or not key:
            print("[Robô] SUPABASE_URL ou SUPABASE_SERVICE_ROLE_KEY não definidos. Coloque no .env em C:\\Users\\ROBO\\Documents\\ROBOS ou na pasta do bot.", file=sys.stderr)
        elif not self._robot_id:
            print("[Robô] Não foi possível registrar no painel. Veja a mensagem de erro acima.", file=sys.stderr)

        central = QWidget()
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(8)

        # Titulo
        top = QFrame()
        top.setStyleSheet(
            "background:qlineargradient(x1:0,y1:0, x2:1,y2:0, stop:0 #1f2f46, stop:1 #2f3f5b);"
            "border-radius:10px; border:1px solid #22344a;"
        )
        title_shadow = QGraphicsDropShadowEffect(self)
        title_shadow.setBlurRadius(26)
        title_shadow.setColor(Qt.black)
        title_shadow.setOffset(0, 10)
        top.setGraphicsEffect(title_shadow)
        htop = QHBoxLayout(top)
        title = QLabel("NFS-e - Download de XML/PDF")
        title.setStyleSheet("color:#ECF0F1; font:12pt 'Verdana'; font-weight:bold;")
        htop.addWidget(title, alignment=Qt.AlignCenter)
        main_layout.addWidget(top)

        # Busca + botoes de selecao
        controls = QVBoxLayout()
        controls.setSpacing(6)
        top_row = QHBoxLayout()
        top_row.setSpacing(6)
        self.search = QLineEdit()
        self.search.setPlaceholderText("Pesquisar empresa...")
        self.search.setStyleSheet("background:#34495E;color:#ECF0F1;border-radius:6px;padding:6px;font-weight:bold;font:9pt 'Verdana';")
        self.search.textChanged.connect(self._filter_items)
        self.search.setMinimumWidth(0)
        self.search.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        top_row.addWidget(self.search, stretch=2)

        dates_widget = QWidget()
        dates_layout = QHBoxLayout(dates_widget)
        dates_layout.setContentsMargins(0, 0, 0, 0)
        dates_layout.setSpacing(4)
        dates_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        start_label = QLabel("Início")
        start_label.setStyleSheet("font:9pt Verdana;")
        dates_layout.addWidget(start_label)

        self.start_date_edit = QDateEdit()
        self.start_date_edit.setDisplayFormat("dd/MM/yyyy")
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setMinimumWidth(118)
        self.start_date_edit.setMaximumWidth(180)
        self.start_date_edit.setMinimumHeight(32)
        self.start_date_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.start_date_edit.setStyleSheet("background:#34495E;color:#ECF0F1;border-radius:6px;padding:6px;font-weight:bold;font:9pt Verdana;")
        self.start_date_edit.setToolTip("Data inicial do período.")
        dates_layout.addWidget(self.start_date_edit)

        end_label = QLabel("Fim")
        end_label.setStyleSheet("font:9pt Verdana;")
        dates_layout.addWidget(end_label)

        self.end_date_edit = QDateEdit()
        self.end_date_edit.setDisplayFormat("dd/MM/yyyy")
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setMinimumWidth(118)
        self.end_date_edit.setMaximumWidth(180)
        self.end_date_edit.setMinimumHeight(32)
        self.end_date_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.end_date_edit.setStyleSheet("background:#34495E;color:#ECF0F1;border-radius:6px;padding:6px;font-weight:bold;font:9pt Verdana;")
        self.end_date_edit.setToolTip("Data final do período.")
        dates_layout.addWidget(self.end_date_edit)

        self.chk_default_prev_day = None
        self.mode_combo = QComboBox()
        self.mode_combo.addItem("Recebidas (padrão)", {"emitidas": False, "recebidas": True})
        self.mode_combo.addItem("Emitidas", {"emitidas": True, "recebidas": False})
        self.mode_combo.addItem("Emitidas + Recebidas", {"emitidas": True, "recebidas": True})
        if self._default_notes_mode == "emitidas":
            self.mode_combo.setCurrentIndex(1)
        elif self._default_notes_mode == "both":
            self.mode_combo.setCurrentIndex(2)
        else:
            self.mode_combo.setCurrentIndex(0)
        self.mode_combo.setMinimumWidth(140)
        self.mode_combo.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.mode_combo.setStyleSheet("background:#34495E;color:#ECF0F1;border-radius:6px;padding:6px;font-weight:bold;font:9pt Verdana;")
        self.mode_combo.setToolTip("Escolha se baixa somente recebidas, somente emitidas ou ambas.")

        self.start_date_edit.dateChanged.connect(self._on_start_date_changed)
        self.end_date_edit.dateChanged.connect(self._on_end_date_changed)
        self.default_date_mode = DATE_MODE_PREVIOUS_MONTH
        self._apply_default_date_range()

        bottom_row = QHBoxLayout()
        bottom_row.setSpacing(6)
        bottom_row.addWidget(dates_widget, 1)
        bottom_row.addWidget(self.mode_combo)

        bottom_row.addStretch(1)

        controls.addLayout(top_row)
        controls.addLayout(bottom_row)
        main_layout.addLayout(controls)

        # Lista de empresas
        wrap = QFrame()
        wrap.setStyleSheet("background:rgba(17,23,39,0.85);border:1px solid #22344a;border-radius:10px;")
        vwrap = QVBoxLayout(wrap)
        vwrap.setContentsMargins(5, 5, 5, 5)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("border:none;")

        cont = QWidget()
        self.list_layout = QVBoxLayout(cont)
        self.list_layout.setContentsMargins(6, 6, 6, 6)
        self._reload_items()
        self.list_layout.addStretch()

        scroll.setWidget(cont)
        vwrap.addWidget(scroll)
        main_layout.addWidget(wrap, stretch=3)

        # Log
        self.log_frame = LogFrame(self._watermark_path, height=230)
        self.log_frame.setMinimumHeight(120)
        self.log_frame.setMaximumHeight(16777215)
        main_layout.addWidget(self.log_frame, stretch=2)

        # Headless toggle
        self.chk_headless = QCheckBox()
        self.chk_headless.setChecked(DEFAULT_HEADLESS)
        self.chk_headless.setStyleSheet("color:#ECF0F1;font:10pt Verdana;")
        self._update_headless_label(DEFAULT_HEADLESS)
        self.chk_headless.toggled.connect(self._update_headless_label)
        main_layout.addWidget(self.chk_headless)

        # Rodape botoes
        hbottom = QHBoxLayout()
        hbottom.setSpacing(10)
        self.btn_start = QPushButton("Iniciar downloads")
        self.btn_start.setStyleSheet(button_style("#27AE60", "#2ECC71", "#1E8449"))
        icon_start = load_png_icon("iniciar")
        if icon_start:
            self.btn_start.setIcon(icon_start)
        attach_glow(self.btn_start, "#2ECC71")
        self.btn_start.clicked.connect(self._start)

        self.btn_stop = QPushButton("Parar")
        self.btn_stop.setStyleSheet(button_style("#C0392B", "#E74C3C", "#922B21"))
        icon_stop = load_png_icon("parar")
        if icon_stop:
            self.btn_stop.setIcon(icon_stop)
        attach_glow(self.btn_stop, "#E74C3C")
        self.btn_stop.clicked.connect(self._stop)
        self.btn_stop.setEnabled(False)

        self.btn_mgr = QPushButton("Gerenciador de Empresas")
        self.btn_mgr.setStyleSheet(button_style("#8E44AD", "#9B59B6", "#6C3483"))
        icon_mgr = load_png_icon("gerenciar")
        if icon_mgr:
            self.btn_mgr.setIcon(icon_mgr)
        attach_glow(self.btn_mgr, "#9B59B6")
        self.btn_mgr.clicked.connect(self._open_manager)

        self.btn_clear = QPushButton("Limpar log")
        self.btn_clear.setStyleSheet(button_style("#1ABC9C", "#16A085", "#117A65"))
        icon_clear = load_png_icon("limpar")
        if icon_clear:
            self.btn_clear.setIcon(icon_clear)
        attach_glow(self.btn_clear, "#16A085")
        self.btn_clear.clicked.connect(lambda: self.log_frame.text.clear())

        hbottom.addWidget(self.btn_start)
        hbottom.addWidget(self.btn_stop)
        hbottom.addWidget(self.btn_mgr)
        hbottom.addWidget(self.btn_clear)
        main_layout.addLayout(hbottom)

        self.setCentralWidget(central)
        self._finish_logged = False

        if url and key and not self.companies:
            QTimer.singleShot(100, lambda: self._log(
                "Nenhuma empresa ativa no dashboard ou habilitada para o Robô NFS. "
                "Cadastre empresas no painel e ative para este robô no Gerenciador (ou use certificado na empresa)."
            ))

        self._tray_icon = QSystemTrayIcon(self)
        if LOGO_ICON.exists():
            self._tray_icon.setIcon(QIcon(str(LOGO_ICON)))
        else:
            from PySide6.QtWidgets import QStyle
            self._tray_icon.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon))
        from PySide6.QtWidgets import QMenu
        menu = QMenu()
        show_act = QAction("Abrir janela", self)
        show_act.triggered.connect(self._show_from_tray)
        menu.addAction(show_act)
        quit_act = QAction("Fechar robô", self)
        quit_act.triggered.connect(self._quit_from_tray)
        menu.addAction(quit_act)
        self._tray_icon.setContextMenu(menu)
        self._tray_icon.activated.connect(self._on_tray_activated)
        self._tray_icon.setToolTip("NFS-e - Download de XML/PDF")

    def _show_from_tray(self) -> None:
        self.showNormal()
        self.raise_()
        self.activateWindow()
        if self._robot_id and self._robot_supabase_url and self._robot_supabase_key:
            current_status = "processing" if self.worker and self.worker.isRunning() else "active"
            update_robot_status(self._robot_supabase_url, self._robot_supabase_key, self._robot_id, current_status)
            if not self._heartbeat_timer.isActive():
                self._heartbeat_timer.start(30000)
            if not self._poll_timer.isActive():
                self._poll_timer.start(5000)
            if not self._display_config_timer.isActive():
                self._display_config_timer.start(2000)

    def _on_tray_activated(self, reason: int) -> None:
        if reason == QSystemTrayIcon.DoubleClick:
            self._show_from_tray()
    def _update_headless_label(self, checked: bool) -> None:
        self.chk_headless.setText("Ocultar navegador" if checked else "Mostrar navegador")

    def _resolve_default_range(self, mode: str) -> Tuple[date, date]:
        today = datetime.now().date()
        if mode == DATE_MODE_PREVIOUS_DAY:
            prev_day = today - timedelta(days=1)
            return prev_day, prev_day
        first_of_month = today.replace(day=1)
        last_prev_month = first_of_month - timedelta(days=1)
        first_prev_month = last_prev_month.replace(day=1)
        return first_prev_month, last_prev_month

    def _apply_default_date_range(self) -> None:
        start_date, end_date = self._resolve_default_range(self.default_date_mode)
        start_qdate = QDate(start_date.year, start_date.month, start_date.day)
        end_qdate = QDate(end_date.year, end_date.month, end_date.day)
        self.start_date_edit.blockSignals(True)
        self.end_date_edit.blockSignals(True)
        try:
            # Libera limites atuais para garantir que as novas datas possam ser aplicadas.
            self.end_date_edit.setMinimumDate(QDate(100, 1, 1))
            self.start_date_edit.setMaximumDate(QDate(9999, 12, 31))
            self.start_date_edit.setDate(start_qdate)
            self.end_date_edit.setDate(end_qdate)
        finally:
            self.start_date_edit.blockSignals(False)
            self.end_date_edit.blockSignals(False)
        self.end_date_edit.setMinimumDate(start_qdate)
        self.start_date_edit.setMaximumDate(end_qdate)

    def _on_start_date_changed(self, new_date: QDate) -> None:
        if self.end_date_edit.date() < new_date:
            self.end_date_edit.setDate(new_date)
        self.end_date_edit.setMinimumDate(new_date)

    def _on_end_date_changed(self, new_date: QDate) -> None:
        if self.start_date_edit.date() > new_date:
            self.start_date_edit.setDate(new_date)
        self.start_date_edit.setMaximumDate(new_date)

    def _reload_items(self) -> None:
        # Limpa layout atual (sem remover o stretch final)
        while self.list_layout.count():
            item = self.list_layout.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()
        self.items = []
        for c in self.companies:
            it = CompanyItem(normalize_company_name(c.get("name", "")), c.get("doc", ""), c.get("auth_mode", AUTH_PASSWORD))
            self.items.append(it)
            self.list_layout.addWidget(it)
        self.list_layout.addStretch()

    def _filter_items(self, text: str) -> None:
        low = text.lower()
        for it in self.items:
            it.setVisible(low in it.label.text().lower())

    def _select_all(self) -> None:
        for it in self.items:
            it.setVisible(True)

    def _deselect_all(self) -> None:
        pass

    # ------------------------------------------------------------------
    # CRUD de empresas
    # ------------------------------------------------------------------
    def _open_manager(self) -> None:
        while True:
            dlg = ManagerDialog(self)
            if dlg.exec() != QDialog.Accepted or not dlg.choice:
                break
            if dlg.choice == "path":
                self._choose_output_base()

    def _choose_output_base(self) -> None:
        current = self.output_base or BASE_DIR
        current_report = self.report_path or current
        dlg = PathDialog(
            current, current_report, self.folder_structure, self, self.preferences,
            read_only=True, segment_path=self._segment_path,
        )
        dlg.exec()
        # Estrutura é somente leitura; path e relatório vêm do departamento/dashboard.

    def _import_companies_excel(self, substituir_lista: bool = True) -> None:
        try:
            from openpyxl import load_workbook
        except ImportError:
            QMessageBox.warning(
                self,
                "Importar Excel",
                "Biblioteca 'openpyxl' nao encontrada. Instale com:\n\npip install openpyxl",
            )
            self._log("Dependencia 'openpyxl' ausente para importar Excel.")
            return

        base_dir = str(self.output_base or BASE_DIR)
        if not os.path.isdir(base_dir):
            base_dir = str(BASE_DIR)
        caminho, _ = QFileDialog.getOpenFileName(
            self,
            "Importar lista de empresas",
            base_dir,
            "Planilha Excel (*.xlsx);;Todos os arquivos (*)",
        )
        if not caminho:
            self._log("Importacao Excel cancelada pelo usuario.")
            return

        try:
            wb = load_workbook(caminho, read_only=True, data_only=True)
            ws = wb.active
        except Exception as exc:
            QMessageBox.critical(self, "Importar Excel", f"Nao foi possivel abrir o arquivo selecionado.\n\nDetalhes: {exc}")
            self._log(f"Falha ao abrir Excel de empresas: {exc}")
            return

        def _digits_from_cell(val: Any, expected_len: Optional[int] = None) -> str:
            if val is None:
                return ""
            was_numeric = isinstance(val, (int, float))
            if isinstance(val, float):
                if abs(val - round(val)) < 1e-6:
                    val = int(round(val))
            txt = str(val).strip()
            if isinstance(val, float) and "e" in txt.lower():
                try:
                    txt = f"{int(val):d}"
                    was_numeric = True
                except Exception:
                    pass
            digits = "".join(ch for ch in txt if ch.isdigit())
            if expected_len and was_numeric and len(digits) == expected_len - 1:
                digits = digits.zfill(expected_len)
            return digits

        def _text(val: Any) -> str:
            return str(val).strip() if val is not None else ""

        novas_empresas: List[Dict[str, str]] = []
        avisos: List[str] = []
        vistos: set[str] = set()

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or ()
        header_map: Dict[str, int] = {}
        for idx, col_name in enumerate(header_row):
            key = _text(col_name).strip().lower()
            if key:
                header_map[key] = idx

        idx_nome = header_map.get("nome da empresa", 0)
        idx_cnpj = header_map.get("cnpj", 1)
        idx_modo = header_map.get("modo de acesso")
        idx_senha = header_map.get("senha")
        idx_cert_senha = header_map.get("senha certificado")
        idx_cert_blob = header_map.get("cert_blob_b64")

        def _cell(row: Tuple[Any, ...], idx: Optional[int]) -> Any:
            if idx is None or idx < 0 or idx >= len(row):
                return None
            return row[idx]

        def _parse_auth_mode(value: str, has_blob: bool) -> str:
            txt = (value or "").strip().lower()
            if txt in {"certificado", "certificate", "cert", AUTH_CERTIFICATE}:
                return AUTH_CERTIFICATE
            if txt in {"senha", "usuario/senha", "usuário/senha", "password", AUTH_PASSWORD}:
                return AUTH_PASSWORD
            return AUTH_CERTIFICATE if has_blob else AUTH_PASSWORD

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            nome = _text(_cell(row, idx_nome))
            cnpj_digits = _digits_from_cell(_cell(row, idx_cnpj), expected_len=14)

            if not cnpj_digits:
                avisos.append(f"Linha {idx}: sem CNPJ, ignorada.")
                continue
            if len(cnpj_digits) != 14 or not is_valid_cnpj(cnpj_digits):
                avisos.append(f"Linha {idx}: CNPJ invalido, ignorado.")
                continue
            if cnpj_digits in vistos:
                avisos.append(f"Linha {idx}: CNPJ duplicado, mantida a primeira ocorrencia.")
                continue
            vistos.add(cnpj_digits)

            display_name = normalize_company_name(nome or format_cnpj(cnpj_digits))
            cert_blob_b64 = _text(_cell(row, idx_cert_blob))
            cert_password = _text(_cell(row, idx_cert_senha))
            password = _text(_cell(row, idx_senha))
            auth_mode_txt = _text(_cell(row, idx_modo))
            auth_mode = _parse_auth_mode(auth_mode_txt, bool(cert_blob_b64))
            if auth_mode == AUTH_CERTIFICATE and not cert_blob_b64:
                avisos.append(f"Linha {idx}: modo certificado sem cert_blob_b64; usando senha.")
                auth_mode = AUTH_PASSWORD
            novas_empresas.append(
                {
                    "name": display_name,
                    "doc": cnpj_digits,
                    "password": password if auth_mode == AUTH_PASSWORD else "",
                    "auth_mode": auth_mode,
                    "cert_path": "",
                    "cert_password": cert_password if auth_mode == AUTH_CERTIFICATE else "",
                    "cert_blob_b64": cert_blob_b64 if auth_mode == AUTH_CERTIFICATE else "",
                }
            )

        try:
            wb.close()
        except Exception:
            pass

        if not novas_empresas:
            QMessageBox.warning(self, "Importar Excel", "Nenhuma linha valida (com CNPJ) foi encontrada na planilha.")
            self._log("Importacao Excel ignorada: planilha sem dados validos.")
            return

        if substituir_lista:
            confirm = ConfirmDialog(
                title="Confirmar importacao",
                message=(
                    f"Isto substituira a lista atual ({len(self.companies)} empresas) "
                    f"por {len(novas_empresas)} registros do Excel. Deseja continuar?"
                ),
                primary_text="Substituir",
                secondary_text="Cancelar",
                parent=self,
            )
            confirm.exec()
            if confirm.choice != "primary":
                self._log("Importacao Excel cancelada pelo usuario.")
                return
            empresas_resultantes = novas_empresas
            adicionadas = len(novas_empresas)
            modo_log = "substituir"
        else:
            existentes = {only_digits(c.get("doc", "")) for c in self.companies if only_digits(c.get("doc", ""))}
            novas_validas: List[Dict[str, str]] = []

            for entry in novas_empresas:
                cnpj = only_digits(entry.get("doc", ""))
                if cnpj in existentes:
                    avisos.append(f"CNPJ ja cadastrado, ignorado: {format_cnpj(cnpj)}.")
                    continue
                novas_validas.append(entry)
                existentes.add(cnpj)

            if not novas_validas:
                QMessageBox.information(
                    self,
                    "Importar Excel",
                    "Nenhuma empresa nova para acrescentar. Os CNPJs da planilha ja estao cadastrados.",
                )
                self._log("Importacao Excel cancelada: nenhum CNPJ novo para acrescentar.")
                return

            confirm = ConfirmDialog(
                title="Confirmar importacao",
                message=(
                    f"Isto acrescentara {len(novas_validas)} novas empresas a lista atual "
                    f"({len(self.companies)} empresas). Deseja continuar?"
                ),
                primary_text="Adicionar",
                secondary_text="Cancelar",
                parent=self,
            )
            confirm.exec()
            if confirm.choice != "primary":
                self._log("Importacao Excel cancelada pelo usuario.")
                return

            empresas_resultantes = self.companies + novas_validas
            adicionadas = len(novas_validas)
            modo_log = "acrescentar"

        self.companies = empresas_resultantes
        self._reload_items()

        if avisos:
            avisos_preview = "\n".join(avisos[:5])
            if len(avisos) > 5:
                avisos_preview += "\n..."
            QMessageBox.information(
                self,
                "Importar Excel",
                f"{adicionadas} empresas importadas.\n\nAvisos:\n{avisos_preview}",
            )
        else:
            QMessageBox.information(self, "Importar Excel", f"{adicionadas} empresas importadas com sucesso.")

        self._log(
            f"Importacao via Excel concluida ({modo_log}): {adicionadas} empresas."
            + (f" Avisos: {len(avisos)}." if avisos else "")
        )

    def _add_company(self) -> None:
        dlg = CompanyEditDialog(parent=self)
        if dlg.exec() != QDialog.Accepted:
            return
        data = dlg.get_data()
        self.companies.append(data)
        self._reload_items()
        self._log(f"Empresa adicionada: {data.get('name', '')}.")

    def _edit_company(self) -> None:
        if not self.items:
            return
        choices = []
        for c in self.companies:
            name = normalize_company_name(c.get("name", ""))
            doc_digits = only_digits(c.get("doc", ""))
            display_doc = format_document(doc_digits)
            meta = "Certificado" if c.get("auth_mode") == AUTH_CERTIFICATE else "Usuario/Senha"
            choices.append((name, display_doc, meta))
        dlg_pick = CompanyPickerDialog(choices, "Editar empresa", "Editar", "#2980B9", "#3498DB", "#2471A3", self)
        if dlg_pick.exec() != QDialog.Accepted:
            return
        idx = dlg_pick.get_index()
        if idx is None:
            return
        curr = self.companies[idx]
        dlg = CompanyEditDialog(
            curr.get("name", ""),
            curr.get("doc", ""),
            curr.get("password", ""),
            curr.get("auth_mode", AUTH_PASSWORD),
            curr.get("cert_path", ""),
            curr.get("cert_password", ""),
            curr.get("cert_blob_b64", ""),
            self,
        )
        if dlg.exec() != QDialog.Accepted:
            return
        updated = dlg.get_data()
        self.companies[idx] = updated
        self._reload_items()
        self._log(f"Empresa atualizada: {updated.get('name', '')}.")

    def _delete_company(self) -> None:
        if not self.items:
            return
        choices = []
        for c in self.companies:
            name = normalize_company_name(c.get("name", ""))
            doc_digits = only_digits(c.get("doc", ""))
            display_doc = format_document(doc_digits)
            meta = "Certificado" if c.get("auth_mode") == AUTH_CERTIFICATE else ""
            choices.append((name, display_doc, meta))
        dlg_pick = CompanyPickerDialog(
            choices,
            "Excluir empresa(s)",
            "Excluir selecionadas",
            "#C0392B",
            "#E74C3C",
            "#922B21",
            multi_select=True,
            parent=self,
        )
        if dlg_pick.exec() != QDialog.Accepted:
            return
        selected_indexes = dlg_pick.get_indexes()
        if not selected_indexes:
            idx = dlg_pick.get_index()
            selected_indexes = [idx] if idx is not None else []
        if not selected_indexes:
            return
        names = [choices[i][0] for i in selected_indexes]
        if len(names) == 1:
            msg = f"Excluir '{names[0]}'?"
        else:
            preview = ", ".join(names[:3])
            if len(names) > 3:
                preview += ", ..."
            msg = f"Excluir {len(names)} empresas?\n{preview}"
        confirm = ConfirmDialog(
            title="Confirmar exclusao",
            message=msg,
            primary_text="Excluir",
            secondary_text="Cancelar",
            primary_colors=("#C0392B", "#E74C3C", "#922B21"),
            secondary_colors=("#2980B9", "#3498DB", "#2471A3"),
            parent=self,
        )
        confirm.exec()
        if confirm.choice != "primary":
            return
        removed_names: List[str] = []
        for idx in sorted(selected_indexes, reverse=True):
            if 0 <= idx < len(self.companies):
                removed_names.append(normalize_company_name(self.companies[idx].get("name", "")))
                self.companies.pop(idx)
        self._reload_items()
        if len(removed_names) == 1:
            self._log(f"Empresa removida: {removed_names[0]}.")
        else:
            self._log(f"Empresas removidas: {len(removed_names)}.")

    # ------------------------------------------------------------------
    # Acoes de bot
    # ------------------------------------------------------------------
    def _start(self, triggered_by_schedule: bool = False) -> None:
        self._log("Preparando iniciar downloads...")
        output_base = self.output_base
        if not output_base:
            self._log("Caminho nao definido. Defina o caminho das empresas no Gerenciador.")
            QMessageBox.warning(self, "Caminho nao definido", "Defina o caminho das empresas no gerenciador antes de iniciar.")
            return
        if not output_base.exists():
            self._log(f"Pasta configurada nao existe: {output_base}")
            QMessageBox.warning(self, "Caminho invalido", "A pasta configurada nao existe. Escolha um caminho valido.")
            return

        selected_records: List[Dict[str, str]] = list(self.companies)
        if not selected_records:
            self._log("Nenhuma empresa marcada para o Robô NFS no dashboard. Ative nas empresas no painel.")
            return

        updated_companies = False
        selected: List[Company] = []
        for rec in selected_records:
            name_norm = normalize_company_name(rec.get("name", ""))
            mode = rec.get("auth_mode", AUTH_PASSWORD)
            doc_digits = only_digits(rec.get("doc", ""))
            if rec.get("name") != name_norm:
                rec["name"] = name_norm
                updated_companies = True
            if mode == AUTH_CERTIFICATE:
                cert_path_txt = rec.get("cert_path", "") or ""
                cert_pass = rec.get("cert_password", "") or ""
                cert_blob_b64 = rec.get("cert_blob_b64", "") or ""
                cert_data = None
                if cert_blob_b64:
                    try:
                        cert_data = base64.b64decode(cert_blob_b64)
                    except Exception:
                        self._log(f"Certificado salvo invalido para {rec.get('name', '')}. Selecione novamente.")
                        return
                if not cert_data and cert_path_txt:
                    cert_path = Path(cert_path_txt)
                    if not cert_path.exists():
                        self._log(f"Certificado nao encontrado para {rec.get('name', '')}: {cert_path}")
                        return
                    try:
                        cert_data = cert_path.read_bytes()
                        rec["cert_blob_b64"] = base64.b64encode(cert_data).decode("ascii")
                        updated_companies = True
                    except Exception as exc:  # noqa: BLE001
                        self._log(f"Falha ao ler certificado para {rec.get('name', '')}: {exc}")
                        return
                if not cert_data:
                    self._log(f"Selecione o certificado para {rec.get('name', '')}.")
                    return
                if not cert_pass:
                    self._log(f"Informe a senha do certificado para {rec.get('name', '')}.")
                    return
                selected.append(
                    Company(
                        name=name_norm,
                        doc=doc_digits,
                        company_id=str(rec.get("id") or ""),
                        auth_mode=AUTH_CERTIFICATE,
                        cert_path=Path(cert_path_txt) if cert_path_txt else None,
                        cert_password=cert_pass,
                        cert_data=cert_data,
                    )
                )
            else:
                password = rec.get("password", "") or ""
                if not password:
                    self._log(f"Informe a senha para {rec.get('name', '')}.")
                    return
                selected.append(
                    Company(
                        name=name_norm,
                        doc=doc_digits,
                        company_id=str(rec.get("id") or ""),
                        password=password,
                        auth_mode=AUTH_PASSWORD,
                    )
                )

        start_date = self.start_date_edit.date()
        end_date = self.end_date_edit.date()
        period_start = datetime(start_date.year(), start_date.month(), start_date.day())
        period_end = datetime(end_date.year(), end_date.month(), end_date.day())
        if period_end < period_start:
            self._log("A data final deve ser igual ou posterior à inicial.")
            return
        selected_mode = self.mode_combo.currentData() or {}
        include_emitidas = bool(selected_mode.get("emitidas", False))
        include_recebidas = bool(selected_mode.get("recebidas", True))
        if not include_emitidas and not include_recebidas:
            self._log("Selecione ao menos um tipo de nota para download.")
            return
        period_label = f"{period_start:%d/%m/%Y} a {period_end:%d/%m/%Y}"

        self._finish_logged = False
        self.last_summary = None
        self.btn_start.setEnabled(False)
        self.btn_stop.setEnabled(True)

        if self._robot_id and self._robot_supabase_url and self._robot_supabase_key:
            update_robot_status(self._robot_supabase_url, self._robot_supabase_key, self._robot_id, "processing")
        headless = self.chk_headless.isChecked()
        central_segment_path, central_date_rule = fetch_central_folder_structure(self._segment_path)
        self.worker = DownloadThread(
            selected,
            headless,
            output_base,
            period_start,
            period_end,
            include_emitidas,
            include_recebidas,
            self.folder_structure,
            central_segment_path,
            central_date_rule,
            segment_path_from_dashboard=self._segment_path,
        )
        self.worker.log.connect(self._log)
        self.worker.summary_ready.connect(self._on_summary_ready)
        self.worker.finished.connect(self._on_finished)
        self.worker.start()
        base_label = friendly_path_display(output_base, keep_parts=2)
        mode_label = self.mode_combo.currentText()
        self._log(f"Downloads: {base_label} | {period_label} | Modo: {mode_label}")

    def _stop(self) -> None:
        w = self.worker
        if not w:
            self._log("Nenhum processo em execucao.")
            return
        if w.isRunning():
            w.request_stop()
            w.wait(5000)
            if w.isRunning():
                w.terminate()
                w.wait(2000)
        if self.last_summary is None and hasattr(w, "summary_data"):
            self.last_summary = getattr(w, "summary_data", None)
            _set_last_json_result_summary(self.last_summary)
        if not self._finish_logged:
            self._on_finished()
        self._log("Processo interrompido.")

    def _on_summary_ready(self, summary: Dict[str, Any]) -> None:
        self.last_summary = summary
        _set_last_json_result_summary(summary)

    def _ensure_report_in_vm_base(self, pdf_path: Path) -> Path:
        if not self.output_base:
            return pdf_path
        if self.companies:
            first_name = self.companies[0].get("name") or self.companies[0].get("doc") or "sem_nome"
            folder_label = safe_folder_name(first_name)
            report_dir = self.output_base / folder_label / (self._segment_path or "FISCAL/NFS").replace("/", os.sep)
        else:
            report_dir = self.output_base / (self._segment_path or "FISCAL/NFS").replace("/", os.sep)
        report_dir.mkdir(parents=True, exist_ok=True)
        target_path = report_dir / pdf_path.name
        try:
            if pdf_path.resolve() == target_path.resolve():
                return target_path
        except Exception:
            if str(pdf_path) == str(target_path):
                return target_path
        shutil.copy2(str(pdf_path), str(target_path))
        return target_path

    def _on_finished(self) -> None:
        if self._finish_logged:
            return
        self._finish_logged = True
        self.btn_start.setEnabled(True)
        self.btn_stop.setEnabled(False)
        summary = self.last_summary
        job_id = self._current_job_id
        job = self._current_job
        if summary:
            _set_last_json_result_summary(summary)
        self._current_job_id = None
        self._current_job = None
        self.worker = None
        if self._robot_id and self._robot_supabase_url and self._robot_supabase_key:
            update_robot_status(self._robot_supabase_url, self._robot_supabase_key, self._robot_id, "active")
        if summary and job and self._robot_id and self._robot_supabase_url and self._robot_supabase_key:
            update_robot_last_period_end(
                self._robot_supabase_url,
                self._robot_supabase_key,
                self._robot_id,
                job.get("period_end") or job.get("period_start"),
            )
        if summary and job and self._robot_supabase_url and self._robot_supabase_key:
            try:
                upsert_nfs_stats(
                    self._robot_supabase_url, self._robot_supabase_key, job, summary
                )
                self._log("[NFS] Totais e ranking enviados para o painel.")
            except Exception as exc:  # noqa: BLE001
                self._log(f"[NFS] Falha ao enviar totais para o painel: {exc}")
        elif summary and self._robot_supabase_url and self._robot_supabase_key and not job:
            cfg = fetch_robot_display_config(self._robot_supabase_url, self._robot_supabase_key)
            company_ids = (cfg.get("company_ids") or []) if cfg else []
            period_start = (summary.get("period_start") or summary.get("period", {}).get("start") or "").strip()
            if company_ids and period_start and len(period_start) >= 7:
                try:
                    fake_job = {
                        "company_ids": company_ids,
                        "period_start": period_start[:10],
                        "period_end": (summary.get("period_end") or summary.get("period", {}).get("end") or period_start)[:10],
                    }
                    upsert_nfs_stats(
                        self._robot_supabase_url, self._robot_supabase_key, fake_job, summary
                    )
                    self._log("[NFS] Totais e ranking enviados para o painel (execução manual).")
                except Exception as exc:  # noqa: BLE001
                    self._log(f"[NFS] Falha ao enviar totais para o painel: {exc}")
        if job_id and self._robot_supabase_url and self._robot_supabase_key:
            complete_execution_request(
                self._robot_supabase_url, self._robot_supabase_key, job_id, True, None
            )
        self._log("Processo concluido.")
        if summary:
            try:
                pdf_path = self._generate_pdf_report(summary)
                if pdf_path:
                    pdf_label = friendly_path_display(pdf_path, keep_parts=3)
                    self._log(f"Relatorio salvo em: {pdf_label}")
                    pdf_vm_base = self._ensure_report_in_vm_base(pdf_path)
                    if str(pdf_vm_base) != str(pdf_path):
                        self._log(f"Relatorio copiado para a pasta base da VM: {friendly_path_display(pdf_vm_base, keep_parts=3)}")
                    # Não abrir o PDF automaticamente; o usuário pode abrir pela pasta se quiser.
            except Exception as exc:  # noqa: BLE001
                self._log(f"Falha ao gerar relatorio: {exc}")

    def _on_robot_heartbeat(self) -> None:
        if self._robot_id:
            update_robot_heartbeat(self._robot_supabase_url, self._robot_supabase_key, self._robot_id)

    def _on_display_config_poll(self) -> None:
        if not self._robot_supabase_url or not self._robot_supabase_key or self.worker and self.worker.isRunning():
            return
        cfg = fetch_robot_display_config(self._robot_supabase_url, self._robot_supabase_key)
        if not cfg:
            return
        updated = (cfg.get("updated_at") or "").strip()
        if updated and updated == self._last_display_config_updated_at:
            return
        company_ids = cfg.get("company_ids") or []
        self._last_display_config_updated_at = updated
        if not company_ids:
            self.companies = []
            self._reload_items()
            self.report_path = None
        else:
            records = load_companies_from_supabase_by_ids(
                self._robot_supabase_url, self._robot_supabase_key, company_ids
            )
            self.companies = records
            self._reload_items()
            if self.output_base:
                seg_slug = (self._segment_path or "FISCAL/NFS").replace("/", os.sep)
                self.report_path = self.output_base / seg_slug
        period_start = (cfg.get("period_start") or "").strip() or None
        period_end = (cfg.get("period_end") or "").strip() or None
        if period_start and period_end:
            try:
                start_dt = datetime.strptime(period_start[:10], "%Y-%m-%d")
                end_dt = datetime.strptime(period_end[:10], "%Y-%m-%d")
                self.start_date_edit.blockSignals(True)
                self.end_date_edit.blockSignals(True)
                try:
                    self.start_date_edit.setDate(QDate(start_dt.year, start_dt.month, start_dt.day))
                    self.end_date_edit.setDate(QDate(end_dt.year, end_dt.month, end_dt.day))
                finally:
                    self.start_date_edit.blockSignals(False)
                    self.end_date_edit.blockSignals(False)
            except Exception:
                pass
        notes_mode = (cfg.get("notes_mode") or "").strip()
        if notes_mode in ("emitidas", "both", "recebidas"):
            if notes_mode == "emitidas":
                self.mode_combo.setCurrentIndex(1)
            elif notes_mode == "both":
                self.mode_combo.setCurrentIndex(2)
            else:
                self.mode_combo.setCurrentIndex(0)

    def _on_robot_poll_job(self) -> None:
        if not self._robot_id:
            return
        if self.worker and self.worker.isRunning():
            return
        job = claim_execution_request(
            self._robot_supabase_url, self._robot_supabase_key, self._robot_id, log_callback=self._log
        )
        if job:
            self._log("[Robô] Job da fila (agendador) iniciado.")
            self._run_job(job)

    def _run_job(self, job: Dict[str, Any]) -> None:
        company_ids = job.get("company_ids") or []
        if not company_ids:
            complete_execution_request(
                self._robot_supabase_url, self._robot_supabase_key, job["id"], False, "Nenhuma empresa no job"
            )
            return
        is_scheduled_job = bool(job.get("schedule_rule_id"))
        if is_scheduled_job:
            yesterday_s = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
            period_start_s = yesterday_s
            period_end_s = yesterday_s
            job["period_start"] = yesterday_s
            job["period_end"] = yesterday_s
        else:
            period_start_s = job.get("period_start")
            period_end_s = job.get("period_end")
            if not period_start_s:
                period_start_s = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
            if not period_end_s:
                period_end_s = period_start_s
        try:
            start_dt = datetime.strptime(period_start_s[:10], "%Y-%m-%d")
            end_dt = datetime.strptime(period_end_s[:10], "%Y-%m-%d")
        except Exception:
            start_dt = datetime.now() - timedelta(days=1)
            end_dt = start_dt
        url = self._robot_supabase_url or self.preferences.get("supabase_url") or ""
        key = self._robot_supabase_key or self.preferences.get("supabase_anon_key") or ""
        records = load_companies_from_supabase_by_ids(url, key, company_ids)
        if not records:
            complete_execution_request(
                self._robot_supabase_url, self._robot_supabase_key, job["id"], False, "Nenhuma empresa encontrada"
            )
            self._log(
                "[Robô] Nenhuma empresa com config no painel (company_robot_config) para este robô. "
                "Habilite as empresas no gerenciador e defina auth_mode/senha."
            )
            return
        selected: List[Company] = []
        skipped_companies = 0
        for rec in records:
            name_norm = normalize_company_name(rec.get("name", ""))
            doc_digits = only_digits(rec.get("doc", ""))
            mode = rec.get("auth_mode") or (AUTH_CERTIFICATE if rec.get("cert_blob_b64") else AUTH_PASSWORD)
            if mode not in (AUTH_PASSWORD, AUTH_CERTIFICATE):
                mode = AUTH_PASSWORD
            if mode == AUTH_CERTIFICATE:
                cert_data = None
                if rec.get("cert_blob_b64"):
                    try:
                        cert_data = base64.b64decode(rec["cert_blob_b64"])
                    except Exception:
                        cert_data = None
                if not cert_data:
                    self._log(f"[Robô] Certificado nao configurado para {name_norm}.")
                    skipped_companies += 1
                    continue
                cert_password = (rec.get("cert_password") or "").strip()
                if not cert_password:
                    self._log(f"[Robô] Senha do certificado nao configurada para {name_norm}.")
                    skipped_companies += 1
                    continue
                selected.append(
                    Company(
                        name=name_norm,
                        doc=doc_digits,
                        company_id=str(rec.get("id") or ""),
                        auth_mode=AUTH_CERTIFICATE,
                        cert_password=cert_password,
                        cert_data=cert_data,
                    )
                )
            else:
                password = (rec.get("password") or "").strip()
                if not password:
                    self._log(f"[Robô] Senha nao configurada para {name_norm}.")
                    skipped_companies += 1
                    continue
                selected.append(
                    Company(
                        name=name_norm,
                        doc=doc_digits,
                        company_id=str(rec.get("id") or ""),
                        password=password,
                        auth_mode=AUTH_PASSWORD,
                    )
                )
        if not selected:
            complete_execution_request(
                self._robot_supabase_url,
                self._robot_supabase_key,
                job["id"],
                False,
                "Nenhuma empresa valida" if skipped_companies == 0 else "Nenhuma empresa valida; todas sem configuracao",
            )
            return
        notes_mode = (job.get("notes_mode") or self._default_notes_mode or "recebidas").strip()
        if notes_mode == "emitidas":
            include_emitidas, include_recebidas = True, False
        elif notes_mode == "both":
            include_emitidas, include_recebidas = True, True
        else:
            include_emitidas, include_recebidas = False, True
        try:
            api_cfg = get_robot_api_config(force_refresh=True) or {}
            if api_cfg.get("segment_path"):
                self._segment_path = str(api_cfg.get("segment_path") or "").strip() or self._segment_path
            if api_cfg.get("notes_mode") in ("recebidas", "emitidas", "both"):
                self._default_notes_mode = api_cfg["notes_mode"]
            refreshed_output_base = get_resolved_output_base()
            if refreshed_output_base:
                self.output_base = refreshed_output_base
            self.companies = records
            self._reload_items()
            self.start_date_edit.blockSignals(True)
            self.end_date_edit.blockSignals(True)
            try:
                self.start_date_edit.setDate(QDate(start_dt.year, start_dt.month, start_dt.day))
                self.end_date_edit.setDate(QDate(end_dt.year, end_dt.month, end_dt.day))
            finally:
                self.start_date_edit.blockSignals(False)
                self.end_date_edit.blockSignals(False)
            if notes_mode == "emitidas":
                self.mode_combo.setCurrentIndex(1)
            elif notes_mode == "both":
                self.mode_combo.setCurrentIndex(2)
            else:
                self.mode_combo.setCurrentIndex(0)
            output_base = get_resolved_output_base() or self.output_base
            if not output_base:
                complete_execution_request(
                    self._robot_supabase_url, self._robot_supabase_key, job["id"], False, "Pasta de saida nao configurada"
                )
                self._log("[Robô] Pasta de saida nao configurada para a execucao agendada.")
                return
            try:
                output_base.mkdir(parents=True, exist_ok=True)
            except Exception as exc:
                complete_execution_request(
                    self._robot_supabase_url,
                    self._robot_supabase_key,
                    job["id"],
                    False,
                    f"Falha ao preparar pasta base: {exc}",
                )
                self._log(f"[Robô] Falha ao preparar pasta base vinda do dashboard: {exc}")
                return
            self._current_job_id = job["id"]
            self._current_job = job
            if self._robot_id:
                update_robot_status(self._robot_supabase_url, self._robot_supabase_key, self._robot_id, "processing")
            self._finish_logged = False
            self.last_summary = None
            self.btn_start.setEnabled(False)
            self.btn_stop.setEnabled(True)
            central_segment_path, central_date_rule = fetch_central_folder_structure(self._segment_path)
            headless = self.chk_headless.isChecked()
            self.worker = DownloadThread(
                selected,
                headless,
                output_base,
                start_dt,
                end_dt,
                include_emitidas,
                include_recebidas,
                self.folder_structure,
                central_segment_path,
                central_date_rule,
                segment_path_from_dashboard=self._segment_path,
            )
            self.worker.log.connect(self._log)
            self.worker.summary_ready.connect(self._on_summary_ready)
            self.worker.finished.connect(self._on_finished)
            self.worker.start()
            self._log(f"[Agendador] Iniciando job para {len(selected)} empresa(s), periodo {start_dt:%d/%m/%Y} a {end_dt:%d/%m/%Y}")
        except Exception as exc:  # noqa: BLE001
            self._log(f"[ERRO] Falha ao preparar job do agendador: {exc}")
            complete_execution_request(
                self._robot_supabase_url,
                self._robot_supabase_key,
                job["id"],
                False,
                f"Falha ao preparar job: {exc}",
            )
            self._current_job_id = None
            self._current_job = None
            self.worker = None
            if self._robot_id:
                update_robot_status(self._robot_supabase_url, self._robot_supabase_key, self._robot_id, "active")

    def closeEvent(self, event) -> None:
        if self.worker and self.worker.isRunning():
            self.worker.request_stop()
            if not self.worker.wait(5000):
                self.worker.terminate()
                self.worker.wait(1000)
        event.ignore()
        self.hide()
        self._tray_icon.show()
        # Não marcar como inativo nem parar o heartbeat: o robô continua em segundo plano (bandeja).

    def _quit_from_tray(self) -> None:
        """Chamado ao escolher 'Fechar robô' na bandeja: marca inativo e encerra o app."""
        if self.worker and self.worker.isRunning():
            self.worker.request_stop()
            if not self.worker.wait(5000):
                self.worker.terminate()
                self.worker.wait(1000)
        if self._robot_id:
            update_robot_status(self._robot_supabase_url, self._robot_supabase_key, self._robot_id, "inactive")
        self._heartbeat_timer.stop()
        self._poll_timer.stop()
        self._display_config_timer.stop()
        QApplication.quit()

    def _find_company_record(self, name: str, doc: str) -> Optional[Dict[str, str]]:
        name_norm = normalize_company_name(name)
        for c in self.companies:
            if normalize_company_name(c.get("name", "")) == name_norm and only_digits(c.get("doc", "")) == doc:
                return c
        return None

    def _log(self, msg: str) -> None:
        line = emit_terminal_log(msg)
        append_runtime_log(line)
        self.log_frame.append(self._format_log(line))

    def _format_log(self, msg: str) -> str:
        """
        Converte a linha de log para HTML, usando negrito em linhas-chave.
        Quebra de linha via <br/> para mensagens multi-linha.
        """
        safe = html.escape(msg).replace("\n", "<br/>")
        bold_prefixes = (
            "🚀",
            "✅",
            "⛔",
            "⚠️",
            "ℹ️",
            "📥",
            "Downloads",
            "Processo concluido",
            "Processo interrompido",
            "Selecione pelo menos",
        )
        if msg.startswith(bold_prefixes):
            safe = f"<b>{safe}</b>"
        margin = "10px" if msg.startswith("🚀") else "4px"
        if not msg:
            margin = "8px"
        return f"<div style='margin-top:{margin}'>{safe}</div>"



    def _status_palette(self, status: str):
        return {
            "success": ("Sucesso", "#27AE60"),
            "partial": ("Parcial", "#F1C40F"),
            "error": ("Erro", "#E74C3C"),
            "empty": ("Sem movimento", "#7F8C8D"),
            "pending": ("Pendente", "#95A5A6"),
        }.get(status, (status.title() if status else "", "#95A5A6"))





    def _generate_pdf_report(self, summary: Dict[str, Any]) -> Optional[Path]:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.lib.utils import ImageReader
            from reportlab.platypus import (
                Image as RLImage,
                LongTable,
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )
        except ImportError:
            self._log("Biblioteca reportlab nao encontrada; relatorio nao gerado.")
            return None

        report_dir = self.output_base / (self._segment_path or "FISCAL/NFS").replace("/", os.sep) if self.output_base else (self.report_path or BASE_DIR)
        if not report_dir:
            self._log("Caminho de relatorio nao definido.")
            return None
        report_dir.mkdir(parents=True, exist_ok=True)
        pdf_path = report_dir / "relatorio nfs.pdf"

        palette = {
            "bg": "#f5f7fb",
            "panel": "#ffffff",
            "surface": "#eef2f7",
            "card": "#eaf1ff",
            "line": "#d8e3f5",
            "muted": "#6b7280",
            "ink": "#1f2937",
            "accent": "#2563eb",
            "accent_soft": "#93c5fd",
            "green": "#16a34a",
            "yellow": "#d97706",
            "red": "#dc2626",
        }

        money_icon_path = find_money_image_path()
        money_icon_cell = None
        if money_icon_path and Path(money_icon_path).exists():
            try:
                money_icon_cell = RLImage(str(money_icon_path), width=12, height=12)
            except Exception:
                money_icon_cell = None

        def fmt_currency(val: Optional[float]) -> str:
            try:
                num = Decimal(str(val or 0)).quantize(Decimal("0.01"))
            except Exception:
                num = Decimal("0.00")
            txt = f"{num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            return f"R$ {txt}"

        logo_path = find_logo_image_path()

        def create_page_background(width: int, height: int, logo_file: Optional[Path]) -> Optional[bytes]:
            try:
                if not logo_file or not Path(logo_file).exists():
                    raise FileNotFoundError("logo ausente")
                base = Image.new("RGBA", (width, height), (255, 255, 255, 255))
                draw = ImageDraw.Draw(base)
                top = (245, 249, 255)
                bottom = (224, 236, 255)
                for yy in range(height):
                    ratio = yy / float(max(1, height - 1))
                    r = int(top[0] * (1 - ratio) + bottom[0] * ratio)
                    g = int(top[1] * (1 - ratio) + bottom[1] * ratio)
                    b = int(top[2] * (1 - ratio) + bottom[2] * ratio)
                    draw.line([(0, yy), (width, yy)], fill=(r, g, b, 255))

                logo_img = Image.open(logo_file).convert("RGBA")
                scale = min(width * 0.95 / logo_img.width, height * 0.95 / logo_img.height, 1.4)
                if scale <= 0:
                    raise ValueError("escala invalida")
                lw = max(1, int(logo_img.width * scale))
                lh = max(1, int(logo_img.height * scale))
                logo_img = logo_img.resize((lw, lh), Image.LANCZOS)
                alpha = logo_img.split()[-1].point(lambda a: int(a * 0.55))
                cx, cy = lw / 2.0, lh / 2.0
                max_r = math.hypot(lw, lh) / 2.0 if lw and lh else 1.0
                inner_r = max_r * 0.82
                outer_r = max_r * 0.97
                mask = Image.new("L", (lw, lh), 0)
                pix = mask.load()
                for y in range(lh):
                    dy = (y + 0.5) - cy
                    for x in range(lw):
                        r0 = math.hypot((x + 0.5) - cx, dy)
                        if r0 <= inner_r:
                            val = 255
                        elif r0 >= outer_r:
                            val = 0
                        else:
                            t = (r0 - inner_r) / (outer_r - inner_r)
                            val = int(255 * (1 - t * t))
                        pix[x, y] = val
                grad_x = Image.new("L", (lw, 1), 0)
                gx = grad_x.load()
                edge_start = 0.64
                edge_end = 0.95
                half = max(1.0, cx)
                for x in range(lw):
                    t = abs((x + 0.5 - cx) / half)
                    t = min(1.0, t)
                    if t <= edge_start:
                        fall = 1.0
                    elif t >= edge_end:
                        fall = 0.0
                    else:
                        tt = (t - edge_start) / (edge_end - edge_start)
                        fall = 1.0 - (tt ** 1.4)
                    gx[x, 0] = int(255 * fall)
                grad_x = grad_x.resize((lw, lh))
                grad_y = Image.new("L", (1, lh), 0)
                gy = grad_y.load()
                edge_start_y = 0.6
                edge_end_y = 0.94
                for y in range(lh):
                    t = abs((y + 0.5 - cy) / half)
                    t = min(1.0, t)
                    if t <= edge_start_y:
                        fall = 1.0
                    elif t >= edge_end_y:
                        fall = 0.0
                    else:
                        tt = (t - edge_start_y) / (edge_end_y - edge_start_y)
                        fall = 1.0 - (tt ** 1.4)
                    gy[0, y] = int(255 * fall)
                grad_y = grad_y.resize((lw, lh))
                alpha = ImageChops.multiply(alpha, mask)
                alpha = ImageChops.multiply(alpha, grad_x)
                alpha = ImageChops.multiply(alpha, grad_y)
                alpha = alpha.filter(ImageFilter.GaussianBlur(max(10, int(min(lw, lh) * 0.03))))
                logo_img.putalpha(alpha)
                overlay = Image.new("RGBA", base.size, (0, 0, 0, 0))
                overlay.paste(logo_img, ((width - lw) // 2, (height - lh) // 2), logo_img)
                composed = Image.alpha_composite(base.convert("RGBA"), overlay)
                buf = BytesIO()
                composed.convert("RGB").save(buf, format="PNG")
                return buf.getvalue()
            except Exception:
                return None

        bg_bytes = create_page_background(int(A4[0]), int(A4[1]), logo_path)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Title"],
            fontSize=20,
            leading=24,
            textColor=colors.HexColor(palette["ink"]),
            spaceAfter=4,
        )
        subtitle_style = ParagraphStyle(
            "Subtitle",
            parent=styles["Normal"],
            fontSize=9.5,
            leading=12,
            textColor=colors.HexColor(palette["muted"]),
        )
        section_style = ParagraphStyle(
            "Section",
            parent=styles["Heading2"],
            fontSize=12,
            leading=16,
            textColor=colors.HexColor("#cbd5f5"),
            spaceBefore=8,
            spaceAfter=6,
        )
        label_style = ParagraphStyle(
            "Label",
            parent=styles["Normal"],
            fontSize=8.5,
            leading=11,
            textColor=colors.HexColor(palette["muted"]),
        )
        value_style = ParagraphStyle(
            "Value",
            parent=styles["Normal"],
            fontSize=10.5,
            leading=13,
            textColor=colors.HexColor(palette["ink"]),
        )
        service_desc_style = ParagraphStyle(
            "ServiceDesc",
            parent=styles["Normal"],
            fontSize=7.7,
            leading=9.4,
            textColor=colors.HexColor(palette["ink"]),
            wordWrap="CJK",
            splitLongWords=1,
        )
        service_code_style = ParagraphStyle(
            "ServiceCode",
            parent=styles["Normal"],
            fontSize=8.6,
            leading=10.4,
            textColor=colors.HexColor(palette["ink"]),
        )
        service_value_style = ParagraphStyle(
            "ServiceValue",
            parent=service_code_style,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor(palette["green"]),
        )
        badge_style = ParagraphStyle(
            "Badge",
            parent=styles["Normal"],
            fontSize=9,
            leading=11,
            alignment=1,
            textColor=colors.white,
        )

        def fmt_dt(value: str) -> str:
            try:
                return datetime.fromisoformat(value).strftime("%d/%m/%Y %H:%M:%S")
            except Exception:
                return value or "-"

        def fmt_duration(start: str, end: str) -> str:
            try:
                dt_start = datetime.fromisoformat(start)
                dt_end = datetime.fromisoformat(end)
                seconds = int((dt_end - dt_start).total_seconds())
                hrs, rem = divmod(seconds, 3600)
                mins, sec = divmod(rem, 60)
                if hrs:
                    return f"{hrs}h {mins:02d}m {sec:02d}s"
                if mins:
                    return f"{mins}m {sec:02d}s"
                return f"{sec}s"
            except Exception:
                return "-"

        def pill(label: str, value: str, color: str) -> Table:
            tbl = Table([[Paragraph(label, label_style), Paragraph(value, value_style)]], colWidths=[70, None])
            tbl.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(palette["card"])),
                        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor(palette["line"])),
                        ("LEFTPADDING", (0, 0), (-1, -1), 7),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor(palette["muted"])),
                        ("TEXTCOLOR", (1, 0), (1, -1), colors.HexColor(color)),
                    ]
                )
            )
            return tbl

        def stat_card(title: str, value: str, helper: str, accent: str) -> Table:
            content = Table([[Paragraph(title, label_style)], [Paragraph(value, value_style)], [Paragraph(helper, label_style)]])
            content.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(palette["card"])),
                        ("LEFTPADDING", (0, 0), (-1, -1), 10),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                        ("TOPPADDING", (0, 0), (-1, -1), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ]
                )
            )
            bar_width = 4
            outer = Table([["", content]], colWidths=[bar_width, None])
            outer.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor(accent)),
                        ("BACKGROUND", (1, 0), (1, -1), colors.HexColor(palette["card"])),
                        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor(palette["line"])),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                        ("TOPPADDING", (0, 0), (-1, -1), 0),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                    ]
                )
            )
            return outer

        def badge(text: str, bg: str) -> Table:
            tbl = Table([[Paragraph(text, badge_style)]])
            tbl.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(bg)),
                        ("LEFTPADDING", (0, 0), (-1, -1), 8),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ]
                )
            )
            return tbl

        def money_box(label: str, value: float) -> Table:
            value_html = f"<b><font color='{palette['green']}'>{fmt_currency(value)}</font></b>"
            value_par = Paragraph(value_html, value_style)
            inner_row = [[value_par]]
            col_widths = [None]
            if money_icon_cell:
                inner_row = [[money_icon_cell, value_par]]
                col_widths = [14, None]
            inner = Table(inner_row, colWidths=col_widths)
            inner.setStyle(
                TableStyle(
                    [
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                        ("TOPPADDING", (0, 0), (-1, -1), 0),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                    ]
                )
            )
            box = Table(
                [[Paragraph(label, label_style)], [inner]],
            )
            box.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(palette["panel"])),
                        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor(palette["line"])),
                        ("LEFTPADDING", (0, 0), (-1, -1), 8),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                        ("TOPPADDING", (0, 0), (-1, -1), 6),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )
            return box

        def service_codes_block(section_label: str, items: List[Dict[str, Any]]) -> Table:
            header_style = ParagraphStyle(
                "service_header",
                parent=label_style,
                fontName="Helvetica-Bold",
                textColor=colors.HexColor(palette["ink"]),
            )
            rank_badges = {
                1: ("TOP 1", "#7a5b00", "#fef3c7"),
                2: ("TOP 2", "#334155", "#e2e8f0"),
                3: ("TOP 3", "#7c2d12", "#fed7aa"),
            }

            rows: List[List[Any]] = [
                [
                    Paragraph("Top", header_style),
                    Paragraph("Codigo de servico", header_style),
                    Paragraph("Descricao", header_style),
                    Paragraph("Valor (R$)", header_style),
                ]
            ]

            ranked_items = sorted(
                items or [],
                key=lambda it: (
                    -float(it.get("total_value", 0) or 0.0),
                    only_digits(str(it.get("code", "")).zfill(6)),
                    str(it.get("description", "")),
                ),
            )

            for pos, entry in enumerate(ranked_items, start=1):
                code_raw = str(entry.get("code", "") or "-").strip() or "-"
                code = format_ctribnac_display(code_raw)
                desc = str(entry.get("description", "") or "-").strip() or "-"
                total_value = float(entry.get("total_value", 0) or 0.0)
                value_label = fmt_currency(total_value)
                if pos <= 3:
                    rank_label, rank_fg, _rank_bg = rank_badges[pos]
                else:
                    rank_label, rank_fg = str(pos), palette["muted"]
                rank_html = f"<b><font color='{rank_fg}'>{rank_label}</font></b>"
                rows.append(
                    [
                        Paragraph(rank_html, service_code_style),
                        Paragraph(html.escape(code), service_code_style),
                        Paragraph(html.escape(desc), service_desc_style),
                        Paragraph(value_label, service_value_style),
                    ]
                )
            if len(rows) == 1:
                rows.append(
                    [
                        Paragraph("-", service_code_style),
                        Paragraph("-", service_code_style),
                        Paragraph("Codigo de servico nao identificado.", label_style),
                        Paragraph("R$ 0,00", service_value_style),
                    ]
                )

            wrapper_padding = 16  # left + right
            content_width = max(260, doc.width - wrapper_padding)
            top_col = max(34, min(42, int(content_width * 0.08)))
            code_col = max(68, min(88, int(content_width * 0.16)))
            value_col = max(86, min(112, int(content_width * 0.20)))
            desc_col = max(100, content_width - top_col - code_col - value_col)

            table = LongTable(
                rows,
                colWidths=[top_col, code_col, desc_col, value_col],
                repeatRows=1,
                splitByRow=1,
            )
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(palette["surface"])),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor(palette["panel"]), colors.HexColor(palette["surface"])]),
                        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor(palette["line"])),
                        ("LEFTPADDING", (0, 0), (-1, -1), 6),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("ALIGN", (3, 0), (3, -1), "RIGHT"),
                        ("ALIGN", (0, 0), (0, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ]
                )
            )
            if len(rows) > 1:
                for pos in range(1, min(3, len(rows) - 1) + 1):
                    _label, _fg, bg = rank_badges.get(pos, ("", "", palette["surface"]))
                    table.setStyle(
                        TableStyle(
                            [
                                ("BACKGROUND", (0, pos), (0, pos), colors.HexColor(bg)),
                                ("BOX", (0, pos), (0, pos), 0.3, colors.HexColor(palette["line"])),
                            ]
                        )
                    )
            wrapper = Table(
                [
                    [Paragraph(f"Codigos de servico - {section_label}", label_style)],
                    [table],
                ],
                colWidths=[doc.width],
            )
            wrapper.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(palette["panel"])),
                        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor(palette["line"])),
                        ("LEFTPADDING", (0, 0), (-1, -1), 8),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                        ("TOPPADDING", (0, 0), (-1, -1), 6),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )
            return wrapper

        def company_separator(idx: int, total: int, comp: Dict[str, Any]) -> Table:
            title_html = (
                f"<b>Empresa {idx + 1} de {total}</b>  "
                f"<font color='{palette['muted']}'>{html.escape(comp.get('name', ''))}</font>"
            )
            tbl = Table([[Paragraph(title_html, label_style)]], colWidths=[doc.width])
            tbl.setStyle(
                TableStyle(
                    [
                        ("LINEABOVE", (0, 0), (-1, 0), 0.8, colors.HexColor(palette["accent_soft"])),
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fbff")),
                        ("LEFTPADDING", (0, 0), (-1, -1), 8),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )
            return tbl

        def company_card(comp: Dict[str, Any]) -> Table:
            status_label, status_color = self._status_palette(comp.get("status"))
            emitidas = comp.get("emitidas", {})
            recebidas = comp.get("recebidas", {})
            emitidas_ok = emitidas.get("downloaded", 0)
            recebidas_ok = recebidas.get("downloaded", 0)
            emitidas_total = emitidas.get("row_count", 0)
            recebidas_total = recebidas.get("row_count", 0)
            emitidas_val = float(emitidas.get("total_value", 0) or 0)
            recebidas_val = float(recebidas.get("total_value", 0) or 0)
            notes_errors = (emitidas.get("notes", []) + emitidas.get("errors", []) + recebidas.get("notes", []) + recebidas.get("errors", [])) or []
            invalid_entries: List[str] = []
            for label, section_data in (("Emitidas", emitidas), ("Recebidas", recebidas)):
                for note in section_data.get("invalid_notes", []) or []:
                    number = note.get("note_number") or "-"
                    access_key = note.get("access_key") or "-"
                    status_pt = "cancelada" if note.get("status") == "cancelled" else "substituída"
                    invalid_entries.append(f"{label}: {number} - {access_key} ({status_pt})")
            if invalid_entries:
                notes_errors.append("Notas canceladas/substituídas:")
                notes_errors.extend(f"  {entry}" for entry in invalid_entries)

            header = Table(
                [
                    [
                        Paragraph(
                            f"<b>{html.escape(comp.get('name', ''))}</b><br/><font size=8.5 color='{palette['muted']}'>{html.escape(comp.get('doc', '-'))}</font>",
                            value_style,
                        ),
                        badge(status_label, status_color),
                    ]
                ],
                colWidths=[doc.width - 110, 85],
            )
            header.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(palette["surface"])),
                        ("LEFTPADDING", (0, 0), (-1, -1), 8),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                        ("TOPPADDING", (0, 0), (-1, -1), 6),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )

            stats = Table(
                [
                    [
                        pill("Emitidas", f"{emitidas_ok}/{emitidas_total}", palette["ink"]),
                        pill("Recebidas", f"{recebidas_ok}/{recebidas_total}", palette["ink"]),
                        pill("Status", status_label, status_color),
                    ]
                ],
                colWidths=[doc.width / 3 - 12] * 3,
            )
            stats.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))

            money = Table(
                [
                    [
                        money_box("Faturamento (Emitidas)", emitidas_val),
                        money_box("Servicos tomados", recebidas_val),
                    ]
                ],
                colWidths=[doc.width / 2 - 12, doc.width / 2 - 12],
            )
            money.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))

            meta = Table(
                [
                    [Paragraph("Pasta", label_style), Paragraph(html.escape(comp.get("folder", "-")), value_style)],
                    [Paragraph("Iniciado", label_style), Paragraph(fmt_dt(comp.get("started_at")), value_style)],
                    [Paragraph("Finalizado", label_style), Paragraph(fmt_dt(comp.get("finished_at")), value_style)],
                ],
                colWidths=[80, doc.width - 80],
            )
            meta.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(palette["panel"])),
                        ("LEFTPADDING", (0, 0), (-1, -1), 8),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )

            notes_block = None
            if notes_errors:
                text_notes = "<br/>".join([f"- {html.escape(str(n))}" for n in notes_errors])
                notes_block = Table([[Paragraph(text_notes, label_style)]])
                notes_block.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(palette["surface"])),
                            ("LEFTPADDING", (0, 0), (-1, -1), 8),
                            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                            ("TOPPADDING", (0, 0), (-1, -1), 4),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                            ("BOX", (0, 0), (-1, -1), 0.35, colors.HexColor(palette["line"])),
                        ]
                    )
                )

            rows = [[header], [stats], [money], [meta]]
            if notes_block:
                rows.append([notes_block])
            card = Table(rows)
            card.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(palette["panel"])),
                        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor(palette["line"])),
                        ("TOPPADDING", (0, 0), (-1, -1), 2),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )
            return card

        def company_service_sections(comp: Dict[str, Any]) -> List[Any]:
            emitidas_codes = comp.get("emitidas", {}).get("service_codes", []) or []
            recebidas_codes = comp.get("recebidas", {}).get("service_codes", []) or []
            blocks: List[Any] = []
            blocks.append(service_codes_block("Emitidas", emitidas_codes))
            blocks.append(Spacer(1, 4))
            blocks.append(service_codes_block("Recebidas", recebidas_codes))
            return blocks

        started_txt = summary.get("started_at", "")
        finished_txt = summary.get("finished_at", "")
        totals = summary.get("totals", {})
        by_status = totals.get("by_status", {})
        downloads = totals.get("downloads", {})
        companies = summary.get("companies", [])

        def decorate(canvas_obj, doc_obj) -> None:
            canvas_obj.saveState()
            if bg_bytes:
                try:
                    canvas_obj.drawImage(ImageReader(BytesIO(bg_bytes)), 0, 0, width=A4[0], height=A4[1], mask="auto")
                except Exception:
                    pass
            canvas_obj.setFillColor(colors.HexColor(palette["bg"]))
            canvas_obj.rect(0, A4[1] - 34, A4[0], 34, stroke=0, fill=1)
            canvas_obj.setFillColor(colors.HexColor(palette["ink"]))
            canvas_obj.setFont("Helvetica-Bold", 10)
            canvas_obj.drawString(doc_obj.leftMargin, A4[1] - 20, "NFS-e | Relatorio de downloads")
            canvas_obj.setFillColor(colors.HexColor(palette["muted"]))
            canvas_obj.setFont("Helvetica", 8)
            canvas_obj.drawString(doc_obj.leftMargin, 20, f"Pagina {doc_obj.page}")
            canvas_obj.setFillColor(colors.HexColor(palette["accent"]))
            canvas_obj.rect(A4[0] - 28 * mm, 16, 16 * mm, 3 * mm, stroke=0, fill=1)
            canvas_obj.restoreState()

        doc = SimpleDocTemplate(
            str(pdf_path),
            pagesize=A4,
            rightMargin=24,
            leftMargin=24,
            topMargin=60,
            bottomMargin=40,
        )

        elements: List[Any] = []

        logo_cell: Optional[Any] = None
        if LOGO_ICON.exists():
            try:
                logo_cell = RLImage(str(LOGO_ICON), width=36, height=36)
            except Exception:
                logo_cell = None

            hero = Table(
                [
                    [
                        logo_cell or Spacer(1, 1),
                        Paragraph("Relatorio de Downloads NFS-e", title_style),
                    Paragraph("Painel visual das notas emitidas e recebidas", subtitle_style),
                ]
            ],
            colWidths=[40, doc.width * 0.45, doc.width * 0.45],
        )
        hero.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(palette["panel"])),
                    ("LEFTPADDING", (0, 0), (-1, -1), 12),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        elements.append(hero)
        elements.append(Spacer(1, 8))

        chips = [
            pill("Iniciado", fmt_dt(started_txt), palette["accent"]),
            pill("Finalizado", fmt_dt(finished_txt), palette["accent"]),
            pill("Duracao", fmt_duration(started_txt, finished_txt), palette["accent"]),
            pill("Pasta XML", html.escape(summary.get("output_base", "-")), palette["ink"]),
        ]
        if self.report_path:
            chips.append(pill("Pasta PDF", html.escape(str(self.report_path)), palette["ink"]))
        for row in [chips[i : i + 3] for i in range(0, len(chips), 3)]:
            while len(row) < 3:
                row.append(Spacer(1, 1))
            chip_table = Table([row], colWidths=[doc.width / 3] * 3)
            chip_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
            elements.append(chip_table)
            elements.append(Spacer(1, 6))

        elements.append(Paragraph("Resumo executivo", section_style))
        cards = [
            stat_card("Sucesso", str(by_status.get("success", 0)), "Empresas ok", palette["green"]),
            stat_card("Parcial", str(by_status.get("partial", 0)), "Com lacunas", palette["yellow"]),
            stat_card("Erro", str(by_status.get("error", 0)), "Falharam", palette["red"]),
            stat_card("Sem movimento", str(by_status.get("empty", 0)), "Sem XML", palette["muted"]),
            stat_card("XML emitidas", str(downloads.get("emitidas", 0)), "Baixadas", palette["accent"]),
            stat_card("XML recebidas", str(downloads.get("recebidas", 0)), "Baixadas", palette["accent"]),
        ]
        for row in [cards[i : i + 3] for i in range(0, len(cards), 3)]:
            cards_table = Table([row], colWidths=[doc.width / 3 - 6] * 3)
            cards_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
            elements.append(cards_table)
            elements.append(Spacer(1, 6))

        elements.append(Paragraph("Overview rapido", section_style))
        if companies:
            overview_rows = [
                [
                    Paragraph("Empresa", label_style),
                    Paragraph("Emitidas", label_style),
                    Paragraph("Recebidas", label_style),
                    Paragraph("Status", label_style),
                    Paragraph("Pasta", label_style),
                ]
            ]
            for comp in companies:
                status_label, status_color = self._status_palette(comp.get("status"))
                overview_rows.append(
                    [
                        Paragraph(html.escape(comp.get("name", "")), value_style),
                        Paragraph(str(comp.get("emitidas", {}).get("downloaded", 0)), value_style),
                        Paragraph(str(comp.get("recebidas", {}).get("downloaded", 0)), value_style),
                        badge(status_label, status_color),
                        Paragraph(html.escape(comp.get("folder", "-")), label_style),
                    ]
                )
            overview = Table(
                overview_rows,
                colWidths=[doc.width * 0.28, doc.width * 0.12, doc.width * 0.12, doc.width * 0.18, doc.width * 0.3],
                repeatRows=1,
            )
            overview.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(palette["surface"])),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor(palette["ink"])),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor(palette["panel"]), colors.HexColor(palette["surface"])]),
                        ("LEFTPADDING", (0, 0), (-1, -1), 8),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                        ("TOPPADDING", (0, 0), (-1, -1), 5),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor(palette["line"])),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ]
                )
            )
            elements.append(overview)
        else:
            elements.append(Paragraph("Nenhuma empresa processada.", value_style))

        elements.append(Spacer(1, 10))

        for idx, comp in enumerate(companies):
            if idx == 0:
                elements.append(Paragraph("Detalhes por empresa", section_style))
            else:
                elements.append(Spacer(1, 2))
                elements.append(company_separator(idx, len(companies), comp))
                elements.append(Spacer(1, 5))
            card = company_card(comp)
            elements.append(card)
            elements.append(Spacer(1, 5))
            for flow in company_service_sections(comp):
                elements.append(flow)
            elements.append(Spacer(1, 10))

        doc.build(elements, onFirstPage=decorate, onLaterPages=decorate)
        return pdf_path
    def _open_pdf(self, pdf_path: Path) -> None:
        try:
            if os.name == "nt":
                os.startfile(str(pdf_path))  # type: ignore[attr-defined]
            else:
                QDesktopServices.openUrl(QUrl.fromLocalFile(str(pdf_path)))
        except Exception:
            # Se nao abrir automaticamente, apenas ignora.
            pass
# --------------------------------------------------------------------
# Main
# --------------------------------------------------------------------


def main():
    try:
        if apply_startup_update_if_available():
            return
    except Exception:
        pass

    scheduler_mode = is_scheduler_mode_enabled()
    app = QApplication(sys.argv)
    if LOGO_ICON.exists():
        app.setWindowIcon(QIcon(str(LOGO_ICON)))
    if not scheduler_mode and not ensure_license_valid(app):
        sys.exit(0)
    app.setStyleSheet(
        """
        QDialog { background:#17202A; }
        QLabel { color:#ECF0F1; font:9pt Verdana; font-weight:bold; }
        QLineEdit { background:#34495E; color:#ECF0F1; border-radius:6px; padding:6px; font-weight:bold; }
        QPushButton { font:9pt Verdana; font-weight:bold; }
        """
    )
    win = MainWindow()
    app.aboutToQuit.connect(lambda: _mark_process_inactive(win, "qt_about_to_quit"))
    _install_process_shutdown_handlers(win)
    if scheduler_mode:
        win.hide()
    else:
        win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
''
