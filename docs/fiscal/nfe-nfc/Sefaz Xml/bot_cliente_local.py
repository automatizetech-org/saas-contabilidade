# -*- coding: utf-8 -*-
"""
CÓPIA para testes em cliente — não altera o script principal "sefaz xml.py".

Se NÃO existir nenhum arquivo .env nos caminhos usados pelo robô (pasta do
Servidor/ROBOTS, data/.env, .env ao lado do script, etc.), este modo ativa:
  - Não carrega python-dotenv a partir de arquivo (evita puxar .env de outro lugar).
  - Supabase fica desligado se não houver variáveis no processo — empresas e login
    vêm dos JSON em data/json (config.json, login_portal.json), como no fluxo antigo.
  - A checagem de licença online (Supabase) é ignorada para abrir a interface.

Coloque um .env válido na pasta esperada para voltar ao comportamento integrado ao SaaS.

Automação SEFAZ-GO com Interface PySide6 + Playwright

- Interface:
  * Seleção de IE (com checkbox Diário)
  * Operação: Entrada / Saída / Ambos
  * Data inicial / final
  * Processar por múltiplos meses
  * Processar por intervalo de dias
  * Agendar início + contagem regressiva
  * Gerenciador de Empresas (Adicionar / Editar / Excluir)
  * Caminhos padrão:
      - Pasta base das empresas (subpasta por empresa, onde vão os XMLs/ZIPs)
      - Pasta para salvar relatórios em PDF
  * Tipo de download: Documentos / Eventos / Documentos + Eventos
  * Log de execução (estruturado com emojis)
  * Ao final: relatório PDF automático + mensagem “Processo finalizado”

- Backend: Playwright (Chrome portátil em pasta "Chrome")
  * Login no portal
  * Acesso restrito
  * Baixar XML NFE
  * Preenche data/IE/tipo de nota (limpa campo antes)
  * 3s de espera e clica Pesquisar
  * Lê alerts (textuais e “gerais” com botão)
  * Classifica alerta e registra no relatório
  * Tenta até 5x em caso de erro geral (sem recarregar página)
  * Solicita geração de ZIP (Baixar todos os arquivos) com seleção robusta do tipo
  * Vai ao Histórico de Downloads, espera “Concluído”
  * Baixa ZIP(s), unifica os XMLs e gera um ZIP unificado renomeado:
      "Unificado - OPERACAO - DATA_INI a DATA_FIM - NOME_EMPRESA.zip"0
"""

# =============================================================================
# IMPORTS - Standard Library
# =============================================================================
import sys
import os
import pathlib
import json
import atexit
import signal
import subprocess
import shutil
import tempfile
import zipfile
import io
import socket
import ctypes
import math
import time
import hashlib
import platform
import uuid
import re
import traceback
from time import sleep
from datetime import datetime, timedelta, date, timezone
from dataclasses import dataclass
from typing import List, Dict, Tuple, Optional, Callable, Any, Set
from collections import defaultdict
from urllib.parse import urlparse
import urllib.request
import urllib.error
import xml.etree.ElementTree as ET

try:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

# =============================================================================
# IMPORTS - Third Party Libraries
# =============================================================================
from PIL import Image, ImageChops, ImageFilter, ImageQt, ImageDraw

# =============================================================================
# CONFIGURAÇÃO Qt (Windows) - Suprimir avisos DPI/monitores
# IMPORTANTE: Deve rodar ANTES de importar PySide6/Qt
# =============================================================================
if sys.platform.startswith('win'):
    _rules = os.environ.get('QT_LOGGING_RULES', '').strip()
    _extra = 'qt.qpa.window=false;qt.qpa.screen=false'
    if _rules:
        if 'qt.qpa.window' not in _rules or 'qt.qpa.screen' not in _rules:
            os.environ['QT_LOGGING_RULES'] = _rules + ';' + _extra
    else:
        os.environ['QT_LOGGING_RULES'] = _extra

# =============================================================================
# IMPORTS - Qt / GUI Framework
# =============================================================================
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QTabWidget, QStackedWidget,
    QScrollArea, QCheckBox, QPushButton, QLabel, QFrame, QPlainTextEdit,
    QRadioButton, QButtonGroup, QLineEdit, QInputDialog, QMessageBox, QSpacerItem,
    QDialog, QDialogButtonBox, QFileDialog, QComboBox, QGraphicsDropShadowEffect, QSizePolicy, QDateEdit, QToolTip,
    QMenu, QSystemTrayIcon,
    QListWidget,
    QListWidgetItem,
    QAbstractItemView,
    QStyleOptionButton, QStyle
)

from PySide6.QtCore import (
    Qt, QThread, Signal, QUrl, QTimer, QPropertyAnimation, QEasingCurve,
    QParallelAnimationGroup, QMarginsF, QRectF, QRect, Property, QPointF,
    QPoint, QSize, QDate, QObject, QEvent
)

from PySide6.QtGui import (
    QAction, QIcon, QPdfWriter, QPainter, QFont, QPageSize, QDesktopServices,
    QColor, QPen, QIntValidator, QPixmap, QImage, QPainterPath, QBrush, QLinearGradient
)

# =============================================================================
# IMPORTS - Playwright (Automação Web)
# =============================================================================
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

# =============================================================================
# CONFIGURAÇÃO SUPABASE / ENV COMPARTILHADO
# =============================================================================
_base_dir = os.path.dirname(os.path.abspath(__file__))
_env_data_path = os.path.join(_base_dir, "data", ".env")
_env_local_path = os.path.join(_base_dir, ".env")
_env_root_path = os.path.abspath(os.path.join(_base_dir, "..", "..", ".env"))
def _resolve_robots_base_env_dir() -> pathlib.Path:
    candidates: list[pathlib.Path] = []
    env_root = (os.getenv("ROBOTS_ROOT_PATH") or "").strip()
    robot_script_dir = (os.getenv("ROBOT_SCRIPT_DIR") or "").strip()

    if env_root:
        candidates.append(pathlib.Path(env_root))
    if robot_script_dir:
        candidates.append(pathlib.Path(robot_script_dir).resolve().parent)
    candidates.append(pathlib.Path(_base_dir).resolve().parent)
    if getattr(sys, "frozen", False):
        exe_dir = pathlib.Path(sys.executable).resolve().parent
        candidates.append(exe_dir.parent)
        candidates.append(exe_dir)

    seen: set[str] = set()
    for candidate in candidates:
        try:
            resolved = candidate.resolve()
        except Exception:
            resolved = candidate
        key = str(resolved).lower()
        if key in seen:
            continue
        seen.add(key)
        if (resolved / ".env").exists() or (resolved / ".env.example").exists():
            return resolved

    return pathlib.Path(env_root).resolve() if env_root else pathlib.Path(_base_dir).resolve().parent


def _json_runtime_utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _json_runtime_ensure_parent(path: pathlib.Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _json_runtime_atomic_write_json(path: pathlib.Path, payload: dict[str, Any]) -> None:
    _json_runtime_ensure_parent(path)
    temp_path = path.with_suffix(path.suffix + ".tmp")
    temp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    temp_path.replace(path)


class JsonRobotRuntime:
    def __init__(self, technical_id: str, display_name: str, base_dir: pathlib.Path) -> None:
        self.technical_id = technical_id
        self.display_name = display_name
        self.base_dir = base_dir
        self.last_status = "inactive"
        self.json_dir = self.base_dir / "data" / "json"
        self.job_path = self.json_dir / "job.json"
        self.result_path = self.json_dir / "result.json"
        self.heartbeat_path = self.json_dir / "heartbeat.json"

    def register_robot(self) -> str:
        self.write_heartbeat(status="active")
        return self.technical_id

    def load_job(self) -> Optional[dict[str, Any]]:
        if not self.job_path.exists():
            return None
        try:
            payload = json.loads(self.job_path.read_text(encoding="utf-8"))
        except Exception:
            return None
        if not isinstance(payload, dict):
            return None
        if not payload.get("id"):
            payload["id"] = payload.get("execution_request_id") or payload.get("job_id")
        if not payload.get("job_id"):
            payload["job_id"] = payload.get("id")
        if not payload.get("execution_request_id"):
            payload["execution_request_id"] = payload.get("id")
        job_execution_id = str(
            payload.get("execution_request_id")
            or payload.get("job_id")
            or payload.get("id")
            or ""
        ).strip()
        if not job_execution_id:
            return None
        if self.result_path.exists():
            try:
                result_payload = json.loads(self.result_path.read_text(encoding="utf-8"))
            except Exception:
                result_payload = None
            if isinstance(result_payload, dict):
                result_execution_id = str(
                    result_payload.get("execution_request_id")
                    or result_payload.get("job_id")
                    or result_payload.get("event_id")
                    or ""
                ).strip()
                if result_execution_id and result_execution_id == job_execution_id:
                    return None
        return payload

    def load_job_companies(
        self,
        job: Optional[dict[str, Any]],
        company_ids: Optional[List[str]] = None,
    ) -> List[dict[str, Any]]:
        if not isinstance(job, dict):
            return []
        companies = job.get("companies")
        if not isinstance(companies, list):
            return []
        wanted = {str(company_id) for company_id in (company_ids or []) if str(company_id).strip()}
        rows: List[dict[str, Any]] = []
        for row in companies:
            if not isinstance(row, dict):
                continue
            company_id = str(row.get("company_id") or row.get("id") or "").strip()
            if wanted and company_id not in wanted:
                continue
            rows.append(row)
        return rows

    def write_heartbeat(
        self,
        *,
        status: str,
        current_job_id: Optional[str] = None,
        current_execution_request_id: Optional[str] = None,
        message: Optional[str] = None,
        progress: Optional[dict[str, Any]] = None,
        extra: Optional[dict[str, Any]] = None,
    ) -> None:
        self.last_status = status
        payload: dict[str, Any] = {
            "robot_technical_id": self.technical_id,
            "display_name": self.display_name,
            "status": status,
            "updated_at": _json_runtime_utc_now_iso(),
            "current_job_id": current_job_id,
            "current_execution_request_id": current_execution_request_id,
            "message": message,
            "progress": progress or {},
        }
        if extra:
            payload.update(extra)
        _json_runtime_atomic_write_json(self.heartbeat_path, payload)

    def write_result(
        self,
        *,
        job: Optional[dict[str, Any]],
        success: bool,
        error_message: Optional[str] = None,
        summary: Optional[dict[str, Any]] = None,
        payload: Optional[dict[str, Any]] = None,
        company_results: Optional[List[dict[str, Any]]] = None,
    ) -> None:
        execution_request_id = None
        job_id = None
        if isinstance(job, dict):
            execution_request_id = str(job.get("execution_request_id") or job.get("id") or "").strip() or None
            job_id = str(job.get("job_id") or job.get("id") or "").strip() or execution_request_id

        event_id = execution_request_id or str(uuid.uuid4())
        result_payload: dict[str, Any] = {
            "event_id": event_id,
            "job_id": job_id or event_id,
            "execution_request_id": execution_request_id,
            "robot_technical_id": self.technical_id,
            "status": "completed" if success else "failed",
            "started_at": _json_runtime_utc_now_iso(),
            "finished_at": _json_runtime_utc_now_iso(),
            "error_message": error_message,
            "summary": summary or {},
            "company_results": company_results or [],
            "payload": payload or {},
        }
        _json_runtime_atomic_write_json(self.result_path, result_payload)
        self.write_heartbeat(
            status="active",
            current_job_id=None,
            current_execution_request_id=None,
            message="result_ready",
        )


_robots_base_env_dir = _resolve_robots_base_env_dir()
_robots_env_path = _robots_base_env_dir / ".env"
_robots_env_example_path = _robots_base_env_dir / ".env.example"


def _any_real_dotenv_file_exists() -> bool:
    """True se existir algum .env real (não usa .env.example para decidir modo local)."""
    paths = (
        _robots_env_path,
        pathlib.Path(_env_data_path),
        pathlib.Path(_env_local_path),
        pathlib.Path(_env_root_path),
    )
    for p in paths:
        try:
            if p.exists() and p.is_file():
                return True
        except Exception:
            continue
    return False


# Sem .env em disco: modo cliente / demonstração — prioriza JSON em data/json.
SEFAZ_LOCAL_JSON_CLIENT_MODE = not _any_real_dotenv_file_exists()


def _get_supabase_service_role_key() -> str:
    for env_name in (
        "SUPABASE_SERVICE_ROLE_KEY",
        "SERVICE_ROLE_KEY",
        "SUPABASE_KEY",
        "SUPABASE_SECRET_KEY",
        "NEXT_PUBLIC_SUPABASE_SERVICE_ROLE_KEY",
    ):
        value = (os.getenv(env_name) or "").strip()
        if value:
            return value
    return ""

try:
    from dotenv import load_dotenv
    from supabase import create_client, Client
    import requests
    import time
    try:
        import httpx  # usado para detectar erros de rede (WinError 10035)
    except Exception:
        httpx = None
    
    if not SEFAZ_LOCAL_JSON_CLIENT_MODE:
        # Ordem: .env compartilhado da VM -> .env local do bot -> fallback padrão.
        if _robots_env_path.exists():
            load_dotenv(_robots_env_path, override=True)
        elif _robots_env_example_path.exists():
            load_dotenv(_robots_env_example_path, override=True)

        if os.path.exists(_env_data_path):
            load_dotenv(_env_data_path, override=True)
        elif os.path.exists(_env_local_path):
            load_dotenv(_env_local_path, override=True)
        elif os.path.exists(_env_root_path):
            load_dotenv(_env_root_path, override=True)
        else:
            load_dotenv(override=True)
    else:
        print(
            "[MODO CLIENTE LOCAL] Nenhum .env encontrado nos caminhos do robô — "
            "não carregando variáveis de arquivo. "
            "Use data/json (config.json, login_portal.json, …). Licença online desligada para abrir a UI."
        )

    # Tenta ler as variáveis do .env (pode ser NEXT_PUBLIC_* ou sem prefixo)
    SUPABASE_URL = os.getenv("SUPABASE_URL") or os.getenv("NEXT_PUBLIC_SUPABASE_URL")
    # Para o Python, PRECISAMOS da SERVICE_ROLE_KEY para fazer INSERT/UPDATE/DELETE
    # A ANON_KEY não funciona para operações de escrita devido ao RLS
    SUPABASE_KEY = _get_supabase_service_role_key()
    
    # Verifica se a Service Role Key está configurada
    if not SUPABASE_KEY:
        if not SEFAZ_LOCAL_JSON_CLIENT_MODE:
            print(f"[SUPABASE] ⚠️ ATENÇÃO: Service Role Key não encontrada no .env")
            print(f"  - Para operações de escrita (INSERT/UPDATE/DELETE), você precisa da Service Role Key")
            print(f"  - Obtenha a Service Role Key em: Supabase Dashboard > Settings > API > service_role key")
            print(f"  - Aceito em qualquer um destes nomes: SUPABASE_SERVICE_ROLE_KEY / SERVICE_ROLE_KEY / SUPABASE_KEY")
        supabase_client = None
        SUPABASE_AVAILABLE = False
    elif SUPABASE_KEY == "COLE_AQUI_SUA_SERVICE_ROLE_KEY":
        print(f"[SUPABASE] ⚠️ ATENÇÃO: Service Role Key ainda está com valor placeholder")
        print(f"  - Substitua 'COLE_AQUI_SUA_SERVICE_ROLE_KEY' pela sua Service Role Key real")
        print(f"  - Obtenha em: Supabase Dashboard > Settings > API > service_role key")
        supabase_client = None
        SUPABASE_AVAILABLE = False
    elif SUPABASE_URL and SUPABASE_KEY:
        supabase_client: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
        SUPABASE_AVAILABLE = True
        # Log de debug (pode remover depois)
        print(f"[SUPABASE] ✅ Configurado: URL={SUPABASE_URL[:30]}...")
        print(f"[SUPABASE] ✅ Service Role Key configurada (primeiros 20 chars): {SUPABASE_KEY[:20]}...")
    else:
        supabase_client = None
        SUPABASE_AVAILABLE = False
        if not SEFAZ_LOCAL_JSON_CLIENT_MODE:
            print(f"[SUPABASE] ❌ Não configurado:")
            print(f"  - SUPABASE_URL: {'✅' if SUPABASE_URL else '❌'}")
            print(f"  - SERVICE_ROLE_KEY: {'✅' if SUPABASE_KEY else '❌'}")
            print(f"  - Caminho .env local tentado: {_env_data_path if os.path.exists(_env_data_path) else (_env_local_path if os.path.exists(_env_local_path) else _env_root_path)}")
            print(f"  - .env compartilhado existe: {_robots_env_path.exists()}")
except ImportError as e:
    SUPABASE_AVAILABLE = False
    supabase_client = None
    print(f"[SUPABASE] ❌ Erro ao importar bibliotecas: {e}")
    print(f"  Execute: pip install supabase python-dotenv requests")
except Exception as e:
    SUPABASE_AVAILABLE = False
    supabase_client = None
    print(f"[SUPABASE] ❌ Erro ao configurar: {e}")
    import traceback
    traceback.print_exc()


# =============================================================================
# SUPABASE - Helper de retry (evita WinError 10035 / flutuação de rede no Windows)
# =============================================================================
def _supabase_retry_execute(fn, log_fn=None, retries: int = 4, base_sleep_s: float = 0.6):
    """
    Executa `fn()` com retry/backoff para erros transitórios do httpx/httpcore no Windows.

    Motivação:
      - Alguns ambientes no Windows geram `httpx.ReadError: [WinError 10035] ...`
        principalmente em conexões HTTP/2. Um retry resolve na prática.
    """
    last_exc = None
    for attempt in range(1, max(1, retries) + 1):
        try:
            return fn()
        except Exception as e:
            last_exc = e
            msg = str(e)
            is_win_10035 = ("WinError 10035" in msg) or ("10035" in msg)
            is_httpx_read = (httpx is not None) and isinstance(e, getattr(httpx, "ReadError", Exception))
            is_httpx_remote_protocol = (httpx is not None) and isinstance(
                e, getattr(httpx, "RemoteProtocolError", Exception)
            )
            # httpcore.RemoteProtocolError pode aparecer encapsulado ou direto
            is_server_disconnected = "Server disconnected" in msg
            is_transient = is_win_10035 or is_httpx_read or is_httpx_remote_protocol or is_server_disconnected

            if not is_transient or attempt >= retries:
                raise

            sleep_s = base_sleep_s * (2 ** (attempt - 1))
            if callable(log_fn):
                log_fn(f"[WARN] ⚠️ Erro de rede (tentativa {attempt}/{retries}). Repetindo em {sleep_s:.1f}s... ({type(e).__name__})")
            time.sleep(sleep_s)

    # fallback (não deve chegar aqui)
    raise last_exc


ROBOT_TECHNICAL_ID = os.getenv("ROBOT_TECHNICAL_ID", "sefaz_xml").strip() or "sefaz_xml"
ROBOT_DISPLAY_NAME_DEFAULT = os.getenv("ROBOT_DISPLAY_NAME", "Sefaz Xml").strip() or "Sefaz Xml"
ROBOT_SEGMENT_PATH_DEFAULT = os.getenv("ROBOT_SEGMENT_PATH", "FISCAL/NFE-NFC").strip() or "FISCAL/NFE-NFC"
_robot_api_config: Optional[Dict[str, Any]] = None
JSON_RUNTIME = JsonRobotRuntime(
    ROBOT_TECHNICAL_ID,
    ROBOT_DISPLAY_NAME_DEFAULT,
    pathlib.Path(__file__).resolve().parent,
)
ACTIVE_JSON_JOB: Optional[Dict[str, Any]] = None
PENDING_RESULT_OPERATIONS: List[Dict[str, Any]] = []

_PROCESS_SHUTDOWN_MARKED_INACTIVE = False
_WINDOWS_CONSOLE_HANDLER = None


def _mark_process_inactive(window: Any | None = None, reason: str = "process_exit") -> None:
    global _PROCESS_SHUTDOWN_MARKED_INACTIVE
    if _PROCESS_SHUTDOWN_MARKED_INACTIVE:
        return
    _PROCESS_SHUTDOWN_MARKED_INACTIVE = True
    try:
        JSON_RUNTIME.write_heartbeat(
            status="inactive",
            current_job_id=None,
            current_execution_request_id=None,
            message=reason,
        )
    except Exception:
        pass
    try:
        robot_id = getattr(window, "robot_id", None) if window is not None else None
        if robot_id:
            update_robot_status(robot_id, "inactive")
    except Exception:
        pass


def _install_process_shutdown_handlers(window: Any | None = None) -> None:
    global _WINDOWS_CONSOLE_HANDLER

    def _handle_signal(signum, _frame) -> None:
        _mark_process_inactive(window, f"signal_{signum}")
        raise SystemExit(0)

    atexit.register(lambda: _mark_process_inactive(window, "atexit"))
    for signal_name in ("SIGINT", "SIGTERM", "SIGBREAK"):
        sig = getattr(signal, signal_name, None)
        if sig is None:
            continue
        try:
            signal.signal(sig, _handle_signal)
        except Exception:
            pass

    if os.name == "nt":
        try:
            handler_type = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_uint)

            @handler_type
            def _console_handler(ctrl_type: int) -> bool:
                _mark_process_inactive(window, f"console_ctrl_{ctrl_type}")
                return False

            ctypes.windll.kernel32.SetConsoleCtrlHandler(_console_handler, True)
            _WINDOWS_CONSOLE_HANDLER = _console_handler
        except Exception:
            pass


def _current_json_job() -> Optional[Dict[str, Any]]:
    return ACTIVE_JSON_JOB if isinstance(ACTIVE_JSON_JOB, dict) else None


def _set_current_json_job(job: Optional[Dict[str, Any]]) -> None:
    global ACTIVE_JSON_JOB
    ACTIVE_JSON_JOB = job if isinstance(job, dict) else None


def _clear_pending_result_operations() -> None:
    PENDING_RESULT_OPERATIONS.clear()


def _append_result_operation(operation: Dict[str, Any]) -> None:
    if isinstance(operation, dict):
        PENDING_RESULT_OPERATIONS.append(operation)


def _consume_result_operations() -> List[Dict[str, Any]]:
    operations = list(PENDING_RESULT_OPERATIONS)
    PENDING_RESULT_OPERATIONS.clear()
    return operations


def _find_current_job_company(cnpj: str, company_name: str = "") -> Optional[Dict[str, Any]]:
    job = _current_json_job()
    if not isinstance(job, dict):
        return None
    target_cnpj = cnpj_somente_digitos(cnpj)
    target_name = normalize_company_name(company_name)
    for row in JSON_RUNTIME.load_job_companies(job):
        row_cnpj = cnpj_somente_digitos(str(row.get("document") or row.get("cnpj") or ""))
        row_name = normalize_company_name(str(row.get("name") or ""))
        if target_cnpj and row_cnpj == target_cnpj:
            return row
        if target_name and row_name == target_name:
            return row
    return None


def get_robot_supabase_credentials() -> Tuple[Optional[str], Optional[str]]:
    url = (os.getenv("SUPABASE_URL") or os.getenv("NEXT_PUBLIC_SUPABASE_URL") or "").strip()
    key = _get_supabase_service_role_key()
    if url and key:
        return (url, key)
    return (None, None)


def fetch_robot_config_from_api() -> Optional[Dict[str, Any]]:
    url_base = (os.getenv("FOLDER_STRUCTURE_API_URL") or os.getenv("SERVER_API_URL") or "").strip().rstrip("/")
    if not url_base:
        return None
    try:
        headers = {}
        if "ngrok" in url_base.lower():
            headers["ngrok-skip-browser-warning"] = "true"
        connector_secret = (os.getenv("CONNECTOR_SECRET") or "").strip()
        if connector_secret:
            headers["Authorization"] = f"Bearer {hashlib.sha256(connector_secret.encode('utf-8')).hexdigest()}"
        response = requests.get(
            f"{url_base}/api/robot-config",
            params={"technical_id": ROBOT_TECHNICAL_ID},
            headers=headers,
            timeout=15,
        )
        response.raise_for_status()
        payload = response.json()
        return payload if isinstance(payload, dict) else None
    except Exception:
        return None


def get_robot_api_config(force_refresh: bool = False) -> Optional[Dict[str, Any]]:
    global _robot_api_config
    if force_refresh or _robot_api_config is None:
        _robot_api_config = fetch_robot_config_from_api()
    return _robot_api_config


def get_resolved_output_base() -> Optional[pathlib.Path]:
    cfg = get_robot_api_config()
    base_path = (cfg or {}).get("base_path") or os.getenv("BASE_PATH") or ""
    base_path = str(base_path).strip()
    if not base_path:
        return None
    return pathlib.Path(base_path)


def resolve_dashboard_output_root() -> Optional[str]:
    base = get_resolved_output_base()
    if not base:
        return None
    # A pasta base da VM definida no dashboard já é a raiz efetiva de gravação.
    # Nada de prefixo fixo como EMPRESAS aqui.
    return str(base)


def get_robot_segment_under_company_parts() -> List[str]:
    """
    Retorna os segmentos do robô para usar *dentro* da pasta da empresa.

    Ex.: se o dashboard usa segment_path = "FISCAL/NFE-NFC",
    queremos salvar em: ...\\<EMPRESA>\\FISCAL\\NFE-NFC\\...
    (respeita a estrutura 3D do dashboard).
    """
    try:
        cfg = get_robot_api_config() or {}
        segment_path = str(cfg.get("segment_path") or ROBOT_SEGMENT_PATH_DEFAULT).strip()
        parts = [safe_folder_name(p.strip()) for p in segment_path.split("/") if p.strip()]
        parts = [p for p in parts if p]
        if not parts:
            return ["NFE-NFC"]

        return parts if parts else ["NFE-NFC"]
    except Exception:
        return ["NFE-NFC"]


def _open_pdf_file(path_str: str, log_fn: Optional[Callable[[str], None]] = None) -> bool:
    """
    Abre um PDF no Windows de forma resiliente.
    Primeiro tenta QDesktopServices; se falhar, faz fallback para os.startfile.
    """
    try:
        if not path_str:
            return False
        p = str(path_str)
        try:
            if os.path.exists(p):
                QDesktopServices.openUrl(QUrl.fromLocalFile(p))
                return True
        except Exception:
            pass
        try:
            if os.name == "nt" and os.path.exists(p):
                os.startfile(p)  # type: ignore[attr-defined]
                return True
        except Exception as e:
            if callable(log_fn):
                log_fn(f"[ERRO] Falha ao abrir PDF automaticamente (fallback): {e}")
    except Exception:
        pass
    return False


def load_companies_from_dashboard() -> List[Dict[str, Any]]:
    if not SUPABASE_AVAILABLE or not supabase_client:
        return []
    try:
        robot_row = fetch_robot_row()
        robot_global_logins = _normalize_sefaz_go_logins((robot_row or {}).get("global_logins"))
        configs = _supabase_retry_execute(
            lambda: supabase_client.table("company_robot_config")
            .select("company_id, enabled, selected_login_cpf")
            .eq("robot_technical_id", ROBOT_TECHNICAL_ID)
            .eq("enabled", True)
            .execute()
        )
        config_rows = getattr(configs, "data", None) or []
        config_by_company = {
            row.get("company_id"): row
            for row in config_rows
            if row.get("company_id")
        }
        enabled_company_ids = list(config_by_company.keys())

        query = (
            supabase_client.table("companies")
            .select("id, name, document, active, state_registration, contador_cpf, sefaz_go_logins")
            .eq("active", True)
            .order("name")
        )
        if enabled_company_ids:
            query = query.in_("id", enabled_company_ids)
        try:
            response = _supabase_retry_execute(lambda: query.execute())
        except Exception:
            fallback_query = (
                supabase_client.table("companies")
                .select("id, name, document, active")
                .eq("active", True)
                .order("name")
            )
            if enabled_company_ids:
                fallback_query = fallback_query.in_("id", enabled_company_ids)
            response = _supabase_retry_execute(lambda: fallback_query.execute())
        rows = getattr(response, "data", None) or []
        normalized: List[Dict[str, Any]] = []
        for row in rows:
            normalized.append(
                {
                    "id": row.get("id"),
                    "name": (row.get("name") or "").strip(),
                    "document": "".join(ch for ch in str(row.get("document") or "") if ch.isdigit()),
                    "state_registration": ie_somente_digitos(row.get("state_registration") or ""),
                    "contador_cpf": cpf_somente_digitos(row.get("contador_cpf") or ""),
                    "selected_login_cpf": cpf_somente_digitos((config_by_company.get(row.get("id")) or {}).get("selected_login_cpf") or ""),
                    "global_logins": robot_global_logins,
                    "legacy_company_logins": _normalize_sefaz_go_logins(row.get("sefaz_go_logins")),
                }
            )
        return normalized
    except Exception:
        return []


def fetch_robot_row() -> Optional[Dict[str, Any]]:
    if not SUPABASE_AVAILABLE or not supabase_client:
        return None
    try:
        try:
            response = _supabase_retry_execute(
                lambda: supabase_client.table("robots")
                .select("id, technical_id, display_name, status, segment_path, initial_period_start, initial_period_end, notes_mode, global_logins")
                .eq("technical_id", ROBOT_TECHNICAL_ID)
                .limit(1)
                .execute()
            )
        except Exception:
            response = _supabase_retry_execute(
                lambda: supabase_client.table("robots")
                .select("id, technical_id, display_name, status, segment_path, initial_period_start, initial_period_end, notes_mode")
                .eq("technical_id", ROBOT_TECHNICAL_ID)
                .limit(1)
                .execute()
            )
        rows = getattr(response, "data", None) or []
        return rows[0] if rows else None
    except Exception:
        return None


def register_robot() -> Optional[str]:
    try:
        return JSON_RUNTIME.register_robot()
    except Exception as exc:
        print(f"[ROBO] Falha ao iniciar runtime JSON: {exc}")
        return None


def update_robot_heartbeat(robot_id: Optional[str]) -> None:
    if not robot_id:
        return
    try:
        job = _current_json_job()
        JSON_RUNTIME.write_heartbeat(
            status="processing" if job else "active",
            current_job_id=(job or {}).get("job_id"),
            current_execution_request_id=(job or {}).get("execution_request_id"),
        )
    except Exception:
        pass


def update_robot_status(robot_id: Optional[str], status: str) -> None:
    if not robot_id:
        return
    try:
        job = _current_json_job()
        JSON_RUNTIME.write_heartbeat(
            status=status,
            current_job_id=(job or {}).get("job_id"),
            current_execution_request_id=(job or {}).get("execution_request_id"),
        )
    except Exception:
        pass


# =============================================================================
# AGENDADOR (fila execution_requests) — compat com AdminScheduler do site
# =============================================================================
def _get_active_schedule_rule_ids_for_queue(client: Any) -> Optional[list]:
    """Lista ids de schedule_rules ativas (status=active, run_daily=true). None em erro."""
    try:
        r = (
            client.table("schedule_rules")
            .select("id")
            .eq("status", "active")
            .eq("run_daily", True)
            .execute()
        )
        rows = getattr(r, "data", None) or []
        return [row["id"] for row in rows if row.get("id")]
    except Exception:
        return None


def claim_execution_request_for_queue(
    *,
    supabase_url: str,
    supabase_service_key: str,
    robot_id: str,
    log_callback: Optional[Callable[[str], None]] = None,
) -> Optional[Dict[str, Any]]:
    """
    Busca um job pending para este robô e marca como running.
    Respeita: regras ativas (schedule_rule_id) + execução sequencial (execution_group_id/execution_order).
    """
    try:
        job = JSON_RUNTIME.load_job()
        if not job:
            return None
        _set_current_json_job(job)
        _clear_pending_result_operations()
        JSON_RUNTIME.write_heartbeat(
            status="processing",
            current_job_id=job.get("job_id"),
            current_execution_request_id=job.get("execution_request_id"),
            message="job_loaded",
        )
        return job
    except Exception as e:
        msg = f"[FILA] Erro ao buscar job do agendador: {e}"
        if callable(log_callback):
            log_callback(msg)
        try:
            print(msg, file=sys.stderr)
        except Exception:
            pass
        return None


def complete_execution_request_for_queue(
    *,
    supabase_url: str,
    supabase_service_key: str,
    request_id: str,
    success: bool,
    error_message: Optional[str] = None,
) -> None:
    try:
        job = _current_json_job()
        operations = _consume_result_operations()
        job_companies = JSON_RUNTIME.load_job_companies(job)
        rows_total = 0
        touched_company_ids: Set[str] = set()
        for operation in operations:
            rows = operation.get("rows") if isinstance(operation, dict) else None
            if isinstance(rows, list):
                rows_total += len(rows)
                for row in rows:
                    if not isinstance(row, dict):
                        continue
                    company_id = str(row.get("company_id") or "").strip()
                    if company_id:
                        touched_company_ids.add(company_id)
            company_id = str((operation or {}).get("company_id") or "").strip()
            if company_id:
                touched_company_ids.add(company_id)
        company_results = [
            {
                "company_id": str(company.get("company_id") or company.get("id") or "").strip() or None,
                "company_name": company.get("name"),
                "document": company.get("document") or company.get("cnpj") or company.get("doc"),
                "status": (
                    "success"
                    if str(company.get("company_id") or company.get("id") or "").strip() in touched_company_ids
                    else "no_data"
                ),
            }
            for company in job_companies
            if isinstance(company, dict)
        ]
        summary = {
            "companies_total": len(company_results),
            "companies_with_data": len([row for row in company_results if row.get("status") == "success"]),
            "companies_without_data": len([row for row in company_results if row.get("status") == "no_data"]),
            "operations_total": len(operations),
            "rows_total": rows_total,
        }
        JSON_RUNTIME.write_result(
            job=job if isinstance(job, dict) and str(job.get("execution_request_id") or job.get("id") or "") == str(request_id) else {"execution_request_id": request_id, "job_id": request_id},
            success=success,
            error_message=error_message,
            summary=summary,
            company_results=company_results,
            payload={"operations": operations},
        )
        _set_current_json_job(None)
    except Exception:
        pass

# =============================================================================
# COMPONENTES Qt - AnimatedCheckBox
# =============================================================================
_BaseQCheckBox = QCheckBox

class NoOutlineCheckBox(QCheckBox):
    """QCheckBox customizado que remove completamente qualquer outline do indicador."""
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setFocusPolicy(Qt.NoFocus)
        try:
            self.setAttribute(Qt.WA_MacShowFocusRect, False)
            self.setAttribute(Qt.WA_Hover, False)
        except Exception:
            pass
        self.setGraphicsEffect(None)
    
    def paintEvent(self, event):
        # Usa QStyleOptionButton para desenhar sem outline de focus
        from PySide6.QtWidgets import QStyleOptionButton
        from PySide6.QtWidgets import QStyle
        option = QStyleOptionButton()
        self.initStyleOption(option)
        # Remove flags de focus para evitar outline
        option.state &= ~QStyle.State_HasFocus
        option.state &= ~QStyle.State_KeyboardFocusChange
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        self.style().drawControl(QStyle.CE_CheckBox, option, painter, self)

class AnimatedCheckBox(_BaseQCheckBox):
    """QCheckBox variant that paints an animated verified indicator."""

    _INDICATOR_SIZE = 20
    _INDICATOR_MARGIN = 6

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setCursor(Qt.ArrowCursor)
        self.setAttribute(Qt.WA_Hover, True)
        self.setMouseTracking(True)
        self._check_progress = 1.0 if self.isChecked() else 0.0
        self._animation = QPropertyAnimation(self, b"checkProgress", self)
        self._animation.setDuration(260)
        self._animation.setEasingCurve(QEasingCurve.OutCubic)
        self.toggled.connect(self._animate_transition)

    @Property(float)
    def checkProgress(self) -> float:
        return self._check_progress

    @checkProgress.setter
    def checkProgress(self, value: float):
        value = max(0.0, min(1.0, value))
        if abs(self._check_progress - value) < 1e-4:
            return
        self._check_progress = value
        self.update()

    def _animate_transition(self, checked: bool):
        self._animation.stop()
        self._animation.setStartValue(self._check_progress)
        self._animation.setEndValue(1.0 if checked else 0.0)
        self._animation.start()

    def _hit_area_contains(self, pos: QPoint) -> bool:
        indicator_rect = QRect(
            self._INDICATOR_MARGIN,
            (self.height() - self._INDICATOR_SIZE) // 2,
            self._INDICATOR_SIZE,
            self._INDICATOR_SIZE,
        )
        if indicator_rect.contains(pos):
            return True
        text = (self.text() or "").strip()
        if not text:
            return False
        font_metrics = self.fontMetrics()
        text_width = font_metrics.horizontalAdvance(text)
        max_width = min(text_width + 8, 260)
        text_rect = QRect(
            indicator_rect.right() + 6,
            indicator_rect.top(),
            max(1, max_width),
            indicator_rect.height(),
        )
        return text_rect.contains(pos)

    def _update_cursor(self, pos: QPoint):
        if self._hit_area_contains(pos):
            self.setCursor(Qt.PointingHandCursor)
        else:
            self.setCursor(Qt.ArrowCursor)

    def enterEvent(self, event):
        pos = event.position().toPoint() if hasattr(event, "position") else event.pos()
        self._update_cursor(pos)
        super().enterEvent(event)

    def leaveEvent(self, event):
        self.setCursor(Qt.ArrowCursor)
        super().leaveEvent(event)

    def mouseMoveEvent(self, event):
        pos = event.position().toPoint() if hasattr(event, "position") else event.pos()
        self._update_cursor(pos)
        super().mouseMoveEvent(event)

    def paintEvent(self, event):
        super().paintEvent(event)
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        indicator_rect = QRect(
            self._INDICATOR_MARGIN,
            (self.height() - self._INDICATOR_SIZE) // 2,
            self._INDICATOR_SIZE,
            self._INDICATOR_SIZE,
        )

        progress = max(0.0, min(1.0, self._check_progress))
        checked = self.isChecked()
        highlight = QColor(116, 185, 255)
        base_bg = QColor(15, 24, 44)
        border = QColor(255, 255, 255, 110)

        painter.setPen(QPen(border, 1.4))
        painter.setBrush(QBrush(base_bg))
        painter.drawRoundedRect(indicator_rect, 8, 8)

        if checked and progress > 0:
            fill_rect = indicator_rect.adjusted(3, 3, -3, -3)
            fill_color = self._blend_color(QColor(34, 57, 108), highlight, progress)
            painter.setPen(Qt.NoPen)
            painter.setBrush(QBrush(fill_color))
            painter.drawRoundedRect(fill_rect, 6, 6)

            pen = QPen(QColor(255, 255, 255, 230), 2.6)
            pen.setCapStyle(Qt.RoundCap)
            pen.setJoinStyle(Qt.RoundJoin)
            painter.setPen(pen)
            start = QPointF(fill_rect.left() + fill_rect.width() * 0.18, fill_rect.center().y() + 1)
            mid = QPointF(fill_rect.left() + fill_rect.width() * 0.45, fill_rect.bottom() - fill_rect.height() * 0.25)
            end = QPointF(fill_rect.right() - fill_rect.width() * 0.18, fill_rect.top() + fill_rect.height() * 0.25)

            first_phase = min(progress / 0.6, 1.0)
            if first_phase > 0:
                painter.drawLine(start, self._lerp_point(start, mid, first_phase))

            if progress > 0.6:
                second_phase = min((progress - 0.6) / 0.4, 1.0)
                painter.drawLine(mid, self._lerp_point(mid, end, second_phase))

    @staticmethod
    def _blend_color(start: QColor, end: QColor, factor: float) -> QColor:
        factor = max(0.0, min(1.0, factor))
        return QColor(
            int(start.red() + (end.red() - start.red()) * factor),
            int(start.green() + (end.green() - start.green()) * factor),
            int(start.blue() + (end.blue() - start.blue()) * factor),
            int(start.alpha() + (end.alpha() - start.alpha()) * factor),
        )

    @staticmethod
    def _lerp_point(src: QPointF, dst: QPointF, factor: float) -> QPointF:
        factor = max(0.0, min(1.0, factor))
        return QPointF(
            src.x() + (dst.x() - src.x()) * factor,
            src.y() + (dst.y() - src.y()) * factor,
        )

    def sizeHint(self):
        base = super().sizeHint()
        min_height = self._INDICATOR_SIZE + self._INDICATOR_MARGIN * 2
        if base.height() < min_height:
            base.setHeight(min_height)
        return base

    def hitButton(self, pos: QPoint) -> bool:
        return self._hit_area_contains(pos)

QCheckBox = AnimatedCheckBox

# =============================================================================
# IMPORTS - Qt Quick (Dashboards/QML) - Opcional
# =============================================================================
try:
    from PySide6.QtQuickWidgets import QQuickWidget
except Exception:
    QQuickWidget = None

# =============================================================================
# IMPORTS - Outras Bibliotecas
# =============================================================================
# pyautogui pode ajustar DPI no Windows; importamos APÓS o Qt para evitar warnings
import pyautogui as py  # opcional; mantido p/ futuros ajustes visuais

# =============================================================================
# CONFIGURAÇÕES GERAIS / UTILITÁRIOS
# =============================================================================

URL_PORTAL = "https://portal.sefaz.go.gov.br/portalsefaz-apps/auth/login-form"

CPF_LOGIN = ""
SENHA_LOGIN = ""
CURRENT_LOGIN_CPF = ""

def _pick_cdp_port(default: int = 9222) -> int:
    """
    Permite sobrescrever a porta via SEFAZ_CDP_PORT.
    Se a porta padrão estiver ocupada, escolhe uma livre automaticamente.
    """
    raw = os.getenv("SEFAZ_CDP_PORT", "").strip()
    if raw:
        try:
            return int(raw)
        except ValueError:
            pass

    # Tenta reservar a porta padrão; se falhar, pega uma livre
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        try:
            s.bind(("127.0.0.1", default))
            return default
        except OSError:
            s.bind(("127.0.0.1", 0))
            return s.getsockname()[1]

CDP_PORT = _pick_cdp_port()
# Proxy opcional (ex.: http://127.0.0.1:8889 para mitmproxy)
SEFAZ_PROXY = os.getenv("SEFAZ_PROXY", "").strip()
DEFAULT_TIMEOUT_MS = 30_000  # 30s
# Portal SEFAZ costuma demorar ou oscilar; login usa timeouts/retries próprios (override por env).
SEFAZ_PORTAL_NAV_TIMEOUT_MS = int(os.environ.get("SEFAZ_PORTAL_NAV_TIMEOUT_MS", "120000"))
SEFAZ_PORTAL_GOTO_RETRIES = max(1, int(os.environ.get("SEFAZ_PORTAL_GOTO_RETRIES", "5")))
try:
    DOWNLOAD_HISTORY_TIMEOUT_SECONDS = max(
        120,
        int((os.getenv("SEFAZ_DOWNLOAD_HISTORY_TIMEOUT_SECONDS") or "300").strip() or "300"),
    )
except Exception:
    DOWNLOAD_HISTORY_TIMEOUT_SECONDS = 300
MAX_TENTATIVAS_EMPRESA = 5   # tentativas por IE/intervalo em erro geral
DEFAULT_VIEWPORT = {"width": 1366, "height": 768}

# Proxy automático: ativa mitmdump local se não houver proxy configurado
DEFAULT_PROXY_URL = "http://127.0.0.1:8889"
_MITMDUMP_PROC = None
_PROXY_AUTO_STARTED = False
def _default_mitmdump_path() -> str:
    """
    Caminho padrão do mitmdump.exe.

    - Primeiro tenta ao lado do .py/.exe em: data/proxy/mitmdump.exe
    - Se estiver empacotado (PyInstaller), também tenta dentro da pasta interna
      (sys._MEIPASS, normalmente algo como *_internal) em: data/proxy/mitmdump.exe
    """
    if getattr(sys, "frozen", False):
        base = os.path.dirname(os.path.abspath(sys.executable))
    else:
        base = os.path.dirname(os.path.abspath(__file__))

    external = os.path.join(base, "data", "proxy", "mitmdump.exe")

    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        try:
            internal_base = getattr(sys, "_MEIPASS")  # type: ignore[attr-defined]
            internal = os.path.join(internal_base, "data", "proxy", "mitmdump.exe")
            if os.path.exists(internal):
                return internal
        except Exception:
            pass

    return external

MITMDUMP_PATH = _default_mitmdump_path()
MITM_CA_PATH = os.path.join(os.path.dirname(MITMDUMP_PATH), "mitmproxy-ca-cert.cer")

# Ajuste para quem quiser desativar o start automático via env
AUTO_START_PROXY = os.getenv("SEFAZ_AUTO_PROXY", "1").strip() not in {"0", "false", "False"}
RUNTIME_FOLDER_NAME = "sefaz_xml"


def _resolve_runtime_base_dir() -> str:
    explicit = os.environ.get("ROBOT_SCRIPT_DIR", "").strip().rstrip(os.sep)
    if explicit:
        return str(pathlib.Path(explicit).resolve())

    current_dir = pathlib.Path(__file__).resolve().parent
    candidates: list[pathlib.Path] = []
    robots_root = (os.getenv("ROBOTS_ROOT_PATH") or "").strip()

    if robots_root:
        candidates.append(pathlib.Path(robots_root) / RUNTIME_FOLDER_NAME)
    candidates.append(pathlib.Path.home() / "Documents" / "ROBOS" / RUNTIME_FOLDER_NAME)
    candidates.append(current_dir)

    seen: set[str] = set()
    for candidate in candidates:
        try:
            resolved = candidate.resolve()
        except Exception:
            resolved = candidate
        key = str(resolved).lower()
        if key in seen:
            continue
        seen.add(key)
        if (resolved / "data" / "json").exists():
            return str(resolved)

    return str(current_dir)

def get_base_dir():
    """
    Pasta 'externa' da aplicação:
      - no .py  → pasta do arquivo .py (sempre; evita usar ~/Documents/ROBOS/sefaz_xml por engano)
      - no .exe → resolve via _resolve_runtime_base_dir() (data/json ao lado ou em pasta de robôs)
    Usada para coisas que o usuário enxerga/grava (perfis, pastas padrão, etc.).
    """
    if getattr(sys, "frozen", False):
        return _resolve_runtime_base_dir()
    return str(pathlib.Path(__file__).resolve().parent)


def get_internal_dir():
    """
    Pasta 'interna' com arquivos empacotados pelo PyInstaller:
      - no .exe → sys._MEIPASS
      - no .py  → mesma pasta do .py
      - em builds com pasta auxiliar "_internal" ao lado do .exe (Inno/auto-py-to-exe)
    Usada para Chrome portátil, ícones, etc.
    """
    if getattr(sys, "frozen", False):
        if hasattr(sys, "_MEIPASS"):
            return sys._MEIPASS  # type: ignore[attr-defined]
        base = get_base_dir()
        candidate_internal = os.path.join(base, "_internal")
        if os.path.isdir(candidate_internal):
            return candidate_internal
        return base
    return get_base_dir()

def get_data_dir():
    """
    Pasta 'data' contendo Chrome, json, ico, image, proxy, etc.
    Procura primeiro ao lado do .exe/.py; se estiver empacotado, tenta _MEIPASS/data também.
    """
    base = get_base_dir()
    candidate = os.path.join(base, "data")
    if os.path.isdir(candidate):
        return candidate
    internal = get_internal_dir()
    candidate_internal = os.path.join(internal, "data")
    if os.path.isdir(candidate_internal):
        return candidate_internal
    return candidate  # fallback

def get_runtime_dir(app_name: str = "Automacao_Sefaz_NFe_GO"):
    """
    Pasta gravável para runtime/cache (perfil do Chrome via CDP, etc.).

    Quando a aplicação é instalada (ex.: Inno Setup), a pasta do app pode ficar em
    "Program Files" (somente leitura). Por isso, usamos LOCALAPPDATA/APPDATA.
    """
    root = os.getenv("LOCALAPPDATA") or os.getenv("APPDATA") or os.path.expanduser("~")
    runtime = os.path.join(root, app_name)
    try:
        os.makedirs(runtime, exist_ok=True)
    except Exception:
        import tempfile
        runtime = os.path.join(tempfile.gettempdir(), app_name)
        os.makedirs(runtime, exist_ok=True)
    return runtime

def _tail_file(path: str, max_lines: int = 40) -> str:
    """
    Retorna as últimas linhas de um arquivo de log.
    Usado para exibir por que o Chrome falhou ao iniciar.
    """
    if not os.path.exists(path):
        return ""
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()
            return "".join(lines[-max_lines:])
    except Exception:
        return ""



BASE_DIR = get_base_dir()          # onde fica o .py / .exe
INTERNAL_DIR = get_internal_dir()  # onde ficam arquivos empacotados (_MEIPASS)
DATA_DIR = get_data_dir()          # onde ficam Chrome, json, ico, proxy, etc.
LOGS_DIR = os.path.join(DATA_DIR, "logs")
RUNTIME_LOG_PATH = os.path.join(LOGS_DIR, "runtime.log")


def append_runtime_log(message: str) -> None:
    try:
        os.makedirs(LOGS_DIR, exist_ok=True)
        with open(RUNTIME_LOG_PATH, "a", encoding="utf-8") as handle:
            handle.write(str(message).rstrip() + "\n")
    except Exception:
        pass


def emit_terminal_log(message: str) -> str:
    text = str(message or "").rstrip()
    timestamp = datetime.now().strftime("%H:%M:%S")
    line = f"[{timestamp}] [SEFAZ XML] {text}" if text else f"[{timestamp}] [SEFAZ XML]"
    try:
        print(line, flush=True)
    except Exception:
        pass
    return line

CHROME_EXE = os.path.join(DATA_DIR, "Chrome", "chrome.exe")
RUNTIME_DIR = get_runtime_dir()
PROFILE_DIR = os.path.join(RUNTIME_DIR, "chrome_cdp_profile")

# Mantemos uma referência ao processo do Chrome desta automação para conseguir
# encerrá-lo com garantia ao reiniciar a sessão (troca de login).
_AUTOMATION_CHROME_PROC: Optional[subprocess.Popen] = None


def _kill_any_mitm_processes(log_fn=None) -> None:
    """Encerra processos mitm (mitmdump/mitmproxy) no Windows.

    Usado quando queremos reiniciar a automacao "do zero" na troca de login,
    evitando reutilizar uma porta/proxy antigo.
    """
    if os.name != "nt":
        return

    def _log(msg: str):
        if callable(log_fn):
            try:
                log_fn(msg)
            except Exception:
                pass

    killed_any = False
    for name in ("mitmdump.exe", "mitmproxy.exe"):
        try:
            run_kwargs = dict(stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, check=False)
            if hasattr(subprocess, "CREATE_NO_WINDOW"):
                run_kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
            subprocess.run(["taskkill", "/IM", name, "/T", "/F"], **run_kwargs)
            killed_any = True
        except Exception:
            continue

    if killed_any:
        _log("[INFO] 🧹 Processos mitm encerrados (taskkill).")


def _clear_chrome_profile_session(profile_dir: str, log_fn=None) -> None:
    """Limpa apenas dados de sessao/autenticacao do perfil do Chrome.

    Mantem o mesmo user-data-dir (PROFILE_DIR), mas remove cookies e storages para
    evitar que o novo login herde a sessao do login anterior.
    """
    if not profile_dir:
        return

    def _log(msg: str):
        if callable(log_fn):
            try:
                log_fn(msg)
            except Exception:
                pass

    default_dir = os.path.join(profile_dir, "Default")
    if not os.path.isdir(default_dir):
        return

    paths_files = [
        os.path.join(default_dir, "Cookies"),
        os.path.join(default_dir, "Cookies-journal"),
        os.path.join(default_dir, "Web Data-journal"),
        os.path.join(default_dir, "Network", "Cookies"),
        os.path.join(default_dir, "Network", "Cookies-journal"),
    ]
    paths_dirs = [
        os.path.join(default_dir, "Local Storage"),
        os.path.join(default_dir, "Session Storage"),
        os.path.join(default_dir, "IndexedDB"),
        os.path.join(default_dir, "Service Worker"),
    ]

    removed_any = False
    for p in paths_files:
        try:
            if os.path.exists(p):
                os.remove(p)
                removed_any = True
        except Exception:
            pass

    for d in paths_dirs:
        try:
            if os.path.isdir(d):
                shutil.rmtree(d, ignore_errors=True)
                removed_any = True
        except Exception:
            pass

    if removed_any:
        _log("[INFO] 🧽 Sessao do perfil limpa (cookies/storage removidos).")
MITMDUMP_LOG_PATH = os.path.join(os.path.dirname(MITMDUMP_PATH), "mitmdump.log")

def _disable_password_leak_detection():
    """
    Desativa a opção de aviso de senha comprometida no Chrome (perfil CDP).
    Força o valor para False em Preferences e no initial_preferences.
    """
    try:
        pref_dir = os.path.join(PROFILE_DIR, "Default")
        pref_path = os.path.join(pref_dir, "Preferences")
        os.makedirs(pref_dir, exist_ok=True)

        prefs = {}
        if os.path.exists(pref_path):
            try:
                with open(pref_path, "r", encoding="utf-8") as f:
                    prefs = json.load(f)
            except Exception:
                prefs = {}
        profile_section = prefs.get("profile", {})
        profile_section["password_manager_leak_detection"] = False
        prefs["profile"] = profile_section
        with open(pref_path, "w", encoding="utf-8") as f:
            json.dump(prefs, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

    # Ajusta também initial_preferences para novos perfis
    try:
        init_path = os.path.join(os.path.dirname(CHROME_EXE), "initial_preferences")
        if os.path.exists(init_path):
            try:
                with open(init_path, "r", encoding="utf-8") as f:
                    init_prefs = json.load(f)
            except Exception:
                init_prefs = {}
            prof = init_prefs.get("profile", {})
            prof["password_manager_leak_detection"] = False
            init_prefs["profile"] = prof
            with open(init_path, "w", encoding="utf-8") as f:
                json.dump(init_prefs, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def _download_portable_chrome(log_fn=None) -> bool:
    """
    Baixa o Chrome oficial (Chrome for Testing, canal Stable) e substitui data/Chrome.
    """
    def _log(msg: str):
        if callable(log_fn):
            try:
                log_fn(msg)
                return
            except Exception:
                pass
        try:
            print(msg)
        except Exception:
            pass

    try:
        meta_url = "https://googlechromelabs.github.io/chrome-for-testing/known-good-versions-with-downloads.json"
        with urllib.request.urlopen(meta_url, timeout=15) as resp:
            data = json.load(resp)
    except Exception as e:
        _log(f"[WARN] Não foi possível obter metadata do Chrome: {e}")
        return False

    try:
        stable = data.get("channels", {}).get("Stable", {})
        downloads = stable.get("downloads", {}).get("chrome", [])
        win = next((d for d in downloads if d.get("platform") == "win64"), None)
        if not win or not win.get("url"):
            _log("[WARN] Não encontrei URL do Chrome win64 em 'Stable'.")
            return False
        url = win["url"]
    except Exception as e:
        _log(f"[WARN] Não foi possível extrair URL do Chrome: {e}")
        return False

    tmp_dir = tempfile.mkdtemp(prefix="chrome_dl_")
    zip_path = os.path.join(tmp_dir, "chrome.zip")
    try:
        _log(f"[INFO] Baixando Chrome portátil (Stable) de {url} ...")
        with urllib.request.urlopen(url, timeout=120) as resp, open(zip_path, "wb") as f:
            shutil.copyfileobj(resp, f)
    except Exception as e:
        _log(f"[WARN] Falha ao baixar Chrome: {e}")
        shutil.rmtree(tmp_dir, ignore_errors=True)
        return False

    try:
        with zipfile.ZipFile(zip_path, "r") as z:
            z.extractall(tmp_dir)
        # Dentro do zip vem uma pasta chrome-win64
        src_dir = os.path.join(tmp_dir, "chrome-win64")
        if not os.path.exists(os.path.join(src_dir, "chrome.exe")):
            raise FileNotFoundError("chrome.exe não encontrado no pacote baixado.")
        if os.path.exists(os.path.dirname(CHROME_EXE)):
            shutil.rmtree(os.path.dirname(CHROME_EXE), ignore_errors=True)
        shutil.copytree(src_dir, os.path.dirname(CHROME_EXE))
        _log(f"[INFO] Chrome portátil atualizado a partir de {url}.")
        return True
    except Exception as e:
        _log(f"[WARN] Falha ao extrair/instalar Chrome: {e}")
        return False
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)
def get_config_dir():
    """
    Onde o JSON de config vai ficar (sempre dentro de data/json).
    """
    return os.path.join(DATA_DIR, "json")

CONFIG_DIR = get_config_dir()
CONFIG_PATH = os.path.join(CONFIG_DIR, "config.json")

def _resolve_ico_path() -> str:
    candidates = [
        pathlib.Path(DATA_DIR) / "ICO" / "app.ico",
        pathlib.Path(DATA_DIR) / "Ico" / "app.ico",
        pathlib.Path(DATA_DIR) / "ico" / "app.ico",
        pathlib.Path(_base_dir) / "data" / "ICO" / "app.ico",
        pathlib.Path(_base_dir) / "data" / "Ico" / "app.ico",
        pathlib.Path(_base_dir) / "data" / "ico" / "app.ico",
        pathlib.Path(_base_dir) / "app.ico",
    ]
    for candidate in candidates:
        if candidate.exists():
            return str(candidate)
    return str(candidates[0])


# ICO (dentro da pasta data/ICO)
ICO_PATH = _resolve_ico_path()

# =============================================================================
# Estilos de botões (mesmo padrão do bot NFS)
# =============================================================================

BTN_COMMON = (
    "font:9pt 'Verdana'; font-weight:bold; color:#fff; padding:8px 14px; "
    "border-radius:8px; border:1px solid rgba(255,255,255,0.08);"
)


def button_style(base: str, hover: str, pressed: str) -> str:
    """
    Retorna stylesheet para um botão no mesmo estilo do NFS (compacto e moderno).
    """
    return f"""
    QPushButton {{
        font: 9pt 'Verdana';
        font-weight: bold;
        color: #E8F4FF;
        padding: 9px 14px;
        border-radius: 10px;
        border: 1px solid rgba(255,255,255,0.12);
        background: qlineargradient(x1:0,y1:0, x2:1,y2:1,
            stop:0 {base}, stop:1 #0F1E35);
    }}
    QPushButton:hover {{
        background: qlineargradient(x1:0,y1:0, x2:1,y2:1,
            stop:0 {hover}, stop:1 #12304E);
    }}
    QPushButton:pressed {{
        background: {pressed};
        padding-top: 11px;
        padding-bottom: 7px;
    }}
    """




def small_button_style(base: str, hover: str, pressed: str) -> str:
    """
    Stylesheet para botões pequenos (ex.: remover item da lista),
    sem padding horizontal grande para não ocultar o texto.
    """
    return f"""
    QPushButton {{
        font: 10pt 'Verdana';
        font-weight: bold;
        color: #E8F4FF;
        padding: 0px;
        border-radius: 10px;
        border: 1px solid rgba(255,255,255,0.12);
        background: qlineargradient(x1:0,y1:0, x2:1,y2:1,
            stop:0 {base}, stop:1 #0F1E35);
    }}
    QPushButton:hover {{
        background: qlineargradient(x1:0,y1:0, x2:1,y2:1,
            stop:0 {hover}, stop:1 #12304E);
    }}
    QPushButton:pressed {{
        background: {pressed};
    }}
    """


class _ButtonGlowFilter(QObject):
    """Event filter that toggles a drop-shadow glow on hover."""

    def eventFilter(self, obj, event):
        effect = getattr(obj, "_button_glow_effect", None)
        if not isinstance(effect, QGraphicsDropShadowEffect):
            return False
        if event.type() == QEvent.Enter:
            effect.setEnabled(True)
        elif event.type() == QEvent.Leave:
            effect.setEnabled(False)
        return False


_BUTTON_GLOW_FILTER = _ButtonGlowFilter()


def _glow_color_from_hex(value: str) -> QColor:
    """Normalize a color string and make it slightly brighter for the glow."""
    try:
        color = QColor(value)
        if not color.isValid():
            raise ValueError()
    except Exception:
        color = QColor("#E8F4FF")
    glow = color.lighter(140)
    glow.setAlpha(200)
    return glow


def install_button_glow(button: QPushButton, base_color: str) -> None:
    """
    Ensures the button receives a glow effect that matches its base color.
    Subsequent calls update the glow color without reinstalling the filter.
    """
    glow = _glow_color_from_hex(base_color)
    effect = getattr(button, "_button_glow_effect", None)
    if isinstance(effect, QGraphicsDropShadowEffect):
        effect.setColor(glow)
    else:
        effect = QGraphicsDropShadowEffect(button)
        effect.setBlurRadius(24)
        effect.setOffset(0, 0)
        effect.setColor(glow)
        effect.setEnabled(False)
        button.setGraphicsEffect(effect)
        button._button_glow_effect = effect
        if not getattr(button, "_button_glow_filter_installed", False):
            button.installEventFilter(_BUTTON_GLOW_FILTER)
            button._button_glow_filter_installed = True
    effect.setEnabled(False)


def set_button_style(button: QPushButton, base: str, hover: str, pressed: str) -> None:
    """Apply the shared stylesheet to a button and wire the glow effect."""
    button.setStyleSheet(button_style(base, hover, pressed))
    install_button_glow(button, base)


def set_small_button_style(button: QPushButton, base: str, hover: str, pressed: str) -> None:
    """Apply the small-button stylesheet and reuse the same glow wiring."""
    button.setStyleSheet(small_button_style(base, hover, pressed))
    install_button_glow(button, base)
# =============================================================================
# Licença (armazenamento local + validação via Supabase)
# =============================================================================

# IMPORTANTE (Supabase):
# - NUNCA use a service_role key dentro do app do cliente.
# - No Supabase, ative RLS na tabela "licenses" e deixe o role "anon" com PERMISSÃO APENAS DE SELECT.
# - Não dê permissão de INSERT/UPDATE/DELETE para "anon" em "licenses" (isso impede o cliente de forjar licença no seu banco).

from supabase import create_client

LICENSE_SUPABASE_URL = (os.environ.get("LICENSE_SUPABASE_URL") or os.environ.get("SUPABASE_URL") or "").strip()
LICENSE_SUPABASE_ANON_KEY = (os.environ.get("LICENSE_SUPABASE_ANON_KEY") or os.environ.get("SUPABASE_ANON_KEY") or "").strip()

# cliente supabase (somente para checar licença)
def _get_supabase_client():
    # cria o client sempre que for checar (evita objeto global fácil de mexer)
    if not LICENSE_SUPABASE_URL or not LICENSE_SUPABASE_ANON_KEY:
        raise RuntimeError(
            "Supabase (licença) não configurado. Defina LICENSE_SUPABASE_URL e LICENSE_SUPABASE_ANON_KEY (ou SUPABASE_URL/SUPABASE_ANON_KEY)."
        )
    return create_client(LICENSE_SUPABASE_URL, LICENSE_SUPABASE_ANON_KEY)

# Arquivo local onde salvaremos a licença (dentro da pasta JSON já usada pelo app)
LICENSE_PATH = os.path.join(CONFIG_DIR, "license.json")

# Arquivo onde vamos guardar login/senha do portal
LOGIN_PATH = os.path.join(CONFIG_DIR, "login_portal.json")

# -------------------------------------------------------------------------
# Helpers de armazenamento local (NÃO CONFIE SÓ NELES – sempre valida no Supabase)
# -------------------------------------------------------------------------
def _load_license_local() -> dict:
    """Lê o JSON local com a licença salva."""
    if os.path.exists(LICENSE_PATH):
        try:
            with open(LICENSE_PATH, "r", encoding="utf-8") as f:
                return json.load(f) or {}
        except Exception:
            return {}
    return {}


def _save_license_local(data: dict) -> None:
    """Grava/atualiza o JSON local de licença."""
    os.makedirs(CONFIG_DIR, exist_ok=True)
    with open(LICENSE_PATH, "w", encoding="utf-8") as f:
        json.dump(data or {}, f, ensure_ascii=False, indent=2)


def get_saved_key() -> str:
    """Devolve a chave de licença já salva localmente (se houver)."""
    return str(_load_license_local().get("license_key", "")).strip()


def _load_login_portal_data() -> dict:
    """Carrega o JSON bruto de logins do portal (compatível com formato antigo)."""
    if os.path.exists(LOGIN_PATH):
        try:
            with open(LOGIN_PATH, "r", encoding="utf-8") as f:
                return json.load(f) or {}
        except Exception:
            return {}
    return {}


def _normalize_sefaz_go_logins(raw: Any) -> List[Dict[str, str]]:
    if not isinstance(raw, list):
        return []
    normalized: List[Dict[str, str]] = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        cpf = cpf_somente_digitos(item.get("cpf", ""))
        senha = str(item.get("password") or item.get("senha") or "").strip()
        if len(cpf) == 11 and senha:
            normalized.append(
                {
                    "cpf": cpf,
                    "senha": senha,
                    "is_default": bool(item.get("is_default")),
                }
            )
    if normalized and not any(item.get("is_default") for item in normalized):
        normalized[0]["is_default"] = True
    return normalized


def _load_dashboard_portal_logins() -> List[Dict[str, str]]:
    robot_row = fetch_robot_row()
    if robot_row:
        logins = _normalize_sefaz_go_logins(robot_row.get("global_logins"))
        if logins:
            return logins

    rows = load_companies_from_dashboard()
    if not rows:
        return []
    aggregated: Dict[str, Dict[str, str]] = {}
    for row in rows:
        for login in _normalize_sefaz_go_logins(row.get("legacy_company_logins")):
            cpf = cpf_somente_digitos(login.get("cpf", ""))
            senha = str(login.get("senha", "")).strip()
            if cpf and senha and cpf not in aggregated:
                aggregated[cpf] = {"cpf": cpf, "senha": senha, "is_default": bool(login.get("is_default"))}
    ordered = list(aggregated.values())
    ordered.sort(key=lambda item: (0 if item.get("is_default") else 1, item.get("cpf", "")))
    return ordered


def load_logins_portal() -> List[Dict[str, str]]:
    """Retorna a lista de logins salvos no portal (cada item: cpf, senha)."""
    if not SEFAZ_LOCAL_JSON_CLIENT_MODE:
        dashboard_logins = _load_dashboard_portal_logins()
        if dashboard_logins:
            return dashboard_logins

    data = _load_login_portal_data()
    logins = data.get("logins")
    if isinstance(logins, list):
        out = []
        for item in logins:
            if not isinstance(item, dict):
                continue
            cpf = cpf_somente_digitos(item.get("cpf", ""))
            senha = str(item.get("senha", "")).strip()
            if cpf or senha:
                out.append({"cpf": cpf, "senha": senha})
        return out

    # Formato antigo (cpf/senha únicos)
    cpf = cpf_somente_digitos(data.get("cpf", ""))
    senha = str(data.get("senha", "")).strip()
    if cpf or senha:
        return [{"cpf": cpf, "senha": senha}]
    return []


def save_logins_portal(logins: List[Dict[str, str]], default_cpf: str | None = None) -> None:
    """Salva lista de logins (cpf/senha) no arquivo login_portal.json."""
    os.makedirs(CONFIG_DIR, exist_ok=True)
    cleaned = []
    for item in (logins or []):
        if not isinstance(item, dict):
            continue
        cpf = cpf_somente_digitos(item.get("cpf", ""))
        senha = str(item.get("senha", "")).strip()
        if cpf or senha:
            cleaned.append({"cpf": cpf, "senha": senha})
    default_cpf_digits = cpf_somente_digitos(default_cpf or "")
    if not default_cpf_digits and cleaned:
        default_cpf_digits = cleaned[0].get("cpf", "")
    data = {
        "logins": cleaned,
        "default_cpf": default_cpf_digits,
    }
    with open(LOGIN_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_login_portal(preferred_cpf: str | None = None, strict_preferred: bool = False) -> Tuple[str, str]:
    """
    Lê o JSON de login do portal e devolve (cpf, senha).
    Se não existir ou der erro, devolve ("", "").
    """
    logins = load_logins_portal()
    if not logins:
        return "", ""

    prefer = cpf_somente_digitos(preferred_cpf or "")
    if prefer:
        for item in logins:
            if cpf_somente_digitos(item.get("cpf", "")) == prefer:
                return item.get("cpf", ""), item.get("senha", "")
        if strict_preferred:
            return "", ""

    default_cpf = ""
    for item in logins:
        if item.get("is_default"):
            default_cpf = cpf_somente_digitos(item.get("cpf", ""))
            break
    if not default_cpf:
        data = _load_login_portal_data()
        default_cpf = cpf_somente_digitos(data.get("default_cpf", ""))
    if default_cpf:
        for item in logins:
            if cpf_somente_digitos(item.get("cpf", "")) == default_cpf:
                return item.get("cpf", ""), item.get("senha", "")

    first = logins[0]
    return first.get("cpf", ""), first.get("senha", "")


def save_login_portal(cpf: str, senha: str) -> None:
    """
    Salva CPF e senha no arquivo login_portal.json dentro da pasta JSON.
    (compatível com o formato antigo)
    """
    save_logins_portal([{"cpf": cpf, "senha": senha}], default_cpf=cpf)

def persist_key(key: str, meta: dict | None = None) -> None:
    """Salva a chave (e metadados) no JSON local (sem amarrar a máquina)."""
    payload = _load_license_local()
    payload["license_key"] = (key or "").strip()
    if meta:
        payload.update(meta)
    _save_license_local(payload)


def check_license_with_supabase(key: str) -> tuple[bool, str, dict]:
    """
    Consulta a função RPC 'verify_license' no Supabase e valida a chave.

    A função verify_license já filtra:
      - is_active = TRUE
      - expires_at >= NOW()

    Retorna: (ok, mensagem, meta)
      meta pode conter: {"expires_at": ..., "licensee": "..."}
    """
    key = (key or "").strip()
    if not key:
        return (False, "Informe a chave de licença.", {})

    try:
        # Cria um cliente Supabase local usando o helper para evitar variáveis globais indefinidas
        client = _get_supabase_client()
        # Chama a função RPC criada no Supabase:
        #   CREATE FUNCTION public.verify_license(p_key text) ...
        res = client.rpc(
            "verify_license",
            {"p_key": key}
        ).execute()

        rows = res.data if hasattr(res, "data") else []
    except Exception as e:
        return (False, f"Falha ao contatar o servidor de licença: {e}", {})

    # Se não voltou nenhuma linha, a chave é inválida / expirada / inativa
    if not rows:
        return (False, "Chave inválida, expirada ou inativa.", {})

    row = rows[0] or {}

    meta = {
        "expires_at": row.get("expires_at"),
        # no seu banco, o nome do cliente está em client_name
        "licensee": row.get("client_name") or "",
    }
    return (True, "Licença válida.", meta)


# -------------------------------------------------------------------------
# Diálogo de licença (UI) – mesma lógica, só usando a função acima
# -------------------------------------------------------------------------
class LicenseDialog(QDialog):
    """Janela estilizada para digitar/ativar a chave de licença."""
    ACCENT = "#7C3AED"   # roxo moderno (mude se quiser)
    BG1    = "#0B1220"   # gradiente início
    BG2    = "#111827"   # gradiente fim
    CARD   = "#0F172A"   # fundo do card
    STROKE = "rgba(255,255,255,0.08)"
    TEXT   = "rgba(255,255,255,0.92)"
    MUTED  = "rgba(255,255,255,0.65)"

    def __init__(self, parent=None, preset_key: str = ""):
        super().__init__(parent)
        self.setObjectName("LicenseDialog")
        self.setModal(True)
        self.setMinimumWidth(520)
        self.setWindowTitle("Ativação da Licença")
        try:
            self.setWindowIcon(QIcon(ICO_PATH))
        except Exception:
            pass

        # ====== ESTILO (QSS) aplicado só neste diálogo ======
        self.setStyleSheet(f"""
        QDialog#LicenseDialog {{
            background: qlineargradient(x1:0,y1:0, x2:1,y2:1,
                        stop:0 {self.BG1}, stop:1 {self.BG2});
        }}
        QLabel, QLineEdit, QPushButton {{
            color: {self.TEXT};
            font-family: "Segoe UI", "Inter", "Roboto", "Ubuntu", "Arial";
            font-size: 14px;
        }}
        QFrame#Card {{
            background: {self.CARD};
            border: 1px solid {self.STROKE};
            border-radius: 18px;
        }}
        QLabel#Title {{
            font-size: 22px;
            font-weight: 700;
            color: {self.TEXT};
        }}
        QLabel#Subtitle {{
            font-size: 13px;
            color: {self.MUTED};
        }}
        QLineEdit#KeyEdit {{
            background: #0B1220;
            border: 1px solid #243041;
            border-radius: 12px;
            padding: 12px 14px;
            selection-background-color: {self.ACCENT};
            selection-color: white;
        }}
        QLineEdit#KeyEdit:focus {{
            border: 1px solid {self.ACCENT};
            box-shadow: 0 0 0 3px rgba(124,58,237,0.25);
        }}
        QPushButton#PrimaryButton {{
            background: {self.ACCENT};
            border: none;
            border-radius: 12px;
            padding: 10px 16px;
            font-weight: 600;
            color: white;
        }}
        QPushButton#PrimaryButton:hover {{
            background: #6D28D9;
        }}
        QPushButton#PrimaryButton:disabled {{
            background: #5B21B6;
            opacity: 0.6;
        }}
        QPushButton#GhostButton {{
            background: transparent;
            border: 1px solid {self.STROKE};
            border-radius: 12px;
            padding: 10px 16px;
            color: {self.MUTED};
        }}
        QPushButton#GhostButton:hover {{
            border: 1px solid #2A374B;
            color: {self.TEXT};
        }}
        QPushButton#LinkButton {{
            background: transparent;
            border: none;
            color: {self.ACCENT};
            font-weight: 600;
        }}
        QPushButton#LinkButton:hover {{
            text-decoration: underline;
        }}
        QLabel#Msg {{
            color: {self.MUTED};
        }}
        """)

        # ====== LAYOUT ROOT ======
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 20, 20, 20)

        # ====== CARD ======
        card = QFrame(self)
        card.setObjectName("Card")
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(28, 28, 28, 28)
        card_layout.setSpacing(16)

        # Sombra suave no card
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(32)
        shadow.setColor(Qt.black)
        shadow.setOffset(0, 12)
        card.setGraphicsEffect(shadow)

        # ====== TOPO: LOGO + TÍTULO ======
        header = QHBoxLayout()
        header.setSpacing(14)

        logo_lbl = QLabel(card)
        logo_pix = None
        try:
            logo_pix = QPixmap(ICO_PATH)
        except Exception:
            pass
        if not logo_pix or logo_pix.isNull():
            logo_pix = QPixmap(64, 64)
            logo_pix.fill(Qt.transparent)
        logo_lbl.setPixmap(logo_pix.scaled(48, 48, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        header.addWidget(logo_lbl, 0, Qt.AlignTop)

        title_box = QVBoxLayout()
        title = QLabel("Ative sua licença", card)
        title.setObjectName("Title")
        self.subtitle = QLabel("Digite sua chave para liberar todas as funcionalidades.", card)
        self.subtitle.setObjectName("Subtitle")
        self.subtitle.setWordWrap(True)
        title_box.addWidget(title)
        title_box.addWidget(self.subtitle)
        header.addLayout(title_box)
        header.addStretch(1)
        card_layout.addLayout(header)

        # ====== CAMPO DA CHAVE + BOTÕES AUXILIARES ======
        key_row = QHBoxLayout()
        key_row.setSpacing(10)

        self.edt = QLineEdit(card)
        self.edt.setObjectName("KeyEdit")
        self.edt.setPlaceholderText("XXXXX-XXXXX-XXXXX-XXXXX")
        if preset_key:
            self.edt.setText(preset_key)
        self.edt.returnPressed.connect(self._on_activate)

        paste_btn = QPushButton("Colar", card)
        set_button_style(paste_btn, "#2980B9", "#2471A3", "#1F618D")
        paste_btn.setObjectName("GhostButton")
        paste_btn.setCursor(Qt.PointingHandCursor)
        paste_btn.clicked.connect(self._paste_from_clipboard)

        key_row.addWidget(self.edt, 1)
        key_row.addWidget(paste_btn, 0)
        card_layout.addLayout(key_row)

        # ====== MENSAGEM (feedback de validação) ======
        self.msg = QLabel("", card)
        self.msg.setObjectName("Msg")
        self.msg.setWordWrap(True)
        card_layout.addWidget(self.msg)

        # Espaço flexível
        card_layout.addItem(QSpacerItem(0, 8, QSizePolicy.Minimum, QSizePolicy.Expanding))

        # ====== BOTÕES PRINCIPAIS ======
        btns_row = QHBoxLayout()
        btns_row.addStretch(1)

        self.btn_cancel = QPushButton("Sair", card)
        set_button_style(self.btn_cancel, "#7F8C8D", "#95A5A6", "#707B7C")
        self.btn_cancel.setObjectName("GhostButton")
        self.btn_cancel.setCursor(Qt.PointingHandCursor)
        self.btn_cancel.clicked.connect(self.reject)

        self.btn_ok = QPushButton("Ativar", card)
        set_button_style(self.btn_ok, "#7C3AED", "#6D28D9", "#5B21B6")
        self.btn_ok.setObjectName("PrimaryButton")
        self.btn_ok.setCursor(Qt.PointingHandCursor)
        self.btn_ok.clicked.connect(self._on_activate)

        btns_row.addWidget(self.btn_cancel)
        btns_row.addWidget(self.btn_ok)
        card_layout.addLayout(btns_row)

        # ====== RODAPÉ (ajuda) ======
        help_row = QHBoxLayout()
        help_row.addStretch(1)
        help_lbl = QLabel(
            '<span style="color:rgba(255,255,255,0.55)">Precisa de ajuda?</span> '
            '<a style="color:#7C3AED; text-decoration:none;" href="https://wa.me/5562993626968"><b>Fale com o suporte</b></a>',
            card
        )
        help_lbl.setTextFormat(Qt.RichText)
        help_lbl.setOpenExternalLinks(True)
        help_row.addWidget(help_lbl, 0, Qt.AlignRight)
        card_layout.addLayout(help_row)

        root.addWidget(card)

        # ====== ANIMAÇÃO DE ENTRADA ======
        self.setWindowOpacity(0.0)
        QTimer.singleShot(0, self._animate_entrance)

    def set_revoked_mode(self, reason: str | None = None):
        """
        Mostra o aviso de licença revogada APENAS no subtítulo do topo.
        Limpa a mensagem inferior para não duplicar.
        """
        base = "Sua licença foi revogada. Entre em contato com o suporte para reativar."
        if reason:
            base = f"{base}\n{reason}"

        try:
            self.subtitle.setText(base)
        except Exception:
            pass

        try:
            self.msg.clear()
        except Exception:
            pass

    # --------- AÇÕES ---------
    def _paste_from_clipboard(self):
        cb = QApplication.clipboard()
        if cb:
            self.edt.setText(cb.text().strip())

    def _animate_entrance(self):
        anim = QPropertyAnimation(self, b"windowOpacity", self)
        anim.setDuration(220)
        anim.setStartValue(0.0)
        anim.setEndValue(1.0)
        anim.setEasingCurve(QEasingCurve.OutCubic)
        anim.start(QPropertyAnimation.DeleteWhenStopped)

    def _on_activate(self):
        key = self.edt.text().strip()
        self._set_busy(True)
        ok, msg, meta = check_license_with_supabase(key)
        self._set_busy(False)

        # feedback visual
        self.msg.setText(msg or "")

        if ok:
            persist_key(key, meta)
            os.environ["AUTOMATIZE_VERIFIED"] = "1"
            self.accept()
            return

        # Se estiver revogada/inativa, muda o subtítulo para o aviso
        low = (msg or "").lower()
        if "desativad" in low or "inativa" in low or "revog" in low:
            self.set_revoked_mode(msg)

    def _set_busy(self, busy: bool):
        self.btn_ok.setEnabled(not busy)
        self.btn_cancel.setEnabled(not busy)
        self.edt.setEnabled(not busy)


def _warn_if_expiring_soon(parent=None):
    """Exibe um aviso se a licença estiver para vencer (<= 7 dias)."""
    meta = _load_license_local()
    exp = str(meta.get("expires_at") or "").strip()
    if not exp:
        return
    try:
        from datetime import datetime as _dt, timedelta as _td
        if "T" in exp:
            exp_dt = _dt.fromisoformat(exp.replace("Z", "+00:00"))
        else:
            exp_dt = _dt.strptime(exp, "%Y-%m-%d")
        days = (exp_dt.date() - _dt.now(exp_dt.tzinfo or None).date()).days
        if days <= 7:
            QMessageBox.warning(
                parent, "Licença perto de vencer",
                f"Sua licença expira em {exp_dt.date().strftime('%d/%m/%Y')} (em {max(days,0)} dia(s))."
            )
    except Exception:
        pass


def ensure_license_valid(parent_app: QApplication) -> bool:
    """
    Garante que há uma licença válida antes de abrir a UI principal.
    Sempre valida no Supabase – não confia apenas no JSON local.
    """
    if SEFAZ_LOCAL_JSON_CLIENT_MODE:
        os.environ["AUTOMATIZE_VERIFIED"] = "1"
        return True

    last_fail_msg = None

    # 1) tenta licença salva (sem apagar se der erro)
    saved = get_saved_key()
    if saved:
        ok, msg, meta = check_license_with_supabase(saved)
        if ok:
            persist_key(saved, meta)
            os.environ["AUTOMATIZE_VERIFIED"] = "1"
            _warn_if_expiring_soon()
            return True
        else:
            last_fail_msg = msg  # guardamos o motivo (p.ex., revogada/inativa)

    # 2) pede no diálogo até validar ou cancelar
    while True:
        dlg = LicenseDialog(preset_key=get_saved_key() or "")
        low = (last_fail_msg or "").lower()
        if "desativad" in low or "inativa" in low or "revog" in low:
            dlg.set_revoked_mode(last_fail_msg)

        if dlg.exec() == QDialog.Accepted:
            _warn_if_expiring_soon()
            return True

        # Cancelou → confirmar saída
        confirm = ConfirmDialog(
            title="É necessário ativar a licença",
            message="Para continuar, você precisa ativar a licença.\n"
                    "Deseja encerrar o aplicativo agora?",
            primary_text="Encerrar",
            secondary_text="Voltar e ativar"
        )
        confirm.exec()
        if confirm.choice == "primary":
            return False
        # se escolher voltar, reabre o diálogo de licença

def is_scheduler_mode_enabled() -> bool:
    return str(os.environ.get("AUTOMATIZE_SCHEDULER_MODE") or "").strip().lower() in {
        "1",
        "true",
        "yes",
        "sim",
        "on",
    }


class ConfirmDialog(QDialog):
    """Confirmação estilizada (mesmo visual da LicenseDialog)."""
    ACCENT = "#7C3AED"   # roxo moderno (troque se quiser)
    BG1    = "#0B1220"
    BG2    = "#111827"
    CARD   = "#0F172A"
    STROKE = "rgba(255,255,255,0.08)"
    TEXT   = "rgba(255,255,255,0.92)"
    MUTED  = "rgba(255,255,255,0.65)"

    def __init__(
        self,
        title: str = "Confirmação",
        message: str = "Deseja continuar?",
        primary_text: str = "OK",
        secondary_text: str = "Cancelar",
        parent=None
    ):
        super().__init__(parent)
        self.setObjectName("ConfirmDialog")
        self.setModal(True)
        self.setMinimumWidth(520)
        self.setWindowTitle(title)
        try:
            self.setWindowIcon(QIcon(ICO_PATH))
        except Exception:
            pass
        self.choice = None  # "primary" ou "secondary"

        # ====== ESTILO (QSS) ======
        self.setStyleSheet(f"""
        QDialog#ConfirmDialog {{
            background: qlineargradient(x1:0,y1:0, x2:1,y2:1,
                        stop:0 {self.BG1}, stop:1 {self.BG2});
        }}
        QLabel, QPushButton {{
            color: {self.TEXT};
            font-family: "Segoe UI", "Inter", "Roboto", "Ubuntu", "Arial";
            font-size: 14px;
        }}
        QFrame#Card {{
            background: {self.CARD};
            border: 1px solid {self.STROKE};
            border-radius: 18px;
        }}
        QLabel#Title {{
            font-size: 22px;
            font-weight: 700;
            color: {self.TEXT};
        }}
        QLabel#Msg {{
            font-size: 14px;
            color: {self.MUTED};
        }}
        QPushButton#PrimaryButton {{
            background: {self.ACCENT};
            border: none;
            border-radius: 12px;
            padding: 10px 16px;
            font-weight: 600;
            color: white;
        }}
        QPushButton#PrimaryButton:hover {{
            background: #6D28D9;
        }}
        QPushButton#GhostButton {{
            background: transparent;
            border: 1px solid {self.STROKE};
            border-radius: 12px;
            padding: 10px 16px;
            color: {self.MUTED};
        }}
        QPushButton#GhostButton:hover {{
            border: 1px solid #2A374B;
            color: {self.TEXT};
        }}
        """)

        # ====== LAYOUT ROOT ======
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 20, 20, 20)

        # ====== CARD ======
        card = QFrame(self)
        card.setObjectName("Card")
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(28, 28, 28, 28)
        card_layout.setSpacing(16)

        # sombra suave
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(32)
        shadow.setColor(Qt.black)
        shadow.setOffset(0, 12)
        card.setGraphicsEffect(shadow)

        # título
        lbl_title = QLabel(title, card)
        lbl_title.setObjectName("Title")
        card_layout.addWidget(lbl_title)

        # mensagem
        lbl_msg = QLabel(message, card)
        lbl_msg.setObjectName("Msg")
        lbl_msg.setWordWrap(True)
        card_layout.addWidget(lbl_msg)

        # espaço
        card_layout.addItem(QSpacerItem(0, 8, QSizePolicy.Minimum, QSizePolicy.Expanding))

        # botões
        btns = QHBoxLayout()
        btns.addStretch(1)

        self.btn_secondary = QPushButton(secondary_text, card)
        set_button_style(self.btn_secondary, "#7F8C8D", "#95A5A6", "#707B7C")
        self.btn_secondary.setObjectName("GhostButton")
        self.btn_secondary.setCursor(Qt.PointingHandCursor)
        self.btn_secondary.clicked.connect(self._on_secondary)

        self.btn_primary = QPushButton(primary_text, card)
        set_button_style(self.btn_primary, "#7C3AED", "#6D28D9", "#5B21B6")
        self.btn_primary.setObjectName("PrimaryButton")
        self.btn_primary.setCursor(Qt.PointingHandCursor)
        self.btn_primary.clicked.connect(self._on_primary)

        btns.addWidget(self.btn_secondary)
        btns.addWidget(self.btn_primary)
        card_layout.addLayout(btns)

        root.addWidget(card)

    def _on_primary(self):
        self.choice = "primary"
        self.accept()

    def _on_secondary(self):
        self.choice = "secondary"
        self.accept()

# Navegação “Baixar XML NFE”
XPATH_MENU = '//*[@id="menuBox"]/ul/li[7]/h3'
XPATH_ITEM = '//*[@id="menuBox"]/ul/li[7]/ul/li'
# Menu superior usado para reabrir a tela "Baixar XML NFE" depois de recarregar a página
XPATH_MENU_SUPERIOR_BAIXAR_XML = '/html/body/div[1]/div[1]/ul/li[4]/a'

# Pesquisa
XPATH_BTN_PESQUISAR = '//*[@id="btnPesquisar"]'
XPATH_ALERT_BUTTON = '//*[@id="message-containter"]/div/button'
XPATH_RECAPTCHA_CHECKBOX = "//input[@type='checkbox']"  # NOVO: marcar recaptcha antes de pesquisar
XPATH_RECAPTCHA_BLOCO = '//div[@data-callback="pegarTokenSuccess"]'  # Bloco do captcha (turnstile) preferencial
XPATH_CAMPO_DATA_INICIAL = '//*[@id="cmpDataInicial"]'
XPATH_CAMPO_DATA_FINAL = '//*[@id="cmpDataFinal"]'
XPATH_CAMPO_IE = '//*[@id="cmpNumIeDest"]'

# Download em lote (modal)
XPATH_BTN_BAIXAR_TODOS = '//*[@id="content"]/div[2]/div/button'
XPATH_MODAL_OPT_DOCS = '/html/body/div[5]/div/div/div[2]/div/label[1]/input'
XPATH_MODAL_OPT_EVENTOS = '/html/body/div[5]/div/div/div[2]/div/label[2]/input'
XPATH_MODAL_OPT_DOCS_EVENTOS = '/html/body/div[5]/div/div/div[2]/div/label[3]/input'
XPATH_MODAL_BTN_OK = '//*[@id="dnwld-all-btn-ok"]'
# Label de falta de permissão após clicar em "Baixar todos os arquivos"
XPATH_LABEL_SEM_PERMISSAO = "//label[normalize-space(text())='Você não tem permissão para acessar esta página.']"
# Botão "Histórico de Downloads" (logo abaixo do "Baixar todos os arquivos")
XPATH_BTN_HISTORICO_DOWNLOADS = '//*[@id="content"]/div[2]/div/a[2]/button'


# =============================================================================
# Intervalos de datas
# =============================================================================

def gerar_intervalos_mensais(data_inicial: str, data_final: str) -> List[Tuple[str, str]]:
    """Gera blocos 01..último dia de cada mês no range informado.
       Se vazio, usa mês anterior completo."""
    if not data_inicial.strip() or not data_final.strip():
        hoje = datetime.now()
        primeiro_dia = (hoje.replace(day=1) - timedelta(days=1)).replace(day=1)
        ultimo_dia = hoje.replace(day=1) - timedelta(days=1)
        return [(primeiro_dia.strftime("%d/%m/%Y"), ultimo_dia.strftime("%d/%m/%Y"))]

    intervalos = []
    inicio = datetime.strptime(data_inicial, "%d/%m/%Y")
    fim = datetime.strptime(data_final, "%d/%m/%Y")

    while inicio <= fim:
        inicio_mes = inicio.replace(day=1)
        proximo_mes = inicio_mes + timedelta(days=32)
        fim_mes = (proximo_mes.replace(day=1) - timedelta(days=1))
        if fim_mes > fim:
            fim_mes = fim
        intervalos.append(
            (inicio_mes.strftime("%d/%m/%Y"), fim_mes.strftime("%d/%m/%Y"))
        )
        inicio = proximo_mes

    return intervalos


def gerar_intervalos_diarios(data_inicial: str, data_final: str, intervalo_dias: int) -> List[Tuple[str, str]]:
    """Quebra o período em blocos de N dias (1..N, N+1..2N, etc.)."""
    intervals = []
    start = datetime.strptime(data_inicial, "%d/%m/%Y")
    end = datetime.strptime(data_final, "%d/%m/%Y")
    while start <= end:
        interval_end = start + timedelta(days=intervalo_dias - 1)
        if interval_end > end:
            interval_end = end
        intervals.append(
            (start.strftime("%d/%m/%Y"), interval_end.strftime("%d/%m/%Y"))
        )
        start = interval_end + timedelta(days=1)
    return intervals


# =============================================================================
# Helpers de IE (máscara / normalização)
# =============================================================================

def ie_somente_digitos(ie: str) -> str:
    """Remove pontos, traços, etc. Deixa só números."""
    return "".join(ch for ch in (ie or "") if ch.isdigit())


def ie_formatada(ie: str) -> str:
    """
    Formata a IE no padrão 10.920.217-1 (2-3-3-1 dígitos).
    Se não tiver 9 dígitos, devolve como veio.
    """
    num = ie_somente_digitos(ie)
    if len(num) == 9:
        return f"{num[0:2]}.{num[2:5]}.{num[5:8]}-{num[8]}"
    return ie


# =============================================================================
# Helpers de CNPJ (normalização / máscara)
# =============================================================================

def cnpj_somente_digitos(cnpj: str) -> str:
    """Remove tudo que não for dígito do CNPJ."""
    return "".join(ch for ch in (cnpj or "") if ch.isdigit())


def cnpj_formatado(cnpj: str) -> str:
    """Formata CNPJ no padrão 00.000.000/0000-00; se inválido, devolve como veio."""
    num = cnpj_somente_digitos(cnpj)
    if len(num) == 14:
        return f"{num[0:2]}.{num[2:5]}.{num[5:8]}/{num[8:12]}-{num[12:14]}"
    return cnpj


# =============================================================================
# Helpers de CPF (normalização / máscara)
# =============================================================================

def cpf_somente_digitos(cpf: str) -> str:
    """Remove tudo que não for dígito do CPF."""
    return "".join(ch for ch in (cpf or "") if ch.isdigit())


def cpf_formatado(cpf: str) -> str:
    """Formata CPF no padrão 000.000.000-00; se inválido, devolve como veio."""
    num = cpf_somente_digitos(cpf)
    if len(num) == 11:
        return f"{num[0:3]}.{num[3:6]}.{num[6:9]}-{num[9:11]}"
    return cpf


def normalize_company_name(name: str) -> str:
    return re.sub(r"\s+", " ", str(name or "").strip()).upper()


def safe_folder_name(name: str) -> str:
    """Remove caracteres inválidos para nome de pasta no Windows (e evita nomes problemáticos)."""
    invalid = '<>:"/\\|?*'
    cleaned = "".join(c for c in (name or "").strip() if c not in invalid)

    # Windows não aceita nomes terminando com ponto/espaço
    cleaned = cleaned.strip().rstrip(". ")
    return cleaned or "Empresa"


_WIN_LONG_PATH_THRESHOLD = 240
_WIN_MAX_FILENAME = 240


def _is_windows() -> bool:
    return sys.platform.startswith("win")


def _normalize_fs_path(path: str) -> str:
    if not path:
        return path
    try:
        return os.path.abspath(os.path.expanduser(path))
    except Exception:
        return path


def _needs_long_path_prefix(path: str) -> bool:
    if not _is_windows() or not path:
        return False
    p = str(path)
    if p.startswith("\\\\?\\"):
        return False
    try:
        p = _normalize_fs_path(p)
    except Exception:
        pass
    return len(p) >= _WIN_LONG_PATH_THRESHOLD


def _to_long_path(path: str) -> str:
    if not _is_windows() or not path:
        return path
    p = str(path)
    if p.startswith("\\\\?\\"):
        return p
    p = _normalize_fs_path(p)
    if p.startswith("\\\\"):
        return "\\\\?\\UNC\\" + p.lstrip("\\")
    return "\\\\?\\" + p


def fs_path(path: str) -> str:
    """Retorna o path com prefixo long-path no Windows quando necessário."""
    if _needs_long_path_prefix(path):
        return _to_long_path(path)
    return path


def safe_file_name(name: str, default: str = "arquivo.zip", max_len: int = _WIN_MAX_FILENAME) -> str:
    """Sanitiza nome de arquivo (Windows) e limita tamanho para evitar erros."""
    invalid = '<>:"/\\|?*'
    cleaned = "".join(c for c in (name or "").strip() if c not in invalid and ord(c) >= 32)
    cleaned = cleaned.strip().rstrip(". ")
    if not cleaned:
        cleaned = default
    if _is_windows() and max_len and len(cleaned) > max_len:
        stem, ext = os.path.splitext(cleaned)
        hash_part = hashlib.md5(cleaned.encode("utf-8", errors="ignore")).hexdigest()[:8]
        keep = max_len - len(ext) - len(hash_part) - 1
        if keep < 1:
            keep = 1
        cleaned = f"{stem[:keep]}-{hash_part}{ext}"
    return cleaned


def build_safe_file_path(dir_path: str, filename: str) -> str:
    """Monta caminho final de arquivo ajustando nome quando o caminho fica longo."""
    name = safe_file_name(filename)
    if _is_windows() and dir_path:
        try:
            base_dir = _normalize_fs_path(dir_path)
            max_name_len = _WIN_LONG_PATH_THRESHOLD - len(base_dir) - 1
            if max_name_len < 40:
                max_name_len = 40
            name = safe_file_name(name, max_len=min(max_name_len, _WIN_MAX_FILENAME))
        except Exception:
            name = safe_file_name(name)
    return os.path.join(dir_path, name) if dir_path else name


def safe_makedirs(path: str) -> None:
    if not path:
        return
    os.makedirs(fs_path(path), exist_ok=True)


def path_exists(path: str) -> bool:
    if not path:
        return False
    try:
        return os.path.exists(fs_path(path))
    except Exception:
        return False


def path_isfile(path: str) -> bool:
    if not path:
        return False
    try:
        return os.path.isfile(fs_path(path))
    except Exception:
        return False


def save_download_with_fallback(download, destino: str, log_fn=None) -> str:
    """Salva download mesmo com caminho longo (usa fallback em temp + move)."""
    if not destino:
        raise ValueError("Destino de download vazio.")

    destino = str(destino)
    dir_dest = os.path.dirname(destino)
    if dir_dest:
        safe_makedirs(dir_dest)

    # Tenta salvar direto
    try:
        download.save_as(destino)
        return destino
    except Exception:
        pass

    # Tenta salvar direto com prefixo long-path
    if _needs_long_path_prefix(destino):
        try:
            download.save_as(fs_path(destino))
            return destino
        except Exception:
            pass

    # Fallback: salva em temp e move para o destino final
    tmp_dir = tempfile.mkdtemp(prefix="sefaz_dl_")
    tmp_path = os.path.join(tmp_dir, os.path.basename(destino) or "download.zip")
    try:
        download.save_as(tmp_path)
        if dir_dest:
            safe_makedirs(dir_dest)
        try:
            os.replace(tmp_path, fs_path(destino))
        except Exception:
            shutil.move(tmp_path, fs_path(destino))
        return destino
    except Exception as e:
        if callable(log_fn):
            log_fn(f"[ERRO] ❌ Falha ao salvar download em caminho longo: {e}")
        raise
    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass
        try:
            if tmp_dir and os.path.isdir(tmp_dir) and not os.listdir(tmp_dir):
                os.rmdir(tmp_dir)
        except Exception:
            pass


def nome_empresa_sem_ie(nome: str) -> str:
    """Remove sufixo ' (IE ...)' do nome, para uso em pastas."""
    n = (nome or "").strip()
    n = re.sub(r"\s*\(IE\s*[^)]+\)\s*$", "", n, flags=re.I).strip()
    return n or (nome or "Empresa").strip() or "Empresa"

# Estrutura padrão das pastas (controla quais níveis serão criados)
DEFAULT_FOLDER_STRUCTURE: Dict[str, Any] = {
    "usar_nome_empresa": True,
    "usar_pasta_cliente": False,
    # Pastas opcionais do cliente (NOVO: lista na ordem).
    # Ex.: ["DEPARTAMENTO FISCAL", "Notas"]
    "pastas_cliente": [],
    # Legacy: compatibilidade com configs antigas (string única).
    "nome_pasta_cliente": "",
    "separar_entrada_saida": False,
    "usar_ano": True,    # cria pasta do ANO (YYYY)
    "usar_mes": True,    # cria pasta do MÊS (MM)
    "case_sensitive": False,
}


def normalizar_estrutura_pastas(raw: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    """Garante que a configuração de estrutura tenha todas as chaves esperadas."""
    cfg = DEFAULT_FOLDER_STRUCTURE.copy()

    def _as_bool(val: Any, default: bool) -> bool:
        if isinstance(val, str):
            v = val.strip().lower()
            if v in {"1", "true", "sim", "yes", "y"}:
                return True
            if v in {"0", "false", "nao", "não", "no", "n"}:
                return False
        if val is None:
            return default
        return bool(val)

    if isinstance(raw, dict):
        cfg["usar_nome_empresa"] = _as_bool(raw.get("usar_nome_empresa"), cfg["usar_nome_empresa"])
        cfg["usar_pasta_cliente"] = _as_bool(raw.get("usar_pasta_cliente"), cfg["usar_pasta_cliente"])
        cfg["separar_entrada_saida"] = _as_bool(raw.get("separar_entrada_saida"), cfg["separar_entrada_saida"])
        cfg["usar_ano"] = _as_bool(raw.get("usar_ano"), cfg["usar_ano"])
        cfg["usar_mes"] = _as_bool(raw.get("usar_mes"), cfg["usar_mes"])
        # Opção removida: Windows trata nomes de pastas como case-insensitive; sempre reutilizamos pastas existentes.
        cfg["case_sensitive"] = False
        # Pastas opcionais do cliente (NOVO: lista)
        pastas = []

        raw_pastas = raw.get("pastas_cliente", None)
        if isinstance(raw_pastas, list):
            for it in raw_pastas:
                s = str(it or "").strip()
                if s:
                    pastas.append(s)
        elif isinstance(raw_pastas, str):
            # permite separar por ';' ou '|' ou quebra de linha
            for p in re.split(r"[;|\n]+", raw_pastas):
                s = (p or "").strip()
                if s:
                    pastas.append(s)

        # Compat: chave antiga 'nome_pasta_cliente' (string única).
        legacy = str(raw.get("nome_pasta_cliente", "") or "").strip()
        if legacy:
            # Se vier como caminho (colado), quebra em segmentos.
            parts = re.split(r"[\\/]+", legacy) if ("\\" in legacy or "/" in legacy) else [legacy]
            parts = [x.strip() for x in parts if x and x.strip()]
            # Inserimos no início preservando ordem.
            for p in reversed(parts):
                if not any((x or "").strip().lower() == p.lower() for x in pastas):
                    pastas.insert(0, p)

        cfg["pastas_cliente"] = pastas
        cfg["nome_pasta_cliente"] = pastas[0] if pastas else ""

    return cfg


def _split_pasta_mes_parts(pasta_mes: Optional[str]) -> Tuple[str, str]:
    """Extrai ano/mês (YYYY, MM) de uma string no formato mmYYYY."""
    if not pasta_mes:
        return "", ""
    txt = re.sub(r"\\D", "", str(pasta_mes))
    if len(txt) >= 6:
        mes = txt[:2]
        ano = txt[2:6]
        return ano, mes
    return "", ""


def _join_segment_case(base: str, segment: str, case_sensitive: bool) -> str:
    """Une um segmento ao caminho, reutilizando pasta existente quando case-insensitive."""
    seg = safe_folder_name(str(segment))
    if not seg:
        return base

    if not case_sensitive and base and os.path.isdir(fs_path(base)):
        try:
            for entry in os.listdir(fs_path(base)):
                if entry.lower() == seg.lower():
                    seg = entry
                    break
        except Exception:
            pass

    return os.path.join(base, seg) if base else seg


def montar_caminho_estruturado(
    base_root: str,
    estrutura_cfg: Optional[Dict[str, Any]],
    empresa_nome: str,
    operacao: Optional[str] = None,
    pasta_mes: Optional[str] = None,
    modelo: Optional[str] = None,
    criar: bool = False,
) -> str:
    """
    Monta o caminho final para salvar arquivos, respeitando as opções marcadas:
      - Nome da empresa
      - Pasta customizada do cliente
      - Separação Entrada/Saída
      - Ano e mês
      - Modelo (55/65)
    """
    cfg = normalizar_estrutura_pastas(estrutura_cfg)
    if not base_root:
        return ""

    path = base_root
    segmentos: List[str] = []

    if cfg["usar_nome_empresa"]:
        segmentos.append(empresa_nome)

    # Segmentos do robô vindos do dashboard, aplicados DENTRO da empresa:
    # ex.: segment_path "FISCAL/NFE-NFC" -> "<EMPRESA>\\NFE-NFC\\..."
    for seg in get_robot_segment_under_company_parts():
        if seg:
            segmentos.append(seg)
    if cfg["usar_pasta_cliente"]:
        for p in (cfg.get("pastas_cliente") or []):
            p = str(p or "").strip()
            if p:
                segmentos.append(p)

    if cfg["separar_entrada_saida"]:
        op = (operacao or "").lower()
        if op.startswith("e"):
            op_label = "Notas Fiscais de Entrada"
        elif op.startswith("s"):
            op_label = "Notas Fiscais de Saida"
        else:
            op_label = "Notas Fiscais"
        segmentos.append(op_label)

    ano, mes = _split_pasta_mes_parts(pasta_mes)
    if cfg["usar_ano"] and ano:
        segmentos.append(ano)
    if cfg["usar_mes"]:
        # Preferimos sempre o mês (MM) quando conseguimos extrair.
        # Assim evitamos criar pastas como '122025' (mmYYYY) quando a estrutura for apenas por mês.
        mes_label = ""
        if mes:
            mes_label = mes
        elif pasta_mes:
            mes_label = str(pasta_mes)
        if mes_label:
            segmentos.append(mes_label)

    if modelo:
        segmentos.append(str(modelo))

    for seg in segmentos:
        path = _join_segment_case(path, seg, cfg["case_sensitive"])

    if criar and path:
        safe_makedirs(path)

    return path

def _local_tag(tag: str) -> str:
    """Retorna o nome do tag sem prefixos/namespaces e em minúsculas."""
    return (tag or "").split("}")[-1].split(":")[-1].lower()

def detectar_modelo_xml_bytes(xml_bytes: bytes, root: Optional[ET.Element] = None) -> str:
    """Detecta o modelo (55/65) a partir do XML.

    Motivo: alguns XMLs vêm com namespaces/prefixos (ex.: <nfe:ide>), e a regex simples
    pode falhar. Primeiro tentamos parsear com ElementTree (robusto a namespaces),
    e só então usamos regex como fallback.
    """
    if not xml_bytes:
        return ""

    # Heurística: remove lixo antes do primeiro '<' (caso venha BOM/bytes extras)
    try:
        i = xml_bytes.find(b"<")
        if i > 0:
            xml_bytes = xml_bytes[i:]
    except Exception:
        pass

    parsed_root = root
    if parsed_root is None:
        try:
            parsed_root = ET.fromstring(xml_bytes)
        except Exception:
            parsed_root = None

    if parsed_root is not None:
        ide_node = None

        for el in parsed_root.iter():
            if _local_tag(getattr(el, "tag", "")) == "ide":
                ide_node = el
                break

        if ide_node is not None:
            for el in ide_node.iter():
                if _local_tag(getattr(el, "tag", "")) == "mod" and getattr(el, "text", None):
                    mod = (el.text or "").strip()
                    if mod in ("55", "65"):
                        return mod

        # Fallback dentro do XML: procura qualquer <mod>
        for el in parsed_root.iter():
            if _local_tag(getattr(el, "tag", "")) == "mod" and getattr(el, "text", None):
                mod = (el.text or "").strip()
                if mod in ("55", "65"):
                    return mod

    # 2) Fallback regex (tenta com/sem prefixo)
    try:
        txt = xml_bytes.decode("utf-8", errors="ignore")
    except Exception:
        try:
            txt = xml_bytes.decode("latin-1", errors="ignore")
        except Exception:
            return ""

    # <ide> e <mod> podem vir com prefixo (ex.: <nfe:ide> <nfe:mod>)
    m = re.search(
        r"<(?:[\w-]+:)?ide\b[^>]*>.*?<(?:[\w-]+:)?mod\b[^>]*>\s*(\d{2})\s*</(?:[\w-]+:)?mod>",
        txt,
        re.I | re.S,
    )
    if m and m.group(1) in ("55", "65"):
        return m.group(1)

    # Último fallback: qualquer <mod>
    m2 = re.search(r"<(?:[\w-]+:)?mod\b[^>]*>\s*(\d{2})\s*</(?:[\w-]+:)?mod>", txt, re.I | re.S)
    if m2 and m2.group(1) in ("55", "65"):
        return m2.group(1)

    return ""


def detectar_ie_emitente(xml_bytes: bytes, root: Optional[ET.Element] = None) -> str:
    """Retorna o valor do <emit><IE> ou <dest><IE> encontrado no XML, se disponível."""
    def _from_root(elem: ET.Element) -> str:
        for child in elem.iter():
            tag_local = _local_tag(getattr(child, "tag", ""))
            if tag_local in ("emit", "dest"):
                for sub in child:
                    if _local_tag(getattr(sub, "tag", "")) == "ie" and getattr(sub, "text", None):
                        return (sub.text or "").strip()
        return ""

    if root is not None:
        value = _from_root(root)
        if value:
            return value

    try:
        txt = xml_bytes.decode("utf-8", errors="ignore")
    except Exception:
        try:
            txt = xml_bytes.decode("latin-1", errors="ignore")
        except Exception:
            return ""

    match = re.search(
        r"<(?:[\w-]+:)?emit\b[^>]*>.*?<((?:[\w-]+:)?ie)\b[^>]*>\s*([^<]+?)\s*</\1>",
        txt,
        re.I | re.S,
    )
    if match and match.group(2):
        return (match.group(2) or "").strip()
    return ""


# --- Extrai XMLs de uma lista de ZIPs e salva por modelo em pastas 55/65 ---
def extrair_e_separar_zips_por_modelo(
    zip_paths,
    empresa_nome: str,
    pasta_mes: str,
    base55: str,
    base65: str,
    operacao: str = "",
    data_inicial: str = "",
    data_final: str = "",
    zipar: bool = True,
    remover_xmls: bool = True,
    log_fn=None,
    empresa_ie: str = "",
    empresas_map: Optional[Dict[str, Dict[str, str]]] = None,
    estrutura_pastas: Optional[Dict[str, Any]] = None,
    deletar_zips_gerados: bool = False,
):
    """
    Extrai XMLs de uma lista de ZIPs e separa por modelo (55/65).

    MODO OTIMIZADO (padrão): quando zipar=True e remover_xmls=True,
    NÃO cria milhares de arquivos .xml em disco.
    Em vez disso, escreve os XMLs diretamente dentro dos ZIPs finais:

      base55\\EMPRESA\\pasta_mes\\Modelo 55 - ...
      base65\\EMPRESA\\pasta_mes\\Modelo 65 - ...

    Mantém um modo de compatibilidade para outros cenários.
    """
    # Pasta da empresa: somente o nome (sem IE). Mantém IE apenas para nomes de arquivos.
    empresa_nome_clean = (empresa_nome or "").strip()
    # Se por algum motivo o nome já vier com "(IE ...)", remove para a pasta.
    empresa_nome_clean = re.sub(r"\s*\(IE\s*[^)]+\)\s*$", "", empresa_nome_clean, flags=re.I).strip()
    empresa_folder = safe_folder_name(empresa_nome_clean) or "Empresa"

    # Para nomear arquivos (ZIP), podemos incluir a IE (como era antes).
    empresa_label = empresa_nome_clean
    ie_digits = ie_somente_digitos(empresa_ie or "")
    if ie_digits:
        formatted_ie = ie_formatada(ie_digits) or ie_digits
        if empresa_label:
            empresa_label = f"{empresa_label} (IE {formatted_ie})"
        else:
            empresa_label = f"IE {formatted_ie}"
    if not empresa_label:
        empresa_label = "Empresa"
    empresa_label_safe = safe_folder_name(empresa_label) or empresa_folder

    if not (base55 or base65):
        if callable(log_fn):
            log_fn("[WARN] Pasta base não definida. Não foi possível separar por modelo (55/65).")
        return (0, 0)

    estrutura_cfg = normalizar_estrutura_pastas(estrutura_pastas)

    safe_makedirs(base55 or base65 or "")  # garante base existente
    count_55 = count_65 = count_unk = 0

    def _dest_dir_for(base_root: str, modelo: Optional[str]) -> str:
        return montar_caminho_estruturado(
            base_root,
            estrutura_cfg,
            empresa_folder,
            operacao=operacao,
            pasta_mes=pasta_mes,
            modelo=modelo,
            criar=False,
        )

    # -------------------------------------------------------------------------
    # MODO OTIMIZADO: zipar=True e remover_xmls=True
    # Não grava XML solto em pasta, já envia direto para o ZIP 55/65.
    # -------------------------------------------------------------------------
    if zipar and remover_xmls:
        dest_dir_55 = _dest_dir_for(base55, "55") if base55 else None
        dest_dir_65 = _dest_dir_for(base65, "65") if base65 else None

        data_ini_safe = (data_inicial or "").replace("/", ".")
        data_fim_safe = (data_final or "").replace("/", ".")

        zip55 = zip65 = None
        zip55_path = zip65_path = None
        names_55 = set()
        names_65 = set()


        def _get_zip_modelo(modelo: str):
            nonlocal zip55, zip65, zip55_path, zip65_path
            if modelo == "55" and dest_dir_55:
                if zip55 is None:
                    # Cria a pasta do modelo somente quando existir XML 55
                    safe_makedirs(dest_dir_55)
                    zip_name = f"Modelo 55 - {operacao} - {data_ini_safe} a {data_fim_safe} - {empresa_label_safe}.zip"
                    zip55_path = build_safe_file_path(dest_dir_55, zip_name)
                    zip55 = zipfile.ZipFile(fs_path(zip55_path), "w", zipfile.ZIP_DEFLATED)
                return zip55
            if modelo == "65" and dest_dir_65:
                if zip65 is None:
                    # Cria a pasta do modelo somente quando existir XML 65
                    safe_makedirs(dest_dir_65)
                    zip_name = f"Modelo 65 - {operacao} - {data_ini_safe} a {data_fim_safe} - {empresa_label_safe}.zip"
                    zip65_path = build_safe_file_path(dest_dir_65, zip_name)
                    zip65 = zipfile.ZipFile(fs_path(zip65_path), "w", zipfile.ZIP_DEFLATED)
                return zip65
            return None

        def _add_xml_to_zip(data: bytes, modelo: str, original_name: str):
            nonlocal count_55, count_65, count_unk
            if modelo not in ("55", "65"):
                count_unk += 1
                return

            zf = _get_zip_modelo(modelo)
            if zf is None:
                count_unk += 1
                return

            base_name = os.path.basename(original_name) or ""
            if not base_name.lower().endswith(".xml"):
                base_name = (base_name + ".xml") if base_name else f"nota_{hashlib.md5(data).hexdigest()}.xml"

            used = names_55 if modelo == "55" else names_65
            name = base_name
            idx = 1
            while name in used:
                stem, ext = os.path.splitext(base_name)
                name = f"{stem}({idx}){ext}"
                idx += 1
            used.add(name)

            zf.writestr(name, data)

            if modelo == "55":
                count_55 += 1
            else:
                count_65 += 1

        def _process_zipfile_stream(zf: zipfile.ZipFile):
            # percorre todos os itens; se for .xml manda direto pro ZIP 55/65,
            # se for .zip interno, abre recursivamente
            for name in zf.namelist():
                try:
                    lower = name.lower()
                    if lower.endswith(".xml"):
                        data = zf.read(name)
                        root = None
                        try:
                            root = ET.fromstring(data)
                        except Exception:
                            root = None
                        modelo = detectar_modelo_xml_bytes(data, root)
                        _add_xml_to_zip(data, modelo, name)
                    elif lower.endswith(".zip"):
                        inner_bytes = zf.read(name)
                        with zipfile.ZipFile(io.BytesIO(inner_bytes), "r") as inner:
                            _process_zipfile_stream(inner)
                    else:
                        continue
                except Exception as e:
                    if callable(log_fn):
                        log_fn(f"[WARN] Falha ao processar item '{name}': {e}")

        # Processa todos os arquivos baixados (ZIPs e possíveis XMLs avulsos)
        for z_path in zip_paths or []:
            if not z_path or not path_isfile(z_path):
                continue

            lower_path = str(z_path).lower()

            # Alguns casos do portal retornam um XML único em vez de um ZIP.
            # Também há situações em que o arquivo vem com nome .zip mas o conteúdo é XML.
            try:
                if lower_path.endswith(".xml"):
                    with open(fs_path(z_path), "rb") as f:
                        data = f.read()
                    root = None
                    try:
                        root = ET.fromstring(data)
                    except Exception:
                        root = None
                    modelo = detectar_modelo_xml_bytes(data, root)
                    _add_xml_to_zip(data, modelo, os.path.basename(z_path))
                    continue

                with zipfile.ZipFile(fs_path(z_path), "r") as z:
                    _process_zipfile_stream(z)

            except Exception as e:
                # Fallback: tenta tratar como XML (arquivo não era ZIP de verdade)
                handled_as_xml = False
                try:
                    with open(fs_path(z_path), "rb") as f:
                        head = f.read(256).lstrip()
                    if head.startswith(b"<?xml") or head.startswith(b"<"):
                        with open(fs_path(z_path), "rb") as f:
                            data = f.read()
                        root = None
                        try:
                            root = ET.fromstring(data)
                        except Exception:
                            root = None
                        modelo = detectar_modelo_xml_bytes(data, root)
                        _add_xml_to_zip(data, modelo, os.path.basename(z_path))
                        handled_as_xml = True
                except Exception:
                    handled_as_xml = False

                if not handled_as_xml and callable(log_fn):
                    log_fn(f"[WARN] Falha ao processar arquivo baixado '{z_path}': {e}")

        # Fecha os ZIPs finais
        for zf in (zip55, zip65):
            try:
                if zf is not None:
                    zf.close()
            except Exception:
                pass

        produced = (count_55 + count_65) > 0

        # Remove os ZIPs originais SOMENTE se conseguimos separar ao menos 1 XML.
        # Se não detectarmos nenhum 55/65, mantemos os ZIPs baixados para conferência
        # (evita "perder" o material em casos de XML com layout inesperado).
        if produced:
            for z_path in zip_paths or []:
                try:
                    if path_isfile(z_path):
                        os.remove(fs_path(z_path))
                except Exception:
                    pass
        else:
            if callable(log_fn):
                log_fn(
                    "[WARN] Nenhum XML modelo 55/65 foi identificado neste(s) ZIP(s). "
                    "Mantendo os ZIPs originais para conferência."
                )

        if callable(log_fn):
            msg = (
                f"[OK] XMLs separados por modelo (streaming): "
                f"55={count_55}, 65={count_65}, indefinidos/ignorados={count_unk}"
            )
            log_fn(msg + ".")
            if zip55_path:
                log_fn(f"[OK] ZIP modelo 55 gerado: {zip55_path}")
            if zip65_path:
                log_fn(f"[OK] ZIP modelo 65 gerado: {zip65_path}")

        if deletar_zips_gerados:
            for generated_zip in (zip55_path, zip65_path):
                try:
                    if generated_zip and path_isfile(generated_zip):
                        os.remove(fs_path(generated_zip))
                        if callable(log_fn):
                            log_fn(f"[INFO] ZIP removido após extração dos XMLs: {generated_zip}")
                except Exception as e:
                    if callable(log_fn):
                        log_fn(f"[WARN] Falha ao remover ZIP gerado '{generated_zip}': {e}")
        return (count_55, count_65)

    # -------------------------------------------------------------------------
    # MODO COMPATIBILIDADE (caso alguém use zipar=False ou remover_xmls=False)
    # Mantém o comportamento antigo: grava XMLs em pastas e depois (opcional) zipa.
    # -------------------------------------------------------------------------
    def _gravar_xml(data: bytes, modelo: str, original_name: str, emit_ie: str):
        nonlocal count_55, count_65, count_unk
        if modelo == "55" and base55:
            base = base55
            count_55 += 1
        elif modelo == "65" and base65:
            base = base65
            count_65 += 1
        else:
            count_unk += 1
            return  # ignora modelos desconhecidos

        dest_dir = _dest_dir_for(base, modelo)
        safe_makedirs(dest_dir)

        # Usa o nome original (normalmente a chave), garantindo extensão .xml
        base_name = os.path.basename(original_name) or ""
        if not base_name.lower().endswith(".xml"):
            if base_name:
                base_name = base_name + ".xml"
            else:
                # fallback se não tiver nome
                base_name = f"nota_{hashlib.md5(data).hexdigest()}.xml"

        dest_path = os.path.join(dest_dir, base_name)
        k = 1
        while path_exists(dest_path):
            stem, ext = os.path.splitext(base_name)
            dest_path = os.path.join(dest_dir, f"{stem}({k}){ext}")
            k += 1

        with open(fs_path(dest_path), "wb") as f:
            f.write(data)


    def _process_zipfile(zf: zipfile.ZipFile):
        for name in zf.namelist():
            try:
                lower = name.lower()
                if lower.endswith(".xml"):
                    data = zf.read(name)
                    root = None
                    try:
                        root = ET.fromstring(data)
                    except Exception:
                        root = None
                    modelo = detectar_modelo_xml_bytes(data, root)
                    emit_ie = detectar_ie_emitente(data, root)
                    _gravar_xml(data, modelo, name, emit_ie)
                elif lower.endswith(".zip"):
                    inner_bytes = zf.read(name)
                    with zipfile.ZipFile(io.BytesIO(inner_bytes), "r") as inner:
                        _process_zipfile(inner)
                else:
                    continue
            except Exception as e:
                if callable(log_fn):
                    log_fn(f"[WARN] Falha ao processar item '{name}': {e}")

    # Extrai todos os XMLs (modo compatibilidade)
    for z_path in zip_paths or []:
        try:
            with zipfile.ZipFile(fs_path(z_path), "r") as z:
                _process_zipfile(z)
        except Exception as e:
            if callable(log_fn):
                log_fn(f"[WARN] Falha ao abrir ZIP '{z_path}': {e}")

    if callable(log_fn):
        msg = f"[OK] XMLs separados por modelo: 55={count_55}, 65={count_65}, indefinidos/ignorados={count_unk}"
        log_fn(msg + ".")

    if zipar:
        def _zipar_pasta(base_root: str, label_modelo: str):
            if not base_root:
                return
            dest_dir = _dest_dir_for(base_root, label_modelo)
            if not os.path.isdir(fs_path(dest_dir)):
                return

            data_ini_safe = (data_inicial or "").replace("/", ".")
            data_fim_safe = (data_final or "").replace("/", ".")
            zip_name = f"Modelo {label_modelo} - {operacao} - {data_ini_safe} a {data_fim_safe} - {empresa_label_safe}.zip"
            zip_path = build_safe_file_path(dest_dir, zip_name)

            added = 0
            with zipfile.ZipFile(fs_path(zip_path), "w", zipfile.ZIP_DEFLATED) as zf:
                for root, _, files in os.walk(dest_dir):
                    for f in files:
                        if f.lower().endswith(".xml"):
                            full = os.path.join(root, f)
                            arc = os.path.relpath(full, dest_dir)
                            zf.write(fs_path(full), arc)
                            added += 1

            if callable(log_fn):
                log_fn(f"[OK] ZIP por modelo gerado: {zip_path} (arquivos: {added})")

            if remover_xmls and added > 0:
                for root, _, files in os.walk(dest_dir):
                    for f in files:
                        if f.lower().endswith(".xml"):
                            try:
                                os.remove(fs_path(os.path.join(root, f)))
                            except Exception:
                                pass

        if count_55 > 0:
            _zipar_pasta(base55, "55")
        if count_65 > 0:
            _zipar_pasta(base65, "65")

        return (count_55, count_65)

# =============================================================================
# Config (JSON: caminhos padrão + empresas)
# =============================================================================

def carregar_config(caminho_config: str = CONFIG_PATH) -> Dict:
    """
    Estrutura esperada:
    {
        "paths": {
            "base_empresas": "C:\\...\\Empresas",
            "relatorios_pdf": "C:\\...\\Relatorios",
            "estrutura_pastas": {...}   # opções de montagem das pastas
        },
        "empresas": {
            "109202171": { "display_name": "Empresa X" },
            ...
        },
        "selecoes_empresas": {
            "marcadas": ["109202171", ...],
            "diario": ["109202171", ...]
        }
    }
    Se existir formato antigo, tenta converter.
    """
    if os.path.exists(caminho_config):
        with open(caminho_config, "r", encoding="utf-8") as f:
            raw = json.load(f)
    else:
        raw = {}

    # Novo formato
    if isinstance(raw, dict) and "empresas" in raw and "paths" in raw:
        paths = raw.get("paths") or {}
        empresas_raw = raw.get("empresas") or {}
        empresas_norm = {}
        for ie, info in empresas_raw.items():
            ie_num = ie_somente_digitos(ie)
            if not isinstance(info, dict):
                info = {"display_name": str(info)}
            display = info.get("display_name", ie_formatada(ie_num))
            entry = {"display_name": display}
            cnpj_digits = cnpj_somente_digitos(info.get("cnpj", ""))
            if cnpj_digits:
                entry["cnpj"] = cnpj_digits
            login_cpf = cpf_somente_digitos(info.get("login_cpf", ""))
            if login_cpf:
                entry["login_cpf"] = login_cpf
            empresas_norm[ie_num] = entry

        # NOVO: normalizar bloco de seleções
        selecoes_raw = raw.get("selecoes_empresas") or {}
        marcadas_raw = selecoes_raw.get("marcadas") or []
        diario_raw = selecoes_raw.get("diario") or []

        marcadas_norm = sorted({
            ie_somente_digitos(ie) for ie in marcadas_raw
            if ie_somente_digitos(ie)
        })
        diario_norm = sorted({
            ie_somente_digitos(ie) for ie in diario_raw
            if ie_somente_digitos(ie)
        })

        selecoes_norm = {
            "marcadas": marcadas_norm,
            "diario": diario_norm,
        }

        estrutura_cfg = normalizar_estrutura_pastas(paths.get("estrutura_pastas"))

        return {
            "paths": {
                "base_empresas": paths.get("base_empresas", ""),
                "relatorios_pdf": paths.get("relatorios_pdf", ""),
                "estrutura_pastas": estrutura_cfg,                # NOVO
                "dashboard_busca_empresa_em_qualquer_pasta": bool(paths.get("dashboard_busca_empresa_em_qualquer_pasta", True)),
                "dashboard_mostrar_quantidade_real_xml": bool(paths.get("dashboard_mostrar_quantidade_real_xml", False)),
            },
            "empresas": empresas_norm,
            "selecoes_empresas": selecoes_norm,
            "enviar_supabase": raw.get("enviar_supabase", True),  # NOVO: flag para enviar ao Supabase
        }
        

    # Formato antigo: {ie: {"path": ..., "display_name": ...}}
    empresas_norm = {}
    if isinstance(raw, dict):
        for ie, info in raw.items():
            if isinstance(info, dict):
                display = info.get("display_name", ie)
                cnpj_digits = cnpj_somente_digitos(info.get("cnpj", ""))
                login_cpf = cpf_somente_digitos(info.get("login_cpf", ""))
            else:
                display = str(info)
                cnpj_digits = ""
                login_cpf = ""
            ie_num = ie_somente_digitos(ie)
            entry = {"display_name": display}
            if cnpj_digits:
                entry["cnpj"] = cnpj_digits
            if login_cpf:
                entry["login_cpf"] = login_cpf
            empresas_norm[ie_num] = entry

    return {
            "paths": {
                "base_empresas": "",
                "relatorios_pdf": "",
                "estrutura_pastas": normalizar_estrutura_pastas(None),  # NOVO
                "dashboard_busca_empresa_em_qualquer_pasta": True,
                "dashboard_mostrar_quantidade_real_xml": False,
            },
        "empresas": empresas_norm,
        "selecoes_empresas": {
            "marcadas": [],
            "diario": [],
        }
    }

def salvar_config(dados: Dict, caminho_config: str = CONFIG_PATH):
    os.makedirs(os.path.dirname(caminho_config), exist_ok=True)
    with open(caminho_config, "w", encoding="utf-8") as f:
        json.dump(dados, f, indent=4, ensure_ascii=False)


# =============================================================================
# Playwright – helpers (seletor, clique, preenchimento)
# =============================================================================

def resolve_selector(selector: str) -> str:
    s = selector.strip()
    if s.startswith("//") or s.startswith("/"):
        return f"xpath={s}"
    if s.startswith(".//") or s.startswith("("):
        return f"xpath={s}"
    if s.startswith("//*[@") or s.startswith("(/"):
        return f"xpath={s}"
    return s

def clicar_recaptcha_cloudflare(page, timeout_ms: int = 20_000, log_fn=None) -> bool:
    """
    Dispara cliques por coordenadas na região do CAPTCHA (sem buscar iframe/elemento)
    e salva uma captura com os pontos clicados enumerados.
    """
    deadline = time.time() + (timeout_ms / 1000.0)
    success = False
    last_error = None
    screenshot_path = None

    def _log(msg: str):
        if callable(log_fn):
            log_fn(msg)

    def _has_token() -> bool:
        try:
            return page.evaluate(
                """() => {
                    const tokens = Array.from(document.querySelectorAll(
                        'input[name=\"cf-turnstile-response\"], input[id^=\"cf-chl-widget\"][type=\"hidden\"], input[name=\"g-recaptcha-response\"]'
                    ));
                    return tokens.some(inp => (inp.value || '').trim().length > 0);
                }"""
            )
        except Exception:
            return False

    def _ensure_overlay():
        try:
            page.evaluate(
                """() => {
                    let ov = document.getElementById('cf-click-overlay');
                    if (!ov) {
                        ov = document.createElement('div');
                        ov.id = 'cf-click-overlay';
                        Object.assign(ov.style, {
                            position: 'fixed',
                            left: '0',
                            top: '0',
                            width: '100vw',
                            height: '100vh',
                            pointerEvents: 'none',
                            zIndex: '2147483647',
                            fontFamily: 'sans-serif'
                        });
                        document.body.appendChild(ov);
                    }
                }"""
            )
        except Exception:
            pass

    def _add_marker(idx: int, x: float, y: float):
        try:
            page.evaluate(
                """(info) => {
                    const { idx, x, y } = info;
                    let ov = document.getElementById('cf-click-overlay');
                    if (!ov) return;
                    const dot = document.createElement('div');
                    dot.className = 'cf-click-marker';
                    dot.textContent = String(idx);
                    Object.assign(dot.style, {
                        position: 'absolute',
                        left: `${x - 12}px`,
                        top: `${y - 12}px`,
                        width: '24px',
                        height: '24px',
                        borderRadius: '12px',
                        border: '2px solid red',
                        background: 'rgba(255,0,0,0.2)',
                        color: '#000',
                        fontSize: '12px',
                        fontWeight: '700',
                        display: 'grid',
                        placeItems: 'center',
                        boxShadow: '0 0 6px rgba(255,0,0,0.4)',
                        pointerEvents: 'none'
                    });
                    ov.appendChild(dot);
                }""",
                {"idx": idx, "x": x, "y": y},
            )
        except Exception:
            pass

    def _clear_overlay():
        try:
            page.evaluate("""() => { const ov = document.getElementById('cf-click-overlay'); if (ov) ov.remove(); }""")
        except Exception:
            pass

    def _save_screenshot():
        nonlocal screenshot_path
        try:
            screenshot_dir = os.path.join(DATA_DIR, "logs")
            os.makedirs(screenshot_dir, exist_ok=True)
            screenshot_path = os.path.join(screenshot_dir, "recaptcha_clicks.png")
            page.screenshot(path=screenshot_path, full_page=True)
        except Exception as e:
            _log(f"[WARN] Falha ao salvar screenshot dos cliques: {e}")

    def _get_recaptcha_bbox():
        """
        Tenta localizar o bloco/iframe do reCAPTCHA para usar como base das coordenadas.
        """
        selectors_locator = [
            XPATH_RECAPTCHA_BLOCO,
            "//form[@id='filtro']//div[@data-callback='pegarTokenSuccess']",
            "//div[@data-callback='pegarTokenSuccess']",
            "form#filtro div.cf-turnstile",
            "div.cf-turnstile",
            "//*[@id='filtro']/div[6]/div",
            "//*[@id='filtro']/div[6]",
            "//form[@id='filtro']//div[contains(@class,'recaptcha') or contains(@class,'captcha')]",
            "//form[@id='filtro']//iframe[contains(@src,'recaptcha') or contains(@src,'turnstile')]",
            "//iframe[contains(@title,'aptcha') or contains(@src,'recaptcha') or contains(@src,'turnstile')]",
        ]

        def _try_locator(frame, sel, timeout=4000):
            try:
                loc = frame.locator(resolve_selector(sel))
                if loc.count() == 0:
                    return None
                target = loc.first
                try:
                    target.wait_for(state="visible", timeout=timeout)
                except Exception:
                    pass
                try:
                    target.scroll_into_view_if_needed(timeout=timeout)
                except Exception:
                    pass
                box = target.bounding_box()
                if box and box.get("width") and box.get("height"):
                    return {**box, "source": sel, "frame_url": frame.url or "about:blank"}
            except Exception:
                return None
            return None

        frames = [page.main_frame] + [f for f in page.frames if f is not page.main_frame]
        for frame in frames:
            for sel in selectors_locator:
                box = _try_locator(frame, sel)
                if box:
                    if callable(log_fn):
                        _log(
                            f"[INFO] Bounding do captcha via locator ({box.get('source')}) no frame {box.get('frame_url')}: "
                            f"x={box.get('x'):.0f}, y={box.get('y'):.0f}, w={box.get('width'):.0f}, h={box.get('height'):.0f}"
                        )
                    return box

        # Fallback: tenta bounding do form e cria área reduzida dentro dele (em cada frame)
        for frame in frames:
            try:
                form_loc = frame.locator("form#filtro")
                if form_loc.count() == 0:
                    continue
                form = form_loc.first
                try:
                    form.wait_for(state="visible", timeout=2000)
                except Exception:
                    pass
                try:
                    form.scroll_into_view_if_needed(timeout=2000)
                except Exception:
                    pass
                fb = form.bounding_box()
                if fb and fb.get("width") and fb.get("height"):
                    fw, fh = fb["width"], fb["height"]
                    if fw > 40 and fh > 40:
                        w = min(360, fw * 0.8)
                        h = min(120, fh * 0.25)
                        x = fb["x"] + (fw - w) / 2
                        y = fb["y"] + fh * 0.45
                        if callable(log_fn):
                            _log(
                                f"[WARN] Usando fallback dentro do form#filtro (frame {frame.url}): "
                                f"x={x:.0f}, y={y:.0f}, w={w:.0f}, h={h:.0f}"
                            )
                        return {"x": x, "y": y, "width": w, "height": h, "source": "form#filtro"}
            except Exception:
                continue

        if callable(log_fn):
            _log("[WARN] Não foi possível localizar bounding do captcha em nenhum frame; os cliques serão pulados.")
        return None

    # Medidas do viewport
    try:
        vp = page.evaluate("() => ({ width: innerWidth || 1366, height: innerHeight || 768 })")
        vw = vp.get("width") or DEFAULT_VIEWPORT["width"]
        vh = vp.get("height") or DEFAULT_VIEWPORT["height"]
    except Exception:
        vw = DEFAULT_VIEWPORT["width"]
        vh = DEFAULT_VIEWPORT["height"]

    def _clamp_point(pt, bounds=None):
        x, y = pt
        if bounds:
            bx, by, bw, bh = bounds
            x = max(bx + 2, min(bx + bw - 2, float(x)))
            y = max(by + 2, min(by + bh - 2, float(y)))
        return (
            max(1, min(vw - 1, float(x))),
            max(1, min(vh - 1, float(y))),
        )

    # Define pontos candidatos tomando o bloco do reCAPTCHA como base (círculo + centro)
    recaptcha_box = _get_recaptcha_bbox()
    if not recaptcha_box:
        if callable(log_fn):
            _log("[WARN] Sem bounding do captcha; cliques não serão disparados para evitar pontos fora do alvo.")
        return False

    x0 = recaptcha_box.get("x", 0) or 0
    y0 = recaptcha_box.get("y", 0) or 0
    w = recaptcha_box.get("width", 0) or 0
    h = recaptcha_box.get("height", 0) or 0
    # Formação circular, ancorada próximo à borda esquerda do bloco
    bounds = (x0, y0, w, h)
    cx = x0 + max(8, min(w * 0.18, 26))  # centro levemente mais à direita
    cy = y0 + h / 2
    radius = max(4, min(w * 0.12, h * 0.30))
    angles = [270, 315, 0, 45, 90, 135, 180, 225]  # graus
    candidates = [
        _clamp_point((cx + radius * math.cos(math.radians(a)), cy + radius * math.sin(math.radians(a))), bounds)
        for a in angles
    ]
    candidates.append(_clamp_point((cx, cy), bounds))  # centro como último recurso
    if callable(log_fn):
        _log(
            "[INFO] Usando bounding box do bloco do reCAPTCHA "
            f"(data-callback='pegarTokenSuccess') em x={x0:.0f}, y={y0:.0f}, w={w:.0f}, h={h:.0f}."
        )

    _ensure_overlay()

    if callable(log_fn):
        _log("[INFO] 🔄 Aguardando cliques no reCAPTCHA...")

    attempt = 0
    while time.time() < deadline and attempt < len(candidates):
        attempt += 1
        x, y = candidates[attempt - 1]
        _add_marker(attempt, x, y)
        try:
            page.mouse.click(x, y, delay=60)
            try:
                page.wait_for_timeout(300)
            except Exception:
                pass
            if _has_token():
                success = True
                break
        except Exception as e:
            last_error = e
            continue

    _save_screenshot()
    _clear_overlay()

    # Considera sucesso após a sequência de cliques, dado viewport fixo
    success = True
    if callable(log_fn):
        _log("[INFO] 🧩 reCAPTCHA marcado via cliques coordenados.")
        if screenshot_path:
            _log("[INFO] screenshot: recaptcha_clicks.png salvo")
    return success


def find_frame_and_locator(page, selector: str, timeout: int = DEFAULT_TIMEOUT_MS):
    sel = resolve_selector(selector)
    deadline = datetime.now() + timedelta(milliseconds=timeout)
    ultimo_erro = None

    while datetime.now() < deadline:
        for frame in page.frames:
            try:
                loc = frame.locator(sel)
                if loc.count() > 0:
                    try:
                        loc.first.wait_for(timeout=1000)
                        return frame, loc.first
                    except PlaywrightTimeoutError:
                        continue
            except Exception as e:
                ultimo_erro = e
                continue
        sleep(0.25)

    msg = f"Timeout esperando seletor {selector} em qualquer frame."
    if ultimo_erro:
        msg += f" Último erro interno: {ultimo_erro}"
    raise TimeoutError(msg)

def resolve_image_path(name: str) -> Optional[str]:
    """
    Resolve o caminho completo de uma imagem considerando variações de caixa
    e caminhos relativos ao DATA_DIR.
    """
    if not name:
        return None
    if os.path.isabs(name) and os.path.exists(name):
        return name
    candidates = [
        os.path.join(DATA_DIR, 'imagens', name),
        os.path.join(DATA_DIR, 'Imagens', name),
        os.path.join(DATA_DIR, 'IMAGENS', name),
        os.path.join(DATA_DIR, 'image', name),
        os.path.join(DATA_DIR, 'Image', name),
        os.path.join(DATA_DIR, 'IMAGE', name),
        os.path.join(DATA_DIR, name),
        name,
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return None

def load_icon(name: str) -> QIcon:
    path = resolve_image_path(name)
    return QIcon(path) if path else QIcon()


class SidebarToggleIconButton(QPushButton):
    """Botão de menu (ícone) com offset fino no desenho do ícone.

    QSS nem sempre desloca o ícone quando o botão está "icon-only"; por isso
    desenhamos o ícone manualmente com um pequeno offset horizontal.
    """

    def __init__(self, icon: QIcon, offset_x: int = -4, offset_y: int = 0, parent=None):
        super().__init__(parent)
        self._icon = icon
        self._off_x = int(offset_x)
        self._off_y = int(offset_y)
        # Evita que o QPushButton desenhe o ícone padrão (vamos desenhar manualmente)
        super().setIcon(QIcon())
        super().setText('')

    def setMenuIcon(self, icon: QIcon):
        self._icon = icon
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        try:
            painter.setRenderHint(QPainter.SmoothPixmapTransform, True)
            opt = QStyleOptionButton()
            opt.initFrom(self)
            opt.rect = self.rect()
            # Não desenhar texto/ícone pelo estilo padrão
            opt.text = ''
            opt.icon = QIcon()
            opt.iconSize = QSize(0, 0)

            self.style().drawControl(QStyle.CE_PushButton, opt, painter, self)

            if self._icon is not None and not self._icon.isNull():
                isz = self.iconSize() if self.iconSize().isValid() else QSize(18, 18)
                pm = self._icon.pixmap(isz)
                # Centraliza dentro da área de conteúdo (evita "subir" ou desalinhamento)
                content = self.style().subElementRect(QStyle.SE_PushButtonContents, opt, self)
                cx = int(content.center().x() - isz.width() / 2) + self._off_x
                cy = int(content.center().y() - isz.height() / 2) + self._off_y
                painter.drawPixmap(cx, cy, pm)
        finally:
            painter.end()

def _any_visible(page, selectors: list[str]) -> bool:
    for sel in selectors:
        try:
            resolved = resolve_selector(sel)
            for frame in page.frames:
                try:
                    loc = frame.locator(resolved)
                    if loc.count() > 0 and loc.first.is_visible():
                        return True
                except Exception:
                    continue
        except Exception:
            continue
    return False

def is_netaccess_login_visible(page) -> bool:
    return _any_visible(page, ['//*[@id="NetAccess.Login"]', '//*[@id="NetAccess.Password"]'])

def ensure_login_ativo(page, log_fn=None, login_cpf: str | None = None):
    """
    Se detectar tela de login (NetAccess), reautentica e reabre tela de filtros.
    Retorna a própria page (mesmo objeto).
    """
    try:
        if is_netaccess_login_visible(page):
            if callable(log_fn):
                log_fn("[WARN] ⚠️ Sessão expirada (NetAccess). Reautenticando para continuar...")
            login_netaccess(page, cpf=login_cpf)
            abrir_menu_baixar_xml_nfe(page)
    except Exception as e:
        if callable(log_fn):
            log_fn(f"[WARN] ⚠️ Tentativa de reautenticação falhou: {e}")
    return page

def ensure_single_tab(page, log_fn=None):
    """Garante que exista apenas 1 aba aberta no mesmo context.

    Alguns fluxos do portal abrem popup/aba nova (Acesso Restrito). Se sobram
    2+ abas, o Playwright pode perder o "foco" e procurar seletores na aba errada.
    Retorna a page "mantida" (a que deve continuar sendo usada).
    """
    try:
        ctx = page.context
        pages = list(getattr(ctx, "pages", []) or [])
    except Exception:
        return page

    if len(pages) <= 1:
        try:
            page.bring_to_front()
        except Exception:
            pass
        return page

    keep = page

    # Preferir a aba onde a tela de filtros (cmpDataInicial) ou login NetAccess/Portal esta visivel.
    try:
        candidates = [
            XPATH_CAMPO_DATA_INICIAL,
            '//*[@id="NetAccess.Login"]',
            '//*[@id="username"]',
        ]
        for p in pages:
            try:
                if _any_visible(p, candidates):
                    keep = p
                    break
            except Exception:
                continue
    except Exception:
        pass

    closed = 0
    for p in pages:
        if p == keep:
            continue
        try:
            p.close()
            closed += 1
        except Exception:
            continue

    try:
        keep.bring_to_front()
    except Exception:
        pass

    if callable(log_fn) and closed:
        try:
            log_fn(f"[INFO] 🧭 Contexto normalizado: {closed} aba(s) extra(s) fechada(s).")
        except Exception:
            pass

    return keep

def wait_netaccess_or_filtros(page, timeout_ms: int = 8000, log_fn=None) -> str:
    """Aguarda aparecer OU a tela de login NetAccess OU a tela de filtros (cmpDataInicial).

    Retorna: 'netaccess', 'filtros' ou 'timeout'.
    """
    deadline = time.time() + (max(0, int(timeout_ms)) / 1000.0)
    while time.time() < deadline:
        try:
            if is_netaccess_login_visible(page):
                return "netaccess"
        except Exception:
            pass
        try:
            if _any_visible(page, [XPATH_CAMPO_DATA_INICIAL]):
                return "filtros"
        except Exception:
            pass
        sleep(0.2)

    # Ultima checagem rapida antes de desistir
    try:
        if is_netaccess_login_visible(page):
            return "netaccess"
    except Exception:
        pass
    try:
        if _any_visible(page, [XPATH_CAMPO_DATA_INICIAL]):
            return "filtros"
    except Exception:
        pass

    if callable(log_fn):
        try:
            log_fn("[WARN] ⚠️ Timeout aguardando NetAccess/login ou tela de filtros. Prosseguindo em modo best-effort.")
        except Exception:
            pass
    return "timeout"

def _parse_proxy_address(proxy_url: str) -> Optional[Tuple[str, int]]:
    """
    Normaliza uma URL de proxy e retorna host/porta válidos. Retorna None se não for possível.
    """
    if not proxy_url:
        return None
    url = proxy_url.strip()
    if not url:
        return None
    if not url.lower().startswith(("http://", "https://")):
        url = "http://" + url
    parsed = urlparse(url)
    host = parsed.hostname
    port = parsed.port
    if not host:
        return None
    if not port:
        port = 443 if parsed.scheme.lower() == "https" else 80
    return host, port


def _is_proxy_reachable(proxy_url: str, timeout: float = 1.0) -> bool:
    addr = _parse_proxy_address(proxy_url)
    if not addr:
        return False
    host, port = addr
    try:
        with socket.create_connection((host, port), timeout=timeout):
            return True
    except Exception:
        return False


def _wait_for_proxy_ready(proxy_url: str, timeout: float = 5.0, interval: float = 0.25) -> bool:
    """
    Aguarda o proxy ouvir na URL fornecida antes de continuar.
    """
    deadline = time.time() + timeout
    while time.time() < deadline:
        if _is_proxy_reachable(proxy_url, timeout=interval):
            return True
        time.sleep(interval)
    return False


def _pick_proxy_port(preferred: int = 8889) -> Tuple[int, bool]:
    """
    Tenta usar a porta preferida; se estiver ocupada, escolhe uma livre.
    Retorna (porta, usando_preferida).
    """
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            s.bind(("127.0.0.1", preferred))
            return preferred, True
    except OSError:
        pass

    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        s.bind(("127.0.0.1", 0))
        return s.getsockname()[1], False


def _is_user_admin() -> bool:
    try:
        return bool(ctypes.windll.shell32.IsUserAnAdmin())
    except Exception:
        return False



def _start_proxy_if_needed(log_fn=None) -> bool:
    """
    Inicia mitmdump local e seta SEFAZ_PROXY. Se nao conseguir subir e nao houver
    proxy ja definido, levanta erro para impedir execucao sem proxy.
    Controlado por SEFAZ_AUTO_PROXY (qualquer valor falsy desativa).
    """
    global SEFAZ_PROXY, _MITMDUMP_PROC, _PROXY_AUTO_STARTED

    def _proxy_log(msg: str):
        if callable(log_fn):
            try:
                log_fn(msg)
                return
            except Exception:
                pass
        try:
            print(msg)
        except Exception:
            pass

    def _install_ca_if_present():
        # Tenta instalar o CA no repositorio do usuario (Windows) se ainda nao estiver.
        if not os.path.exists(MITM_CA_PATH):
            return
        stores = ["Cert:\\CurrentUser\\Root"]
        if _is_user_admin():
            stores.append("Cert:\\LocalMachine\\Root")
        for store in stores:
            try:
                cmd = [
                    "powershell",
                    "-NoProfile",
                    "-ExecutionPolicy",
                    "Bypass",
                    "-Command",
                    f"Import-Certificate -FilePath '{MITM_CA_PATH}' -CertStoreLocation {store} | Out-Null"
                ]
                run_kwargs = dict(stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, check=False)
                if os.name == "nt" and hasattr(subprocess, "CREATE_NO_WINDOW"):
                    run_kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
                subprocess.run(cmd, **run_kwargs)
            except Exception:
                continue

    def _kill_existing_mitm():
        """Encerra instancias antigas do mitm para liberar a porta antes de subir uma nova."""
        if os.name != "nt":
            return
        for name in ("mitmdump.exe", "mitmproxy.exe"):
            try:
                run_kwargs = dict(stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, check=False)
                if hasattr(subprocess, "CREATE_NO_WINDOW"):
                    run_kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
                subprocess.run(["taskkill", "/IM", name, "/F"], **run_kwargs)
            except Exception:
                continue

    _PROXY_AUTO_STARTED = False
    if SEFAZ_PROXY:
        reachable = _is_proxy_reachable(SEFAZ_PROXY)
        if not AUTO_START_PROXY:
            if reachable:
                return True
            _proxy_log(
                f"[WARN] Proxy '{SEFAZ_PROXY}' detectado, mas nao respondeu. Executando sem proxy."
            )
            SEFAZ_PROXY = ""
            os.environ.pop("SEFAZ_PROXY", None)
            return False

        if reachable:
            _proxy_log(
                f"[WARN] Proxy '{SEFAZ_PROXY}' detectado, mas a automacao vai usar o proxy local."
            )
        else:
            _proxy_log(
                f"[WARN] Proxy '{SEFAZ_PROXY}' detectado, mas nao respondeu. Usando proxy local."
            )
        SEFAZ_PROXY = ""
        os.environ.pop("SEFAZ_PROXY", None)
    else:
        # Nenhum proxy definido por env: se já houver um mitm rodando na porta padrão, reaproveitamos
        if AUTO_START_PROXY:
            if _is_proxy_reachable(DEFAULT_PROXY_URL):
                SEFAZ_PROXY = DEFAULT_PROXY_URL
                os.environ["SEFAZ_PROXY"] = SEFAZ_PROXY
                _proxy_log(f"[INFO] Proxy já está ativo em {SEFAZ_PROXY}; não será reiniciado.")
                return True
    if not AUTO_START_PROXY:
        return False
    if _MITMDUMP_PROC and _MITMDUMP_PROC.poll() is None:
        _PROXY_AUTO_STARTED = True
        return True

    # Garante que nao ha outro mitm segurando a porta
    _kill_existing_mitm()
    _install_ca_if_present()

    bin_path = os.getenv("SEFAZ_MITMDUMP", "").strip() or MITMDUMP_PATH

    if not os.path.exists(bin_path):
        msg = (
            f"mitmdump.exe nao encontrado (procurei em {bin_path}). ",
            "Coloque o binario em data/proxy ou defina SEFAZ_MITMDUMP."
        )
        joined = "".join(msg)
        _proxy_log(f"[ERRO] {joined}")
        raise FileNotFoundError(joined)

    log_handle = None
    log_tmp_path = MITMDUMP_LOG_PATH
    error_happened = False
    try:
        port, using_preferred = _pick_proxy_port(8889)
        if not using_preferred:
            _proxy_log(f"[WARN] Porta 8889 ocupada. Usando porta livre {port} para o mitmdump.")

        try:
            os.makedirs(os.path.dirname(log_tmp_path), exist_ok=True)
            log_handle = open(log_tmp_path, "wb", buffering=0)
            stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            header = f"\\n--- mitmdump start {stamp} porta={port} bin={bin_path}\\n"
            log_handle.write(header.encode("utf-8", errors="ignore"))
        except Exception:
            log_handle = None

        args = [
            bin_path,
            "--listen-port", str(port),
            "--allow-hosts", "portal.sefaz.go.gov.br",
            "--set", "block_global=false",
        ]
        _MITMDUMP_PROC = subprocess.Popen(
            args,
            stdout=log_handle or subprocess.DEVNULL,
            stderr=log_handle or subprocess.DEVNULL,
            stdin=subprocess.DEVNULL,
            shell=False,
            creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, "CREATE_NO_WINDOW") else 0,
        )
        proxy_target = f"http://127.0.0.1:{port}"
        ready_timeout = 6.0
        if not _wait_for_proxy_ready(proxy_target, timeout=ready_timeout):
            _proxy_log(f"[WARN] Proxy mitmdump iniciado, mas nao respondeu em {ready_timeout:.1f}s.")
            try:
                _MITMDUMP_PROC.terminate()
            except Exception:
                pass
            _MITMDUMP_PROC = None
            error_happened = True

            tail_source = log_tmp_path if log_tmp_path and os.path.exists(log_tmp_path) else MITMDUMP_LOG_PATH
            tail = _tail_file(tail_source, max_lines=120)
            if log_tmp_path and os.path.exists(log_tmp_path):
                try:
                    os.replace(log_tmp_path, MITMDUMP_LOG_PATH)
                except Exception:
                    pass
            if tail:
                _proxy_log(f"[DEBUG] mitmdump.log (ultimas linhas):\n{tail}")
            else:
                _proxy_log(f"[DEBUG] Consulte {MITMDUMP_LOG_PATH} para detalhes.")
            return False
        _PROXY_AUTO_STARTED = True
        SEFAZ_PROXY = proxy_target
        os.environ["SEFAZ_PROXY"] = SEFAZ_PROXY
        _proxy_log(f"[INFO] Proxy mitmdump iniciado em {SEFAZ_PROXY} (PID: {_MITMDUMP_PROC.pid}).")
        if log_tmp_path and os.path.exists(log_tmp_path):
            try:
                os.remove(log_tmp_path)
            except Exception:
                pass
        return True
    except Exception as e:
        _proxy_log(f"[WARN] Falha ao iniciar proxy mitmdump: {e}")
        _MITMDUMP_PROC = None
        error_happened = True
        return False
    finally:
        if log_handle:
            try:
                log_handle.close()
            except Exception:
                pass
        if error_happened and log_tmp_path and os.path.exists(log_tmp_path):
            try:
                os.replace(log_tmp_path, MITMDUMP_LOG_PATH)
            except Exception:
                pass
        if not error_happened and log_tmp_path and os.path.exists(log_tmp_path):
            try:
                os.remove(log_tmp_path)
            except Exception:
                pass

def _stop_proxy_if_running(log_fn=None):
    """
    Encerra o mitmdump iniciado automaticamente (se houver) e limpa as variáveis.
    """
    global SEFAZ_PROXY, _MITMDUMP_PROC, _PROXY_AUTO_STARTED
    stopped = False

    if _MITMDUMP_PROC:
        try:
            _MITMDUMP_PROC.terminate()
            try:
                _MITMDUMP_PROC.wait(timeout=3)
            except subprocess.TimeoutExpired:
                try:
                    _MITMDUMP_PROC.kill()
                except Exception:
                    pass
        except Exception:
            pass
        finally:
            _MITMDUMP_PROC = None
            stopped = True

    if _PROXY_AUTO_STARTED:
        if SEFAZ_PROXY == DEFAULT_PROXY_URL:
            SEFAZ_PROXY = ""
        os.environ.pop("SEFAZ_PROXY", None)
        _PROXY_AUTO_STARTED = False
        stopped = True

    if stopped and callable(log_fn):
        log_fn("[INFO] Proxy mitmdump encerrado.")


def wait_and_fill(page, selector: str, value: str,
                  timeout: int = DEFAULT_TIMEOUT_MS,
                  clear_first: bool = False):
    frame, loc = find_frame_and_locator(page, selector, timeout)

    if clear_first:
        try:
            loc.click()
            loc.press("Control+A")
            loc.press("Delete")
        except Exception:
            try:
                loc.fill("")
            except Exception:
                pass

    loc.fill(value)


def wait_and_click(
    page,
    selector: str,
    timeout: int = DEFAULT_TIMEOUT_MS,
    retries: int = 3,
):
    """
    Tenta achar o elemento e clicar até 'retries' vezes antes de estourar erro.
    Isso ajuda em casos de pequeno delay/carregamento do portal.
    """
    last_exc = None
    for tentativa in range(1, retries + 1):
        try:
            frame, loc = find_frame_and_locator(page, selector, timeout)
            loc.click()
            return
        except Exception as e:
            last_exc = e
            # Pequeno intervalo entre tentativas
            sleep(1)

    # Se chegou aqui, todas as tentativas falharam -> propaga o último erro
    raise last_exc

def type_date(page, selector: str, date_str: str,
              timeout: int = DEFAULT_TIMEOUT_MS):
    """
    Preenche o campo de data digitando apenas números (o site insere as barras)
    e CONFERE se o valor realmente ficou no input. Se o portal “apagar” a data
    ao perder o foco, tentamos mais algumas vezes.
    """
    digits = "".join(ch for ch in (date_str or "") if ch.isdigit())
    frame, loc = find_frame_and_locator(page, selector, timeout)

    # Se vier vazio, apenas limpa o campo e sai
    if not digits:
        try:
            loc.click()
            loc.press("Control+A")
            loc.press("Delete")
        except Exception:
            try:
                loc.fill("")
            except Exception:
                pass
        return

    # Até 3 tentativas para garantir que o valor ficou no campo
    for _ in range(3):
        try:
            # Limpa o campo
            try:
                loc.click()
                loc.press("Control+A")
                loc.press("Delete")
            except Exception:
                try:
                    loc.fill("")
                except Exception:
                    pass

            # Digita só os dígitos (deixa o portal colocar as barras)
            try:
                loc.type(digits, delay=80)
            except Exception:
                # fallback: manda já formatado
                loc.fill(date_str)

            # Força “commit” da máscara / blur
            try:
                loc.press("Tab")
            except Exception:
                pass

            sleep(0.3)

            # Lê o valor atual do campo
            try:
                current_value = loc.input_value()
            except Exception:
                try:
                    current_value = loc.evaluate("el => el.value")
                except Exception:
                    current_value = ""

            if not current_value:
                # portal apagou / não aceitou, tenta de novo
                continue

            current_digits = "".join(ch for ch in current_value if ch.isdigit())
            if current_digits == digits:
                # Data ficou gravada corretamente
                return
        except Exception:
            # Qualquer erro, espera um pouquinho e tenta de novo
            sleep(0.2)

    # Se chegar aqui, seguimos assim mesmo.
    # Se o portal ainda reclamar (“A data final é obrigatória”), o código
    # abaixo trata isso como erro transitório (DATA_OBRIGATÓRIA) e re-tenta.

def selecionar_tipo_nota(page, tipo: str, timeout: int = DEFAULT_TIMEOUT_MS):
    """Entrada => input[value='0']; Saída => input[value='1']"""
    if tipo.lower().startswith("e"):
        valor = "0"   # Entrada
    else:
        valor = "1"   # Saída
    selector = f'//input[@value="{valor}"]'
    frame, loc = find_frame_and_locator(page, selector, timeout)
    try:
        loc.check()
    except Exception:
        loc.click()


def fechar_popup_certificado(page):
    """Fecha popups de certificado caso apareça."""
    try:
        possiveis_botoes = [
            "Cancelar",
            "Fechar",
            "Não utilizar certificado",
            "Acessar com usuário e senha",
        ]
        for texto in possiveis_botoes:
            try:
                btn = page.get_by_role("button", name=texto)
                if btn.is_visible():
                    btn.click()
                    return
            except Exception:
                continue

        possiveis_xpaths = [
            "//button[contains(@class,'close')]",
            "//button[@aria-label='Close']",
        ]
        for xp in possiveis_xpaths:
            sel = resolve_selector(xp)
            loc = page.locator(sel)
            if loc.count() > 0:
                el = loc.first
                if el.is_visible():
                    el.click()
                    return
    except Exception:
        pass

def _kill_automation_chrome_instances(log_fn=None) -> int:
    """Encerra apenas instancias do Chrome desta automacao (mesmo exe/perfil).

    Usado para garantir que ao "reiniciar navegador" nao fique um Chrome antigo vivo.
    """
    if os.name != "nt":
        return 0
    killed = 0
    def _log(msg: str):
        if callable(log_fn):
            try:
                log_fn(msg)
            except Exception:
                pass
    try:
        run_kwargs = dict(stdout=subprocess.PIPE, stderr=subprocess.DEVNULL, check=False, text=True)
        if hasattr(subprocess, "CREATE_NO_WINDOW"):
            run_kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
        out = subprocess.run(
            ["wmic", "process", "where", "name='chrome.exe'", "get", "ProcessId,ExecutablePath,CommandLine", "/format:csv"],
            **run_kwargs
        ).stdout or ""
        alvo = os.path.normcase(os.path.abspath(CHROME_EXE))
        alvo_l = alvo.lower()
        perfil = os.path.normcase(os.path.abspath(PROFILE_DIR))
        perfil_l = perfil.lower()
        pids = []
        for line in out.splitlines():
            parts = [p.strip() for p in line.split(",") if p.strip()]
            if len(parts) < 2:
                continue
            pid_str = parts[-1]
            try:
                pid = int(pid_str)
            except ValueError:
                continue
            cmdline = ",".join(parts[:-1]).lower()
            if alvo_l in cmdline or perfil_l in cmdline:
                pids.append(str(pid))
        if not pids:
            # Fallback: tenta matar por ExecutablePath exato (ainda limitado ao nosso Chrome portátil).
            try:
                escaped_alvo = alvo.replace("\\", "\\\\")
                q = f"ExecutablePath='{escaped_alvo}'"
                out2 = subprocess.run(
                    ["wmic", "process", "where", q, "get", "ProcessId", "/format:csv"],
                    **run_kwargs
                ).stdout or ""
                for line in out2.splitlines():
                    parts = [p.strip() for p in line.split(",") if p.strip()]
                    if not parts:
                        continue
                    pid_str = parts[-1]
                    if pid_str.isdigit():
                        pids.append(pid_str)
            except Exception:
                pass

        if pids:
            subprocess.run(
                ["taskkill", "/PID"] + pids + ["/F"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                check=False,
                creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, "CREATE_NO_WINDOW") else 0,
            )
            killed = len(pids)
    except Exception:
        pass
    if killed and callable(log_fn):
        _log(f"[INFO] 🧹 Chrome antigo encerrado: {killed} PID(s).")
    return killed


def _kill_automation_chrome_proc(log_fn=None) -> bool:
    """Tenta encerrar o processo principal do Chrome desta automação (se conhecido).

    Retorna True se tentou encerrar (processo existia), False se não havia processo.
    """
    global _AUTOMATION_CHROME_PROC
    proc = _AUTOMATION_CHROME_PROC
    if proc is None:
        return False

    def _log(msg: str):
        if callable(log_fn):
            try:
                log_fn(msg)
            except Exception:
                pass

    try:
        pid = int(getattr(proc, "pid", 0) or 0)
    except Exception:
        pid = 0

    try:
        alive = proc.poll() is None
    except Exception:
        alive = False

    if not alive:
        _AUTOMATION_CHROME_PROC = None
        return True

    if os.name == "nt" and pid:
        try:
            # /T mata a árvore de processos (subprocessos do Chrome).
            subprocess.run(
                ["taskkill", "/PID", str(pid), "/T", "/F"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                check=False,
                creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, "CREATE_NO_WINDOW") else 0,
            )
            try:
                proc.wait(timeout=3)
            except Exception:
                pass
            _log(f"[INFO] 🧹 Chrome principal encerrado (PID {pid}).")
        except Exception:
            pass
    else:
        try:
            proc.terminate()
        except Exception:
            pass

    _AUTOMATION_CHROME_PROC = None
    return True


def iniciar_chrome_com_cdp(max_espera_segundos: int = 30, retries: int = 1):
    """
    Inicia o Chrome portátil com porta de depuração remota (CDP) e
    tenta conectar o Playwright com retries até max_espera_segundos.

    Se não conseguir conectar, levanta um erro com mensagem clara.
    """
    # Conferir se o Chrome portátil existe
    if not os.path.exists(CHROME_EXE):
        raise FileNotFoundError(f"Chrome não encontrado em: {CHROME_EXE}")

    def _kill_existing_chrome():
        """Encerra apenas instancias do Chrome desta automacao (mesmo exe/perfil)."""
        if os.name != "nt":
            return
        try:
            run_kwargs = dict(stdout=subprocess.PIPE, stderr=subprocess.DEVNULL, check=False, text=True)
            if hasattr(subprocess, "CREATE_NO_WINDOW"):
                run_kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
            # Lista chrome.exe com linha de comando para filtrar só os nossos
            out = subprocess.run(
                ["wmic", "process", "where", "name='chrome.exe'", "get", "ProcessId,ExecutablePath,CommandLine", "/format:csv"],
                **run_kwargs
            ).stdout or ""
            alvo = os.path.normcase(os.path.abspath(CHROME_EXE))
            perfil = os.path.normcase(os.path.abspath(PROFILE_DIR))
            pids = []
            for line in out.splitlines():
                parts = [p.strip() for p in line.split(",") if p.strip()]
                if len(parts) < 2:
                    continue
                pid_str = parts[-1]
                try:
                    pid = int(pid_str)
                except ValueError:
                    continue
                cmdline = ",".join(parts[:-1]).lower()
                if alvo in cmdline or perfil.lower() in cmdline:
                    pids.append(str(pid))
            if pids:
                subprocess.run(["taskkill", "/PID"] + pids + ["/F"], stdout=subprocess.DEVNULL,
                               stderr=subprocess.DEVNULL, check=False,
                               creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, "CREATE_NO_WINDOW") else 0)
        except Exception:
            pass

    # Importante ao "reiniciar navegador" na troca de login:
    # `browser.close()` desconecta do CDP, mas o processo do Chrome pode continuar vivo.
    # Aqui garantimos que apenas as instancias desta automacao (mesmo EXE/perfil) sejam encerradas.
    _kill_automation_chrome_proc()
    _kill_automation_chrome_instances()

    def _ensure_password_leak_off():
        try:
            pref_dir = os.path.join(PROFILE_DIR, "Default")
            pref_path = os.path.join(pref_dir, "Preferences")
            os.makedirs(pref_dir, exist_ok=True)
            prefs = {}
            if os.path.exists(pref_path):
                with open(pref_path, "r", encoding="utf-8") as f:
                    prefs = json.load(f)
            profile_section = prefs.get("profile", {})
            profile_section["password_manager_leak_detection"] = False
            prefs["profile"] = profile_section
            with open(pref_path, "w", encoding="utf-8") as f:
                json.dump(prefs, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

        try:
            init_path = os.path.join(os.path.dirname(CHROME_EXE), "initial_preferences")
            if os.path.exists(init_path):
                with open(init_path, "r", encoding="utf-8") as f:
                    init_prefs = json.load(f)
                prof = init_prefs.get("profile", {})
                prof["password_manager_leak_detection"] = False
                init_prefs["profile"] = prof
                with open(init_path, "w", encoding="utf-8") as f:
                    json.dump(init_prefs, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    # Sobe proxy mitmdump automaticamente (se configurado) e exige proxy valido.
    _start_proxy_if_needed()
    if AUTO_START_PROXY:
        proxy_ativo = (SEFAZ_PROXY or "").strip()
        if not proxy_ativo:
            raise RuntimeError(
                "Proxy obrigatorio nao ficou disponivel. "
                "Chrome nao sera iniciado sem proxy para evitar tela de certificado."
            )
        if not _is_proxy_reachable(proxy_ativo, timeout=2.0):
            raise RuntimeError(
                f"Proxy '{proxy_ativo}' nao respondeu. "
                "Chrome nao sera iniciado sem proxy para evitar tela de certificado."
            )
    _ensure_password_leak_off()
    _ensure_password_leak_off()

    # Garante que a pasta de perfil é gravável
    os.makedirs(PROFILE_DIR, exist_ok=True)

    # Remove arquivo de porta antigo para evitar ler porta obsoleta
    devtools_port_file = os.path.join(PROFILE_DIR, "DevToolsActivePort")
    try:
        if os.path.exists(devtools_port_file):
            os.remove(devtools_port_file)
    except Exception:
        pass

    chrome_log_path = os.path.join(RUNTIME_DIR, "chrome_start.log")

    chrome_cmd = [
        CHROME_EXE,
        f"--remote-debugging-port={CDP_PORT}",
        f"--user-data-dir={PROFILE_DIR}",
        "--no-first-run",
        "--no-default-browser-check",
        "--ignore-certificate-errors",
        "--enable-logging=stderr",
        "--v=1",
    ]
    if SEFAZ_PROXY:
        chrome_cmd.append(f"--proxy-server={SEFAZ_PROXY}")

    global _AUTOMATION_CHROME_PROC
    chrome_proc = None
    try:
        # Inicia o Chrome em background e captura o log de inicialização
        with open(chrome_log_path, "w", encoding="utf-8", errors="replace") as chrome_log:
            chrome_log.write(
                "=== Chrome bootstrap ===\n"
                f"EXE: {CHROME_EXE}\n"
                f"DATA_DIR: {DATA_DIR}\n"
                f"PROFILE_DIR: {PROFILE_DIR}\n"
                f"CDP_PORT: {CDP_PORT}\n"
                f"Proxy: {SEFAZ_PROXY or 'sem proxy'}\n"
                f"CMD: {' '.join(chrome_cmd)}\n\n"
            )
            chrome_proc = subprocess.Popen(
                chrome_cmd,
                stdout=chrome_log,
                stderr=subprocess.STDOUT,
                cwd=os.path.dirname(CHROME_EXE),
                shell=False,
            )
            _AUTOMATION_CHROME_PROC = chrome_proc
    except Exception as e:
        raise RuntimeError(f"Falha ao iniciar o Chrome portátil. Erro: {e}") from e

    # Inicia o Playwright
    pw = sync_playwright().start()

    url_cdp = f"http://127.0.0.1:{CDP_PORT}"
    ultimo_erro = None
    inicio = time.time()

    def _porta_devtools_ativa() -> Optional[int]:
        try:
            with open(devtools_port_file, "r", encoding="utf-8") as f:
                primeira_linha = f.readline().strip()
                if primeira_linha.isdigit():
                    return int(primeira_linha)
        except Exception:
            pass
        return None

    # Tenta conectar até max_espera_segundos
    while time.time() - inicio < max_espera_segundos:
        # Se o Chrome morreu, não adianta continuar tentando
        if chrome_proc and chrome_proc.poll() is not None:
            rc = chrome_proc.returncode
            # Em alguns cenários (quando o app está elevado), o Chrome relança
            # o processo de forma "de-elevated" e o primeiro processo sai com rc=0.
            # Nessa situação, seguimos esperando a porta CDP abrir.
            if rc != 0:
                log_tail = _tail_file(chrome_log_path)
                pw.stop()
                raise RuntimeError(
                    f"Chrome finalizou com código {rc} antes de abrir a porta CDP.\n"
                    f"Log (últimas linhas) em {chrome_log_path}:\n{log_tail}"
                )
            chrome_proc = None  # rc==0: assume relançamento e continua aguardando

        porta_encontrada = _porta_devtools_ativa()
        if porta_encontrada:
            url_cdp = f"http://127.0.0.1:{porta_encontrada}"

        try:
            browser = pw.chromium.connect_over_cdp(url_cdp)

            if browser.contexts:
                context = browser.contexts[0]
            else:
                context = browser.new_context(ignore_https_errors=True)

            if context.pages:
                page = context.pages[0]
            else:
                page = context.new_page()

            page.set_default_timeout(DEFAULT_TIMEOUT_MS)
            try:
                page.set_viewport_size(DEFAULT_VIEWPORT)
            except Exception:
                pass
            return pw, browser, page

        except Exception as e:
            ultimo_erro = e
            # Espera 1 segundo e tenta de novo
            time.sleep(1)

    # Se chegou aqui, não conseguiu conectar
    if chrome_proc and chrome_proc.poll() is None:
        chrome_proc.terminate()
    pw.stop()
    log_tail = _tail_file(chrome_log_path)
    raise RuntimeError(
        f"Não foi possível conectar ao Chrome via CDP em {url_cdp} "
        f"após {max_espera_segundos}s. Último erro: {ultimo_erro}\n"
        f"Log (últimas linhas) em {chrome_log_path}:\n{log_tail}"
    )

def _set_current_login_cpf(cpf: str) -> None:
    global CURRENT_LOGIN_CPF
    CURRENT_LOGIN_CPF = cpf_somente_digitos(cpf or "")


def _get_current_login_cpf() -> str:
    return cpf_somente_digitos(CURRENT_LOGIN_CPF or "")


def login_portal(
    page,
    cpf: str | None = None,
    senha: str | None = None,
    log_fn: Callable[..., Any] | None = None,
):
    def _log(msg: str) -> None:
        if callable(log_fn):
            try:
                log_fn(msg)
            except Exception:
                pass

    last_err: Exception | None = None
    for attempt in range(1, SEFAZ_PORTAL_GOTO_RETRIES + 1):
        try:
            _log(
                f"[INFO] Portal SEFAZ: carregando login "
                f"(tentativa {attempt}/{SEFAZ_PORTAL_GOTO_RETRIES}, "
                f"timeout {SEFAZ_PORTAL_NAV_TIMEOUT_MS // 1000}s)..."
            )
            # "domcontentloaded" costuma bastar e evita travar no evento "load" do gov.
            page.goto(
                URL_PORTAL,
                wait_until="domcontentloaded",
                timeout=SEFAZ_PORTAL_NAV_TIMEOUT_MS,
            )
            try:
                page.wait_for_load_state(
                    "load",
                    timeout=min(90_000, SEFAZ_PORTAL_NAV_TIMEOUT_MS),
                )
            except Exception:
                _log(
                    "[WARN] ⚠️ Evento 'load' completo não ocorreu a tempo; "
                    "seguindo se a página de login estiver utilizável."
                )
            break
        except PlaywrightTimeoutError as e:
            last_err = e
            _log(f"[WARN] ⚠️ Timeout ao abrir portal SEFAZ (tentativa {attempt}): {e}")
        except Exception as e:
            last_err = e
            _log(f"[WARN] ⚠️ Falha ao abrir portal SEFAZ (tentativa {attempt}): {e}")
        if attempt >= SEFAZ_PORTAL_GOTO_RETRIES:
            raise RuntimeError(
                f"Falha no login do portal SEFAZ após {SEFAZ_PORTAL_GOTO_RETRIES} tentativas "
                f"de navegação: {last_err}"
            ) from last_err
        backoff = min(120, 8 * attempt)
        _log(f"[INFO] Aguardando {backoff}s antes de tentar abrir o portal novamente...")
        time.sleep(float(backoff))
        try:
            page.goto("about:blank", wait_until="domcontentloaded", timeout=15_000)
        except Exception:
            pass

    fechar_popup_certificado(page)

    if not cpf or not senha:
        cpf, senha = load_login_portal(preferred_cpf=cpf, strict_preferred=bool(cpf))
    if not cpf or not senha:
        # Mensagem mais especifica para diagnostico (arquivo/CPF solicitado).
        pref = cpf_somente_digitos(cpf or "")
        raise RuntimeError(
            "Login do portal não configurado. Cadastre em 'Login do Portal'. "
            + (f"(CPF solicitado: {cpf_formatado(pref)}) " if pref else "")
            + f"(arquivo: {LOGIN_PATH})"
        )

    wait_and_fill(page, '//*[@id="username"]', cpf)
    wait_and_fill(page, '//*[@id="password"]', senha)
    wait_and_click(page, '//*[@id="btnAuthenticate"]')
    try:
        page.wait_for_load_state(
            "networkidle",
            timeout=min(90_000, SEFAZ_PORTAL_NAV_TIMEOUT_MS),
        )
    except Exception:
        try:
            page.wait_for_load_state("load", timeout=45_000)
        except Exception:
            _log(
                "[WARN] ⚠️ Pós-login: networkidle/load não confirmados; "
                "continuando (portal pode estar lento)."
            )
    fechar_popup_certificado(page)
    _set_current_login_cpf(cpf)

def portal_logout_via_pta(page, log_fn=None, timeout_ms: int = 15_000) -> bool:
    """Faz logout do portal pelo menu "PTA" -> "Log Out" -> "Voltar para a pagina de login".

    Esse fluxo é necessário para trocar de CPF/senha em algumas sessões em que o
    portal "gruda" no login atual e não aceita apenas navegar para a tela de login.
    Retorna True se conseguiu chegar na tela de login (campo username visível).
    """
    def _log(msg: str):
        if callable(log_fn):
            try:
                log_fn(msg)
            except Exception:
                pass

    try:
        page = ensure_single_tab(page, log_fn=log_fn)
    except Exception:
        pass

    # Se já estiver na tela de login do portal, não faz nada.
    try:
        if _any_visible(page, ['//*[@id="username"]', '//*[@id="password"]', '//*[@id="btnAuthenticate"]']):
            return True
    except Exception:
        pass

    deadline = time.time() + (max(1, int(timeout_ms)) / 1000.0)

    # 1) Clicar em "PTA"
    pta_selectors = [
        'text="PTA"',
        "text=PTA",
        'xpath=//*[normalize-space()="PTA"]',
        'xpath=//a[normalize-space()="PTA"]',
        'xpath=//button[normalize-space()="PTA"]',
    ]
    clicked_pta = False
    for sel in pta_selectors:
        try:
            wait_and_click(page, sel, timeout=3_000, retries=2)
            clicked_pta = True
            break
        except Exception:
            continue
    if not clicked_pta:
        _log("[WARN] ⚠️ Não encontrei o menu 'PTA' para fazer logout (vou tentar prosseguir).")
        return False

    # 2) Clicar em "Log Out"
    logout_selectors = [
        'text="Log Out"',
        "text=Log Out",
        'text="Logout"',
        "text=Logout",
        'xpath=//*[contains(normalize-space(), "Log Out")]',
        'xpath=//*[contains(normalize-space(), "Logout")]',
    ]
    clicked_logout = False
    for sel in logout_selectors:
        try:
            wait_and_click(page, sel, timeout=5_000, retries=2)
            clicked_logout = True
            break
        except Exception:
            continue
    if not clicked_logout:
        _log("[WARN] ⚠️ Clique em 'Log Out' não foi possível (vou tentar prosseguir).")
        return False

    # 3) Esperar mensagem de desconexão (não obrigatório)
    msg_selectors = [
        "text=desconectado com sucesso",
        "text=Logout, você foi desconectado com sucesso",
        "text=Logout, voce foi desconectado com sucesso",
        'xpath=//*[contains(translate(normalize-space(), "ÁÀÃÂÉÊÍÓÔÕÚÇáàãâéêíóôõúç", "AAAAEEIOOOUCaaaaeeiooouc"), "DESCON") and contains(translate(normalize-space(), "ÁÀÃÂÉÊÍÓÔÕÚÇáàãâéêíóôõúç", "AAAAEEIOOOUCaaaaeeiooouc"), "SUCESSO")]',
    ]
    try:
        while time.time() < deadline:
            if _any_visible(page, msg_selectors):
                break
            sleep(0.2)
    except Exception:
        pass

    # 4) Clicar em "Voltar para a página de login"
    back_selectors = [
        'text="Voltar para a página de login"',
        'text="Voltar para a pagina de login"',
        "text=Voltar para a página de login",
        "text=Voltar para a pagina de login",
        'xpath=//*[contains(normalize-space(), "Voltar para a") and contains(normalize-space(), "login")]',
    ]
    clicked_back = False
    for sel in back_selectors:
        try:
            wait_and_click(page, sel, timeout=8_000, retries=2)
            clicked_back = True
            break
        except Exception:
            continue
    if not clicked_back:
        _log("[WARN] ⚠️ Não encontrei o botão 'Voltar para a página de login' (vou tentar prosseguir).")
        return False

    # 5) Confirmar que voltou para a tela de login
    try:
        while time.time() < deadline:
            if _any_visible(page, ['//*[@id="username"]', '//*[@id="password"]', '//*[@id="btnAuthenticate"]']):
                return True
            sleep(0.2)
    except Exception:
        pass

    return False

def portal_sair_duplo_fechar_aba(page, log_fn=None, timeout_ms: int = 12_000):
    """Fluxo adicional antes do logout via PTA.

    Em alguns cenários, para trocar de login é preciso:
      1) clicar em "Sair" (duas vezes)
      2) a aba atual se encerra (ou deve ser fechada)
      3) na aba restante aparece "PTA" e então seguimos com o logout padrão

    Retorna a page ativa após o fechamento/troca de aba (best-effort).
    """
    def _log(msg: str):
        if callable(log_fn):
            try:
                log_fn(msg)
            except Exception:
                pass

    try:
        page = ensure_single_tab(page, log_fn=log_fn)
    except Exception:
        pass

    deadline = time.time() + (max(1, int(timeout_ms)) / 1000.0)

    sair_selectors = [
        'text="Sair"',
        "text=Sair",
        'xpath=//*[normalize-space()="Sair"]',
        'xpath=//a[normalize-space()="Sair"]',
        'xpath=//button[normalize-space()="Sair"]',
    ]

    def _click_sair_once() -> bool:
        for sel in sair_selectors:
            try:
                wait_and_click(page, sel, timeout=4_000, retries=2)
                return True
            except Exception:
                continue
        return False

    # Clica 2x em "Sair" (como você descreveu)
    ok1 = _click_sair_once()
    try:
        page.wait_for_timeout(400)
    except Exception:
        sleep(0.4)
    ok2 = _click_sair_once()

    if not (ok1 or ok2):
        _log("[WARN] ⚠️ Não encontrei o botão/link 'Sair' (vou tentar seguir com o logout via PTA).")
        return page

    # Tenta fechar a aba atual (algumas vezes ela já se fecha sozinha)
    try:
        ctx = page.context
        pages_before = list(getattr(ctx, "pages", []) or [])
    except Exception:
        pages_before = []
        ctx = None

    # Aguarda um pouco para o portal encerrar/alterar a aba
    while time.time() < deadline:
        try:
            if ctx is not None:
                pages_now = list(getattr(ctx, "pages", []) or [])
                if len(pages_now) < len(pages_before):
                    break
        except Exception:
            break
        sleep(0.2)

    # Fecha manualmente se ainda estiver aberta
    try:
        page.close()
    except Exception:
        pass

    # Seleciona uma aba remanescente (se houver)
    try:
        if ctx is not None:
            pages = [p for p in (getattr(ctx, "pages", []) or []) if p is not None]
        else:
            pages = []
    except Exception:
        pages = []

    if pages:
        new_page = pages[0]
        try:
            new_page.set_default_timeout(DEFAULT_TIMEOUT_MS)
        except Exception:
            pass
        try:
            new_page.set_viewport_size(DEFAULT_VIEWPORT)
        except Exception:
            pass
        try:
            new_page.bring_to_front()
        except Exception:
            pass
        _log("[INFO] 🧭 Aba atual fechada. Continuando na aba restante para iniciar logout via PTA.")
        return new_page

    # Se não achou outra aba, devolve a mesma referência (best-effort)
    return page

def ir_para_acesso_restrito(page, log_fn=None):
    xpath_acesso_restrito = "/html/body/div/div/div[3]/section/div[1]/div/div/div[2]/a[1]"
    # Se ja estamos na area (tela de filtros ou NetAccess), nao precisa clicar.
    try:
        if _any_visible(page, [XPATH_CAMPO_DATA_INICIAL, '//*[@id="NetAccess.Login"]']):
            return ensure_single_tab(page, log_fn=log_fn)
    except Exception:
        pass

    def _log(msg: str):
        if callable(log_fn):
            try:
                log_fn(msg)
            except Exception:
                pass

    selectors = [
        xpath_acesso_restrito,
        'text="Acesso Restrito"',
        "text=Acesso Restrito",
        'xpath=//a[contains(normalize-space(), "Acesso Restrito")]',
        'xpath=//button[contains(normalize-space(), "Acesso Restrito")]',
    ]

    for sel in selectors:
        try:
            # tenta popup
            with page.expect_popup(timeout=5_000) as popup_info:
                wait_and_click(page, sel, timeout=8_000, retries=2)
            new_page = popup_info.value
            new_page.set_default_timeout(DEFAULT_TIMEOUT_MS)
            try:
                new_page.set_viewport_size(DEFAULT_VIEWPORT)
            except Exception:
                pass
            return ensure_single_tab(new_page, log_fn=log_fn)
        except Exception:
            pass

        try:
            wait_and_click(page, sel, timeout=8_000, retries=2)
            try:
                page.wait_for_load_state("networkidle")
            except Exception:
                pass
            return ensure_single_tab(page, log_fn=log_fn)
        except Exception as e:
            _log(f"[WARN] ⚠️ Não foi possível abrir Acesso Restrito usando seletor '{sel}': {e}")
            continue

    # Fallback final: segue com a page atual, sem derrubar o Worker.
    return ensure_single_tab(page, log_fn=log_fn)


def abrir_menu_baixar_xml_nfe(page):
    """
    Garante a tela de 'Baixar XML NFE':
      - Se o campo cmpDataInicial já estiver visível, não clica em nada.
      - Caso contrário, clica no menu e item (timeouts curtos).
    """
    sel_data = resolve_selector(XPATH_CAMPO_DATA_INICIAL)
    for frame in page.frames:
        try:
            el = frame.query_selector(sel_data)
            if el and el.is_visible():
                return
        except Exception:
            continue

    try:
        wait_and_click(page, XPATH_MENU, timeout=5_000)
    except Exception:
        pass
    try:
        wait_and_click(page, XPATH_ITEM, timeout=5_000)
    except Exception:
        pass
    sleep(1)


def login_netaccess(page, cpf: str | None = None, senha: str | None = None):
    if not cpf or not senha:
        cpf_pref = cpf or _get_current_login_cpf()
        cpf, senha = load_login_portal(preferred_cpf=cpf_pref, strict_preferred=bool(cpf_pref))
    if not cpf or not senha:
        raise RuntimeError(
            "Login do portal não configurado. Cadastre em 'Login do Portal'."
        )
    wait_and_fill(page, '//*[@id="NetAccess.Login"]', cpf)
    wait_and_fill(page, '//*[@id="NetAccess.Password"]', senha)
    wait_and_click(page, '//*[@id="btnAuthenticate"]')
    page.wait_for_load_state("networkidle")
    _set_current_login_cpf(cpf)

def preencher_filtros_pesquisa(page, data_inicial: str, data_final: str,
                               tipo: str, ie: str):
    # Datas: limpa e digita somente dígitos (site insere barras)
    type_date(page, XPATH_CAMPO_DATA_INICIAL, data_inicial)
    type_date(page, XPATH_CAMPO_DATA_FINAL, data_final)

    # IE -> SEM máscara (apenas números)
    ie_limpa = ie_somente_digitos(ie)
    wait_and_fill(page, XPATH_CAMPO_IE, ie_limpa, clear_first=True)

    # Tipo de nota
    selecionar_tipo_nota(page, tipo)
# =============================================================================

def esperar_alertas(
    page,
    timeout_ms: int = 15000,
    should_stop: Optional[Callable[[], bool]] = None,
) -> Tuple[str, Optional[str]]:
    """
    Retorna:
      ('textual', texto) – div.alert[role='alert'] visível
      ('geral', texto)   – botão de fechar message-container visível
      ('nenhum', None)   – nada apareceu dentro do tempo
      ('INTERRUPCAO_USUARIO', None) – parada manual solicitada
    """
    deadline = datetime.now() + timedelta(milliseconds=timeout_ms)
    sel_alert_btn = resolve_selector(XPATH_ALERT_BUTTON)
    sel_sem_permissao = resolve_selector(XPATH_LABEL_SEM_PERMISSAO)

    while datetime.now() < deadline:
        # Parada manual durante a espera
        if callable(should_stop) and should_stop():
            return "INTERRUPCAO_USUARIO", None

        # Página de "Você não tem permissão..." visível em qualquer frame
        try:
            el = page.query_selector(sel_sem_permissao)
            if el and el.is_visible():
                return "SEM_PERMISSAO", None
        except Exception:
            pass
        for frame in page.frames:
            try:
                el = frame.query_selector(sel_sem_permissao)
                if el and el.is_visible():
                    return "SEM_PERMISSAO", None
            except Exception:
                continue

        # alerta textual
        try:
            alert = page.query_selector("div.alert[role='alert']")
            if alert and alert.is_visible():
                texto = (alert.inner_text() or "").strip()
                return "textual", texto
        except Exception:
            pass

        for frame in page.frames:
            try:
                alert = frame.query_selector("div.alert[role='alert']")
                if alert and alert.is_visible():
                    texto = (alert.inner_text() or "").strip()
                    return "textual", texto
            except Exception:
                continue

        # alerta geral (botão)
        for frame in page.frames:
            try:
                btn = frame.query_selector(sel_alert_btn)
                if btn and btn.is_visible():
                    try:
                        texto = btn.evaluate("node => node.parentElement.innerText") or ""
                        texto = texto.strip()
                    except Exception:
                        texto = ""
                    return "geral", texto
            except Exception:
                continue

        sleep(0.25)

    return "nenhum", None

def clicar_alert_button(page):
    sel = resolve_selector(XPATH_ALERT_BUTTON)
    for frame in page.frames:
        try:
            btn = frame.query_selector(sel)
            if btn and btn.is_visible():
                btn.click()
                return True
        except Exception:
            continue
    return False


def detectar_erro_503(page) -> bool:
    """
    Detecta a página de erro do portal (Service Unavailable / HTTP 503),
    como no caso do SEFAZ retornar uma página HTML simples de indisponibilidade.
    """
    try:
        if page.query_selector("text=Service Unavailable") or page.query_selector("text=HTTP Error 503"):
            return True
        if page.query_selector("text=Erro 503") or page.query_selector("text=HTTP 503"):
            return True
    except Exception:
        pass

    try:
        html = page.content() or ""
        low = html.lower()
        return ("service unavailable" in low and "503" in low) or ("http error 503" in low)
    except Exception:
        return False


def detectar_erro_sem_dados_enviados_portal(page) -> str:
    """Detecta erros transitórios do iframe do portal após a pesquisa."""
    padroes = (
        (
            "nenhum dado foi enviado por",
            "nfeweb.sefaz.go.gov.br",
            "Erro 'Nenhum dado foi enviado' detectado",
        ),
        (
            "sub-frame-error",
            "a conexao foi redefinida.",
            "Erro 'A conexao foi redefinida' detectado",
        ),
    )

    def _texto_doc(frame) -> str:
        try:
            txt = frame.evaluate(
                """() => {
                    const body = document.body?.innerText || "";
                    const root = document.documentElement?.innerText || "";
                    return `${body}\n${root}`.trim();
                }"""
            )
            return normalizar_txt(txt or "")
        except Exception:
            return ""

    try:
        frames = [page] + list(getattr(page, "frames", []) or [])
    except Exception:
        frames = [page]

    for frame in frames:
        texto = _texto_doc(frame)
        if not texto:
            continue
        for marcador_1, marcador_2, descricao in padroes:
            if marcador_1 in texto and marcador_2 in texto:
                try:
                    nome_frame = getattr(frame, "name", "") or "principal"
                except Exception:
                    nome_frame = "principal"
                return f"{descricao} no frame {nome_frame}."

    return ""


def classificar_alerta(texto: Optional[str]) -> str:
    t = (texto or "").lower()

    # IE
    if "inscrição estadual inválida" in t:
        return "IE_INVÁLIDA"
    if "inscrição estadual é obrigatória" in t:
        return "IE_OBRIGATÓRIA"

    # Datas obrigatórias (ex.: "A data inicial é obrigatória")
    if (
        "data inicial é obrigatória" in t
        or "data final é obrigatória" in t
        or "data inicial obrigatória" in t
        or "data final obrigatória" in t
    ):
        return "DATA_OBRIGATÓRIA"

    # Datas inválidas
    if "data inválida" in t or ("data final" in t and "maior" in t):
        return "DATA_INVÁLIDA"

    # Captcha
    if "captcha" in t or "recaptcha" in t:
        return "CAPTCHA"

    # Sem movimento
    if "nenhuma nota encontrada" in t or "sem resultado" in t:
        return "SEM_RESULTADO"

    return "OUTRO"

def limpar_texto_alerta(texto: Optional[str]) -> str:
    """
    Compacta o texto dos alertas do portal em uma única linha, removendo
    ruídos como '×' e 'Fechar' e linhas vazias.

    Exemplo:
        \"\"\"
        ×
        Fechar
        Sem Resultados!
        \"\"\"
    vira: "Sem Resultados!"
    """
    if not texto:
        return ""

    # quebra em linhas e remove lixos
    linhas = []
    for linha in str(texto).splitlines():
        l = linha.strip()
        if not l:
            continue
        if l in {"×", "Fechar", "Fechar.", "Fechar "}:
            continue
        linhas.append(l)

    # remove duplicadas mantendo ordem
    compactadas = []
    for l in linhas:
        if l not in compactadas:
            compactadas.append(l)

    return " | ".join(compactadas)

def pesquisar_intervalo(
    page,
    ie: str,
    data_inicial: str,
    data_final: str,
    tipo: str,
    max_tentativas: int = MAX_TENTATIVAS_EMPRESA,
    log_fn: Optional[Callable[[str], None]] = None,
    should_stop: Optional[Callable[[], bool]] = None,
):
    """
    Realiza a pesquisa (Entrada/Saída) para IE + intervalo.
    Retorna dict com sucesso/alertas/motivo/tentativas.
    Agora com:
      - logs detalhados e mais compactos
      - checagem de parada manual (should_stop)
      - re-tentativa também para DATA_OBRIGATÓRIA, CAPTCHA e SEM_RESULTADO
      - PRIORIDADE de classificação final: se em alguma tentativa houve SEM_RESULTADO,
        a falha final reporta SEM_RESULTADO (sem movimento), mesmo que a última
        tenha sido CAPTCHA/outro alerta transitório.
    """
    # NOVO: rastrear se em ALGUMA tentativa apareceu SEM_RESULTADO (sem movimento)
    teve_sem_resultado = False
    texto_sem_resultado = ""
    teve_sem_permissao = False
    teve_503 = False
    ultimo_alert_class = ""
    ultimo_alert_text = ""
    ultimo_motivo = ""

    # Reautentica se perdeu login
    page = ensure_login_ativo(page, log_fn=log_fn, login_cpf=_get_current_login_cpf())

    for tentativa in range(1, max_tentativas + 1):
        # Se o portal devolver página de indisponibilidade (503), recarrega e tenta de novo
        if detectar_erro_503(page):
            teve_503 = True
            ultimo_alert_class = "SERVICE_503"
            ultimo_alert_text = "Portal indisponível (HTTP 503)"
            ultimo_motivo = "Portal indisponível (HTTP 503) durante a pesquisa."
            if callable(log_fn):
                log_fn(
                    f"[WARN] ⚠️ Portal indisponível (Service Unavailable / HTTP 503). "
                    f"Tentativa {tentativa}/{max_tentativas}. Recarregando e tentando novamente..."
                )
            try:
                page.reload(wait_until="networkidle", timeout=60_000)
            except Exception:
                try:
                    page.reload()
                except Exception:
                    pass
            sleep(3)
            continue

        # Parada solicitada antes de começar a tentativa
        if callable(should_stop) and should_stop():
            if callable(log_fn):
                log_fn(
                    "[WARN] ⏹️ Pesquisa interrompida pelo usuário antes de enviar requisição ao portal."
                )
            return {
                "success": False,
                "tentativas": tentativa,
                "tipo_alerta": "INTERRUPCAO_USUARIO",
                "alert_text": "",
                "alert_class": "INTERRUPCAO_USUARIO",
                "motivo": "Pesquisa interrompida pelo usuário.",
            }

        if callable(log_fn):
            log_fn(
                f"[INFO] 🔍 Iniciando pesquisa no portal "
                f"(tentativa {tentativa}/{max_tentativas}) | "
                f"IE {ie_formatada(ie)} | tipo: {tipo.upper()} | "
                f"período: {data_inicial} a {data_final}"
            )

        try:
            abrir_menu_baixar_xml_nfe(page)

            if callable(log_fn):
                log_fn("[INFO] 🧾 Preenchendo filtros de pesquisa (datas, IE e tipo de nota)...")

            preencher_filtros_pesquisa(page, data_inicial, data_final, tipo, ie)

            if callable(log_fn):
                log_fn("[INFO] 🔎 Enviando pesquisa (clique em 'Pesquisar') e aguardando retorno do portal.")
        except TimeoutError as e:
            msg = str(e)
            if "cmpDataInicial" in msg and "Frame was detached" in msg:
                if is_netaccess_login_visible(page):
                    if callable(log_fn):
                        log_fn("[WARN] ⚠️ Sessão expirada (frame desmontado). Reautenticando no segundo login e retomando pesquisa...")
                    ensure_login_ativo(page, log_fn=log_fn, login_cpf=_get_current_login_cpf())
                    abrir_menu_baixar_xml_nfe(page)
                    continue
                else:
                    if callable(log_fn):
                        log_fn("[WARN] ⚠️ Timeout no campo de filtro, mas sem tela de login visível. Repetindo tentativa...")
                    continue
            raise

        # Tentar marcar o reCAPTCHA depois de preencher os filtros (cliques por coordenada)
        try:
            if callable(log_fn):
                log_fn("[INFO] Tentando marcar o reCAPTCHA (após filtros)...")
            ok_cf = clicar_recaptcha_cloudflare(page, timeout_ms=20_000, log_fn=log_fn)
            if not ok_cf and callable(log_fn):
                log_fn("[WARN] Não consegui marcar o reCAPTCHA; seguindo para evitar cliques cegos.")
        except Exception as e:
            if callable(log_fn):
                log_fn(f"[WARN] Erro ao tentar clicar no reCAPTCHA: {e}")

        sleep(10)
        wait_and_click(page, XPATH_BTN_PESQUISAR)

        erro_envio_portal = detectar_erro_sem_dados_enviados_portal(page)
        if erro_envio_portal:
            ultimo_alert_class = "SEM_DADOS_ENVIADOS"
            ultimo_alert_text = erro_envio_portal
            ultimo_motivo = "Portal retornou a pagina 'Nenhum dado foi enviado'."
            if callable(log_fn):
                log_fn(
                    f"[WARN] Portal retornou erro de envio vazio apos a pesquisa "
                    f"(tentativa {tentativa}/{max_tentativas}): {erro_envio_portal}"
                )
            return {
                "success": False,
                "tentativas": tentativa,
                "tipo_alerta": "ERRO_PORTAL_SEM_DADOS",
                "alert_text": erro_envio_portal,
                "alert_class": "SEM_DADOS_ENVIADOS",
                "motivo": "Portal retornou a pagina 'Nenhum dado foi enviado'.",
            }


        tipo_alerta, texto_alerta = esperar_alertas(
            page,
            timeout_ms=15000,
            should_stop=should_stop,
        )

        erro_envio_portal = detectar_erro_sem_dados_enviados_portal(page)
        if erro_envio_portal:
            ultimo_alert_class = "SEM_DADOS_ENVIADOS"
            ultimo_alert_text = erro_envio_portal
            ultimo_motivo = "Portal retornou a pagina 'Nenhum dado foi enviado'."
            if callable(log_fn):
                log_fn(
                    f"[WARN] Portal retornou erro de envio vazio durante o retorno da pesquisa "
                    f"(tentativa {tentativa}/{max_tentativas}): {erro_envio_portal}"
                )
            return {
                "success": False,
                "tentativas": tentativa,
                "tipo_alerta": "ERRO_PORTAL_SEM_DADOS",
                "alert_text": erro_envio_portal,
                "alert_class": "SEM_DADOS_ENVIADOS",
                "motivo": "Portal retornou a pagina 'Nenhum dado foi enviado'.",
            }

        if tipo_alerta == "SEM_PERMISSAO":
            teve_sem_permissao = True
            ultimo_alert_class = "SEM_PERMISSAO"
            ultimo_alert_text = "Você não tem permissão para acessar esta página."
            ultimo_motivo = "Sem permissão para acessar a consulta no portal."
            if callable(log_fn):
                log_fn(
                    f"[WARN] 🚫 Mensagem de falta de permissão detectada na tentativa "
                    f"{tentativa}/{max_tentativas}. Recarregando e tentando de novo."
                )
            try:
                page.reload()
                page.wait_for_load_state("networkidle")
            except Exception:
                pass
            continue

        # Parada manual enquanto esperava o retorno do portal
        if tipo_alerta == "INTERRUPCAO_USUARIO":
            if callable(log_fn):
                log_fn(
                    "[WARN] ⏹️ Pesquisa interrompida pelo usuário durante o aguardo de resposta do portal."
                )
            return {
                "success": False,
                "tentativas": tentativa,
                "tipo_alerta": "INTERRUPCAO_USUARIO",
                "alert_text": "",
                "alert_class": "INTERRUPCAO_USUARIO",
                "motivo": "Pesquisa interrompida pelo usuário.",
            }

        # ALERTA GERAL (normalmente erro interno / captcha / etc.)
        if tipo_alerta == "geral":
            alert_class = classificar_alerta(texto_alerta)
            detalhes = limpar_texto_alerta(texto_alerta)
            ultimo_alert_class = alert_class or "OUTRO"
            ultimo_alert_text = detalhes or ""
            ultimo_motivo = "Alerta geral retornado pelo sistema"

            if callable(log_fn):
                log_fn(
                    f"[WARN] ⚠️ Portal retornou ALERTA GERAL "
                    f"(classe: {alert_class}) na tentativa {tentativa}/{max_tentativas}: "
                    f"{detalhes or 'sem texto.'}"
                )
            clicar_alert_button(page)

            if tentativa == max_tentativas:
                # NOVO: priorizar SEM_RESULTADO se ocorreu em alguma tentativa
                final_alert_class = alert_class
                final_alert_text = limpar_texto_alerta(texto_alerta)
                final_motivo = "Alerta geral mesmo após múltiplas tentativas"

                # Só prioriza sem movimento sobre erros transitórios (ex.: CAPTCHA/OUTRO/DATA_OBRIGATÓRIA)
                if teve_sem_resultado and final_alert_class in ("CAPTCHA", "OUTRO", "DATA_OBRIGATÓRIA", None):
                    final_alert_class = "SEM_RESULTADO"
                    final_alert_text = texto_sem_resultado or "Sem Resultados!"
                    final_motivo = "Sem movimento detectado em uma das tentativas"

                if callable(log_fn):
                    log_fn(
                        "[ERRO] ❌ Limite máximo de tentativas atingido para este período. "
                        f"Pesquisa será marcada como falha ({final_alert_class})."
                    )
                return {
                    "success": False,
                    "tentativas": tentativa,
                    "tipo_alerta": tipo_alerta,
                    "alert_text": final_alert_text,
                    "alert_class": final_alert_class or "OUTRO",
                    "motivo": final_motivo,
                }
            else:
                if callable(log_fn):
                    log_fn(
                        f"[INFO] 🔁 Nova tentativa será realizada para o mesmo período "
                        f"({tentativa + 1}/{max_tentativas})."
                    )
            continue

        # ALERTA TEXTUAL
        if tipo_alerta == "textual":
            alert_class = classificar_alerta(texto_alerta)
            detalhes = limpar_texto_alerta(texto_alerta)
            ultimo_alert_class = alert_class or "OUTRO"
            ultimo_alert_text = detalhes or ""
            ultimo_motivo = "Alerta textual retornado pelo sistema"

            # NOVO: se apareceu SEM_RESULTADO em qualquer tentativa, guardamos
            if alert_class == "SEM_RESULTADO":
                teve_sem_resultado = True
                if detalhes:
                    texto_sem_resultado = detalhes

            # RE-TENTAR quando for um alerta transitório:
            # - DATA_OBRIGATÓRIA (instabilidade do portal/filtros)
            # - CAPTCHA (captcha inválido/expirado)
            # - SEM_RESULTADO (falso negativo do portal acontece; tentar novamente)
            if alert_class in ("DATA_OBRIGATÓRIA", "CAPTCHA", "SEM_RESULTADO") and tentativa < max_tentativas:
                if callable(log_fn):
                    log_fn(
                        f"[WARN] 📢 Portal retornou ALERTA TEXTUAL (classe: {alert_class}) "
                        f"na tentativa {tentativa}/{max_tentativas}: {detalhes or 'sem texto.'}"
                    )
                    log_fn(
                        f"[INFO] 🔁 Alerta tratável ({alert_class}). "
                        f"Será feita nova tentativa ({tentativa + 1}/{max_tentativas}) "
                        f"repreenchendo os filtros."
                    )
                clicar_alert_button(page)
                # pequena pausa para o portal estabilizar/renovar captcha
                sleep(1)
                continue

            # Demais alertas textuais encerram a pesquisa do período
            # (inclui o caso de alerta "transitório" na última tentativa)
            # Aplicar PRIORIDADE do SEM_RESULTADO se ele ocorreu antes.
            preferir_sem_mov = (
                teve_sem_resultado
                and alert_class in ("CAPTCHA", "OUTRO", "DATA_OBRIGATÓRIA", "SEM_RESULTADO")
            )

            final_alert_class = "SEM_RESULTADO" if preferir_sem_mov else alert_class
            final_alert_text = (texto_sem_resultado or "Sem Resultados!") if preferir_sem_mov else detalhes
            final_motivo = "Sem movimento detectado em uma das tentativas" if preferir_sem_mov else "Alerta textual retornado pelo sistema"

            if callable(log_fn):
                log_fn(
                    f"[WARN] 📢 Portal retornou ALERTA TEXTUAL "
                    f"(classe: {final_alert_class}). Pesquisa para este período será encerrada. "
                    f"Detalhes: {final_alert_text or 'sem texto.'}"
                )
            clicar_alert_button(page)
            return {
                "success": False,
                "tentativas": tentativa,
                "tipo_alerta": tipo_alerta,
                "alert_text": final_alert_text,
                "alert_class": final_alert_class,
                "motivo": final_motivo,
            }

        # Sem alerta -> SUCESSO
        if callable(log_fn):
            log_fn(
                f"[OK] ✅ Pesquisa concluída sem alertas "
                f"(tentativa {tentativa}/{max_tentativas})."
            )
        return {
            "success": True,
            "tentativas": tentativa,
            "tipo_alerta": None,
            "alert_text": "",
            "alert_class": "",
            "motivo": "",
        }

    # fallback (não houve retorno dentro do loop por algum motivo)
    if callable(log_fn):
        log_fn("[ERRO] ❌ Máximo de tentativas atingido sem sucesso (fallback).")
    # NOVO: se em alguma tentativa houve SEM_RESULTADO, prioriza isso no fallback também
    if teve_sem_resultado:
        return {
            "success": False,
            "tentativas": max_tentativas,
            "tipo_alerta": "textual",
            "alert_text": texto_sem_resultado or "Sem Resultados!",
            "alert_class": "SEM_RESULTADO",
            "motivo": "Sem movimento detectado em uma das tentativas",
        }
    if teve_sem_permissao:
        return {
            "success": False,
            "tentativas": max_tentativas,
            "tipo_alerta": "SEM_PERMISSAO",
            "alert_text": "Você não tem permissão para acessar esta página.",
            "alert_class": "SEM_PERMISSAO",
            "motivo": "Sem permissão para acessar a consulta no portal.",
        }
    if ultimo_alert_class:
        return {
            "success": False,
            "tentativas": max_tentativas,
            "tipo_alerta": "geral",
            "alert_text": ultimo_alert_text,
            "alert_class": ultimo_alert_class,
            "motivo": ultimo_motivo or "Falha na pesquisa após tentativas.",
        }
    return {
        "success": False,
        "tentativas": max_tentativas,
        "tipo_alerta": "geral",
        "alert_text": "Falha sem classificação específica.",
        "alert_class": "SERVICE_503" if teve_503 else "OUTRO",
        "motivo": ("Portal indisponível (HTTP 503) durante as tentativas." if teve_503 else "Falha sem classificação específica após tentativas."),
    }

# =============================================================================
# Downloads: selecionar opção no modal + solicitar + aguardar histórico
# =============================================================================

def selecionar_opcao_download_modal(
    page,
    download_option: str,
    timeout: int = DEFAULT_TIMEOUT_MS,
    log_fn: Optional[Callable[[str], None]] = None,
) -> bool:
    """
    Seleciona a opção do modal 'Baixar todos os arquivos':
      download_option: 'docs', 'eventos', 'ambos'
    Tenta por XPath fixo e por texto do label.
    """
    textos_por_opcao = {
        "docs": ["somente documentos", "documentos apenas"],
        "eventos": ["somente eventos", "eventos apenas"],
        "ambos": ["documentos e eventos", "docs e eventos", "documentos e  eventos"],
    }

    if download_option == "docs":
        xpaths_preferenciais = [XPATH_MODAL_OPT_DOCS]
    elif download_option == "eventos":
        xpaths_preferenciais = [XPATH_MODAL_OPT_EVENTOS]
    else:
        xpaths_preferenciais = [XPATH_MODAL_OPT_DOCS_EVENTOS]

    deadline = datetime.now() + timedelta(milliseconds=timeout)

    while datetime.now() < deadline:
        for frame in page.frames:
            # 1) XPaths fixos (inputs)
            for xp in xpaths_preferenciais:
                try:
                    sel = resolve_selector(xp)
                    el = frame.query_selector(sel)
                    if el and el.is_visible():
                        try:
                            el.check()
                        except Exception:
                            try:
                                el.click()
                            except Exception:
                                el.evaluate("e => e.click()")
                        if callable(log_fn):
                            log_fn("[INFO] ✅ Opção de download selecionada no modal via XPath fixo.")
                        return True
                except Exception:
                    continue

            # 2) Por texto de label
            for txt in textos_por_opcao.get(download_option, []):
                try:
                    xp_label = (
                        "//div[contains(@class,'modal') and contains(@style,'display: block')]"
                        f"//label[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'), '{txt.lower()}')]"
                    )
                    sel_label = resolve_selector(xp_label)
                    label_el = frame.query_selector(sel_label)
                    if label_el and label_el.is_visible():
                        input_el = label_el.query_selector("input")
                        if input_el:
                            try:
                                input_el.check()
                            except Exception:
                                try:
                                    input_el.click()
                                except Exception:
                                    input_el.evaluate("e => e.click()")
                        else:
                            label_el.click()
                        if callable(log_fn):
                            log_fn("[INFO] ✅ Opção de download selecionada no modal via texto do label.")
                        return True
                except Exception:
                    continue

        sleep(0.3)

    if callable(log_fn):
        log_fn("[ERRO] ❌ Não foi possível selecionar nenhuma opção de download no modal dentro do tempo limite.")
    return False

def solicitar_geracao_zip(
    page,
    download_option: str,
    tentativas: int = 3,
    log_fn: Optional[Callable[[str], None]] = None,
) -> bool:
    """
    Na tela de resultado:
      - Clica 'Baixar todos os arquivos'
      - Se aparecer a label 'Você não tem permissão para acessar esta página.',
        tenta novamente até 'tentativas' vezes.
      - Seleciona a opção (docs/eventos/ambos)
      - Clica 'Baixar'
    O site redireciona para o Histórico de Downloads.

    NOVO:
      - Se, ao clicar em 'Baixar' no modal, aparecer o alerta
        "Já existe um arquivo com os mesmos parâmetros solicitados em processamento...",
        NÃO faz nova solicitação. Em vez disso, usa o pedido já existente no
        Histórico de Downloads (a próxima chamada de aguardar_e_baixar_arquivos
        continua a lógica normal de esperar 'Concluído' e baixar).
    """
    for tentativa in range(1, tentativas + 1):
        if callable(log_fn):
            log_fn(
                f"[INFO] 📦 Iniciando solicitação de geração de ZIP "
                f"(tentativa {tentativa}/{tentativas}) | opção: {download_option}."
            )
        # 1) Clicar em "Baixar todos os arquivos"
        try:
            wait_and_click(page, XPATH_BTN_BAIXAR_TODOS, timeout=10_000)
            if callable(log_fn):
                log_fn("[INFO] 📥 Botão 'Baixar todos os arquivos' clicado. Verificando permissões...")
        except Exception as e:
            if tentativa < tentativas:
                if callable(log_fn):
                    log_fn(
                        f"[WARN] ⚠️ Falha ao clicar em 'Baixar todos os arquivos' "
                        f"(tentativa {tentativa}/{tentativas}): {e} | tentando de novo..."
                    )
                sleep(2)
                continue  # volta para o for tentativa in range(...)
            else:
                if callable(log_fn):
                    log_fn(
                        f"[ERRO] ❌ Falha ao clicar em 'Baixar todos os arquivos' "
                        f"na última tentativa ({tentativa}/{tentativas}): {e} "
                        "(não será possível solicitar o ZIP)."
                    )
                return False


        # 2) Verificar se apareceu a mensagem de "sem permissão"
        try:
            sel_label = resolve_selector(XPATH_LABEL_SEM_PERMISSAO)
            encontrou = False
            deadline = datetime.now() + timedelta(seconds=3)

            while datetime.now() < deadline and not encontrou:
                for frame in page.frames:
                    try:
                        el = frame.query_selector(sel_label)
                        if el and el.is_visible():
                            encontrou = True
                            break
                    except Exception:
                        continue
                if not encontrou:
                    sleep(0.2)

            if encontrou:
                if callable(log_fn):
                    log_fn(
                        f"[WARN] 🚫 Mensagem de 'sem permissão' encontrada ao solicitar ZIP "
                        f"(tentativa {tentativa}/{tentativas})."
                    )
                # Se encontrou a mensagem de permissão, tenta de novo
                if tentativa == tentativas:
                    if callable(log_fn):
                        log_fn(
                            "[ERRO] ❌ Limite de tentativas ao solicitar ZIP com mensagem de permissão. "
                            "Desistindo desta solicitação."
                        )
                    return False
                if callable(log_fn):
                    log_fn("[INFO] 🔁 Aguardando para tentar solicitar o ZIP novamente...")
                sleep(1)
                continue
        except Exception:
            # Se der erro nessa verificação, simplesmente segue o fluxo normal
            if callable(log_fn):
                log_fn(
                    "[WARN] ⚠️ Erro ao verificar mensagem de permissão, "
                    "seguindo fluxo normal do download."
                )

        # 3) Selecionar opção do modal
        ok_select = selecionar_opcao_download_modal(
            page,
            download_option,
            timeout=10_000,
            log_fn=log_fn
        )
        if not ok_select:
            if callable(log_fn):
                log_fn(
                    "[ERRO] ❌ Não foi possível selecionar a opção de download no modal. "
                    "Solicitação de ZIP será abortada."
                )
            return False

        # 4) Confirmar (clicar em "Baixar" / "OK" do modal)
        try:
            if callable(log_fn):
                log_fn("[INFO] ⏬ Confirmando modal de download (clicando em 'Baixar/OK')...")
            wait_and_click(page, XPATH_MODAL_BTN_OK, timeout=10_000)

            tipo_alerta, texto_alerta = esperar_alertas(page, timeout_ms=8_000)
            if texto_alerta:
                detalhes = limpar_texto_alerta(texto_alerta)
                # normaliza texto para ignorar acentos
                texto_lower = detalhes.lower()
                try:
                    import unicodedata
                    texto_norm = unicodedata.normalize("NFD", texto_lower)
                    texto_norm = "".join(
                        ch for ch in texto_norm
                        if unicodedata.category(ch) != "Mn"
                    )
                except Exception:
                    texto_norm = texto_lower

                if "ja existe um arquivo com os mesmos parametros solicitados em processamento" in texto_norm:
                    if callable(log_fn):
                        log_fn(
                            "[INFO] 🔁 Já existe um arquivo com os mesmos parâmetros em processamento. "
                            "Usando o pedido existente no Histórico de Downloads."
                        )

                    # Fecha o alerta
                    clicar_alert_button(page)

                    # Tenta clicar diretamente no botão "Histórico de Downloads"
                    try:
                        if callable(log_fn):
                            log_fn("[INFO] 📜 Clicando no botão 'Histórico de Downloads' após alerta de solicitação já existente.")
                        wait_and_click(page, XPATH_BTN_HISTORICO_DOWNLOADS, timeout=10_000)
                        page.wait_for_load_state("networkidle")
                        sleep(1)
                    except Exception as e:
                        # Se der erro, cai no plano B: usar a função genérica para garantir a página
                        if callable(log_fn):
                            log_fn(
                                f"[WARN] ⚠️ Falha ao clicar no botão 'Histórico de Downloads' após alerta: {e}. "
                                "Tentando garantir página via busca de texto."
                            )

                    # Garante que estamos, de fato, na página de histórico
                    garantir_pagina_historico_downloads(page, log_fn=log_fn)
                    return True

                elif callable(log_fn) and tipo_alerta != "nenhum":
                    log_fn(
                        f"[WARN] ⚠️ Alerta inesperado após confirmar o modal de download: "
                        f"{detalhes}"
                    )

            if callable(log_fn):
                log_fn("[OK] ✅ Solicitação de geração de ZIP enviada com sucesso. Redirecionando ao histórico...")
            return True
        except Exception as e:
            if tentativa == tentativas:
                if callable(log_fn):
                    log_fn(
                        f"[ERRO] ❌ Falha ao confirmar o modal de download na última tentativa: {e} "
                        "Solicitação de ZIP será encerrada."
                    )
                return False
            else:
                if callable(log_fn):
                    log_fn(
                        f"[WARN] ⚠️ Falha ao confirmar modal de download: {e}. "
                        f"Tentando novamente ({tentativa + 1}/{tentativas})..."
                    )
            sleep(1)

    return False

def obter_linhas_historico(page):
    """Retorna elementos <tr> do histórico."""
    linhas = []
    for frame in page.frames:
        try:
            rs = frame.query_selector_all("tr.tbody-row.paginated-row")
            if rs:
                linhas.extend(rs)
        except Exception:
            continue
    return linhas

def garantir_pagina_historico_downloads(
    page,
    log_fn: Optional[Callable[[str], None]] = None,
    timeout_segundos: int = 15,
) -> bool:
    """
    Garante que estamos na tela de 'Histórico de Downloads'.

    Se a tabela do histórico já estiver visível, não faz nada.
    Caso contrário, tenta clicar em algum link/texto 'Histórico de Downloads'.
    """

    # Já está no histórico?
    if obter_linhas_historico(page):
        if callable(log_fn):
            log_fn("[INFO] 📜 Página de Histórico de Downloads já está aberta.")
        return True

    textos_busca = [
        "Histórico de Downloads",
        "Histórico de downloads",
        "Histórico de Download",
        "Histórico de download",
        "Histórico de Arquivos",
        "Histórico de arquivos",
    ]

    deadline = datetime.now() + timedelta(seconds=timeout_segundos)

    while datetime.now() < deadline:
        # Se, em algum momento, a tabela aparecer, ok.
        if obter_linhas_historico(page):
            if callable(log_fn):
                log_fn("[INFO] 📜 Página de Histórico de Downloads aberta com sucesso.")
            return True

        try:
            for txt in textos_busca:
                try:
                    loc = page.get_by_text(txt, exact=False)
                    if loc.is_visible():
                        loc.click()
                        page.wait_for_load_state("networkidle")
                        sleep(1)
                        break
                except Exception:
                    continue
        except Exception:
            pass

        if obter_linhas_historico(page):
            if callable(log_fn):
                log_fn("[INFO] 📜 Página de Histórico de Downloads aberta com sucesso.")
            return True

        sleep(0.5)

    if callable(log_fn):
        log_fn(
            "[WARN] ⚠️ Não foi possível garantir navegação para a página de "
            "Histórico de Downloads. Aguardando assim mesmo."
        )
    return False

def reabrir_historico_via_pesquisa(
    page,
    ie: str,
    data_inicial: str,
    data_final: str,
    tipo: str,
    should_stop=None,
    inicio: Optional[datetime] = None,
    timeout_total: Optional[int] = None,
    log_fn: Optional[Callable[[str], None]] = None,
) -> bool:
    """
    Após 1 minuto sem encontrar 'Concluído' no Histórico de Downloads, esta rotina:
      - Recarrega a página
      - Clica no menu superior '/html/body/div[1]/div[1]/ul/li[4]/a'
      - Preenche novamente período, IE e tipo de nota
      - Clica em 'Pesquisar'
          * Se der erro de CAPTCHA/recaptcha => tenta de novo, sem limite de 5x
          * Para outros alertas, encerra com erro
      - Na tela de resultado, clica em 'Histórico de Downloads' (não clica em 'Baixar todos')

    NOVO:
      - Se após o reload aparecer o label
        "Você não tem permissão para acessar esta página.",
        considera como erro transitório: tenta recarregar de novo e repetir o fluxo.
    """

    if not (ie and data_inicial and data_final and tipo):
        if callable(log_fn):
            log_fn(
                "[WARN] ⚠️ Dados insuficientes (IE/período/tipo) para refazer "
                "a pesquisa e reabrir o Histórico de Downloads."
            )
        return False

    if inicio is None:
        inicio = datetime.now()

    while True:
        # Parada externa / timeout global
        if callable(should_stop) and should_stop():
            if callable(log_fn):
                log_fn(
                    "[WARN] ⏹️ Parada solicitada enquanto tentava reabrir o Histórico de Downloads."
                )
            return False

        if timeout_total and (datetime.now() - inicio).total_seconds() >= timeout_total:
            if callable(log_fn):
                log_fn(
                    "[ERRO] ⏰ Tempo máximo geral atingido enquanto tentava reabrir o Histórico de Downloads."
                )
            return False

        # 1) Recarrega a página
        try:
            if callable(log_fn):
                log_fn("[INFO] 🔄 Recarregando página para tentar reabrir o Histórico de Downloads...")
            page.reload()
            page.wait_for_load_state("networkidle")
        except Exception as e:
            if callable(log_fn):
                log_fn(f"[WARN] ⚠️ Erro ao recarregar a página: {e}")
            # Mesmo com erro de reload, tenta seguir o fluxo

        # 1.1) Verifica se caiu na página de "Você não tem permissão..."
        try:
            sel_label_perm = resolve_selector(XPATH_LABEL_SEM_PERMISSAO)
            encontrou_perm = False
            for frame in page.frames:
                try:
                    el = frame.query_selector(sel_label_perm)
                    if el and el.is_visible():
                        encontrou_perm = True
                        break
                except Exception:
                    continue

            if encontrou_perm:
                if callable(log_fn):
                    log_fn(
                        "[WARN] 🚫 Página de 'Você não tem permissão para acessar esta página.' "
                        "após recarregar. Nova recarga será tentada automaticamente."
                    )
                sleep(1)
                # volta para o início do while True, respeitando timeout_total
                continue
        except Exception:
            # Se der erro nessa verificação, segue o fluxo normal
            pass

        # 2) Clica no menu superior "Baixar XML NFE"
        try:
            if callable(log_fn):
                log_fn("[INFO] 📄 Acessando novamente a tela 'Baixar XML NFE' pelo menu superior.")
            try:
                wait_and_click(page, XPATH_MENU_SUPERIOR_BAIXAR_XML, timeout=15_000)
            except Exception:
                sel_menu = resolve_selector(XPATH_MENU_SUPERIOR_BAIXAR_XML)
                clicked = False
                for frame in page.frames:
                    try:
                        el = frame.query_selector(sel_menu)
                        if el and el.is_visible():
                            el.click()
                            clicked = True
                            break
                    except Exception:
                        continue
                if not clicked:
                    if callable(log_fn):
                        log_fn(
                            "[ERRO] ❌ Não foi possível clicar no menu superior "
                            "'Baixar XML NFE' para reabrir o histórico."
                        )
                    return False
        except Exception as e:
            if callable(log_fn):
                log_fn(f"[ERRO] ❌ Falha ao acessar menu superior 'Baixar XML NFE': {e}")
            return False

        # 3) Preencher filtros (período, IE, tipo)
        try:
            if callable(log_fn):
                log_fn(
                    f"[INFO] 🧾 Reprenchendo filtros (IE, período e operação) "
                    f"para reabrir o histórico. IE: {ie_formatada(ie)}, "
                    f"Período: {data_inicial} a {data_final}, Tipo: {tipo}."
                )
            preencher_filtros_pesquisa(page, data_inicial, data_final, tipo, ie)
        except Exception as e:
            if callable(log_fn):
                log_fn(f"[ERRO] ❌ Erro ao preencher filtros para reabrir histórico: {e}")
            return False

        # 3.1) Espera 10 segundos antes de clicar em Pesquisar
        if callable(log_fn):
            log_fn("[INFO] ⏲️ Aguardando 10 segundos antes de clicar em 'Pesquisar' após a recarga...")
        sleep(10)

        # 4) Loop interno: clicar em Pesquisar e tratar alertas (incluindo CAPTCHA infinito)
        while True:
            if callable(should_stop) and should_stop():
                if callable(log_fn):
                    log_fn(
                        "[WARN] ⏹️ Parada solicitada enquanto aguardava o resultado "
                        "da pesquisa para reabrir o histórico."
                    )
                return False

            if timeout_total and (datetime.now() - inicio).total_seconds() >= timeout_total:
                if callable(log_fn):
                    log_fn(
                        "[ERRO] ⏰ Tempo máximo geral atingido enquanto aguardava "
                        "resultado da pesquisa para reabrir histórico."
                    )
                return False

            # Clicar em Pesquisar
            try:
                if callable(log_fn):
                    log_fn("[INFO] 🔍 Enviando pesquisa (somente para reabrir o Histórico de Downloads)...")
                wait_and_click(page, XPATH_BTN_PESQUISAR, timeout=DEFAULT_TIMEOUT_MS)
            except Exception as e:
                if callable(log_fn):
                    log_fn(f"[ERRO] ❌ Falha ao clicar em 'Pesquisar' ao tentar reabrir histórico: {e}")
                # Sai do while interno e recomeça todo o ciclo (reload + menu + filtros)
                break

            # Verificar alertas
            tipo_alerta, texto_alerta = esperar_alertas(page, timeout_ms=15000)

            if tipo_alerta == "SEM_PERMISSAO":
                if callable(log_fn):
                    log_fn(
                        "[WARN] 🚫 Mensagem de falta de permissão detectada ao reabrir histórico. "
                        "Recarregando e reiniciando fluxo."
                    )
                try:
                    page.reload()
                    page.wait_for_load_state("networkidle")
                except Exception:
                    pass
                break  # volta ao início do while externo (reload + filtros + pesquisar)

            if tipo_alerta == "geral":
                alert_class = classificar_alerta(texto_alerta)
                detalhes = limpar_texto_alerta(texto_alerta)
                if callable(log_fn):
                    log_fn(
                        f"[WARN] ⚠️ Alerta geral ao tentar reabrir histórico "
                        f"(classe: {alert_class}): {detalhes or 'sem texto.'}"
                    )
                clicar_alert_button(page)

                # CAPTCHA ou DATA_OBRIGATÓRIA -> trata como transitório
                if alert_class in ("CAPTCHA", "DATA_OBRIGATÓRIA"):
                    if callable(log_fn):
                        if alert_class == "CAPTCHA":
                            log_fn(
                                "[INFO] 🔁 CAPTCHA detectado ao tentar reabrir histórico. "
                                "Reiniciando fluxo (reload + menu + filtros + pesquisar)..."
                            )
                        else:
                            log_fn(
                                "[INFO] 🔁 Alerta de data obrigatória ao tentar reabrir histórico. "
                                "Reiniciando fluxo (reload + menu + filtros + pesquisar)..."
                            )
                    # Sai do while interno e volta para o while externo
                    break
                else:
                    if callable(log_fn):
                        log_fn(
                            "[ERRO] ❌ Alerta geral não tratável ao reabrir histórico. "
                            "Rotina será encerrada."
                        )
                    return False

            elif tipo_alerta == "textual":
                alert_class = classificar_alerta(texto_alerta)
                detalhes = limpar_texto_alerta(texto_alerta)
                if callable(log_fn):
                    log_fn(
                        f"[WARN] 📢 Alerta textual ao tentar reabrir histórico "
                        f"(classe: {alert_class}): {detalhes or 'sem texto.'}"
                    )
                clicar_alert_button(page)
                return False

            else:
                # Sem alerta -> estamos na tela de resultado
                try:
                    if callable(log_fn):
                        log_fn("[INFO] 📜 Pesquisa concluída. Abrindo 'Histórico de Downloads' (sem solicitar novo ZIP).")
                    wait_and_click(page, XPATH_BTN_HISTORICO_DOWNLOADS, timeout=15_000)
                    page.wait_for_load_state("networkidle")
                    garantir_pagina_historico_downloads(page, log_fn=log_fn)
                    if callable(log_fn):
                        log_fn("[OK] ✅ Histórico de Downloads reaberto com sucesso.")
                    return True
                except Exception as e:
                    if callable(log_fn):
                        log_fn(
                            f"[ERRO] ❌ Falha ao clicar em 'Histórico de Downloads' "
                            f"após a pesquisa: {e}"
                        )
                    return False

def aguardar_e_baixar_arquivos(
    page,
    caminho_destino: str,
    quantidade: int,
    timeout_total: int = DOWNLOAD_HISTORY_TIMEOUT_SECONDS,
    should_stop=None,
    log_fn: Optional[Callable[[str], None]] = None,
    ie: Optional[str] = None,
    data_inicial: Optional[str] = None,
    data_final: Optional[str] = None,
    tipo: Optional[str] = None,
    download_option: Optional[str] = None,
    max_sem_resultados: int = 5,
    max_reaberturas_historico: int = 4,
) -> List[str]:

    """
    Na página de Histórico:
      - Espera N linhas mais recentes com:
          * col-situacao = 'Concluído'
          * link 'Baixar XML' visível
      - Baixa cada ZIP via expect_download() salvando em caminho_destino
      - Retorna lista de caminhos completos dos arquivos ZIP baixados.

    NOVA ROTINA DE ESPERA (sem clicar mais no XPath //*[@id="menuBox"]/ul/li[7]/ul/li):
      - Enquanto não tiver 'Concluído':
          * Fica checando normalmente (a cada 2s)
          * Se passar 1 minuto sem achar 'Concluído':
              - Recarrega a página
              - Clica no menu superior '/html/body/div[1]/div[1]/ul/li[4]/a'
              - Preenche período, IE e tipo
              - Clica em Pesquisar
                  · Se der CAPTCHA/recaptcha → tenta de novo, sem limite, até conseguir
              - Na tela de resultado, clica em 'Histórico de Downloads'
              - Volta a checar 'Concluído' por mais 1 minuto, e assim sucessivamente
    """

    inicio = datetime.now()
    ultimo_refresh = datetime.now()
    intervalo_refresh = 60  # segundos
    tentativas_reabertura_historico = 0
    tentativas_sem_resultados = 0  # contador específico para o bug "Sem resultados"

    safe_makedirs(caminho_destino)


    if callable(log_fn):
        log_fn(
            f"[INFO] ⏳ Aguardando {quantidade} arquivo(s) no histórico de downloads "
            f"com situação 'Concluído' (timeout total: {timeout_total}s)..."
        )

    while True:
        # Parada externa
        if callable(should_stop) and should_stop():
            if callable(log_fn):
                log_fn(
                    "[WARN] ⏹️ Parada solicitada durante a espera de downloads no histórico. "
                    "Cancelando aguardo e não baixando arquivos."
                )
            return []

        # Se o portal mostrar erro 503 (Service Unavailable), recarrega e continua aguardando
        if detectar_erro_503(page):
            if callable(log_fn):
                log_fn("[WARN] ⚠️ Portal retornou Service Unavailable (HTTP 503) no histórico. Recarregando e tentando novamente...")
            try:
                page.reload(wait_until="networkidle", timeout=60_000)
            except Exception:
                try:
                    page.reload()
                except Exception:
                    pass
            sleep(3)
            continue

        # Timeout total (em segundos)
        if timeout_total and (datetime.now() - inicio).total_seconds() >= timeout_total:
            if callable(log_fn):
                log_fn(
                    "[ERRO] ⏰ Tempo máximo atingido ao aguardar arquivos 'Concluído' "
                    "no histórico de downloads."
                )
            return []

        # -------- ROTINA DE REABERTURA DO HISTÓRICO APÓS 1 MINUTO --------
        if (datetime.now() - ultimo_refresh).total_seconds() >= intervalo_refresh:
            tentativas_reabertura_historico += 1
            if tentativas_reabertura_historico > max_reaberturas_historico:
                if callable(log_fn):
                    log_fn(
                        "[ERRO] ❌ Limite de reaberturas do Histórico de Downloads atingido. "
                        "Desistindo deste download para evitar loop infinito."
                    )
                return []
            if callable(log_fn):
                log_fn(
                    "[INFO] 🔄 Mais de 1 minuto aguardando download sem 'Concluído'. "
                    f"Recarregando página e reabrindo o Histórico de Downloads através de nova pesquisa "
                    f"({tentativas_reabertura_historico}/{max_reaberturas_historico})."
                )
            ultimo_refresh = datetime.now()

            # Só faz sentido se tivermos os dados do intervalo/IE/tipo
            if ie and data_inicial and data_final and tipo:
                ok_hist = reabrir_historico_via_pesquisa(
                    page=page,
                    ie=ie,
                    data_inicial=data_inicial,
                    data_final=data_final,
                    tipo=tipo,
                    should_stop=should_stop,
                    inicio=inicio,
                    timeout_total=timeout_total,
                    log_fn=log_fn,
                )
                if not ok_hist:
                    if callable(log_fn):
                        log_fn(
                            "[ERRO] ❌ Não foi possível reabrir o Histórico de Downloads "
                            "durante a espera de conclusão dos arquivos."
                        )
                    return []
                # Se deu certo, estamos de novo no histórico → volta ao loop normal
                continue
            else:
                if callable(log_fn):
                    log_fn(
                        "[WARN] ⚠️ IE/período/tipo não informados para reabrir histórico. "
                        "Ignorando rotina de recarga e continuando a aguardar."
                    )
                # Não tenta nada especial, só segue checando
                # (mantém o while principal rodando)
                # NÃO dá continue aqui, para já cair na checagem das linhas
                # após atualizar o ultimo_refresh.

        # -------- VERIFICAÇÃO NORMAL DO HISTÓRICO --------
        linhas = obter_linhas_historico(page)

        if len(linhas) >= quantidade:
            links_para_baixar = []
            ok = True
            linha_sem_resultados = False  # flag para o bug do SEFAZ

            for row in linhas[:quantidade]:
                try:
                    # Situação da linha
                    situacao_el = row.query_selector("td.col-situacao")
                    situacao = (situacao_el.inner_text() or "").strip() if situacao_el else ""
                    if "Concluído" not in situacao:
                        ok = False
                        break

                    # NOVO: checa a coluna de observações
                    obs_el = row.query_selector("td.col-observacoes")
                    observacao = (obs_el.inner_text() or "").strip() if obs_el else ""
                    if observacao and "sem resultados" in observacao.lower():
                        linha_sem_resultados = True
                        ok = False
                        break

                    # Link de download
                    link = row.query_selector("td.col-acoes a.btn.btn-info")
                    if not link or not link.is_visible():
                        ok = False
                        break

                    links_para_baixar.append(link)
                except Exception:
                    ok = False
                    break

            # Tratamento específico para o bug "Sem resultados" no histórico
            if linha_sem_resultados:
                tentativas_sem_resultados += 1
                if callable(log_fn):
                    log_fn(
                        f"[WARN] ⚠️ Linha do histórico com observação 'Sem resultados' detectada "
                        f"(tentativa {tentativas_sem_resultados}/{max_sem_resultados}). "
                        "Considerando como bug do SEFAZ e tentando gerar o ZIP novamente."
                    )

                if tentativas_sem_resultados > max_sem_resultados:
                    if callable(log_fn):
                        log_fn(
                            "[ERRO] ❌ Limite de tentativas por 'Sem resultados' no histórico atingido. "
                            "Desistindo deste download."
                        )
                    return []

                if download_option:
                    # Tenta solicitar um novo ZIP para o mesmo filtro
                    if not solicitar_geracao_zip(
                        page,
                        download_option,
                        log_fn=log_fn,
                    ):
                        if callable(log_fn):
                            log_fn(
                                "[ERRO] ❌ Falha ao solicitar novo ZIP após 'Sem resultados' no histórico."
                            )
                        return []

                    # Garante que estamos de volta no histórico
                    garantir_pagina_historico_downloads(page, log_fn=log_fn)
                else:
                    if callable(log_fn):
                        log_fn(
                            "[WARN] ⚠️ download_option não informado; não é possível "
                            "re-solicitar o ZIP após 'Sem resultados'."
                        )

                # Volta ao início do while para checar novamente
                sleep(2)
                continue

            # Se já estiver tudo "Concluído" e com link visível, prossegue IMEDIATAMENTE
            if ok and len(links_para_baixar) == quantidade:
                baixados = []
                if callable(log_fn):
                    log_fn(
                        f"[INFO] ⬇️ Iniciando download de {len(links_para_baixar)} arquivo(s) ZIP "
                        f"para a pasta: {caminho_destino}"
                    )

                for link in links_para_baixar:
                    with page.expect_download() as download_info:
                        link.click()
                    download = download_info.value
                    nome_raw = download.suggested_filename or "download.zip"
                    destino = build_safe_file_path(caminho_destino, nome_raw)
                    nome_final = os.path.basename(destino)

                    if callable(log_fn):
                        if nome_final != nome_raw:
                            log_fn(f"[WARN] ⚠️ Nome do ZIP ajustado por limite de caminho: '{nome_raw}' -> '{nome_final}'")
                        log_fn(f"[INFO] 📁 Baixando arquivo: {nome_final}")

                    saved_path = save_download_with_fallback(download, destino, log_fn=log_fn)
                    baixados.append(saved_path)

                if callable(log_fn):
                    log_fn("[OK] ✅ Todos os arquivos ZIP foram baixados com sucesso.")

                return baixados

        # Pequena espera antes da próxima checagem
        sleep(2)

def unificar_zips_xml(
    zip_paths: List[str],
    dest_dir: str,
    ie: str,
    display_name: str,
    operacao: str,
    data_inicial: str,
    data_final: str,
    log_fn=None
) -> Optional[str]:
    """
    Lê todos os ZIPs em zip_paths e gera um único ZIP final
    diretamente em dest_dir, sem usar pasta temporária em disco.

      "Unificado - OPERACAO - DATA_INI a DATA_FIM - DISPLAY_NAME.zip"

    Remove os ZIPs originais após unificação.
    """
    if not zip_paths:
        if log_fn:
            log_fn(
                f"[WARN] 📂 Nenhum arquivo ZIP informado para unificação "
                f"({display_name} - {ie_formatada(ie)})."
            )
        return None

    safe_makedirs(dest_dir)

    if log_fn:
        log_fn(
            f"[INFO] 📦 Iniciando unificação de {len(zip_paths)} arquivo(s) ZIP "
            f"para [{ie_formatada(ie)} - {display_name}] | Operação: {operacao} "
            f"| Período: {data_inicial} a {data_final}"
        )

    data_ini_safe = (data_inicial or "").replace("/", ".")
    data_fim_safe = (data_final or "").replace("/", ".")
    unified_name = f"Unificado - {operacao} - {data_ini_safe} a {data_fim_safe} - {display_name}.zip"
    unified_path = build_safe_file_path(dest_dir, unified_name)

    xml_count = 0
    existing_names = set()

    def _unique_name(original: str) -> str:
        base, ext = os.path.splitext(original)
        candidate = original
        idx = 1
        while candidate in existing_names:
            candidate = f"{base}({idx}){ext}"
            idx += 1
        existing_names.add(candidate)
        return candidate

    try:
        # Cria o ZIP final e já vai escrevendo os XMLs direto nele
        with zipfile.ZipFile(fs_path(unified_path), "w", zipfile.ZIP_DEFLATED) as unified_zip:
            for z_path in zip_paths:
                if not path_isfile(z_path):
                    continue
                try:
                    with zipfile.ZipFile(fs_path(z_path), "r") as zf:
                        for name in zf.namelist():
                            if name.lower().endswith(".xml"):
                                data = zf.read(name)
                                arcname = name or os.path.basename(name)
                                if arcname in existing_names:
                                    arcname = _unique_name(os.path.basename(arcname) or "nota.xml")
                                else:
                                    existing_names.add(arcname)
                                unified_zip.writestr(arcname, data)
                                xml_count += 1
                except Exception as e:
                    if log_fn:
                        log_fn(
                            f"[ERRO] [{ie_formatada(ie)} - {display_name}] "
                            f"Erro ao ler ZIP '{z_path}': {e}"
                        )

        if log_fn:
            log_fn(
                f"[OK] [{ie_formatada(ie)} - {display_name}] "
                f"ZIP unificado criado: {unified_name} ({xml_count} XMLs)"
            )

        # Remove os ZIPs originais
        for z_path in zip_paths:
            try:
                if path_isfile(z_path):
                    os.remove(fs_path(z_path))
            except Exception:
                pass

        return unified_path

    except Exception as e:
        if log_fn:
            log_fn(
                f"[ERRO] [{ie_formatada(ie)} - {display_name}] "
                f"Falha na unificação dos ZIPs: {e}"
            )
        return None

# =============================================================================
# Estrutura de resultado / relatório
# =============================================================================

# =============================================================================
# INTEGRAÇÃO SUPABASE
# Funções para processar XMLs e enviar dados agregados para o Supabase
# =============================================================================

def get_company_state_from_cnpj_supabase(cnpj: str) -> Optional[str]:
    """
    Consulta CNPJ na API da Receita Federal (gratuita) e retorna o estado (UF)
    
    Args:
        cnpj: CNPJ sem formatação (apenas números)
    
    Returns:
        Sigla do estado (ex: 'SP', 'RJ') ou None se não encontrar
    """
    if not SUPABASE_AVAILABLE:
        return None
    try:
        cnpj_clean = cnpj_somente_digitos(cnpj)
        if len(cnpj_clean) != 14:
            return None
        
        url = f"https://www.receitaws.com.br/v1/cnpj/{cnpj_clean}"
        response = requests.get(url, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            uf = data.get('uf')
            if uf and len(uf) == 2:
                return uf.upper()
    except Exception:
        pass
    
    return None


def ensure_company_exists_supabase(cnpj: str, company_name: str) -> Optional[str]:
    """
    Garante que a empresa existe no Supabase, cria se não existir
    Atualiza o estado se o CNPJ for consultado
    
    Returns:
        UUID da empresa ou None se Supabase não disponível
    """
    current_job_company = _find_current_job_company(cnpj, company_name)
    if current_job_company:
        return str(current_job_company.get("company_id") or current_job_company.get("id") or "").strip() or None
    if not SUPABASE_AVAILABLE or not supabase_client:
        return None
    
    try:
        cnpj_clean = cnpj_somente_digitos(cnpj)
        if not cnpj_clean or len(cnpj_clean) != 14:
            return None
        
        # Busca empresa pelo documento (schema atual do dashboard)
        result = _supabase_retry_execute(
            lambda: supabase_client.table('companies')
                .select('id')
                .eq('document', cnpj_clean)
                .execute()
        )
        
        if result.data and len(result.data) > 0:
            company_id = result.data[0]['id']

            # Atualiza nome se mudou
            _supabase_retry_execute(
                lambda: supabase_client.table('companies')
                    # `companies` neste projeto não tem `updated_at` (evita PGRST204)
                    .update({'name': company_name})
                    .eq('id', company_id)
                    .execute()
            )
            
            return company_id
        else:
            # Cria nova empresa compatível com o schema atual do dashboard
            new_company = {
                'name': company_name,
                'document': cnpj_clean,
                'active': True,
            }
            
            try:
                # Tenta inserir
                result = _supabase_retry_execute(
                    lambda: supabase_client.table('companies').insert(new_company).execute()
                )
                if result.data and len(result.data) > 0:
                    company_id = result.data[0]['id']
                    return company_id
            except Exception as insert_error:
                # Se der erro de duplicata, busca novamente (pode ter sido criado por outra thread)
                error_str = str(insert_error)
                if 'duplicate key' in error_str.lower() or '23505' in error_str:
                    # Busca novamente
                    result_retry = _supabase_retry_execute(
                        lambda: supabase_client.table('companies')
                            .select('id')
                            .eq('document', cnpj_clean)
                            .execute()
                    )
                    
                    if result_retry.data and len(result_retry.data) > 0:
                        company_id = result_retry.data[0]['id']
                        _supabase_retry_execute(
                            lambda: supabase_client.table('companies')
                                .update({'name': company_name})
                                .eq('id', company_id)
                                .execute()
                        )
                        return company_id
                else:
                    # Outro tipo de erro, re-lança
                    raise
            
            return None
    except Exception as e:
        import traceback
        traceback.print_exc()
        return None


def sincronizar_empresas_supabase(empresas: Dict[str, Dict[str, str]], log_fn: Optional[Callable] = None) -> int:
    """
    Sincroniza todas as empresas do config.json com o Supabase
    Cria/atualiza empresas que têm CNPJ
    
    Returns:
        Número de empresas sincronizadas
    """
    if not SUPABASE_AVAILABLE or not supabase_client:
        if log_fn:
            log_fn("[WARN] ⚠️ Supabase não configurado. Empresas não serão sincronizadas.")
        return 0
    
    count = 0
    for ie, info in empresas.items():
        cnpj = info.get("cnpj", "")
        if not cnpj:
            continue
        
        company_name = info.get("display_name", f"Empresa {ie}")
        company_id = ensure_company_exists_supabase(cnpj, company_name)
        
        if company_id:
            count += 1
            if log_fn:
                log_fn(f"[OK] ✅ Empresa sincronizada: {company_name} (CNPJ: {cnpj})")
        else:
            if log_fn:
                log_fn(f"[WARN] ⚠️ Falha ao sincronizar empresa: {company_name} (CNPJ: {cnpj})")
    
    return count


def extrair_info_nfe_simples_supabase(xml_bytes: bytes) -> Dict:
    """
    Extrai informações básicas do XML para Supabase (compatível com sefaz xml.py)
    Retorna: {model: '55'/'65', vNF: float, dtEmi: date, cnpjEmit: str}
    """
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return {}
    
    model = ''
    v_nf = 0.0
    dt_emi = None
    cnpj_emit = ''
    
    if root is None:
        return {}
    
    # Detecta modelo
    for el in root.iter():
        tag = _local_tag(el.tag)
        if tag == 'mod' and el.text:
            mod = el.text.strip()
            if mod in ('55', '65'):
                model = mod
                break
    
    # Valor total (vNF)
    for el in root.iter():
        tag = _local_tag(el.tag)
        if tag == 'vnf' and el.text:
            try:
                v_nf = float(el.text.strip().replace(',', '.'))
            except Exception:
                pass
            break
    
    # Data de emissão (procura em <ide>)
    ide = None
    for el in root.iter():
        if _local_tag(el.tag) == 'ide':
            ide = el
            break
    
    if ide is not None:
        for el in ide.iter():
            tag = _local_tag(el.tag)
            if tag in ('dhemi', 'demi') and el.text:
                raw = el.text.strip()
                try:
                    if 'T' in raw:
                        raw = raw.split('T')[0]
                    dt_emi = datetime.strptime(raw, '%Y-%m-%d').date()
                except Exception:
                    try:
                        dt_emi = datetime.strptime(raw, '%d/%m/%Y').date()
                    except Exception:
                        pass
                break
    
    # CNPJ emitente
    for el in root.iter():
        tag = _local_tag(el.tag)
        if tag == 'emit':
            for sub in el.iter():
                sub_tag = _local_tag(sub.tag)
                if sub_tag == 'cnpj' and sub.text and not cnpj_emit:
                    cnpj_emit = cnpj_somente_digitos(sub.text)
                    break
            break
    
    return {
        'model': model,
        'vNF': v_nf,
        'dtEmi': dt_emi,
        'cnpjEmit': cnpj_emit,
    }


def processar_xmls_e_agregar_supabase(zip_paths: List[str], empresa_cnpj: str, empresa_nome: str) -> Dict[date, Dict]:
    """
    Processa XMLs dos ZIPs e agrega por data para Supabase
    
    Returns:
        Dict[date, {
            'xml_count': int,
            'nf_count': int,  # modelo 55
            'nfc_count': int, # modelo 65
            'total_amount': float
        }]
    """
    dados_por_data: Dict[date, Dict] = defaultdict(lambda: {
        'xml_count': 0,
        'nf_count': 0,
        'nfc_count': 0,
        'total_amount': 0.0
    })
    
    def processar_zip(zf: zipfile.ZipFile):
        for name in zf.namelist():
            try:
                if not name.lower().endswith('.xml'):
                    continue
                
                xml_bytes = zf.read(name)
                info = extrair_info_nfe_simples_supabase(xml_bytes)
                
                if not info.get('dtEmi'):
                    continue
                
                dt = info['dtEmi']
                model = info.get('model', '')
                vnf = float(info.get('vNF', 0.0))
                
                dados_por_data[dt]['xml_count'] += 1
                if model == '55':
                    dados_por_data[dt]['nf_count'] += 1
                elif model == '65':
                    dados_por_data[dt]['nfc_count'] += 1
                dados_por_data[dt]['total_amount'] += vnf
                
            except Exception:
                continue
    
    # Processa todos os ZIPs
    for zip_path in zip_paths:
        if not path_exists(zip_path):
            continue
        try:
            with zipfile.ZipFile(fs_path(zip_path), 'r') as zf:
                processar_zip(zf)
        except Exception:
            continue
    
    return dict(dados_por_data)


def _build_fiscal_document_storage_key(info: Dict, xml_bytes: bytes) -> str:
    try:
        chave = re.sub(r'[^0-9]', '', str(info.get('chave') or '').strip())
        if len(chave) == 44:
            return chave
    except Exception:
        pass
    return hashlib.sha1(xml_bytes or b'').hexdigest()


def _save_nfe_nfc_xml_for_dashboard(
    *,
    company_folder_name: str,
    storage_key: str,
    xml_bytes: bytes,
    document_date: date,
    operacao: str,
    model: str,
    base_empresas_dir: Optional[str],
    folder_structure: Optional[Dict[str, Any]],
    separar_modelo_xml: bool,
) -> Optional[str]:
    base_root = get_resolved_output_base()
    if not base_root:
        return None

    empresa_folder = safe_folder_name(nome_empresa_sem_ie(company_folder_name or ""))
    if not empresa_folder:
        return None

    destino_dir = montar_caminho_estruturado(
        str(base_empresas_dir or base_root),
        folder_structure,
        empresa_folder,
        operacao=operacao,
        pasta_mes=document_date.strftime("%m%Y"),
        modelo=(model if separar_modelo_xml else None),
        criar=True,
    )
    if not destino_dir:
        return None

    abs_path = pathlib.Path(destino_dir) / f"{storage_key}.xml"
    safe_makedirs(str(abs_path.parent))
    with open(fs_path(str(abs_path)), "wb") as fp:
        fp.write(xml_bytes)

    try:
        rel_path = abs_path.relative_to(base_root)
        return rel_path.as_posix()
    except Exception:
        return None


def sincronizar_fiscal_documents_nfe_nfc(
    company_id: str,
    company_name: str,
    zip_paths: List[str],
    data_inicial: date,
    data_final: date,
    operacao: str,
    base_empresas_dir: Optional[str],
    folder_structure: Optional[Dict[str, Any]],
    separar_modelo_xml: bool,
    log_fn: Optional[Callable] = None
) -> Dict[str, int]:
    """
    Salva os XMLs em disco e registra cada nota em fiscal_documents.
    """
    result = {
        "saved_files": 0,
        "inserted": 0,
        "updated": 0,
        "errors": 0,
    }
    pending_rows: List[Dict[str, Any]] = []
    using_json_runtime = _current_json_job() is not None

    if not company_id:
        return result
    if not using_json_runtime and (not SUPABASE_AVAILABLE or not supabase_client):
        return result

    if not get_resolved_output_base():
        if log_fn:
            log_fn("[WARN] ⚠️ BASE_PATH não configurado; fiscal_documents de NFE/NFC não serão sincronizados.")
        return result

    office_id_for_row = ""
    if supabase_client:
        try:
            cr = _supabase_retry_execute(
                lambda cid=company_id: supabase_client.table("companies")
                .select("office_id")
                .eq("id", cid)
                .limit(1)
                .execute(),
                log_fn=log_fn,
            )
            row0 = (getattr(cr, "data", None) or [None])[0]
            if isinstance(row0, dict):
                office_id_for_row = str(row0.get("office_id") or "").strip()
        except Exception:
            pass
    if not office_id_for_row and callable(log_fn):
        log_fn(
            "[WARN] ⚠️ office_id não resolvido para fiscal_documents (NFE/NFC). "
            "O painel do SaaS pode ficar zerado até o office_id ser gravado nas linhas."
        )

    seen_docs: Set[Tuple[str, str]] = set()

    for zip_path in zip_paths:
        if not path_exists(zip_path):
            continue

        try:
            with zipfile.ZipFile(fs_path(zip_path), "r") as zf:
                for name in zf.namelist():
                    if not name.lower().endswith(".xml"):
                        continue

                    try:
                        xml_bytes = zf.read(name)
                        info = extrair_info_nfe(xml_bytes)
                        model = str(info.get("model") or "").strip()
                        dt_emi = info.get("dtEmi")

                        if model not in ("55", "65") or dt_emi is None:
                            continue
                        if dt_emi < data_inicial or dt_emi > data_final:
                            continue

                        doc_type = "NFE" if model == "55" else "NFC"
                        periodo = dt_emi.strftime("%Y-%m")
                        storage_key = _build_fiscal_document_storage_key(info, xml_bytes)
                        dedupe_key = (doc_type, storage_key)

                        if dedupe_key in seen_docs:
                            continue
                        seen_docs.add(dedupe_key)

                        relative_file_path = _save_nfe_nfc_xml_for_dashboard(
                            company_folder_name=company_name,
                            storage_key=storage_key,
                            xml_bytes=xml_bytes,
                            document_date=dt_emi,
                            operacao=operacao,
                            model=model,
                            base_empresas_dir=base_empresas_dir,
                            folder_structure=folder_structure,
                            separar_modelo_xml=separar_modelo_xml,
                        )
                        if not relative_file_path:
                            result["errors"] += 1
                            continue

                        result["saved_files"] += 1
                        payload = {
                            "company_id": company_id,
                            "type": doc_type,
                            "chave": storage_key,
                            "periodo": periodo,
                            "status": "novo",
                            "document_date": dt_emi.isoformat(),
                            "file_path": relative_file_path,
                        }
                        if office_id_for_row:
                            payload["office_id"] = office_id_for_row

                        if using_json_runtime:
                            pending_rows.append(payload)
                            result["inserted"] += 1
                            continue
                        # Sempre substituir o que já existe: UPDATE por (company_id, file_path); se nenhuma linha afetada, INSERT
                        upd = _supabase_retry_execute(
                            lambda p=payload, rfpath=relative_file_path, cid=company_id: supabase_client.table("fiscal_documents")
                            .update(p)
                            .eq("company_id", cid)
                            .eq("file_path", rfpath)
                            .execute(),
                            log_fn=log_fn,
                        )
                        if upd.data and len(upd.data) > 0:
                            result["updated"] += 1
                        else:
                            try:
                                _supabase_retry_execute(
                                    lambda p=payload: supabase_client.table("fiscal_documents").insert(p).execute(),
                                    log_fn=log_fn,
                                )
                                result["inserted"] += 1
                            except Exception as insert_err:
                                err_str = str(getattr(insert_err, "message", insert_err) or insert_err)
                                if "23505" in err_str or "duplicate key" in err_str.lower() or "company_file_path_key" in err_str:
                                    _supabase_retry_execute(
                                        lambda p=payload, rfpath=relative_file_path, cid=company_id: supabase_client.table("fiscal_documents")
                                        .update(p)
                                        .eq("company_id", cid)
                                        .eq("file_path", rfpath)
                                        .execute(),
                                        log_fn=log_fn,
                                    )
                                    result["updated"] += 1
                                else:
                                    raise
                    except Exception as doc_error:
                        result["errors"] += 1
                        if log_fn:
                            log_fn(f"[WARN] ⚠️ Falha ao sincronizar XML '{name}' no dashboard: {doc_error}")
        except Exception as zip_error:
            result["errors"] += 1
            if log_fn:
                log_fn(f"[WARN] ⚠️ Falha ao abrir ZIP para sincronizar fiscal_documents: {zip_error}")

    if using_json_runtime and pending_rows:
        _append_result_operation(
            {
                "kind": "upsert_rows",
                "table": "fiscal_documents",
                "on_conflict": "company_id,file_path",
                "rows": pending_rows,
            }
        )

    return result


def enviar_dados_supabase_smart_upsert(
    empresa_cnpj: str,
    empresa_nome: str,
    zip_paths: List[str],
    data_inicial: date,
    data_final: date,
    operacao: str,
    base_empresas_dir: Optional[str],
    folder_structure: Optional[Dict[str, Any]],
    separar_modelo_xml: bool,
    log_fn: Optional[Callable] = None
) -> bool:
    """
    Processa XMLs, agrega por data e envia para Supabase com lógica inteligente:
    
    1. Se o período for exatamente o mesmo: SUBSTITUI todos os dados do período
    2. Se houver sobreposição parcial: REMOVE o sobreposto e ADICIONA o novo
    3. Se for período completamente novo: ADICIONA normalmente
    
    Args:
        empresa_cnpj: CNPJ da empresa
        empresa_nome: Nome da empresa
        zip_paths: Lista de caminhos para ZIPs com XMLs
        data_inicial: Data inicial do período processado
        data_final: Data final do período processado
        log_fn: Função de log opcional
    
    Returns:
        True se sucesso, False caso contrário
    """
    if not SUPABASE_AVAILABLE or not supabase_client:
        if log_fn:
            log_fn(f"[WARN] ⚠️ Supabase não configurado. Dados não serão enviados.")
        return False
    
    # Série diária (uma linha por data)
    try:
        if log_fn:
            log_fn(f"[INFO] 🔄 Processando XMLs e enviando para Supabase...")
        
        # Garante que empresa existe
        company_id = ensure_company_exists_supabase(empresa_cnpj, empresa_nome)
        
        if not company_id:
            if log_fn:
                log_fn(f"[ERRO] ❌ Não foi possível criar/encontrar empresa no Supabase")
            return False
        
        if log_fn:
            log_fn(f"[OK] ✅ Empresa {empresa_nome} encontrada/criada no Supabase (ID: {company_id})")

        docs_sync_result = sincronizar_fiscal_documents_nfe_nfc(
            company_id=company_id,
            company_name=empresa_nome,
            zip_paths=zip_paths,
            data_inicial=data_inicial,
            data_final=data_final,
            operacao=operacao,
            base_empresas_dir=base_empresas_dir,
            folder_structure=folder_structure,
            separar_modelo_xml=separar_modelo_xml,
            log_fn=log_fn,
        )
        if log_fn:
            log_fn(
                "[OK] ✅ fiscal_documents NFE/NFC sincronizado "
                f"(arquivos={docs_sync_result['saved_files']}, "
                f"novos={docs_sync_result['inserted']}, atualizados={docs_sync_result['updated']}, "
                f"erros={docs_sync_result['errors']})"
            )
        
        return bool(
            docs_sync_result["saved_files"]
            or docs_sync_result["inserted"]
            or docs_sync_result["updated"]
        )

        # Processa XMLs e agrega por data
        dados_por_data = processar_xmls_e_agregar_supabase(zip_paths, empresa_cnpj, empresa_nome)
        
        if not dados_por_data:
            if log_fn:
                log_fn(f"[WARN] ⚠️ Nenhum XML válido encontrado nos ZIPs")
            return False
        
        if log_fn:
            log_fn(f"[INFO] 📊 Processados {len(dados_por_data)} dias com dados")
        
        # Verifica dados existentes no período (usando estrutura flexível automation_data)
        existing_data = _supabase_retry_execute(
            lambda: supabase_client.table('automation_data')
            .select('date')
            .eq('company_id', company_id)
            .eq('automation_id', automation_id)
            .gte('date', data_inicial.isoformat())
            .lte('date', data_final.isoformat())
            .execute(),
            log_fn=log_fn,
        )
        
        existing_dates = {datetime.fromisoformat(row['date']).date() for row in existing_data.data} if existing_data.data else set()
        new_dates = set(dados_por_data.keys())
        
        # Lógica inteligente de upsert
        if existing_dates == new_dates:
            # Caso 1: Período idêntico - SUBSTITUI tudo
            if log_fn:
                log_fn(f"[INFO] 🔄 Período idêntico detectado. Substituindo dados de {data_inicial} até {data_final}")
            
            # Remove todos os dados do período
            _supabase_retry_execute(
                lambda: supabase_client.table('automation_data')
                .delete()
                .eq('company_id', company_id)
                .eq('automation_id', automation_id)
                .gte('date', data_inicial.isoformat())
                .lte('date', data_final.isoformat())
                .execute(),
                log_fn=log_fn,
            )
            
            # Insere novos dados (mapeando para estrutura flexível)
            for dt, dados in dados_por_data.items():
                _supabase_retry_execute(
                    lambda dt=dt, dados=dados: supabase_client.table('automation_data').insert(
                        {
                            'company_id': company_id,
                            'automation_id': automation_id,
                            'date': dt.isoformat(),
                            'count_1': dados['xml_count'],      # xml_count → count_1
                            'count_2': dados['nf_count'],       # nf_count → count_2 (modelo 55)
                            'count_3': dados['nfc_count'],     # nfc_count → count_3 (modelo 65)
                            'amount_1': dados['total_amount'], # total_amount → amount_1
                            'metadata': {},
                        }
                    ).execute(),
                    log_fn=log_fn,
                )
            
            if log_fn:
                log_fn(f"[OK] ✅ {len(dados_por_data)} registros substituídos no Supabase")
            
        elif existing_dates and new_dates.intersection(existing_dates):
            # Caso 2: Sobreposição parcial - REMOVE sobreposto, ADICIONA tudo
            if log_fn:
                log_fn(f"[INFO] 🔀 Sobreposição parcial detectada. Unindo períodos...")
            
            overlap_dates = existing_dates.intersection(new_dates)
            
            # Remove dados sobrepostos
            for overlap_date in overlap_dates:
                _supabase_retry_execute(
                    lambda overlap_date=overlap_date: supabase_client.table('automation_data')
                    .delete()
                    .eq('company_id', company_id)
                    .eq('automation_id', automation_id)
                    .eq('date', overlap_date.isoformat())
                    .execute(),
                    log_fn=log_fn,
                )
            
            # Insere todos os dados (novos + não sobrepostos)
            for dt, dados in dados_por_data.items():
                _supabase_retry_execute(
                    lambda dt=dt, dados=dados: supabase_client.table('automation_data').insert(
                        {
                            'company_id': company_id,
                            'automation_id': automation_id,
                            'date': dt.isoformat(),
                            'count_1': dados['xml_count'],
                            'count_2': dados['nf_count'],
                            'count_3': dados['nfc_count'],
                            'amount_1': dados['total_amount'],
                            'metadata': {},
                        }
                    ).execute(),
                    log_fn=log_fn,
                )
            
            if log_fn:
                log_fn(f"[OK] ✅ Dados unidos: {len(overlap_dates)} substituídos, {len(dados_por_data)} inseridos")
            
        else:
            # Caso 3: Período novo - ADICIONA normalmente
            if log_fn:
                log_fn(f"[INFO] ➕ Período novo. Adicionando {len(dados_por_data)} registros...")
            
            for dt, dados in dados_por_data.items():
                _supabase_retry_execute(
                    lambda dt=dt, dados=dados: supabase_client.table('automation_data').insert(
                        {
                            'company_id': company_id,
                            'automation_id': automation_id,
                            'date': dt.isoformat(),
                            'count_1': dados['xml_count'],
                            'count_2': dados['nf_count'],
                            'count_3': dados['nfc_count'],
                            'amount_1': dados['total_amount'],
                            'metadata': {},
                        }
                    ).execute(),
                    log_fn=log_fn,
                )
            
            if log_fn:
                log_fn(f"[OK] ✅ {len(dados_por_data)} registros adicionados ao Supabase")
        
        return True
        
    except Exception as e:
        if log_fn:
            log_fn(f"[ERRO] ❌ Erro ao enviar dados para Supabase: {e}")
        import traceback
        traceback.print_exc()
        return False


# =============================================================================
# FUNÇÕES DE ENVIO SUPABASE - DASHBOARD
# =============================================================================

def enviar_dados_supabase_dashboard(
    empresa_cnpj: str,
    empresa_nome: str,
    daily_data: Dict[date, Dict],
    data_inicial: date,
    data_final: date,
    log_fn: Optional[Callable] = None
) -> bool:
    """
    Envia dados agregados do dashboard para Supabase
    UMA ÚNICA LINHA por empresa com todos os dados consolidados em JSON no metadata
    
    Args:
        empresa_cnpj: CNPJ da empresa
        empresa_nome: Nome da empresa
        daily_data: Dict[date, {xml_count, nf_count, nfc_count, faturamento, despesa, resultado}]
                   Deve conter apenas UMA entrada com os totais do período
        data_inicial: Data inicial do período
        data_final: Data final do período
        log_fn: Função de log opcional
    
    Returns:
        True se sucesso, False caso contrário
    """
    if not SUPABASE_AVAILABLE or not supabase_client:
        if _current_json_job():
            pass
        elif log_fn:
            log_fn("[WARN] ⚠️ Supabase não configurado. Dados não serão enviados.")
        if not _current_json_job():
            return False
    
    # Linha consolidada (uma linha por empresa)
    automation_id = 'xml-sefaz-dashboard'
    
    try:
        if log_fn:
            log_fn(f"[INFO] 🔄 Enviando dados consolidados (uma linha por empresa) para Supabase...")
        
        # Garante que empresa existe
        company_id = ensure_company_exists_supabase(empresa_cnpj, empresa_nome)
        
        if not company_id:
            if log_fn:
                log_fn(f"[ERRO] ❌ Não foi possível criar/encontrar empresa no Supabase")
            return False
        
        if not daily_data:
            if log_fn:
                log_fn(f"[WARN] ⚠️ Nenhum dado para enviar")
            return False
        
        # Pega a única entrada (período agregado)
        periodo_date = list(daily_data.keys())[0]
        dados_periodo = daily_data[periodo_date]
        
        if log_fn:
            log_fn(f"[INFO] 📊 Consolidando dados do período {data_inicial} até {data_final}")
            log_fn(f"[INFO] 📊 Totais: XML={dados_periodo.get('xml_count', 0)}, NF={dados_periodo.get('nf_count', 0)}, NFC={dados_periodo.get('nfc_count', 0)}")
            log_fn(f"[INFO] 📊 Financeiro: Faturamento={dados_periodo.get('faturamento', 0.0):.2f}, Despesa={dados_periodo.get('despesa', 0.0):.2f}, Resultado={dados_periodo.get('resultado', 0.0):.2f}")
        
        if _current_json_job():
            _append_result_operation(
                {
                    "kind": "replace_rows",
                    "table": "automation_data",
                    "filters": {
                        "company_id": company_id,
                        "automation_id": automation_id,
                    },
                    "rows": [
                        {
                            "company_id": company_id,
                            "automation_id": automation_id,
                            "date": data_final.isoformat(),
                            "metadata": metadata_consolidado,
                        }
                    ],
                }
            )
            if log_fn:
                log_fn("[OK] ✅ Dados consolidados enfileirados para ingestão pelo connector.")
            return True

        # Verifica se já existe registro para esta empresa (UMA LINHA POR EMPRESA)
        existing_data = _supabase_retry_execute(
            lambda: supabase_client.table('automation_data')
            .select('id')
            .eq('company_id', company_id)
            .eq('automation_id', automation_id)
            .execute(),
            log_fn=log_fn,
        )
        
        # Prepara metadata consolidado com TODOS os dados em JSON
        metadata_consolidado = {
            'periodo_agregado': {
                'data_inicial': data_inicial.isoformat(),
                'data_final': data_final.isoformat(),
                'xml_count': dados_periodo.get('xml_count', 0),
                'nf_count': dados_periodo.get('nf_count', 0),
                'nfc_count': dados_periodo.get('nfc_count', 0),
                'faturamento': dados_periodo.get('faturamento', 0.0),
                'despesa': dados_periodo.get('despesa', 0.0),
                'resultado': dados_periodo.get('resultado', 0.0),
                'total_amount': dados_periodo.get('total_amount', 0.0)
            }
        }
        
        if existing_data.data and len(existing_data.data) > 0:
            # ATUALIZA registro existente (mesma empresa, mesma linha)
            record_id = existing_data.data[0]['id']
            if log_fn:
                log_fn(f"[INFO] 🔄 Atualizando registro consolidado da empresa...")
            
            # Busca metadata existente para preservar dados de evolução se houver
            existing_record = _supabase_retry_execute(
                lambda: supabase_client.table('automation_data')
                .select('metadata')
                .eq('id', record_id)
                .execute(),
                log_fn=log_fn,
            )
            
            if existing_record.data and existing_record.data[0].get('metadata'):
                existing_metadata = existing_record.data[0]['metadata']
                # Preserva dados de evolução se existirem
                if 'evolucao_diaria' in existing_metadata:
                    metadata_consolidado['evolucao_diaria'] = existing_metadata['evolucao_diaria']
            
            _supabase_retry_execute(
                lambda: supabase_client.table('automation_data')
                .update(
                    {
                        'date': data_final.isoformat(),  # Data final do período mais recente
                        'metadata': metadata_consolidado,
                    }
                )
                .eq('id', record_id)
                .execute(),
                log_fn=log_fn,
            )
            
            if log_fn:
                log_fn(f"[OK] ✅ Registro consolidado atualizado no Supabase")
        else:
            # INSERE novo registro (UMA LINHA POR EMPRESA)
            if log_fn:
                log_fn(f"[INFO] ➕ Criando novo registro consolidado para a empresa...")
            
            _supabase_retry_execute(
                lambda: supabase_client.table('automation_data').insert(
                    {
                        'company_id': company_id,
                        'automation_id': automation_id,
                        'date': data_final.isoformat(),  # Data final do período
                        'metadata': metadata_consolidado,
                    }
                ).execute(),
                log_fn=log_fn,
            )
            
            if log_fn:
                log_fn(f"[OK] ✅ Novo registro consolidado criado no Supabase")
        
        return True
        
    except Exception as e:
        if log_fn:
            log_fn(f"[ERRO] ❌ Erro ao enviar dados consolidados para Supabase: {e}")
        import traceback
        traceback.print_exc()
        return False


def enviar_dados_evolucao_supabase(
    empresa_cnpj: str,
    empresa_nome: str,
    serie_labels: List[str],
    serie_values: List[float],
    serie_metric: str,
    daily_in: Dict[date, float],
    daily_out: Dict[date, float],
    daily_xml_count: Dict[date, int],
    daily_nf_count: Dict[date, int],
    daily_nfc_count: Dict[date, int],
    data_inicial: date,
    data_final: date,
    log_fn: Optional[Callable] = None
) -> bool:
    """
    Envia dados diários de evolução para Supabase consolidados em UMA LINHA por empresa
    Todos os dados diários ficam em JSON no metadata
    
    Args:
        empresa_cnpj: CNPJ da empresa
        empresa_nome: Nome da empresa
        serie_labels: Lista de labels (formato '%d/%m')
        serie_values: Lista de valores diários (conforme operação)
        serie_metric: Métrica da série ('Despesa diária', 'Faturamento diário', 'Resultado diário')
        daily_in: Dict[date, float] - valores de entrada por data
        daily_out: Dict[date, float] - valores de saída por data
        daily_xml_count: Dict[date, int] - contagem de XMLs por data
        daily_nf_count: Dict[date, int] - contagem de NFs por data
        daily_nfc_count: Dict[date, int] - contagem de NFCs por data
        data_inicial: Data inicial do período
        data_final: Data final do período
        log_fn: Função de log opcional
    
    Returns:
        True se sucesso, False caso contrário
    """
    if not SUPABASE_AVAILABLE or not supabase_client:
        if _current_json_job():
            pass
        elif log_fn:
            log_fn("[WARN] ⚠️ Supabase não configurado. Dados de evolução não serão enviados.")
        if not _current_json_job():
            return False
    
    # Evolução diária consolidada (uma linha por empresa)
    automation_id = 'xml-sefaz-evolucao'
    
    try:
        if log_fn:
            log_fn(f"[INFO] 🔄 Consolidando dados diários de evolução (uma linha por empresa)...")
        
        # Garante que empresa existe
        company_id = ensure_company_exists_supabase(empresa_cnpj, empresa_nome)
        
        if not company_id:
            if log_fn:
                log_fn(f"[ERRO] ❌ Não foi possível criar/encontrar empresa no Supabase")
            return False
        
        # Usa todas as datas disponíveis (documentos + valores)
        all_dates = sorted(set(daily_in.keys()) | set(daily_out.keys()) | set(daily_xml_count.keys()))
        if not all_dates:
            if log_fn:
                log_fn(f"[WARN] ⚠️ Nenhuma data encontrada para evolução")
            return False
        
        # Mapeia labels para datas (se disponível)
        label_to_date = {}
        if serie_labels:
            date_idx = 0
            for label in serie_labels:
                try:
                    if date_idx < len(all_dates):
                        label_to_date[label] = all_dates[date_idx]
                        date_idx += 1
                except Exception:
                    continue
        
        # Calcula valores acumulados e prepara dados diários em JSON
        acumulado = 0.0
        evolucao_diaria = []
        
        for data_dia in all_dates:
            # Obtém valores financeiros para este dia
            faturamento_dia = float(daily_out.get(data_dia, 0.0) or 0.0)
            despesa_dia = float(daily_in.get(data_dia, 0.0) or 0.0)
            resultado_dia = faturamento_dia - despesa_dia
            
            # Obtém contagens de documentos para este dia
            xml_count = int(daily_xml_count.get(data_dia, 0) or 0)
            nf_count = int(daily_nf_count.get(data_dia, 0) or 0)
            nfc_count = int(daily_nfc_count.get(data_dia, 0) or 0)
            
            # Calcula valor diário e acumulado (se série disponível)
            valor_diario = None
            if serie_labels and serie_values:
                # Tenta encontrar o valor correspondente a esta data
                for i, label in enumerate(serie_labels):
                    if label in label_to_date and label_to_date[label] == data_dia:
                        valor_diario = float(serie_values[i] or 0.0)
                        acumulado += valor_diario
                        break
            
            # Adiciona dados deste dia ao array de evolução
            evolucao_diaria.append({
                'date': data_dia.isoformat(),
                'xml_count': xml_count,
                'nf_count': nf_count,
                'nfc_count': nfc_count,
                'faturamento': faturamento_dia,
                'despesa': despesa_dia,
                'resultado': resultado_dia,
                **({'valor_diario': valor_diario, 'valor_acumulado': acumulado} if valor_diario is not None else {})
            })
        
        if not evolucao_diaria:
            if log_fn:
                log_fn(f"[WARN] ⚠️ Nenhum dado de evolução para enviar")
            return False
        
        if _current_json_job():
            _append_result_operation(
                {
                    "kind": "replace_rows",
                    "table": "automation_data",
                    "filters": {
                        "company_id": company_id,
                        "automation_id": automation_id,
                    },
                    "rows": [
                        {
                            "company_id": company_id,
                            "automation_id": automation_id,
                            "date": data_final.isoformat(),
                            "metadata": metadata_consolidado,
                        }
                    ],
                }
            )
            if log_fn:
                log_fn("[OK] ✅ Dados de evolução enfileirados para ingestão pelo connector.")
            return True

        # Verifica se já existe registro para esta empresa (UMA LINHA POR EMPRESA)
        existing_data = supabase_client.table('automation_data')\
            .select('id, metadata')\
            .eq('company_id', company_id)\
            .eq('automation_id', automation_id)\
            .execute()
        
        # Prepara metadata consolidado
        metadata_consolidado = {
            'evolucao_diaria': evolucao_diaria,
            'serie_metric': serie_metric if serie_metric else '',
            'data_inicial': data_inicial.isoformat(),
            'data_final': data_final.isoformat(),
            'is_evolution_data': True
        }
        
        # Preserva dados de período agregado se existirem
        if existing_data.data and len(existing_data.data) > 0:
            existing_metadata = existing_data.data[0].get('metadata', {})
            if 'periodo_agregado' in existing_metadata:
                metadata_consolidado['periodo_agregado'] = existing_metadata['periodo_agregado']
        
        if existing_data.data and len(existing_data.data) > 0:
            # ATUALIZA registro existente (mesma empresa, mesma linha)
            record_id = existing_data.data[0]['id']
            if log_fn:
                log_fn(f"[INFO] 🔄 Atualizando dados de evolução consolidados...")
            
            supabase_client.table('automation_data')\
                .update({
                    'date': data_final.isoformat(),  # Data final do período mais recente
                    'metadata': metadata_consolidado,
                })\
                .eq('id', record_id)\
                .execute()
            
            if log_fn:
                log_fn(f"[OK] ✅ Dados de evolução consolidados atualizados: {len(evolucao_diaria)} dias")
        else:
            # INSERE novo registro (UMA LINHA POR EMPRESA)
            if log_fn:
                log_fn(f"[INFO] ➕ Criando novo registro consolidado com dados de evolução...")
            
            supabase_client.table('automation_data').insert({
                'company_id': company_id,
                'automation_id': automation_id,
                'date': data_final.isoformat(),
                'metadata': metadata_consolidado
            }).execute()
            
            if log_fn:
                log_fn(f"[OK] ✅ Novo registro consolidado criado: {len(evolucao_diaria)} dias")
        
        return True
        
    except Exception as e:
        if log_fn:
            log_fn(f"[ERRO] ❌ Erro ao enviar dados de evolução consolidados para Supabase: {e}")
        import traceback
        traceback.print_exc()
        return False


# =============================================================================
# CLASSES E DATACLASSES
# =============================================================================

@dataclass
class ResultadoExecucao:
    ie: str
    display_name: str
    operacao: str
    data_inicial: str
    data_final: str
    success: bool
    tentativas: int
    alert_type: Optional[str] = None
    alert_class: Optional[str] = None
    alert_text: str = ""
    motivo: str = ""

    def resumo(self) -> str:
        status = "SUCESSO" if self.success else "ERRO"
        base = f"{self.display_name} - {ie_formatada(self.ie)} | {self.operacao} | {self.data_inicial} a {self.data_final} | {status}"
        if not self.success:
            extra = f" | ({self.alert_class or self.alert_type or ''}) {self.alert_text or self.motivo}"
            return base + extra
        return base


# =============================================================================
# GERAÇÃO DE RELATÓRIO PDF
# =============================================================================
def gerar_relatorio_pdf(resultados: List[ResultadoExecucao],
                        empresas_data: Dict[str, Dict[str, str]],
                        companies_daily_flags: Dict[str, bool],
                        daily_interval: Optional[int],
                        pasta_relatorios: str) -> Optional[str]:
    """
    Gera um PDF consolidando os resultados por empresa e retorna o caminho do arquivo,
    com:
      - Cabeçalho em todas as páginas
      - Legenda colorida (verde / cinza / vermelho)
      - Cartão por empresa com borda suave e linhas horizontais só dentro do cartão
      - 'Modo diário' sempre abaixo do fundo do nome da empresa
      - Cores:
          ✔ / sucesso  -> verde
          ○ / sem mov. -> cinza
          ✖ / erro     -> vermelho
          'XML baixado (com movimento)' -> texto preto
    """
    if not pasta_relatorios or not resultados:
        return None

    os.makedirs(pasta_relatorios, exist_ok=True)
    nome_arquivo = f"Relatorio_SEFAZ_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.pdf"
    caminho_pdf = os.path.join(pasta_relatorios, nome_arquivo)

    writer = QPdfWriter(caminho_pdf)
    writer.setPageSize(QPageSize(QPageSize.A4))
    writer.setResolution(96)
    try:
        writer.setPageMargins(QMarginsF(0, 0, 0, 0))
    except Exception:
        pass

    painter = QPainter(writer)

    def load_qimage_from_path(path: Optional[str]) -> Optional[QImage]:
        if not path:
            return None
        img = QImage(path)
        return img if not img.isNull() else None

    logo_path = resolve_image_path("logo.png")
    success_icon_img = load_qimage_from_path(resolve_image_path("selecionar.png"))
    sem_mov_icon_img = load_qimage_from_path(resolve_image_path("sem-movimento.png"))
    error_icon_img = load_qimage_from_path(
        resolve_image_path("erro.png") or resolve_image_path("desmarcar-cancelar.png")
    )
    company_icon_img = load_qimage_from_path(resolve_image_path("predio.png"))

    def create_page_background(width: int, height: int, logo_file: Optional[str]) -> Optional[QImage]:
        try:
            if not logo_file:
                raise FileNotFoundError("logo ausente")
            base = Image.new("RGBA", (width, height), (255, 255, 255, 255))
            draw = ImageDraw.Draw(base)
            top = (245, 249, 255)
            bottom = (224, 236, 255)
            for yy in range(height):
                ratio = yy / float(max(1, height - 1))
                r = int(top[0] * (1 - ratio) + bottom[0] * ratio)
                g = int(top[1] * (1 - ratio) + bottom[1] * ratio)
                b = int(top[2] * (1 - ratio) + bottom[2] * ratio)
                draw.line([(0, yy), (width, yy)], fill=(r, g, b, 255))

            logo_img = Image.open(logo_file).convert("RGBA")
            # escala mais contida para a marca d'água (próxima ao tamanho usado em certidoes_bot)
            scale = min(width * 0.95 / logo_img.width, height * 0.95 / logo_img.height, 1.4)
            if scale <= 0:
                raise ValueError("escala inválida")
            lw = max(1, int(logo_img.width * scale))
            lh = max(1, int(logo_img.height * scale))
            logo_img = logo_img.resize((lw, lh), Image.LANCZOS)
            alpha = logo_img.split()[-1].point(lambda a: int(a * 0.55))
            cx, cy = lw / 2.0, lh / 2.0
            max_r = math.hypot(lw, lh) / 2.0 if lw and lh else 1.0
            inner_r = max_r * 0.82
            outer_r = max_r * 0.97
            mask = Image.new("L", (lw, lh), 0)
            pix = mask.load()
            for y in range(lh):
                dy = (y + 0.5) - cy
                for x in range(lw):
                    r0 = math.hypot((x + 0.5) - cx, dy)
                    if r0 <= inner_r:
                        val = 255
                    elif r0 >= outer_r:
                        val = 0
                    else:
                        t = (r0 - inner_r) / (outer_r - inner_r)
                        val = int(255 * (1 - t * t))
                    pix[x, y] = val
            grad_x = Image.new("L", (lw, 1), 0)
            gx = grad_x.load()
            edge_start = 0.64
            edge_end = 0.95
            half = max(1.0, cx)
            for x in range(lw):
                t = abs((x + 0.5 - cx) / half)
                t = min(1.0, t)
                if t <= edge_start:
                    fall = 1.0
                elif t >= edge_end:
                    fall = 0.0
                else:
                    tt = (t - edge_start) / (edge_end - edge_start)
                    fall = 1.0 - (tt ** 1.4)
                gx[x, 0] = int(255 * fall)
            grad_x = grad_x.resize((lw, lh))
            grad_y = Image.new("L", (1, lh), 0)
            gy = grad_y.load()
            edge_start_y = 0.6
            edge_end_y = 0.94
            for y in range(lh):
                t = abs((y + 0.5 - cy) / half)
                t = min(1.0, t)
                if t <= edge_start_y:
                    fall = 1.0
                elif t >= edge_end_y:
                    fall = 0.0
                else:
                    tt = (t - edge_start_y) / (edge_end_y - edge_start_y)
                    fall = 1.0 - (tt ** 1.4)
                gy[0, y] = int(255 * fall)
            grad_y = grad_y.resize((lw, lh))
            alpha = ImageChops.multiply(alpha, mask)
            alpha = ImageChops.multiply(alpha, grad_x)
            alpha = ImageChops.multiply(alpha, grad_y)
            alpha = alpha.filter(ImageFilter.GaussianBlur(max(10, int(min(lw, lh) * 0.03))))
            logo_img.putalpha(alpha)
            overlay = Image.new("RGBA", base.size, (0, 0, 0, 0))
            overlay.paste(logo_img, ((width - lw) // 2, (height - lh) // 2), logo_img)
            composed = Image.alpha_composite(base.convert("RGBA"), overlay)
            return QImage(ImageQt.ImageQt(composed))
        except Exception:
            return None

    page_background = create_page_background(writer.width(), writer.height(), logo_path)

    def desenhar_background():
        painter.save()
        painter.fillRect(0, 0, writer.width(), writer.height(), Qt.white)
        if page_background and not page_background.isNull():
            target = QRectF(0, 0, float(writer.width()), float(writer.height()))
            painter.drawImage(target, page_background)
        painter.restore()

    def draw_icon_only(x_pos: float, y_pos: float, img: Optional[QImage], size: int = 14) -> float:
        if img is None or img.isNull():
            return x_pos
        metrics = painter.fontMetrics()
        text_top = y_pos - metrics.ascent()
        icon_top = text_top + max(0, (metrics.height() - size) // 2)
        scaled = img.scaled(size, size, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        painter.drawImage(int(x_pos), int(icon_top), scaled)
        return x_pos + size + 6

    def draw_icon_with_text(
        x_pos: float,
        y_pos: float,
        img: Optional[QImage],
        text: str,
        color: QColor,
        size: int = 14,
        gap_after: int = 26,
    ) -> float:
        painter.setPen(color)
        next_x = draw_icon_only(x_pos, y_pos, img, size)
        painter.drawText(int(next_x), int(y_pos), text)
        return next_x + painter.fontMetrics().horizontalAdvance(text) + gap_after

    # ---------------- CORES ----------------
    color_text = QColor(44, 62, 80)
    color_subtitle = QColor(52, 73, 94)
    color_header_line = QColor(189, 195, 199)
    color_card_border = QColor(210, 214, 217)
    color_card_bg = QColor(245, 247, 250)

    color_success = QColor(46, 204, 113)   # verde
    color_sem_mov = QColor(127, 140, 141)  # cinza
    color_error = QColor(231, 76, 60)      # vermelho

    # ---------------- FONTES ----------------
    font_title = QFont("Helvetica", 16, QFont.Bold)
    font_subtitle = QFont("Helvetica", 11, QFont.Bold)
    font_header = QFont("Helvetica", 9, QFont.Bold)
    font_body = QFont("Helvetica", 9)
    font_small = QFont("Helvetica", 8)

    margin_x = 40
    margin_y = 40
    line_height = 18
    y = margin_y

    total_empresas = len({r.ie for r in resultados})
    total_intervalos = len(resultados)
    total_sucesso = sum(1 for r in resultados if r.success)
    total_sem_mov = sum(1 for r in resultados if r.alert_class == "SEM_RESULTADO")
    total_erro = max(0, total_intervalos - total_sucesso - total_sem_mov)
    gerado_em = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    header_info = f"Gerado em {gerado_em} | Empresas: {total_empresas} | Intervalos: {total_intervalos}"

    table_width = writer.width() - margin_x * 2
    card_left = margin_x
    card_right = margin_x + table_width

    def desenhar_cabecalho():
        nonlocal y
        desenhar_background()
        y = margin_y

        # Título
        painter.setFont(font_title)
        painter.setPen(color_text)
        painter.drawText(margin_x, y, "Relatório de Execução - Automação SEFAZ-GO")
        y += line_height + 4

        # Subtitulo (info)
        painter.setFont(font_subtitle)
        painter.setPen(color_subtitle)
        painter.drawText(margin_x, y, header_info)
        y += line_height

        # Linha horizontal do cabeçalho
        painter.setPen(QPen(color_header_line, 0.8))
        painter.drawLine(margin_x, y, card_right, y)
        y += 12

        # Legenda colorida com ícones PNG
        painter.setFont(font_small)
        base_y = y + 6  # levemente mais abaixo

        x = margin_x
        x = draw_icon_with_text(x, base_y, success_icon_img, "Sucesso (XML baixado)", color_success, 14, 32)
        x = draw_icon_with_text(x, base_y, sem_mov_icon_img, "Sem movimento", color_sem_mov, 14, 30)
        draw_icon_with_text(x, base_y, error_icon_img, "Erro", color_error, 15, 0)

        y += int(line_height * 1.6)

    def desenhar_dashboard():
        """Card de percentuais gerais antes da listagem das empresas."""
        nonlocal y
        if total_intervalos <= 0:
            return

        card_h = 120
        card_pad = 14
        card_radius = 10
        card_rect = (card_left, y, card_right - card_left, card_h)

        painter.save()
        painter.setPen(QPen(QColor(210, 214, 217), 0.8))
        painter.setBrush(QColor(246, 248, 252))
        painter.drawRoundedRect(*card_rect, card_radius, card_radius)
        painter.restore()

        # Título
        painter.setFont(font_subtitle)
        painter.setPen(color_text)
        painter.drawText(card_left + card_pad, y + 28, "Visão geral da execução")

        # Percentuais
        def _pct(val: int) -> float:
            return round((val / total_intervalos) * 100, 1) if total_intervalos else 0.0

        metrics = [
            ("Movimento", success_icon_img, color_success, total_sucesso, _pct(total_sucesso)),
            ("Sem movimento", sem_mov_icon_img, color_sem_mov, total_sem_mov, _pct(total_sem_mov)),
            ("Erro", error_icon_img, color_error, total_erro, _pct(total_erro)),
        ]

        bar_y = y + 52
        bar_h = 18
        bar_w = card_right - card_left - (card_pad * 2)
        bar_x = card_left + card_pad

        # Barra empilhada
        painter.save()
        painter.setPen(Qt.NoPen)
        x_cursor = bar_x
        for _, _, color, count, _ in metrics:
            if count <= 0:
                continue
            w = int(bar_w * (count / total_intervalos))
            painter.setBrush(color)
            painter.drawRoundedRect(x_cursor, bar_y, w, bar_h, 6, 6)
            x_cursor += w
        painter.restore()

        # Labels individuais
        painter.setFont(font_small)
        label_y = bar_y + bar_h + 18
        x_label = card_left + card_pad
        gap = 34
        for label, icon_img, color, count, pct in metrics:
            painter.setPen(color)
            x_next = draw_icon_only(x_label, label_y, icon_img, size=14)
            txt = f"{label}: {count} ({pct:.1f}%)"
            painter.drawText(x_next, label_y, txt)
            x_label = x_next + painter.fontMetrics().horizontalAdvance(txt) + gap

        y += card_h + 16

    def ensure_space(lines_needed: int):
        """Garante espaço para o cartão da empresa; se não tiver, vai pra nova página."""
        nonlocal y
        needed = lines_needed * line_height + 40
        if y + needed > writer.height() - margin_y:
            writer.newPage()
            desenhar_cabecalho()

    # Cabeçalho da primeira página
    desenhar_cabecalho()

    # Dashboard geral antes dos cartões
    desenhar_dashboard()

    # Agrupa por IE
    grouped: Dict[str, List[ResultadoExecucao]] = defaultdict(list)
    for r in resultados:
        grouped[r.ie].append(r)

    ies_ordenadas = sorted(
        grouped.keys(),
        key=lambda ie: empresas_data.get(ie, {}).get("display_name", ie).lower()
    )

    for ie in ies_ordenadas:
        res_list = grouped[ie]
        display_name = empresas_data.get(ie, {}).get("display_name", ie)
        diario_flag = companies_daily_flags.get(ie, False) if companies_daily_flags else False

        # Quantidade aproximada de linhas pra esse cartão (pra não quebrar feio entre páginas)
        linhas_cartao = (
            1  # título
            + (1 if diario_flag and daily_interval else 0)
            + 1  # cabeçalho da tabela
            + len(res_list)  # linhas da tabela
            + 1  # resumo
            + 1  # respiro
        )
        ensure_space(linhas_cartao)

        # ---------------- CARTÃO DA EMPRESA ----------------
        card_top = y
        card_height = linhas_cartao * line_height + 20

        # Fundo do cartão
        painter.save()
        painter.setPen(QPen(color_card_border, 0.8))
        painter.setBrush(color_card_bg)
        painter.drawRoundedRect(card_left, card_top, card_right - card_left, card_height, 6, 6)
        painter.restore()

        # Faixa mais escura só pro nome da empresa
        header_bar_height = line_height + 10
        painter.save()
        painter.setPen(Qt.NoPen)
        painter.setBrush(QColor(52, 73, 94))
        painter.drawRoundedRect(card_left, card_top, card_right - card_left, header_bar_height, 6, 6)
        painter.restore()

        # Texto do nome da empresa dentro da faixa
        painter.setFont(font_subtitle)
        painter.setPen(Qt.white)
        titulo_emp = f"{display_name}  (IE {ie_formatada(ie)})"
        titulo_y = card_top + header_bar_height - 6
        titulo_x = draw_icon_only(card_left + 10, titulo_y, company_icon_img, size=18)
        painter.drawText(titulo_x, titulo_y, titulo_emp)

        # y agora DESCE um pouco pra ficar claramente abaixo do fundo do nome
        y = card_top + header_bar_height + 10

        # Modo diário (se tiver) – sempre abaixo do fundo do nome
        painter.setFont(font_small)
        painter.setPen(color_subtitle)
        if diario_flag and daily_interval:
            painter.drawText(
                card_left + 15,
                y,
                f"Modo diário: blocos de {daily_interval} dia(s)."
            )
            y += line_height

        # Cabeçalho da tabela
        painter.setFont(font_header)
        painter.setPen(color_subtitle)

        col_op = card_left + 15
        col_periodo = card_left + 120
        col_resultado = card_left + 270

        # 🔽 ajuste fino de altura dos títulos
        header_offset = 4  # valor positivo = desce | negativo = sobe
        y += header_offset

        painter.drawText(col_op, y, "Operação")
        painter.drawText(col_periodo, y, "Período")
        painter.drawText(col_resultado, y, "Resultado")

        # deixa a linha mais centralizada entre o cabeçalho e o primeiro registro
        y += int(line_height * 0.98)  # metade do line_height
        painter.setPen(QPen(color_header_line, 0.8))
        painter.drawLine(card_left + 10, y, card_right - 10, y)
        y += int(line_height * 0.98)

        # Corpo da tabela
        painter.setFont(font_body)

        total_sucesso = 0
        total_sem_mov = 0
        total_erro = 0

        for r in res_list:
            # Se estiver estourando a página, vai pra próxima e redesenha o cabeçalho
            if y > writer.height() - margin_y - 40:
                writer.newPage()
                desenhar_cabecalho()
                # Recomeça o cartão na nova página
                card_top = y
                painter.save()
                painter.setPen(QPen(color_card_border, 0.8))
                painter.setBrush(color_card_bg)
                painter.drawRoundedRect(card_left, card_top, card_right - card_left, card_height, 6, 6)
                painter.restore()

                painter.save()
                painter.setPen(Qt.NoPen)
                painter.setBrush(QColor(52, 73, 94))
                painter.drawRoundedRect(card_left, card_top, card_right - card_left, header_bar_height, 6, 6)
                painter.restore()

                painter.setFont(font_subtitle)
                painter.setPen(Qt.white)
                titulo_y = card_top + header_bar_height - 6
                titulo_x = draw_icon_only(card_left + 10, titulo_y, company_icon_img, size=18)
                painter.drawText(titulo_x, titulo_y, titulo_emp)
                y = card_top + header_bar_height + 10

            periodo_txt = f"{r.data_inicial} a {r.data_final}"

            if r.success:
                status_tipo = "SUCESSO"
                total_sucesso += 1
            elif r.alert_class == "SEM_RESULTADO":
                status_tipo = "SEM_MOV"
                total_sem_mov += 1
            else:
                status_tipo = "ERRO"
                total_erro += 1

            if status_tipo == "SUCESSO":
                painter.setPen(color_text)
                painter.drawText(col_op, y, r.operacao)
                painter.drawText(col_periodo, y, periodo_txt)

                x_res = draw_icon_only(col_resultado, y, success_icon_img, size=14)
                painter.setPen(Qt.black)
                painter.drawText(x_res, y, "XML baixado (com movimento)")

            elif status_tipo == "SEM_MOV":
                painter.setPen(color_text)
                painter.drawText(col_op, y, r.operacao)
                painter.drawText(col_periodo, y, periodo_txt)

                x_res = draw_icon_only(col_resultado, y, sem_mov_icon_img, size=14)
                painter.setPen(color_sem_mov)
                painter.drawText(x_res, y, "Sem movimento")

            else:  # ERRO
                painter.setPen(color_text)
                painter.drawText(col_op, y, r.operacao)
                painter.drawText(col_periodo, y, periodo_txt)

                txt = (r.alert_text or r.motivo).strip()
                if not txt:
                    txt = "Erro não especificado"
                max_len = 90
                if len(txt) > max_len:
                    txt = txt[: max_len - 3] + "..."

                painter.setPen(color_error)
                x_res = draw_icon_only(col_resultado, y, error_icon_img, size=14)
                painter.drawText(x_res, y, txt)

            y += line_height

        # Resumo por empresa – com cores também
        y += 2
        painter.setFont(font_small)

        painter.setPen(color_subtitle)
        painter.drawText(col_op, y, "Resumo:")

        x_resumo = col_op + 70

        resumo_txt = f"{total_sucesso} sucesso(s)"
        painter.setPen(color_success)
        x_resumo = draw_icon_only(x_resumo, y, success_icon_img, size=13)
        painter.drawText(x_resumo, y, resumo_txt)
        x_resumo += painter.fontMetrics().horizontalAdvance(resumo_txt) + 30

        resumo_txt = f"{total_sem_mov} sem movimento"
        painter.setPen(color_sem_mov)
        x_resumo = draw_icon_only(x_resumo, y, sem_mov_icon_img, size=13)
        painter.drawText(x_resumo, y, resumo_txt)
        x_resumo += painter.fontMetrics().horizontalAdvance(resumo_txt) + 30

        resumo_txt = f"{total_erro} erro(s)"
        painter.setPen(color_error)
        x_resumo = draw_icon_only(x_resumo, y, error_icon_img, size=13)
        painter.drawText(x_resumo, y, resumo_txt)

        y += int(line_height * 1.7)

    painter.end()
    return caminho_pdf

# =============================================================================
# Log com marca d'água (mesmo estilo do certidoes_bot)
# =============================================================================
class WatermarkLog(QFrame):
    def __init__(self, image_path: str, height=260):
        super().__init__()
        self.setStyleSheet("border:2px solid #34495E; border-radius:4px;")
        self.setFixedHeight(height)

        def _soften(img: Image.Image) -> Image.Image:
            if img.mode != "RGBA":
                img = img.convert("RGBA")
            w, h = img.size
            alpha = img.split()[-1].point(lambda p: int(p * 0.22))
            cx, cy = w / 2.0, h / 2.0
            half_x = max(1.0, cx)
            grad_x = Image.new("L", (w, 1), 0)
            gx = grad_x.load()
            for x in range(w):
                t = abs((x + 0.5 - cx) / half_x)
                t = min(1.0, t)
                fall = 1.0 - (t ** 1.8)
                gx[x, 0] = int(255 * fall)
            grad_x = grad_x.resize((w, h))
            alpha = ImageChops.multiply(alpha, grad_x)
            alpha = alpha.filter(ImageFilter.GaussianBlur(max(8, int(min(w, h) * 0.05))))
            img.putalpha(alpha)
            return img

        try:
            img = Image.open(image_path)
            img = _soften(img)
            self.pixmap = QPixmap.fromImage(ImageQt.ImageQt(img))
        except Exception:
            self.pixmap = QPixmap()

        self.text = QPlainTextEdit(self)
        self.text.setReadOnly(True)
        self.text.setStyleSheet(
            "QPlainTextEdit{background:transparent;color:#E8F4FF;font:10pt Verdana;padding:5px;}"
        )
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.addWidget(self.text)

    def paintEvent(self, e):
        super().paintEvent(e)
        if getattr(self, "pixmap", None) and not self.pixmap.isNull():
            p = QPainter(self)
            sz = self.size()
            target_w = int(sz.width() * 1.35)
            target_h = int(sz.height() * 1.35)
            scaled = self.pixmap.scaled(
                target_w, target_h, Qt.KeepAspectRatio, Qt.SmoothTransformation
            )
            x = (sz.width() - scaled.width()) // 2
            y = (sz.height() - scaled.height()) // 2
            p.drawPixmap(x, y, scaled)

# =============================================================================
# Widget de Empresa (interface)
# =============================================================================

class EmpresaItem(QWidget):
    def __init__(self, ie, display_name, parent=None):
        super().__init__(parent)
        # sempre guarda IE interna só com dígitos
        self.ie = ie_somente_digitos(ie)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(2, 2, 2, 2)
        layout.setSpacing(2)  # deixa bem “enfileirado”

        # exibe "Nome - IE"
        self.checkbox = QCheckBox(f"{display_name} - {ie_formatada(self.ie)}")
        self.daily_checkbox = QCheckBox("Diário")
        layout.addWidget(self.checkbox)
        layout.addWidget(self.daily_checkbox)
        layout.addStretch()
        self.setLayout(layout)

# =============================================================================
# WORKERS E THREADS
# =============================================================================

# =============================================================================
# AutomationWorker - Thread principal de automação (Playwright + download)
# =============================================================================
class AutomationWorker(QThread):
    log_signal = Signal(str)
    metrics_signal = Signal(int, int)

    def __init__(self,
                 empresas: List[str],
                 empresas_data: Dict[str, Dict[str, str]],
                 base_empresas_dir: str,
                 operacoes: List[str],
                 intervalos_meses: List[Tuple[str, str]],
                 data_inicial: str, data_final: str,
                 daily_interval: Optional[int],
                 companies_daily_flags: Dict[str, bool],
                 multiplos_meses: bool,
                 download_option: str,
                 enviar_supabase: bool = True,
                  separar_modelo_xml: bool = True,
                  folder_structure: Optional[Dict[str, Any]] = None,
                  parent=None):
        super().__init__(parent)
        self.empresas = empresas
        self.empresas_data = empresas_data
        self.base_empresas_dir = base_empresas_dir
        self.operacoes = operacoes
        self.intervalos_meses = intervalos_meses
        self.data_inicial = data_inicial
        self.data_final = data_final
        self.daily_interval = daily_interval
        self.companies_daily_flags = companies_daily_flags
        self.multiplos_meses = multiplos_meses
        self.download_option = download_option  # 'docs', 'eventos', 'ambos'
        self.enviar_supabase = bool(enviar_supabase)
        # --- Controle de separação por modelo (55/65) ---
        # IMPORTANTE: a base sempre é a pasta definida em 'Caminhos Padrão' (base_empresas_dir).
        # Quando a opção de separar estiver marcada, as pastas 55/65 são criadas APÓS o mês.
        self.separar_modelo_xml = bool(separar_modelo_xml)
        self.folder_structure = normalizar_estrutura_pastas(folder_structure)

        # Totais para atualizar o dashboard (modelo 55/65)
        self.total_modelo_55 = 0
        self.total_modelo_65 = 0

        self.stop_requested = False
        self.resultados: List[ResultadoExecucao] = []

        # Referências para poder matar o navegador imediatamente ao clicar em "Parar"
        self.pw = None
        self.browser = None
        self.page = None

    def _log(self, msg: str):
        self.log_signal.emit(msg)

    def stop(self):
        self.stop_requested = True
        try:
            if self.page:
                self.page.close()
        except Exception:
            pass
        try:
            if self.browser:
                self._log("[WARN] ⏹️ Parada solicitada pelo usuário. Encerrando navegador imediatamente...")
                self.browser.close()   # ← FECHA O NAVEGADOR NA HORA
        except Exception:
            pass

        try:
            if self.pw:
                self.pw.stop()        # ← ENCERRA O PLAYWRIGHT
        except Exception:
            pass

        self.page = None
        self.browser = None
        self.pw = None

    def run(self):
        pw = browser = page = None
        # Também zera nas propriedades de instância
        self.pw = None
        self.browser = None
        self.page = None

        # ⬇️ NOVO: acumular os ZIPs por empresa + operação + intervalo-base (mes_ini/mes_fim)
        zips_por_emp_op_intervalo = defaultdict(list)  # chave: (ie, operacao, mes_ini, mes_fim)

        try:
            if not self.base_empresas_dir:
                self._log("[ERRO] 💾 Pasta base das empresas não definida. Abortando execução do Worker.")
                return
            os.makedirs(self.base_empresas_dir, exist_ok=True)

            # ✅ Rodar por LOGIN (em lotes) para evitar trocar login a cada empresa.
            # - Cada login executa suas empresas em ordem alfabética (display_name).
            # - Ao trocar para outro login, reinicia o navegador e autentica novamente.
            default_cpf, _ = load_login_portal(preferred_cpf=None)
            default_cpf = cpf_somente_digitos(default_cpf)

            _cpf_not_found_warned: set[str] = set()
            invalid_login_companies: List[Tuple[str, str, str]] = []
            login_groups: Dict[str, List[str]] = {}

            for ie_tmp in (self.empresas or []):
                info_tmp = self.empresas_data.get(ie_tmp, {}) or {}
                cpf_tmp = cpf_somente_digitos(info_tmp.get("login_cpf", ""))
                if cpf_tmp:
                    cpf_check, _ = load_login_portal(preferred_cpf=cpf_tmp, strict_preferred=True)
                    if cpf_somente_digitos(cpf_check) != cpf_tmp:
                        if cpf_tmp not in _cpf_not_found_warned:
                            _cpf_not_found_warned.add(cpf_tmp)
                            self._log(
                                f"[WARN] ⚠️ Login '{cpf_formatado(cpf_tmp)}' não encontrado em login_portal.json. "
                                "Essas empresas serão executadas com o login padrão."
                            )
                        invalid_login_companies.append(
                            (
                                ie_tmp,
                                str(info_tmp.get("display_name") or ie_tmp),
                                cpf_tmp,
                            )
                        )
                        continue

                cpf_group = cpf_tmp or default_cpf
                login_groups.setdefault(cpf_group, []).append(ie_tmp)

            if invalid_login_companies:
                for ie_tmp, display_name_tmp, cpf_tmp in invalid_login_companies:
                    self._log(
                        f"[ERRO] âŒ A empresa '{display_name_tmp}' (IE {ie_formatada(ie_tmp)}) "
                        f"estÃ¡ vinculada ao login {cpf_formatado(cpf_tmp)}, mas esse CPF nÃ£o foi encontrado "
                        "nas credenciais disponÃ­veis do portal."
                    )
                self._log(
                    "[ERRO] âŒ ExecuÃ§Ã£o cancelada para evitar processar empresas com login do contador errado. "
                    "Revise os logins globais do Sefaz Xml e tente novamente."
                )
                return

            def _emp_sort_key(_ie: str):
                try:
                    inf = self.empresas_data.get(_ie, {}) or {}
                    nm = (inf.get("display_name") or str(_ie)).strip().lower()
                except Exception:
                    nm = str(_ie).strip().lower()
                return (nm, str(_ie))

            for _cpf, _ies in login_groups.items():
                _ies.sort(key=_emp_sort_key)

            ordered_cpfs = sorted([c for c in login_groups.keys() if c])
            if default_cpf and default_cpf in ordered_cpfs:
                ordered_cpfs = [default_cpf] + [c for c in ordered_cpfs if c != default_cpf]

            if not ordered_cpfs:
                self._log("[ERRO] ❌ Nenhum login do portal configurado. Cadastre em 'Login do Portal'.")
                return

            # Ordem final das empresas: por login, e dentro do login por ordem alfabética.
            cpf_por_ie: Dict[str, str] = {}
            ies_ordenados: List[str] = []
            for _cpf in ordered_cpfs:
                for _ie in (login_groups.get(_cpf, []) or []):
                    cpf_por_ie[_ie] = _cpf
                ies_ordenados.extend(login_groups.get(_cpf, []) or [])

            current_login_cpf = ""

            def _teardown_current_session(wipe_profile: bool = False, wait_s: float = 3.0) -> None:
                nonlocal pw, browser, page

                # Fecha handles ativos antes de matar processos, para evitar sobra de contexto/aba.
                try:
                    if page:
                        page.close()
                except Exception:
                    pass
                try:
                    if browser:
                        for ctx in list(getattr(browser, "contexts", []) or []):
                            try:
                                ctx.close()
                            except Exception:
                                pass
                        browser.close()
                except Exception:
                    pass
                try:
                    if pw:
                        pw.stop()
                except Exception:
                    pass

                try:
                    _kill_automation_chrome_proc(log_fn=self._log)
                    _kill_automation_chrome_instances()
                except Exception:
                    pass

                try:
                    _stop_proxy_if_running(log_fn=self._log)
                except Exception:
                    pass
                try:
                    _kill_any_mitm_processes(log_fn=self._log)
                except Exception:
                    pass

                try:
                    global SEFAZ_PROXY
                    if (SEFAZ_PROXY or "").strip().lower().startswith("http://127.0.0.1:"):
                        SEFAZ_PROXY = ""
                        os.environ.pop("SEFAZ_PROXY", None)
                except Exception:
                    pass

                # Na troca de login, o perfil precisa ser zerado por completo para não herdar sessão.
                if wipe_profile:
                    try:
                        if os.path.isdir(PROFILE_DIR):
                            shutil.rmtree(PROFILE_DIR, ignore_errors=True)
                            self._log("[INFO] 🧹 Perfil do Chrome removido para iniciar o próximo login do zero.")
                    except Exception as e:
                        self._log(f"[WARN] ⚠️ Falha ao remover perfil do Chrome: {e}")
                else:
                    try:
                        _clear_chrome_profile_session(PROFILE_DIR, log_fn=self._log)
                    except Exception:
                        pass

                try:
                    sleep(float(wait_s or 0))
                except Exception:
                    time.sleep(float(wait_s or 0))

                pw = browser = page = None
                self.pw = None
                self.browser = None
                self.page = None

            def _start_session_for_login(cpf_digits: str):
                nonlocal pw, browser, page, current_login_cpf

                cpf_digits = cpf_somente_digitos(cpf_digits)
                if not cpf_digits:
                    self._log("[ERRO] ❌ CPF de login vazio para iniciar sessão.")
                    return False

                cpf_use, senha_use = load_login_portal(preferred_cpf=cpf_digits, strict_preferred=True)
                if not cpf_use or not senha_use:
                    # Não trata como exceção "não tratada": mostra claramente o que falta no JSON.
                    self._log(
                        f"[ERRO] ❌ Credenciais do portal não encontradas para {cpf_formatado(cpf_digits)}. "
                        f"Verifique se a senha está preenchida em {LOGIN_PATH}."
                    )
                    return False

                # Na troca de login, derruba tudo e apaga o perfil para impedir herança da sessão anterior.
                _teardown_current_session(wipe_profile=bool(current_login_cpf), wait_s=3.0)

                self._log("[INFO] 🔧 Preparando ambiente de automação (Playwright + Chrome via CDP)...")
                try:
                    pw, browser, page = iniciar_chrome_com_cdp()
                except Exception as e:
                    self._log(f"[WARN] ⚠️ Falha ao iniciar Chrome/CDP ({e}). Tentando novamente com perfil limpo...")
                    try:
                        subprocess.run(["taskkill", "/IM", "chrome.exe", "/F"], stdout=subprocess.DEVNULL,
                                       stderr=subprocess.DEVNULL, check=False)
                    except Exception:
                        pass
                    try:
                        if os.path.isdir(PROFILE_DIR):
                            shutil.rmtree(PROFILE_DIR, ignore_errors=True)
                    except Exception:
                        pass
                    pw, browser, page = iniciar_chrome_com_cdp()

                self.pw, self.browser, self.page = pw, browser, page
                self._log("[OK] ✅ Playwright/Chrome inicializados e conectados com sucesso.")

                self._log("[INFO] 🔐 Acessando portal SEFAZ e realizando login...")
                try:
                    login_portal(page, cpf=cpf_use, senha=senha_use, log_fn=self._log)
                except Exception as e:
                    self._log(
                        f"[WARN] ⚠️ Login do portal falhou ({e}). Aguardando 12s e tentando novamente uma vez..."
                    )
                    time.sleep(12)
                    try:
                        login_portal(
                            page, cpf=cpf_use, senha=senha_use, log_fn=self._log
                        )
                    except Exception as e2:
                        self._log(f"[ERRO] ❌ Falha no login do portal SEFAZ: {e2}")
                        return False
                self._log("[OK] ✅ Login no portal SEFAZ concluído.")
                current_login_cpf = _get_current_login_cpf()

                self._log("[INFO] 🔗 Abrindo área de Acesso Restrito...")
                page = ir_para_acesso_restrito(page, log_fn=self._log)
                self.page = page
                self._log("[OK] ✅ Área de Acesso Restrito pronta.")

                self._log("[INFO] 📄 Acessando menu 'Baixar XML NFE' pela primeira vez...")
                abrir_menu_baixar_xml_nfe(page)

                self._log("[INFO] 🔑 Autenticando no NetAccess com usuário/senha configurados...")
                page = ensure_single_tab(page, log_fn=self._log)
                estado = wait_netaccess_or_filtros(page, timeout_ms=8000, log_fn=self._log)
                if estado == "netaccess":
                    login_netaccess(page)
                    self._log("[OK] ✅ Autenticação NetAccess concluída.")
                else:
                    self._log("[INFO] ℹ️ NetAccess não foi solicitado. Prosseguindo.")

                self._log("[INFO] 📄 Garantindo tela 'Baixar XML NFE' após autenticação...")
                abrir_menu_baixar_xml_nfe(page)
                self._log("[INFO] 🟢 Ambiente SEFAZ pronto para iniciar o processamento das empresas.")
                return True

            def _restart_session_same_login(reason: str = "", wait_s: float = 5.0) -> bool:
                """Reinicia o navegador e reautentica mantendo o mesmo login atual.

                Regra solicitada: quando 2 empresas consecutivas falharem em baixar Entrada e Saida,
                fecha o navegador, espera 5s e abre novamente.
                """
                nonlocal pw, browser, page, current_login_cpf

                cpf_digits = cpf_somente_digitos(current_login_cpf or "")
                if not cpf_digits:
                    cpf_digits = cpf_somente_digitos(default_cpf or "")
                if not cpf_digits:
                    self._log("[ERRO] Nao foi possivel reiniciar a sessao: CPF atual vazio e sem CPF padrao.")
                    return False

                cpf_use, senha_use = load_login_portal(preferred_cpf=cpf_digits, strict_preferred=True)
                if not cpf_use or not senha_use:
                    self._log(
                        f"[ERRO] Credenciais nao encontradas para reiniciar sessao no login {cpf_formatado(cpf_digits)}."
                    )
                    return False

                if reason:
                    self._log(f"[WARN] Reiniciando navegador por regra de falhas consecutivas. Motivo: {reason}")
                else:
                    self._log("[WARN] Reiniciando navegador por regra de falhas consecutivas.")

                _teardown_current_session(wipe_profile=True, wait_s=float(wait_s or 5.0))

                # Reabre Chrome e reautentica
                try:
                    pw, browser, page = iniciar_chrome_com_cdp()
                except Exception as e:
                    self._log(f"[ERRO] Falha ao iniciar Chrome/CDP no restart: {e}")
                    return False

                self.pw, self.browser, self.page = pw, browser, page
                try:
                    login_portal(page, cpf=cpf_use, senha=senha_use, log_fn=self._log)
                except Exception as e:
                    self._log(
                        f"[WARN] ⚠️ Login após restart falhou ({e}). Nova tentativa em 12s..."
                    )
                    time.sleep(12)
                    try:
                        login_portal(
                            page, cpf=cpf_use, senha=senha_use, log_fn=self._log
                        )
                    except Exception as e2:
                        self._log(f"[ERRO] Falha ao refazer login do portal no restart: {e2}")
                        return False

                current_login_cpf = _get_current_login_cpf()

                try:
                    page = ir_para_acesso_restrito(page, log_fn=self._log)
                    self.page = page
                except Exception as e:
                    self._log(f"[ERRO] Falha ao abrir Acesso Restrito no restart: {e}")
                    return False

                try:
                    abrir_menu_baixar_xml_nfe(page)
                    page = ensure_single_tab(page, log_fn=self._log)
                    estado = wait_netaccess_or_filtros(page, timeout_ms=8000, log_fn=self._log)
                    if estado == "netaccess":
                        login_netaccess(page)
                    abrir_menu_baixar_xml_nfe(page)
                except Exception as e:
                    self._log(f"[ERRO] Falha ao preparar tela 'Baixar XML NFE' no restart: {e}")
                    return False

                return True

            def _empresa_tem_baixa_entrada_saida(ie_digits: str) -> tuple[bool, bool]:
                """Retorna (tem_entrada, tem_saida) considerando todos os resultados ate agora."""
                ie_digits = ie_somente_digitos(ie_digits or "")
                tem_entrada = False
                tem_saida = False
                try:
                    for rr in (self.resultados or []):
                        if ie_somente_digitos(getattr(rr, "ie", "") or "") != ie_digits:
                            continue
                        if not bool(getattr(rr, "success", False)):
                            continue
                        op = (getattr(rr, "operacao", "") or "").strip().lower()
                        if op.startswith("e"):
                            tem_entrada = True
                        elif op.startswith("s"):
                            tem_saida = True
                        if tem_entrada and tem_saida:
                            break
                except Exception:
                    pass
                return tem_entrada, tem_saida

            def _empresa_sem_download_ignorar_consecutivo(ie_digits: str) -> tuple[bool, str]:
                """
                Retorna se a empresa deve ser ignorada na contagem de consecutivas quando
                nao baixou Entrada/Saida. Isso acontece quando os resultados sao apenas de
                "sem movimento" ou "sem permissao" (nao erro tecnico real).
                """
                ie_digits = ie_somente_digitos(ie_digits or "")
                if not ie_digits:
                    return False, ""

                classes_ignoraveis = {"SEM_RESULTADO", "SEM_PERMISSAO"}
                classes_reais: set[str] = set()
                classes_vistas: set[str] = set()

                try:
                    for rr in (self.resultados or []):
                        if ie_somente_digitos(getattr(rr, "ie", "") or "") != ie_digits:
                            continue
                        op = (getattr(rr, "operacao", "") or "").strip().lower()
                        if not (op.startswith("e") or op.startswith("s")):
                            continue
                        if bool(getattr(rr, "success", False)):
                            return False, ""

                        raw_class = (getattr(rr, "alert_class", "") or "").strip()
                        cls = raw_class.upper()
                        classes_vistas.add(cls or "SEM_CLASSE")
                        if cls not in classes_ignoraveis:
                            classes_reais.add(cls or "SEM_CLASSE")
                except Exception:
                    return False, ""

                if not classes_vistas:
                    return False, ""
                if classes_reais:
                    return False, ""

                return True, ", ".join(sorted(classes_vistas))

            # Inicia com o primeiro login da fila.
            first_login_cpf = cpf_somente_digitos(ordered_cpfs[0] or "")
            self._log(f"[INFO] 🔐 Iniciando execucao pelo login {cpf_formatado(first_login_cpf)}.")
            if not _start_session_for_login(first_login_cpf):
                return

            processa_entrada = any((op or "").strip().lower().startswith("e") for op in (self.operacoes or []))
            processa_saida = any((op or "").strip().lower().startswith("s") for op in (self.operacoes or []))
            consecutivas_sem_entrada_saida = 0
            falhas_sem_dados_enviados_consecutivas = 0

            def _tratar_sem_dados_enviados(contexto: str) -> bool:
                nonlocal page, falhas_sem_dados_enviados_consecutivas

                falhas_sem_dados_enviados_consecutivas += 1
                self._log(
                    f"[WARN] Erro 'Nenhum dado foi enviado' detectado em {contexto} "
                    f"(consecutivas: {falhas_sem_dados_enviados_consecutivas}/3)."
                )

                if falhas_sem_dados_enviados_consecutivas > 2:
                    ok_restart = _restart_session_same_login(
                        reason="Erro 'Nenhum dado foi enviado por nfeweb.sefaz.go.gov.br' repetido",
                        wait_s=5.0,
                    )
                    if not ok_restart:
                        self._log("[FATAL] Nao foi possivel reiniciar a sessao apos repeticao do erro de envio vazio.")
                        return False
                    page = self.page or page
                    falhas_sem_dados_enviados_consecutivas = 0
                else:
                    try:
                        abrir_menu_baixar_xml_nfe(page)
                    except Exception:
                        pass

                sleep(2)
                return True

            for ie in ies_ordenados:
                if self.stop_requested:
                    self._log(
                        "[WARN] ⏹️ Parada solicitada pelo usuário. "
                        "Encerrando processamento das empresas restantes."
                    )
                    break

                cpf_needed = cpf_somente_digitos(cpf_por_ie.get(ie, "") or default_cpf)
                if cpf_needed and cpf_needed != current_login_cpf:
                    self._log(
                        f"[INFO] 🔁 Mudando para o próximo login {cpf_formatado(cpf_needed)} "
                        "(fechando e reabrindo navegador)."
                    )
                    if not _start_session_for_login(cpf_needed):
                        # Se não consegue iniciar sessão do próximo login, interrompe para não processar com login errado.
                        return
                    # Troca de login ja reinicia o navegador; zera o contador para evitar restart redundante.
                    consecutivas_sem_entrada_saida = 0

                info_emp = self.empresas_data.get(ie, {}) or {}

                display_name = info_emp.get("display_name", ie)
                empresa_pasta_nome = nome_empresa_sem_ie(display_name)
                baixou_entrada_empresa = False
                baixou_saida_empresa = False
                nao_contar_consecutivo = False
                # Flag: modo "separar por modelo" ativo
                separar_por_modelo_ativo = self.separar_modelo_xml
                folder_name = safe_folder_name(empresa_pasta_nome)
                caminho_emp = montar_caminho_estruturado(
                    self.base_empresas_dir,
                    self.folder_structure,
                    empresa_pasta_nome,
                    criar=not separar_por_modelo_ativo,
                ) or os.path.join(self.base_empresas_dir, folder_name)

                # Só cria a pasta da empresa na base se NÃO estiver separando por modelo
                if not separar_por_modelo_ativo and caminho_emp:
                    os.makedirs(caminho_emp, exist_ok=True)

                # Cabeçalho estruturado para a empresa (consumido pelo update_log)
                header_payload = "|".join([
                    display_name,
                    ie_formatada(ie),
                    caminho_emp,
                    ", ".join(self.operacoes),
                    self.download_option or "",
                ])
                self._log(f"::EMPRESA_HEADER::{header_payload}")

                # Mensagem complementar mais “limpa”
                self._log(
                    "[INFO] Iniciando fluxo de intervalos para esta empresa."
                )

                # Intervalos mensais/dias
                if self.multiplos_meses:
                    meses = self.intervalos_meses
                else:
                    if self.data_inicial and self.data_final:
                        meses = [(self.data_inicial, self.data_final)]
                    else:
                        meses = gerar_intervalos_mensais("", "")

                if self.multiplos_meses:
                    self._log(
                        f"[INFO] 📅 Modo 'múltiplos meses' ativo. "
                        f"Serão processados {len(meses)} intervalo(s) base para esta empresa."
                    )
                else:
                    self._log(
                        f"[INFO] 📅 Modo de período único. "
                        f"Intervalo base: {meses[0][0]} a {meses[0][1]}"
                    )

                parar_empresa = False

                for (mes_ini, mes_fim) in meses:
                    if self.stop_requested:
                        self._log(
                            "[WARN] ⏹️ Parada solicitada. Encerrando processamento do intervalo atual."
                        )
                        parar_empresa = True
                        break
                    if parar_empresa:
                        break

                    self._log(
                        f"[INFO] 🧩 Intervalo-base para a empresa: {mes_ini} a {mes_fim}"
                    )

                    # Pasta do mês no formato mmYYYY (ex.: 012025)
                    mes_inicio_dt = datetime.strptime(mes_ini, "%d/%m/%Y")
                    pasta_mes = mes_inicio_dt.strftime("%m%Y")

                    download_dirs_por_operacao: Dict[str, str] = {}
                    for op in self.operacoes:
                        caminho_op = montar_caminho_estruturado(
                            self.base_empresas_dir,
                            self.folder_structure,
                            empresa_pasta_nome,
                            operacao=op,
                            pasta_mes=pasta_mes,
                            criar=True,
                        ) or self.base_empresas_dir
                        download_dirs_por_operacao[op] = caminho_op
                        self._log(
                            f"[INFO] 🗂️ Pasta de trabalho para {op} ({pasta_mes}): {caminho_op}"
                        )

                    # se marcado "Diário" e tiver intervalo de dias, quebra o mês
                    if self.companies_daily_flags.get(ie, False) and self.daily_interval:
                        intervalos = gerar_intervalos_diarios(mes_ini, mes_fim, self.daily_interval)
                        self._log(
                            f"[INFO] 🗓️ Empresa marcada como 'Diário' com blocos de "
                            f"{self.daily_interval} dia(s). Intervalo {mes_ini} a {mes_fim} "
                            f"foi quebrado em {len(intervalos)} bloco(s)."
                        )
                    else:
                        intervalos = [(mes_ini, mes_fim)]
                        self._log(
                            f"[INFO] 🗓️ Intervalo sem quebra diária. "
                            f"Será usado o período completo: {mes_ini} a {mes_fim}."
                        )

                    for (dt_ini, dt_fim) in intervalos:
                        if self.stop_requested:
                            self._log(
                                "[WARN] ⏹️ Parada solicitada. Finalizando antes de iniciar novo intervalo."
                            )
                            # registra intervalos restantes deste período como interrompidos
                            for op in self.operacoes:
                                res_abort = ResultadoExecucao(
                                    ie=ie,
                                    display_name=display_name,
                                    operacao=op,
                                    data_inicial=dt_ini,
                                    data_final=dt_fim,
                                    success=False,
                                    tentativas=0,
                                    alert_type="INTERRUPCAO_USUARIO",
                                    alert_class="INTERRUPCAO_USUARIO",
                                    alert_text="",
                                    motivo="Processo interrompido pelo usuário antes da operação."
                                )
                                self.resultados.append(res_abort)
                            parar_empresa = True
                            break

                        for operacao in self.operacoes:
                            if self.stop_requested:
                                self._log(
                                    "[WARN] ⏹️ Parada solicitada. Encerrando processamento da operação atual."
                                )
                                res_abort = ResultadoExecucao(
                                    ie=ie,
                                    display_name=display_name,
                                    operacao=operacao,
                                    data_inicial=dt_ini,
                                    data_final=dt_fim,
                                    success=False,
                                    tentativas=0,
                                    alert_type="INTERRUPCAO_USUARIO",
                                    alert_class="INTERRUPCAO_USUARIO",
                                    alert_text="",
                                    motivo="Processo interrompido pelo usuário durante a operação."
                                )
                                self.resultados.append(res_abort)
                                parar_empresa = True
                                break

                            if parar_empresa:
                                break

                            download_base_dir_mes = download_dirs_por_operacao.get(
                                operacao, self.base_empresas_dir
                            )
                            tipo_nota = "entrada" if operacao.lower().startswith("e") else "saida"
                            self._log(
                                f"[INFO] ▶️ Empresa: {display_name} (IE {ie_formatada(ie)}) | "
                                f"Operação: {operacao} | Período: {dt_ini} a {dt_fim}"
                            )

                            while True:
                                try:
                                    resultado_dict = pesquisar_intervalo(
                                        page,
                                        ie,
                                        dt_ini,
                                        dt_fim,
                                        tipo_nota,
                                        log_fn=self._log,
                                        should_stop=lambda: self.stop_requested,
                                    )
                                except Exception as e:
                                    self._log(
                                        f"[ERRO] 💥 Exceção durante a pesquisa "
                                        f"({display_name} | {operacao} | {dt_ini} a {dt_fim}): {e}"
                                    )
                                    res = ResultadoExecucao(
                                        ie=ie,
                                        display_name=display_name,
                                        operacao=operacao,
                                        data_inicial=dt_ini,
                                        data_final=dt_fim,
                                        success=False,
                                        tentativas=0,
                                        alert_type="EXCECAO",
                                        alert_class="EXCECAO_WORKER",
                                        alert_text="",
                                        motivo=f"Exceção durante a pesquisa: {e}",
                                    )
                                    self.resultados.append(res)

                                    if not self.stop_requested:
                                        try:
                                            abrir_menu_baixar_xml_nfe(page)
                                        except Exception:
                                            pass
                                    break

                                if (resultado_dict.get("alert_class") or "") == "SEM_DADOS_ENVIADOS":
                                    if not _tratar_sem_dados_enviados(
                                        f"{display_name} | {operacao} | {dt_ini} a {dt_fim}"
                                    ):
                                        return
                                    continue

                                falhas_sem_dados_enviados_consecutivas = 0
                                res = ResultadoExecucao(
                                    ie=ie,
                                    display_name=display_name,
                                    operacao=operacao,
                                    data_inicial=dt_ini,
                                    data_final=dt_fim,
                                    success=resultado_dict["success"],
                                    tentativas=resultado_dict["tentativas"],
                                    alert_type=resultado_dict.get("tipo_alerta"),
                                    alert_class=resultado_dict.get("alert_class"),
                                    alert_text=resultado_dict.get("alert_text", ""),
                                    motivo=resultado_dict.get("motivo", ""),
                                )
                                self.resultados.append(res)
                                break

                            if res.success:
                                self._log(
                                    f"[OK] ✅ Pesquisa bem-sucedida para "
                                    f"{operacao} {dt_ini} a {dt_fim} "
                                    f"(tentativa {res.tentativas})."
                                )
                                self._log(
                                    "[INFO] ⏬ Solicitando geração e download do ZIP para este resultado..."
                                )
                                try:
                                    ok_zip_req = solicitar_geracao_zip(
                                        page,
                                        self.download_option,
                                        log_fn=self._log
                                    )
                                except Exception as e:
                                    self._log(f"[ERRO] 💥 Exceção ao solicitar geração do ZIP: {e}")
                                    ok_zip_req = False

                                if not ok_zip_req:
                                    self._log(
                                        "[ERRO] ❌ Falha ao solicitar geração do ZIP "
                                        "para este intervalo."
                                    )
                                    res.success = False
                                    res.alert_class = res.alert_class or "DOWNLOAD_ERRO"
                                    res.motivo = "Erro ao solicitar geração de ZIP."
                                else:
                                    self._log(
                                        "[INFO] 📜 Solicitação de geração de ZIP enviada. "
                                        "Aguardando disponibilidade no Histórico de Downloads."
                                    )
                                    # ALTERADO: salvar diretamente na pasta configurada para o mês
                                    try:
                                        zip_paths = aguardar_e_baixar_arquivos(
                                            page,
                                            download_base_dir_mes,  # ← ANTES era caminho_mes
                                            quantidade=1,
                                            should_stop=lambda: self.stop_requested,
                                            log_fn=self._log,
                                            ie=ie,
                                            data_inicial=dt_ini,
                                            data_final=dt_fim,
                                            tipo=tipo_nota,
                                            download_option=self.download_option,  # NOVO: para poder re-solicitar o ZIP
                                        )
                                    except Exception as e:
                                        self._log(f"[ERRO] 💥 Exceção ao aguardar/baixar ZIP no Histórico: {e}")
                                        zip_paths = []

                                    if not zip_paths:
                                        if self.stop_requested:
                                            self._log(
                                                "[WARN] ⏹️ Parada solicitada durante o aguardo de downloads. "
                                                "ZIP não foi baixado."
                                            )
                                            # Marca explicitamente como interrupção do usuário
                                            res.success = False
                                            res.alert_class = "INTERRUPCAO_USUARIO"
                                            res.motivo = "Processo interrompido pelo usuário durante o download."
                                        else:
                                            self._log(
                                                "[ERRO] ❌ Erro ao aguardar/baixar ZIP no Histórico."
                                            )
                                            # Marca o intervalo como ERRO no resumo
                                            res.success = False
                                            if not res.alert_class:
                                                res.alert_class = "DOWNLOAD_ERRO"
                                            res.motivo = "Erro ao aguardar/baixar ZIP no Histórico."
                                    else:
                                        if self.enviar_supabase and SUPABASE_AVAILABLE:
                                            try:
                                                empresa_cnpj = info_emp.get("cnpj", "")
                                                if empresa_cnpj:
                                                    try:
                                                        dt_sync_ini = datetime.strptime(dt_ini, "%d/%m/%Y").date()
                                                        dt_sync_fim = datetime.strptime(dt_fim, "%d/%m/%Y").date()
                                                        enviar_dados_supabase_smart_upsert(
                                                            empresa_cnpj=empresa_cnpj,
                                                            empresa_nome=display_name,
                                                            zip_paths=zip_paths,
                                                            data_inicial=dt_sync_ini,
                                                            data_final=dt_sync_fim,
                                                            operacao=operacao,
                                                            base_empresas_dir=self.base_empresas_dir,
                                                            folder_structure=self.folder_structure,
                                                            separar_modelo_xml=self.separar_modelo_xml,
                                                            log_fn=self._log
                                                        )
                                                    except ValueError as e:
                                                        self._log(f"[WARN] âš ï¸ Erro ao converter datas do intervalo para Supabase: {e}")
                                                else:
                                                    self._log(f"[WARN] âš ï¸ CNPJ nÃ£o encontrado para empresa {display_name}. Dados nÃ£o enviados ao Supabase.")
                                            except Exception as e:
                                                self._log(f"[ERRO] âŒ Erro ao sincronizar ZIP baixado no Supabase: {e}")

                                        chave = (ie, operacao, mes_ini, mes_fim)
                                        if self.separar_modelo_xml:
                                            # processa e limpa imediatamente em modo separacao
                                            try:
                                                c55, c65 = extrair_e_separar_zips_por_modelo(
                                                    zip_paths=zip_paths,
                                                    empresa_nome=empresa_pasta_nome,
                                                    empresa_ie=ie,
                                                    empresas_map=self.empresas,
                                                    pasta_mes=pasta_mes,
                                                    base55=self.base_empresas_dir,
                                                    base65=self.base_empresas_dir,
                                                    estrutura_pastas=self.folder_structure,
                                                    operacao=operacao,
                                                    data_inicial=mes_ini,
                                                    data_final=mes_fim,
                                                    zipar=True,
                                                    remover_xmls=True,
                                                    deletar_zips_gerados=True,
                                                    log_fn=self._log
                                                )
                                                self.total_modelo_55 += int(c55 or 0)
                                                self.total_modelo_65 += int(c65 or 0)
                                                self.metrics_signal.emit(self.total_modelo_55, self.total_modelo_65)

                                            except Exception as e:
                                                self._log(f"[WARN] ⚠️ Falha ao separar ZIPs imediatamente: {e}")
                                                zips_por_emp_op_intervalo[chave].extend(zip_paths)
                                        else:
                                            # acumula os ZIPs deste bloco diario no intervalo-base (mes)
                                            zips_por_emp_op_intervalo[chave].extend(zip_paths)

                                        # Marca que esta empresa conseguiu baixar pelo menos algo nesta operacao
                                        if operacao.lower().startswith("e"):
                                            baixou_entrada_empresa = True
                                        else:
                                            baixou_saida_empresa = True

                                # volta para a tela de filtro
                                if not self.stop_requested:
                                    self._log(
                                        "[INFO] ↩️ Retornando para a tela de filtros 'Baixar XML NFE' "
                                        "para o proximo intervalo."
                                    )
                                    abrir_menu_baixar_xml_nfe(page)
                            else:
                                self._log(
                                    f"[ERRO] ❌ Falha na pesquisa para este intervalo: {res.resumo()}"
                                )
                                if res.alert_class in ("IE_INVÁLIDA", "IE_OBRIGATÓRIA"):
                                    self._log(
                                        "[WARN] 🚫 IE inválida ou obrigatória. "
                                        "A empresa será pulada para evitar novas tentativas."
                                    )
                                    nao_contar_consecutivo = True
                                    parar_empresa = True
                                    break

                    # AO FINAL DO MÊS
                    if not parar_empresa:
                        for operacao in self.operacoes:
                            chave = (ie, operacao, mes_ini, mes_fim)
                            zip_list = zips_por_emp_op_intervalo.get(chave)
                            if not zip_list:
                                continue
                            
                            if self.separar_modelo_xml:
                                # ✅ MODO SEPARAÇÃO: salva SOMENTE nas pastas dos modelos
                                c55, c65 = extrair_e_separar_zips_por_modelo(
                                    zip_paths=zip_list,
                                    empresa_nome=empresa_pasta_nome,
                                    empresa_ie=ie,
                                    empresas_map=self.empresas,
                                    pasta_mes=pasta_mes,
                                    base55=self.base_empresas_dir,
                                        base65=self.base_empresas_dir,
                                    estrutura_pastas=self.folder_structure,
                                    operacao=operacao,          # para nome do .zip por modelo
                                    data_inicial=mes_ini,       # para nome do .zip por modelo
                                    data_final=mes_fim,         # para nome do .zip por modelo
                                    zipar=True,                 # gera o .zip por modelo
                                    remover_xmls=True,          # não deixa XMLs soltos nos modelos
                                    deletar_zips_gerados=True,
                                    log_fn=self._log
                                )
                                self.total_modelo_55 += int(c55 or 0)
                                self.total_modelo_65 += int(c65 or 0)
                                self.metrics_signal.emit(self.total_modelo_55, self.total_modelo_65)
                                
                                # Integração Supabase: envia dados processados (modo separação)
                                if False and self.enviar_supabase and SUPABASE_AVAILABLE and zip_list:
                                    try:
                                        empresa_cnpj = info_emp.get("cnpj", "")
                                        if empresa_cnpj:
                                            try:
                                                dt_ini = datetime.strptime(mes_ini, "%d/%m/%Y").date()
                                                dt_fim = datetime.strptime(mes_fim, "%d/%m/%Y").date()
                                                enviar_dados_supabase_smart_upsert(
                                                    empresa_cnpj=empresa_cnpj,
                                                    empresa_nome=display_name,
                                                    zip_paths=zip_list,
                                                    data_inicial=dt_ini,
                                                    data_final=dt_fim,
                                                    operacao=operacao,
                                                    base_empresas_dir=self.base_empresas_dir,
                                                    folder_structure=self.folder_structure,
                                                    separar_modelo_xml=self.separar_modelo_xml,
                                                    log_fn=self._log
                                                )
                                            except ValueError as e:
                                                self._log(f"[WARN] ⚠️ Erro ao converter datas para Supabase: {e}")
                                        else:
                                            self._log(f"[WARN] ⚠️ CNPJ não encontrado para {display_name}. Dados não enviados ao Supabase.")
                                    except Exception as e:
                                        self._log(f"[ERRO] ❌ Erro ao enviar dados para Supabase: {e}")

                                
                            else:
                                # ✅ MODO NORMAL (sem separação): mantém UNIFICAÇÃO na pasta da empresa
                                destino_unificado = download_dirs_por_operacao.get(operacao)
                                if not destino_unificado:
                                    destino_unificado = montar_caminho_estruturado(
                                        self.base_empresas_dir,
                                        self.folder_structure,
                                        empresa_pasta_nome,
                                        operacao=operacao,
                                        pasta_mes=pasta_mes,
                                        criar=True,
                                    ) or self.base_empresas_dir
                                if False and self.enviar_supabase and SUPABASE_AVAILABLE and zip_list:
                                    try:
                                        empresa_cnpj = info_emp.get("cnpj", "")
                                        if empresa_cnpj:
                                            try:
                                                dt_ini = datetime.strptime(mes_ini, "%d/%m/%Y").date()
                                                dt_fim = datetime.strptime(mes_fim, "%d/%m/%Y").date()
                                                enviar_dados_supabase_smart_upsert(
                                                    empresa_cnpj=empresa_cnpj,
                                                    empresa_nome=display_name,
                                                    zip_paths=zip_list,
                                                    data_inicial=dt_ini,
                                                    data_final=dt_fim,
                                                    operacao=operacao,
                                                    base_empresas_dir=self.base_empresas_dir,
                                                    folder_structure=self.folder_structure,
                                                    separar_modelo_xml=self.separar_modelo_xml,
                                                    log_fn=self._log
                                                )
                                            except ValueError as e:
                                                self._log(f"[WARN] âš ï¸ Erro ao converter datas para Supabase: {e}")
                                        else:
                                            self._log(f"[WARN] âš ï¸ CNPJ nÃ£o encontrado para empresa {display_name}. Dados nÃ£o enviados ao Supabase.")
                                    except Exception as e:
                                        self._log(f"[ERRO] âŒ Erro ao enviar dados para Supabase: {e}")
                                unified_zip_path = unificar_zips_xml(
                                    zip_paths=zip_list,
                                    dest_dir=destino_unificado,
                                    ie=ie,
                                    display_name=display_name,
                                    operacao=operacao,
                                    data_inicial=mes_ini,
                                    data_final=mes_fim,
                                    log_fn=self._log
                                )
                                try:
                                    if unified_zip_path and path_isfile(unified_zip_path):
                                        os.remove(fs_path(unified_zip_path))
                                        self._log(f"[INFO] ZIP removido após extração dos XMLs: {unified_zip_path}")
                                except Exception as e:
                                    self._log(f"[WARN] Falha ao remover ZIP final '{unified_zip_path}': {e}")
                                
                                # Integração Supabase: envia dados processados
                                if False and self.enviar_supabase and SUPABASE_AVAILABLE and zip_list:
                                    try:
                                        empresa_cnpj = info_emp.get("cnpj", "")
                                        if empresa_cnpj:
                                            # Converte mes_ini e mes_fim para date
                                            try:
                                                dt_ini = datetime.strptime(mes_ini, "%d/%m/%Y").date()
                                                dt_fim = datetime.strptime(mes_fim, "%d/%m/%Y").date()
                                                enviar_dados_supabase_smart_upsert(
                                                    empresa_cnpj=empresa_cnpj,
                                                    empresa_nome=display_name,
                                                    zip_paths=zip_list,
                                                    data_inicial=dt_ini,
                                                    data_final=dt_fim,
                                                    log_fn=self._log
                                                )
                                            except ValueError as e:
                                                self._log(f"[WARN] ⚠️ Erro ao converter datas para Supabase: {e}")
                                        else:
                                            self._log(f"[WARN] ⚠️ CNPJ não encontrado para empresa {display_name}. Dados não enviados ao Supabase.")
                                    except Exception as e:
                                        self._log(f"[ERRO] ❌ Erro ao enviar dados para Supabase: {e}")
                    
                            # Limpa a chave (os ZIPs originais já são apagados na unificação/separação)
                            zips_por_emp_op_intervalo.pop(chave, None)

                # Regra solicitada: se 2 empresas consecutivas nao baixarem Entrada nem Saida,
                # reinicia navegador (fecha, espera 5s e abre novamente).
                if (not self.stop_requested) and processa_entrada and processa_saida and (not nao_contar_consecutivo):
                    if (not baixou_entrada_empresa) and (not baixou_saida_empresa):
                        ignora_consecutivo, motivo_ignorar = _empresa_sem_download_ignorar_consecutivo(ie)
                        if ignora_consecutivo:
                            if consecutivas_sem_entrada_saida:
                                self._log(
                                    "[INFO] Empresa sem download por motivo nao-erro "
                                    f"({motivo_ignorar}). Zerando contador de consecutivas."
                                )
                            else:
                                self._log(
                                    "[INFO] Empresa sem download por motivo nao-erro "
                                    f"({motivo_ignorar}). Nao entra na regra de reinicio."
                                )
                            consecutivas_sem_entrada_saida = 0
                        else:
                            consecutivas_sem_entrada_saida += 1
                            self._log(
                                f"[WARN] Empresa sem download de Entrada e Saida: {display_name} "
                                f"(consecutivas: {consecutivas_sem_entrada_saida}/2)."
                            )
                            if consecutivas_sem_entrada_saida >= 2:
                                ok_restart = _restart_session_same_login(
                                    reason="2 empresas consecutivas sem baixar Entrada/Saida",
                                    wait_s=5.0,
                                )
                                if not ok_restart:
                                    self._log("[FATAL] Nao foi possivel reiniciar o navegador apos falhas consecutivas.")
                                    return
                                consecutivas_sem_entrada_saida = 0
                    else:
                        consecutivas_sem_entrada_saida = 0

                self._log(
                    f"[INFO] 🏁 Finalizado processamento da empresa: {display_name} "
                    f"(IE {ie_formatada(ie)})."
                )

            if self.stop_requested:
                self._log(
                    "[WARN] ⏹️ Processamento geral interrompido pelo usuário durante a execução."
                )


            # =============================================================
            # RETENTATIVA AUTOMÁTICA (INFINITA) PARA ERROS DE DOWNLOAD/PORTAL
            # - Reprocessa apenas intervalos que falharam por erro (não inclui "SEM_RESULTADO")
            # - Continua tentando até baixar tudo (ou até o usuário solicitar parar)
            # =============================================================
            if not self.stop_requested:
                NAO_REPETIR_CLASSES = {
                    "SEM_RESULTADO",
                    "IE_INVÁLIDA",
                    "IE_OBRIGATÓRIA",
                    "INTERRUPCAO_USUARIO",
                }
                rodada_retry = 0
                max_rodadas_retry = 2
                while True:
                    if self.stop_requested:
                        break

                    pendentes = [
                        r for r in self.resultados
                        if (not r.success) and ((r.alert_class or "") not in NAO_REPETIR_CLASSES)
                    ]
                    if not pendentes:
                        break

                    if rodada_retry >= max_rodadas_retry:
                        self._log(
                            f"[WARN] 🔁 Limite de {max_rodadas_retry} rodada(s) de retentativa atingido. "
                            "As pendências restantes serão mantidas como falha."
                        )
                        break

                    rodada_retry += 1
                    self._log(
                        f"[WARN] 🔁 Rodada de retentativa #{rodada_retry}: "
                        f"{len(pendentes)} pendência(s) com erro detectada(s)."
                    )

                    pendentes_por_ie: Dict[str, List[ResultadoExecucao]] = {}
                    for rr in pendentes:
                        pendentes_por_ie.setdefault(rr.ie, []).append(rr)

                    # Mantem uma ordem previsivel: ordem original das empresas, depois qualquer IE extra.
                    ies_no_retry = []
                    try:
                        base_order = list(ies_ordenados or [])
                    except Exception:
                        base_order = []
                    for ie0 in base_order:
                        if ie0 in pendentes_por_ie:
                            ies_no_retry.append(ie0)
                    for ie0 in pendentes_por_ie.keys():
                        if ie0 not in ies_no_retry:
                            ies_no_retry.append(ie0)

                    for ie_retry in ies_no_retry:
                        if self.stop_requested:
                            break
                        lista = pendentes_por_ie.get(ie_retry) or []
                        if not lista:
                            continue

                        # Processa todos os intervalos pendentes da empresa na rodada atual
                        for r in lista:
                            if self.stop_requested:
                                break

                            try:
                                while True:
                                    try:
                                        abrir_menu_baixar_xml_nfe(page)
                                    except Exception:
                                        pass

                                    tipo_nota = "entrada" if r.operacao.lower().startswith("e") else "saida"

                                    resultado_dict = pesquisar_intervalo(
                                        page,
                                        r.ie,
                                        r.data_inicial,
                                        r.data_final,
                                        tipo_nota,
                                        max_tentativas=MAX_TENTATIVAS_EMPRESA,
                                        should_stop=lambda: self.stop_requested,
                                        log_fn=self._log,
                                    )

                                    if (resultado_dict.get("alert_class") or "") == "SEM_DADOS_ENVIADOS":
                                        if not _tratar_sem_dados_enviados(
                                            f"retentativa {r.display_name} | {r.operacao} | {r.data_inicial} a {r.data_final}"
                                        ):
                                            return
                                        continue

                                    falhas_sem_dados_enviados_consecutivas = 0
                                    r.success = bool(resultado_dict.get("success"))
                                    r.tentativas = int(resultado_dict.get("tentativas") or 0)
                                    r.alert_type = resultado_dict.get("tipo_alerta")
                                    r.alert_class = resultado_dict.get("alert_class")
                                    r.alert_text = resultado_dict.get("alert_text", "")
                                    r.motivo = resultado_dict.get("motivo", "")
                                    break

                                # Se a pesquisa não foi bem-sucedida, deixa para a próxima rodada
                                if not r.success:
                                    if r.alert_class == "SERVICE_503":
                                        sleep(3)
                                    continue

                                # Pesquisa OK -> tenta gerar e baixar o ZIP
                                try:
                                    ok_zip_req = solicitar_geracao_zip(
                                        page,
                                        self.download_option,
                                        log_fn=self._log
                                    )
                                except Exception as e:
                                    self._log(f"[ERRO] Excecao ao solicitar geracao do ZIP (retry): {e}")
                                    ok_zip_req = False

                                if not ok_zip_req:
                                    r.success = False
                                    r.alert_class = "DOWNLOAD_ERRO"
                                    r.motivo = "Erro ao solicitar geracao de ZIP (retry)."
                                    continue

                                # Destino do ZIP: mesmo criterio do fluxo principal (respeita estrutura de pastas)
                                separar_por_modelo_ativo = self.separar_modelo_xml
                                dt_ini_dt = datetime.strptime(r.data_inicial, "%d/%m/%Y")
                                pasta_mes2 = dt_ini_dt.strftime("%m%Y")
                                empresa_pasta_nome_retry = nome_empresa_sem_ie(r.display_name)
                                download_dir = montar_caminho_estruturado(
                                    self.base_empresas_dir,
                                    self.folder_structure,
                                    empresa_pasta_nome_retry,
                                    operacao=r.operacao,
                                    pasta_mes=pasta_mes2,
                                    criar=True,
                                ) or self.base_empresas_dir

                                try:
                                    zip_paths = aguardar_e_baixar_arquivos(
                                        page,
                                        download_dir,
                                        quantidade=1,
                                        should_stop=lambda: self.stop_requested,
                                        log_fn=self._log,
                                        ie=r.ie,
                                        data_inicial=r.data_inicial,
                                        data_final=r.data_final,
                                        tipo=tipo_nota,
                                        download_option=self.download_option,
                                    )
                                except Exception as e:
                                    self._log(f"[ERRO] Excecao ao aguardar/baixar ZIP (retry): {e}")
                                    zip_paths = []

                                if not zip_paths:
                                    r.success = False
                                    r.alert_class = "DOWNLOAD_ERRO"
                                    r.motivo = "Erro ao aguardar/baixar ZIP no Historico (retry)."
                                    continue

                                # Sucesso no download
                                r.success = True
                                r.alert_type = None
                                r.alert_class = None
                                r.alert_text = ""
                                r.motivo = ""

                                # Processa imediatamente o ZIP baixado na retentativa
                                try:
                                    if separar_por_modelo_ativo:
                                        pasta_mes_retry = pasta_mes2
                                        c55, c65 = extrair_e_separar_zips_por_modelo(
                                            zip_paths=zip_paths,
                                            empresa_nome=empresa_pasta_nome_retry,
                                            empresa_ie=r.ie,
                                            empresas_map=self.empresas,
                                            pasta_mes=pasta_mes_retry,
                                            base55=self.base_empresas_dir,
                                            base65=self.base_empresas_dir,
                                            estrutura_pastas=self.folder_structure,
                                            operacao=r.operacao,
                                            data_inicial=r.data_inicial,
                                            data_final=r.data_final,
                                            zipar=True,
                                            remover_xmls=True,
                                            log_fn=self._log
                                        )
                                        self.total_modelo_55 += int(c55 or 0)
                                        self.total_modelo_65 += int(c65 or 0)
                                        self.metrics_signal.emit(self.total_modelo_55, self.total_modelo_65)

                                    else:
                                        unificar_zips_xml(
                                            zip_paths=zip_paths,
                                            dest_dir=download_dir,
                                            ie=r.ie,
                                            display_name=r.display_name,
                                            operacao=r.operacao,
                                            data_inicial=r.data_inicial,
                                            data_final=r.data_final,
                                            log_fn=self._log
                                        )
                                except Exception as e:
                                    self._log(f"[WARN] Falha ao processar ZIP baixado na retentativa: {e}")

                                # Volta para a tela de filtros
                                if not self.stop_requested:
                                    try:
                                        abrir_menu_baixar_xml_nfe(page)
                                    except Exception:
                                        pass

                            except Exception as e:
                                self._log(
                                    f"[ERRO] Excecao durante retentativa "
                                    f"({r.display_name} | {r.operacao} | {r.data_inicial} a {r.data_final}): {e}"
                                )
                                r.success = False
                                r.alert_class = "EXCECAO_WORKER"
                                r.motivo = f"Excecao durante retentativa: {e}"
                                sleep(2)

                            # Pequena pausa para evitar loop apertado em caso de instabilidade do portal
                            if not self.stop_requested:
                                sleep(5)

                        # Ao terminar a empresa nesta rodada, aplica regra de reinicio por 2 falhas consecutivas
                        if (not self.stop_requested) and processa_entrada and processa_saida:
                            tem_e, tem_s = _empresa_tem_baixa_entrada_saida(ie_retry)
                            if (not tem_e) and (not tem_s):
                                ignora_consecutivo, motivo_ignorar = _empresa_sem_download_ignorar_consecutivo(ie_retry)
                                if ignora_consecutivo:
                                    if consecutivas_sem_entrada_saida:
                                        self._log(
                                            "[INFO] Retentativa: empresa sem download por motivo nao-erro "
                                            f"({motivo_ignorar}). Zerando contador de consecutivas."
                                        )
                                    else:
                                        self._log(
                                            "[INFO] Retentativa: empresa sem download por motivo nao-erro "
                                            f"({motivo_ignorar}). Nao entra na regra de reinicio."
                                        )
                                    consecutivas_sem_entrada_saida = 0
                                else:
                                    consecutivas_sem_entrada_saida += 1
                                    self._log(
                                        f"[WARN] Retentativa: empresa ainda sem download de Entrada e Saida: "
                                        f"{(lista[0].display_name if lista else ie_retry)} "
                                        f"(consecutivas: {consecutivas_sem_entrada_saida}/2)."
                                    )
                                    if consecutivas_sem_entrada_saida >= 2:
                                        ok_restart = _restart_session_same_login(
                                            reason="2 empresas consecutivas sem baixar Entrada/Saida (retentativa)",
                                            wait_s=5.0,
                                        )
                                        if not ok_restart:
                                            self._log("[FATAL] Nao foi possivel reiniciar o navegador apos falhas consecutivas (retentativa).")
                                            return
                                        consecutivas_sem_entrada_saida = 0
                            else:
                                consecutivas_sem_entrada_saida = 0

            self._log("[OK] ✅ Processamento concluído (ou interrompido).")
            self._log("[INFO] 📊 ===== RESUMO FINAL POR INTERVALO =====")
            for r in self.resultados:
                self._log(f"[INFO] {r.resumo()}")

        except Exception as e:
            self._log(f"[FATAL] 💥 Exceção não tratada no Worker: {e}")
        finally:
            try:
                self._log("[INFO] 🔻 Encerrando sessao final do navegador e proxy...")
                _teardown_current_session(wipe_profile=False, wait_s=0.0)
            except Exception:
                pass
            try:
                _kill_automation_chrome_proc(log_fn=self._log)
            except Exception:
                pass
            try:
                _kill_automation_chrome_instances(log_fn=self._log)
            except Exception:
                pass
            try:
                _stop_proxy_if_running(log_fn=self._log)
            except Exception:
                pass
            try:
                _kill_any_mitm_processes(log_fn=self._log)
            except Exception:
                pass
            self._log("[INFO] 🔚 Worker finalizado.")





# =============================================================================
# Dashboard (Widgets nativos) – Donut Gauges (NF-e / NFC-e) + KPIs
# =============================================================================

def _parse_xml_root(xml_bytes: bytes):
    if not xml_bytes:
        return None
    try:
        i = xml_bytes.find(b"<")
        if i > 0:
            xml_bytes = xml_bytes[i:]
    except Exception:
        pass
    try:
        return ET.fromstring(xml_bytes)
    except Exception:
        return None


def _parse_nfe_emission_date_str(raw: str) -> Optional[date]:
    """Converte dhEmi/dEmi (ISO, BR ou YYYYMMDD) em date."""
    if not raw:
        return None
    s = (raw or "").strip()
    if not s:
        return None
    if "T" in s:
        s_date = s.split("T", 1)[0]
    else:
        s_date = s[:10] if len(s) >= 10 and s[4:5] == "-" and s[7:8] == "-" else s
    try:
        if len(s_date) >= 10 and s_date[4] == "-" and s_date[7] == "-":
            return datetime.strptime(s_date[:10], "%Y-%m-%d").date()
    except Exception:
        pass
    try:
        return datetime.strptime(s, "%d/%m/%Y").date()
    except Exception:
        pass
    dig = re.sub(r"[^0-9]", "", s)
    if len(dig) >= 8:
        try:
            return datetime.strptime(dig[:8], "%Y%m%d").date()
        except Exception:
            pass
    return None


def _sniff_nfe_emission_date_from_bytes(xml_bytes: bytes, max_scan: int = 98304) -> Optional[date]:
    """Só os primeiros ~96 KB: regex dhEmi/dEmi (com prefixo de namespace), sem depender de <ide>."""
    if not xml_bytes:
        return None
    head = xml_bytes if len(xml_bytes) <= max_scan else xml_bytes[:max_scan]
    try:
        txt = head.decode("utf-8", errors="ignore")
    except Exception:
        return None
    for pat in (
        r"<(?:[\w.-]+:)?dhEmi\b[^>]*>\s*([^<]+?)\s*</(?:[\w.-]+:)?dhEmi>",
        r"<(?:[\w.-]+:)?dEmi\b[^>]*>\s*([^<]+?)\s*</(?:[\w.-]+:)?dEmi>",
    ):
        m = re.search(pat, txt, re.I | re.S)
        if m:
            d = _parse_nfe_emission_date_str(m.group(1))
            if d:
                return d
    return None


def extrair_info_nfe(xml_bytes: bytes):
    """Extrai modelo (55/65), tpNF (0/1), valor vNF e data de emissão do XML.

    Retorna dict:
      {model: '55'/'65'/'', tpNF: '0'/'1'/'', vNF: float, dhEmi: date|None}

    Observação: funciona com namespaces/prefixos.
    """
    root = _parse_xml_root(xml_bytes)
    model = detectar_modelo_xml_bytes(xml_bytes, root)

    tp_nf = ""
    v_nf = 0.0
    dt_emi = None
    ie_emit = ''
    ie_dest = ''
    cnpj_emit = ''
    cnpj_dest = ''
    cfops: List[str] = []

    if root is not None:
        # tpNF e data dentro de <ide>
        ide = None
        for el in root.iter():
            if _local_tag(getattr(el, 'tag', '')) == 'ide':
                ide = el
                break

        if ide is not None:
            for el in ide.iter():
                t = _local_tag(getattr(el, 'tag', ''))
                if t == 'tpnf' and getattr(el, 'text', None):
                    val = (el.text or '').strip()
                    if val in ('0', '1'):
                        tp_nf = val
                elif t in ('dhemi', 'demi') and getattr(el, 'text', None):
                    dt_emi = _parse_nfe_emission_date_str((el.text or '').strip())

        if dt_emi is None and root is not None:
            for el in root.iter():
                t = _local_tag(getattr(el, 'tag', ''))
                if t in ('dhemi', 'demi') and getattr(el, 'text', None):
                    dt_emi = _parse_nfe_emission_date_str((el.text or '').strip())
                    if dt_emi:
                        break

        # vNF (total da nota)
        for el in root.iter():
            if _local_tag(getattr(el, 'tag', '')) == 'vnf' and getattr(el, 'text', None):
                raw = (el.text or '').strip().replace(',', '.')
                try:
                    v_nf = float(raw)
                except Exception:
                    v_nf = 0.0
                break
        # IE/CNPJ emitente e destinatário (prioriza IE, fallback por CNPJ)
        try:
            for el in root.iter():
                tag = _local_tag(getattr(el, 'tag', ''))
                if tag == 'emit':
                    for sub in el.iter():
                        t = _local_tag(getattr(sub, 'tag', ''))
                        if t == 'ie' and getattr(sub, 'text', None) and not ie_emit:
                            ie_emit = (sub.text or '').strip()
                        elif t == 'cnpj' and getattr(sub, 'text', None) and not cnpj_emit:
                            cnpj_emit = (sub.text or '').strip()
                    break
        except Exception:
            pass

        try:
            for el in root.iter():
                tag = _local_tag(getattr(el, 'tag', ''))
                if tag == 'dest':
                    for sub in el.iter():
                        t = _local_tag(getattr(sub, 'tag', ''))
                        if t == 'ie' and getattr(sub, 'text', None) and not ie_dest:
                            ie_dest = (sub.text or '').strip()
                        elif t == 'cnpj' and getattr(sub, 'text', None) and not cnpj_dest:
                            cnpj_dest = (sub.text or '').strip()
                    break
        except Exception:
            pass

        # CFOPs (pode haver mais de um item na nota)
        try:
            seen = set()
            for el in root.iter():
                if _local_tag(getattr(el, 'tag', '')) == 'cfop' and getattr(el, 'text', None):
                    v = (el.text or '').strip()
                    v = re.sub(r'[^0-9]', '', v)
                    if v and v not in seen:
                        seen.add(v)
                        cfops.append(v)
        except Exception:
            pass

    if dt_emi is None:
        dt_emi = _sniff_nfe_emission_date_from_bytes(xml_bytes)

    # Chave de acesso (44 dígitos) – usada para deduplicar no Dashboard
    chave = ''
    try:
        # 1) <chNFe> (comum em protNFe)
        for el in root.iter() if root is not None else []:
            if _local_tag(getattr(el, 'tag', '')) == 'chnfe' and getattr(el, 'text', None):
                dig = re.sub(r'[^0-9]', '', (el.text or '').strip())
                if len(dig) == 44:
                    chave = dig
                    break
        # 2) atributo Id="NFe..." em <infNFe>
        if not chave and root is not None:
            for el in root.iter():
                if _local_tag(getattr(el, 'tag', '')) == 'infnfe':
                    idv = (el.attrib.get('Id') or el.attrib.get('id') or '').strip()
                    dig = re.sub(r'[^0-9]', '', idv)
                    if len(dig) >= 44:
                        chave = dig[-44:]
                        break
    except Exception:
        chave = ''




    return {
        'model': model or '',
        'tpNF': tp_nf,
        'vNF': float(v_nf or 0.0),
        'dtEmi': dt_emi,
        'chave': chave,
        'ieEmit': ie_emit,
        'ieDest': ie_dest,
        'cnpjEmit': cnpj_emit,
        'cnpjDest': cnpj_dest,
        'cfops': cfops,
    }


def _format_brl(valor: float) -> str:
    try:
        v = float(valor or 0.0)
    except Exception:
        v = 0.0
    # Formato BR manual (sem locale)
    s = f"{v:,.2f}"
    s = s.replace(',', 'X').replace('.', ',').replace('X', '.')
    return f"R$ {s}"


class DonutGaugeWidget(QWidget):
    """Gauge circular estilo dashboard, desenhado via QPainter (sem QtQuick).

    value: 0..1
    """

    def __init__(self, label: str = 'NF-e', sublabel: str = '', accent: str = '#3B82F6', parent=None):
        super().__init__(parent)
        self._value = 0.0
        self._label = label
        self._sublabel = sublabel
        self._accent = QColor(accent)

        self._anim = QPropertyAnimation(self, b"value")
        self._anim.setDuration(550)
        self._anim.setEasingCurve(QEasingCurve.OutCubic)

        # responsivo: permite reduzir em telas menores sem precisar de scroll
        self.setMinimumSize(160, 160)
        self.setAttribute(Qt.WA_TranslucentBackground, True)

    def sizeHint(self):
        return QSize(200, 200)

    def getValue(self):
        return float(self._value)

    def setValue(self, v: float):
        try:
            v = float(v)
        except Exception:
            v = 0.0
        v = max(0.0, min(1.0, v))
        if abs(v - self._value) < 1e-6:
            return
        self._value = v
        self.update()

    value = Property(float, getValue, setValue)

    def setLabel(self, t: str):
        self._label = str(t or '')
        self.update()

    def setSubLabel(self, t: str):
        self._sublabel = str(t or '')
        self.update()

    def setAccent(self, c):
        self._accent = QColor(c)
        self.update()

    def animate_to(self, v: float):
        self._anim.stop()
        self._anim.setStartValue(self._value)
        self._anim.setEndValue(max(0.0, min(1.0, float(v or 0.0))))
        self._anim.start()

    def paintEvent(self, ev):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)

        # margem interna: evita qualquer recorte do arco/ticks quando o layout fica mais "apertado"
        r = self.rect().adjusted(10, 10, -10, -10)
        w = r.width()
        h = r.height()
        side = min(w, h)

        # fundo transparente (mantém o tema do app)
        p.setPen(Qt.NoPen)

        cx = r.center().x()
        cy = r.center().y()

        radius = side * 0.36
        thickness = max(12.0, min(18.0, radius * 0.18))

        start_deg = -162.0
        span_deg_total = 324.0
        span_deg_val = span_deg_total * self._value

        arc_rect = QRectF(cx - radius, cy - radius, radius * 2, radius * 2)

        # background ring
        pen_bg = QPen(QColor(120, 140, 170, int(255 * 0.18)))
        pen_bg.setWidthF(thickness)
        pen_bg.setCapStyle(Qt.RoundCap)
        p.setPen(pen_bg)
        p.drawArc(arc_rect, int(start_deg * 16), int(span_deg_total * 16))

        # glow (desenha uma vez mais grossa com alpha)
        glow = QColor(self._accent)
        glow.setAlpha(int(255 * 0.28))
        pen_glow = QPen(glow)
        pen_glow.setWidthF(thickness + 6)
        pen_glow.setCapStyle(Qt.RoundCap)
        p.setPen(pen_glow)
        p.drawArc(arc_rect, int(start_deg * 16), int(span_deg_val * 16))

        # active ring
        pen_fg = QPen(self._accent)
        pen_fg.setWidthF(thickness)
        pen_fg.setCapStyle(Qt.RoundCap)
        p.setPen(pen_fg)
        p.drawArc(arc_rect, int(start_deg * 16), int(span_deg_val * 16))

        # tick dots
        ticks = 28
        tick_r = radius + thickness * 0.8
        start_rad = math.radians(start_deg)
        span_rad_total = math.radians(span_deg_total)
        for i in range(ticks + 1):
            t = start_rad + span_rad_total * (i / ticks)
            x = cx + math.cos(t) * tick_r
            y = cy + math.sin(t) * tick_r
            active = (i / ticks) <= self._value + 1e-9
            if active:
                c = QColor(self._accent)
                c.setAlpha(int(255 * 0.65))
            else:
                c = QColor(160, 175, 200, int(255 * 0.10))
            p.setPen(Qt.NoPen)
            p.setBrush(c)
            p.drawEllipse(QPointF(x, y), 1.3, 1.3)

        # textos (responsivos): tamanho dos textos acompanha o widget
        muted = QColor('#9AA3B2')
        textc = QColor('#E8ECF6')

        # calcula tamanhos em pixels com base no lado útil
        # (clamps evitam ficar gigante ou minúsculo em telas extremas)
        label_px = max(9, min(14, int(side * 0.070)))
        pct_px = max(16, min(44, int(side * 0.185)))
        sub_px = max(9, min(13, int(side * 0.060)))

        # alturas de linha (em px) + gaps
        label_h = int(label_px * 1.6)
        pct_h = int(pct_px * 1.35)
        sub_h = int(sub_px * 1.55) if self._sublabel else 0
        gap1 = max(2, int(side * 0.012))
        gap2 = max(2, int(side * 0.010)) if self._sublabel else 0

        total_h = label_h + gap1 + pct_h + gap2 + sub_h
        start_y = cy - (total_h / 2.0)

        # label
        p.setPen(muted)
        f1 = QFont('Verdana')
        f1.setPixelSize(label_px)
        p.setFont(f1)
        p.drawText(QRectF(r.left(), start_y, r.width(), label_h), Qt.AlignHCenter | Qt.AlignVCenter, self._label)

        # percent
        p.setPen(textc)
        f2 = QFont('Verdana')
        f2.setBold(True)
        f2.setPixelSize(pct_px)
        p.setFont(f2)
        pct = f"{int(round(self._value * 100))}%"
        p.drawText(QRectF(r.left(), start_y + label_h + gap1, r.width(), pct_h), Qt.AlignHCenter | Qt.AlignVCenter, pct)

        # sublabel
        if self._sublabel:
            p.setPen(muted)
            f3 = QFont('Verdana')
            f3.setPixelSize(sub_px)
            p.setFont(f3)
            p.drawText(
                QRectF(r.left(), start_y + label_h + gap1 + pct_h + gap2, r.width(), sub_h),
                Qt.AlignHCenter | Qt.AlignVCenter,
                self._sublabel,
            )


class DonutDashboard(QWidget):
    """Painel com 2 donuts: NF-e (55) e NFC-e (65)."""

    def __init__(self, parent=None):
        super().__init__(parent)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(18)

        self._g55 = DonutGaugeWidget(label='NF-e', sublabel='Modelo 55', accent='#3B82F6')
        self._g65 = DonutGaugeWidget(label='NFC-e', sublabel='Modelo 65', accent='#3B82F6')

        for g in (self._g55, self._g65):
            g.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        layout.addStretch(1)
        layout.addWidget(self._g55, 1)
        layout.addWidget(self._g65, 1)
        layout.addStretch(1)

    def set_counts(self, total_55: int, total_65: int):
        total_55 = int(total_55 or 0)
        total_65 = int(total_65 or 0)
        total = max(0, total_55) + max(0, total_65)
        v55 = (total_55 / total) if total else 0.0
        v65 = (total_65 / total) if total else 0.0

        self._g55.setSubLabel(f"{total_55} XML")
        self._g65.setSubLabel(f"{total_65} XML")

        self._g55.animate_to(v55)
        self._g65.animate_to(v65)


class DualBarKpiWidget(QWidget):
    """Mini gráfico de barras (Entrada vs Saída), desenhado via QPainter.

    Evita QML/QQuickWidget (sem quadrados pretos), e mantém animação leve.
    """

    def __init__(self, title: str = "Entrada x Saída", accent: str = "#3B82F6", parent=None):
        super().__init__(parent)
        self._title = title
        self._accent = QColor(accent)
        self._entrada = 0.0
        self._saida = 0.0

        self._anim_e = QPropertyAnimation(self, b"entradaValue")
        self._anim_s = QPropertyAnimation(self, b"saidaValue")
        for a in (self._anim_e, self._anim_s):
            a.setDuration(520)
            a.setEasingCurve(QEasingCurve.OutCubic)

        # responsivo: permite reduzir em telas menores
        self.setMinimumHeight(140)
        self.setAttribute(Qt.WA_TranslucentBackground, True)

    def sizeHint(self):
        return QSize(520, 190)

    def getEntradaValue(self):
        return float(self._entrada)

    def setEntradaValue(self, v: float):
        try:
            v = float(v)
        except Exception:
            v = 0.0
        if abs(v - self._entrada) < 1e-6:
            return
        self._entrada = max(0.0, v)
        self.update()

    entradaValue = Property(float, getEntradaValue, setEntradaValue)

    def getSaidaValue(self):
        return float(self._saida)

    def setSaidaValue(self, v: float):
        try:
            v = float(v)
        except Exception:
            v = 0.0
        if abs(v - self._saida) < 1e-6:
            return
        self._saida = max(0.0, v)
        self.update()

    saidaValue = Property(float, getSaidaValue, setSaidaValue)

    def animate_to(self, entrada: float, saida: float):
        self._anim_e.stop()
        self._anim_s.stop()

        self._anim_e.setStartValue(self._entrada)
        self._anim_e.setEndValue(max(0.0, float(entrada or 0.0)))

        self._anim_s.setStartValue(self._saida)
        self._anim_s.setEndValue(max(0.0, float(saida or 0.0)))

        self._anim_e.start()
        self._anim_s.start()

    def paintEvent(self, ev):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing, True)

        r = self.rect().adjusted(10, 10, -10, -10)
        if r.width() <= 10 or r.height() <= 10:
            return

        # Paleta
        textc = QColor('#E8F4FF')
        muted = QColor('#9AA3B2')
        rail = QColor(160, 175, 200, int(255 * 0.10))
        accent1 = QColor(self._accent)
        accent2 = QColor(self._accent)
        accent2.setAlpha(int(255 * 0.55))

        # título
        p.setPen(muted)
        p.setFont(QFont('Verdana', 9, QFont.Bold))
        p.drawText(QRectF(r.left(), r.top(), r.width(), 16), Qt.AlignLeft | Qt.AlignVCenter, self._title)

        # valores
        vmax = max(self._entrada, self._saida, 1.0)
        bar_h = 16.0
        gap = 30.0
        x0 = r.left()
        y0 = r.top() + 28
        w = r.width()

        def _draw_bar(y, label, value, color):
            # label
            p.setPen(muted)
            p.setFont(QFont('Verdana', 9))
            p.drawText(QRectF(x0, y - 10, 110, 18), Qt.AlignLeft | Qt.AlignVCenter, label)

            # trilho
            rail_rect = QRectF(x0 + 110, y, w - 110, bar_h)
            p.setPen(Qt.NoPen)
            p.setBrush(rail)
            p.drawRoundedRect(rail_rect, 7, 7)

            # barra
            frac = (value / vmax) if vmax else 0.0
            fill_w = max(8.0, (rail_rect.width() * frac)) if value > 0 else 0.0
            fill_rect = QRectF(rail_rect.left(), rail_rect.top(), fill_w, rail_rect.height())

            # glow
            glow = QColor(color)
            glow.setAlpha(int(255 * 0.22))
            p.setBrush(glow)
            p.drawRoundedRect(QRectF(fill_rect.left(), fill_rect.top() - 2, fill_rect.width(), fill_rect.height() + 4), 8, 8)

            p.setBrush(color)
            p.drawRoundedRect(fill_rect, 7, 7)

            # valor (sempre branco)
            p.setPen(textc)
            p.setFont(QFont('Verdana', 10, QFont.Bold))
            p.drawText(QRectF(rail_rect.left(), y - 22, rail_rect.width(), 18), Qt.AlignRight | Qt.AlignVCenter, _format_brl(value))

        _draw_bar(y0, 'Entrada', float(self._entrada), accent2)
        _draw_bar(y0 + bar_h + gap, 'Saída', float(self._saida), accent1)



class MountainAreaChartWidget(QWidget):
    """Gráfico de área ("montanha") desenhado com QPainter.

    - Leve e sem dependências de QtCharts/QtQuick.
    - Anima atualizações sem "piscar"/sumir.
    - Mostra eixos (Y à esquerda, X embaixo) com marcadores.

    Extra (seleção de intervalo):
      - 1º clique: fixa o dia inicial
      - mover o mouse: prévia do intervalo até o dia sob o mouse
      - 2º clique: fixa o dia final
      - 3º clique: desfaz

    Emite rangeChanged(str) para a UI atualizar a métrica do cabeçalho.
    """

    rangeChanged = Signal(str)

    def __init__(self, accent: str = "#3B82F6", parent=None):
        super().__init__(parent)
        self._accent = QColor(accent)
        self._labels = []
        self._values = []
        self._prev_values = []
        self._anim_t = 1.0

        # Hover/tooltip
        self._hover_idx: int | None = None
        self.setMouseTracking(True)

        # Seleção de intervalo
        self._range_start_idx: int | None = None
        self._range_end_idx: int | None = None

        self._anim = QPropertyAnimation(self, b"animT")
        self._anim.setDuration(520)
        self._anim.setEasingCurve(QEasingCurve.OutCubic)

        # Mantém o gráfico sempre visível dentro do QScrollArea
        self.setMinimumHeight(200)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.setAttribute(Qt.WA_TranslucentBackground, True)

    def _compute_outer_plot(self):
        outer = self.rect().adjusted(12, 12, -12, -12)
        left_w = 54
        bottom_h = 26
        plot = outer.adjusted(left_w, 10, -10, -bottom_h)
        return outer, plot, left_w, bottom_h

    def _idx_from_xy(self, x: float, y: float):
        try:
            if not self._values:
                return None
            _, plot, _left_w, _bottom_h = self._compute_outer_plot()
            if plot.width() < 40 or plot.height() < 40:
                return None
            if (x < plot.left()) or (x > plot.right()) or (y < plot.top()) or (y > plot.bottom()):
                return None
            vals = self._interp_vals()
            n = len(vals)
            if n <= 0:
                return None
            denom = (n - 1) if n > 1 else 1
            rel = (x - plot.left()) / plot.width() if plot.width() > 0 else 0.0
            idx = int(round(rel * denom))
            idx = max(0, min(n - 1, idx))
            return idx
        except Exception:
            return None

    def _range_bounds(self, end_idx: int | None, n: int):
        if self._range_start_idx is None or n <= 0:
            return None
        a = int(self._range_start_idx)
        b = int(end_idx if end_idx is not None else a)
        a = max(0, min(n - 1, a))
        b = max(0, min(n - 1, b))
        return (a, b) if a <= b else (b, a)

    def _range_metric_line(self, end_idx: int | None, vals: list[float]) -> str:
        n = len(vals)
        bounds = self._range_bounds(end_idx, n)
        if not bounds:
            return ''
        a, b = bounds

        def _lab(i: int) -> str:
            if self._labels and len(self._labels) == n:
                try:
                    return str(self._labels[i])
                except Exception:
                    pass
            return f"{i+1}"
        la, lb = _lab(a), _lab(b)
        # A série exibida é acumulada; para o intervalo, usamos o delta do acumulado
        start = a
        end = b
        prev = float(vals[start - 1]) if (start > 0 and start - 1 < n) else 0.0
        total = float(vals[end]) - prev
        dias = (end - start + 1)
        media = (total / dias) if dias else 0.0

        if self._range_end_idx is None:
            # prévia (ainda não fixou o fim)
            if a == b:
                return f"Seleção: {la} • {_format_brl(total)} • clique em outro dia para fechar (3º clique desfaz)"
            return f"Intervalo {la} a {lb} • {dias} dias • Total {_format_brl(total)} • Média/dia {_format_brl(media)}"

        # fixo
        if a == b:
            return f"Intervalo fixo: {la} • {_format_brl(total)} • 3º clique desfaz"
        return f"Intervalo fixo {la} a {lb} • {dias} dias • Total {_format_brl(total)} • Média/dia {_format_brl(media)}"

    def _range_tooltip_text(self, hover_idx: int, vals: list[float]) -> str:
        if self._range_start_idx is None:
            return ''
        end_idx = self._range_end_idx if self._range_end_idx is not None else hover_idx
        n = len(vals)
        bounds = self._range_bounds(end_idx, n)
        if not bounds:
            return ''
        a, b = bounds

        def _lab(i: int) -> str:
            if self._labels and len(self._labels) == n:
                try:
                    return str(self._labels[i])
                except Exception:
                    pass
            return f"{i+1}"
        la, lb = _lab(a), _lab(b)
        # A série exibida é acumulada; para o intervalo, usamos o delta do acumulado
        start = a
        end = b
        prev = float(vals[start - 1]) if (start > 0 and start - 1 < n) else 0.0
        total = float(vals[end]) - prev
        dias = (end - start + 1)
        media = (total / dias) if dias else 0.0

        if self._range_end_idx is None:
            return (
                f"Intervalo (prévia): {la} → {lb} ({dias} dias)\n"
                f"Total: {_format_brl(total)}\n"
                f"Média/dia: {_format_brl(media)}\n\n"
                "Clique para fixar o fim (3º clique desfaz)"
            )

        return (
            f"Intervalo (fixo): {la} → {lb} ({dias} dias)\n"
            f"Total: {_format_brl(total)}\n"
            f"Média/dia: {_format_brl(media)}\n\n"
            "3º clique desfaz"
        )

    def _emit_range_metric(self, end_idx=None):
        try:
            if self._range_start_idx is None:
                self.rangeChanged.emit('')
                return
            vals = self._interp_vals()
            if not vals:
                self.rangeChanged.emit('')
                return
            if end_idx is None:
                end_idx = self._range_end_idx if self._range_end_idx is not None else self._hover_idx
            self.rangeChanged.emit(self._range_metric_line(end_idx, vals))
        except Exception:
            pass

    def _hover_text(self, idx: int, vals: list[float]) -> str:
        n = len(vals)
        lab = ''
        if self._labels and len(self._labels) == n:
            try:
                lab = str(self._labels[idx])
            except Exception:
                lab = ''
        if not lab:
            lab = f"Ponto {idx+1}"
        val = float(vals[idx] if idx < n else 0.0)

        prev = float(vals[idx - 1]) if (idx > 0 and (idx - 1) < n) else 0.0
        dia = val - prev

        base = f"{lab}\nAcumulado: {_format_brl(val)}\nNo dia: {_format_brl(dia)}"
        try:
            if self._range_start_idx is not None:
                extra = self._range_tooltip_text(idx, vals)
                if extra:
                    base = base + "\n\n" + extra
        except Exception:
            pass
        return base

    def mousePressEvent(self, ev):
        try:
            # 3º clique (quando já tem início e fim) desfaz
            if self._range_start_idx is not None and self._range_end_idx is not None:
                self._range_start_idx = None
                self._range_end_idx = None
                try:
                    self.rangeChanged.emit('')
                except Exception:
                    pass
                self.update()
                return

            x = ev.position().x() if hasattr(ev, 'position') else ev.pos().x()
            y = ev.position().y() if hasattr(ev, 'position') else ev.pos().y()
            idx = self._idx_from_xy(x, y)
            if idx is None:
                return

            if self._range_start_idx is None:
                # 1º clique: fixa o início
                self._range_start_idx = int(idx)
                self._range_end_idx = None
            else:
                # 2º clique: fixa o fim
                self._range_end_idx = int(idx)

            self._emit_range_metric(idx)
            self.update()
        finally:
            try:
                super().mousePressEvent(ev)
            except Exception:
                pass

    def mouseMoveEvent(self, ev):
        try:
            if not self._values:
                return

            _, plot, _left_w, _bottom_h = self._compute_outer_plot()
            if plot.width() < 40 or plot.height() < 40:
                return

            x = ev.position().x() if hasattr(ev, 'position') else ev.pos().x()
            y = ev.position().y() if hasattr(ev, 'position') else ev.pos().y()

            if (x < plot.left() - 6) or (x > plot.right() + 6) or (y < plot.top() - 6) or (y > plot.bottom() + 6):
                if self._hover_idx is not None:
                    self._hover_idx = None
                    QToolTip.hideText()
                    self.update()
                # se está selecionando, volta para a métrica de "apenas início"
                if self._range_start_idx is not None and self._range_end_idx is None:
                    self._emit_range_metric(None)
                return

            vals = self._interp_vals()
            n = len(vals)
            if n <= 0:
                return

            denom = (n - 1) if n > 1 else 1
            rel = (x - plot.left()) / plot.width() if plot.width() > 0 else 0.0
            idx = int(round(rel * denom))
            idx = max(0, min(n - 1, idx))

            if idx != self._hover_idx:
                self._hover_idx = idx
                gp = ev.globalPosition().toPoint() if hasattr(ev, 'globalPosition') else ev.globalPos()
                QToolTip.showText(gp, self._hover_text(idx, vals), self)

                # prévia do intervalo enquanto não fixou o fim
                if self._range_start_idx is not None and self._range_end_idx is None:
                    self._emit_range_metric(idx)

                self.update()
        except Exception:
            return

    def leaveEvent(self, ev):
        try:
            self._hover_idx = None
            QToolTip.hideText()
            self.update()
            if self._range_start_idx is not None and self._range_end_idx is None:
                self._emit_range_metric(None)
        except Exception:
            pass
        try:
            super().leaveEvent(ev)
        except Exception:
            pass

    def getAnimT(self):
        return float(self._anim_t)

    def setAnimT(self, v: float):
        try:
            v = float(v)
        except Exception:
            v = 1.0
        self._anim_t = max(0.0, min(1.0, v))
        self.update()

    animT = Property(float, getAnimT, setAnimT)

    def set_series(self, labels, values, force_clear: bool = False):
        """Define a série do gráfico.

        - Por padrão, evita zerar em atualizações vazias para não piscar.
        - Use force_clear=True para limpar (ex.: empresa sem dados).

        Obs.: atualizar a série reseta a seleção de intervalo.
        """
        labels = list(labels or [])
        new_vals = [float(x or 0.0) for x in (values or [])]

        # reseta seleção (mudou a base do gráfico)
        self._range_start_idx = None
        self._range_end_idx = None
        try:
            self.rangeChanged.emit('')
        except Exception:
            pass

        if force_clear:
            self._labels = labels
            self._values = list(new_vals)
            self._prev_values = list(new_vals)
            self._hover_idx = None
            try:
                self._anim.stop()
            except Exception:
                pass
            self._anim_t = 1.0
            self.update()
            return

        # Se vier vazio durante uma atualização, NÃO zera (evita "sumir").
        if (not labels or not new_vals) and self._values:
            if labels and len(labels) == len(self._values):
                self._labels = labels
            return

        self._labels = labels

        if not self._values:
            self._prev_values = list(new_vals)
        else:
            self._prev_values = list(self._values)

        # alinha tamanhos para interpolação
        if len(self._prev_values) != len(new_vals):
            last = self._prev_values[-1] if self._prev_values else 0.0
            if len(self._prev_values) < len(new_vals):
                self._prev_values += [last] * (len(new_vals) - len(self._prev_values))
            else:
                self._prev_values = self._prev_values[:len(new_vals)]

        self._values = new_vals

        self._anim.stop()
        self._anim.setStartValue(0.0)
        self._anim.setEndValue(1.0)
        self._anim.start()

    def _interp_vals(self):
        if not self._values:
            return []
        t = float(self._anim_t)
        if not self._prev_values:
            return list(self._values)
        return [pv * (1 - t) + nv * t for pv, nv in zip(self._prev_values, self._values)]

    @staticmethod
    def _short_money(v: float) -> str:
        try:
            v = float(v)
        except Exception:
            v = 0.0
        abs_v = abs(v)
        if abs_v >= 1_000_000_000:
            return f"{v/1_000_000_000:.1f}B".replace('.0B', 'B')
        if abs_v >= 1_000_000:
            return f"{v/1_000_000:.1f}M".replace('.0M', 'M')
        if abs_v >= 1_000:
            return f"{v/1_000:.1f}k".replace('.0k', 'k')
        return f"{v:.0f}"

    def paintEvent(self, ev):
        p = QPainter(self)
        try:
            p.setRenderHint(QPainter.Antialiasing, True)

            outer, plot, left_w, _bottom_h = self._compute_outer_plot()
            if outer.width() < 60 or outer.height() < 60:
                return
            if plot.width() < 40 or plot.height() < 40:
                return

            # Fundo sutil
            try:
                bg = QColor(15, 23, 42, int(255 * 0.18))
                bd = QColor(232, 244, 255, int(255 * 0.05))
                p.setPen(QPen(bd, 1))
                p.setBrush(QBrush(bg))
                rr = QRectF(outer)
                p.drawRoundedRect(rr, 12, 12)
            except Exception:
                pass

            vals = self._interp_vals()
            if not vals:
                p.setPen(QColor('#9AA3B2'))
                p.setFont(QFont('Verdana', 10))
                p.drawText(plot, Qt.AlignCenter, 'Sem dados')
                return

            vmax = max(vals)
            vmin = min(vals)
            if vmax <= 0:
                vmax, vmin = 1.0, 0.0

            pad = max((vmax - vmin) * 0.12, vmax * 0.10, 1e-9)
            hi = vmax + pad
            lo = (vmin - pad)

            n = len(vals)
            denom = (n - 1) if n > 1 else 1
            xs = [plot.left() + plot.width() * (i / denom) for i in range(n)]
            ys = [plot.bottom() - (((v - lo) / (hi - lo) if hi > lo else 0.0) * plot.height()) for v in vals]

            # grade
            grid_pen = QPen(QColor(160, 175, 200, int(255 * 0.06)), 1)
            p.setPen(grid_pen)
            for i in range(1, 4):
                y = plot.top() + plot.height() * i / 4
                p.drawLine(plot.left(), y, plot.right(), y)

            # eixos
            axis_pen = QPen(QColor(160, 175, 200, int(255 * 0.12)), 1)
            p.setPen(axis_pen)
            p.drawLine(plot.left(), plot.bottom(), plot.right(), plot.bottom())
            p.drawLine(plot.left(), plot.top(), plot.left(), plot.bottom())

            # ticks Y
            p.setFont(QFont('Verdana', 8))
            p.setPen(QColor(159, 182, 214, int(255 * 0.82)))
            for j in range(0, 5):
                y = plot.top() + (plot.height() * j / 4)
                v = hi - (hi - lo) * (j / 4)
                p.setPen(QColor(160, 175, 200, int(255 * 0.18)))
                p.drawLine(plot.left() - 4, y, plot.left(), y)
                p.setPen(QColor(159, 182, 214, int(255 * 0.82)))
                rr = QRectF(outer.left(), y - 8, left_w - 10, 16)
                p.drawText(rr, Qt.AlignRight | Qt.AlignVCenter, self._short_money(v))

            # smooth line
            path = QPainterPath()
            path.moveTo(xs[0], ys[0])
            for i in range(1, n):
                x1, y1 = xs[i - 1], ys[i - 1]
                x2, y2 = xs[i], ys[i]
                cx = (x1 + x2) / 2
                path.cubicTo(cx, y1, cx, y2, x2, y2)

            # baseline (suporta negativos)
            y0 = plot.bottom()
            if lo < 0 < hi:
                y0 = plot.bottom() - (((0.0 - lo) / (hi - lo)) * plot.height())

            if lo < 0 < hi:
                p.setPen(QPen(QColor(160, 175, 200, int(255 * 0.10)), 1, Qt.DashLine))
                p.drawLine(plot.left(), y0, plot.right(), y0)

            # area fill
            area = QPainterPath(path)
            area.lineTo(xs[-1], y0)
            area.lineTo(xs[0], y0)
            area.closeSubpath()

            grad = QLinearGradient(plot.left(), plot.top(), plot.left(), plot.bottom())
            c_top = QColor(self._accent)
            c_top.setAlpha(int(255 * 0.26))
            c_bot = QColor(self._accent)
            c_bot.setAlpha(0)
            grad.setColorAt(0.0, c_top)
            grad.setColorAt(1.0, c_bot)

            p.setPen(Qt.NoPen)
            p.setBrush(QBrush(grad))
            p.drawPath(area)

            # seleção de intervalo (faixa)
            try:
                sel_end = self._range_end_idx if self._range_end_idx is not None else self._hover_idx
                if self._range_start_idx is not None and sel_end is not None and n > 0:
                    a = int(self._range_start_idx)
                    b = int(sel_end)
                    a = max(0, min(n - 1, a))
                    b = max(0, min(n - 1, b))
                    lo_i, hi_i = (a, b) if a <= b else (b, a)

                    x0 = xs[lo_i]
                    x1 = xs[hi_i]
                    left = min(x0, x1)
                    width = abs(x1 - x0)
                    if width < 2:
                        width = 2

                    sel = QColor(self._accent)
                    sel.setAlpha(int(255 * 0.10))
                    p.setPen(Qt.NoPen)
                    p.setBrush(sel)
                    p.drawRoundedRect(QRectF(left, plot.top(), width, plot.height()), 6, 6)

                    linec = QColor(self._accent)
                    linec.setAlpha(int(255 * 0.35))
                    p.setPen(QPen(linec, 1))
                    p.drawLine(QPointF(xs[lo_i], plot.top()), QPointF(xs[lo_i], plot.bottom()))
                    p.drawLine(QPointF(xs[hi_i], plot.top()), QPointF(xs[hi_i], plot.bottom()))
            except Exception:
                pass

            # glow + stroke
            glow_col = QColor(self._accent)
            glow_col.setAlpha(int(255 * 0.18))
            p.setPen(QPen(glow_col, 5, Qt.SolidLine, Qt.RoundCap, Qt.RoundJoin))
            p.drawPath(path)
            p.setPen(QPen(QColor(self._accent), 2.2, Qt.SolidLine, Qt.RoundCap, Qt.RoundJoin))
            p.drawPath(path)

            # Hover marker
            if self._hover_idx is not None and 0 <= int(self._hover_idx) < n:
                hi_idx = int(self._hover_idx)
                hx = xs[hi_idx]
                hy = ys[hi_idx]

                vcol = QColor(self._accent)
                vcol.setAlpha(int(255 * 0.22))
                p.setPen(QPen(vcol, 1))
                p.drawLine(QPointF(hx, plot.top()), QPointF(hx, plot.bottom()))

                glow = QColor(self._accent)
                glow.setAlpha(int(255 * 0.25))
                p.setPen(Qt.NoPen)
                p.setBrush(glow)
                p.drawEllipse(QPointF(hx, hy), 7.0, 7.0)
                p.setBrush(QColor(self._accent))
                p.drawEllipse(QPointF(hx, hy), 4.0, 4.0)
                p.setBrush(QColor('#0B1220'))
                p.drawEllipse(QPointF(hx, hy), 1.6, 1.6)

            # labels X
            if self._labels and len(self._labels) == n:
                max_lbl = 6
                step = max(1, int((n - 1) / (max_lbl - 1))) if n > 1 else 1
                idxs = list(range(0, n, step))
                if idxs and idxs[-1] != n - 1:
                    idxs.append(n - 1)

                p.setFont(QFont('Verdana', 8))
                p.setPen(QColor(159, 182, 214, int(255 * 0.72)))
                for i in idxs:
                    lab = str(self._labels[i])
                    if len(lab) > 8:
                        lab = lab[-8:]
                    x = xs[i]
                    rr = QRectF(x - 42, plot.bottom() + 6, 84, 16)
                    p.drawText(rr, Qt.AlignHCenter | Qt.AlignVCenter, lab)

            # último valor
            try:
                last = float(vals[-1])
                p.setPen(QColor('#E8F4FF'))
                p.setFont(QFont('Verdana', 9, QFont.Bold))
                txt = _format_brl(last).replace('R$ ', '')
                p.drawText(plot.adjusted(0, -6, -6, 0), int(Qt.AlignTop | Qt.AlignRight), txt)
            except Exception:
                pass

        finally:
            p.end()


def _norm_txt(s: str) -> str:
    try:
        import unicodedata
        s = unicodedata.normalize('NFD', s or '')
        s = ''.join(ch for ch in s if unicodedata.category(ch) != 'Mn')
    except Exception:
        s = s or ''
    return (s or '').lower()


def _infer_operacao_from_path(path: str):
    t = _norm_txt(path)
    # tenta reconhecer os nomes usados na estrutura
    if 'notas fiscais de entrada' in t or re.search(r'(^|\\|/)(entrada)(\\|/|$)', t):
        return 'Entrada'
    if 'notas fiscais de saida' in t or 'notas fiscais de saída' in t or re.search(r'(^|\\|/)(saida|saída)(\\|/|$)', t):
        return 'Saída'
    return None


# =============================================================================
# CFOP – classificação para auditoria de valores (Dashboard)
#
# - Conta TODOS os XMLs (quantidade 55/65) como antes.
# - Valores:
#     * Saída (faturamento): soma apenas CFOPs classificados como "venda".
#     * Entrada (despesa): soma tudo que for Entrada, EXCETO CFOPs "ignorar" (ex.: 5949).
#
# O arquivo esperado fica na pasta JSON do app (data/json): cfop.json.
# (Aceita também FLP.json por compatibilidade.)
# =============================================================================

_CFOP_CACHE: Optional[dict] = None


def _load_cfop_map() -> dict:
    """Carrega e cacheia o JSON de CFOPs (cfop.json/FLP.json)."""
    global _CFOP_CACHE
    if isinstance(_CFOP_CACHE, dict) and _CFOP_CACHE.get('cfops'):
        return _CFOP_CACHE

    cand_paths = [
        os.path.join(CONFIG_DIR, 'cfop.json'),
        os.path.join(CONFIG_DIR, 'FLP.json'),
        os.path.join(CONFIG_DIR, 'cfops.json'),
    ]

    data = {'cfops': {}}
    for p in cand_paths:
        try:
            if os.path.exists(p):
                with open(p, 'r', encoding='utf-8') as f:
                    data = json.load(f) or {'cfops': {}}
                if isinstance(data, dict) and isinstance(data.get('cfops'), dict) and data['cfops']:
                    break
        except Exception:
            continue

    # garante estrutura
    if not isinstance(data, dict):
        data = {'cfops': {}, 'flags': {}}
    if not isinstance(data.get('cfops'), dict):
        data['cfops'] = {}
    flags = data.get('flags')
    if not isinstance(flags, dict):
        data['flags'] = {}

    _CFOP_CACHE = data
    return _CFOP_CACHE




def _get_cfop_json_path() -> str:
    """Retorna o caminho do arquivo JSON de CFOPs que o app deve salvar."""
    cand_paths = [
        os.path.join(CONFIG_DIR, 'cfop.json'),
        os.path.join(CONFIG_DIR, 'FLP.json'),
        os.path.join(CONFIG_DIR, 'cfops.json'),
    ]
    for p in cand_paths:
        try:
            if os.path.exists(p):
                return p
        except Exception:
            continue
    return cand_paths[0]
def _norm_cnpj(v: str) -> str:
    return cnpj_somente_digitos(v or '')


def _norm_cfop(v: str) -> str:
    v = (v or '').strip()
    v = re.sub(r'[^0-9]', '', v)
    # CFOP é 4 dígitos; se vier maior (raríssimo), mantemos assim mesmo
    return v


def _cfop_classe(cfop_codes: List[str]) -> str:
    """Retorna a classe agregada do documento a partir da lista de CFOPs."""
    cfop_map = _load_cfop_map().get('cfops', {})
    classes = []
    for c in (cfop_codes or []):
        c = _norm_cfop(c)
        if not c:
            continue
        info = cfop_map.get(c) or {}
        cls = (info.get('classe') or '').strip().lower()
        if cls:
            classes.append(cls)

    # prioridade: venda > despesa > outro > ignorar
    if 'venda' in classes:
        return 'venda'
    if 'despesa' in classes:
        return 'despesa'
    if 'outro' in classes:
        return 'outro'
    if 'ignorar' in classes:
        return 'ignorar'
    return 'outro'


def _infer_direcao_emitente_por_cfop(cfop_codes: List[str]) -> Optional[str]:
    """Infere se a operação (do ponto de vista do EMITENTE) é Entrada ou Saída pelo dígito do CFOP."""
    if not cfop_codes:
        return None
    saida = 0
    entrada = 0
    for c in cfop_codes:
        c = _norm_cfop(c)
        if not c:
            continue
        d = c[0]
        if d in {'5', '6', '7'}:
            saida += 1
        elif d in {'1', '2', '3'}:
            entrada += 1
    if saida == 0 and entrada == 0:
        return None
    return 'Saída' if saida >= entrada else 'Entrada'


def _norm_ie(v: str) -> str:
    v = (v or '').strip()
    # IE pode ter pontos/hífen etc
    return re.sub(r'[^0-9A-Za-z]', '', v).upper()


def _note_has_same_emit_dest(info: dict) -> bool:
    """True se o emitente e destinatário forem o mesmo (IE ou CNPJ)."""
    if not isinstance(info, dict):
        return False
    ie_emit = _norm_ie(info.get('ieEmit') or '')
    ie_dest = _norm_ie(info.get('ieDest') or '')
    if ie_emit and ie_dest and ie_emit == ie_dest:
        return True
    cnpj_emit = _norm_cnpj(info.get('cnpjEmit') or '')
    cnpj_dest = _norm_cnpj(info.get('cnpjDest') or '')
    if cnpj_emit and cnpj_dest and cnpj_emit == cnpj_dest:
        return True
    return False


def _infer_operacao_empresa(info: dict, empresa_ie: str, empresa_cnpj: str, path_hint: str | None):
    """Infere Entrada/Saída **do ponto de vista da empresa selecionada**.

    Regras (ordem):
      1) Identificação da empresa no XML por IE (prioridade) ou CNPJ.
      2) Direção do emitente por CFOP (mais confiável) e, se faltar, tpNF.
      3) Se não der pra identificar a empresa, usa tpNF e/ou dica de pasta.

    Observação importante:
      - tpNF e CFOP refletem o ponto de vista do *emitente*.
      - Se a empresa for o destinatário, a operação é o oposto.
    """

    ie_ref = _norm_ie(empresa_ie)
    cnpj_ref = _norm_cnpj(empresa_cnpj)

    ie_emit = _norm_ie(info.get('ieEmit') or '')
    ie_dest = _norm_ie(info.get('ieDest') or '')
    cnpj_emit = _norm_cnpj(info.get('cnpjEmit') or '')
    cnpj_dest = _norm_cnpj(info.get('cnpjDest') or '')

    # Quem é a empresa no XML?
    is_emit = False
    is_dest = False

    if ie_ref:
        is_emit = bool(ie_emit and ie_emit == ie_ref)
        is_dest = bool(ie_dest and ie_dest == ie_ref)

    if (not is_emit and not is_dest) and cnpj_ref:
        is_emit = bool(cnpj_emit and cnpj_emit == cnpj_ref)
        is_dest = bool(cnpj_dest and cnpj_dest == cnpj_ref)

    # Regra prioritária: se a empresa só é DESTINATÁRIA, trata como Entrada
    # independentemente de CFOP/tpNF, evitando classificar como Saída quando
    # outra empresa emite um CFOP de venda para nós.
    if is_dest and not is_emit:
        return 'Entrada'

    # Direção da operação do ponto de vista do EMITENTE
    cfops = info.get('cfops') or []
    dir_emit = _infer_direcao_emitente_por_cfop(cfops)

    # fallback tpNF (0 = entrada do emitente, 1 = saída do emitente)
    tp = (info.get('tpNF') or '').strip()
    if dir_emit is None:
        if tp == '0':
            dir_emit = 'Entrada'
        elif tp == '1':
            dir_emit = 'Saída'

    # Se a empresa for o emitente, ela segue a direção do emitente.
    if is_emit and dir_emit in ('Entrada', 'Saída'):
        return dir_emit

    # Se a empresa for o destinatário, a operação é o inverso.
    if is_dest and dir_emit in ('Entrada', 'Saída'):
        return 'Saída' if dir_emit == 'Entrada' else 'Entrada'

    # Se não bateu IE/CNPJ, tenta dica de pasta (se existir)
    if path_hint in ('Entrada', 'Saída'):
        return path_hint

    # Último fallback: tpNF como estava (assumindo que o XML está "do ponto de vista" do usuário)
    if tp == '0':
        return 'Entrada'
    if tp == '1':
        return 'Saída'
    return None

# =============================================================================
# DashboardScanWorker - Thread para varrer pastas e agregar métricas
# =============================================================================
class DashboardScanWorker(QThread):
    """Varre pastas, lê ZIP/XML e agrega métricas para o dashboard.

    Ajustes importantes:
      - Deduplica notas pela chave de acesso (quando disponível) para evitar contagem duplicada
        quando o mesmo XML existe solto e também dentro de ZIP, ou quando existem cópias.
      - Pode operar em modo recursivo (os.walk) ou estrito (apenas arquivos diretos do diretório).
        * recursive=True  : usado quando o usuário escolhe "varrer qualquer subpasta".
        * recursive=False : usado quando o usuário escolhe "respeitar a estrutura atual".

      - Pode opcionalmente considerar XMLs duplicados (sem deduplicar por chave).
        Isso altera tanto as contagens quanto os somatórios, pois o documento passa a
        ser processado mais de uma vez.
    """

    resultReady = Signal(object)

    def __init__(
        self,
        scan_dirs,
        dt_ini,
        dt_fim,
        operacao,
        modelo,
        empresa_ie=None,
        empresa_cnpj=None,
        consider_duplicates: bool = False,
        recursive: bool = True,
        count_real_qty: bool = False,
        parent=None,
    ):
        super().__init__(parent)
        self.scan_dirs = list(scan_dirs or [])
        self.dt_ini = dt_ini
        self.dt_fim = dt_fim
        self.operacao = operacao  # 'Entrada'/'Saída'/'Ambos'
        self.modelo = modelo      # 'Todos'/'55'/'65'
        self.empresa_ie = empresa_ie or ''
        self.empresa_cnpj = empresa_cnpj or ''
        self.consider_duplicates = bool(consider_duplicates)
        self.recursive = bool(recursive)
        self.count_real_qty = bool(count_real_qty)

    def run(self):
        log_fn = getattr(self, "_log", None)
        try:
            self._run_impl(log_fn)
        except Exception as e:
            try:
                if callable(log_fn):
                    log_fn(f"[ERRO] 💥 Falha ao atualizar dashboard: {e}")
            except Exception:
                pass
            try:
                traceback.print_exc()
            except Exception:
                pass
            try:
                self.resultReady.emit({
                    'c55': 0,
                    'c65': 0,
                    'total_xml': 0,
                    'soma_entrada': 0.0,
                    'soma_saida': 0.0,
                    'serie_metric': '',
                    'serie_labels': [],
                    'serie_values': [],
                    'cfops_fora_param': [],
                })
            except Exception:
                pass

    def _run_impl(self, log_fn):
        cfop_data = _load_cfop_map()
        cfop_map = cfop_data.get('cfops', {})
        cfop_flags = cfop_data.get('flags') or {}
        ignore_same_by_op = {
            'Entrada': bool(cfop_flags.get('despesa_ignore_same_party')),
            'Saída': bool(cfop_flags.get('venda_ignore_same_party')),
        }

        def _should_ignore_same_party(info: dict, op_real: Optional[str]) -> bool:
            if not (
                op_real in ignore_same_by_op
                and ignore_same_by_op[op_real]
            ):
                return False
            # Ignora quando o próprio XML traz emitente e destinatário iguais
            # (IE ou CNPJ), alinhado ao rótulo do checkbox.
            return _note_has_same_emit_dest(info)

        c55 = c65 = 0
        total_xml = 0
        raw_c55 = raw_c65 = 0
        soma_entrada = 0.0
        soma_saida = 0.0
        daily_in = {}   # date -> soma vNF (Entrada)
        daily_out = {}  # date -> soma vNF (Saída)
        # Agregação por data para Supabase (XML, NF, NFC)
        daily_xml_count = {}  # date -> xml_count
        daily_nf_count = {}  # date -> nf_count (modelo 55)
        daily_nfc_count = {}  # date -> nfc_count (modelo 65)
        daily_total_amount = {}  # date -> total_amount

        # CFOPs encontrados no conjunto (para avisar quando não estiverem parametrizados)
        seen_cfops = set()

        # Deduplicação (nota pode aparecer como XML solto e também dentro de ZIP)
        # Se consider_duplicates=True, não deduplica e processa tudo.
        seen_docs = set() if not self.consider_duplicates else None

        def _in_range(d):
            if d is None:
                return True
            if self.dt_ini and d < self.dt_ini:
                return False
            if self.dt_fim and d > self.dt_fim:
                return False
            return True

        def _uid_from(info: dict, xml_bytes: bytes) -> str:
            try:
                chave = (info.get('chave') or '').strip()
                chave = re.sub(r'\D', '', chave)
                if len(chave) == 44:
                    return 'K' + chave
            except Exception:
                pass
            try:
                return 'H' + hashlib.sha1(xml_bytes or b'').hexdigest()
            except Exception:
                return 'H'

        def _iter_dir(base: str):
            if self.recursive:
                for root_dir, _, files in os.walk(base):
                    yield root_dir, files
            else:
                try:
                    files = list(os.listdir(base))
                except Exception:
                    files = []
                yield base, files

        for base in self.scan_dirs:
            if self.isInterruptionRequested():
                return
            if not base or not os.path.isdir(base):
                continue

            for root_dir, files in _iter_dir(base):
                if self.isInterruptionRequested():
                    return
                for fn in files:
                    if self.isInterruptionRequested():
                        return
                    low = (fn or '').lower()
                    path = os.path.join(root_dir, fn)

                    # XML solto
                    if low.endswith('.xml'):
                        try:
                            xml_bytes = pathlib.Path(fs_path(path)).read_bytes()
                        except Exception:
                            continue
                        if self.isInterruptionRequested():
                            return

                        info = extrair_info_nfe(xml_bytes)
                        model = info.get('model')
                        vnf = float(info.get('vNF') or 0.0)
                        dt = info.get('dtEmi')

                        if model not in ('55', '65'):
                            continue
                        if self.modelo in ('55', '65') and model != self.modelo:
                            continue
                        if not _in_range(dt):
                            continue

                        # coleta CFOPs (mesmo que este XML não entre nos valores)
                        for _c in (info.get('cfops') or []):
                            _c = _norm_cfop(_c)
                            if _c:
                                seen_cfops.add(_c)

                        path_hint = _infer_operacao_from_path(root_dir)
                        op_real = _infer_operacao_empresa(info, self.empresa_ie, self.empresa_cnpj, path_hint)

                        # Contagem "real" de XMLs (antes da deduplicação)
                        if model == '55':
                            raw_c55 += 1
                        elif model == '65':
                            raw_c65 += 1

                        if _should_ignore_same_party(info, op_real):
                            if callable(log_fn):
                                chave = (info.get('chave') or '').strip()
                                resumo = chave or "documento sem chave"
                                log_fn(
                                    f"[INFO] ☑️ Ignorando {resumo} (emitente=destinatário) "
                                    f"para op. {op_real or 'desconhecida'}."
                                )
                            continue

                        if self.operacao == 'Entrada' and op_real != 'Entrada':
                            continue
                        if self.operacao == 'Saída' and op_real != 'Saída':
                            continue

                        if not self.consider_duplicates:
                            uid = _uid_from(info, xml_bytes)
                            if uid in seen_docs:
                                continue
                            seen_docs.add(uid)

                        # Auditoria de valores por CFOP (não altera contagem de XMLs)
                        classe_doc = _cfop_classe(info.get('cfops') or [])

                        total_xml += 1
                        if model == '55':
                            c55 += 1
                        else:
                            c65 += 1

                        # Agregação por data para Supabase (XMLs soltos)
                        if dt is not None:
                            daily_xml_count[dt] = daily_xml_count.get(dt, 0) + 1
                            if model == '55':
                                daily_nf_count[dt] = daily_nf_count.get(dt, 0) + 1
                            elif model == '65':
                                daily_nfc_count[dt] = daily_nfc_count.get(dt, 0) + 1
                            daily_total_amount[dt] = daily_total_amount.get(dt, 0.0) + vnf

                        # Valores:
                        # - Entrada: soma tudo, exceto CFOPs marcados como "ignorar" (ex.: 5949)
                        # - Saída: soma apenas CFOPs classificados como "venda"
                        if op_real == 'Entrada':
                            if classe_doc != 'ignorar':
                                soma_entrada += vnf
                        elif op_real == 'Saída':
                            if classe_doc == 'venda':
                                soma_saida += vnf

                        if dt is not None:
                            if op_real == 'Entrada':
                                if classe_doc != 'ignorar':
                                    daily_in[dt] = daily_in.get(dt, 0.0) + vnf
                            elif op_real == 'Saída':
                                if classe_doc == 'venda':
                                    daily_out[dt] = daily_out.get(dt, 0.0) + vnf
                        continue

                    # ZIP com XML
                    if not low.endswith('.zip'):
                        continue

                    try:
                        with zipfile.ZipFile(fs_path(path), 'r') as zf:
                            for name in zf.namelist():
                                if self.isInterruptionRequested():
                                    return
                                if not name.lower().endswith('.xml'):
                                    continue
                                try:
                                    xml_bytes = zf.read(name)
                                except Exception:
                                    continue
                                if self.isInterruptionRequested():
                                    return

                                info = extrair_info_nfe(xml_bytes)
                                model = info.get('model')
                                vnf = float(info.get('vNF') or 0.0)
                                dt = info.get('dtEmi')

                                if model not in ('55', '65'):
                                    continue
                                if self.modelo in ('55', '65') and model != self.modelo:
                                    continue
                                if not _in_range(dt):
                                    continue

                                # coleta CFOPs (mesmo que este XML não entre nos valores)
                                for _c in (info.get('cfops') or []):
                                    _c = _norm_cfop(_c)
                                    if _c:
                                        seen_cfops.add(_c)

                                path_hint = _infer_operacao_from_path(root_dir)
                                op_real = _infer_operacao_empresa(info, self.empresa_ie, self.empresa_cnpj, path_hint)

                                # Contagem "real" de XMLs (antes da deduplicação)
                                if model == '55':
                                    raw_c55 += 1
                                elif model == '65':
                                    raw_c65 += 1

                                if _should_ignore_same_party(info, op_real):
                                    if callable(log_fn):
                                        chave = (info.get('chave') or '').strip()
                                        resumo = chave or "documento sem chave"
                                        log_fn(
                                            f"[INFO] ☑️ Ignorando {resumo} (emitente=destinatário) "
                                            f"para op. {op_real or 'desconhecida'}."
                                        )
                                    continue

                                if self.operacao == 'Entrada' and op_real != 'Entrada':
                                    continue
                                if self.operacao == 'Saída' and op_real != 'Saída':
                                    continue

                                if not self.consider_duplicates:
                                    uid = _uid_from(info, xml_bytes)
                                    if uid in seen_docs:
                                        continue
                                    seen_docs.add(uid)

                                # Auditoria de valores por CFOP (não altera contagem de XMLs)
                                classe_doc = _cfop_classe(info.get('cfops') or [])

                                total_xml += 1
                                if model == '55':
                                    c55 += 1
                                else:
                                    c65 += 1

                                if op_real == 'Entrada':
                                    if classe_doc != 'ignorar':
                                        soma_entrada += vnf
                                elif op_real == 'Saída':
                                    if classe_doc == 'venda':
                                        soma_saida += vnf

                                # Agregação por data para Supabase
                                if dt is not None:
                                    daily_xml_count[dt] = daily_xml_count.get(dt, 0) + 1
                                    if model == '55':
                                        daily_nf_count[dt] = daily_nf_count.get(dt, 0) + 1
                                    elif model == '65':
                                        daily_nfc_count[dt] = daily_nfc_count.get(dt, 0) + 1
                                    daily_total_amount[dt] = daily_total_amount.get(dt, 0.0) + vnf
                                    
                                    if op_real == 'Entrada':
                                        if classe_doc != 'ignorar':
                                            daily_in[dt] = daily_in.get(dt, 0.0) + vnf
                                    elif op_real == 'Saída':
                                        if classe_doc == 'venda':
                                            daily_out[dt] = daily_out.get(dt, 0.0) + vnf
                    except Exception:
                        continue

        # monta série conforme operação (mais intuitivo)
        all_dates = sorted(set(daily_in.keys()) | set(daily_out.keys()))

        if self.operacao == 'Entrada':
            metric = 'Despesa diária (R$) • Entrada'
            serie = [float(daily_in.get(d, 0.0) or 0.0) for d in all_dates]
        elif self.operacao == 'Saída':
            metric = 'Faturamento diário (R$) • Saída'
            serie = [float(daily_out.get(d, 0.0) or 0.0) for d in all_dates]
        else:
            metric = 'Resultado diário (R$) • Saída - Entrada'
            serie = [float(daily_out.get(d, 0.0) or 0.0) - float(daily_in.get(d, 0.0) or 0.0) for d in all_dates]

        labels = [d.strftime('%d/%m') for d in all_dates]
        values = [float(x or 0.0) for x in serie]

        # CFOPs fora da parametrização: ausentes no JSON ou com classe 'outro'
        try:
            _cfop_map = (_load_cfop_map().get('cfops') or {})
        except Exception:
            _cfop_map = {}

        cfops_fora_param = []
        try:
            for _code in sorted(seen_cfops):
                _info = (_cfop_map.get(_code) or {})
                _cls = (_info.get('classe') or '').strip().lower()
                if _cls not in {'venda', 'despesa', 'ignorar'}:
                    _desc = (_info.get('desc') or '').strip()
                    cfops_fora_param.append(f"{_code} - {_desc}" if _desc else _code)
        except Exception:
            pass

        emit_c55 = raw_c55 if self.count_real_qty else c55
        emit_c65 = raw_c65 if self.count_real_qty else c65
        emit_total = (raw_c55 + raw_c65) if self.count_real_qty else total_xml

        # Prepara dados agregados do PERÍODO COMPLETO para Supabase
        # UMA LINHA por período (não por data) com totais agregados
        # Soma todos os dados do período em uma única linha
        total_xml_periodo = sum(daily_xml_count.values())
        total_nf_periodo = sum(daily_nf_count.values())
        total_nfc_periodo = sum(daily_nfc_count.values())
        total_faturamento_periodo = sum(daily_out.values())
        total_despesa_periodo = sum(daily_in.values())
        total_resultado_periodo = total_faturamento_periodo - total_despesa_periodo
        
        # Cria uma única entrada para o período (usa data inicial do período)
        periodo_data = {
            'xml_count': total_xml_periodo,
            'nf_count': total_nf_periodo,
            'nfc_count': total_nfc_periodo,
            'faturamento': total_faturamento_periodo,
            'despesa': total_despesa_periodo,
            'resultado': total_resultado_periodo,
            'total_amount': total_faturamento_periodo + total_despesa_periodo
        }
        
        # Usa a data inicial do período como identificador
        # Se dt_ini não estiver disponível, usa a primeira data encontrada
        if self.dt_ini:
            periodo_date = self.dt_ini
        elif daily_in or daily_out:
            periodo_date = sorted(set(daily_in.keys()) | set(daily_out.keys()))[0]
        elif daily_xml_count:
            periodo_date = sorted(daily_xml_count.keys())[0]
        else:
            periodo_date = None
        
        daily_data_for_supabase = {}
        if periodo_date:
            daily_data_for_supabase[periodo_date] = periodo_data

        self.resultReady.emit({
            'c55': emit_c55,
            'c65': emit_c65,
            'total_xml': emit_total,
            'soma_entrada': soma_entrada,
            'soma_saida': soma_saida,
            'serie_metric': metric,
            'serie_labels': labels,
            'serie_values': values,
            'cfops_fora_param': cfops_fora_param,
            # Dados agregados por data para Supabase
            'daily_data': daily_data_for_supabase,
            # Dados diários de evolução (para enviar ao Supabase)
            # Mantém como date objects - serão serializados automaticamente pelo Signal
            'daily_in': daily_in,
            'daily_out': daily_out,
            'daily_xml_count': daily_xml_count,
            'daily_nf_count': daily_nf_count,
            'daily_nfc_count': daily_nfc_count,
            'empresa_cnpj': self.empresa_cnpj,
            'empresa_ie': self.empresa_ie,
            'dt_ini': self.dt_ini,
            'dt_fim': self.dt_fim,
        })


# =============================================================================
# Receita (CNPJ -> Nome) – usado no cadastro de empresas (sem travar UI)
# =============================================================================

class CNPJLookupWorker(QThread):
    """Busca o nome da empresa pelo CNPJ em uma API pública.

    Observação: a Receita Federal não disponibiliza uma API pública simples.
    Este worker usa um endpoint público compatível (ex.: receitaws), e já
    trata erros/limites. A chamada roda fora da UI.
    """

    finishedLookup = Signal(object)  # {'ok': bool, 'nome': str, 'erro': str}

    def __init__(self, cnpj_digits: str, parent=None):
        super().__init__(parent)
        self.cnpj_digits = cnpj_somente_digitos(cnpj_digits or '')

    def run(self):
        res = {'ok': False, 'nome': '', 'erro': ''}
        cnpj = self.cnpj_digits
        if len(cnpj) != 14:
            res['erro'] = 'CNPJ inválido.'
            self.finishedLookup.emit(res)
            return

        # Endpoint público (pode responder 429/limite)
        url = f"https://www.receitaws.com.br/v1/cnpj/{cnpj}"
        try:
            req = urllib.request.Request(
                url,
                headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) DashboardCFOP/1.0',
                    'Accept': 'application/json',
                },
                method='GET',
            )
            with urllib.request.urlopen(req, timeout=12) as resp:
                raw = resp.read() or b''
            data = json.loads(raw.decode('utf-8', errors='replace') or '{}')

            # Padrão do receitaws: {'status':'OK','nome':'...'} ou {'status':'ERROR','message':'...'}
            status = (data.get('status') or '').upper()
            if status == 'ERROR':
                msg = (data.get('message') or data.get('motivo') or 'Falha ao consultar a Receita.').strip()
                res['erro'] = msg
            else:
                nome = (data.get('nome') or '').strip()
                if not nome:
                    nome = (data.get('fantasia') or '').strip()
                if nome:
                    res['ok'] = True
                    res['nome'] = nome
                else:
                    res['erro'] = 'Não foi possível obter o nome pelo CNPJ.'
        except urllib.error.HTTPError as e:
            code = getattr(e, 'code', None)
            if code == 429:
                res['erro'] = 'Limite de consultas excedido. Aguarde um pouco e tente novamente.'
            else:
                res['erro'] = f'Erro HTTP ao consultar: {code}' if code else 'Erro HTTP ao consultar.'
        except Exception:
            res['erro'] = 'Falha na consulta do CNPJ (sem conexão ou serviço indisponível).'

        self.finishedLookup.emit(res)


# =============================================================================
# Interface – MyCompactUI
# =============================================================================
# INTERFACE GRÁFICA PRINCIPAL
# =============================================================================

# =============================================================================
# MyCompactUI - Janela principal da aplicação
# =============================================================================
class MyCompactUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Automação Sefaz NFe (GO)")
        self.worker_thread: Optional[AutomationWorker] = None
        self._tray_icon: Optional[QSystemTrayIcon] = None
        self._quitting_from_tray = False
        self.robot_id: Optional[str] = None
        self.dashboard_company_cache: List[Dict[str, Any]] = []
        self.runtime_paths: Dict[str, str] = {}

        # flags de controle de execução
        self.automation_running = False
        self._stop_requested_by_user = False

        # Carrega config (paths + empresas + seleções)
        self.config = carregar_config()
        self.paths = self.config.get("paths", {})
        # {ie: {"display_name": ..., "cnpj": ""}}
        local_empresas = self.config.get("empresas", {})
        self._local_empresas_metadata = dict(local_empresas)
        # Com .env em disco + Supabase, a lista vem do dashboard; em modo cliente (só JSON) usa sempre config.json.
        use_dashboard_empresas = SUPABASE_AVAILABLE and not SEFAZ_LOCAL_JSON_CLIENT_MODE
        if use_dashboard_empresas:
            self.empresas = {}
            self.produtos = []
        else:
            self.empresas = local_empresas
            self.produtos = list(self.empresas.keys())

        if use_dashboard_empresas:
            print("[SUPABASE] ℹ️ Sefaz Xml em modo dashboard: empresas locais não serão sincronizadas para o banco.")
        elif SEFAZ_LOCAL_JSON_CLIENT_MODE and local_empresas:
            print(f"[MODO CLIENTE LOCAL] Empresas carregadas do config.json ({len(local_empresas)}).")

        # rate-limit simples p/ consulta de CNPJ (evita travas e 429)
        self._cnpj_lookup_calls: List[float] = []
        self.folder_structure = normalizar_estrutura_pastas(self.paths.get("estrutura_pastas"))
        self.paths["estrutura_pastas"] = self.folder_structure

        # UI state: lista de empresas (precisa existir antes de qualquer sync inicial)
        self.empresa_items = []
        self.empresa_items_layout = None
        self._empresa_list_has_stretch = False

        # NOVO: seleção de empresas (marcadas + diário) vinda do config.json
        selecoes = self.config.get("selecoes_empresas", {}) or {}
        marcadas_ini = selecoes.get("marcadas", [])
        diario_ini = selecoes.get("diario", [])

        # guarda como sets internos já normalizados (só dígitos)
        self._cfg_empresas_marcadas = set(ie_somente_digitos(ie) for ie in marcadas_ini)
        self._cfg_empresas_diario = set(ie_somente_digitos(ie) for ie in diario_ini)

        self._stop_requested_by_user = False
        self._scheduled_pending = False
        self._scheduled_start_dt = None
        self.schedule_timer = QTimer(self)
        self.schedule_timer.timeout.connect(self._update_schedule_countdown)
        self.robot_presence_timer = QTimer(self)
        self.robot_presence_timer.timeout.connect(self._refresh_robot_presence)
        self.dashboard_sync_timer = QTimer(self)
        self.dashboard_sync_timer.timeout.connect(self._sync_dashboard_runtime_state)
        # Poll da fila do agendador (execution_requests) criada pelo site
        self.queue_poll_timer = QTimer(self)
        self.queue_poll_timer.timeout.connect(self._poll_execution_queue)
        self._active_queue_job: Optional[Dict[str, Any]] = None
        self._queue_supabase_url: Optional[str] = None
        self._queue_supabase_key: Optional[str] = None

        # Garante janela redimensionável (alguns ambientes no Windows podem herdar hint de tamanho fixo)
        try:
            flags = self.windowFlags()
            if hasattr(Qt, 'MSWindowsFixedSizeDialogHint'):
                flags &= ~Qt.MSWindowsFixedSizeDialogHint
            # garante botões de minimizar/maximizar e permite resize
            flags |= Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint | Qt.WindowSystemMenuHint
            self.setWindowFlags(flags)
        except Exception:
            pass

        # tamanho inicial, mas com mínimo baixo o suficiente para permitir reduzir/aumentar
        self.resize(980, 780)
        self.setMinimumSize(760, 600)
        self.setStyleSheet("background: qlineargradient(x1:0,y1:0,x2:1,y2:1, stop:0 #0B1220, stop:1 #0F172A);")

        if os.path.exists(ICO_PATH):
            self.setWindowIcon(QIcon(ICO_PATH))
        self._setup_tray_icon()

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(8, 8, 8, 8)
        main_layout.setSpacing(6)

        # Top bar
        top_bar = QFrame()
        top_bar.setStyleSheet("background: qlineargradient(x1:0,y1:0,x2:1,y2:1, stop:0 #12304E, stop:1 #0F1E35); border-radius: 5px;")
        top_bar_layout = QHBoxLayout(top_bar)
        top_bar_layout.setContentsMargins(10, 5, 10, 5)
        title_label = QLabel("Seleção de Inscrições Estaduais")
        title_label.setStyleSheet(
            "color: #E8F4FF; font-family: Verdana; font-size: 10pt; font-weight: bold;"
        )
        top_bar_layout.addWidget(title_label, alignment=Qt.AlignCenter)
        main_layout.addWidget(top_bar)

        # Opções (Entrada/Saída/Ambos + datas)
        options_frame = QFrame()
        options_frame.setStyleSheet("background: qlineargradient(x1:0,y1:0,x2:1,y2:1, stop:0 #0F172A, stop:1 #0F1E35); border-radius: 5px;")
        options_layout = QHBoxLayout(options_frame)
        options_layout.setContentsMargins(8, 8, 8, 8)
        options_layout.setSpacing(20)

        radio_frame = QWidget()
        radio_layout = QVBoxLayout(radio_frame)
        radio_layout.setContentsMargins(0, 0, 0, 0)

        self.rbEntrada = QRadioButton("Entrada")
        self.rbEntrada.setStyleSheet(
            "QRadioButton { color: #E8F4FF; font-family: Verdana; font-size: 10pt; }"
        )
        self.rbSaida = QRadioButton("Saída")
        self.rbSaida.setStyleSheet(
            "QRadioButton { color: #E8F4FF; font-family: Verdana; font-size: 10pt; }"
        )
        self.rbAmbos = QRadioButton("Ambos")
        self.rbAmbos.setStyleSheet(
            "QRadioButton { color: #E8F4FF; font-family: Verdana; font-size: 10pt; }"
        )
        self.rbAmbos.setChecked(True)

        self.group_operacao = QButtonGroup()
        self.group_operacao.addButton(self.rbEntrada)
        self.group_operacao.addButton(self.rbSaida)
        self.group_operacao.addButton(self.rbAmbos)

        radio_layout.addWidget(self.rbEntrada)
        radio_layout.addWidget(self.rbSaida)
        radio_layout.addWidget(self.rbAmbos)

        date_frame = QWidget()
        date_layout = QVBoxLayout(date_frame)
        date_layout.setContentsMargins(0, 0, 0, 0)
        date_layout.setSpacing(2)

        row_data_inicial = QHBoxLayout()
        row_data_inicial.setSpacing(5)
        row_data_inicial.addStretch(3)
        lbl_data_inicial = QLabel("Data Inicial:")
        lbl_data_inicial.setStyleSheet(
            "color: #E8F4FF; font-family: Verdana; font-size: 10pt;"
        )
        self.line_data_inicial = QLineEdit()
        self.line_data_inicial.setInputMask("00/00/0000")
        self.line_data_inicial.setFixedWidth(120)
        self.line_data_inicial.setStyleSheet(
            "QLineEdit { background-color: #0F1E35; color: #E8F4FF; "
            "border-radius: 6px; font-family: Verdana; font-size: 10pt; padding: 4px; }"
        )
        row_data_inicial.addWidget(lbl_data_inicial)
        row_data_inicial.addWidget(self.line_data_inicial)
        date_layout.addLayout(row_data_inicial)

        row_data_final = QHBoxLayout()
        row_data_final.setSpacing(5)
        row_data_final.addStretch(3)
        lbl_data_final = QLabel("Data Final:")
        lbl_data_final.setStyleSheet(
            "color: #E8F4FF; font-family: Verdana; font-size: 10pt;"
        )
        self.line_data_final = QLineEdit()
        self.line_data_final.setInputMask("00/00/0000")
        self.line_data_final.setFixedWidth(120)
        self.line_data_final.setStyleSheet(
            "QLineEdit { background-color: #0F1E35; color: #E8F4FF; "
            "border-radius: 6px; font-family: Verdana; font-size: 10pt; padding: 4px; }"
        )
        row_data_final.addWidget(lbl_data_final)
        row_data_final.addWidget(self.line_data_final)
        date_layout.addLayout(row_data_final)

        # 👉 Preencher automaticamente com o mês anterior inteiro
        hoje = datetime.now()
        primeiro_deste_mes = hoje.replace(day=1)
        ultimo_dia_mes_anterior = primeiro_deste_mes - timedelta(days=1)
        primeiro_dia_mes_anterior = ultimo_dia_mes_anterior.replace(day=1)

        self.line_data_inicial.setText(primeiro_dia_mes_anterior.strftime("%d/%m/%Y"))
        self.line_data_final.setText(ultimo_dia_mes_anterior.strftime("%d/%m/%Y"))

        options_layout.addWidget(radio_frame)
        options_layout.addWidget(date_frame)
        options_frame.setLayout(options_layout)
        main_layout.addWidget(options_frame)

        # -----------------------------
        # Navegação lateral (Menu) – Dashboard / Execução
        # -----------------------------
        content_frame = QFrame()
        content_frame.setStyleSheet("background: transparent;")
        content_h = QHBoxLayout(content_frame)
        content_h.setContentsMargins(0, 0, 0, 0)
        content_h.setSpacing(10)
        main_layout.addWidget(content_frame, stretch=1)

        # Sidebar
        self.sidebar = QFrame()
        self._sidebar_expanded_w = 190
        self._sidebar_collapsed_w = 64
        self._sidebar_collapsed = False
        self.sidebar.setFixedWidth(self._sidebar_expanded_w)
        self.sidebar.setStyleSheet("""
            QFrame { background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #0F1E35, stop:1 #0B1220);
                    border-radius: 14px; border: 1px solid #1F2A44; }
        """)
        sb = QVBoxLayout(self.sidebar)
        sb.setContentsMargins(12, 12, 12, 12)
        sb.setSpacing(10)

        # Header do sidebar: 'Painel' + botão (menu) para recolher/expandir
        self.sb_header = QFrame()
        self.sb_header.setObjectName('sbHeader')
        self.sb_header.setStyleSheet(
            "QFrame { color:#E8F4FF; font-family: Verdana; font-size: 10pt; font-weight: 900; "
            "padding:0px; background: rgba(18,48,78,0.22); border: 1px solid rgba(232,244,255,0.10); "
            "border-radius: 12px; }"
        )
        sb_header_h = QHBoxLayout(self.sb_header)
        self._sb_header_h = sb_header_h
        sb_header_h.setContentsMargins(10, 6, 10, 6)
        sb_header_h.setSpacing(6)

        self.lbl_sidebar_title = QLabel('Painel')
        self.lbl_sidebar_title.setObjectName('lblSidebarTitle')
        self.lbl_sidebar_title.setStyleSheet('color:#E8F4FF; background: transparent; border: none; padding:0;')

        # Botão de recolher/expandir o painel (prioriza menu.png; senão usa ☰)
        _menu_icon = load_icon('menu.png')
        if _menu_icon is not None and not _menu_icon.isNull():
            # QSS nem sempre desloca ícone em botão "icon-only"; usamos pintura manual com offset
            # Offset leve para compensar padding interno do PNG, mantendo centralização no "pill"
            self.btn_sidebar_toggle = SidebarToggleIconButton(_menu_icon, offset_x=-1, offset_y=1, parent=self.sb_header)
        else:
            self.btn_sidebar_toggle = QPushButton('☰', parent=self.sb_header)
            self.btn_sidebar_toggle.setStyleSheet('QPushButton { font-size: 15pt; font-weight: 900; }')

        self.btn_sidebar_toggle.setObjectName('btnSidebarToggle')
        self.btn_sidebar_toggle.setCursor(Qt.PointingHandCursor)
        self.btn_sidebar_toggle.setFocusPolicy(Qt.NoFocus)
        self.btn_sidebar_toggle.setFixedSize(28, 28)
        self.btn_sidebar_toggle.setIconSize(QSize(18, 18))

        # Sem linhas/bordas em volta (hover pill)
        _base_css = self.btn_sidebar_toggle.styleSheet()
        self.btn_sidebar_toggle.setStyleSheet(_base_css + """
            #btnSidebarToggle { border: none; background: transparent; color:#E8F4FF; padding: 0px; }
            #btnSidebarToggle:hover { background: transparent; }
            #btnSidebarToggle:pressed { background: transparent; }
        """)

        sb_header_h.addWidget(self.lbl_sidebar_title)
        sb_header_h.addStretch(1)
        sb_header_h.addWidget(self.btn_sidebar_toggle)

        sb.addWidget(self.sb_header)


        nav_btn_css = """
            QPushButton { text-align: left; padding: 10px 12px; border-radius: 12px;
                         background: transparent; color: #9FB6D6; border: 1px solid transparent;
                         font-family: Verdana; font-size: 9pt; font-weight: 800; }
            QPushButton:hover { background-color: rgba(18, 48, 78, 0.35); color: #E8F4FF; }
            QPushButton:checked { background-color: #0F172A; color: #E8F4FF; border: 1px solid #1F2A44; }
        """

        self._nav_btn_css_expanded = nav_btn_css
        self._nav_btn_css_collapsed = nav_btn_css + """
            QPushButton { text-align: center; padding: 10px 0px; }
        """

        self.btn_nav_dashboard = QPushButton("Dashboard")
        self.btn_nav_dashboard.setCheckable(True)
        self.btn_nav_dashboard.setCursor(Qt.PointingHandCursor)
        self.btn_nav_dashboard.setStyleSheet(nav_btn_css)
        try:
            self.btn_nav_dashboard.setIcon(load_icon("dashboard.png"))
        except Exception:
            pass

        self.btn_nav_execucao = QPushButton("Execução")
        self.btn_nav_execucao.setCheckable(True)
        self.btn_nav_execucao.setCursor(Qt.PointingHandCursor)
        self.btn_nav_execucao.setStyleSheet(nav_btn_css)
        try:
            self.btn_nav_execucao.setIcon(load_icon("iniciar.png"))
        except Exception:
            pass


        # textos completos e emojis (fallback) para o modo recolhido
        self.btn_nav_dashboard._full_text = self.btn_nav_dashboard.text()
        self.btn_nav_dashboard._collapsed_text = '📊'
        self.btn_nav_execucao._full_text = self.btn_nav_execucao.text()
        self.btn_nav_execucao._collapsed_text = '🚀'
        self._sidebar_nav_buttons = [self.btn_nav_dashboard, self.btn_nav_execucao]

        self._nav_group = QButtonGroup(self)
        self._nav_group.setExclusive(True)
        self._nav_group.addButton(self.btn_nav_dashboard, 0)
        self._nav_group.addButton(self.btn_nav_execucao, 1)

        sb.addWidget(self.btn_nav_dashboard)
        sb.addWidget(self.btn_nav_execucao)
        sb.addStretch(1)

        # Toggle do sidebar (recolher/expandir)
        try:
            self.btn_sidebar_toggle.clicked.connect(self._toggle_sidebar)
        except Exception:
            pass

        content_h.addWidget(self.sidebar)

        # Pages (stack)
        self.pages = QStackedWidget()
        self.pages.setStyleSheet("QStackedWidget { background: transparent; }")
        content_h.addWidget(self.pages, 1)

        self.page_dashboard = QWidget()
        self.page_execucao = QWidget()
        self.pages.addWidget(self.page_dashboard)
        self.pages.addWidget(self.page_execucao)

        dash_layout = QVBoxLayout(self.page_dashboard)
        dash_layout.setContentsMargins(10, 10, 10, 10)
        dash_layout.setSpacing(10)

        exec_layout = QVBoxLayout(self.page_execucao)
        exec_layout.setContentsMargins(10, 10, 10, 10)
        exec_layout.setSpacing(8)

        # Dashboard primeiro
        self._init_dashboard_ui(dash_layout)

        # troca de telas
        self.btn_nav_dashboard.clicked.connect(lambda: self._nav_set_page(0))
        self.btn_nav_execucao.clicked.connect(lambda: self._nav_set_page(1))
        self._nav_set_page(0)

        # Painel recolhido por padrão
        try:
            self._set_sidebar_collapsed(True, animate=False)
        except Exception:
            pass

        # A partir daqui, todo o restante da UI vai para a tela Execução
        main_layout = exec_layout

        # Agendamento
        agendamento_frame = QFrame()
        agendamento_frame.setStyleSheet("background-color: #0F172A; border-radius: 5px;")
        agendamento_layout = QHBoxLayout(agendamento_frame)
        agendamento_layout.setContentsMargins(8, 6, 8, 6)
        agendamento_layout.setSpacing(10)

        self.cbAgendar = QCheckBox("Agendar início")
        self.cbAgendar.toggled.connect(self.on_agendar_toggled)

        self.line_agenda_data = QLineEdit()
        self.line_agenda_data.setInputMask("00/00/0000")
        self.line_agenda_data.setFixedWidth(110)
        self.line_agenda_data.setStyleSheet(
            "QLineEdit { background-color: #0F1E35; color: #E8F4FF; "
            "border-radius: 6px; font-family: Verdana; font-size: 10pt; padding: 4px; }"
        )

        self.line_agenda_hora = QLineEdit()
        self.line_agenda_hora.setInputMask("00")
        self.line_agenda_hora.setFixedWidth(40)
        self.line_agenda_hora.setStyleSheet(
            "QLineEdit { background-color: #0F1E35; color: #E8F4FF; "
            "border-radius: 6px; font-family: Verdana; font-size: 10pt; padding: 4px; }"
        )

        self.line_agenda_min = QLineEdit()
        self.line_agenda_min.setInputMask("00")
        self.line_agenda_min.setFixedWidth(40)
        self.line_agenda_min.setStyleSheet(
            "QLineEdit { background-color: #0F1E35; color: #E8F4FF; "
            "border-radius: 6px; font-family: Verdana; font-size: 10pt; padding: 4px; }"
        )

        lbl_data = QLabel("Data:")
        lbl_data.setStyleSheet("color: #E8F4FF; font-family: Verdana; font-size: 9pt;")
        lbl_hora = QLabel("Hora:")
        lbl_hora.setStyleSheet("color: #E8F4FF; font-family: Verdana; font-size: 9pt;")
        lbl_min = QLabel("Min:")
        lbl_min.setStyleSheet("color: #E8F4FF; font-family: Verdana; font-size: 9pt;")

        self.lbl_agenda_countdown = QLabel("")
        self.lbl_agenda_countdown.setStyleSheet("color: #F1C40F; font-family: Verdana; font-size: 9pt;")

        agendamento_layout.addWidget(self.cbAgendar)
        agendamento_layout.addWidget(lbl_data)
        agendamento_layout.addWidget(self.line_agenda_data)
        agendamento_layout.addWidget(lbl_hora)
        agendamento_layout.addWidget(self.line_agenda_hora)
        agendamento_layout.addWidget(lbl_min)
        agendamento_layout.addWidget(self.line_agenda_min)
        agendamento_layout.addStretch()
        agendamento_layout.addWidget(self.lbl_agenda_countdown)

        main_layout.addWidget(agendamento_frame)
        self._init_schedule_defaults()

        # Tipo de download
        download_frame = QFrame()
        download_frame.setFrameShape(QFrame.NoFrame)
        download_frame.setFrameShadow(QFrame.Plain)
        download_frame.setStyleSheet("QFrame { background-color: #0F172A; border-radius: 5px; border: none; }")
        download_layout = QHBoxLayout(download_frame)
        download_layout.setContentsMargins(8, 4, 8, 4)
        download_layout.setSpacing(10)

        lbl_download = QLabel("Download:")
        lbl_download.setStyleSheet(
            "color: #E8F4FF; font-family: Verdana; font-size: 9pt;"
        )

        self.rbDownDocs = QRadioButton("Documentos")
        self.rbDownDocs.setStyleSheet(
            "QRadioButton { color: #E8F4FF; font-family: Verdana; font-size: 9pt; }"
        )
        self.rbDownDocs.setChecked(True)

        self.rbDownEventos = QRadioButton("Eventos")
        self.rbDownEventos.setStyleSheet(
            "QRadioButton { color: #E8F4FF; font-family: Verdana; font-size: 9pt; }"
        )
        self.rbDownAmbos = QRadioButton("Docs + Eventos")
        self.rbDownAmbos.setStyleSheet(
            "QRadioButton { color: #E8F4FF; font-family: Verdana; font-size: 9pt; }"
        )

        self.group_download = QButtonGroup()
        self.group_download.addButton(self.rbDownDocs)
        self.group_download.addButton(self.rbDownEventos)
        self.group_download.addButton(self.rbDownAmbos)

        download_layout.addWidget(lbl_download)
        download_layout.addWidget(self.rbDownDocs)
        download_layout.addWidget(self.rbDownEventos)
        download_layout.addWidget(self.rbDownAmbos)
        main_layout.addWidget(download_frame)

        # Processar por meses / dias
        self.cbProcessarPorMeses = QCheckBox("Processar por múltiplos meses")
        main_layout.addWidget(self.cbProcessarPorMeses)

        self.cbProcessarPorDias = QCheckBox("Processar por intervalo de dias")

        self.line_intervalo_dias = QLineEdit()
        self.line_intervalo_dias.setPlaceholderText("Dias por bloco")
        # ❌ tira a máscara
        # self.line_intervalo_dias.setInputMask("00")
        # ✅ usa validador numérico (1 a 99, por exemplo)
        self.line_intervalo_dias.setValidator(QIntValidator(1, 99, self))

        self.line_intervalo_dias.setFixedWidth(80)
        self.line_intervalo_dias.setStyleSheet(
            "QLineEdit { background-color: #0F1E35; color: #E8F4FF; "
            "border-radius: 6px; font-family: Verdana; font-size: 10pt; padding: 4px; }"
        )
        self.line_intervalo_dias.setEnabled(False)

        # ao invés de só setEnabled, usamos um slot pra também limpar o campo
        self.cbProcessarPorDias.toggled.connect(self.on_intervalo_dias_toggled)

        layout_dias = QHBoxLayout()
        layout_dias.addWidget(self.cbProcessarPorDias)
        layout_dias.addWidget(self.line_intervalo_dias)
        main_layout.addLayout(layout_dias)

        self.cbSepararModelos = QCheckBox("Separar XML por modelo (55/65)")
        self.cbSepararModelos.setChecked(True)
        main_layout.addWidget(self.cbSepararModelos)

        # Pesquisa/seleção empresas
        search_layout = QHBoxLayout()
        self.search_bar = QLineEdit()
        self.search_bar.setPlaceholderText("Pesquisar Empresa")
        self.search_bar.setStyleSheet(
            "QLineEdit { background-color: #0F1E35; color: #E8F4FF; "
            "border-radius: 6px; font-family: Verdana; font-size: 10pt; padding: 4px; }"
        )
        try:
            self.search_bar.addAction(load_icon("pesquisar.png"), QLineEdit.LeadingPosition)
        except Exception:
            pass
        self.search_bar.textChanged.connect(self.filtrar_empresas)
        search_layout.addWidget(self.search_bar)

        btn_select_all = QPushButton("Selecionar Todas")
        set_button_style(btn_select_all, "#27AE60", "#2ECC71", "#1E8449")
        btn_select_all.setIcon(load_icon("selecionar.png"))
        btn_select_all.clicked.connect(self.select_all)
        btn_select_all.setContextMenuPolicy(Qt.CustomContextMenu)
        btn_select_all.customContextMenuRequested.connect(self._on_select_all_context_menu)
        self.btn_select_all = btn_select_all
        search_layout.addWidget(btn_select_all)

        btn_deselect_all = QPushButton("Desmarcar Todas")
        set_button_style(btn_deselect_all, "#C0392B", "#E74C3C", "#922B21")
        btn_deselect_all.setIcon(load_icon("desmarcar-cancelar.png"))
        btn_deselect_all.clicked.connect(self.deselect_all)
        search_layout.addWidget(btn_deselect_all)
        main_layout.addLayout(search_layout)

        # Lista de empresas (mesmo padrão do certidões_bot)
        # 🔹 agora a QScrollArea é atributo da janela
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setStyleSheet(
            "QScrollArea{border:1px solid #1F2A44;border-radius:12px;background:transparent;}"
            "QScrollArea QWidget#qt_scrollarea_viewport{background:qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #0F172A, stop:1 #0B1220);border-radius:12px;}"
        )

        scroll_widget = QWidget()
        self.empresa_items_layout = QVBoxLayout(scroll_widget)
        self.empresa_items_layout.setContentsMargins(2, 2, 2, 2)
        self.empresa_items_layout.setSpacing(2)
        self.empresa_items: List[EmpresaItem] = []
        self._empresa_list_has_stretch = True
        self.empresa_items_layout.addStretch()

        # cria lista já em ordem alfabética (por nome da empresa)
        self._recarregar_lista_empresas()

        self.scroll_area.setWidget(scroll_widget)
        main_layout.addWidget(self.scroll_area, stretch=5)        # Log (com logo de fundo)
        img = None
        bases = []
        if hasattr(sys, "_MEIPASS"):
            bases.append(sys._MEIPASS)  # type: ignore[attr-defined]
        if getattr(sys, "frozen", False):
            bases.append(os.path.dirname(os.path.abspath(sys.executable)))
        bases.append(os.path.dirname(os.path.abspath(__file__)))

        seen = set()
        bases = [b for b in bases if not (b in seen or seen.add(b))]

        for b in bases:
            for rel in (
                "data/IMAGE/logo 2.png",
                "data/Image/logo 2.png",
                "data/image/logo 2.png",
                "IMAGE/logo 2.png",
                "Image/logo 2.png",
                "image/logo 2.png",
                "data/IMAGE/logo.png",
                "data/Image/logo.png",
                "data/image/logo.png",
                "IMAGE/logo.png",
                "Image/logo.png",
                "image/logo.png",
            ):
                p = os.path.join(b, rel)
                if os.path.exists(p):
                    img = p
                    break
            if img:
                break

        self.log_frame = WatermarkLog(img or "", height=240)
        # mantém o comportamento anterior do layout (redimensionável)
        self.log_frame.setMinimumHeight(190)
        self.log_frame.setMaximumHeight(16777215)
        self.log_frame.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.log_text = self.log_frame.text
        self.log_text.setPlaceholderText("Logs da Execução...")

        main_layout.addWidget(self.log_frame, stretch=3)
        # Barra inferior
        bottom_bar = QFrame()
        bottom_bar.setStyleSheet("background-color: #0F1E35; border-radius: 5px;")
        bottom_bar_layout = QHBoxLayout(bottom_bar)
        bottom_bar_layout.setContentsMargins(8, 5, 8, 5)
        bottom_bar_layout.setSpacing(8)

        self.btn_iniciar = QPushButton("Iniciar")
        set_button_style(self.btn_iniciar, "#27AE60", "#2ECC71", "#1E8449")
        self.btn_iniciar.setIcon(load_icon("iniciar.png"))
        self.btn_iniciar.clicked.connect(self.on_start)
        bottom_bar_layout.addWidget(self.btn_iniciar)

        self.btn_parar = QPushButton("Parar")
        set_button_style(self.btn_parar, "#C0392B", "#E74C3C", "#922B21")
        self.btn_parar.setIcon(load_icon("parar.png"))
        self.btn_parar.clicked.connect(self.on_stop)
        bottom_bar_layout.addWidget(self.btn_parar)

        btn_gerenciar = QPushButton("Gerenciar Empresas")
        set_button_style(btn_gerenciar, "#8E44AD", "#9B59B6", "#6C3483")
        btn_gerenciar.setIcon(load_icon("gerenciar.png"))
        btn_gerenciar.clicked.connect(self.abrir_gerenciador_empresas)
        bottom_bar_layout.addWidget(btn_gerenciar)

        btn_login_portal = QPushButton("Login do Portal")
        set_button_style(btn_login_portal, "#2980B9", "#3498DB", "#2471A3")
        btn_login_portal.setIcon(load_icon("login.png"))
        btn_login_portal.clicked.connect(self.abrir_login_portal)
        bottom_bar_layout.addWidget(btn_login_portal)

        btn_clear_log = QPushButton("Limpar Log")
        set_button_style(btn_clear_log, "#1ABC9C", "#16A085", "#117A65")
        btn_clear_log.setIcon(load_icon("limpar-editar.png"))
        btn_clear_log.clicked.connect(self.limpar_log)
        bottom_bar_layout.addWidget(btn_clear_log)

        main_layout.addWidget(bottom_bar)

        footer_label = QLabel("© Automatize Tech")
        footer_label.setAlignment(Qt.AlignCenter)
        footer_label.setStyleSheet(
            "QLabel { color: #E8F4FF; font-family: Verdana; font-size: 6pt; margin-top: 1px; }"
        )
        main_layout.addWidget(footer_label)

        self.btn_parar.setEnabled(False)

        # para gerar PDF depois
        self._last_companies_daily_flags: Dict[str, bool] = {}
        self._last_daily_interval: Optional[int] = None

    def _on_dashboard_metrics(self, total_55: int, total_65: int):
        if getattr(self, 'dashboard_widget', None) is None:
            return
        try:
            self.dashboard_widget.set_counts(total_55, total_65)
        except Exception:
            pass

    def update_log(self, message: str):
        """

        Formata e envia mensagens para o log visual.

        Recursos:
          - Ícones por nível: [INFO]/[ERRO]/[WARN]/[FATAL]/[OK]
          - Timestamp HH:MM:SS
          - Alinhamento de linhas multi-linha
          - Cabeçalho especial por empresa (::EMPRESA_HEADER::...)
        """
        if not message:
            return

        msg = str(message)
        line = emit_terminal_log(msg)
        append_runtime_log(line)

        # =====================================================================
        # 1) CABEÇALHO ESPECIAL POR EMPRESA
        #    Enviado pelo Worker no formato:
        #    ::EMPRESA_HEADER::NOME|IE_FORMATADA|PASTA|OPERACOES|DOWNLOAD_OPTION
        # =====================================================================
        if msg.startswith("::EMPRESA_HEADER::"):
            try:
                payload = msg.split("::EMPRESA_HEADER::", 1)[1]
                parts = payload.split("|")

                display_name = parts[0].strip() if len(parts) > 0 else ""
                ie_fmt = parts[1].strip() if len(parts) > 1 else ""
                pasta = parts[2].strip() if len(parts) > 2 else ""
                ops = parts[3].strip() if len(parts) > 3 else ""
                download_opt = parts[4].strip() if len(parts) > 4 else ""
            except Exception:
                # Se der qualquer problema de parse, cai no fluxo normal de log
                display_name = ie_fmt = pasta = ops = download_opt = ""
            else:
                # Mapeia código interno do tipo de download para um rótulo amigável
                mapa_download = {
                    "docs": "Documentos",
                    "eventos": "Eventos",
                    "ambos": "Documentos + Eventos",
                }
                download_label = mapa_download.get(download_opt, download_opt)

                # Pequena separação visual antes do cabeçalho
                self.log_text.appendPlainText("")

                timestamp = datetime.now().strftime("%H:%M:%S")
                sep = "─" * 72

                self.log_text.appendPlainText(f"⏱️ {timestamp} | {sep}")
                if display_name or ie_fmt:
                    self.log_text.appendPlainText(
                        f"🏢 EMPRESA : {display_name} (IE {ie_fmt})"
                    )
                if ops or download_label:
                    if ops and download_label:
                        self.log_text.appendPlainText(
                            f"⚙️  OPERAÇÃO: {ops} | Download: {download_label}"
                        )
                    elif ops:
                        self.log_text.appendPlainText(f"⚙️  OPERAÇÃO: {ops}")
                    else:
                        self.log_text.appendPlainText(
                            f"⚙️  Download: {download_label}"
                        )
                if pasta:
                    self.log_text.appendPlainText(f"📂 PASTA   : {pasta}")

                self.log_text.appendPlainText(sep)
                self.log_text.appendPlainText("")
                return  # já tratamos esse tipo de mensagem, não segue para o fluxo normal

        # =====================================================================
        # 2) FORMATAÇÃO PADRÃO DOS LOGS ([INFO], [ERRO], [WARN], [FATAL], [OK])
        # =====================================================================
        icon = "ℹ️"
        clean = msg

        if msg.startswith("[INFO]"):
            icon = "ℹ️"
            clean = msg[6:].lstrip()
        elif msg.startswith("[ERRO]"):
            icon = "❌"
            clean = msg[6:].lstrip()
        elif msg.startswith("[WARN]"):
            icon = "⚠️"
            clean = msg[6:].lstrip()
        elif msg.startswith("[FATAL]"):
            icon = "💥"
            clean = msg[7:].lstrip()
        elif msg.startswith("[OK]"):
            icon = "✅"
            clean = msg[4:].lstrip()

        # Alguns marcadores "fortes" ganham uma linha em branco antes
        texto_check = clean.strip()
        if (
            "===== RESUMO FINAL" in texto_check
            or texto_check.startswith("📊 ")
            or texto_check.startswith("🏁 Finalizado processamento da empresa")
            or texto_check.startswith("🟢 Ambiente SEFAZ pronto")
        ):
            self.log_text.appendPlainText("")

        timestamp = datetime.now().strftime("%H:%M:%S")

        # Suporte a mensagens multi-linha alinhadas
        lines = clean.splitlines() or [""]
        header = f"{icon} {timestamp} | "
        indent = " " * len(header)

        for idx, line in enumerate(lines):
            if idx == 0:
                formatted = f"{header}{line}"
            else:
                formatted = f"{indent}{line}"
            self.log_text.appendPlainText(formatted)
            
    # ---------- Filtros / Seleção de Empresas ----------

    @staticmethod
    def _normalize_text(text: str) -> str:
        """
        Normaliza texto para busca:
          - minúsculas
          - sem acentos
        """
        import unicodedata

        if not text:
            return ""

        # remove acentos
        nfkd = unicodedata.normalize("NFD", text)
        sem_acentos = "".join(ch for ch in nfkd if unicodedata.category(ch) != "Mn")
        return sem_acentos.lower()

    def filtrar_empresas(self, text: str):
        text = (text or "").lower().strip()

        for item in self.empresa_items:
            # se não tiver texto, mostra tudo; se tiver, faz o contains
            if not text or text in item.checkbox.text().lower():
                item.show()
            else:
                item.hide()

        # depois de filtrar, joga o scroll da lista para o começo,
        if hasattr(self, "scroll_area") and self.scroll_area is not None:
            barra = self.scroll_area.verticalScrollBar()
            barra.setValue(barra.minimum())

    def select_all(self):
        """Seleciona todas as empresas VISÍVEIS (de acordo com o filtro atual)."""
        for item in self.empresa_items:
            if item.isVisible():
                item.checkbox.setChecked(True)

    def _on_select_all_context_menu(self, pos):
        """Menu de clique-direito do botão 'Selecionar Todas'."""
        menu = QMenu(self)
        act_pendentes = menu.addAction("Selecionar empresas com notas pendentes (Entrada/Saída)")
        menu.addSeparator()
        act_visiveis = menu.addAction("Selecionar todas visíveis")

        btn = getattr(self, "btn_select_all", None)
        if btn is None:
            return

        chosen = menu.exec(btn.mapToGlobal(pos))
        if chosen == act_pendentes:
            self.select_empresas_com_notas_pendentes()
        elif chosen == act_visiveis:
            self.select_all()

    def _empresa_tem_zip_entrada_saida(self, ie: str) -> Tuple[bool, bool]:
        """
        Verifica na estrutura de pastas configurada se a empresa possui ZIP de
        Entrada e de Saída.
        """
        ie = ie_somente_digitos(ie or "")
        if not ie:
            return False, False

        info = (self.empresas or {}).get(ie, {}) or {}
        display_name = (info.get("display_name") or "").strip()
        if not display_name:
            return False, False

        base_root = (self.paths.get("base_empresas") or "").strip()
        if not base_root or not os.path.isdir(fs_path(base_root)):
            return False, False

        empresa_nome = safe_folder_name(nome_empresa_sem_ie(display_name))
        estrutura_cfg = normalizar_estrutura_pastas(self.paths.get("estrutura_pastas", self.folder_structure))

        cfg_base = dict(estrutura_cfg or {})
        cfg_base["separar_entrada_saida"] = False
        cfg_base["usar_ano"] = False
        cfg_base["usar_mes"] = False

        root = montar_caminho_estruturado(
            base_root,
            cfg_base,
            empresa_nome,
            criar=False,
        )

        search_roots: List[str] = []
        if root and os.path.isdir(fs_path(root)):
            search_roots.append(root)
        elif bool(cfg_base.get("usar_nome_empresa", True)):
            # fallback case-insensitive para pasta da empresa
            try:
                target = empresa_nome.lower()
                for entry in os.listdir(fs_path(base_root)):
                    full = os.path.join(base_root, entry)
                    if os.path.isdir(fs_path(full)) and entry.lower() == target:
                        search_roots.append(full)
                        break
            except Exception:
                pass
        else:
            # Estrutura sem pasta por empresa: usa a base inteira.
            search_roots.append(base_root)

        if not search_roots:
            return False, False

        tem_entrada = False
        tem_saida = False

        for root_dir in search_roots:
            try:
                for dirpath, _, filenames in os.walk(fs_path(root_dir)):
                    for fn in filenames:
                        if not fn.lower().endswith(".zip"):
                            continue
                        alvo = self._normalize_text(f"{dirpath} {fn}")
                        if ("entrada" in alvo) or ("notas fiscais de entrada" in alvo):
                            tem_entrada = True
                        if ("saida" in alvo) or ("notas fiscais de saida" in alvo):
                            tem_saida = True
                        if tem_entrada and tem_saida:
                            return True, True
            except Exception:
                continue

        return tem_entrada, tem_saida

    def select_empresas_com_notas_pendentes(self):
        """
        Marca apenas empresas que ainda não possuem os dois tipos de notas
        (falta Entrada ou falta Saída).
        """
        if not getattr(self, "empresa_items", None):
            return
        base_root = (self.paths.get("base_empresas") or "").strip()
        if not base_root or not os.path.isdir(fs_path(base_root)):
            QMessageBox.warning(
                self,
                "Caminho não definido",
                "Defina a pasta base das empresas em 'Caminhos Padrão' para usar esta seleção.",
            )
            return

        faltando = 0
        completas = 0

        for item in self.empresa_items:
            tem_e, tem_s = self._empresa_tem_zip_entrada_saida(item.ie)
            marcar = not (tem_e and tem_s)
            item.checkbox.setChecked(marcar)
            if marcar:
                faltando += 1
            else:
                completas += 1

        try:
            self._salvar_selecao_empresas_config()
        except Exception:
            pass

        self.update_log(
            f"[INFO] Seleção por pendência de notas concluída: "
            f"{faltando} empresa(s) com falta de Entrada/Saída selecionada(s), "
            f"{completas} com ambos os tipos desmarcada(s)."
        )

    def deselect_all(self):
        """Desmarca todas as empresas (independente do filtro)."""
        for item in self.empresa_items:
            item.checkbox.setChecked(False)

    def _recarregar_lista_empresas(self):
        """Recria a lista visual de empresas em ordem alfabética e reaplica marcações."""
        if not getattr(self, "empresa_items_layout", None):
            return
        # remove widgets antigos
        for item in self.empresa_items:
            self.empresa_items_layout.removeWidget(item)
            item.setParent(None)
        self.empresa_items.clear()

        # garante que existe um stretch no final (para empurrar itens para o topo)
        if not hasattr(self, "_empresa_list_has_stretch") or not self._empresa_list_has_stretch:
            self.empresa_items_layout.addStretch()
            self._empresa_list_has_stretch = True

        # conjuntos com o que está marcado (vindo do config / memória interna)
        marcadas = set(getattr(self, "_cfg_empresas_marcadas", set()))
        diario = set(getattr(self, "_cfg_empresas_diario", set()))

        for ie in sorted(self.produtos, key=lambda x: self.empresas[x]['display_name'].lower()):
            display_name = self.empresas[ie]['display_name']
            item = EmpresaItem(ie, display_name)

            # aplica o que veio do config.json
            if ie in marcadas:
                item.checkbox.setChecked(True)
            if ie in diario:
                item.daily_checkbox.setChecked(True)

            # conecta sinais para manter _cfg_* sempre atualizados
            item.checkbox.toggled.connect(
                lambda checked, ie=ie: self._on_empresa_checkbox_toggled(ie, checked)
            )
            item.daily_checkbox.toggled.connect(
                lambda checked, ie=ie: self._on_empresa_diario_toggled(ie, checked)
            )

            self.empresa_items.append(item)
            insert_pos = max(0, self.empresa_items_layout.count() - 1) if getattr(self, "_empresa_list_has_stretch", False) else self.empresa_items_layout.count()

            self.empresa_items_layout.insertWidget(insert_pos, item)

        # sincroniza combo do Dashboard
        try:
            self._dash_sync_empresas_combo()
        except Exception:
            pass

# reaplica filtro atual, se tiver texto na busca
        if hasattr(self, "search_bar"):
            self.filtrar_empresas(self.search_bar.text())



    # -----------------------------
    # Dashboard (aba 1)
    # -----------------------------
    def _init_dashboard_ui(self, dash_layout: QVBoxLayout):
        """Monta a aba Dashboard (primeira tela)."""
        self._dash_worker = None
        self._dash_debounce = QTimer(self)
        self._dash_debounce.setSingleShot(True)
        self._dash_debounce.timeout.connect(self._start_dashboard_refresh)
        self._dash_syncing_op = False

        header = QFrame()
        header.setStyleSheet("""
            QFrame { background: qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 #0F1E35, stop:1 #0B1220);
                    border-radius: 12px; border: 1px solid rgba(232,244,255,0.08); border-top: 1px solid rgba(59,130,246,0.22); }
        """)
        gl = QGridLayout(header)
        gl.setContentsMargins(12, 10, 12, 10)
        gl.setHorizontalSpacing(10)
        gl.setVerticalSpacing(8)
        self._dash_header_gl = gl
        self._dash_header_mode = None
        self._dash_header_frame = header

        # Widgets do header
        self._dash_hdr_lbl_emp = QLabel("Empresa:")
        self._dash_hdr_lbl_emp.setStyleSheet("color:#E8F4FF; font-family: Verdana; font-size: 9pt; font-weight: 700;")

        self.cmb_dash_empresa = QComboBox()
        self.cmb_dash_empresa.setStyleSheet("""
            QComboBox { background-color:#0B1220; color:#E8F4FF; border:1px solid #1F2A44; border-radius:10px; padding:6px; font-family: Verdana; font-size: 9pt; }
            QComboBox QAbstractItemView { background:#0B1220; color:#E8F4FF; selection-background-color:#12304E; }
        """)
        self.cmb_dash_empresa.setMinimumWidth(220)
        self.cmb_dash_empresa.setMaximumWidth(900)
        try:
            self.cmb_dash_empresa.setSizeAdjustPolicy(QComboBox.AdjustToMinimumContentsLengthWithIcon)
            self.cmb_dash_empresa.setMinimumContentsLength(28)
        except Exception:
            pass

        self._dash_hdr_lbl_mod = QLabel("Modelos:")
        self._dash_hdr_lbl_mod.setStyleSheet("color:#9FB6D6; font-family: Verdana; font-size: 9pt;")

        self.cmb_dash_modelo = QComboBox()
        self.cmb_dash_modelo.addItems(["Todos", "55", "65"])
        self.cmb_dash_modelo.setStyleSheet(self.cmb_dash_empresa.styleSheet())
        self.cmb_dash_modelo.setMinimumWidth(86)
        self.cmb_dash_modelo.setMaximumWidth(140)

        self._dash_hdr_lbl_op = QLabel("Operação:")
        self._dash_hdr_lbl_op.setStyleSheet("color:#9FB6D6; font-family: Verdana; font-size: 9pt;")

        self.cmb_dash_operacao = QComboBox()
        self.cmb_dash_operacao.addItems(["Ambos", "Entrada", "Saída"])
        self.cmb_dash_operacao.setStyleSheet(self.cmb_dash_empresa.styleSheet())
        self.cmb_dash_operacao.setMinimumWidth(108)
        self.cmb_dash_operacao.setMaximumWidth(170)

        self._dash_hdr_lbl_per = QLabel('Período:')
        self._dash_hdr_lbl_per.setStyleSheet('color:#9FB6D6; font-family: Verdana; font-size: 9pt; font-weight: 800; padding:3px 10px; background: rgba(11,18,32,0.55); border: 1px solid rgba(232,244,255,0.10); border-radius: 10px;')

        self.dash_date_ini = QDateEdit()
        self.dash_date_ini.setCalendarPopup(True)
        self.dash_date_ini.setDisplayFormat('dd/MM/yyyy')
        self.dash_date_ini.setStyleSheet(self.cmb_dash_empresa.styleSheet())
        self.dash_date_ini.setMinimumWidth(105)
        self.dash_date_ini.setMaximumWidth(140)

        self._dash_hdr_lbl_ate = QLabel('até')
        self._dash_hdr_lbl_ate.setStyleSheet('color:#9FB6D6; font-family: Verdana; font-size: 9pt; font-weight: 800; padding:3px 10px; background: rgba(11,18,32,0.55); border: 1px solid rgba(232,244,255,0.10); border-radius: 10px;')

        self.dash_date_fim = QDateEdit()
        self.dash_date_fim.setCalendarPopup(True)
        self.dash_date_fim.setDisplayFormat('dd/MM/yyyy')
        self.dash_date_fim.setStyleSheet(self.cmb_dash_empresa.styleSheet())
        self.dash_date_fim.setMinimumWidth(105)
        self.dash_date_fim.setMaximumWidth(140)

        # Inicializa com as datas globais (se houver)
        try:
            di = datetime.strptime(self.line_data_inicial.text().strip(), '%d/%m/%Y').date()
            self.dash_date_ini.setDate(QDate(di.year, di.month, di.day))
        except Exception:
            self.dash_date_ini.setDate(QDate.currentDate().addMonths(-1))
        try:
            df = datetime.strptime(self.line_data_final.text().strip(), '%d/%m/%Y').date()
            self.dash_date_fim.setDate(QDate(df.year, df.month, df.day))
        except Exception:
            self.dash_date_fim.setDate(QDate.currentDate())

        self.btn_dash_refresh = QPushButton("Atualizar")
        self.btn_dash_refresh.setCursor(Qt.PointingHandCursor)
        set_button_style(self.btn_dash_refresh, "#12304E", "#16406A", "#0B355B")

        # Checkbox para habilitar/desabilitar envio ao Supabase
        # Usa o mesmo estilo das outras checkboxes do sistema (AnimatedCheckBox)
        self.chk_enviar_supabase = QCheckBox("Enviar ao Supabase")
        # Carrega estado do config.json
        try:
            config = carregar_config()
            enviar_supabase = config.get("enviar_supabase", True)  # Default: True (comportamento atual)
            self.chk_enviar_supabase.setChecked(enviar_supabase)
        except Exception:
            self.chk_enviar_supabase.setChecked(True)  # Default: True
        
        # Salva estado quando mudar
        self.chk_enviar_supabase.stateChanged.connect(self._salvar_config_enviar_supabase)

        self.lbl_dash_status = QLabel("")
        self.lbl_dash_status.setStyleSheet('color:#9AA3B2; font-family: Verdana; font-size: 8pt; font-weight: 700; padding:3px 10px; background: rgba(11,18,32,0.45); border: 1px solid rgba(232,244,255,0.08); border-radius: 10px;')
        self.lbl_dash_status.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)

        # Atualiza ao mudar datas
        self.dash_date_ini.dateChanged.connect(lambda _=None: self._request_dashboard_refresh())
        self.dash_date_fim.dateChanged.connect(lambda _=None: self._request_dashboard_refresh())

        # Monta layout inicial (modo wide/compacto definido no resize)
        self._dash_set_header_mode('wide')
        QTimer.singleShot(0, self._dash_apply_header_responsive)

        dash_layout.addWidget(header)

        body = QFrame()
        body.setStyleSheet("""
            QFrame { background-color:#0B1220; border-radius: 14px; border: 1px solid rgba(232,244,255,0.08); border-top: 1px solid rgba(59,130,246,0.22); }
        """)
        # Evita que o QScrollArea comprima demais os cards (o que fazia o gráfico sumir)
        body.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        body.setMinimumHeight(0)
        grid = QGridLayout(body)
        grid.setContentsMargins(14, 14, 14, 14)
        grid.setHorizontalSpacing(14)
        grid.setVerticalSpacing(14)

        # Gauges
        gauge_card = QFrame()
        gauge_card.setStyleSheet("""
            QFrame { background: qlineargradient(x1:0,y1:0,x2:1,y2:1, stop:0 #0F172A, stop:1 #0F1E35);
                    border-radius: 14px; border: 1px solid rgba(232,244,255,0.08); border-top: 1px solid rgba(59,130,246,0.22); }
        """)
        gvl = QVBoxLayout(gauge_card)
        gvl.setContentsMargins(14, 12, 14, 12)
        gvl.setSpacing(10)
        gauge_card.setMinimumHeight(0)
        gt = QLabel("Dashboards (55 / 65)")
        gt.setStyleSheet("color:#E8F4FF; font-family: Verdana; font-size: 9pt; font-weight: 900; padding:4px 10px; background: rgba(18,48,78,0.18); border: 1px solid rgba(232,244,255,0.10); border-radius: 10px;")
        gvl.addWidget(gt)
        self.dashboard_widget = DonutDashboard()
        self.dashboard_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.dashboard_widget.setMinimumHeight(0)
        gvl.addWidget(self.dashboard_widget, 1)

        # Barras (Entrada x Saída)
        barras_card = QFrame()
        barras_card.setStyleSheet("""
            QFrame { background: qlineargradient(x1:0,y1:0,x2:1,y2:1, stop:0 #0F172A, stop:1 #0F1E35);
                    border-radius: 14px; border: 1px solid rgba(232,244,255,0.08); border-top: 1px solid rgba(59,130,246,0.22); }
        """)
        bvl = QVBoxLayout(barras_card)
        bvl.setContentsMargins(14, 12, 14, 12)
        bvl.setSpacing(10)
        barras_card.setMinimumHeight(0)
        bt = QLabel("Entrada x Saída")
        bt.setStyleSheet("color:#E8F4FF; font-family: Verdana; font-size: 9pt; font-weight: 900; padding:4px 10px; background: rgba(18,48,78,0.18); border: 1px solid rgba(232,244,255,0.10); border-radius: 10px;")
        bvl.addWidget(bt)
        self.bar_widget = DualBarKpiWidget(title="")
        bvl.addWidget(self.bar_widget)


        # Montanha (evolução)
        mont_card = QFrame()
        mont_card.setStyleSheet("""
            QFrame { background: qlineargradient(x1:0,y1:0,x2:1,y2:1, stop:0 #0F172A, stop:1 #0F1E35);
                    border-radius: 14px; border: 1px solid rgba(232,244,255,0.08); border-top: 1px solid rgba(59,130,246,0.22); }
        """)
        mvl = QVBoxLayout(mont_card)
        mvl.setContentsMargins(14, 12, 14, 12)
        mvl.setSpacing(10)
        mont_card.setMinimumHeight(0)
        mt_row = QHBoxLayout()
        mt_row.setContentsMargins(0,0,0,0)
        mt_row.setSpacing(8)

        self.lbl_mont_title = QLabel("Evolução")
        self.lbl_mont_title.setStyleSheet("color:#E8F4FF; font-family: Verdana; font-size: 9pt; font-weight: 900; padding:4px 10px; background: rgba(18,48,78,0.18); border: 1px solid rgba(232,244,255,0.10); border-radius: 10px;")

        self.lbl_mont_metric = QLabel("Resultado diário (R$) • Saída - Entrada")
        self.lbl_mont_metric.setStyleSheet("color:#9FB6D6; font-family: Verdana; font-size: 8pt; font-weight: 800; padding:3px 10px; background: rgba(11,18,32,0.65); border: 1px solid rgba(232,244,255,0.08); border-radius: 10px;")

        mt_row.addWidget(self.lbl_mont_title)
        mt_row.addStretch(1)
        mt_row.addWidget(self.lbl_mont_metric)
        mvl.addLayout(mt_row)
        self.mountain_widget = MountainAreaChartWidget(accent='#3B82F6')
        try:
            self.mountain_widget.rangeChanged.connect(self._on_mountain_range_changed)
        except Exception:
            pass
        # Dá prioridade de espaço para o gráfico (sem reflow agressivo)
        mvl.addWidget(self.mountain_widget, 1)

        # KPI cards
        def _emoji_icon(emoji: str, png_name: str, size: int = 18):
            lb = QLabel()
            lb.setFixedSize(size + 4, size + 2)
            lb.setAlignment(Qt.AlignCenter)
            # nunca desenhar borda/caixa no ícone (remove "linhas" em volta do emoji/imagem)
            lb.setStyleSheet('background: transparent; border: none; padding: 0px; margin: 0px;')
            try:
                p = resolve_image_path(png_name)
                if p and os.path.exists(p):
                    pm = QPixmap(p)
                    if not pm.isNull():
                        pm = pm.scaled(size, size, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                        lb.setPixmap(pm)
                        return lb
            except Exception:
                pass
            lb.setText(emoji or '')
            lb.setStyleSheet('background: transparent; border: none; padding: 0px; color:#E8F4FF; font-size: 12pt;')
            return lb
        def _kpi_card(title: str, emoji: str = '', png_name: str = '', value_color: str = '#E8F4FF'):
            card = QFrame()
            card.setStyleSheet('''
                QFrame { background: qlineargradient(x1:0,y1:0,x2:1,y2:1, stop:0 #0F1E35, stop:1 #0B1220);
                        border-radius: 14px; border: 1px solid rgba(232,244,255,0.08); border-top: 1px solid rgba(59,130,246,0.22); }
            ''')
            v = QVBoxLayout(card)
            v.setContentsMargins(12, 12, 12, 12)
            v.setSpacing(6)

            # Header fixo (não expande verticalmente)
            header_w = QWidget()
            header_w.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
            header_w.setFixedHeight(28)
            header = QHBoxLayout(header_w)
            header.setContentsMargins(0, 0, 0, 0)
            header.setSpacing(8)

            # Badge do título: bem justo no texto
            badge = QFrame()
            badge.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
            badge.setMaximumHeight(26)
            badge.setStyleSheet(
                "QFrame { background: rgba(18, 48, 78, 0.20); border: 1px solid rgba(232,244,255,0.10); border-radius: 10px; }"
            )
            bl = QHBoxLayout(badge)
            bl.setContentsMargins(8, 3, 8, 3)
            bl.setSpacing(6)
            if emoji or png_name:
                bl.addWidget(_emoji_icon(emoji, png_name or '', 14))
            lt = QLabel(title)
            lt.setStyleSheet('background: transparent; border: none; padding: 0px; color:#9FB6D6; font-family: Verdana; font-size: 8pt; font-weight: 900;')
            lt.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
            bl.addWidget(lt)

            header.addWidget(badge, 0, Qt.AlignLeft | Qt.AlignVCenter)
            header.addStretch(1)
            v.addWidget(header_w, 0)

            lv = QLabel('–')
            lv.setStyleSheet(f'background: transparent; border: none; color:{value_color}; font-family: Verdana; font-size: 16pt; font-weight: 900;')
            lv.setWordWrap(False)
            lv.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            lv.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            lv.setFixedHeight(36)

            ls = QLabel('')
            ls.setStyleSheet('background: transparent; border: none; padding: 0px; color:#9AA3B2; font-family: Verdana; font-size: 8pt;')
            ls.setWordWrap(False)
            ls.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            ls.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            ls.setFixedHeight(18)

            # Distribui o espaço interno sem esticar o header
            v.addWidget(lv, 0)
            v.addStretch(1)
            v.addWidget(ls, 0)

            card.setMinimumHeight(120)
            card.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
            return card, lv, ls

        self.card_total_xml, self.lbl_total_xml, self.lbl_total_xml_sub = _kpi_card('Total de XMLs', emoji='📄', png_name='arquivo.png', value_color='#E8F4FF')
        self.card_saida, self.lbl_saida_val, self.lbl_saida_sub = _kpi_card('Faturamento (Saída)', emoji='💰', png_name='dinheiro.png', value_color='#2ECC71')
        self.card_entrada, self.lbl_entrada_val, self.lbl_entrada_sub = _kpi_card('Despesa (Entrada)', emoji='📉', png_name='despesa.png', value_color='#E74C3C')

        self.card_saldo, self.lbl_saldo_val, self.lbl_saldo_sub = _kpi_card('Resultado (Saída - Entrada)', emoji='🧮', png_name='saldo.png', value_color='#E8F4FF')

        grid.addWidget(gauge_card, 0, 0, 2, 1)
        grid.addWidget(barras_card, 2, 0)
        grid.addWidget(mont_card, 3, 0)

        kpi_wrap = QWidget()
        kpi_v = QVBoxLayout(kpi_wrap)
        kpi_v.setContentsMargins(0, 0, 0, 0)
        kpi_v.setSpacing(6)
        kpi_v.addWidget(self.card_total_xml)
        kpi_v.addWidget(self.card_saida)
        kpi_v.addWidget(self.card_entrada)
        kpi_v.addWidget(self.card_saldo)

        # distribui o espaço vertical extra entre os 4 cards (evita "lacuna" grande embaixo)
        for i in range(4):
            try:
                kpi_v.setStretch(i, 1)
            except Exception:
                pass

        grid.addWidget(kpi_wrap, 0, 1, 4, 1)

        grid.setRowStretch(0, 3)
        grid.setRowStretch(1, 3)
        grid.setRowStretch(2, 2)
        grid.setRowStretch(3, 4)
        grid.setColumnStretch(0, 2)
        grid.setColumnStretch(1, 1)
        # Layout adaptativo (igual Execução): corpo dentro de QScrollArea
        # Isso evita sobreposição em resoluções menores e garante que nada seja cortado.
        self.dash_scroll_area = QScrollArea()
        self.dash_scroll_area.setWidgetResizable(True)
        self.dash_scroll_area.setFrameShape(QFrame.NoFrame)
        self.dash_scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.dash_scroll_area.setStyleSheet(
            "QScrollArea{border:none;background:transparent;}"
            "QScrollArea QWidget#qt_scrollarea_viewport{background:transparent;}"
        )
        self.dash_scroll_area.setWidget(body)
        dash_layout.addWidget(self.dash_scroll_area, 1)

        # Sinais
        self.btn_dash_refresh.clicked.connect(lambda: self._request_dashboard_refresh(force=True))
        self.cmb_dash_empresa.currentIndexChanged.connect(lambda *_: self._request_dashboard_refresh())
        self.cmb_dash_modelo.currentIndexChanged.connect(lambda *_: self._request_dashboard_refresh())
        self.cmb_dash_operacao.currentIndexChanged.connect(self._on_dash_operacao_changed)

        # sincroniza com os radios globais (Entrada/Saída/Ambos) e datas
        try:
            self.group_operacao.buttonClicked.connect(lambda *_: self._on_global_operacao_changed())
        except Exception:
            pass
        try:
            self.line_data_inicial.textChanged.connect(lambda *_: self._request_dashboard_refresh())
            self.line_data_final.textChanged.connect(lambda *_: self._request_dashboard_refresh())
        except Exception:
            pass

        # Preenche combo de empresas e atualiza dashboard inicial
        self._dash_sync_empresas_combo()
        self._on_global_operacao_changed()
        self._request_dashboard_refresh(force=True)
        self.robot_id = register_robot()
        self._sync_dashboard_runtime_state(log_changes=False)
        update_robot_status(self.robot_id, "active")
        self.robot_presence_timer.start(15_000)
        self.dashboard_sync_timer.start(10_000)
        # habilita poll da fila apenas se houver credenciais service role
        try:
            url, key = get_robot_supabase_credentials()
            self._queue_supabase_url, self._queue_supabase_key = url or "", key or ""
            if self.robot_id:
                self.queue_poll_timer.start(6_000)
                self.update_log("[FILA] Poll do agendador habilitado (job.json).")
            else:
                self.update_log("[FILA] Poll do agendador desabilitado.")
        except Exception:
            pass

    # -----------------------------
    # Navegação (menu lateral)
    # -----------------------------
    def _match_local_company_metadata(self, company_name: str, document: str) -> Tuple[Optional[str], Optional[Dict[str, Any]]]:
        company_name_norm = normalize_company_name(company_name)
        document_digits = cnpj_somente_digitos(document or "")
        fallback_match: Tuple[Optional[str], Optional[Dict[str, Any]]] = (None, None)
        for ie, info in (self._local_empresas_metadata or {}).items():
            info_cnpj = cnpj_somente_digitos(info.get("cnpj", ""))
            if document_digits and info_cnpj and info_cnpj == document_digits:
                return ie, dict(info)
            if not fallback_match[0] and normalize_company_name(info.get("display_name", "")) == company_name_norm:
                fallback_match = (ie, dict(info))
        return fallback_match

    def _build_dashboard_empresas_map(self, rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
        merged: Dict[str, Dict[str, Any]] = {}
        for row in rows:
            company_name = (row.get("name") or "").strip()
            company_id = row.get("id")
            document = cnpj_somente_digitos(row.get("document", ""))
            state_registration = ie_somente_digitos(row.get("state_registration") or "")
            contador_cpf = cpf_somente_digitos(row.get("contador_cpf") or "")
            selected_login_cpf = cpf_somente_digitos(row.get("selected_login_cpf") or "")
            logins = _normalize_sefaz_go_logins(row.get("global_logins"))
            if not logins:
                logins = _normalize_sefaz_go_logins(row.get("legacy_company_logins"))
            default_login = next((item for item in logins if item.get("is_default")), logins[0] if logins else None)

            ie, local = self._match_local_company_metadata(company_name, document)
            effective_ie = state_registration or ie
            if not effective_ie:
                continue
            entry = {
                "display_name": company_name or (local or {}).get("display_name") or effective_ie,
                "cnpj": document or (local or {}).get("cnpj", ""),
                "company_id": company_id,
                "portal_logins": logins,
            }
            login_cpf = (
                selected_login_cpf
                or contador_cpf
                or cpf_somente_digitos((default_login or {}).get("cpf", ""))
                or cpf_somente_digitos((local or {}).get("login_cpf", ""))
            )
            if login_cpf:
                entry["login_cpf"] = login_cpf
            merged[effective_ie] = entry
        return merged

    def _apply_robot_period_from_dashboard(self) -> None:
        row = fetch_robot_row()
        if not row:
            return
        period_start = (row.get("initial_period_start") or "").strip()
        period_end = (row.get("initial_period_end") or "").strip()
        try:
            if period_start:
                self.line_data_inicial.setText(datetime.strptime(period_start, "%Y-%m-%d").strftime("%d/%m/%Y"))
            if period_end:
                self.line_data_final.setText(datetime.strptime(period_end, "%Y-%m-%d").strftime("%d/%m/%Y"))
        except Exception:
            pass

    def _resolve_runtime_paths(self) -> Dict[str, str]:
        runtime = dict(self.paths or {})
        dashboard_root = resolve_dashboard_output_root()
        if dashboard_root:
            runtime["base_empresas"] = dashboard_root
            # O relatório final deve seguir a mesma pasta base definida no dashboard.
            runtime["relatorios_pdf"] = dashboard_root
        return runtime

    def _sync_dashboard_runtime_state(self, log_changes: bool = True) -> None:
        if SEFAZ_LOCAL_JSON_CLIENT_MODE:
            return
        self.runtime_paths = self._resolve_runtime_paths()
        if self.runtime_paths.get("base_empresas"):
            self.paths["base_empresas"] = self.runtime_paths["base_empresas"]
        if self.runtime_paths.get("relatorios_pdf"):
            self.paths["relatorios_pdf"] = self.runtime_paths["relatorios_pdf"]
        self._apply_robot_period_from_dashboard()
        rows = load_companies_from_dashboard()
        if not rows:
            return
        merged = self._build_dashboard_empresas_map(rows)
        if not merged:
            return
        previous_keys = set(self.empresas.keys())
        self.dashboard_company_cache = rows
        self.empresas = merged
        self.produtos = list(self.empresas.keys())
        self._recarregar_lista_empresas()
        if log_changes and set(self.empresas.keys()) != previous_keys:
            self.update_log(f"[PAINEL] Lista de empresas sincronizada em tempo real ({len(self.empresas)} empresa(s) com IE vinculada).")

    def _refresh_robot_presence(self) -> None:
        update_robot_heartbeat(self.robot_id)
        if self.worker_thread and self.worker_thread.isRunning():
            update_robot_status(self.robot_id, "processing")
        else:
            update_robot_status(self.robot_id, "active")

    def _poll_execution_queue(self) -> None:
        """
        Consome a fila `execution_requests` criada pelo agendador do site.
        Quando encontrar um job para este robô, seleciona as empresas correspondentes e inicia a automação.
        """
        if not self.robot_id:
            return
        if self.automation_running or (self.worker_thread and self.worker_thread.isRunning()):
            return
        if self._active_queue_job:
            return

        job = claim_execution_request_for_queue(
            supabase_url=self._queue_supabase_url,
            supabase_service_key=self._queue_supabase_key,
            robot_id=self.robot_id,
            log_callback=self.update_log,
        )
        if not job:
            return
        self._active_queue_job = job
        try:
            self._start_automation_from_queue_job(job)
        except Exception as e:
            self.update_log(f"[FILA] ERRO ao iniciar job da fila: {e}")
            try:
                complete_execution_request_for_queue(
                    supabase_url=self._queue_supabase_url,
                    supabase_service_key=self._queue_supabase_key,
                    request_id=str(job.get("id")),
                    success=False,
                    error_message=f"Falha ao iniciar job: {e}",
                )
            except Exception:
                pass
            self._active_queue_job = None

    def _start_automation_from_queue_job(self, job: Dict[str, Any]) -> None:
        company_ids = job.get("company_ids") or []
        period_start = str(job.get("period_start") or "").strip()[:10]
        period_end = str(job.get("period_end") or "").strip()[:10]
        job_companies = JSON_RUNTIME.load_job_companies(job, company_ids)

        # garante estado sincronizado (empresas + paths)
        self._sync_dashboard_runtime_state(log_changes=False)
        if job_companies:
            self.empresas = self._build_dashboard_empresas_map(
                [
                    {
                        "id": row.get("company_id") or row.get("id"),
                        "name": row.get("name"),
                        "document": row.get("document") or row.get("cnpj"),
                        "state_registration": row.get("state_registration"),
                        "contador_cpf": row.get("contador_cpf"),
                        "selected_login_cpf": row.get("selected_login_cpf"),
                        "global_logins": (job.get("robot") or {}).get("global_logins") or [],
                        "legacy_company_logins": row.get("sefaz_go_logins") or [],
                    }
                    for row in job_companies
                ]
            )
            self.produtos = list(self.empresas.keys())
            self._recarregar_lista_empresas()

        # aplica datas do período (job usa yyyy-mm-dd; UI usa dd/mm/yyyy)
        def _iso_to_br(s: str) -> str:
            try:
                if not s or len(s) < 10:
                    return ""
                y, m, d = s[:10].split("-")
                return f"{d}/{m}/{y}"
            except Exception:
                return ""

        if period_start:
            self.line_data_inicial.setText(_iso_to_br(period_start))
        if period_end:
            self.line_data_final.setText(_iso_to_br(period_end))

        # seleciona empresas por company_id
        wanted = set(str(x) for x in company_ids if x)
        matched_ies: List[str] = []
        for ie, data in (self.empresas or {}).items():
            cid = str((data or {}).get("company_id") or "")
            if cid and cid in wanted:
                matched_ies.append(ie)

        # aplica seleção na UI (checkbox)
        for item in self.empresa_items:
            item.checkbox.setChecked(item.ie in set(matched_ies))

        if not matched_ies:
            raise Exception("Nenhuma empresa do job corresponde às empresas sincronizadas no robô (verifique company_id/IE).")

        self.update_log(f"[FILA] Job recebido. Empresas: {len(matched_ies)}. Iniciando automação...")
        # inicia automação diretamente (sem usar o agendamento local)
        self.automation_running = True
        self.btn_iniciar.setEnabled(False)
        self.btn_parar.setEnabled(True)
        self.iniciar_automacao()

    def _nav_set_page(self, idx: int):
        """Alterna entre Dashboard (0) e Execução (1)."""
        try:
            idx = int(idx)
        except Exception:
            idx = 0
        idx = 0 if idx <= 0 else 1

        if hasattr(self, 'pages') and self.pages is not None:
            self.pages.setCurrentIndex(idx)

        # mantém estado visual dos botões
        try:
            self.btn_nav_dashboard.setChecked(idx == 0)
            self.btn_nav_execucao.setChecked(idx == 1)
        except Exception:
            pass

        # sempre que voltar ao dashboard, atualiza (sem travar)
        if idx == 0:
            try:
                self._request_dashboard_refresh()
            except Exception:
                pass

    def _on_mountain_range_changed(self, text: str):
        """Atualiza a métrica do cabeçalho do gráfico de evolução conforme a seleção."""
        try:
            if not hasattr(self, 'lbl_mont_metric') or self.lbl_mont_metric is None:
                return
            if text:
                self.lbl_mont_metric.setText(text)
            else:
                base = getattr(self, '_mont_metric_base', '') or ''
                self.lbl_mont_metric.setText(base)
        except Exception:
            pass





    def _animate_sidebar_width(self, target_w: int):
        """Anima a largura da barra lateral de forma leve/clean."""
        if not hasattr(self, 'sidebar') or self.sidebar is None:
            return
        try:
            target_w = int(target_w)
        except Exception:
            return

        start_w = int(self.sidebar.width() or 0)
        if start_w <= 0:
            try:
                start_w = int(self.sidebar.maximumWidth() or 0)
            except Exception:
                start_w = target_w
        if start_w <= 0:
            start_w = target_w

        # Para animações anteriores
        try:
            if hasattr(self, '_sidebar_anim_group') and self._sidebar_anim_group is not None:
                self._sidebar_anim_group.stop()
        except Exception:
            pass

        # trava no valor atual e anima até o alvo
        try:
            self.sidebar.setMinimumWidth(start_w)
            self.sidebar.setMaximumWidth(start_w)
        except Exception:
            pass

        grp = QParallelAnimationGroup(self)
        a_min = QPropertyAnimation(self.sidebar, b"minimumWidth")
        a_max = QPropertyAnimation(self.sidebar, b"maximumWidth")

        for a in (a_min, a_max):
            a.setDuration(200)
            a.setStartValue(start_w)
            a.setEndValue(target_w)
            a.setEasingCurve(QEasingCurve.InOutCubic)
            grp.addAnimation(a)

        def _finish():
            try:
                self.sidebar.setMinimumWidth(target_w)
                self.sidebar.setMaximumWidth(target_w)
            except Exception:
                pass

        grp.finished.connect(_finish)
        self._sidebar_anim_group = grp
        grp.start()

    # -----------------------------
    # Colapsar/Expandir sidebar
    # -----------------------------
    def _toggle_sidebar(self):
        try:
            self._set_sidebar_collapsed(not getattr(self, '_sidebar_collapsed', False))
        except Exception:
            pass

    def _set_sidebar_collapsed(self, collapsed: bool, animate: bool = True):
        collapsed = bool(collapsed)
        self._sidebar_collapsed = collapsed

        if not hasattr(self, 'sidebar') or self.sidebar is None:
            return

        target_w = (getattr(self, '_sidebar_collapsed_w', 64) if collapsed
             else getattr(self, '_sidebar_expanded_w', 190))
        if animate:
            try:
                self._animate_sidebar_width(target_w)
            except Exception:
                animate = False

        if not animate:
            # aplica direto (sem animação) – útil no startup
            try:
                self.sidebar.setMinimumWidth(target_w)
                self.sidebar.setMaximumWidth(target_w)
            except Exception:
                pass

        # Título/Hint
        try:
            if hasattr(self, 'lbl_sidebar_title') and self.lbl_sidebar_title is not None:
                self.lbl_sidebar_title.setVisible(not collapsed)
        except Exception:
            pass
        try:
            if hasattr(self, 'lbl_sidebar_hint') and self.lbl_sidebar_hint is not None:
                self.lbl_sidebar_hint.setVisible(not collapsed)
        except Exception:
            pass

        try:
            h = getattr(self, '_sb_header_h', None)
            if h is not None and hasattr(self, 'btn_sidebar_toggle'):
                if collapsed:
                    # no modo recolhido, centraliza o botão/menu dentro do header
                    try:
                        h.setStretch(1, 0)
                    except Exception:
                        pass
                    try:
                        h.setContentsMargins(6, 6, 6, 6)
                    except Exception:
                        pass
                    try:
                        h.setAlignment(Qt.AlignCenter)
                        h.setAlignment(self.btn_sidebar_toggle, Qt.AlignCenter)
                    except Exception:
                        pass
                else:
                    # no modo aberto, título à esquerda e menu à direita
                    try:
                        h.setStretch(1, 1)
                    except Exception:
                        pass
                    try:
                        h.setContentsMargins(10, 6, 10, 6)
                    except Exception:
                        pass
                    try:
                        h.setAlignment(Qt.AlignVCenter)
                        h.setAlignment(self.btn_sidebar_toggle, Qt.AlignRight | Qt.AlignVCenter)
                        if hasattr(self, 'lbl_sidebar_title') and self.lbl_sidebar_title is not None:
                            h.setAlignment(self.lbl_sidebar_title, Qt.AlignLeft | Qt.AlignVCenter)
                    except Exception:
                        pass
        except Exception:
            pass

        # Botões de navegação: texto some e ficam ícones/emojis
        for btn in getattr(self, '_sidebar_nav_buttons', []) or []:
            if btn is None:
                continue
            full = getattr(btn, '_full_text', None)
            if not full:
                try:
                    btn._full_text = btn.text()
                    full = btn._full_text
                except Exception:
                    full = ''

            if collapsed:
                try:
                    if btn.icon() is not None and not btn.icon().isNull():
                        btn.setText('')
                    else:
                        btn.setText(getattr(btn, '_collapsed_text', ''))
                    btn.setStyleSheet(getattr(self, '_nav_btn_css_collapsed', btn.styleSheet()))
                    btn.setToolTip(full)
                except Exception:
                    pass
            else:
                try:
                    btn.setText(full)
                    btn.setStyleSheet(getattr(self, '_nav_btn_css_expanded', btn.styleSheet()))
                    btn.setToolTip('')
                except Exception:
                    pass


    def _dash_sync_empresas_combo(self):
        """Atualiza o combo do dashboard com base na lista de empresas do app."""
        if not hasattr(self, 'cmb_dash_empresa'):
            return
        combo = self.cmb_dash_empresa
        current_ie = combo.currentData() if combo.count() else None

        combo.blockSignals(True)
        combo.clear()

        try:
            # usa a mesma ordenação da lista
            ies = sorted(self.produtos, key=lambda x: self.empresas[x]['display_name'].lower())
        except Exception:
            ies = list(getattr(self, 'produtos', []) or [])

        for ie in ies:
            info = self.empresas.get(ie, {})
            name = info.get('display_name', ie)
            combo.addItem(name, ie)

        # tenta restaurar seleção
        if current_ie is not None:
            idx = combo.findData(current_ie)
            if idx >= 0:
                combo.setCurrentIndex(idx)
        if combo.count() and combo.currentIndex() < 0:
            combo.setCurrentIndex(0)

        combo.blockSignals(False)

    def _on_dash_operacao_changed(self):
        """Quando o usuário muda a operação no dashboard, sincroniza os radios globais."""
        if self._dash_syncing_op:
            return
        self._dash_syncing_op = True
        try:
            op = self.cmb_dash_operacao.currentText().strip()
            if op == 'Entrada':
                self.rbEntrada.setChecked(True)
            elif op == 'Saída':
                self.rbSaida.setChecked(True)
            else:
                self.rbAmbos.setChecked(True)
        finally:
            self._dash_syncing_op = False
        self._request_dashboard_refresh()

    def _on_global_operacao_changed(self):
        """Quando o usuário muda Entrada/Saída/Ambos nos radios, reflete no dashboard."""
        if not hasattr(self, 'cmb_dash_operacao'):
            return
        if self._dash_syncing_op:
            return
        self._dash_syncing_op = True
        try:
            if self.rbEntrada.isChecked():
                target = 'Entrada'
            elif self.rbSaida.isChecked():
                target = 'Saída'
            else:
                target = 'Ambos'
            idx = self.cmb_dash_operacao.findText(target)
            if idx >= 0:
                self.cmb_dash_operacao.setCurrentIndex(idx)
        finally:
            self._dash_syncing_op = False

    def _request_dashboard_refresh(self, force: bool = False):
        """Debounce: evita travadinhas quando mexe em filtros."""
        try:
            if not hasattr(self, '_dash_debounce'):
                # Se não tem debounce, tenta iniciar diretamente
                try:
                    self._start_dashboard_refresh()
                except Exception as e:
                    print(f"[ERRO] Falha ao iniciar refresh sem debounce: {e}")
                return
            if force:
                self._dash_debounce.stop()
                try:
                    self._start_dashboard_refresh()
                except Exception as e:
                    print(f"[ERRO] Falha ao iniciar refresh forçado: {e}")
                return
            self._dash_debounce.start(220)
        except Exception as e:
            print(f"[ERRO] Erro em _request_dashboard_refresh: {e}")
            # Em caso de erro no debounce, tenta iniciar diretamente
            try:
                self._start_dashboard_refresh()
            except Exception:
                pass
    

    def _get_dashboard_scan_dirs(self, empresa_nome: str, dt_ini: datetime, dt_fim: datetime, operacao: str, modelo: str = "Todos"):
        """Retorna (scan_dirs, recursive).

        Regras:
          - Se a opção "Varrer qualquer subpasta" estiver marcada, varre recursivamente a pasta da empresa.
          - Se estiver desmarcada, respeita **apenas** a estrutura atual configurada em "Caminhos Padrão"
            (incluindo separar por Entrada/Saída, ano, mês e a opção "Separar XML por modelo (55/65)").
          - No modo estrito (estrutura), NÃO faz fallback para varrer a empresa inteira.
        """

        base_root = (self.paths.get("base_empresas") or "").strip()
        if not base_root or not os.path.isdir(base_root):
            return ([], True)

        empresa_pasta = safe_folder_name(empresa_nome)
        empresa_root = ""

        # tenta resolver a pasta da empresa (case-insensitive)
        cand = os.path.join(base_root, empresa_pasta)
        if os.path.isdir(cand):
            empresa_root = cand
        else:
            try:
                target = empresa_pasta.lower()
                for entry in os.listdir(base_root):
                    full = os.path.join(base_root, entry)
                    if os.path.isdir(full) and entry.lower() == target:
                        empresa_root = full
                        break
            except Exception:
                empresa_root = ""

        if not empresa_root:
            return ([], True)

        usar_varredura_livre = bool(self.paths.get("dashboard_busca_empresa_em_qualquer_pasta", True))
        if usar_varredura_livre:
            return ([empresa_root], True)

        # MODO ESTRITO: respeita a estrutura configurada
        try:
            estrutura_cfg = normalizar_estrutura_pastas(self.paths.get("estrutura_pastas"))
        except Exception:
            estrutura_cfg = normalizar_estrutura_pastas(None)

        # Já estamos ancorados na pasta da empresa
        cfg_scan = dict(estrutura_cfg or {})
        cfg_scan["usar_nome_empresa"] = False

        # Se datas não estiverem válidas, não arriscamos varrer a empresa inteira no modo estrito
        if not dt_ini or not dt_fim:
            return ([], False)

        # Operações a incluir (somente se a estrutura separar Entrada/Saída)
        op_txt = (operacao or "Ambos").strip().lower()
        if bool(cfg_scan.get("separar_entrada_saida")):
            if op_txt.startswith('e'):
                ops = ["Entrada"]
            elif op_txt.startswith('s'):
                ops = ["Saída"]
            else:
                ops = ["Entrada", "Saída"]
        else:
            ops = ["Ambos"]

        # Pastas de modelo (55/65) só entram se a opção estiver marcada
        separar_modelo = False
        try:
            separar_modelo = bool(getattr(self, 'cbSepararModelos').isChecked())
        except Exception:
            separar_modelo = False

        mod_txt = (modelo or "Todos").strip()
        if separar_modelo:
            if mod_txt in ("55", "65"):
                modelos = [mod_txt]
            else:
                modelos = ["55", "65"]
        else:
            modelos = [None]

        # Define quais chaves de data vão compor o caminho conforme a estrutura
        usar_ano = bool(cfg_scan.get('usar_ano'))
        usar_mes = bool(cfg_scan.get('usar_mes'))

        datas_para_caminho = []
        if usar_ano and usar_mes:
            # mmYYYY
            y, m = dt_ini.year, dt_ini.month
            end_y, end_m = dt_fim.year, dt_fim.month
            while (y < end_y) or (y == end_y and m <= end_m):
                datas_para_caminho.append(f"{m:02d}{y:04d}")
                m += 1
                if m == 13:
                    y += 1
                    m = 1
        elif usar_ano and not usar_mes:
            # apenas anos
            anos = list(range(dt_ini.year, dt_fim.year + 1))
            datas_para_caminho = [f"01{y:04d}" for y in anos]
        elif (not usar_ano) and usar_mes:
            # apenas mês (MM)
            meses = []
            y, m = dt_ini.year, dt_ini.month
            end_y, end_m = dt_fim.year, dt_fim.month
            while (y < end_y) or (y == end_y and m <= end_m):
                mm = f"{m:02d}"
                if mm not in meses:
                    meses.append(mm)
                m += 1
                if m == 13:
                    y += 1
                    m = 1
            datas_para_caminho = meses
        else:
            datas_para_caminho = [None]

        dirs = []
        for pasta_mes in datas_para_caminho:
            for op in ops:
                for mo in modelos:
                    d = montar_caminho_estruturado(
                        empresa_root,
                        cfg_scan,
                        empresa_pasta,
                        operacao=op,
                        pasta_mes=pasta_mes,
                        modelo=(mo if separar_modelo else None),
                        criar=False,
                    )
                    if d and os.path.isdir(d):
                        dirs.append(d)

        # Dedup preservando ordem
        seen = set()
        out = []
        for d in dirs:
            if d not in seen:
                seen.add(d)
                out.append(d)

        return (out, False)

    def _start_dashboard_refresh(self):
        """Inicia atualização do dashboard com tratamento robusto de erros."""
        try:
            if not hasattr(self, 'cmb_dash_empresa'):
                return

            ie = self.cmb_dash_empresa.currentData()
            if not ie:
                return

            info = self.empresas.get(ie, {})
            empresa_nome_display = info.get('display_name', str(ie))
            empresa_cnpj = info.get('cnpj', '')
            empresa_nome = safe_folder_name(nome_empresa_sem_ie(empresa_nome_display))
            # datas (dashboard)
            dt_ini = None
            dt_fim = None
            try:
                if hasattr(self, 'dash_date_ini') and hasattr(self, 'dash_date_fim'):
                    try:
                        dt_ini = self.dash_date_ini.date().toPython()
                        dt_fim = self.dash_date_fim.date().toPython()
                        dt_ini = datetime(dt_ini.year, dt_ini.month, dt_ini.day)
                        dt_fim = datetime(dt_fim.year, dt_fim.month, dt_fim.day)
                    except Exception as e:
                        print(f"[ERRO] Falha ao obter datas do QDateEdit: {e}")
                        dt_ini = None
                        dt_fim = None
                else:
                    try:
                        dt_ini = datetime.strptime(self.line_data_inicial.text().strip(), '%d/%m/%Y')
                        dt_fim = datetime.strptime(self.line_data_final.text().strip(), '%d/%m/%Y')
                    except Exception as e:
                        print(f"[ERRO] Falha ao parsear datas dos campos de texto: {e}")
                        dt_ini = None
                        dt_fim = None
            except Exception as e:
                print(f"[ERRO] Erro geral ao processar datas: {e}")
                dt_ini = None
                dt_fim = None

            operacao = (self.cmb_dash_operacao.currentText() if hasattr(self, 'cmb_dash_operacao') else 'Ambos').strip()
            modelo = (self.cmb_dash_modelo.currentText() if hasattr(self, 'cmb_dash_modelo') else 'Todos').strip()

            self._dash_refresh_token = getattr(self, '_dash_refresh_token', 0) + 1
            token = self._dash_refresh_token

            try:
                scan_dirs, dash_recursive = self._get_dashboard_scan_dirs(empresa_nome, dt_ini, dt_fim, operacao, modelo)
            except Exception as e:
                print(f"[ERRO] Falha ao obter diretórios de scan: {e}")
                scan_dirs = []
                dash_recursive = True

            # UI: estado de carregamento
            try:
                self.lbl_dash_status.setText('Atualizando...')
            except Exception:
                pass
            # Mantém valores anteriores durante a atualização para não mexer no layout.

            # mata worker anterior
            try:
                if hasattr(self, '_dash_worker') and self._dash_worker and self._dash_worker.isRunning():
                    try:
                        self._dash_worker.requestInterruption()
                        self._dash_worker.wait(1500)
                    except Exception:
                        pass
                    try:
                        if self._dash_worker.isRunning():
                            self._dash_worker.terminate()
                    except Exception:
                        pass
            except Exception:
                pass

            try:
                dt_ini_d = dt_ini.date() if dt_ini else None
                dt_fim_d = dt_fim.date() if dt_fim else None
            except Exception as e:
                print(f"[ERRO] Falha ao converter datas para date: {e}")
                dt_ini_d = None
                dt_fim_d = None

            try:
                self._dash_worker = DashboardScanWorker(
                    scan_dirs=scan_dirs,
                    dt_ini=dt_ini_d,
                    dt_fim=dt_fim_d,
                    operacao=operacao,
                    modelo=modelo,
                    empresa_ie=str(ie),
                    empresa_cnpj=str(empresa_cnpj or ''),
                    consider_duplicates=bool(self.paths.get("dashboard_considerar_xml_duplicados", False)),
                    recursive=bool(dash_recursive),
                    count_real_qty=bool(self.paths.get("dashboard_mostrar_quantidade_real_xml", False)),
                    parent=self,
                )
                self._dash_worker.resultReady.connect(lambda data, tok=token: self._on_dashboard_result(data, tok))
                self._dash_worker.start()
            except Exception as e:
                print(f"[ERRO] Falha ao iniciar worker do dashboard: {e}")
                import traceback
                traceback.print_exc()
                # Restaura status em caso de erro
                try:
                    self.lbl_dash_status.setText('Erro ao atualizar')
                except Exception:
                    pass
        except Exception as e:
            print(f"[ERRO] Erro crítico em _start_dashboard_refresh: {e}")
            import traceback
            traceback.print_exc()
            # Garante que o status seja atualizado mesmo em caso de erro crítico
            try:
                if hasattr(self, 'lbl_dash_status'):
                    self.lbl_dash_status.setText('Erro ao atualizar')
            except Exception:
                pass
    def _on_dashboard_result(self, data: dict, token: int = None):
        """Aplica o resultado do worker no dashboard.

        Ajustes:
          - Ignora resultados atrasados (token) para evitar "pulos" quando o usuário troca
            de empresa rapidamente.
          - Garante que as animações sejam visíveis mesmo quando o refresh demorou.
          - Se houver aviso de CFOPs não parametrizados, ele é exibido *depois* das animações
            para não interromper a transição.
        """
        try:
            try:
                if token is not None and token != getattr(self, "_dash_refresh_token", None):
                    return
            except Exception:
                pass

            data = data or {}

            try:
                c55 = int(data.get("c55") or 0)
                c65 = int(data.get("c65") or 0)
                total_xml = int(data.get("total_xml") or (c55 + c65))
                soma_entrada = float(data.get("soma_entrada") or 0.0)
                soma_saida = float(data.get("soma_saida") or 0.0)
            except Exception:
                c55 = c65 = total_xml = 0
                soma_entrada = soma_saida = 0.0

            # Série (montanha): acumulado, mas exibindo apenas os últimos N pontos
            labels = data.get("serie_labels") or []
            values = data.get("serie_values") or []
            metric = (data.get("serie_metric") or "").strip()

            MAX_PTS = 30
            labels_disp = []
            cum_disp = []
            try:
                if labels and values and len(labels) == len(values):
                    vals = []
                    for v in values:
                        try:
                            vals.append(float(v or 0.0))
                        except Exception:
                            vals.append(0.0)

                    if len(vals) <= MAX_PTS:
                        acc = 0.0
                        for v in vals:
                            acc += v
                            cum_disp.append(acc)
                        labels_disp = list(labels)
                    else:
                        acc = 0.0
                        for v in vals[:-MAX_PTS]:
                            acc += v
                        labels_disp = list(labels[-MAX_PTS:])
                        for v in vals[-MAX_PTS:]:
                            acc += v
                            cum_disp.append(acc)
            except Exception:
                labels_disp = []
                cum_disp = []

            metric_disp = metric
            if metric_disp:
                metric_disp = (metric_disp
                               .replace("diária", "acumulada")
                               .replace("diário", "acumulado")
                               .replace("diaria", "acumulada")
                               .replace("diario", "acumulado"))

            fora = data.get("cfops_fora_param") or []
        
            # Envia dados agregados para Supabase quando dashboard é atualizado
            daily_data = data.get("daily_data", {})
            empresa_cnpj = data.get("empresa_cnpj", "")
            empresa_ie = data.get("empresa_ie", "")
            dt_ini = data.get("dt_ini")
            dt_fim = data.get("dt_fim")
            
            if False and daily_data and empresa_cnpj and dt_ini and dt_fim and isinstance(dt_ini, date) and isinstance(dt_fim, date):
                # Busca nome da empresa
                empresa_nome = ""
                empresa_cnpj_clean = cnpj_somente_digitos(empresa_cnpj) if empresa_cnpj else ""
                
                if empresa_ie and empresa_ie in self.empresas:
                    empresa_nome = self.empresas[empresa_ie].get("display_name", f"Empresa {empresa_ie}")
                    # Se não tem CNPJ no dict, tenta pegar do config
                    if not empresa_cnpj_clean:
                        empresa_cnpj_clean = cnpj_somente_digitos(self.empresas[empresa_ie].get("cnpj", ""))
                elif empresa_cnpj_clean:
                    # Tenta encontrar pelo CNPJ
                    for ie, info in self.empresas.items():
                        cnpj_info = cnpj_somente_digitos(info.get("cnpj", ""))
                        if cnpj_info == empresa_cnpj_clean:
                            empresa_nome = info.get("display_name", f"Empresa {ie}")
                            break
                
                # Se ainda não encontrou nome, usa CNPJ como fallback
                if not empresa_nome and empresa_cnpj_clean:
                    empresa_nome = f"Empresa {empresa_cnpj_clean}"
                
                if empresa_nome:
                    # Envia dados para Supabase (em background, não bloqueia UI) - apenas se checkbox estiver marcado
                    if hasattr(self, 'chk_enviar_supabase') and self.chk_enviar_supabase.isChecked():
                        try:
                            from threading import Thread
                            def _enviar_supabase():
                                try:
                                    # Envia dados agregados do período
                                    enviar_dados_supabase_dashboard(
                                        empresa_cnpj=empresa_cnpj_clean if empresa_cnpj_clean else empresa_cnpj,
                                        empresa_nome=empresa_nome,
                                        daily_data=daily_data,
                                        data_inicial=dt_ini,
                                        data_final=dt_fim,
                                        log_fn=lambda msg: self.update_log(f"[SUPABASE] {msg}")
                                    )
                                    
                                    # Envia dados diários de evolução (documentos + valores)
                                    serie_labels = data.get("serie_labels", [])
                                    serie_values = data.get("serie_values", [])
                                    serie_metric = data.get("serie_metric", "")
                                    daily_in_raw = data.get("daily_in", {})
                                    daily_out_raw = data.get("daily_out", {})
                                    daily_xml_count_raw = data.get("daily_xml_count", {})
                                    daily_nf_count_raw = data.get("daily_nf_count", {})
                                    daily_nfc_count_raw = data.get("daily_nfc_count", {})
                                    
                                    # Converte para Dict[date, float/int] se necessário
                                    daily_in_dict = {}
                                    daily_out_dict = {}
                                    daily_xml_count_dict = {}
                                    daily_nf_count_dict = {}
                                    daily_nfc_count_dict = {}
                                    try:
                                        for d, value in daily_in_raw.items():
                                            try:
                                                if isinstance(d, str):
                                                    d = datetime.fromisoformat(d).date()
                                                elif isinstance(d, datetime):
                                                    d = d.date()
                                                daily_in_dict[d] = float(value)
                                            except Exception:
                                                pass
                                        for d, value in daily_out_raw.items():
                                            try:
                                                if isinstance(d, str):
                                                    d = datetime.fromisoformat(d).date()
                                                elif isinstance(d, datetime):
                                                    d = d.date()
                                                daily_out_dict[d] = float(value)
                                            except Exception:
                                                pass
                                        for d, value in daily_xml_count_raw.items():
                                            try:
                                                if isinstance(d, str):
                                                    d = datetime.fromisoformat(d).date()
                                                elif isinstance(d, datetime):
                                                    d = d.date()
                                                daily_xml_count_dict[d] = int(value)
                                            except Exception:
                                                pass
                                        for d, value in daily_nf_count_raw.items():
                                            try:
                                                if isinstance(d, str):
                                                    d = datetime.fromisoformat(d).date()
                                                elif isinstance(d, datetime):
                                                    d = d.date()
                                                daily_nf_count_dict[d] = int(value)
                                            except Exception:
                                                pass
                                        for d, value in daily_nfc_count_raw.items():
                                            try:
                                                if isinstance(d, str):
                                                    d = datetime.fromisoformat(d).date()
                                                elif isinstance(d, datetime):
                                                    d = d.date()
                                                daily_nfc_count_dict[d] = int(value)
                                            except Exception:
                                                pass
                                    except Exception as e:
                                        print(f"[ERRO] Falha ao converter datas de evolução: {e}")
                                    
                                    # Envia dados de evolução (documentos + valores) apenas se houver dados
                                    if (daily_in_dict or daily_out_dict or daily_xml_count_dict):
                                        enviar_dados_evolucao_supabase(
                                            empresa_cnpj=empresa_cnpj_clean if empresa_cnpj_clean else empresa_cnpj,
                                            empresa_nome=empresa_nome,
                                            serie_labels=serie_labels,
                                            serie_values=serie_values,
                                            serie_metric=serie_metric,
                                            daily_in=daily_in_dict,
                                            daily_out=daily_out_dict,
                                            daily_xml_count=daily_xml_count_dict,
                                            daily_nf_count=daily_nf_count_dict,
                                            daily_nfc_count=daily_nfc_count_dict,
                                            data_inicial=dt_ini,
                                            data_final=dt_fim,
                                            log_fn=lambda msg: self.update_log(f"[SUPABASE-EVOL] {msg}")
                                        )
                                except Exception as e:
                                    print(f"[ERRO] Falha ao enviar dados do dashboard para Supabase: {e}")
                            
                            thread = Thread(target=_enviar_supabase, daemon=True)
                            thread.start()
                        except Exception:
                            pass
                    else:
                        try:
                            self.update_log("[INFO] Envio ao Supabase desabilitado (checkbox desmarcado)")
                        except Exception:
                            pass

            def _apply_ui():
                try:
                    if token is not None and token != getattr(self, "_dash_refresh_token", None):
                        return
                except Exception:
                    pass

                # Donuts (animação)
                try:
                    if hasattr(self, "dashboard_widget") and self.dashboard_widget is not None:
                        self.dashboard_widget.set_counts(c55, c65)
                except Exception:
                    pass

                # Textos
                try:
                    self.lbl_total_xml.setText(str(total_xml))
                    self.lbl_total_xml_sub.setText(f"55: {c55} | 65: {c65}")

                    self.lbl_saida_val.setText(_format_brl(soma_saida))
                    self.lbl_saida_sub.setText("Somatório vNF (Saída)")

                    self.lbl_entrada_val.setText(_format_brl(soma_entrada))
                    self.lbl_entrada_sub.setText("Somatório vNF (Entrada)")
                except Exception:
                    pass

                # Barras (animação)
                try:
                    if hasattr(self, "bar_widget") and self.bar_widget is not None:
                        self.bar_widget.animate_to(soma_entrada, soma_saida)
                except Exception:
                    pass

                # Saldo
                try:
                    saldo = float(soma_saida) - float(soma_entrada)
                    self.lbl_saldo_val.setText(_format_brl(saldo))
                    col = "#2ECC71" if saldo >= 0 else "#E74C3C"
                    self.lbl_saldo_val.setStyleSheet(
                        f"color:{col}; font-family: Verdana; font-size: 16pt; font-weight: 900;"
                    )
                    self.lbl_saldo_sub.setText("Saldo do período")
                except Exception:
                    pass

                # Métrica (texto)
                try:
                    self._mont_metric_base = (metric_disp or "")
                    if metric_disp:
                        self.lbl_mont_metric.setText(metric_disp)
                    else:
                        self.lbl_mont_metric.setText("")
                except Exception:
                    pass

                # Montanha (animação)
                try:
                    if hasattr(self, "mountain_widget") and self.mountain_widget is not None:
                        if (not labels_disp) or (not cum_disp) or (float(total_xml) <= 0):
                            self.mountain_widget.set_series([], [], force_clear=True)
                            try:
                                self._mont_metric_base = ""
                            except Exception:
                                pass
                            try:
                                self.lbl_mont_metric.setText("")
                            except Exception:
                                pass
                        else:
                            self.mountain_widget.set_series(labels_disp, cum_disp)
                except Exception:
                    pass

                # Status
                try:
                    self.lbl_dash_status.setText(
                        f"Atualizado em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
                    )
                except Exception:
                    pass

            # Deixa a UI respirar e iniciar animações no próximo tick
            try:
                QTimer.singleShot(0, _apply_ui)
            except Exception:
                _apply_ui()

            # Aviso de CFOPs não parametrizados (depois das animações)
            try:
                if fora:
                    last = getattr(self, "_dash_last_cfops_fora", None)
                    key = tuple(fora)
                    if key != last:
                        self._dash_last_cfops_fora = key

                        def _show_warn():
                            try:
                                if token is not None and token != getattr(self, "_dash_refresh_token", None):
                                    return
                            except Exception:
                                pass

                            MAX_SHOW = 80
                            shown = fora[:MAX_SHOW]
                            extra = len(fora) - len(shown)

                            msg = (
                                "Dashboard atualizado, porém alguns CFOPs não estão parametrizados no Gerenciar CFOPs.\n"
                                "Eles foram interpretados pela regra automática (1/2/3 = Entrada, 5/6/7 = Saída).\n\n"
                                + "\n".join(shown)
                            )
                            if extra > 0:
                                msg += f"\n\n... (+{extra} outros)"

                            QMessageBox.information(self, "CFOPs fora da parametrização", msg)

                            try:
                                self.update_log(
                                    "[WARN] Dashboard atualizado com CFOPs fora da parametrização: "
                                    + ", ".join(fora[:20])
                                    + (" ..." if len(fora) > 20 else "")
                                )
                            except Exception:
                                pass

                        # tempo suficiente para o usuário ver a transição
                        QTimer.singleShot(900, _show_warn)
            except Exception:
                pass
        except Exception as e:
            # Tratamento de erro geral para evitar tela branca
            print(f"[ERRO] Erro crítico em _on_dashboard_result: {e}")
            import traceback
            traceback.print_exc()
            # Tenta restaurar status mesmo em caso de erro
            try:
                if hasattr(self, 'lbl_dash_status'):
                    self.lbl_dash_status.setText('Erro ao atualizar dashboard')
            except Exception:
                pass


    def _on_empresa_checkbox_toggled(self, ie: str, checked: bool):
        """Atualiza cache interno de empresas marcadas quando a checkbox principal muda."""
        if not hasattr(self, "_cfg_empresas_marcadas"):
            self._cfg_empresas_marcadas = set()
        if checked:
            self._cfg_empresas_marcadas.add(ie)
        else:
            self._cfg_empresas_marcadas.discard(ie)

    def _on_empresa_diario_toggled(self, ie: str, checked: bool):
        """Atualiza cache interno de empresas em modo 'Diário'."""
        if not hasattr(self, "_cfg_empresas_diario"):
            self._cfg_empresas_diario = set()
        if checked:
            self._cfg_empresas_diario.add(ie)
        else:
            self._cfg_empresas_diario.discard(ie)

    def _salvar_config_enviar_supabase(self):
        """Salva a preferência de enviar ao Supabase no config.json"""
        try:
            if not hasattr(self, 'config'):
                self.config = carregar_config()
            self.config["enviar_supabase"] = self.chk_enviar_supabase.isChecked()
            salvar_config(self.config)
        except Exception as e:
            print(f"[ERRO] Falha ao salvar config enviar_supabase: {e}")

    def _salvar_selecao_empresas_config(self):
        """
        Salva no config.json quais empresas estão marcadas e quais estão com 'Diário' marcado.
        Estrutura:
        "selecoes_empresas": {
            "marcadas": ["ie1", "ie2"],
            "diario":   ["ie1", ...]
        }
        """
        marcadas = []
        diario = []

        for item in self.empresa_items:
            if item.checkbox.isChecked():
                marcadas.append(item.ie)
            if item.daily_checkbox.isChecked():
                diario.append(item.ie)

        # atualiza caches internos
        self._cfg_empresas_marcadas = set(marcadas)
        self._cfg_empresas_diario = set(diario)

        selecoes = self.config.get("selecoes_empresas")
        if not isinstance(selecoes, dict):
            selecoes = {}

        selecoes["marcadas"] = marcadas
        selecoes["diario"] = diario

        self.config["selecoes_empresas"] = selecoes
        salvar_config(self.config)

    def _salvar_empresas_config(self):
        """
        Sincroniza o mapa de empresas com o config e persiste no disco.
        """
        self.config["empresas"] = self.empresas
        salvar_config(self.config)

    # ---------- Helpers IE (máscara) ----------

    def _pedir_ie(self, titulo: str, label: str, valor_inicial: str = "") -> Tuple[str, bool]:
        """
        Abre um diálogo com máscara de IE.
        Retorna (ie_sem_mascara, ok).
        """
        dlg = QDialog(self)
        dlg.setWindowTitle(titulo)
        dlg.setModal(True)
        dlg.setMinimumWidth(360)
        dlg.setStyleSheet("""
            QDialog {
                background-color: #0B1220;
            }
            QLabel {
                color: #E8F4FF;
                font-family: Verdana;
                font-size: 10pt;
            }
            QLineEdit {
                background-color: #0F1E35;
                color: #E8F4FF;
                border-radius: 6px;
                padding: 4px;
                font-family: Verdana;
                font-size: 10pt;
            }
        """)

        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        lbl = QLabel(label)
        layout.addWidget(lbl)

        line = QLineEdit()
        line.setInputMask("00.000.000-0;_")  # máscara 10.920.217-1
        if valor_inicial:
            line.setText(ie_formatada(valor_inicial))
        layout.addWidget(line)

        btns = QHBoxLayout()
        btn_ok = QPushButton("OK")
        btn_cancel = QPushButton("Cancelar")

        # ESTILO ESPECÍFICO DOS BOTÕES
        set_button_style(btn_ok, "#27AE60", "#2ECC71", "#1E8449")
        btn_ok.setIcon(load_icon("selecionar.png"))
        set_button_style(btn_cancel, "#C0392B", "#E74C3C", "#922B21")
        btn_cancel.setIcon(load_icon("desmarcar-cancelar.png"))

        btns.addWidget(btn_ok)
        btns.addWidget(btn_cancel)
        layout.addLayout(btns)

        def aceitar():
            if ie_somente_digitos(line.text()):
                dlg.accept()
            else:
                QMessageBox.warning(self, "IE inválida", "Informe uma IE válida (somente dígitos).")

        def rejeitar():
            dlg.reject()

        btn_ok.clicked.connect(aceitar)
        btn_cancel.clicked.connect(rejeitar)

        ok = dlg.exec() == QDialog.Accepted
        ie_limpa = ie_somente_digitos(line.text())
        return ie_limpa, ok

    # ---------- Diálogos estilizados auxiliares ----------

    def _dialog_nome_empresa(self, titulo: str, label: str, valor_inicial: str = "") -> tuple[str, bool]:
        """Caixinha estilizada para digitar o NOME da empresa."""
        dlg = QDialog(self)
        dlg.setWindowTitle(titulo)
        dlg.setModal(True)
        dlg.setMinimumWidth(360)
        dlg.setStyleSheet("""
            QDialog {
                background-color: #0B1220;
            }
            QLabel {
                color: #E8F4FF;
                font-family: Verdana;
                font-size: 10pt;
            }
            QLineEdit {
                background-color: #0F1E35;
                color: #E8F4FF;
                border-radius: 6px;
                padding: 4px;
                font-family: Verdana;
                font-size: 10pt;
            }
        """)

        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(8)

        lbl = QLabel(label)
        layout.addWidget(lbl)

        line = QLineEdit()
        if valor_inicial:
            line.setText(valor_inicial)
        layout.addWidget(line)

        btns = QHBoxLayout()
        btns.addStretch()
        btn_ok = QPushButton("OK")
        btn_cancel = QPushButton("Cancelar")

        set_button_style(btn_ok, "#27AE60", "#2ECC71", "#1E8449")
        btn_ok.setIcon(load_icon("selecionar.png"))
        set_button_style(btn_cancel, "#C0392B", "#E74C3C", "#922B21")
        btn_cancel.setIcon(load_icon("desmarcar-cancelar.png"))

        btns.addWidget(btn_ok)
        btns.addWidget(btn_cancel)
        layout.addLayout(btns)

        def aceitar():
            dlg.accept()

        def rejeitar():
            dlg.reject()

        btn_ok.clicked.connect(aceitar)
        btn_cancel.clicked.connect(rejeitar)

        ok = dlg.exec() == QDialog.Accepted
        return line.text(), ok

    def _cnpj_lookup_allow(self) -> Tuple[bool, int]:
        """Rate-limit simples: permite até 3 consultas por janela de 20s.

        Retorna (allowed, wait_seconds).
        """
        now = time.time()
        window = 20.0
        calls = [t for t in (self._cnpj_lookup_calls or []) if (now - t) < window]
        self._cnpj_lookup_calls = calls
        if len(calls) >= 3:
            oldest = min(calls) if calls else now
            wait = int(math.ceil(max(0.0, window - (now - oldest))))
            return False, max(1, wait)
        self._cnpj_lookup_calls.append(now)
        return True, 0

    def _dialog_empresa_form(self, titulo: str, nome_init: str = "", cnpj_init: str = "", ie_init: str = "", login_cpf_init: str = "") -> Tuple[str, str, str, str, bool]:
        """Diálogo único: Nome + CNPJ + Inscrição Estadual.

        - Ao sair do campo de CNPJ (ou Enter), tenta buscar o nome via API.
        - Só preenche o nome automaticamente se o campo de nome estiver vazio.
        """
        dlg = QDialog(self)
        dlg.setWindowTitle(titulo)
        dlg.setModal(True)
        dlg.setStyleSheet("""
            QDialog { background-color: #0B1220; }
            QLabel { color: #E8F4FF; font-family: Verdana; font-size: 10pt; }
            QLineEdit {
                background-color: #0F1E35;
                color: #E8F4FF;
                border-radius: 6px;
                padding: 6px;
                font-family: Verdana;
                font-size: 10pt;
            }
        """)

        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(10)

        grid = QGridLayout()
        grid.setHorizontalSpacing(10)
        grid.setVerticalSpacing(8)
        grid.setColumnStretch(1, 1)

        lbl_nome = QLabel("Nome")
        lbl_cnpj = QLabel("CNPJ")
        lbl_ie = QLabel("Inscrição Estadual")

        line_nome = QLineEdit()
        line_nome.setPlaceholderText("Digite o nome da empresa")
        line_nome.setMinimumWidth(250)
        line_nome.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        if nome_init:
            line_nome.setText(nome_init)

        line_cnpj = QLineEdit()
        line_cnpj.setInputMask("00.000.000/0000-00;_")
        line_cnpj.setPlaceholderText("00.000.000/0000-00")
        line_cnpj.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        if cnpj_init:
            line_cnpj.setText(cnpj_formatado(cnpj_init))

        line_ie = QLineEdit()
        line_ie.setInputMask("00.000.000-0;_")  # mantém padrão GO (como já era)
        line_ie.setPlaceholderText("10.920.217-1")
        line_ie.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        if ie_init:
            line_ie.setText(ie_formatada(ie_init))

        grid.addWidget(lbl_nome, 0, 0)
        grid.addWidget(line_nome, 0, 1)
        grid.addWidget(lbl_cnpj, 1, 0)
        grid.addWidget(line_cnpj, 1, 1)
        grid.addWidget(lbl_ie, 2, 0)
        grid.addWidget(line_ie, 2, 1)

        # Login do portal (opcional por empresa)
        logins = load_logins_portal()
        has_logins = bool(logins)
        login_cpf_init_digits = cpf_somente_digitos(login_cpf_init)

        cb_login_especifico = QCheckBox("Selecionar login do portal para esta empresa")
        cb_login_especifico.setChecked(bool(login_cpf_init_digits))
        if not has_logins:
            cb_login_especifico.setChecked(False)
            cb_login_especifico.setEnabled(False)
            cb_login_especifico.setToolTip("Nenhum login cadastrado. Use 'Login do Portal' para adicionar.")

        lbl_login = QLabel("Login do Portal")
        combo_login = QComboBox()
        combo_login.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        combo_login.setMinimumWidth(250)
        combo_login.setStyleSheet("""
            QComboBox, QComboBox QListView {
                background-color: #0F1E35;
                color: #E8F4FF;
                border-radius: 6px;
                padding: 4px;
                font-family: Verdana;
                font-size: 9pt;
            }
        """)

        if has_logins:
            for item in logins:
                cpf_digits = cpf_somente_digitos(item.get("cpf", ""))
                if cpf_digits:
                    combo_login.addItem(cpf_formatado(cpf_digits), cpf_digits)
            if login_cpf_init_digits:
                idx = combo_login.findData(login_cpf_init_digits)
                if idx >= 0:
                    combo_login.setCurrentIndex(idx)
        else:
            combo_login.addItem("Nenhum login cadastrado", "")
            combo_login.setEnabled(False)

        grid.addWidget(cb_login_especifico, 3, 0, 1, 2)
        grid.addWidget(lbl_login, 4, 0)
        grid.addWidget(combo_login, 4, 1)

        def _update_login_visibility(checked: bool):
            show = bool(checked and has_logins)
            lbl_login.setVisible(show)
            combo_login.setVisible(show)

        cb_login_especifico.toggled.connect(_update_login_visibility)
        _update_login_visibility(cb_login_especifico.isChecked())

        layout.addLayout(grid)

        lbl_status = QLabel("")
        lbl_status.setStyleSheet("color:#9AA3B2; font-family: Verdana; font-size: 8pt;")
        layout.addWidget(lbl_status)

        btns = QHBoxLayout()
        btns.addStretch()
        btn_ok = QPushButton("OK")
        btn_cancel = QPushButton("Cancelar")

        set_button_style(btn_ok, "#27AE60", "#2ECC71", "#1E8449")
        btn_ok.setIcon(load_icon("selecionar.png"))
        set_button_style(btn_cancel, "#C0392B", "#E74C3C", "#922B21")
        btn_cancel.setIcon(load_icon("desmarcar-cancelar.png"))

        btns.addWidget(btn_ok)
        btns.addWidget(btn_cancel)
        layout.addLayout(btns)

        def _finish_lookup(res: dict):
            try:
                lbl_status.setText("")
            except Exception:
                pass

            if not isinstance(res, dict):
                return

            if res.get('ok'):
                # só preenche se o nome estiver vazio
                if not (line_nome.text() or '').strip():
                    line_nome.setText((res.get('nome') or '').strip())
                return

            err = (res.get('erro') or '').strip()
            if err:
                QMessageBox.information(dlg, "Consulta CNPJ", err)

        def _try_lookup():
            # não sobrescreve o nome se já estiver preenchido
            if (line_nome.text() or '').strip():
                return
            cnpj_digits = cnpj_somente_digitos(line_cnpj.text())
            if len(cnpj_digits) != 14:
                return

            allowed, wait = self._cnpj_lookup_allow()
            if not allowed:
                QMessageBox.information(
                    dlg,
                    "Consulta CNPJ",
                    f"Aguarde {wait}s para nova consulta de CNPJ (limite temporário)."
                )
                return

            lbl_status.setText("Consultando CNPJ...")
            worker = CNPJLookupWorker(cnpj_digits, parent=dlg)
            dlg._cnpj_worker = worker  # evita GC
            worker.finishedLookup.connect(_finish_lookup)
            worker.start()

        # dispara ao desfocar/enter
        line_cnpj.editingFinished.connect(_try_lookup)

        def aceitar():
            nome = (line_nome.text() or '').strip()
            cnpj_digits = cnpj_somente_digitos(line_cnpj.text())
            ie_digits = ie_somente_digitos(line_ie.text())
            if not nome:
                QMessageBox.warning(dlg, "Dados inválidos", "Informe o nome da empresa.")
                return
            if len(cnpj_digits) != 14:
                QMessageBox.warning(dlg, "Dados inválidos", "Informe um CNPJ válido.")
                return
            if not ie_digits:
                QMessageBox.warning(dlg, "Dados inválidos", "Informe uma Inscrição Estadual válida (somente dígitos).")
                return
            if cb_login_especifico.isChecked():
                if not has_logins:
                    QMessageBox.warning(dlg, "Dados inválidos", "Nenhum login do portal cadastrado. Cadastre em 'Login do Portal'.")
                    return
                if not (combo_login.currentData() or ""):
                    QMessageBox.warning(dlg, "Dados inválidos", "Selecione um login do portal.")
                    return
            dlg.accept()

        def rejeitar():
            dlg.reject()

        btn_ok.clicked.connect(aceitar)
        btn_cancel.clicked.connect(rejeitar)

        ok = dlg.exec() == QDialog.Accepted
        login_cpf = ""
        if ok and cb_login_especifico.isChecked() and has_logins:
            login_cpf = cpf_somente_digitos(combo_login.currentData() or "")
        return (
            (line_nome.text() or '').strip(),
            cnpj_somente_digitos(line_cnpj.text()),
            ie_somente_digitos(line_ie.text()),
            login_cpf,
            ok,
        )

    def _dialog_selecionar_empresa(self, titulo: str, label: str, itens: list[str]) -> tuple[str, bool]:
        """Caixinha estilizada com combo para escolher a empresa (Editar / Excluir)."""
        dlg = QDialog(self)
        dlg.setWindowTitle(titulo)
        dlg.setModal(True)
        dlg.setStyleSheet("""
            QDialog {
                background-color: #0B1220;
            }
            QLabel {
                color: #E8F4FF;
                font-family: Verdana;
                font-size: 10pt;
            }
            QComboBox, QComboBox QListView {
                background-color: #0F1E35;
                color: #E8F4FF;
                border-radius: 6px;
                padding: 4px;
                font-family: Verdana;
                font-size: 9pt;
            }
        """)

        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(8)

        lbl = QLabel(label)
        layout.addWidget(lbl)

        combo = QComboBox()
        combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        combo.setMinimumWidth(260)
        combo.addItems(itens)
        layout.addWidget(combo)

        btns = QHBoxLayout()
        btns.addStretch()
        btn_ok = QPushButton("OK")
        btn_cancel = QPushButton("Cancelar")

        set_button_style(btn_ok, "#27AE60", "#2ECC71", "#1E8449")
        btn_ok.setIcon(load_icon("selecionar.png"))
        set_button_style(btn_cancel, "#C0392B", "#E74C3C", "#922B21")
        btn_cancel.setIcon(load_icon("desmarcar-cancelar.png"))

        btns.addWidget(btn_ok)
        btns.addWidget(btn_cancel)
        layout.addLayout(btns)

        def aceitar():
            dlg.accept()

        def rejeitar():
            dlg.reject()

        btn_ok.clicked.connect(aceitar)
        btn_cancel.clicked.connect(rejeitar)

        ok = dlg.exec() == QDialog.Accepted
        if ok and combo.currentIndex() >= 0:
            return combo.currentText(), True
        return "", False

    def _dialog_selecionar_empresas_multiplas(
        self,
        titulo: str,
        label: str,
        empresas: list[dict],
        login_cpfs: Optional[List[str]] = None,
    ) -> list[str]:
        """Dialog com busca para selecionar VARIAS empresas.

        Retorna lista de IEs (somente digitos). Se cancelar, retorna [].
        """
        dlg = QDialog(self)
        dlg.setWindowTitle(titulo)
        dlg.setModal(True)
        dlg.setMinimumSize(680, 520)
        dlg.setStyleSheet("""
            QDialog { background-color: #0B1220; }
            QLabel { color: #E8F4FF; font-family: Verdana; font-size: 10pt; }
            QLineEdit {
                background-color: #0F1E35; color: #E8F4FF; border-radius: 6px;
                padding: 6px; font-family: Verdana; font-size: 9pt;
                border: 1px solid rgba(255,255,255,0.08);
            }
            QListWidget {
                background-color: #0F1E35; color: #E8F4FF; border-radius: 8px;
                padding: 6px; border: 1px solid rgba(255,255,255,0.08);
                font-family: Verdana; font-size: 9pt;
            }
            QPushButton {
                font:9pt 'Verdana'; font-weight:bold; color:#fff; padding:8px 14px;
                border-radius:8px; border:1px solid rgba(255,255,255,0.08);
                background:#102A4C;
            }
            QPushButton:hover { background:#12304E; }
            QPushButton:pressed { background:#0F1E35; padding-top:11px; padding-bottom:9px; }
        """)

        root = QVBoxLayout(dlg)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(10)

        lbl = QLabel(label)
        lbl.setWordWrap(True)
        root.addWidget(lbl)

        le = QLineEdit()
        le.setPlaceholderText("Buscar por nome, IE ou CNPJ (digite parte do texto)")
        root.addWidget(le)

        login_filter_row = QHBoxLayout()
        lbl_login = QLabel("Filtrar por login:")
        combo_login = QComboBox()
        combo_login.addItem("Selecione login", "")
        for cpf in (login_cpfs or []):
            cpf_d = cpf_somente_digitos(cpf or "")
            if len(cpf_d) == 11:
                combo_login.addItem(cpf_formatado(cpf_d), cpf_d)
        login_filter_row.addWidget(lbl_login)
        login_filter_row.addWidget(combo_login, 1)
        root.addLayout(login_filter_row)

        lst = QListWidget()
        lst.setSelectionMode(QAbstractItemView.ExtendedSelection)
        try:
            lst.setSelectionRectVisible(True)
        except Exception:
            pass
        root.addWidget(lst, 1)

        actions = QHBoxLayout()
        btn_sel_all = QPushButton("Selecionar tudo (filtrado)")
        set_button_style(btn_sel_all, "#2980B9", "#3498DB", "#2471A3")
        btn_clear = QPushButton("Limpar selecao")
        set_button_style(btn_clear, "#C0392B", "#E74C3C", "#922B21")
        actions.addWidget(btn_sel_all)
        actions.addWidget(btn_clear)
        actions.addStretch(1)
        root.addLayout(actions)

        btns = QHBoxLayout()
        btns.addStretch(1)
        b_ok = QPushButton("OK")
        set_button_style(b_ok, "#27AE60", "#2ECC71", "#1E8449")
        b_ok.setIcon(load_icon("selecionar.png"))
        b_cancel = QPushButton("Cancelar")
        set_button_style(b_cancel, "#102A4C", "#12304E", "#0F1E35")
        b_cancel.setIcon(load_icon("desmarcar-cancelar.png"))
        btns.addWidget(b_ok)
        btns.addWidget(b_cancel)
        root.addLayout(btns)

        # Prepara lista base (ordenada por nome) e normaliza campos para busca
        base: list[dict] = []
        for e in (empresas or []):
            try:
                ie = ie_somente_digitos((e or {}).get("ie", ""))
            except Exception:
                ie = ""
            if not ie:
                continue
            try:
                nome = ((e or {}).get("display_name", "") or "").strip()
            except Exception:
                nome = ""
            try:
                cnpj = cnpj_somente_digitos((e or {}).get("cnpj", ""))
            except Exception:
                cnpj = ""
            try:
                login_cpf = cpf_somente_digitos((e or {}).get("login_cpf", ""))
            except Exception:
                login_cpf = ""
            base.append(
                {
                    "ie": ie,
                    "nome": nome,
                    "cnpj": cnpj,
                    "login_cpf": login_cpf,
                    "label": f"{nome} - IE {ie_formatada(ie)}" + (f" - CNPJ {cnpj_formatado(cnpj)}" if cnpj else ""),
                }
            )
        base.sort(key=lambda x: (x.get("nome") or "").lower())

        selected_ies: set[str] = set()

        def _selected_now() -> set[str]:
            out: set[str] = set()
            for it in (lst.selectedItems() or []):
                ie = it.data(Qt.UserRole) or ""
                ie = ie_somente_digitos(ie)
                if ie:
                    out.add(ie)
            return out

        def refresh():
            nonlocal selected_ies
            # Mantem selecao ao filtrar
            selected_ies |= _selected_now()

            q = (le.text() or "").strip().lower()
            q_digits = re.sub(r"\\D", "", q)
            login_sel = cpf_somente_digitos(combo_login.currentData() or "")

            lst.clear()
            for e in base:
                label_txt = e.get("label") or ""
                nome_l = (e.get("nome") or "").lower()
                ie = e.get("ie") or ""
                cnpj = e.get("cnpj") or ""
                login_cpf = cpf_somente_digitos(e.get("login_cpf") or "")

                if login_sel and login_cpf != login_sel:
                    continue

                ok = True
                if q:
                    ok = (q in nome_l) or (q in label_txt.lower())
                    if (not ok) and q_digits:
                        ok = (q_digits in ie) or (q_digits in cnpj)
                if not ok:
                    continue

                it = QListWidgetItem(label_txt)
                it.setData(Qt.UserRole, ie)
                lst.addItem(it)
                if ie in selected_ies:
                    it.setSelected(True)

            if lst.count() > 0 and lst.currentRow() < 0:
                lst.setCurrentRow(0)

        def select_all_filtered():
            for i in range(lst.count()):
                it = lst.item(i)
                if it:
                    it.setSelected(True)

        def clear_selection():
            lst.clearSelection()
            selected_ies.clear()

        def accept_selection():
            ies = list(selected_ies | _selected_now())
            ies = [ie_somente_digitos(x) for x in ies if ie_somente_digitos(x)]
            # dedupe mantendo ordem
            out: list[str] = []
            seen = set()
            for ie in ies:
                if ie not in seen:
                    seen.add(ie)
                    out.append(ie)
            dlg._chosen_ies = out
            dlg.accept()

        def on_return_pressed():
            # Se so existe 1 item filtrado, confirma direto.
            if lst.count() == 1:
                it = lst.item(0)
                if it:
                    it.setSelected(True)
                accept_selection()
            else:
                lst.setFocus(Qt.TabFocusReason)

        le.textChanged.connect(refresh)
        combo_login.currentIndexChanged.connect(lambda _: refresh())
        le.returnPressed.connect(on_return_pressed)
        lst.itemDoubleClicked.connect(lambda _: accept_selection())
        btn_sel_all.clicked.connect(select_all_filtered)
        btn_clear.clicked.connect(clear_selection)
        b_ok.clicked.connect(accept_selection)
        b_cancel.clicked.connect(dlg.reject)

        refresh()

        if dlg.exec() == QDialog.Accepted:
            return list(getattr(dlg, "_chosen_ies", []) or [])
        return []

    def _dialog_confirmacao(self, titulo: str, mensagem: str) -> bool:
        """Caixinha estilizada de confirmação (Sim / Não)."""
        dlg = QDialog(self)
        dlg.setWindowTitle(titulo)
        dlg.setModal(True)
        dlg.setStyleSheet("""
            QDialog {
                background-color: #0B1220;
            }
            QLabel {
                color: #E8F4FF;
                font-family: Verdana;
                font-size: 10pt;
            }
        """)

        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(10)

        lbl = QLabel(mensagem)
        lbl.setWordWrap(True)
        layout.addWidget(lbl)

        btns = QHBoxLayout()
        btns.addStretch()
        btn_yes = QPushButton("Sim")
        btn_no = QPushButton("Não")

        set_button_style(btn_yes, "#27AE60", "#2ECC71", "#1E8449")
        btn_yes.setIcon(load_icon("selecionar.png"))
        set_button_style(btn_no, "#C0392B", "#E74C3C", "#922B21")
        btn_no.setIcon(load_icon("desmarcar-cancelar.png"))

        btns.addWidget(btn_yes)
        btns.addWidget(btn_no)
        layout.addLayout(btns)

        def aceitar():
            dlg.accept()

        def rejeitar():
            dlg.reject()

        btn_yes.clicked.connect(aceitar)
        btn_no.clicked.connect(rejeitar)

        return dlg.exec() == QDialog.Accepted

    def limpar_log(self):
        if self._dialog_confirmacao("Confirmação", "Deseja limpar o log?"):
            self.log_text.clear()
            self.update_log("🧹 [LOG] Log limpo.")
        else:
            self.update_log("ℹ️ [LOG] Limpeza de log cancelada.")

    def _setup_tray_icon(self):
        self._tray_icon = QSystemTrayIcon(self)
        app_icon = QIcon(ICO_PATH) if os.path.exists(ICO_PATH) else QIcon()
        if app_icon.isNull():
            app_icon = self.windowIcon()
        if app_icon.isNull():
            app_icon = self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon)
        self._tray_icon.setIcon(app_icon)
        menu = QMenu()
        show_act = QAction("Abrir janela", self)
        show_act.triggered.connect(self._show_from_tray)
        menu.addAction(show_act)
        quit_act = QAction("Fechar robô", self)
        quit_act.triggered.connect(self._quit_from_tray)
        menu.addAction(quit_act)
        self._tray_icon.setContextMenu(menu)
        self._tray_icon.activated.connect(self._on_tray_activated)
        self._tray_icon.setToolTip("Sefaz Xml - NFe / NFC-e GO")

    def _show_from_tray(self):
        self.showNormal()
        self.raise_()
        self.activateWindow()
        self._refresh_robot_presence()
        if not self.robot_presence_timer.isActive():
            self.robot_presence_timer.start(30000)
        if not self.dashboard_sync_timer.isActive():
            self.dashboard_sync_timer.start(60000)

    def _on_tray_activated(self, reason: int):
        if reason == QSystemTrayIcon.DoubleClick:
            self._show_from_tray()

    def _quit_from_tray(self):
        self._quitting_from_tray = True
        if self.worker_thread and self.worker_thread.isRunning():
            self.worker_thread.stop()
            if not self.worker_thread.wait(5000):
                self.worker_thread.terminate()
                self.worker_thread.wait(1000)
        self.robot_presence_timer.stop()
        self.dashboard_sync_timer.stop()
        update_robot_status(self.robot_id, "inactive")
        try:
            _kill_automation_chrome_proc(log_fn=self.update_log)
        except Exception:
            pass
        try:
            _kill_automation_chrome_instances(log_fn=self.update_log)
        except Exception:
            pass
        _stop_proxy_if_running(log_fn=self.update_log)
        if self._tray_icon is not None:
            self._tray_icon.hide()
        QApplication.quit()

    def closeEvent(self, event):
        # Salva a marcação atual das empresas/diário no config.json ao fechar
        try:
            self._salvar_selecao_empresas_config()
        except Exception:
            # aqui não vamos travar o fechamento por causa de erro de gravação
            pass
        if self._quitting_from_tray:
            if self.worker_thread and self.worker_thread.isRunning():
                self.worker_thread.stop()
                self.worker_thread.wait()
                self.worker_thread = None
            self.robot_presence_timer.stop()
            self.dashboard_sync_timer.stop()
            update_robot_status(self.robot_id, "inactive")
            _stop_proxy_if_running(log_fn=self.update_log)
            event.accept()
            return

        event.ignore()
        self.hide()
        if self._tray_icon is not None and not self._tray_icon.icon().isNull():
            self._tray_icon.show()

    # ---------- Início / Agendamento ----------

    def on_start(self):
        if self.automation_running:
            self.update_log("[INFO] Já existe uma automação em execução. Ignorando novo comando de início.")
            return

        # limpamos o estado de parada manual para um novo ciclo
        self._stop_requested_by_user = False

        self._sync_dashboard_runtime_state(log_changes=False)
        base_emp = (self.runtime_paths.get("base_empresas") or self.paths.get("base_empresas") or "").strip()
        if not base_emp:
            QMessageBox.warning(
                self,
                "Caminho não definido",
                "Defina a pasta base das empresas em 'Caminhos Padrão' antes de iniciar."
            )
            self.update_log("[ERRO] Pasta base das empresas não definida. Configure em 'Caminhos Padrão'.")
            return

        # Agendamento
        if self.cbAgendar.isChecked():
            dt = self._parse_schedule_datetime()
            if not dt:
                self.update_log("[ERRO] Data/hora do agendamento inválidas.")
                return
            if dt <= datetime.now():
                self.update_log("[WARN] Agendamento em horário passado/atual. Iniciando imediatamente.")
            else:
                self._scheduled_pending = True
                self._scheduled_start_dt = dt
                self.automation_running = True
                self.btn_iniciar.setEnabled(False)
                self.btn_parar.setEnabled(True)
                self.schedule_timer.start(1000)
                self._update_schedule_countdown()
                self.update_log(f"[INFO] Agendado para {dt.strftime('%d/%m/%Y %H:%M')}. Contagem regressiva iniciada.")
                return

        self.automation_running = True
        self.btn_iniciar.setEnabled(False)
        self.btn_parar.setEnabled(True)
        self.iniciar_automacao()

    def iniciar_automacao(self):
        # operações
        operacoes = []
        if self.rbEntrada.isChecked():
            operacoes = ["Entrada"]
        elif self.rbSaida.isChecked():
            operacoes = ["Saída"]
        elif self.rbAmbos.isChecked():
            operacoes = ["Entrada", "Saída"]

        if not operacoes:
            self.update_log("[WARN] Nenhuma operação selecionada (Entrada/Saída/Ambos).")
            self.automation_running = False
            self.btn_iniciar.setEnabled(True)
            self.btn_parar.setEnabled(False)
            return

        # tipo de download
        if self.rbDownDocs.isChecked():
            download_option = "docs"
        elif self.rbDownEventos.isChecked():
            download_option = "eventos"
        else:
            download_option = "ambos"

        data_inicial = self.line_data_inicial.text().strip()
        data_final = self.line_data_final.text().strip()

        # intervalos padrão (mensais ou único)
        if self.cbProcessarPorMeses.isChecked():
            intervalos_padrao = gerar_intervalos_mensais(data_inicial, data_final)
        else:
            if not data_inicial and not data_final:
                intervalos_padrao = gerar_intervalos_mensais("", "")
            else:
                intervalos_padrao = [(data_inicial, data_final)]

        selecionados = []
        companies_daily_flags = {}
        for item in self.empresa_items:
            if item.checkbox.isChecked():
                selecionados.append(item.ie)
                companies_daily_flags[item.ie] = item.daily_checkbox.isChecked()

        if not selecionados:
            self.update_log("[WARN] Nenhuma empresa selecionada para processamento.")
            self.automation_running = False
            self.btn_iniciar.setEnabled(True)
            self.btn_parar.setEnabled(False)
            return

        # 🔽🔽🔽 É AQUI QUE VOCÊ COLOCA O BLOCO 🔽🔽🔽

        # Lista amigável das empresas selecionadas
        lista_empresas_exe = []
        for ie in selecionados:
            dados = self.empresas.get(ie, {})
            nome = dados.get("display_name", ie)
            lista_empresas_exe.append(f"{nome} - IE {ie_formatada(ie)}")

        if lista_empresas_exe:
            linhas = "\n".join(f"   • {linha}" for linha in lista_empresas_exe)
            self.update_log(
                "[INFO] Empresas selecionadas para esta execução:\n" + linhas
            )

        # 🔼🔼🔼 FIM DO BLOCO NOVO 🔼🔼🔼

        daily_interval = None
        if self.cbProcessarPorDias.isChecked():
            txt = self.line_intervalo_dias.text().strip()
            if not txt:
                self.update_log("[WARN] Intervalo de dias não informado. Preencha antes de iniciar.")
                self.automation_running = False
                self.btn_iniciar.setEnabled(True)
                self.btn_parar.setEnabled(False)
                return
            try:
                daily_interval = int(txt)
            except ValueError:
                self.update_log("[ERRO] Número de dias inválido para o intervalo diário.")
                self.automation_running = False
                self.btn_iniciar.setEnabled(True)
                self.btn_parar.setEnabled(False)
                return

        self._sync_dashboard_runtime_state(log_changes=False)
        base_empresas_dir = (self.runtime_paths.get("base_empresas") or self.paths.get("base_empresas") or "").strip()
        relatorios_pdf_dir = (self.runtime_paths.get("relatorios_pdf") or self.paths.get("relatorios_pdf") or "").strip()

        # Salva seleção atual (empresas marcadas + diário) no config.json
        try:
            self._salvar_selecao_empresas_config()
        except Exception as e:
            self.update_log(f"[WARN] Não foi possível salvar marcação das empresas: {e}")

        self.worker_thread = AutomationWorker(
            empresas=selecionados,
            empresas_data=self.empresas,
            base_empresas_dir=base_empresas_dir,
            operacoes=operacoes,
            intervalos_meses=intervalos_padrao,
            data_inicial=data_inicial,
            data_final=data_final,
            daily_interval=daily_interval,
            companies_daily_flags=companies_daily_flags,
            multiplos_meses=self.cbProcessarPorMeses.isChecked(),
            download_option=download_option,
            enviar_supabase=bool(getattr(self, "chk_enviar_supabase", None) and self.chk_enviar_supabase.isChecked()),
            # --- Separação por modelo (55/65) ---
            separar_modelo_xml=self.cbSepararModelos.isChecked(),
            folder_structure=self.paths.get("estrutura_pastas", self.folder_structure),
        )
        update_robot_status(self.robot_id, "processing")

        # guarda flags para uso no PDF
        self._last_companies_daily_flags = companies_daily_flags
        self._last_daily_interval = daily_interval
        self._last_relatorios_pdf_dir = relatorios_pdf_dir

        # 🔴 O QUE FALTAVA: conectar sinais e iniciar a thread
        self.worker_thread.log_signal.connect(self.update_log)
        if getattr(self, 'dashboard_widget', None) is not None and hasattr(self.worker_thread, 'metrics_signal'):
            try:
                self.worker_thread.metrics_signal.connect(self._on_dashboard_metrics)
            except Exception:
                pass
        self.worker_thread.finished.connect(self.on_process_finished)
        self.worker_thread.start()

        # opcional: um logzinho de confirmação
        self.update_log("[INFO] Automação iniciada (Worker em execução).")

        # guarda flags para uso no PDF
        self._last_companies_daily_flags = companies_daily_flags
        self._last_daily_interval = daily_interval

    def on_process_finished(self):
        pdf_path = None
        stopped_by_user = getattr(self, "_stop_requested_by_user", False)
        queue_job = getattr(self, "_active_queue_job", None)

        if self.worker_thread:
            resultados = self.worker_thread.resultados
            try:
                pasta_rel = self.paths.get("relatorios_pdf", "").strip()
                pasta_rel = (getattr(self, "_last_relatorios_pdf_dir", "") or self.runtime_paths.get("relatorios_pdf") or pasta_rel).strip()
                # Agora SEMPRE tenta gerar relatório, mesmo se foi parada manualmente
                if resultados and pasta_rel:
                    try:
                        os.makedirs(pasta_rel, exist_ok=True)
                    except Exception:
                        pass
                    pdf_path = gerar_relatorio_pdf(
                        resultados=resultados,
                        empresas_data=self.empresas,
                        companies_daily_flags=getattr(self, "_last_companies_daily_flags", {}),
                        daily_interval=getattr(self, "_last_daily_interval", None),
                        pasta_relatorios=pasta_rel
                    )
            except Exception as e:
                self.update_log(f"[ERRO] Falha ao gerar relatório PDF: {e}")
            self.worker_thread = None

        _stop_proxy_if_running(log_fn=self.update_log)

        self.automation_running = False
        self.btn_iniciar.setEnabled(True)
        self.btn_parar.setEnabled(False)
        self._scheduled_pending = False
        self._scheduled_start_dt = None
        self.schedule_timer.stop()
        self.lbl_agenda_countdown.clear()

        if stopped_by_user:
            if pdf_path:
                self.update_log(f"[OK] Execução interrompida pelo usuário (relatório parcial gerado em: {pdf_path}).")
            else:
                self.update_log("[OK] Execução interrompida pelo usuário (sem dados para relatório).")
        else:
            if pdf_path:
                self.update_log(f"[OK] Execução concluída. Relatório salvo em: {pdf_path}")
            else:
                self.update_log("[OK] Execução concluída.")
            self.update_log("[OK] Execução concluída (Worker finalizado e GUI liberada).")

        # atualiza o Dashboard (pode ter novos XMLs baixados)
        try:
            self._request_dashboard_refresh(force=True)
        except Exception:
            pass

        # reseta flag para próxima execução
        self._stop_requested_by_user = False
        update_robot_status(self.robot_id, "active")

        # Finaliza job da fila (agendador do site), se houver
        if queue_job and self._queue_supabase_url and self._queue_supabase_key:
            try:
                complete_execution_request_for_queue(
                    supabase_url=self._queue_supabase_url,
                    supabase_service_key=self._queue_supabase_key,
                    request_id=str(queue_job.get("id")),
                    success=not stopped_by_user,
                    error_message=("Interrompido pelo usuário" if stopped_by_user else None),
                )
            except Exception:
                pass
        self._active_queue_job = None

    def on_stop(self):
        if self._scheduled_pending and (not self.worker_thread or not self.worker_thread.isRunning()):
            self._scheduled_pending = False
            self._scheduled_start_dt = None
            self.schedule_timer.stop()
            self.lbl_agenda_countdown.clear()
            self.automation_running = False
            self.btn_iniciar.setEnabled(True)
            self.btn_parar.setEnabled(False)
            self.update_log("[WARN] ⏹️ Agendamento cancelado pelo usuário.")
            _stop_proxy_if_running(log_fn=self.update_log)
            update_robot_status(self.robot_id, "active")
            return

        # Se o worker estiver rodando, dispara parada
        if self.worker_thread and self.worker_thread.isRunning():
            self._stop_requested_by_user = True
            # seta flag de parada e fecha navegador/playwright
            self.worker_thread.stop()

            # ainda mantendo o "parou, parou" radical
            self.worker_thread.terminate()

            self.btn_parar.setEnabled(False)
            self.update_log("[WARN] ⏹️ Execução interrompida pelo usuário.")
        else:
            self.update_log(
                "[INFO] Nenhuma automação em execução no momento."
            )

        self.automation_running = False
        self.btn_iniciar.setEnabled(True)
        self.btn_parar.setEnabled(False)
        update_robot_status(self.robot_id, "active")
        _stop_proxy_if_running(log_fn=self.update_log)

    # ---------- Gestão de Empresas (Adicionar / Editar / Excluir) ----------

    def abrir_gerenciador_empresas(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Gerenciador de Empresas")
        dlg.setModal(True)
        dlg.setFixedWidth(330)
        dlg.setStyleSheet("""
            QDialog {
                background-color: #0B1220;
            }
            QLabel {
                color: #E8F4FF;
                font-family: Verdana;
                font-size: 10pt;
            }
        """)
        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        frame = QFrame()
        frame.setStyleSheet("QFrame { background-color: #0F172A; border-radius: 10px; }")
        frame_layout = QVBoxLayout(frame)
        frame_layout.setContentsMargins(15, 15, 15, 15)
        frame_layout.setSpacing(8)

        title = QLabel("Gerenciar Empresas")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("QLabel { font-size: 11pt; font-weight: bold; }")
        frame_layout.addWidget(title)
        frame_layout.addSpacing(6)

        btn_add = QPushButton("Adicionar Empresa")
        btn_add.setMinimumHeight(32)
        btn_add.setMinimumWidth(220)
        set_button_style(btn_add, "#27AE60", "#2ECC71", "#1E8449")
        btn_add.setIcon(load_icon("adicionar.png"))

        btn_edit = QPushButton("Editar Empresa")
        btn_edit.setMinimumHeight(32)
        btn_edit.setMinimumWidth(220)
        set_button_style(btn_edit, "#2980B9", "#3498DB", "#2471A3")
        btn_edit.setIcon(load_icon("limpar-editar.png"))

        btn_del = QPushButton("Excluir Empresa")
        btn_del.setMinimumHeight(32)
        btn_del.setMinimumWidth(220)
        set_button_style(btn_del, "#C0392B", "#E74C3C", "#922B21")
        btn_del.setIcon(load_icon("excluir.png"))

        btn_paths = QPushButton("Caminhos Padrão")
        btn_paths.setMinimumHeight(32)
        btn_paths.setMinimumWidth(220)
        set_button_style(btn_paths, "#F1C40F", "#F4D03F", "#D4AC0D")
        btn_paths.setIcon(load_icon("caminhos.png"))

        btn_excel = QPushButton("Importar / Exportar")
        btn_excel.setMinimumHeight(32)
        btn_excel.setMinimumWidth(220)
        set_button_style(btn_excel, "#1E8449", "#229954", "#196F3D")
        btn_excel.setIcon(load_icon("excel.png"))

        btn_cfops = QPushButton("Gerenciar CFOPs")
        btn_cfops.setMinimumHeight(32)
        btn_cfops.setMinimumWidth(220)
        set_button_style(btn_cfops, "#9B59B6", "#AF7AC5", "#884EA0")
        btn_cfops.setIcon(load_icon("cfop.png"))

        btn_close = QPushButton("Fechar")
        btn_close.setMinimumHeight(28)
        set_button_style(btn_close, "#102A4C", "#12304E", "#0F1E35")
        btn_close.setIcon(load_icon("desmarcar-cancelar.png"))

        frame_layout.addWidget(btn_add, alignment=Qt.AlignCenter)
        frame_layout.addWidget(btn_edit, alignment=Qt.AlignCenter)
        frame_layout.addWidget(btn_del, alignment=Qt.AlignCenter)
        frame_layout.addWidget(btn_paths, alignment=Qt.AlignCenter)
        frame_layout.addWidget(btn_excel, alignment=Qt.AlignCenter)
        frame_layout.addWidget(btn_cfops, alignment=Qt.AlignCenter)
        frame_layout.addSpacing(4)
        frame_layout.addWidget(btn_close, alignment=Qt.AlignRight)

        layout.addWidget(frame)

        btn_add.clicked.connect(self.adicionar_ie)
        btn_edit.clicked.connect(self.editar_ie)
        btn_del.clicked.connect(self.excluir_ie)
        btn_paths.clicked.connect(self.abrir_caminhos_padrao)
        btn_excel.clicked.connect(self.abrir_opcoes_excel_empresas)
        btn_cfops.clicked.connect(self.abrir_gerenciador_cfops)
        btn_close.clicked.connect(dlg.accept)

        dlg.exec()


    def abrir_opcoes_excel_empresas(self):
        """Mostra as opções de importar/exportar a lista de empresas em Excel."""
        dlg = QDialog(self)
        dlg.setWindowTitle("Importar / Exportar Empresas")
        dlg.setModal(True)
        dlg.setFixedWidth(440)
        dlg.setStyleSheet("""
            QDialog { background-color: #0B1220; }
            QLabel { color: #E8F4FF; font-family: Verdana; font-size: 10pt; }
        """)

        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        frame = QFrame()
        frame.setStyleSheet("QFrame { background-color: #0F172A; border-radius: 10px; }")
        frame_layout = QVBoxLayout(frame)
        frame_layout.setContentsMargins(14, 14, 14, 14)
        frame_layout.setSpacing(10)

        lbl = QLabel("Escolha importar ou exportar a lista (Nome, CNPJ, Inscrição Estadual, Login CPF) em formato Excel.")
        lbl.setWordWrap(True)
        frame_layout.addWidget(lbl)

        cb_substituir = QCheckBox("Substituir a lista atual ao importar")
        cb_substituir.setChecked(False)
        cb_substituir.setToolTip("Desmarque para apenas acrescentar novas empresas. CNPJs já cadastrados serão ignorados.")
        cb_substituir.setStyleSheet("QCheckBox { color: #E8F4FF; font-family: Verdana; font-size: 10pt; }")
        frame_layout.addWidget(cb_substituir)

        actions = QHBoxLayout()
        actions.setSpacing(10)

        btn_import = QPushButton("Importar do Excel")
        btn_import.setMinimumHeight(32)
        btn_import.setMinimumWidth(170)
        btn_import.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        set_button_style(btn_import, "#1E8449", "#229954", "#196F3D")
        btn_import.setIcon(load_icon("excel.png"))
        actions.addWidget(btn_import)

        btn_export = QPushButton("Exportar para Excel")
        btn_export.setMinimumHeight(32)
        btn_export.setMinimumWidth(170)
        btn_export.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        set_button_style(btn_export, "#0E6655", "#148F77", "#0B5345")
        btn_export.setIcon(load_icon("excel.png"))
        actions.addWidget(btn_export)

        frame_layout.addLayout(actions)

        btn_close = QPushButton("Fechar")
        btn_close.setMinimumHeight(28)
        set_button_style(btn_close, "#102A4C", "#12304E", "#0F1E35")
        btn_close.setIcon(load_icon("desmarcar-cancelar.png"))
        frame_layout.addWidget(btn_close, alignment=Qt.AlignRight)

        layout.addWidget(frame)

        def _do_import():
            dlg.accept()
            self.importar_empresas_excel(substituir_lista=cb_substituir.isChecked())

        def _do_export():
            dlg.accept()
            self.exportar_empresas_excel()

        btn_import.clicked.connect(_do_import)
        btn_export.clicked.connect(_do_export)
        btn_close.clicked.connect(dlg.reject)

        dlg.exec()


    def abrir_gerenciador_cfops(self):
        """Abre o gerenciador de CFOPs (Vendas / Compras / Ignorados).

        Observação:
          - O botão Salvar fica DENTRO de cada categoria.
          - Este dialog principal tem apenas Fechar.
        """
        dlg = QDialog(self)
        dlg.setWindowTitle("Gerenciar CFOPs")
        dlg.setModal(True)
        dlg.setFixedWidth(360)
        dlg.setStyleSheet("""
            QDialog { background-color: #0B1220; }
            QLabel { color: #E8F4FF; font-family: Verdana; font-size: 10pt; }
            QPushButton {
                font:9pt 'Verdana'; font-weight:bold; color:#fff; padding:8px 14px;
                border-radius:8px; border:1px solid rgba(255,255,255,0.08);
                background:#102A4C;
            }
            QPushButton:hover { background:#12304E; }
            QPushButton:pressed { background:#0F1E35; padding-top:11px; padding-bottom:9px; }
        """)

        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        frame = QFrame()
        frame.setStyleSheet("QFrame { background-color: #0F172A; border-radius: 10px; }")
        frame_layout = QVBoxLayout(frame)
        frame_layout.setContentsMargins(15, 15, 15, 15)
        frame_layout.setSpacing(8)

        title = QLabel("Gerenciar CFOPs")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("QLabel { font-size: 11pt; font-weight: bold; }")
        frame_layout.addWidget(title)
        frame_layout.addSpacing(6)

        btn_vendas = QPushButton("Vendas (Saídas)")
        btn_vendas.setMinimumHeight(32)
        btn_vendas.setMinimumWidth(240)
        set_button_style(btn_vendas, "#27AE60", "#2ECC71", "#1E8449")

        btn_compras = QPushButton("Compras/Despesas (Entradas)")
        btn_compras.setMinimumHeight(32)
        btn_compras.setMinimumWidth(240)
        set_button_style(btn_compras, "#2980B9", "#3498DB", "#2471A3")

        btn_ign = QPushButton("Ignorados (não somar)")
        btn_ign.setMinimumHeight(32)
        btn_ign.setMinimumWidth(240)
        set_button_style(btn_ign, "#C0392B", "#E74C3C", "#922B21")

        btn_close = QPushButton("Fechar")
        btn_close.setMinimumHeight(28)
        set_button_style(btn_close, "#102A4C", "#12304E", "#0F1E35")

        frame_layout.addWidget(btn_vendas, alignment=Qt.AlignCenter)
        frame_layout.addWidget(btn_compras, alignment=Qt.AlignCenter)
        frame_layout.addWidget(btn_ign, alignment=Qt.AlignCenter)
        frame_layout.addSpacing(6)
        frame_layout.addWidget(btn_close, alignment=Qt.AlignRight)

        layout.addWidget(frame)

        btn_vendas.clicked.connect(lambda: self._abrir_editor_cfops_categoria('venda'))
        btn_compras.clicked.connect(lambda: self._abrir_editor_cfops_categoria('despesa'))
        btn_ign.clicked.connect(lambda: self._abrir_editor_cfops_categoria('ignorar'))
        btn_close.clicked.connect(dlg.accept)

        dlg.exec()

    def _abrir_editor_cfops_categoria(self, categoria: str):
        """Editor de uma categoria (venda/despesa/ignorar) com + / - e Salvar.

        - Fechar: descarta alterações.
        - Salvar: grava no JSON e mantém a tela aberta.
        """
        # carrega mapa e faz cópia local (para fechar sem salvar)
        data = _load_cfop_map()
        local = json.loads(json.dumps(data, ensure_ascii=False))  # deep copy sem depender de copy
        cfops = local.get('cfops', {})
        if not isinstance(cfops, dict):
            cfops = {}
            local['cfops'] = cfops

        cat_label = {'venda': 'Vendas', 'despesa': 'Compras/Despesas', 'ignorar': 'Ignorados'}.get(categoria, categoria)

        dlg = QDialog(self)
        dlg.setWindowTitle(f"CFOPs - {cat_label}")
        dlg.setModal(True)
        dlg.setMinimumSize(560, 420)
        dlg.setStyleSheet("""
            QDialog { background-color: #0B1220; }
            QLabel { color: #E8F4FF; font-family: Verdana; font-size: 10pt; }
            QLineEdit {
                background-color: #0F1E35; color: #E8F4FF; border-radius: 6px;
                padding: 6px; font-family: Verdana; font-size: 9pt;
                border: 1px solid rgba(255,255,255,0.08);
            }
            QListWidget {
                background-color: #0F1E35; color: #E8F4FF; border-radius: 8px;
                padding: 6px; border: 1px solid rgba(255,255,255,0.08);
                font-family: Verdana; font-size: 9pt;
            }
            QPushButton {
                font:9pt 'Verdana'; font-weight:bold; color:#fff; padding:8px 14px;
                border-radius:8px; border:1px solid rgba(255,255,255,0.08);
                background:#102A4C;
            }
            QPushButton:hover { background:#12304E; }
            QPushButton:pressed { background:#0F1E35; padding-top:11px; padding-bottom:9px; }
        """)

        root = QVBoxLayout(dlg)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(10)

        header = QHBoxLayout()
        lbl = QLabel(f"CFOPs em: {cat_label}")
        lbl.setStyleSheet("QLabel { font-size: 11pt; font-weight: bold; }")
        header.addWidget(lbl)
        header.addStretch(1)
        root.addLayout(header)

        flags_dict = local.setdefault('flags', {})
        if categoria in {'venda', 'despesa'}:
            flag_key = f"{categoria}_ignore_same_party"
            _cb_ignore_same = QCheckBox("Ignorar notas com emitente e destinatário iguais")
            _cb_ignore_same.setChecked(bool(flags_dict.get(flag_key)))
            _cb_ignore_same.setToolTip(
                "Quando marcado, notas em que emitente e destinatário coincidem são descartadas "
                "das métricas desta categoria."
            )
            _cb_ignore_same.setStyleSheet("QCheckBox { color: #E8F4FF; font-size: 10pt; }")
            def _sync_ignore_same(checked: bool, fd=flags_dict, key=flag_key):
                fd[key] = bool(checked)
            _cb_ignore_same.toggled.connect(_sync_ignore_same)
            root.addWidget(_cb_ignore_same)

        # lista + botoes +/−
        mid = QHBoxLayout()

        lst = QListWidget()
        # ✅ multi seleção (arrastar/shift/ctrl)
        lst.setSelectionMode(QAbstractItemView.ExtendedSelection)
        try:
            lst.setSelectionRectVisible(True)
        except Exception:
            pass

        btn_col = QVBoxLayout()
        btn_add = QPushButton("+")
        btn_add.setFixedSize(48, 40)
        set_button_style(btn_add, "#27AE60", "#2ECC71", "#1E8449")
        btn_del = QPushButton("-")
        btn_del.setFixedSize(48, 40)
        set_button_style(btn_del, "#C0392B", "#E74C3C", "#922B21")
        btn_col.addWidget(btn_add)
        btn_col.addWidget(btn_del)
        btn_col.addStretch(1)

        mid.addWidget(lst, 1)
        mid.addLayout(btn_col)
        root.addLayout(mid, 1)

        # botoes (Mover / Salvar / Fechar)
        btn_row = QHBoxLayout()

        def _mk_move_btn(_txt: str, c1: str, c2: str, c3: str):
            b = QPushButton(_txt)
            set_button_style(b, c1, c2, c3)
            b.setMinimumHeight(30)
            return b

        move_buttons = []  # [(QPushButton, target_class)]
        if categoria == 'venda':
            b1 = _mk_move_btn('Mover → Compras/Despesas', '#2980B9', '#3498DB', '#2471A3')
            b2 = _mk_move_btn('Mover → Ignorados', '#C0392B', '#E74C3C', '#922B21')
            move_buttons = [(b1, 'despesa'), (b2, 'ignorar')]
        elif categoria == 'despesa':
            b1 = _mk_move_btn('Mover → Vendas', '#27AE60', '#2ECC71', '#1E8449')
            b2 = _mk_move_btn('Mover → Ignorados', '#C0392B', '#E74C3C', '#922B21')
            move_buttons = [(b1, 'venda'), (b2, 'ignorar')]
        elif categoria == 'ignorar':
            b1 = _mk_move_btn('Mover → Vendas', '#27AE60', '#2ECC71', '#1E8449')
            b2 = _mk_move_btn('Mover → Compras/Despesas', '#2980B9', '#3498DB', '#2471A3')
            move_buttons = [(b1, 'venda'), (b2, 'despesa')]

        for _b, _ in move_buttons:
            _b.setEnabled(False)
            btn_row.addWidget(_b)

        btn_row.addStretch(1)

        btn_save = QPushButton('Salvar')
        set_button_style(btn_save, '#2980B9', '#3498DB', '#2471A3')
        btn_close = QPushButton('Fechar')
        set_button_style(btn_close, '#102A4C', '#12304E', '#0F1E35')
        btn_row.addWidget(btn_save)
        btn_row.addWidget(btn_close)
        root.addLayout(btn_row)

        def _item_text(code: str) -> str:
            info = cfops.get(code) or {}
            desc = (info.get('desc') or '').strip()
            return f"{code} - {desc}" if desc else code

        def _refresh_list():
            lst.clear()
            items = []
            for code, info in cfops.items():
                try:
                    cls = (info.get('classe') or 'outro').strip().lower()
                except Exception:
                    cls = 'outro'
                if cls == categoria:
                    items.append(code)
            for code in sorted(items):
                it = QListWidgetItem(_item_text(code))
                it.setData(Qt.UserRole, code)
                lst.addItem(it)

        def _classe_humana(cls: str) -> str:
            return {'venda': 'Vendas', 'despesa': 'Compras/Despesas', 'ignorar': 'Ignorados'}.get(cls, cls)

        def _add_cfop():
            codes = self._dialog_buscar_cfops(cfops, titulo=f"Adicionar CFOP(s) em {cat_label}")
            if not codes:
                return

            added = 0
            skipped_same: List[str] = []
            skipped_other: List[Tuple[str, str]] = []  # (cfop, categoria_existente)

            # dedupe mantendo ordem
            seen = set()
            uniq_codes: List[str] = []
            for c in codes:
                c = _norm_cfop(c)
                if not c or c in seen:
                    continue
                seen.add(c)
                uniq_codes.append(c)

            for code in uniq_codes:
                info = cfops.get(code)
                if info is None:
                    # se não existir no mapa, cria (sem desc)
                    info = {'classe': 'outro', 'desc': '', 'tipo': '', 'grupo': ''}
                    cfops[code] = info

                existing = (info.get('classe') or 'outro').strip().lower()
                if existing == categoria:
                    skipped_same.append(code)
                    continue
                if existing in {'venda', 'despesa', 'ignorar'} and existing != categoria:
                    skipped_other.append((code, existing))
                    continue

                info['classe'] = categoria
                added += 1

            _refresh_list()

            if skipped_other or skipped_same:
                parts = []
                if skipped_other:
                    lines = [f"- {c} (já está em {_classe_humana(cat)})" for c, cat in skipped_other]
                    parts.append("Alguns CFOPs não foram adicionados porque já estão em outra categoria:\n" + "\n".join(lines))
                if skipped_same:
                    parts.append("Alguns CFOPs já estavam nesta categoria e foram ignorados:\n- " + "\n- ".join(skipped_same))
                QMessageBox.information(dlg, "Aviso", "\n\n".join(parts))

        def _remove_cfops():
            sel = lst.selectedItems() or []
            if not sel:
                QMessageBox.information(dlg, "Remover", "Selecione um ou mais CFOPs para remover.")
                return
            for it in sel:
                code = it.data(Qt.UserRole)
                if not code:
                    # tenta parse
                    code = (it.text().split(' - ', 1)[0] or '').strip()
                code = _norm_cfop(code)
                if not code:
                    continue
                if code in cfops:
                    cfops[code]['classe'] = 'outro'
            _refresh_list()

        def _move_selected_to(target_cls: str):
            sel = lst.selectedItems() or []
            if not sel:
                QMessageBox.information(dlg, 'Mover', 'Selecione um ou mais CFOPs para mover.')
                return

            for it in sel:
                code = it.data(Qt.UserRole)
                if not code:
                    code = (it.text().split(' - ', 1)[0] or '').strip()
                code = _norm_cfop(code)
                if not code:
                    continue
                if code in cfops:
                    cfops[code]['classe'] = target_cls
            _refresh_list()

        def _sync_action_buttons():
            has_sel = bool(lst.selectedItems())
            try:
                btn_del.setEnabled(has_sel)
            except Exception:
                pass
            for _b, _ in move_buttons:
                try:
                    _b.setEnabled(has_sel)
                except Exception:
                    pass

        def _save():
            try:
                p = _get_cfop_json_path()
                os.makedirs(os.path.dirname(p), exist_ok=True)
                with open(p, 'w', encoding='utf-8') as f:
                    json.dump(local, f, ensure_ascii=False, indent=2)
                # atualiza cache
                global _CFOP_CACHE
                _CFOP_CACHE = local
                QMessageBox.information(dlg, "Salvo", "CFOPs salvos com sucesso.")
                try:
                    self._request_dashboard_refresh(force=True)
                except Exception:
                    pass
            except Exception as e:
                QMessageBox.critical(dlg, "Erro ao salvar", f"Não foi possível salvar o JSON de CFOPs.\n\n{e}")

        btn_add.clicked.connect(_add_cfop)
        btn_del.clicked.connect(_remove_cfops)
        btn_save.clicked.connect(_save)
        btn_close.clicked.connect(dlg.reject)

        # move buttons
        for _b, _target in move_buttons:
            try:
                _b.clicked.connect(lambda _, t=_target: _move_selected_to(t))
            except Exception:
                pass

        # habilita/desabilita ações conforme seleção
        try:
            lst.itemSelectionChanged.connect(_sync_action_buttons)
        except Exception:
            pass
        _sync_action_buttons()

        _refresh_list()
        dlg.exec()

    def _dialog_buscar_cfops(self, cfops_dict: dict, titulo: str = "Selecionar CFOPs") -> List[str]:
        """Dialog para buscar CFOP por código/descrição e selecionar VÁRIOS.

        Suporta seleção estilo Explorer:
          - arrastar para selecionar
          - Shift + clique (intervalo)
          - Ctrl + clique (multi)
        """
        dlg = QDialog(self)
        dlg.setWindowTitle(titulo)
        dlg.setModal(True)
        dlg.setMinimumSize(560, 420)
        dlg.setStyleSheet("""
            QDialog { background-color: #0B1220; }
            QLabel { color: #E8F4FF; font-family: Verdana; font-size: 10pt; }
            QLineEdit {
                background-color: #0F1E35; color: #E8F4FF; border-radius: 6px;
                padding: 6px; font-family: Verdana; font-size: 9pt;
                border: 1px solid rgba(255,255,255,0.08);
            }
            QListWidget {
                background-color: #0F1E35; color: #E8F4FF; border-radius: 8px;
                padding: 6px; border: 1px solid rgba(255,255,255,0.08);
                font-family: Verdana; font-size: 9pt;
            }
            QPushButton {
                font:9pt 'Verdana'; font-weight:bold; color:#fff; padding:8px 14px;
                border-radius:8px; border:1px solid rgba(255,255,255,0.08);
                background:#102A4C;
            }
            QPushButton:hover { background:#12304E; }
            QPushButton:pressed { background:#0F1E35; padding-top:11px; padding-bottom:9px; }
        """)

        root = QVBoxLayout(dlg)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(10)

        le = QLineEdit()
        le.setPlaceholderText("Digite o CFOP (ex: 5102 ou 51-02) ou parte da descrição")
        root.addWidget(le)

        lst = QListWidget()
        lst.setSelectionMode(QAbstractItemView.ExtendedSelection)
        try:
            lst.setSelectionRectVisible(True)
        except Exception:
            pass
        root.addWidget(lst, 1)

        btns = QHBoxLayout()
        btns.addStretch(1)
        b_ok = QPushButton("OK")
        set_button_style(b_ok, "#2980B9", "#3498DB", "#2471A3")
        b_cancel = QPushButton("Cancelar")
        set_button_style(b_cancel, "#102A4C", "#12304E", "#0F1E35")
        btns.addWidget(b_ok)
        btns.addWidget(b_cancel)
        root.addLayout(btns)

        chosen: Dict[str, Any] = {'codes': []}

        # prepara lista base ordenada (para performance)
        items: List[Tuple[str, str]] = []
        for code, info in (cfops_dict or {}).items():
            c = _norm_cfop(code)
            if not c:
                continue
            desc = (info.get('desc') or '').strip()
            items.append((c, desc))
        items.sort(key=lambda x: x[0])

        def refresh():
            q_raw = (le.text() or '').strip().lower()
            q_digits = re.sub(r'\D', '', q_raw)
            lst.clear()
            for code, desc in items:
                label = f"{code} - {desc}" if desc else code

                ok_code = (not q_raw) or (q_raw in code)
                if (not ok_code) and q_digits:
                    ok_code = q_digits in code

                ok_desc = (desc and q_raw and q_raw in desc.lower())

                if not q_raw or ok_code or ok_desc:
                    it = QListWidgetItem(label)
                    it.setData(Qt.UserRole, code)
                    lst.addItem(it)
            if lst.count() > 0 and lst.currentRow() < 0:
                lst.setCurrentRow(0)

        def accept_selection():
            sel = lst.selectedItems() or []
            if not sel:
                it = lst.currentItem()
                if it:
                    sel = [it]
            if not sel:
                return
            codes: List[str] = []
            for it in sel:
                c = it.data(Qt.UserRole) or (it.text().split(' - ', 1)[0] if it.text() else None)
                c = _norm_cfop(c)
                if c:
                    codes.append(c)
            chosen['codes'] = codes
            dlg.accept()

        def on_return_pressed():
            # se só existe 1 item, confirma direto
            if lst.count() == 1:
                lst.setCurrentRow(0)
                it = lst.item(0)
                if it:
                    it.setSelected(True)
                accept_selection()
            else:
                lst.setFocus(Qt.TabFocusReason)

        le.textChanged.connect(refresh)
        le.returnPressed.connect(on_return_pressed)
        lst.itemDoubleClicked.connect(lambda _: accept_selection())
        b_ok.clicked.connect(accept_selection)
        b_cancel.clicked.connect(dlg.reject)

        refresh()

        if dlg.exec() == QDialog.Accepted:
            # dedupe mantendo ordem
            out: List[str] = []
            seen = set()
            for c in (chosen.get('codes') or []):
                c = _norm_cfop(c)
                if c and c not in seen:
                    seen.add(c)
                    out.append(c)
            return out
        return []

    def _dialog_buscar_cfop(self, cfops_dict: dict, titulo: str = "Selecionar CFOP") -> Optional[str]:
        """Compat: retorna apenas 1 CFOP (o primeiro selecionado)."""
        codes = self._dialog_buscar_cfops(cfops_dict, titulo=titulo)
        return codes[0] if codes else None

    def exportar_empresas_excel(self):
        """Exporta a lista de empresas para uma planilha Excel (Nome, CNPJ, IE, Login CPF)."""
        if not self.empresas:
            QMessageBox.information(self, "Exportar Excel", "Nenhuma empresa cadastrada para exportar.")
            self.update_log("[INFO] Exportação Excel cancelada: lista vazia.")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.warning(
                self,
                "Exportar Excel",
                "Biblioteca 'openpyxl' nÇœo encontrada. Instale com:\n\npip install openpyxl"
            )
            self.update_log("[ERRO] Dependência 'openpyxl' ausente para exportar Excel.")
            return

        base_dir = self.paths.get("base_empresas", "") or BASE_DIR
        if not os.path.isdir(base_dir):
            base_dir = BASE_DIR
        sugestao = os.path.join(base_dir, "empresas.xlsx")

        caminho, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar lista de empresas",
            sugestao,
            "Planilha Excel (*.xlsx);;Todos os arquivos (*)"
        )
        if not caminho:
            self.update_log("[INFO] Exportação Excel cancelada pelo usuário.")
            return
        if not caminho.lower().endswith(".xlsx"):
            caminho += ".xlsx"

        try:
            os.makedirs(os.path.dirname(caminho), exist_ok=True)
        except Exception:
            pass

        wb = Workbook()
        ws = wb.active
        ws.title = "Empresas"
        ws.append(["Nome da Empresa", "CNPJ", "Inscrição Estadual", "Login CPF"])

        for ie in sorted(self.produtos, key=lambda x: self.empresas[x]['display_name'].lower()):
            info = self.empresas.get(ie, {})
            nome = (info.get("display_name") or "").strip()
            cnpj_raw = info.get("cnpj", "")
            cnpj_fmt = cnpj_formatado(cnpj_raw) if cnpj_raw else ""
            login_cpf_raw = cpf_somente_digitos(info.get("login_cpf", ""))
            login_cpf_fmt = cpf_formatado(login_cpf_raw) if login_cpf_raw else ""
            ws.append([nome, cnpj_fmt, ie_formatada(ie), login_cpf_fmt])

        try:
            widths = [32, 22, 22, 18]
            for idx, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = width
        except Exception:
            pass

        try:
            wb.save(caminho)
        except Exception as e:
            QMessageBox.critical(self, "Exportar Excel", f"Não foi possível salvar o arquivo.\n\nDetalhes: {e}")
            self.update_log(f"[ERRO] Falha ao exportar empresas para Excel: {e}")
            return

        self.update_log(f"[OK] Lista de empresas exportada para '{caminho}'.")
        QMessageBox.information(self, "Exportar Excel", f"Arquivo salvo com sucesso:\n{caminho}")

    def importar_empresas_excel(self, substituir_lista: bool = True):
        """Importa empresas a partir de um arquivo Excel (Nome, CNPJ, IE, Login CPF).

        Args:
            substituir_lista: quando True substitui toda a lista atual; quando False apenas
                acrescenta novos CNPJs que ainda não existem.
        """
        QMessageBox.information(
            self,
            "Cadastro centralizado",
            "As empresas agora são gerenciadas no site. Use o painel para cadastrar empresa, IE e logins da SEFAZ GO.",
        )
        return
        try:
            from openpyxl import load_workbook
        except ImportError:
            QMessageBox.warning(
                self,
                "Importar Excel",
                "Biblioteca 'openpyxl' não encontrada. Instale com:\n\npip install openpyxl"
            )
            self.update_log("[ERRO] Dependência 'openpyxl' ausente para importar Excel.")
            return

        base_dir = self.paths.get("base_empresas", "") or BASE_DIR
        if not os.path.isdir(base_dir):
            base_dir = BASE_DIR
        caminho, _ = QFileDialog.getOpenFileName(
            self,
            "Importar lista de empresas",
            base_dir,
            "Planilha Excel (*.xlsx);;Todos os arquivos (*)"
        )
        if not caminho:
            self.update_log("[INFO] Importação Excel cancelada pelo usuário.")
            return

        try:
            wb = load_workbook(caminho, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            QMessageBox.critical(self, "Importar Excel", f"Não foi possível abrir o arquivo selecionado.\n\nDetalhes: {e}")
            self.update_log(f"[ERRO] Falha ao abrir Excel de empresas: {e}")
            return

        def _digits_from_cell(val: Any, expected_len: Optional[int] = None) -> str:
            if val is None:
                return ""
            was_numeric = isinstance(val, (int, float))
            if isinstance(val, float):
                if abs(val - round(val)) < 1e-6:
                    val = int(round(val))
            txt = str(val).strip()
            if isinstance(val, float) and "e" in txt.lower():
                try:
                    txt = f"{int(val):d}"
                    was_numeric = True
                except Exception:
                    pass
            digits = "".join(ch for ch in txt if ch.isdigit())
            if expected_len and was_numeric and len(digits) == expected_len - 1:
                digits = digits.zfill(expected_len)
            return digits

        def _text(val: Any) -> str:
            return str(val).strip() if val is not None else ""

        novas_empresas: Dict[str, Dict[str, str]] = {}
        avisos: List[str] = []

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            nome = _text(row[0] if len(row) > 0 else None)
            cnpj_digits = _digits_from_cell(row[1] if len(row) > 1 else None, expected_len=14)
            ie_digits = _digits_from_cell(row[2] if len(row) > 2 else None, expected_len=9)
            login_cpf_digits = _digits_from_cell(row[3] if len(row) > 3 else None, expected_len=11)

            if not ie_digits:
                avisos.append(f"Linha {idx}: sem IE, ignorada.")
                continue

            if ie_digits in novas_empresas:
                avisos.append(f"Linha {idx}: IE duplicada, mantida a primeira ocorrência.")
                continue

            entry = {"display_name": nome or ie_formatada(ie_digits)}
            if cnpj_digits:
                if len(cnpj_digits) == 14:
                    entry["cnpj"] = cnpj_digits
                else:
                    avisos.append(f"Linha {idx}: CNPJ com tamanho inválido, registrado sem CNPJ.")
            if login_cpf_digits:
                if len(login_cpf_digits) == 11:
                    entry["login_cpf"] = login_cpf_digits
                else:
                    avisos.append(f"Linha {idx}: Login CPF com tamanho inválido, vínculo ignorado.")
            novas_empresas[ie_digits] = entry

        try:
            wb.close()
        except Exception:
            pass

        if not novas_empresas:
            QMessageBox.warning(self, "Importar Excel", "Nenhuma linha válida (com IE) foi encontrada na planilha.")
            self.update_log("[WARN] Importação Excel ignorada: planilha sem dados válidos.")
            return

        if substituir_lista:
            if not self._dialog_confirmacao(
                "Confirmar importação",
                f"Isto substituirá a lista atual ({len(self.empresas)} empresas) por {len(novas_empresas)} registros do Excel. Deseja continuar?"
            ):
                self.update_log("[INFO] Importação Excel cancelada pelo usuário.")
                return
            empresas_resultantes = novas_empresas
            adicionadas = len(novas_empresas)
            modo_log = "substituir"
        else:
            cnpjs_existentes = {info.get("cnpj") for info in self.empresas.values() if info.get("cnpj")}
            novas_validas: Dict[str, Dict[str, str]] = {}
            cnpjs_novos = set()

            for ie, entry in novas_empresas.items():
                cnpj = entry.get("cnpj")
                if ie in self.empresas:
                    avisos.append(f"IE já cadastrada, ignorada: {ie_formatada(ie)}.")
                    continue
                if cnpj and (cnpj in cnpjs_existentes or cnpj in cnpjs_novos):
                    avisos.append(f"CNPJ já cadastrado, ignorado: {cnpj_formatado(cnpj)}.")
                    continue
                novas_validas[ie] = entry
                if cnpj:
                    cnpjs_novos.add(cnpj)

            if not novas_validas:
                QMessageBox.information(
                    self,
                    "Importar Excel",
                    "Nenhuma empresa nova para acrescentar. Os CNPJs da planilha já estão cadastrados."
                )
                self.update_log("[INFO] Importação Excel cancelada: nenhum CNPJ novo para acrescentar.")
                return

            if not self._dialog_confirmacao(
                "Confirmar importação",
                f"Isto acrescentará {len(novas_validas)} novas empresas à lista atual ({len(self.empresas)} empresas). Deseja continuar?"
            ):
                self.update_log("[INFO] Importação Excel cancelada pelo usuário.")
                return

            empresas_resultantes = {**self.empresas, **novas_validas}
            adicionadas = len(novas_validas)
            modo_log = "acrescentar"

        self.empresas = empresas_resultantes
        self.produtos = sorted(self.empresas.keys(), key=lambda ie: self.empresas[ie]['display_name'].lower())
        self._cfg_empresas_marcadas = {ie for ie in getattr(self, "_cfg_empresas_marcadas", set()) if ie in self.empresas}
        self._cfg_empresas_diario = {ie for ie in getattr(self, "_cfg_empresas_diario", set()) if ie in self.empresas}

        self._recarregar_lista_empresas()
        self._salvar_empresas_config()
        self._salvar_selecao_empresas_config()

        if avisos:
            avisos_preview = "\n".join(avisos[:5])
            if len(avisos) > 5:
                avisos_preview += "\n..."
            QMessageBox.information(
                self,
                "Importar Excel",
                f"{adicionadas} empresas importadas.\n\nAvisos:\n{avisos_preview}"
            )
        else:
            QMessageBox.information(self, "Importar Excel", f"{adicionadas} empresas importadas com sucesso.")

        self.update_log(
            f"[OK] Importação via Excel concluída ({modo_log}): {adicionadas} empresas. "
            + (f"Avisos: {len(avisos)}." if avisos else "")
        )

    def adicionar_ie(self):
        QMessageBox.information(
            self,
            "Cadastro centralizado",
            "As empresas agora são cadastradas no site. Use o painel para incluir empresa, IE e logins da SEFAZ GO.",
        )
        return
        nome, cnpj_digits, ie_digits, login_cpf, ok = self._dialog_empresa_form(
            "Adicionar Empresa",
            nome_init="",
            cnpj_init="",
            ie_init="",
        )
        if not ok:
            self.update_log("ℹ️ [EMPRESAS] Adição cancelada.")
            return

        ie = ie_somente_digitos(ie_digits)
        cnpj_digits = cnpj_somente_digitos(cnpj_digits)

        if ie in self.empresas:
            QMessageBox.warning(self, "Erro", f"IE '{ie_formatada(ie)}' já cadastrada.")
            self.update_log(f"⚠️ [EMPRESAS] IE '{ie_formatada(ie)}' já existe, operação cancelada.")
            return

        display_name = nome.strip()
        entry = {"display_name": display_name, "cnpj": cnpj_digits}
        if login_cpf:
            entry["login_cpf"] = login_cpf
        self.empresas[ie] = entry
        self.produtos.append(ie)

        # atualiza lista visual já em ordem alfabética
        self._recarregar_lista_empresas()
        self._salvar_empresas_config()
        
        # Envia empresa para Supabase (apenas se checkbox estiver marcado)
        try:
            config = carregar_config()
            enviar_supabase = config.get("enviar_supabase", True)
        except Exception:
            enviar_supabase = True
        
        if cnpj_digits and SUPABASE_AVAILABLE and enviar_supabase:
            try:
                from threading import Thread
                def _enviar_empresa():
                    try:
                        company_id = ensure_company_exists_supabase(cnpj_digits, display_name)
                        if company_id:
                            print(f"[SUPABASE] ✅ Empresa adicionada ao Supabase: {display_name}")
                    except Exception as e:
                        print(f"[ERRO] Falha ao enviar empresa para Supabase: {e}")
                
                thread = Thread(target=_enviar_empresa, daemon=True)
                thread.start()
            except Exception:
                pass
        
        self.update_log(f"✅ [EMPRESAS] Nova empresa adicionada: {display_name} - {ie_formatada(ie)}")

    def editar_ie(self):
        QMessageBox.information(
            self,
            "Cadastro centralizado",
            "A edição das empresas agora é feita no site.",
        )
        return
        if not self.produtos:
            self.update_log("ℹ️ [EMPRESAS] Nenhuma empresa para editar.")
            return

        itens_lista = [
            f"{self.empresas[ie]['display_name']} - {ie_formatada(ie)}"
            for ie in self.produtos
        ]

        selecionado, ok = self._dialog_selecionar_empresa(
            "Editar Empresa",
            "Selecione a empresa para editar:",
            itens_lista
        )
        if not (ok and selecionado):
            self.update_log("ℹ️ [EMPRESAS] Edição cancelada.")
            return

        try:
            nome_sel, ie_masc = selecionado.rsplit(" - ", 1)
        except ValueError:
            self.update_log("❌ [EMPRESAS] Não foi possível identificar a empresa selecionada.")
            return

        ie_selecionada = ie_somente_digitos(ie_masc)
        dados_atual = self.empresas.get(ie_selecionada, {})
        nome_atual = dados_atual.get("display_name", "")
        cnpj_atual = dados_atual.get("cnpj", "")

        novo_nome, novo_cnpj, nova_ie, novo_login_cpf, ok_edit = self._dialog_empresa_form(
            "Editar Empresa",
            nome_init=nome_atual,
            cnpj_init=cnpj_atual,
            ie_init=ie_selecionada,
            login_cpf_init=dados_atual.get("login_cpf", ""),
        )
        if not ok_edit:
            self.update_log("ℹ️ [EMPRESAS] Edição cancelada.")
            return

        nova_ie = ie_somente_digitos(nova_ie)
        novo_cnpj = cnpj_somente_digitos(novo_cnpj)

        if nova_ie != ie_selecionada and nova_ie in self.empresas:
            QMessageBox.warning(self, "Erro", f"IE '{ie_formatada(nova_ie)}' já cadastrada.")
            self.update_log(f"⚠️ [EMPRESAS] IE '{ie_formatada(nova_ie)}' já existe, operação cancelada.")
            return

        # Atualiza estrutura
        del self.empresas[ie_selecionada]
        entry = {"display_name": novo_nome.strip(), "cnpj": novo_cnpj}
        if novo_login_cpf:
            entry["login_cpf"] = novo_login_cpf
        self.empresas[nova_ie] = entry

        if ie_selecionada in self.produtos:
            idx = self.produtos.index(ie_selecionada)
            self.produtos[idx] = nova_ie

        # Recarrega lista visual já em ordem alfabética
        self._recarregar_lista_empresas()
        self._salvar_empresas_config()
        
        # Atualiza empresa no Supabase (apenas se checkbox estiver marcado)
        try:
            config = carregar_config()
            enviar_supabase = config.get("enviar_supabase", True)
        except Exception:
            enviar_supabase = True
        
        if novo_cnpj and SUPABASE_AVAILABLE and enviar_supabase:
            try:
                from threading import Thread
                def _atualizar_empresa():
                    try:
                        company_id = ensure_company_exists_supabase(novo_cnpj, novo_nome.strip())
                        if company_id:
                            print(f"[SUPABASE] ✅ Empresa atualizada no Supabase: {novo_nome.strip()}")
                    except Exception as e:
                        print(f"[ERRO] Falha ao atualizar empresa no Supabase: {e}")
                
                thread = Thread(target=_atualizar_empresa, daemon=True)
                thread.start()
            except Exception:
                pass
        
        self.update_log(
            f"✏️ [EMPRESAS] Editado: {nome_atual} - {ie_formatada(ie_selecionada)} "
            f"→ {novo_nome.strip()} - {ie_formatada(nova_ie)}"
        )

    def excluir_ie(self):
        QMessageBox.information(
            self,
            "Cadastro centralizado",
            "A exclusão das empresas agora é feita no site.",
        )
        return
        if not self.produtos:
            self.update_log("ℹ️ [EMPRESAS] Nenhuma empresa para excluir.")
            return

        itens_lista = [
            f"{self.empresas[ie]['display_name']} - {ie_formatada(ie)}"
            for ie in self.produtos
        ]

        selecionado, ok = self._dialog_selecionar_empresa(
            "Excluir Empresa",
            "Selecione a empresa para excluir:",
            itens_lista
        )
        if not (ok and selecionado):
            self.update_log("ℹ️ [EMPRESAS] Exclusão cancelada.")
            return

        try:
            nome_sel, ie_masc = selecionado.rsplit(" - ", 1)
        except ValueError:
            self.update_log("❌ [EMPRESAS] Não foi possível identificar a empresa selecionada.")
            return

        ie_selecionada = ie_somente_digitos(ie_masc)

        if not self._dialog_confirmacao(
            "Confirmar exclusão",
            f"Confirma excluir a empresa '{nome_sel}' (IE {ie_formatada(ie_selecionada)})?"
        ):
            self.update_log("ℹ️ [EMPRESAS] Exclusão cancelada pelo usuário.")
            return

        if ie_selecionada in self.produtos:
            self.produtos.remove(ie_selecionada)
        if ie_selecionada in self.empresas:
            del self.empresas[ie_selecionada]

        # remove item visual
        for item in list(self.empresa_items):
            if item.ie == ie_selecionada:
                self.empresa_items_layout.removeWidget(item)
                item.deleteLater()
                self.empresa_items.remove(item)
                break

        self._salvar_empresas_config()
        self.update_log(f"🗑️ [EMPRESAS] Empresa '{nome_sel}' (IE {ie_formatada(ie_selecionada)}) excluída.")

    # ---------- Caminhos padrão ----------

    def abrir_caminhos_padrao(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Caminhos Padrão")
        dlg.setModal(True)
        dlg.setMinimumWidth(560)
        dlg.setStyleSheet("""
            QDialog {
                background-color: #0B1220;
            }
            QLabel {
                color: #E8F4FF;
                font-family: Verdana;
                font-size: 10pt;
            }
            QLineEdit {
                background-color: #0F1E35;
                color: #E8F4FF;
                border-radius: 6px;
                padding: 4px;
                font-family: Verdana;
                font-size: 9pt;
            }
            QPushButton {
                font:9pt 'Verdana'; font-weight:bold; color:#fff; padding:8px 14px;
                border-radius:8px; border:1px solid rgba(255,255,255,0.08);
                background:#102A4C;
            }
            QPushButton:hover {
                background:#12304E;
            }
            QPushButton:pressed {
                background:#0F1E35;
                padding-top:11px;
                padding-bottom:9px;
            }
        """)

        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        frame = QFrame()
        frame.setStyleSheet("QFrame { background-color: #0F172A; border-radius: 10px; }")
        frame_layout = QVBoxLayout(frame)
        frame_layout.setContentsMargins(15, 15, 15, 15)
        frame_layout.setSpacing(10)

        title = QLabel("Definir Caminhos Padrão")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("QLabel { font-size: 11pt; font-weight: bold; }")
        frame_layout.addWidget(title)

        # Pasta base das empresas
        lbl_emp = QLabel("Pasta base das empresas (XMLs / ZIPs unificados):")
        frame_layout.addWidget(lbl_emp)
        row_emp = QHBoxLayout()
        le_emp = QLineEdit(self.paths.get("base_empresas", ""))
        btn_emp = QPushButton("...")
        btn_emp.setFixedWidth(64)
        set_button_style(btn_emp, "#2980B9", "#3498DB", "#2471A3")
        row_emp.addWidget(le_emp)
        row_emp.addWidget(btn_emp)
        frame_layout.addLayout(row_emp)

        def selecionar_pasta_emp():
            caminho = QFileDialog.getExistingDirectory(
                dlg, "Selecionar pasta base das empresas",
                self.paths.get("base_empresas", "") or BASE_DIR
            )
            if caminho:
                le_emp.setText(caminho)

        btn_emp.clicked.connect(selecionar_pasta_emp)

        # Pasta de relatórios PDF
        lbl_rel = QLabel("Pasta para salvar relatórios em PDF:")
        frame_layout.addWidget(lbl_rel)
        row_rel = QHBoxLayout()
        le_rel = QLineEdit(self.paths.get("relatorios_pdf", ""))
        btn_rel = QPushButton("...")
        btn_rel.setFixedWidth(64)
        set_button_style(btn_rel, "#2980B9", "#3498DB", "#2471A3")
        row_rel.addWidget(le_rel)
        row_rel.addWidget(btn_rel)
        frame_layout.addLayout(row_rel)

        def selecionar_pasta_rel():
            caminho = QFileDialog.getExistingDirectory(
                dlg, "Selecionar pasta de relatórios",
                self.paths.get("relatorios_pdf", "") or BASE_DIR
            )
            if caminho:
                le_rel.setText(caminho)

        btn_rel.clicked.connect(selecionar_pasta_rel)

        # Estrutura de pastas
        estrutura_cfg = normalizar_estrutura_pastas(self.paths.get("estrutura_pastas"))
        lbl_struct = QLabel("Estrutura de pastas para salvar XMLs:")
        frame_layout.addWidget(lbl_struct)

        struct_layout = QVBoxLayout()
        struct_layout.setSpacing(4)

        cb_nome_emp = QCheckBox("Incluir nome da empresa")
        cb_nome_emp.setChecked(estrutura_cfg.get("usar_nome_empresa", True))
        struct_layout.addWidget(cb_nome_emp)

        cb_pasta_cliente = QCheckBox("Pasta de opção do cliente")
        cb_pasta_cliente.setChecked(estrutura_cfg.get("usar_pasta_cliente", False))
        struct_layout.addWidget(cb_pasta_cliente)

        # Pastas opcionais do cliente (lista) + botão '+'
        pastas_cliente_state = []
        pastas_cliente_initial = list(estrutura_cfg.get("pastas_cliente") or [])
        if not pastas_cliente_initial:
            legacy_one = (estrutura_cfg.get("nome_pasta_cliente") or "").strip()
            if legacy_one:
                pastas_cliente_initial = [legacy_one]

        row_custom = QHBoxLayout()
        row_custom.addSpacing(18)
        le_pasta_cliente = QLineEdit("")
        le_pasta_cliente.setPlaceholderText("Adicionar pasta (ex.: DEPARTAMENTO FISCAL)")
        le_pasta_cliente.setEnabled(cb_pasta_cliente.isChecked())

        btn_add_pasta = QPushButton("+")
        btn_add_pasta.setFixedWidth(38)
        set_button_style(btn_add_pasta, "#2980B9", "#3498DB", "#2471A3")
        btn_add_pasta.setEnabled(cb_pasta_cliente.isChecked())

        row_custom.addWidget(le_pasta_cliente)
        row_custom.addWidget(btn_add_pasta)
        struct_layout.addLayout(row_custom)

        # Lista visual das pastas adicionadas
        w_lista_pastas = QWidget()
        lista_layout = QVBoxLayout(w_lista_pastas)
        lista_layout.setContentsMargins(30, 0, 0, 0)
        lista_layout.setSpacing(4)

        def _pasta_existe(n: str) -> bool:
            return any((x or "").strip().lower() == (n or "").strip().lower() for x in pastas_cliente_state)

        def _add_pasta_item(nome: str):
            nome = (nome or "").strip()
            if not nome or _pasta_existe(nome):
                return

            pastas_cliente_state.append(nome)

            item = QWidget()
            item_row = QHBoxLayout(item)
            item_row.setContentsMargins(0, 0, 0, 0)
            item_row.setSpacing(6)

            lbl = QLabel(nome)
            lbl.setStyleSheet("QLabel { color: #BFE6FF; font-family: Verdana; }")

            btn_del = QPushButton("-")
            btn_del.setFixedSize(34, 28)
            set_small_button_style(btn_del, "#34495E", "#5D6D7E", "#2C3E50")
            item_row.addWidget(lbl)
            item_row.addStretch()
            item_row.addWidget(btn_del)

            def _remove():
                try:
                    for i, v in enumerate(list(pastas_cliente_state)):
                        if (v or "").strip().lower() == nome.lower():
                            pastas_cliente_state.pop(i)
                            break
                except Exception:
                    pass
                item.setParent(None)
                item.deleteLater()

            btn_del.clicked.connect(_remove)
            lista_layout.addWidget(item)

        for p in pastas_cliente_initial:
            _add_pasta_item(p)

        struct_layout.addWidget(w_lista_pastas)

        def _on_add_click():
            raw_txt = le_pasta_cliente.text().strip()
            if not raw_txt:
                return
            # Permite adicionar várias de uma vez separando por ';', '|', ou quebra de linha.
            parts = re.split(r"[;|\n]+", raw_txt)
            for part in parts:
                part = (part or "").strip()
                if not part:
                    continue
                # Se o usuário colar um caminho, quebra em segmentos.
                segs = re.split(r"[\\/]+", part) if ("\\" in part or "/" in part) else [part]
                for s in segs:
                    s = (s or "").strip()
                    if s:
                        _add_pasta_item(s)
            le_pasta_cliente.clear()

        btn_add_pasta.clicked.connect(_on_add_click)

        def _toggle_pasta_cliente(enabled: bool):
            le_pasta_cliente.setEnabled(enabled)
            btn_add_pasta.setEnabled(enabled)
            w_lista_pastas.setEnabled(enabled)

        cb_pasta_cliente.toggled.connect(_toggle_pasta_cliente)
        _toggle_pasta_cliente(cb_pasta_cliente.isChecked())

        cb_sep_op = QCheckBox("Separar por Entrada/Saída")
        cb_sep_op.setChecked(estrutura_cfg.get("separar_entrada_saida", False))
        struct_layout.addWidget(cb_sep_op)

        row_data_struct = QHBoxLayout()
        cb_ano = QCheckBox("Criar pasta do ano (YYYY)")
        cb_ano.setChecked(estrutura_cfg.get("usar_ano", False))
        cb_ano.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        cb_mes = QCheckBox("Criar pasta do mês")
        cb_mes.setChecked(estrutura_cfg.get("usar_mes", True))
        cb_mes.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        row_data_struct.addWidget(cb_ano)
        row_data_struct.addWidget(cb_mes)
        row_data_struct.addStretch()
        struct_layout.addLayout(row_data_struct)

        # Dashboard: modo de busca dos XMLs
        lbl_dash_scan = QLabel("Busca de XMLs do Dashboard:")
        lbl_dash_scan.setStyleSheet("QLabel { color: #9AA3B2; font-family: Verdana; font-size: 9pt; }")
        struct_layout.addWidget(lbl_dash_scan)

        # Se marcado, o dashboard NÃO deduplica pela chave de acesso (44 dígitos).
        # Útil quando o cliente quer ver exatamente a quantidade de XMLs/ZIPs encontrados.
        cb_dash_considerar_dup = QCheckBox("Considerar XMLs duplicados")
        cb_dash_considerar_dup.setChecked(bool(self.paths.get("dashboard_considerar_xml_duplicados", False)))
        cb_dash_considerar_dup.setToolTip("Se marcado, não remove duplicados pela chave de acesso.")
        struct_layout.addWidget(cb_dash_considerar_dup)

        cb_dash_quantidade_real = QCheckBox("Mostrar quantidade real de XMLs (sem deduplicar contagem)")
        cb_dash_quantidade_real.setChecked(bool(self.paths.get("dashboard_mostrar_quantidade_real_xml", False)))
        cb_dash_quantidade_real.setToolTip(
            "Afeta apenas a contagem exibida (total/55/65). Valores de faturamento/despesa permanecem deduplicados."
        )
        cb_dash_quantidade_real.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        cb_dash_quantidade_real.setMinimumWidth(480)
        struct_layout.addWidget(cb_dash_quantidade_real)

        cb_dash_busca_empresa = QCheckBox("Varrer qualquer subpasta dentro da pasta da empresa (mais compatível, porém mais lento)")
        cb_dash_busca_empresa.setChecked(bool(self.paths.get("dashboard_busca_empresa_em_qualquer_pasta", True)))
        cb_dash_busca_empresa.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        cb_dash_busca_empresa.setMinimumWidth(480)
        struct_layout.addWidget(cb_dash_busca_empresa)

        frame_layout.addLayout(struct_layout)

        # botões
        btns = QHBoxLayout()
        btns.setSpacing(8)
        btn_ok = QPushButton("Salvar")
        set_button_style(btn_ok, "#27AE60", "#2ECC71", "#1E8449")
        btn_ok.setIcon(load_icon("salvar.png"))
        btn_cancel = QPushButton("Cancelar")
        set_button_style(btn_cancel, "#C0392B", "#E74C3C", "#922B21")
        btn_cancel.setIcon(load_icon("desmarcar-cancelar.png"))
        btns.addStretch()
        btns.addWidget(btn_ok)
        btns.addWidget(btn_cancel)

        frame_layout.addSpacing(8)
        frame_layout.addLayout(btns)

        layout.addWidget(frame)

        def salvar():
            self.paths["base_empresas"] = le_emp.text().strip()
            self.paths["relatorios_pdf"] = le_rel.text().strip()
            self.paths["dashboard_considerar_xml_duplicados"] = cb_dash_considerar_dup.isChecked()
            self.paths["dashboard_mostrar_quantidade_real_xml"] = cb_dash_quantidade_real.isChecked()
            self.paths["dashboard_busca_empresa_em_qualquer_pasta"] = cb_dash_busca_empresa.isChecked()
            estrutura_cfg = {
                "usar_nome_empresa": cb_nome_emp.isChecked(),
                "usar_pasta_cliente": cb_pasta_cliente.isChecked(),
                "pastas_cliente": pastas_cliente_state,
                "nome_pasta_cliente": pastas_cliente_state[0] if pastas_cliente_state else "",  # legacy
                "separar_entrada_saida": cb_sep_op.isChecked(),
                "usar_ano": cb_ano.isChecked(),
                "usar_mes": cb_mes.isChecked(),
            }
            self.folder_structure = normalizar_estrutura_pastas(estrutura_cfg)
            self.paths["estrutura_pastas"] = self.folder_structure
            self.config["paths"] = self.paths
            salvar_config(self.config)
            self.update_log("[OK] Caminhos padrão atualizados.")
            dlg.accept()

        def cancelar():
            dlg.reject()

        btn_ok.clicked.connect(salvar)
        btn_cancel.clicked.connect(cancelar)

        dlg.exec()

    def _dialog_login_portal_form(self, titulo: str, cpf_init: str = "", senha_init: str = "") -> Tuple[str, str, bool]:
        """Diálogo simples para CPF + senha do portal."""
        dlg = QDialog(self)
        dlg.setWindowTitle(titulo)
        dlg.setModal(True)
        dlg.setFixedWidth(340)
        dlg.setStyleSheet("""
            QDialog { background-color: #0B1220; }
            QLabel { color: #E8F4FF; font-family: Verdana; font-size: 10pt; }
            QLineEdit {
                background-color: #0A1324;
                color: #E8F4FF;
                border-radius: 4px;
                padding: 4px 6px;
                font-family: Verdana;
                font-size: 10pt;
            }
            QPushButton {
                font:9pt 'Verdana'; font-weight:bold; color:#fff; padding:8px 14px;
                border-radius:8px; border:1px solid rgba(255,255,255,0.08);
                background:#102A4C;
            }
            QPushButton:hover { background:#12304E; }
            QPushButton:pressed { background:#0F1E35; padding-top:11px; padding-bottom:9px; }
        """)

        layout = QVBoxLayout(dlg)

        lbl_info = QLabel("Informe o CPF (usuário) e a senha que o robô vai usar no login do portal.")
        lbl_info.setWordWrap(True)
        layout.addWidget(lbl_info)

        lbl_cpf = QLabel("CPF / Usuário:")
        edt_cpf = QLineEdit()
        edt_cpf.setInputMask("000.000.000-00;_")
        if cpf_init:
            edt_cpf.setText(cpf_formatado(cpf_init))

        lbl_senha = QLabel("Senha:")
        edt_senha = QLineEdit()
        edt_senha.setEchoMode(QLineEdit.Password)
        if senha_init:
            edt_senha.setText(senha_init)

        layout.addWidget(lbl_cpf)
        layout.addWidget(edt_cpf)
        layout.addWidget(lbl_senha)
        layout.addWidget(edt_senha)

        buttons_layout = QHBoxLayout()
        btn_salvar = QPushButton("Salvar")
        set_button_style(btn_salvar, "#27AE60", "#2ECC71", "#1E8449")
        btn_salvar.setIcon(load_icon("salvar.png"))
        btn_cancelar = QPushButton("Cancelar")
        set_button_style(btn_cancelar, "#C0392B", "#E74C3C", "#922B21")
        btn_cancelar.setIcon(load_icon("desmarcar-cancelar.png"))

        buttons_layout.addWidget(btn_salvar)
        buttons_layout.addWidget(btn_cancelar)
        layout.addLayout(buttons_layout)

        def on_salvar():
            cpf_digits = cpf_somente_digitos(edt_cpf.text())
            senha = edt_senha.text().strip()
            if len(cpf_digits) != 11:
                QMessageBox.warning(dlg, "Dados inválidos", "Informe um CPF válido.")
                return
            if not senha:
                QMessageBox.warning(dlg, "Dados inválidos", "Informe a senha.")
                return
            dlg.accept()

        btn_salvar.clicked.connect(on_salvar)
        btn_cancelar.clicked.connect(dlg.reject)

        ok = dlg.exec() == QDialog.Accepted
        return (
            cpf_somente_digitos(edt_cpf.text()),
            edt_senha.text().strip(),
            ok,
        )

    def abrir_login_portal(self):
        """Gerenciador de logins do portal (Adicionar / Editar / Excluir)."""
        if not SEFAZ_LOCAL_JSON_CLIENT_MODE:
            QMessageBox.information(
                self,
                "Cadastro centralizado",
                "Os logins da SEFAZ GO agora são gerenciados no site, dentro do cadastro da empresa.",
            )
            return
        dlg = QDialog(self)
        dlg.setWindowTitle("Login do Portal SEFAZ-GO")
        dlg.setModal(True)
        dlg.setFixedWidth(330)
        dlg.setStyleSheet("""
            QDialog { background-color: #0B1220; }
            QLabel { color: #E8F4FF; font-family: Verdana; font-size: 10pt; }
        """)
        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        frame = QFrame()
        frame.setStyleSheet("QFrame { background-color: #0F172A; border-radius: 10px; }")
        frame_layout = QVBoxLayout(frame)
        frame_layout.setContentsMargins(15, 15, 15, 15)
        frame_layout.setSpacing(8)

        title = QLabel("Gerenciar Logins")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("QLabel { font-size: 11pt; font-weight: bold; }")
        frame_layout.addWidget(title)
        frame_layout.addSpacing(6)

        BTN_W = 240
        BTN_H = 32

        btn_add = QPushButton("Adicionar Login")
        btn_add.setFixedHeight(BTN_H)
        btn_add.setFixedWidth(BTN_W)
        set_button_style(btn_add, "#27AE60", "#2ECC71", "#1E8449")
        btn_add.setIcon(load_icon("adicionar.png"))

        btn_edit = QPushButton("Editar Login")
        btn_edit.setFixedHeight(BTN_H)
        btn_edit.setFixedWidth(BTN_W)
        set_button_style(btn_edit, "#2980B9", "#3498DB", "#2471A3")
        btn_edit.setIcon(load_icon("limpar-editar.png"))

        btn_default = QPushButton("Definir Login Padrão")
        btn_default.setFixedHeight(BTN_H)
        btn_default.setFixedWidth(BTN_W)
        set_button_style(btn_default, "#16A085", "#1ABC9C", "#0E6251")
        btn_default.setIcon(load_icon("selecionar.png"))

        btn_move_emp = QPushButton("Mover Empresas para Login")
        btn_move_emp.setFixedHeight(BTN_H)
        btn_move_emp.setFixedWidth(BTN_W)
        set_button_style(btn_move_emp, "#B9770E", "#F5B041", "#7D6608")
        btn_move_emp.setIcon(load_icon("selecionar.png"))

        btn_del = QPushButton("Excluir Login")
        btn_del.setFixedHeight(BTN_H)
        btn_del.setFixedWidth(BTN_W)
        set_button_style(btn_del, "#C0392B", "#E74C3C", "#922B21")
        btn_del.setIcon(load_icon("excluir.png"))

        btn_close = QPushButton("Fechar")
        btn_close.setMinimumHeight(28)
        set_button_style(btn_close, "#102A4C", "#12304E", "#0F1E35")
        btn_close.setIcon(load_icon("desmarcar-cancelar.png"))

        frame_layout.addWidget(btn_add, alignment=Qt.AlignCenter)
        frame_layout.addWidget(btn_edit, alignment=Qt.AlignCenter)
        frame_layout.addWidget(btn_default, alignment=Qt.AlignCenter)
        frame_layout.addWidget(btn_move_emp, alignment=Qt.AlignCenter)
        frame_layout.addWidget(btn_del, alignment=Qt.AlignCenter)
        frame_layout.addSpacing(4)
        frame_layout.addWidget(btn_close, alignment=Qt.AlignRight)

        layout.addWidget(frame)

        def _get_default_cpf() -> str:
            data = _load_login_portal_data()
            return cpf_somente_digitos(data.get("default_cpf", ""))

        def _save_logins_preserving_default(logins: List[Dict[str, str]]):
            default_cpf = _get_default_cpf()
            if default_cpf and any(cpf_somente_digitos(l.get("cpf", "")) == default_cpf for l in logins):
                save_logins_portal(logins, default_cpf=default_cpf)
            else:
                save_logins_portal(logins, default_cpf=(logins[0].get("cpf") if logins else ""))

        def _load_logins_for_ui():
            logins = load_logins_portal()
            items = [cpf_formatado(l.get("cpf", "")) for l in logins]
            return logins, items

        def _set_default_login():
            logins, items = _load_logins_for_ui()
            if not items:
                QMessageBox.information(dlg, "Login Padrão", "Nenhum login cadastrado.")
                return

            cpf_atual = _get_default_cpf()
            label = "Selecione qual login será o padrão:"
            if cpf_atual:
                label = f"{label}\n\nAtual: {cpf_formatado(cpf_atual)}"

            selecionado, ok = self._dialog_selecionar_empresa(
                "Definir Login Padrão",
                label,
                items
            )
            if not (ok and selecionado):
                return

            cpf_sel = cpf_somente_digitos(selecionado)
            if len(cpf_sel) != 11:
                QMessageBox.warning(dlg, "Dados inválidos", "Não foi possível identificar o CPF selecionado.")
                return

            save_logins_portal(logins, default_cpf=cpf_sel)
            QMessageBox.information(
                dlg,
                "Sucesso",
                f"Login padrão definido como {cpf_formatado(cpf_sel)}."
            )

        def _move_empresas_to_login():
            logins, items = _load_logins_for_ui()
            if not items:
                QMessageBox.information(dlg, "Mover Empresas", "Nenhum login cadastrado.")
                return

            empresas_map = getattr(self, "empresas", None)
            if not isinstance(empresas_map, dict) or not empresas_map:
                QMessageBox.information(dlg, "Mover Empresas", "Nenhuma empresa cadastrada.")
                return

            # Mantem a mesma ordem usada no dashboard (quando existir).
            try:
                ies = list(getattr(self, "produtos", []) or [])
            except Exception:
                ies = []
            if not ies:
                ies = list(empresas_map.keys())

            empresas_list: list[dict] = []
            for ie in ies:
                ie_digits = ie_somente_digitos(str(ie))
                if not ie_digits:
                    continue
                info = empresas_map.get(ie_digits) or empresas_map.get(str(ie)) or {}
                if not isinstance(info, dict):
                    info = {}
                empresas_list.append(
                    {
                        "ie": ie_digits,
                        "display_name": (info.get("display_name") or "").strip(),
                        "cnpj": info.get("cnpj") or "",
                        "login_cpf": cpf_somente_digitos(info.get("login_cpf", "") or ""),
                    }
                )

            login_cpfs = [
                cpf_somente_digitos((l or {}).get("cpf", ""))
                for l in (logins or [])
                if len(cpf_somente_digitos((l or {}).get("cpf", ""))) == 11
            ]

            sel_ies = self._dialog_selecionar_empresas_multiplas(
                "Selecionar Empresas",
                "Selecione uma ou mais empresas para mover para um login do portal:",
                empresas_list,
                login_cpfs=login_cpfs,
            )
            if not sel_ies:
                return

            selecionado_login, ok = self._dialog_selecionar_empresa(
                "Selecionar Login",
                "Para qual login voce quer mover essas empresas?",
                items,
            )
            if not (ok and selecionado_login):
                return

            cpf_dest = cpf_somente_digitos(selecionado_login)
            if len(cpf_dest) != 11:
                QMessageBox.warning(dlg, "Dados invalidos", "Nao foi possivel identificar o CPF do login selecionado.")
                return

            if not self._dialog_confirmacao(
                "Confirmar",
                f"Vincular {len(sel_ies)} empresa(s) ao login {cpf_formatado(cpf_dest)}?",
            ):
                return

            changed = 0
            for ie in sel_ies:
                ie = ie_somente_digitos(ie)
                if not ie:
                    continue
                if ie not in self.empresas:
                    continue
                entry = self.empresas.get(ie) or {}
                if not isinstance(entry, dict):
                    # Nao deve acontecer, mas evita quebrar.
                    entry = {"display_name": str(entry), "cnpj": ""}
                entry["login_cpf"] = cpf_dest
                self.empresas[ie] = entry
                changed += 1

            self._salvar_empresas_config()
            try:
                self._recarregar_lista_empresas()
            except Exception:
                pass

            QMessageBox.information(
                dlg,
                "Sucesso",
                f"{changed} empresa(s) vinculada(s) ao login {cpf_formatado(cpf_dest)} com sucesso.",
            )
            try:
                self.update_log(
                    f"[OK] [EMPRESAS] Vinculadas ao login {cpf_formatado(cpf_dest)}: {changed} empresa(s)."
                )
            except Exception:
                pass

        def _add_login():
            cpf, senha, ok = self._dialog_login_portal_form("Adicionar Login")
            if not ok:
                return
            logins, _ = _load_logins_for_ui()
            if any(cpf_somente_digitos(l.get("cpf", "")) == cpf for l in logins):
                QMessageBox.warning(dlg, "Erro", f"CPF '{cpf_formatado(cpf)}' já cadastrado.")
                return
            logins.append({"cpf": cpf, "senha": senha})
            _save_logins_preserving_default(logins)
            QMessageBox.information(dlg, "Sucesso", "Login do portal adicionado com sucesso.")

        def _edit_login():
            logins, items = _load_logins_for_ui()
            if not items:
                QMessageBox.information(dlg, "Editar Login", "Nenhum login cadastrado.")
                return
            selecionado, ok = self._dialog_selecionar_empresa(
                "Editar Login",
                "Selecione o login para editar:",
                items
            )
            if not (ok and selecionado):
                return
            cpf_sel = cpf_somente_digitos(selecionado)
            idx = next((i for i, l in enumerate(logins) if cpf_somente_digitos(l.get("cpf", "")) == cpf_sel), -1)
            if idx < 0:
                QMessageBox.warning(dlg, "Erro", "Não foi possível identificar o login selecionado.")
                return
            cpf_new, senha_new, ok_edit = self._dialog_login_portal_form(
                "Editar Login",
                cpf_init=logins[idx].get("cpf", ""),
                senha_init=logins[idx].get("senha", "")
            )
            if not ok_edit:
                return
            if cpf_new != cpf_sel and any(cpf_somente_digitos(l.get("cpf", "")) == cpf_new for l in logins):
                QMessageBox.warning(dlg, "Erro", f"CPF '{cpf_formatado(cpf_new)}' já cadastrado.")
                return
            logins[idx] = {"cpf": cpf_new, "senha": senha_new}
            _save_logins_preserving_default(logins)
            QMessageBox.information(dlg, "Sucesso", "Login do portal atualizado com sucesso.")

        def _del_login():
            logins, items = _load_logins_for_ui()
            if not items:
                QMessageBox.information(dlg, "Excluir Login", "Nenhum login cadastrado.")
                return
            selecionado, ok = self._dialog_selecionar_empresa(
                "Excluir Login",
                "Selecione o login para excluir:",
                items
            )
            if not (ok and selecionado):
                return
            cpf_sel = cpf_somente_digitos(selecionado)
            if QMessageBox.question(
                dlg,
                "Excluir Login",
                f"Confirma excluir o login {cpf_formatado(cpf_sel)}?"
            ) != QMessageBox.Yes:
                return
            logins = [l for l in logins if cpf_somente_digitos(l.get("cpf", "")) != cpf_sel]
            _save_logins_preserving_default(logins)
            QMessageBox.information(dlg, "Sucesso", "Login do portal excluído com sucesso.")

        btn_add.clicked.connect(_add_login)
        btn_edit.clicked.connect(_edit_login)
        btn_default.clicked.connect(_set_default_login)
        btn_move_emp.clicked.connect(_move_empresas_to_login)
        btn_del.clicked.connect(_del_login)
        btn_close.clicked.connect(dlg.accept)

        dlg.exec()

    def on_intervalo_dias_toggled(self, checked: bool):
        """Habilita/desabilita o campo e mantém ele limpinho ao ativar."""
        self.line_intervalo_dias.setEnabled(checked)
        if checked:
            # quando o usuário marca a opção, o campo vem limpo, sem lixo visual
            self.line_intervalo_dias.clear()

    # ---------- Agendamento ----------

    def _init_schedule_defaults(self):
        dt = datetime.now() + timedelta(minutes=2)
        self.line_agenda_data.setText(dt.strftime("%d/%m/%Y"))
        self.line_agenda_hora.setText(dt.strftime("%H"))
        self.line_agenda_min.setText(dt.strftime("%M"))
        self.lbl_agenda_countdown.clear()
        self.line_agenda_data.setEnabled(False)
        self.line_agenda_hora.setEnabled(False)
        self.line_agenda_min.setEnabled(False)

    def on_agendar_toggled(self, checked: bool):
        self.line_agenda_data.setEnabled(checked)
        self.line_agenda_hora.setEnabled(checked)
        self.line_agenda_min.setEnabled(checked)
        if not checked:
            self._scheduled_pending = False
            self._scheduled_start_dt = None
            self.schedule_timer.stop()
            self.lbl_agenda_countdown.clear()
        else:
            self._init_schedule_defaults()
            self.line_agenda_data.setEnabled(True)
            self.line_agenda_hora.setEnabled(True)
            self.line_agenda_min.setEnabled(True)

    def _parse_schedule_datetime(self):
        try:
            data = self.line_agenda_data.text().strip()
            hora = int(self.line_agenda_hora.text().strip() or 0)
            minuto = int(self.line_agenda_min.text().strip() or 0)
            dt = datetime.strptime(data, "%d/%m/%Y")
            dt = dt.replace(hour=hora, minute=minuto, second=0, microsecond=0)
            return dt
        except Exception:
            return None

    def _update_schedule_countdown(self):
        if not self._scheduled_pending or not self._scheduled_start_dt:
            self.schedule_timer.stop()
            self.lbl_agenda_countdown.clear()
            return
        now = datetime.now()
        delta = (self._scheduled_start_dt - now).total_seconds()
        if delta <= 0:
            self.schedule_timer.stop()
            self.lbl_agenda_countdown.setText("Iniciando...")
            self._scheduled_pending = False
            self._scheduled_start_dt = None
            self.iniciar_automacao()
            return
        horas = int(delta // 3600)
        minutos = int((delta % 3600) // 60)
        segundos = int(delta % 60)
        self.lbl_agenda_countdown.setText(f"Início em {horas:02d}:{minutos:02d}:{segundos:02d}")

    # -----------------------------
    # Header responsivo (somente botoes/filtros do dashboard)
    # -----------------------------
    def _dash_clear_grid(self, grid: QGridLayout):
        try:
            while grid.count():
                grid.takeAt(0)
        except Exception:
            pass

    def _dash_set_header_mode(self, mode: str):
        if not hasattr(self, '_dash_header_gl') or self._dash_header_gl is None:
            return
        mode = 'compact' if (mode or '').lower().startswith('c') else 'wide'
        if getattr(self, '_dash_header_mode', None) == mode:
            return
        self._dash_header_mode = mode
        gl = self._dash_header_gl
        self._dash_clear_grid(gl)

        ws = [
            self._dash_hdr_lbl_emp, self.cmb_dash_empresa,
            self._dash_hdr_lbl_mod, self.cmb_dash_modelo,
            self._dash_hdr_lbl_op, self.cmb_dash_operacao,
            self._dash_hdr_lbl_per, self.dash_date_ini,
            self._dash_hdr_lbl_ate, self.dash_date_fim,
            self.btn_dash_refresh, self.chk_enviar_supabase, self.lbl_dash_status,
        ]
        for w in ws:
            try:
                w.setVisible(False)
            except Exception:
                pass

        if mode == 'wide':
            gl.addWidget(self._dash_hdr_lbl_emp, 0, 0)
            gl.addWidget(self.cmb_dash_empresa, 0, 1)
            gl.addWidget(self._dash_hdr_lbl_mod, 0, 2)
            gl.addWidget(self.cmb_dash_modelo, 0, 3)
            gl.addWidget(self._dash_hdr_lbl_op, 0, 4)
            gl.addWidget(self.cmb_dash_operacao, 0, 5)
            gl.addWidget(self._dash_hdr_lbl_per, 0, 6)
            gl.addWidget(self.dash_date_ini, 0, 7)
            gl.addWidget(self._dash_hdr_lbl_ate, 0, 8)
            gl.addWidget(self.dash_date_fim, 0, 9)
            gl.addWidget(self.btn_dash_refresh, 0, 10)
            gl.addWidget(self.chk_enviar_supabase, 0, 11)
            gl.addWidget(self.lbl_dash_status, 0, 12, 1, 2)
            try:
                gl.setColumnStretch(1, 5)
                gl.setColumnStretch(11, 1)
            except Exception:
                pass
        else:
            gl.addWidget(self._dash_hdr_lbl_emp, 0, 0)
            gl.addWidget(self.cmb_dash_empresa, 0, 1, 1, 4)
            gl.addWidget(self.btn_dash_refresh, 0, 5)
            gl.addWidget(self.chk_enviar_supabase, 0, 6)

            gl.addWidget(self._dash_hdr_lbl_mod, 1, 0)
            gl.addWidget(self.cmb_dash_modelo, 1, 1)
            gl.addWidget(self._dash_hdr_lbl_op, 1, 2)
            gl.addWidget(self.cmb_dash_operacao, 1, 3, 1, 2)

            gl.addWidget(self._dash_hdr_lbl_per, 2, 0)
            gl.addWidget(self.dash_date_ini, 2, 1)
            gl.addWidget(self._dash_hdr_lbl_ate, 2, 2)
            gl.addWidget(self.dash_date_fim, 2, 3)
            gl.addWidget(self.lbl_dash_status, 2, 4, 1, 3)
            try:
                gl.setColumnStretch(1, 3)
                gl.setColumnStretch(4, 2)
            except Exception:
                pass

        for w in ws:
            try:
                w.setVisible(True)
            except Exception:
                pass

    def _dash_apply_header_responsive(self):
        try:
            w = self._dash_header_frame.width() if hasattr(self, '_dash_header_frame') else self.width()
        except Exception:
            w = self.width()
        self._dash_set_header_mode('compact' if w < 980 else 'wide')

    def resizeEvent(self, event):
        try:
            super().resizeEvent(event)
        except Exception:
            pass
        try:
            if hasattr(self, '_dash_header_gl') and self._dash_header_gl is not None:
                self._dash_apply_header_responsive()
        except Exception:
            pass

# =============================================================================
# Global exception safety helpers
# =============================================================================

_SAFE_APP_EXCEPTION_ACTIVE = False

def _report_unhandled_exception(exc_type, exc_value, exc_tb):
    """Mostra/registras erros não tratados sem encerrar a aplicação."""
    if exc_type is None:
        return
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_tb)
        return

    global _SAFE_APP_EXCEPTION_ACTIVE
    if _SAFE_APP_EXCEPTION_ACTIVE:
        return
    _SAFE_APP_EXCEPTION_ACTIVE = True

    try:
        stack = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
        print("[UNHANDLED EXCEPTION]\n" + stack)
        app = QApplication.instance()
        if app is not None and not app.closingDown():
            summary = str(exc_value) or exc_type.__name__
            QMessageBox.critical(
                None,
                "Erro inesperado",
                (
                    "Ocorreu um erro interno inesperado.\n\n"
                    f"{summary}\n\n"
                    "Você pode continuar usando a aplicação, mas verifique os logs "
                    "no console para mais detalhes."
                ),
            )
    except Exception:
        pass
    finally:
        _SAFE_APP_EXCEPTION_ACTIVE = False


def _handle_unraisable_hook(unraisable):
    exc_type = getattr(unraisable, "exc_type", None)
    if exc_type:
        _report_unhandled_exception(
            exc_type,
            getattr(unraisable, "exc_value", None),
            getattr(unraisable, "exc_traceback", None),
        )


class SafeApplication(QApplication):
    """Wrapper que captura exceções não tratadas dentro dos eventos do Qt."""

    def notify(self, receiver, event):
        try:
            return super().notify(receiver, event)
        except Exception:
            exc_type, exc_value, exc_tb = sys.exc_info()
            if exc_type and issubclass(exc_type, KeyboardInterrupt):
                raise
            _report_unhandled_exception(exc_type, exc_value, exc_tb)
            return False


# =============================================================================
# Main
# =============================================================================

# =============================================================================
# ENTRY POINT - Ponto de entrada da aplicação
# =============================================================================
if __name__ == "__main__":
    scheduler_mode = is_scheduler_mode_enabled()
    app = SafeApplication(sys.argv)
    sys.excepthook = _report_unhandled_exception
    try:
        sys.unraisablehook = _handle_unraisable_hook
    except AttributeError:
        pass
    try:
        app.setWindowIcon(QIcon(ICO_PATH))
    except Exception:
        pass

    # === NOVO: validar licença antes de abrir a UI ===
    if not scheduler_mode and not ensure_license_valid(app):

        # usuário saiu ou licença inválida
        sys.exit(0)
    style_text = """
        /* QInputDialog (nome da empresa, seleção, etc.) */
        QInputDialog {
            background-color: #0B1220;
        }
        QInputDialog QLabel {
            color: #E8F4FF;
            font-family: Verdana;
            font-size: 10pt;
        }
        QInputDialog QLineEdit {
            background-color: #0F1E35;
            color: #E8F4FF;
            border-radius: 6px;
            padding: 4px;
            font-family: Verdana;
            font-size: 9pt;
        }
        QInputDialog QPushButton {
            font:9pt 'Verdana'; font-weight:bold; color:#fff; padding:8px 14px;
            border-radius:8px; border:1px solid rgba(255,255,255,0.08);
            background:#102A4C;
        }
        QInputDialog QPushButton:hover {
            background:#12304E;
        }
        QInputDialog QPushButton:pressed {
            background:#0F1E35;
            padding-top:11px;
            padding-bottom:9px;
        }

        /* QMessageBox (confirmação de excluir, avisos, etc.) */
        QMessageBox {
            background-color: #0B1220;
        }
        QMessageBox QLabel {
            color: #E8F4FF;
            font-family: Verdana;
            font-size: 10pt;
        }
        QMessageBox QPushButton {
            font:9pt 'Verdana'; font-weight:bold; color:#fff; padding:8px 14px;
            border-radius:8px; border:1px solid rgba(255,255,255,0.08);
            background:#102A4C;
        }
        QMessageBox QPushButton:hover {
            background:#12304E;
        }
        QMessageBox QPushButton:pressed {
            background:#0F1E35;
            padding-top:11px;
            padding-bottom:9px;
        }

        /* Dialogos genéricos */
        QDialog {
            background-color: #0B1220;
        }
        QDialog QLabel {
            color: #E8F4FF;
            font-family: Verdana;
            font-size: 10pt;
        }
        QDialog QLineEdit {
            background-color: #0F1E35;
            color: #E8F4FF;
            border-radius: 6px;
            padding: 4px;
            font-family: Verdana;
            font-size: 9pt;
        }

        /* Inputs padrão */
        QLineEdit {
            background-color: #0F1E35;
            color: #E8F4FF;
        }
        QComboBox, QComboBox QListView {
            color: #E8F4FF;
            background-color: #0F1E35;
        }
        /* Animated checkbox indicators */
        QCheckBox {
            color: #E8F4FF;
            font-family: Verdana;
            font-size: 10pt;
            padding-left: 34px;
            spacing: 6px;
        }
        QCheckBox::indicator {
            width: 0;
            height: 0;
            margin-left: 0;
            border: none;
            background: transparent;
        }
        QCheckBox:hover {
            color: #F5FBFF;
        }
        QRadioButton {
            color: #E8F4FF;
            font-family: Verdana;
            font-size: 10pt;
        }
        QRadioButton::indicator {
            width: 16px;
            height: 16px;
            border: 1px solid rgba(255,255,255,0.4);
            background: #0F1E35;
            border-radius: 8px;
        }
        QRadioButton::indicator:checked {
            border-color: #74B9FF;
            background: qradialgradient(cx:0.5, cy:0.5, radius: 0.6,
                fx:0.5, fy:0.5, stop:0 #74B9FF, stop:1 #234B8A);
        }
        QRadioButton::indicator:unchecked {
            border-color: rgba(255,255,255,0.35);
        }
        QRadioButton::indicator:hover {
            border-color: #A7C7FF;
        }
        /* Botões padrão (quando não tiver estilo específico) */
        QPushButton {
            font:9pt 'Verdana'; font-weight:bold; color:#fff; padding:8px 14px;
            border-radius:8px; border:1px solid rgba(255,255,255,0.08);
            background:#102A4C;
        }
        QPushButton:hover {
            background:#12304E;
        }
        QPushButton:pressed {
            background:#0F1E35;
            padding-top:11px;
            padding-bottom:9px;
        }
    """
    app.setStyleSheet(style_text)

    # =============================================================================
    # INICIALIZAÇÃO DA APLICAÇÃO
    # =============================================================================
    window = MyCompactUI()
    app.aboutToQuit.connect(lambda: _mark_process_inactive(window, "qt_about_to_quit"))
    _install_process_shutdown_handlers(window)
    if scheduler_mode:
        window.hide()
    else:
        window.show()
    sys.exit(app.exec())
